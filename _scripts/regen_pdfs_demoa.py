"""会社名置換後、残存する76個のPDFを再生成

対象PDF:
1. 規程_職務権限規程_R18.pdf (5 locations)
2. 注文書_ORD-XXXX.pdf + PLC-S-001_注文書_*.pdf (28 files) from VA05 xlsx
3. 請求書_INV-XXXXXX.pdf (13 files) from VF05 csv
4. 発注書_PO-XXXX.pdf (PLC-P 27, ELC 3) from ME2N csv
5. 2025年度内部監査計画書.pdf (1 file)
6. SOC1_TypeII_Report_SIerA_FY2024.pdf ELC copy from ITGC
"""
import csv
import os
import sys
import io
import shutil
import warnings
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from fpdf import FPDF

warnings.filterwarnings('ignore')
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

ROOT = Path(r"C:\Users\nyham\work\demo_data")
FONT_REG = r"C:\Windows\Fonts\YuGothM.ttc"
FONT_BLD = r"C:\Windows\Fonts\YuGothB.ttc"

COMPANY = 'デモA株式会社'
CEO = '山本 健一'


def new_pdf(margin=15):
    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.set_margins(margin, margin, margin)
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_font('YG', '', FONT_REG, uni=True)
    pdf.add_font('YGB', '', FONT_BLD, uni=True)
    return pdf


# ==============================================================
# 1. 職務権限規程 R18
# ==============================================================
def gen_r18(out_path):
    pdf = new_pdf()
    pdf.add_page()

    # Cover
    pdf.set_font('YGB', '', 28)
    pdf.ln(40)
    pdf.cell(0, 20, '職 務 権 限 規 程', ln=1, align='C')
    pdf.set_font('YGB', '', 16)
    pdf.cell(0, 10, '（規程番号 R18）', ln=1, align='C')
    pdf.ln(30)
    pdf.set_font('YG', '', 14)
    pdf.cell(0, 8, COMPANY, ln=1, align='C')
    pdf.ln(50)
    pdf.set_font('YG', '', 11)
    pdf.cell(0, 7, '制定： 1998年4月1日', ln=1, align='C')
    pdf.cell(0, 7, '最終改訂： 2025年4月1日（第15回改訂）', ln=1, align='C')
    pdf.cell(0, 7, '主管部門： 総務部', ln=1, align='C')

    # TOC
    pdf.add_page()
    pdf.set_font('YGB', '', 16)
    pdf.cell(0, 10, '目 次', ln=1, align='C')
    pdf.ln(4)
    pdf.set_font('YG', '', 11)
    toc = [
        ('第1章 総則', '3'),
        ('  第1条 目的', '3'),
        ('  第2条 適用範囲', '3'),
        ('  第3条 定義', '3'),
        ('  第4条 権限行使の原則', '4'),
        ('第2章 経営に関する権限', '4'),
        ('  第5条 取締役会の決議事項', '4'),
        ('  第6条 代表取締役の権限', '5'),
        ('  第7条 管理本部長(CFO)の権限', '5'),
        ('第3章 業務分野別の承認権限', '6'),
        ('  第8条 販売関連の承認権限', '6'),
        ('  第9条 購買関連の承認権限', '7'),
        ('  第10条 人事関連の承認権限', '8'),
        ('  第11条 IT関連の承認権限', '9'),
        ('  第12条 財務・会計関連の承認権限', '9'),
        ('第4章 例外手続', '10'),
        ('  第13条 緊急時の権限代行', '10'),
        ('  第14条 権限逸脱発生時の手続', '10'),
        ('第5章 職務分掌の原則', '11'),
        ('  第15条 職務分掌の基本原則', '11'),
        ('  第16条 併任禁止の職務', '11'),
        ('附則', '11'),
    ]
    for t, p in toc:
        pdf.cell(150, 7, t)
        pdf.cell(0, 7, f'... {p}', ln=1, align='R')

    # Ch.1
    pdf.add_page()
    def header_blue(text):
        pdf.set_x(pdf.l_margin)
        pdf.set_font('YGB', '', 13)
        pdf.set_fill_color(214, 224, 240)
        pdf.cell(0, 9, text, ln=1, fill=True)
    def subheader(text):
        pdf.set_x(pdf.l_margin)
        pdf.set_font('YGB', '', 11)
        pdf.set_fill_color(235, 242, 249)
        pdf.cell(0, 8, text, ln=1, fill=True)
    def body(text):
        pdf.set_x(pdf.l_margin)
        pdf.set_font('YG', '', 10)
        pdf.multi_cell(0, 6, text)

    header_blue('第1章 総則')
    subheader('第1条（目的）')
    body(f'本規程は、{COMPANY}（以下「当社」という）における業務執行上の職務権限及び責任の所在を明確にし、業務の効率的かつ適正な執行を図ることを目的とする。')
    subheader('第2条（適用範囲）')
    body('本規程は当社のすべての役員及び従業員に適用される。子会社については、本規程の趣旨に従い、各社が別途定める。')
    subheader('第3条（定義）')
    body('本規程において使用する用語の定義は次のとおりとする。\n'
         '(1)「業務執行」とは、当社の事業に関する一切の行為をいう。\n'
         '(2)「決裁」とは、権限を有する者が承認することをいう。\n'
         '(3)「稟議」とは、ワークフローシステム（S04）を通じて行う決裁手続をいう。\n'
         '(4)「承認金額」とは、消費税を含まない取引金額（税抜）をいう。')
    subheader('第4条（権限行使の原則）')
    body('職務権限の行使にあたっては、次の原則を遵守しなければならない。\n'
         '(1) 権限の範囲内で行使すること\n'
         '(2) 業務の必要性に基づくこと\n'
         '(3) 関連規程及び法令を遵守すること\n'
         '(4) 職務分掌を守ること（第15条参照）')

    pdf.add_page()
    header_blue('第2章 経営に関する権限')
    subheader('第5条（取締役会の決議事項）')
    body('取締役会は次の事項を決議する。\n'
         '(1) 中長期経営計画及び年度予算の承認\n'
         '(2) 100億円超の設備投資・M&A\n'
         '(3) 10億円超の借入・社債発行\n'
         '(4) 定款変更及び株主総会議案\n'
         '(5) 取締役及び執行役員の選任・解任\n'
         '(6) 規程の制定・改廃（重要なもの）\n'
         '(7) 開示すべき重要な不備の認定')
    subheader('第6条（代表取締役の権限）')
    body('代表取締役は取締役会決議事項以外の経営上重要な事項について決裁する。特に次の事項について個別承認権を有する。\n'
         '(1) 1件¥1億円超の設備投資\n'
         '(2) 1件¥1億円超の契約締結\n'
         '(3) 重要人事（部長以上）\n'
         '(4) 規程の制定・改廃（軽微なもの）')
    subheader('第7条（管理本部長(CFO)の権限）')
    body('管理本部長は経理・財務・人事・総務・法務に関する業務を統括し、次の事項について承認権を有する。\n'
         '(1) 1件¥1億円以下の財務取引\n'
         '(2) 1件¥1億円以下の発注承認\n'
         '(3) 見積りの評価・承認\n'
         '(4) 月次・四半期・年次決算承認\n'
         '(5) 税務申告の承認')

    pdf.add_page()
    header_blue('第3章 業務分野別の承認権限')
    subheader('第8条（販売関連の承認権限）')
    body('販売関連の承認権限は次のとおり定める。')
    # table 8
    pdf.set_font('YGB', '', 10)
    pdf.set_fill_color(48, 84, 150)
    pdf.set_text_color(255, 255, 255)
    pdf.set_x(pdf.l_margin)
    pdf.cell(70, 8, '業務', border=1, align='C', fill=True)
    pdf.cell(60, 8, '承認者', border=1, align='C', fill=True)
    pdf.cell(50, 8, '承認上限', border=1, align='C', fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('YG', '', 10)
    for row in [
        ('通常受注（与信枠内）', '自動承認（SAP）', '－'),
        ('与信限度超過の受注', '営業本部長', '超過全件対象'),
        ('新規顧客登録', '営業本部長+CFO', '全件対象'),
        ('与信限度引上', 'CFO', '年次見直し'),
        ('¥100M超の個別受注', '代表取締役', '全件対象'),
        ('顧客別価格変更（既存）', '営業本部長', '全件対象'),
        ('新規品目の初期価格設定', '営業本部長+CFO', '全件対象'),
        ('返品・値引き（¥5M超）', '営業本部長', '全件対象'),
        ('回収不能認定', 'CFO', '全件対象'),
    ]:
        pdf.set_x(pdf.l_margin)
        pdf.cell(70, 7, row[0], border=1)
        pdf.cell(60, 7, row[1], border=1)
        pdf.cell(50, 7, row[2], border=1)
        pdf.ln()

    pdf.add_page()
    subheader('第9条（購買関連の承認権限）')
    body('購買関連の承認権限は次のとおり定める。金額区分に応じて承認者が自動判定され、SAPワークフロー（S04）経由で承認手続を行う。')
    pdf.set_font('YGB', '', 10)
    pdf.set_fill_color(48, 84, 150)
    pdf.set_text_color(255, 255, 255)
    pdf.set_x(pdf.l_margin)
    pdf.cell(55, 8, '金額区分', border=1, align='C', fill=True)
    pdf.cell(70, 8, '承認者', border=1, align='C', fill=True)
    pdf.cell(55, 8, 'SAPロール', border=1, align='C', fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('YG', '', 10)
    for row in [
        ('〜¥500,000', '購買部担当（主任）', 'PO_CREATE のみ'),
        ('〜¥5,000,000', '購買部課長', 'PO_APPROVE'),
        ('〜¥20,000,000', '購買部長', 'PO_APPROVE'),
        ('〜¥100,000,000', '管理本部長（CFO）', 'PO_APPROVE'),
        ('¥100,000,000超', '代表取締役', 'PO_APPROVE'),
    ]:
        pdf.set_x(pdf.l_margin)
        pdf.cell(55, 7, row[0], border=1)
        pdf.cell(70, 7, row[1], border=1)
        pdf.cell(55, 7, row[2], border=1)
        pdf.ln()
    body('\nその他の購買関連権限：\n'
         '(1) 新規仕入先登録：購買部長（反社チェック・信用調査完了後）\n'
         '(2) 仕入先評価：購買部長が年1回実施\n'
         '(3) 継続契約解除：購買部長\n'
         '(4) 外注加工契約締結：購買部長（¥50M超はCFO追加承認）')

    pdf.add_page()
    subheader('第10条（人事関連の承認権限）')
    pdf.set_font('YGB', '', 10)
    pdf.set_fill_color(48, 84, 150); pdf.set_text_color(255,255,255)
    pdf.set_x(pdf.l_margin)
    pdf.cell(65, 8, '業務', border=1, align='C', fill=True)
    pdf.cell(115, 8, '承認者', border=1, align='C', fill=True)
    pdf.ln()
    pdf.set_text_color(0,0,0); pdf.set_font('YG', '', 10)
    for row in [
        ('採用（一般社員）', '人事部長'),
        ('採用（課長以上）', '社長'),
        ('異動・昇格', '人事部長→関連部門長→社長'),
        ('給与改定（年次）', '人事部長→CFO→社長'),
        ('賞与支給', '人事部長→CFO→社長'),
        ('退職手続', '人事部長'),
        ('懲戒処分', '人事部長→コンプラ委員会→社長'),
    ]:
        pdf.set_x(pdf.l_margin)
        pdf.cell(65, 7, row[0], border=1)
        pdf.cell(115, 7, row[1], border=1)
        pdf.ln()

    pdf.add_page()
    subheader('第11条（IT関連の承認権限）')
    pdf.set_font('YGB', '', 10)
    pdf.set_fill_color(48, 84, 150); pdf.set_text_color(255,255,255)
    pdf.set_x(pdf.l_margin)
    pdf.cell(65, 8, '業務', border=1, align='C', fill=True)
    pdf.cell(115, 8, '承認者', border=1, align='C', fill=True)
    pdf.ln()
    pdf.set_text_color(0,0,0); pdf.set_font('YG', '', 10)
    for row in [
        ('ユーザID新規登録', '所属部門長 + 情シス部アプリリーダー'),
        ('特権ID付与', '情シス部長 + CFO'),
        ('ユーザID削除（退職）', '人事部 → 情シス部担当'),
        ('プログラム変更', '情シス部アプリリーダー + 業務部門長'),
        ('緊急変更', '情シス部長（事後承認）'),
        ('本番移送', '情シス部アプリリーダー（専任者）'),
        ('IT投資計画', '情シス部長 → CFO → 取締役会'),
    ]:
        pdf.set_x(pdf.l_margin)
        pdf.cell(65, 7, row[0], border=1)
        pdf.cell(115, 7, row[1], border=1)
        pdf.ln()

    subheader('第12条（財務・会計関連の承認権限）')
    pdf.set_font('YGB', '', 10)
    pdf.set_fill_color(48, 84, 150); pdf.set_text_color(255,255,255)
    pdf.set_x(pdf.l_margin)
    pdf.cell(65, 8, '業務', border=1, align='C', fill=True)
    pdf.cell(115, 8, '承認者', border=1, align='C', fill=True)
    pdf.ln()
    pdf.set_text_color(0,0,0); pdf.set_font('YG', '', 10)
    for row in [
        ('月次決算承認', '経理部長'),
        ('四半期決算承認', '経理部長 → CFO → 取締役会'),
        ('年次決算承認', 'CFO → 取締役会 → 監査等委員会'),
        ('連結仕訳（非定型）', '経理部長 → CFO'),
        ('会計上の見積（引当金等）', 'CFO → 監査等委員会'),
    ]:
        pdf.set_x(pdf.l_margin)
        pdf.cell(65, 7, row[0], border=1)
        pdf.cell(115, 7, row[1], border=1)
        pdf.ln()

    pdf.add_page()
    header_blue('第4章 例外手続')
    subheader('第13条（緊急時の権限代行）')
    body('通常の承認者が不在等により承認が困難な場合、以下の代行者による承認を認める。\n'
         '(1) 営業本部長不在時：営業副本部長または経営企画部長\n'
         '(2) 購買部長不在時：管理本部長（CFO）\n'
         '(3) 代表取締役不在時：取締役会議長\n'
         '代行承認は、通常承認者の復帰後に事後報告を行う。')
    subheader('第14条（権限逸脱発生時の手続）')
    body('承認権限を超過した承認が発覚した場合、次の手続をとる。\n'
         '(1) 権限者による事後承認の取得\n'
         '(2) 内部監査室への報告\n'
         '(3) 再発防止策の策定と実施\n'
         '(4) 監査等委員会への報告（重要な場合）')

    header_blue('第5章 職務分掌の原則')
    subheader('第15条（職務分掌の基本原則）')
    body('権限の乱用及び誤謬を防止するため、次の職務は同一人物が兼任してはならない。')
    subheader('第16条（併任禁止の職務）')
    pdf.set_font('YGB', '', 10)
    pdf.set_fill_color(48, 84, 150); pdf.set_text_color(255,255,255)
    pdf.set_x(pdf.l_margin)
    pdf.cell(55, 8, '業務', border=1, align='C', fill=True)
    pdf.cell(125, 8, '禁止される兼任', border=1, align='C', fill=True)
    pdf.ln()
    pdf.set_text_color(0,0,0); pdf.set_font('YG', '', 10)
    for row in [
        ('発注', '発注作成 ⇔ 発注承認'),
        ('検収', '発注承認 ⇔ 検収'),
        ('買掛・支払', '検収 ⇔ 買掛計上 ⇔ 支払実行'),
        ('受注・売上', '受注登録 ⇔ 与信限度設定'),
        ('入金・消込', '現金受領 ⇔ 消込記帳 ⇔ 照合'),
        ('ユーザ管理', 'ユーザ作成 ⇔ ユーザ申請承認'),
        ('変更管理', '変更開発 ⇔ 本番移送'),
    ]:
        pdf.set_x(pdf.l_margin)
        pdf.cell(55, 7, row[0], border=1)
        pdf.cell(125, 7, row[1], border=1)
        pdf.ln()

    pdf.ln(5)
    pdf.set_font('YGB', '', 14); pdf.cell(0, 10, '附 則', ln=1)
    pdf.set_font('YG', '', 10)
    body('1. 本規程は1998年4月1日より施行する。\n'
         '2. 本規程の改廃は取締役会の決議による。\n'
         '3. 本規程に定めのない事項については、関連規程又は総務部長の判断による。')
    pdf.ln(4)
    body('改訂履歴：\n'
         '第1回改訂  1999年10月1日\n'
         '（中略）\n'
         '第14回改訂 2024年4月1日\n'
         '第15回改訂 2025年4月1日（金額区分の見直し、職務分掌の明文化）')
    pdf.ln(10)
    pdf.set_x(pdf.l_margin)
    pdf.cell(0, 7, '承認： 2025年4月1日施行', ln=1, align='R')
    pdf.add_page()
    pdf.set_font('YG', '', 11)
    pdf.ln(5)
    pdf.cell(0, 10, f'代表取締役社長 {CEO} [印]', ln=1, align='R')

    pdf.output(str(out_path))


# ==============================================================
# 2. 注文書 (ORD-*)
# ==============================================================
def gen_chumon(out_path, ord_no, cust_po, issue_date, cust_name, cust_code, items, delivery_date, delivery_place='貴社指定倉庫', payment_term='月末締 翌月末払'):
    pdf = new_pdf()
    pdf.add_page()

    pdf.set_font('YGB', '', 24)
    pdf.cell(0, 15, '注 文 書', ln=1, align='C')
    pdf.ln(2)
    pdf.set_font('YG', '', 11)
    pdf.cell(0, 6, f'(顧客側) 注文書番号: {cust_po}', ln=1, align='R')
    pdf.cell(0, 6, f'発行日: {issue_date}', ln=1, align='R')
    pdf.ln(3)

    pdf.set_font('YGB', '', 13)
    pdf.cell(0, 8, f'{COMPANY} 御中', ln=1)
    pdf.set_font('YG', '', 10)
    pdf.cell(0, 6, '営業本部 担当者殿', ln=1)
    pdf.ln(3)

    pdf.set_font('YGB', '', 11)
    pdf.cell(0, 7, f'発注元: {cust_name}', ln=1, align='R')
    pdf.set_font('YG', '', 10)
    pdf.cell(0, 6, f'顧客コード: {cust_code}', ln=1, align='R')
    pdf.ln(3)

    pdf.set_font('YG', '', 10)
    pdf.multi_cell(0, 6, '下記のとおり発注致します。ご確認のうえ、納期に間に合うよう手配をお願い致します。')
    pdf.ln(2)

    # Items table
    pdf.set_font('YGB', '', 10)
    pdf.set_fill_color(48, 84, 150); pdf.set_text_color(255,255,255)
    pdf.cell(30, 8, '品目コード', border=1, align='C', fill=True)
    pdf.cell(70, 8, '品名', border=1, align='C', fill=True)
    pdf.cell(20, 8, '数量', border=1, align='C', fill=True)
    pdf.cell(30, 8, '単価(円)', border=1, align='C', fill=True)
    pdf.cell(30, 8, '金額(円)', border=1, align='C', fill=True)
    pdf.ln()
    pdf.set_text_color(0,0,0); pdf.set_font('YG', '', 10)
    subtotal = 0
    for item in items:
        code, name, qty, unit_price = item['code'], item['name'], item['qty'], item['unit_price']
        amount = qty * unit_price
        subtotal += amount
        pdf.cell(30, 7, code, border=1)
        pdf.cell(70, 7, name, border=1)
        pdf.cell(20, 7, f'{qty:,}', border=1, align='R')
        pdf.cell(30, 7, f'{unit_price:,}', border=1, align='R')
        pdf.cell(30, 7, f'{amount:,}', border=1, align='R')
        pdf.ln()

    tax = subtotal // 10
    total = subtotal + tax
    pdf.cell(150, 7, '小計', border=1, align='R')
    pdf.cell(30, 7, f'¥ {subtotal:,}', border=1, align='R')
    pdf.ln()
    pdf.cell(150, 7, '消費税 (10%)', border=1, align='R')
    pdf.cell(30, 7, f'¥ {tax:,}', border=1, align='R')
    pdf.ln()
    pdf.set_fill_color(255, 242, 204)
    pdf.set_font('YGB', '', 11)
    pdf.cell(150, 8, '合計', border=1, align='R', fill=True)
    pdf.cell(30, 8, f'¥ {total:,}', border=1, align='R', fill=True)
    pdf.ln(10)

    # Delivery info
    pdf.set_font('YGB', '', 11)
    pdf.cell(0, 7, '■ 納入条件', ln=1)
    pdf.set_font('YG', '', 10)
    pdf.cell(40, 7, '納期', border=1)
    pdf.cell(140, 7, delivery_date, border=1)
    pdf.ln()
    pdf.cell(40, 7, '納入場所', border=1)
    pdf.cell(140, 7, delivery_place, border=1)
    pdf.ln()
    pdf.cell(40, 7, '支払条件', border=1)
    pdf.cell(140, 7, payment_term, border=1)
    pdf.ln()

    pdf.output(str(out_path))


# ==============================================================
# 3. 請求書 (INV-*)
# ==============================================================
def gen_seikyu(out_path, inv_no, issue_date, cust_code, cust_name, amount_before_tax, tax, amount_with_tax, due_date):
    pdf = new_pdf()
    pdf.add_page()

    pdf.set_font('YGB', '', 24)
    pdf.cell(0, 15, '請 求 書', ln=1, align='C')
    pdf.ln(2)
    pdf.set_font('YG', '', 11)
    pdf.cell(0, 6, f'請求書番号: {inv_no}', ln=1, align='R')
    pdf.cell(0, 6, f'発行日: {issue_date}', ln=1, align='R')
    pdf.ln(3)

    pdf.set_font('YGB', '', 13)
    pdf.cell(0, 8, f'{cust_name} 御中', ln=1)
    pdf.set_font('YG', '', 10)
    pdf.cell(0, 6, f'顧客コード: {cust_code}', ln=1)
    pdf.ln(5)

    pdf.set_font('YGB', '', 11)
    pdf.cell(0, 7, f'発行元: {COMPANY}', ln=1, align='R')
    pdf.set_font('YG', '', 10)
    pdf.cell(0, 6, '経理部 AR担当', ln=1, align='R')
    pdf.ln(3)

    pdf.multi_cell(0, 6, '下記のとおりご請求申し上げます。')
    pdf.ln(2)

    # Amount block
    pdf.set_font('YG', '', 10)
    pdf.cell(60, 8, '請求金額(税抜)', border=1)
    pdf.cell(120, 8, f'¥ {amount_before_tax:,}', border=1, align='R')
    pdf.ln()
    pdf.cell(60, 8, '消費税 (10%)', border=1)
    pdf.cell(120, 8, f'¥ {tax:,}', border=1, align='R')
    pdf.ln()
    pdf.set_fill_color(255, 242, 204)
    pdf.set_font('YGB', '', 12)
    pdf.cell(60, 10, '税込合計金額', border=1, fill=True)
    pdf.cell(120, 10, f'¥ {amount_with_tax:,}', border=1, align='R', fill=True)
    pdf.ln(15)

    # Payment info
    pdf.set_font('YGB', '', 11)
    pdf.cell(0, 7, '■ お支払条件', ln=1)
    pdf.set_font('YG', '', 10)
    pdf.cell(40, 7, '支払期日', border=1)
    pdf.cell(140, 7, due_date, border=1)
    pdf.ln()
    pdf.cell(40, 7, '振込先', border=1)
    pdf.cell(140, 7, 'A銀行 本店営業部 普通 1234567 デモA株式会社', border=1)
    pdf.ln()
    pdf.ln(5)
    pdf.multi_cell(0, 6, '※振込手数料は貴社にてご負担ください。\n※ご不明点は経理部AR担当までお問合せください。')

    pdf.output(str(out_path))


# ==============================================================
# 4. 発注書 (PO-*)
# ==============================================================
def gen_hatchu(out_path, po_no, po_date, vendor_code, vendor_name, items, amount, approver, delivery_date):
    pdf = new_pdf()
    pdf.add_page()

    pdf.set_font('YGB', '', 24)
    pdf.cell(0, 15, '発 注 書', ln=1, align='C')
    pdf.ln(2)
    pdf.set_font('YG', '', 11)
    pdf.cell(0, 6, f'発注番号: {po_no}', ln=1, align='R')
    pdf.cell(0, 6, f'発注日: {po_date}', ln=1, align='R')
    pdf.ln(3)

    pdf.set_font('YGB', '', 13)
    pdf.cell(0, 8, f'{vendor_name} 御中', ln=1)
    pdf.set_font('YG', '', 10)
    pdf.cell(0, 6, f'仕入先コード: {vendor_code}', ln=1)
    pdf.ln(5)

    pdf.set_font('YGB', '', 11)
    pdf.cell(0, 7, f'発注元: {COMPANY}', ln=1, align='R')
    pdf.set_font('YG', '', 10)
    pdf.cell(0, 6, '購買部', ln=1, align='R')
    pdf.ln(3)

    pdf.multi_cell(0, 6, '下記のとおり発注致します。納期に間に合うよう手配をお願い致します。')
    pdf.ln(2)

    # Items
    pdf.set_font('YGB', '', 10)
    pdf.set_fill_color(48, 84, 150); pdf.set_text_color(255,255,255)
    pdf.cell(40, 8, '品目分類', border=1, align='C', fill=True)
    pdf.cell(110, 8, '品名/仕様', border=1, align='C', fill=True)
    pdf.cell(30, 8, '金額(円)', border=1, align='C', fill=True)
    pdf.ln()
    pdf.set_text_color(0,0,0); pdf.set_font('YG', '', 10)
    for item in items:
        pdf.cell(40, 7, item['class'], border=1)
        pdf.cell(110, 7, item['name'], border=1)
        pdf.cell(30, 7, f'{item["amount"]:,}', border=1, align='R')
        pdf.ln()
    pdf.set_fill_color(255, 242, 204)
    pdf.set_font('YGB', '', 11)
    pdf.cell(150, 8, '発注金額合計 (税抜)', border=1, align='R', fill=True)
    pdf.cell(30, 8, f'¥ {amount:,}', border=1, align='R', fill=True)
    pdf.ln(10)

    pdf.set_font('YGB', '', 11)
    pdf.cell(0, 7, '■ 納入・支払条件', ln=1)
    pdf.set_font('YG', '', 10)
    pdf.cell(40, 7, '納期', border=1); pdf.cell(140, 7, delivery_date, border=1); pdf.ln()
    pdf.cell(40, 7, '納入場所', border=1); pdf.cell(140, 7, 'デモA株式会社 指定工場/倉庫', border=1); pdf.ln()
    pdf.cell(40, 7, '支払条件', border=1); pdf.cell(140, 7, '月末締 翌月末払', border=1); pdf.ln()
    pdf.ln(8)

    # Approval
    pdf.set_font('YGB', '', 11)
    pdf.cell(0, 7, '■ 社内承認', ln=1)
    pdf.set_font('YG', '', 10)
    pdf.set_fill_color(48, 84, 150); pdf.set_text_color(255,255,255)
    pdf.cell(50, 8, '役割', border=1, align='C', fill=True)
    pdf.cell(60, 8, '氏名', border=1, align='C', fill=True)
    pdf.cell(40, 8, '承認日', border=1, align='C', fill=True)
    pdf.cell(30, 8, '承認印', border=1, align='C', fill=True)
    pdf.ln()
    pdf.set_text_color(0,0,0)
    pdf.cell(50, 10, '申請者', border=1, align='C')
    pdf.cell(60, 10, '清水 智明 (PUR003)', border=1, align='C')
    pdf.cell(40, 10, po_date, border=1, align='C')
    pdf.set_text_color(200,0,0); pdf.cell(30, 10, '申請', border=1, align='C'); pdf.set_text_color(0,0,0); pdf.ln()
    pdf.cell(50, 10, '承認者', border=1, align='C')
    pdf.cell(60, 10, approver, border=1, align='C')
    pdf.cell(40, 10, po_date, border=1, align='C')
    pdf.set_text_color(200,0,0); pdf.cell(30, 10, '承認', border=1, align='C'); pdf.set_text_color(0,0,0); pdf.ln()

    pdf.output(str(out_path))


# ==============================================================
# 5. 内部監査計画書
# ==============================================================
def gen_audit_plan(out_path):
    pdf = new_pdf()
    pdf.add_page()

    pdf.set_font('YGB', '', 24)
    pdf.cell(0, 15, '2025年度 内部監査計画書', ln=1, align='C')
    pdf.ln(5)
    pdf.set_font('YG', '', 11)
    pdf.cell(0, 7, f'作成: {COMPANY} 内部監査室', ln=1, align='R')
    pdf.cell(0, 7, '作成日: 2025年4月10日', ln=1, align='R')
    pdf.cell(0, 7, '承認: 取締役会 2025年4月15日', ln=1, align='R')
    pdf.ln(8)

    pdf.set_font('YGB', '', 13)
    pdf.set_fill_color(214, 224, 240)
    pdf.cell(0, 9, '1. 監査対象', ln=1, fill=True)
    pdf.set_font('YG', '', 10)
    pdf.multi_cell(0, 6, f'{COMPANY}および連結子会社(デモA東北株式会社/デモA物流サービス/Demo-A Thailand/デモAトレーディング)を対象とする内部統制の整備状況・運用状況評価')
    pdf.ln(3)

    pdf.set_font('YGB', '', 13); pdf.set_fill_color(214, 224, 240)
    pdf.cell(0, 9, '2. 監査スコープ', ln=1, fill=True)
    pdf.set_font('YG', '', 10)
    pdf.multi_cell(0, 6, 'JSOX (金融商品取引法第24条の4の4) に基づく財務報告に係る内部統制の評価:\n'
                          '  (1) 全社的な統制 (ELC) 8統制\n'
                          '  (2) 業務プロセス統制 (PLC) 21統制\n'
                          '  (3) IT全般統制 (ITGC) 10統制\n'
                          '  (4) IT業務処理統制 (ITAC) 5統制\n'
                          '  (5) 決算・財務報告プロセス統制 (FCRP) 5統制\n'
                          '  計 49統制 (補完4統制含め計53統制)')
    pdf.ln(3)

    pdf.set_font('YGB', '', 13); pdf.set_fill_color(214, 224, 240)
    pdf.cell(0, 9, '3. 評価期間', ln=1, fill=True)
    pdf.set_font('YG', '', 10)
    pdf.multi_cell(0, 6, '整備状況評価: 2025年6月〜7月 / 運用状況評価: 2025年10月〜2026年3月')
    pdf.ln(3)

    pdf.set_font('YGB', '', 13); pdf.set_fill_color(214, 224, 240)
    pdf.cell(0, 9, '4. 監査体制', ln=1, fill=True)
    pdf.set_font('YG', '', 10)
    pdf.multi_cell(0, 6, '内部監査室長 長谷川 剛 (IA001) 統括\n'
                          '内部監査担当 大塚 美穂 (IA002) 実地評価\n'
                          '監査法人 独立外部監査人 (XYZ CPA Firm相当) 外部検証')
    pdf.ln(3)

    pdf.set_font('YGB', '', 13); pdf.set_fill_color(214, 224, 240)
    pdf.cell(0, 9, '5. 報告先', ln=1, fill=True)
    pdf.set_font('YG', '', 10)
    pdf.multi_cell(0, 6, '監査等委員会 / CFO / 代表取締役 / 取締役会')

    pdf.output(str(out_path))


# ==============================================================
# Data extraction from CSV/xlsx
# ==============================================================
def load_va05_orders():
    """Load sales orders from VA05 xlsx"""
    path = ROOT / "4.evidence" / "PLC-S" / "SAP_VA05_SalesOrderList_FY2025.xlsx"
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    orders = {}
    for r in range(5, ws.max_row + 1):
        ord_no = ws.cell(r, 2).value
        if not ord_no or not str(ord_no).startswith('ORD-'): continue
        ord_date = ws.cell(r, 3).value
        cust_code = ws.cell(r, 5).value
        cust_name = ws.cell(r, 6).value
        amount = ws.cell(r, 8).value
        ship_date = ws.cell(r, 14).value
        cust_po = ws.cell(r, 15).value or f'CUST-PO-2025-{hash(ord_no) % 100000:05d}'
        if isinstance(ord_date, datetime):
            ord_date = ord_date.strftime('%Y年%m月%d日')
        if isinstance(ship_date, datetime):
            ship_date = ship_date.strftime('%Y年%m月%d日')
        orders[str(ord_no)] = {
            'date': ord_date, 'cust_code': cust_code, 'cust_name': cust_name,
            'amount': amount, 'ship_date': ship_date, 'cust_po': cust_po
        }
    return orders


def load_vf05_invoices():
    path = ROOT / "4.evidence" / "PLC-S" / "SAP_VF05_InvoiceRegister_FY2025.csv"
    invoices = {}
    with open(path, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for row in reader:
            if not row or row[0].startswith('#') or row[0] == '請求書番号': continue
            if len(row) < 10: continue
            inv_no, issue_date, cust_code, cust_name, currency, amt_ex, tax, amt_inc, due_date, method = row[:10]
            invoices[inv_no] = {
                'issue_date': issue_date, 'cust_code': cust_code, 'cust_name': cust_name,
                'amount_ex': int(amt_ex), 'tax': int(tax), 'amount_inc': int(amt_inc),
                'due_date': due_date
            }
    return invoices


def load_me2n_orders():
    path = ROOT / "4.evidence" / "PLC-P" / "SAP_ME2N_PurchaseOrder_Detail_FY2025Samples.csv"
    pos = {}
    with open(path, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for row in reader:
            if not row or row[0].startswith('#') or row[0] == 'サンプル№': continue
            if len(row) < 12: continue
            sno, po_no, po_date, vendor_code, vendor_name, item_class, amount, approver, limit, route, status, memo = row[:12]
            pos[po_no] = {
                'date': po_date, 'vendor_code': vendor_code, 'vendor_name': vendor_name,
                'item_class': item_class, 'amount': int(amount),
                'approver': approver, 'route': route
            }
    return pos


# ==============================================================
# Main
# ==============================================================
def main():
    # 1. R18 (5 copies)
    r18_locations = [
        ROOT / "0.profile" / "規程_職務権限規程_R18.pdf",
        ROOT / "4.evidence" / "ELC" / "規程_職務権限規程_R18.pdf",
        ROOT / "4.evidence" / "ITGC" / "規程_職務権限規程_R18.pdf",
        ROOT / "4.evidence" / "PLC-P" / "規程_職務権限規程_R18.pdf",
        ROOT / "4.evidence" / "PLC-S" / "規程_職務権限規程_R18.pdf",
    ]
    # Generate once then copy
    master = r18_locations[0]
    gen_r18(master)
    for loc in r18_locations[1:]:
        shutil.copy(master, loc)
    print(f"[R18] Generated 1 master + copied to {len(r18_locations)-1} locations")

    # 2. 注文書 (ORD-*)
    orders = load_va05_orders()
    plc_s_dir = ROOT / "4.evidence" / "PLC-S"
    existing_ord_pdfs = [f for f in os.listdir(plc_s_dir) if f.startswith('注文書_ORD') and f.endswith('.pdf')]
    prefixed_pdfs = [f for f in os.listdir(plc_s_dir) if f.startswith('PLC-S-001_注文書_') and f.endswith('.pdf')]

    gen_cnt = 0
    for pdf_name in existing_ord_pdfs:
        # 注文書_ORD-2025-XXXX.pdf
        ord_no = pdf_name.replace('注文書_', '').replace('.pdf', '')
        if ord_no in orders:
            o = orders[ord_no]
            items = [{'code': 'P-30027', 'name': f'ロボットアーム外装パネル (受注単位)', 'qty': max(1, o['amount']//5000 if o['amount'] else 100), 'unit_price': 5000 if o['amount'] else 5000}]
            # Re-calculate to match amount
            if o['amount']:
                # single item at total amount
                items = [{'code': 'P-30027', 'name': '受注品 (一式)', 'qty': 1, 'unit_price': o['amount']}]
            gen_chumon(plc_s_dir / pdf_name, ord_no, o['cust_po'], o['date'],
                       o['cust_name'], o['cust_code'], items, o['ship_date'])
            gen_cnt += 1

    # 与信超過3件のフォールバックデータ (VA05にない)
    fallback_credit = {
        'ORD-2025-0412': {'cust_code': 'C-10008', 'cust_name': 'サンプル顧客H社',
                           'date': '2025年05月15日', 'amount': 12500000,
                           'ship_date': '2025年06月05日', 'cust_po': 'CUST-PO-2025-04120'},
        'ORD-2025-1420': {'cust_code': 'C-10002', 'cust_name': 'サンプル顧客B社',
                           'date': '2025年08月10日', 'amount': 15800000,
                           'ship_date': '2025年08月30日', 'cust_po': 'CUST-PO-2025-14200'},
        'ORD-2025-1876': {'cust_code': 'C-10012', 'cust_name': 'サンプル顧客L社',
                           'date': '2025年09月22日', 'amount': 8600000,
                           'ship_date': '2025年10月12日', 'cust_po': 'CUST-PO-2025-18760'},
    }

    for pdf_name in prefixed_pdfs:
        m = pdf_name.replace('PLC-S-001_注文書_', '').replace('.pdf', '')
        parts = m.split('_')
        ord_no = parts[0]
        o = orders.get(ord_no) or fallback_credit.get(ord_no)
        if o:
            items = [{'code': 'P-30027', 'name': '受注品 (一式)', 'qty': 1, 'unit_price': o['amount'] if o['amount'] else 1000000}]
            gen_chumon(plc_s_dir / pdf_name, ord_no, o['cust_po'], o['date'],
                       o['cust_name'], o['cust_code'], items, o['ship_date'])
            gen_cnt += 1

    print(f"[ORD] Regenerated {gen_cnt} 注文書 PDFs")

    # 3. 請求書 (INV-*)
    invoices = load_vf05_invoices()
    inv_cnt = 0
    for pdf_name in os.listdir(plc_s_dir):
        if not pdf_name.startswith('請求書_INV-') or not pdf_name.endswith('.pdf'): continue
        inv_no = pdf_name.replace('請求書_', '').replace('.pdf', '')
        if inv_no in invoices:
            i = invoices[inv_no]
            gen_seikyu(plc_s_dir / pdf_name, inv_no, i['issue_date'], i['cust_code'],
                       i['cust_name'], i['amount_ex'], i['tax'], i['amount_inc'], i['due_date'])
            inv_cnt += 1
        else:
            # use filename info only
            gen_seikyu(plc_s_dir / pdf_name, inv_no, '2025-XX-XX', 'C-XXXXX',
                       'サンプル顧客 (詳細不明)', 1000000, 100000, 1100000, '2025-XX-XX')
            inv_cnt += 1

    print(f"[INV] Regenerated {inv_cnt} 請求書 PDFs")

    # 4. 発注書 (PO-*)
    pos = load_me2n_orders()
    plc_p_dir = ROOT / "4.evidence" / "PLC-P"
    elc_dir = ROOT / "4.evidence" / "ELC"
    po_cnt = 0

    for d in [plc_p_dir, elc_dir]:
        for pdf_name in os.listdir(d):
            if not pdf_name.startswith('発注書_PO-') or not pdf_name.endswith('.pdf'): continue
            po_no = pdf_name.replace('発注書_', '').replace('.pdf', '')
            if po_no in pos:
                p = pos[po_no]
                items = [{'class': p['item_class'], 'name': f'{p["item_class"]} 一式', 'amount': p['amount']}]
                # compute delivery = po_date + 2 weeks
                try:
                    dt = datetime.strptime(p['date'], '%Y-%m-%d')
                    from datetime import timedelta
                    ship = (dt + timedelta(days=14)).strftime('%Y-%m-%d')
                except:
                    ship = p['date']
                gen_hatchu(d / pdf_name, po_no, p['date'], p['vendor_code'], p['vendor_name'],
                           items, p['amount'], p['approver'], ship)
                po_cnt += 1
            else:
                items = [{'class': '一般購買', 'name': '(詳細は社内システム参照)', 'amount': 1000000}]
                gen_hatchu(d / pdf_name, po_no, '2025-XX-XX', 'V-XXXXX', 'サンプル仕入先 (詳細不明)',
                           items, 1000000, '購買部課長', '2025-XX-XX')
                po_cnt += 1

    print(f"[PO] Regenerated {po_cnt} 発注書 PDFs")

    # 5. 内部監査計画書
    gen_audit_plan(ROOT / "4.evidence" / "ELC" / "2025年度内部監査計画書.pdf")
    print("[AUDIT_PLAN] Regenerated")

    # 6. SOC1 SIer-A (copy from ITGC to ELC)
    src = ROOT / "4.evidence" / "ITGC" / "SOC1_TypeII_Report_SIerA_FY2024.pdf"
    dst = ROOT / "4.evidence" / "ELC" / "SOC1_TypeII_Report_SIerA_FY2024.pdf"
    shutil.copy(src, dst)
    print("[SOC1_ELC] Copied from ITGC")

    print("\n=== ALL PDF REGENERATION COMPLETED ===")


if __name__ == '__main__':
    main()
