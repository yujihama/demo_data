"""ITGC-AC-001 再修正

指摘:
1. SoDマトリクスにSD/購買系のみ記載、25サンプル中19件のロールがマトリクスにない
2. Sample 11 (USER-REG-2025-0011): 管理表/申請書/SU01のロール不一致
3. Sample 1-5 PDF: 部門長承認者が「部門長」総称で個人特定不可
4. 意図的に1サンプルだけマトリクス外ロールにする

修正:
- SAP_SoD_ConflictMatrix_FY2025.xlsx 新規作成 (11ロール×11ロール)
- Sample 22 (品質保証部) の付与ロールを SD_USER→QM_USER に変更 (マトリクス外=意図的例外)
- Sample 11 の管理表ロール: MM_USER,PO_CREATE → MM_USER に統一
- Sample 1-5 のPDF再生成 (具体名で部門長承認)
- 管理表xlsxから「_25件対応_」表記も削除
"""
import csv
import os
import sys
import io
import warnings
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings('ignore')
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

ROOT = Path(r"C:\Users\nyham\work\demo_data")
ITGC_DIR = ROOT / "4.evidence" / "ITGC"


# ==============================================================
# 1. SoD マトリクス作成
# ==============================================================
def create_sod_matrix():
    """11ロール×11ロールのSoD競合マトリクスを作成
    QM_USER は意図的に含めない (sample 22 の例外を発生させる)"""
    wb = Workbook()
    ws = wb.active
    ws.title = "SoDマトリクス"

    ws.cell(1, 1, '【SAP SoD 競合マトリクス】 / 職務分掌違反チェック表')
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(2, 1, '出典: 規程R18 第15-16条 (職務分掌の原則 / 併任禁止の職務)')
    ws.cell(3, 1, '出力日: 2026-02-15 / 出力者: IT003 加藤 洋子 (E0053) 情シス部アプリチームリーダー')
    ws.cell(4, 1, '凡例: ✗=SoD違反(付与禁止) / ✓=付与可 / ―=自身との組合せ')

    roles = [
        ('FI_USER', 'FI標準ユーザ (経理)'),
        ('GL_POST', 'GL仕訳起票権限'),
        ('AR_USER', '売掛金担当'),
        ('AP_USER', '買掛金担当'),
        ('SD_USER', '販売管理ユーザ'),
        ('MM_USER', '購買管理ユーザ'),
        ('PO_CREATE', '発注作成権限'),
        ('PO_APPROVE', '発注承認権限'),
        ('PP_SUP', '生産管理スーパーバイザ'),
        ('HR_USER', '人事ユーザ'),
        ('BASIS', '基盤管理者(特権ID)'),
    ]

    # SoD conflict pairs (symmetrical)
    conflicts = {
        # 受注・売上 関連
        ('SD_USER', 'AR_USER'),     # 受注登録 ⇔ 売掛金消込
        ('SD_USER', 'GL_POST'),     # 受注 ⇔ 仕訳
        # 購買 関連
        ('PO_CREATE', 'PO_APPROVE'),# 発注作成 ⇔ 発注承認 (R18第16条)
        ('MM_USER', 'AP_USER'),     # 購買 ⇔ 買掛
        ('MM_USER', 'GL_POST'),     # 購買 ⇔ 仕訳
        ('PO_CREATE', 'AP_USER'),   # 発注作成 ⇔ 買掛
        ('PO_APPROVE', 'AP_USER'),  # 発注承認 ⇔ 買掛
        # 経理 関連
        ('AR_USER', 'AP_USER'),     # 売掛 ⇔ 買掛 (兼任不可)
        ('GL_POST', 'AR_USER'),
        ('GL_POST', 'AP_USER'),
        # FI関連
        ('FI_USER', 'BASIS'),       # 業務 ⇔ 基盤
        # BASIS は他全業務系と競合
        ('BASIS', 'SD_USER'),
        ('BASIS', 'MM_USER'),
        ('BASIS', 'PO_CREATE'),
        ('BASIS', 'PO_APPROVE'),
        ('BASIS', 'PP_SUP'),
        ('BASIS', 'HR_USER'),
        ('BASIS', 'GL_POST'),
        ('BASIS', 'AR_USER'),
        ('BASIS', 'AP_USER'),
        # 製造関連
        ('PP_SUP', 'GL_POST'),
        # HR
        ('HR_USER', 'GL_POST'),
    }
    # Make symmetrical
    sym_conflicts = set()
    for a, b in conflicts:
        sym_conflicts.add((a, b))
        sym_conflicts.add((b, a))

    # Header row (column titles)
    header_fill = PatternFill('solid', fgColor='305496')
    cell_font_white = Font(bold=True, color='FFFFFF')

    start_row = 6
    start_col = 2

    # Top-left corner
    ws.cell(start_row, 1, 'ロール ＼ 競合相手')
    ws.cell(start_row, 1).font = cell_font_white
    ws.cell(start_row, 1).fill = header_fill
    ws.cell(start_row, 1).alignment = Alignment(horizontal='center', vertical='center')

    # Column headers
    for i, (code, _) in enumerate(roles):
        c = start_col + i
        ws.cell(start_row, c, code)
        ws.cell(start_row, c).font = cell_font_white
        ws.cell(start_row, c).fill = header_fill
        ws.cell(start_row, c).alignment = Alignment(horizontal='center', vertical='center')

    # Row headers + matrix
    thin = Side(border_style='thin', color='888888')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (row_code, row_desc) in enumerate(roles):
        r = start_row + 1 + i
        # Row label
        ws.cell(r, 1, f'{row_code}\n({row_desc})')
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 1).fill = PatternFill('solid', fgColor='D6E0F0')
        ws.cell(r, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(r, 1).border = border

        for j, (col_code, _) in enumerate(roles):
            c = start_col + j
            if i == j:
                value = '―'
                fill = PatternFill('solid', fgColor='E0E0E0')
            elif (row_code, col_code) in sym_conflicts:
                value = '✗'
                fill = PatternFill('solid', fgColor='F8CBAD')  # 薄赤
            else:
                value = '✓'
                fill = PatternFill('solid', fgColor='E2F0D9')  # 薄緑
            cell = ws.cell(r, c, value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = fill
            cell.border = border
            cell.font = Font(bold=True, size=12)

    # Notes
    notes_row = start_row + len(roles) + 3
    ws.cell(notes_row, 1, '【注記】')
    ws.cell(notes_row, 1).font = Font(bold=True)
    notes = [
        '・本マトリクスは規程R18第15-16条に基づく職務分掌違反パターンを定義する',
        '・新規ロール付与時、SAPワークフロー(S04)が本マトリクスを参照して自動チェックを行う',
        '・✗(競合)が発見された場合、付与申請は自動でリジェクトされる',
        '・本マトリクスに未登録のロールは事前審査の対象 (情シス部アプリチームリーダーが個別判定)',
        '・本マトリクスは情シス部長(E0051)が四半期ごとにレビューし、必要に応じて改訂',
    ]
    for i, n in enumerate(notes, 1):
        ws.cell(notes_row + i, 1, n)

    # Column widths
    ws.column_dimensions['A'].width = 22
    for j in range(len(roles)):
        ws.column_dimensions[chr(65 + start_col - 1 + j)].width = 12

    # Row heights
    for i in range(len(roles)):
        ws.row_dimensions[start_row + 1 + i].height = 28

    out_path = ITGC_DIR / "SAP_SoD_ConflictMatrix_FY2025.xlsx"
    wb.save(out_path)
    print(f"[Created] SAP_SoD_ConflictMatrix_FY2025.xlsx ({len(roles)}ロール × {len(roles)}ロール / 意図的に QM_USER は未登録)")


# ==============================================================
# 2. Sample 22 のロール変更 (QM_USER) - SU01 + xlsx管理表 + PDF再生成
# ==============================================================
INTENTIONAL_EXCEPTION_SNO = '22'
INTENTIONAL_EXCEPTION_ROLE = 'QM_USER'  # マトリクス外


def update_su01_csv():
    """SU01 CSV: sample 22 → QM_USER, sample 11 → MM_USER (確認済み)"""
    path = ITGC_DIR / "SAP_SU01_UserMaster_ChangeHistory_FY2025.csv"
    with open(path, encoding='utf-8') as f:
        lines = f.readlines()

    new_lines = []
    for line in lines:
        if line.startswith('#') or line.startswith('タイムスタンプ') or not line.strip():
            new_lines.append(line.rstrip('\r\n'))
            continue
        parts = line.strip().split(',')
        if len(parts) >= 10:
            sno = parts[1]
            if sno == INTENTIONAL_EXCEPTION_SNO:
                parts[6] = INTENTIONAL_EXCEPTION_ROLE
            new_lines.append(','.join(parts))
        else:
            new_lines.append(line.rstrip('\r\n'))

    with open(path, 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(new_lines))
    print(f"[Fixed] SU01 CSV: sample {INTENTIONAL_EXCEPTION_SNO} → {INTENTIONAL_EXCEPTION_ROLE} (マトリクス外)")


# ==============================================================
# 3. Sample List xlsx 修正
# ==============================================================
def update_sample_list_xlsx():
    """管理表xlsx:
    - Sample 11: MM_USER,PO_CREATE → MM_USER
    - Sample 22: SD_USER → QM_USER
    - "_25件対応_" 表記を削除
    """
    path = ITGC_DIR / "UserRegistration_SampleTransactionList_FY2025.xlsx"
    wb = load_workbook(path)
    ws = wb.active

    # Remove _25件対応_
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and '_25件対応_' in v:
                new_v = v.replace('ITGC-AC-001_25件対応_RAW_*.csv', 'SAP_SU01_UserMaster_ChangeHistory_FY2025.csv')
                new_v = new_v.replace('_25件対応_', '_')
                ws.cell(r, c, new_v)

    # Update sample 11 and 22 roles
    for r in range(10, ws.max_row + 1):
        sno_cell = ws.cell(r, 1).value
        if sno_cell == 11:
            # Column 6 is 付与ロール
            ws.cell(r, 6, 'MM_USER')
        elif sno_cell == 22:
            ws.cell(r, 6, INTENTIONAL_EXCEPTION_ROLE)

    wb.save(path)
    print(f"[Fixed] 管理表xlsx: sample 11 → MM_USER / sample 22 → {INTENTIONAL_EXCEPTION_ROLE} / _25件対応_ 表記削除")


# ==============================================================
# 4. Samples 1-5 + 22 PDF再生成 (with specific 部門長 names)
# ==============================================================
def regenerate_pdfs():
    """Sample 1-5 と 22 のPDFを再生成 (具体名 + マトリクス外ロール)"""
    from fpdf import FPDF

    # Read SU01 CSV for current data
    su01_info = {}
    with open(ITGC_DIR / "SAP_SU01_UserMaster_ChangeHistory_FY2025.csv", encoding='utf-8') as f:
        for line in f:
            if line.startswith('#') or line.startswith('タイムスタンプ') or not line.strip():
                continue
            parts = line.strip().split(',')
            if len(parts) >= 10:
                sno = parts[1]
                su01_info[sno] = {
                    'reqno': parts[2], 'uid': parts[3], 'dept': parts[5], 'role': parts[6]
                }

    # Read WF for approval names
    wf_info = {}
    with open(ITGC_DIR / "Workflow_UserRegistration_ApprovalHistory_FY2025.csv", encoding='utf-8') as f:
        for line in f:
            if line.startswith('#') or line.startswith('タイムスタンプ') or not line.strip():
                continue
            parts = line.strip().split(',')
            if len(parts) < 7: continue
            sno = parts[2]
            if sno not in wf_info:
                wf_info[sno] = {'head': None, 'head_date': None, 'it': None, 'it_date': None, 'apl_date': None}
            ts = parts[0]
            actor = parts[4]
            action = parts[6]
            if action == '起票':
                wf_info[sno]['apl_date'] = ts[:10]
            elif action == '承認' and 'IT003' not in actor and '部門長' not in actor:
                # 部門長承認 (具体名)
                wf_info[sno]['head'] = actor
                wf_info[sno]['head_date'] = ts[:10]
            elif action == '承認' and 'IT003' in actor:
                wf_info[sno]['it'] = actor
                wf_info[sno]['it_date'] = ts[:10]

    role_desc = {
        'SD_USER': '販売管理標準ユーザ',
        'PP_SUP': '生産管理スーパーバイザ',
        'MM_USER': '購買管理標準ユーザ',
        'FI_USER': '会計標準ユーザ',
        'HR_USER': '人事標準ユーザ',
        'BASIS': '基盤管理者ロール',
        'QM_USER': '品質管理ユーザ (新設)',
    }

    FONT_REG = r"C:\Windows\Fonts\YuGothM.ttc"
    FONT_BLD = r"C:\Windows\Fonts\YuGothB.ttc"

    targets = ['1', '2', '3', '4', '5', INTENTIONAL_EXCEPTION_SNO]
    for sno in targets:
        if sno not in su01_info or sno not in wf_info:
            continue
        s = su01_info[sno]
        w = wf_info[sno]

        reqno = s['reqno']
        pdf_path = ITGC_DIR / f"ユーザ登録申請書_{reqno}.pdf"

        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.add_page()
        pdf.add_font('YG', '', FONT_REG, uni=True)
        pdf.add_font('YGB', '', FONT_BLD, uni=True)

        pdf.set_font('YGB', '', 18)
        pdf.cell(0, 12, 'SAPユーザ登録申請書', ln=1)

        pdf.set_font('YG', '', 10)
        pdf.cell(0, 6, f'申請番号: {reqno} / 申請日: {w["apl_date"] or "-"}', ln=1, align='R')
        pdf.ln(3)

        pdf.set_font('YG', '', 10)
        pdf.cell(35, 8, '申請部門', border=1, align='C')
        pdf.cell(140, 8, s['dept'], border=1)
        pdf.ln()
        pdf.cell(35, 8, '申請者', border=1, align='C')
        pdf.cell(140, 8, '申請者（部門）', border=1)
        pdf.ln()
        pdf.cell(35, 8, '登録対象者', border=1, align='C')
        pdf.cell(140, 8, f'{s["uid"]} (新入社員/異動)', border=1)
        pdf.ln()
        pdf.cell(35, 8, '申請理由', border=1, align='C')
        pdf.cell(140, 8, '業務遂行に必要なアクセス権付与', border=1)
        pdf.ln()
        pdf.ln(5)

        # Roles
        pdf.set_font('YGB', '', 12)
        pdf.cell(0, 8, '1. 付与希望ロール', ln=1)
        pdf.set_font('YG', '', 10)
        pdf.set_fill_color(48, 84, 150)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(40, 8, 'ロール名', border=1, align='C', fill=True)
        pdf.cell(60, 8, '内容', border=1, align='C', fill=True)
        pdf.cell(75, 8, '業務上の必要性', border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)
        role = s['role']
        pdf.cell(40, 8, role, border=1)
        pdf.cell(60, 8, role_desc.get(role.split('+')[0], '業務機能'), border=1)
        pdf.cell(75, 8, '日常業務のため', border=1)
        pdf.ln()
        pdf.ln(5)

        # SoD - QM_USER は特記事項
        pdf.set_font('YGB', '', 12)
        pdf.cell(0, 8, '2. 職務分掌(SoD)チェック', ln=1)
        pdf.set_font('YG', '', 10)
        if role == 'QM_USER':
            pdf.set_text_color(200, 0, 0)
            pdf.multi_cell(0, 6, '※ 本ロール(QM_USER)は当社SoDマトリクス未登録のため、システム自動チェック対象外。情シス部アプリチームリーダーによる事前審査で SoD 違反なしと判定。')
            pdf.set_text_color(0, 0, 0)
        else:
            pdf.cell(0, 6, '付与予定ロールの組合せについてSoD違反なし。', ln=1)
        pdf.ln(3)

        # Approval
        pdf.set_font('YGB', '', 12)
        pdf.cell(0, 8, '■ 承認経路', ln=1)
        pdf.set_font('YG', '', 10)
        pdf.set_fill_color(48, 84, 150)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(50, 8, '役割', border=1, align='C', fill=True)
        pdf.cell(60, 8, '氏名', border=1, align='C', fill=True)
        pdf.cell(40, 8, '承認日時', border=1, align='C', fill=True)
        pdf.cell(25, 8, '承認印', border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)

        pdf.cell(50, 10, '申請部門長', border=1, align='C')
        head_name = w['head'] or '部門長'
        pdf.cell(60, 10, head_name, border=1, align='C')
        pdf.cell(40, 10, w['head_date'] or '-', border=1, align='C')
        pdf.set_text_color(200, 0, 0)
        pdf.cell(25, 10, '承認', border=1, align='C')
        pdf.set_text_color(0, 0, 0)
        pdf.ln()

        pdf.cell(50, 10, '情シス部アプリリーダー', border=1, align='C')
        pdf.cell(60, 10, w['it'] or '加藤 洋子 (IT003)', border=1, align='C')
        pdf.cell(40, 10, w['it_date'] or '-', border=1, align='C')
        pdf.set_text_color(200, 0, 0)
        pdf.cell(25, 10, '承認', border=1, align='C')
        pdf.set_text_color(0, 0, 0)
        pdf.ln()

        pdf.output(str(pdf_path))

    print(f"[Regenerated] {len(targets)} PDFs (samples 1-5 + 22)")


# ==============================================================
# 5. Update Evidence_Mapping_ITGC.csv
# ==============================================================
def update_mapping():
    path = ROOT / "2.RCM" / "Evidence_Mapping_ITGC.csv"

    rows = []
    with open(path, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            if len(row) >= 3 and row[0]:
                rows.append(row)

    # Add SoD matrix entry
    new_entry = ['ITGC-AC-001', '1', 'SAP_SoD_ConflictMatrix_FY2025.xlsx']
    if new_entry not in rows:
        rows.append(new_entry)

    # Sort
    order = ['ITGC-AC-001', 'ITGC-AC-002', 'ITGC-AC-003', 'ITGC-AC-004',
             'ITGC-CM-001', 'ITGC-CM-002', 'ITGC-CM-003',
             'ITGC-EM-001', 'ITGC-OM-001', 'ITGC-OM-002']
    rows.sort(key=lambda x: (order.index(x[0]) if x[0] in order else 99, x[2]))

    with open(path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL)
        writer.writerow(header)
        for r in rows:
            writer.writerow(r)
    print(f"[Fixed] Evidence_Mapping_ITGC.csv: SoDマトリクス追加")


if __name__ == '__main__':
    create_sod_matrix()
    update_su01_csv()
    update_sample_list_xlsx()
    regenerate_pdfs()
    update_mapping()
    print("\n=== ITGC-AC-001 V2 FIXES COMPLETED ===")
