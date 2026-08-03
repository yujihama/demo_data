# -*- coding: utf-8 -*-
"""連結決算ダミーデータ生成用の openpyxl スタイルヘルパー。"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
BLUE = "2F5597"
LIGHT = "D6E4F0"
GRAY = "F2F2F2"
YELLOW = "FFF2CC"
ORANGE = "FCE4D6"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

F_TITLE = Font(name="Yu Gothic", size=14, bold=True, color=NAVY)
F_SUB = Font(name="Yu Gothic", size=10, color="595959")
F_HEAD = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
F_BODY = Font(name="Yu Gothic", size=10)
F_BOLD = Font(name="Yu Gothic", size=10, bold=True)
F_SMALL = Font(name="Yu Gothic", size=9, color="595959")

FILL_HEAD = PatternFill("solid", fgColor=BLUE)
FILL_SECTION = PatternFill("solid", fgColor=LIGHT)
FILL_TOTAL = PatternFill("solid", fgColor=GRAY)
FILL_WARN = PatternFill("solid", fgColor=YELLOW)

NUM = "#,##0;[Red]△#,##0"
NUM2 = "#,##0.00;[Red]△#,##0.00"


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title(ws, row, text, subtitle=None, ncols=6):
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_TITLE
    if subtitle:
        s = ws.cell(row=row + 1, column=1, value=subtitle)
        s.font = F_SUB
        return row + 3
    return row + 2


def header_row(ws, row, headers, fill=FILL_HEAD):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = F_HEAD
        c.fill = fill
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return row + 1


def data_row(ws, row, values, numcols=(), bold=False, fill=None, indent=None):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = F_BOLD if bold else F_BODY
        c.border = BORDER
        if fill:
            c.fill = fill
        if i in numcols:
            c.number_format = NUM
            c.alignment = Alignment(horizontal="right")
        elif i == 1 and indent:
            c.alignment = Alignment(horizontal="left", indent=indent)
    return row + 1


def kv_block(ws, row, pairs, key_width_col=1):
    """表紙用の項目:値ブロック"""
    for k, v in pairs:
        ck = ws.cell(row=row, column=1, value=k)
        ck.font = F_BOLD
        ck.fill = FILL_SECTION
        ck.border = BORDER
        cv = ws.cell(row=row, column=2, value=v)
        cv.font = F_BODY
        cv.border = BORDER
        cv.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    return row


def note(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_SMALL
    c.alignment = Alignment(wrap_text=True, vertical="top")
    return row + 1
