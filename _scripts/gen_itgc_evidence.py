"""
ITGC（IT全般統制）エビデンス生成
【真の不備：ITGC-AC-003 退職者アカウント停止遅延2件】を含む
4ドメイン: AC(アクセス)/CM(変更)/OM(運用)/EM(外部委託)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
from pathlib import Path
import random
import sys

sys.path.insert(0, str(Path(__file__).parent))
from pdf_util import JPPDF
from image_util import sap_screenshot, workflow_screenshot, table_image

BASE_AC = Path(r"C:\Users\nyham\work\demo_data\4.evidence\ITGC\AC_アクセス管理")
BASE_CM = Path(r"C:\Users\nyham\work\demo_data\4.evidence\ITGC\CM_変更管理")
BASE_OM = Path(r"C:\Users\nyham\work\demo_data\4.evidence\ITGC\OM_運用管理")
BASE_EM = Path(r"C:\Users\nyham\work\demo_data\4.evidence\ITGC\EM_外部委託管理")
for p in [BASE_AC, BASE_CM, BASE_OM, BASE_EM]:
    p.mkdir(parents=True, exist_ok=True)

HF = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
BFONT = Font(name="Yu Gothic", size=10)
BBOLD = Font(name="Yu Gothic", size=10, bold=True)
C_ = Alignment(horizontal="center", vertical="center", wrap_text=True)
L_ = Alignment(horizontal="left", vertical="center", wrap_text=True)
R_ = Alignment(horizontal="right", vertical="center")
T_ = Side("thin", color="888888")
BRD = Border(left=T_, right=T_, top=T_, bottom=T_)
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")
FILL_NG = PatternFill("solid", fgColor="FCE4D6")
FILL_OK = PatternFill("solid", fgColor="E2EFDA")


# ============================================================
# AC-001 新規ユーザ登録申請（PDF）
# ============================================================
def gen_ac_001_user_application():
    pdf = JPPDF()
    pdf.add_page()

    pdf.h1("SAP ユーザ登録申請書")
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, "申請番号: USER-REG-2025-0087 / 申請日: 2025/11/10",
             align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.kv("申請部門", "営業本部")
    pdf.kv("申請者", "斎藤 次郎（営業部課長 SLS002）")
    pdf.kv("登録対象者", "新入社員 SLS006（2025/11採用）")
    pdf.kv("申請理由", "新入社員配属に伴う通常のアクセス権付与")
    pdf.ln(5)

    pdf.h2("1. 付与希望ロール")
    pdf.table_header(["ロール名", "内容", "業務上の必要性"], [30, 60, 90])
    pdf.table_row(["SD_USER", "販売管理モジュール利用者",
                   "受注登録・顧客情報照会のため"], [30, 60, 90])
    pdf.ln(3)

    pdf.h2("2. 職務分掌(SoD)チェック")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
                   "・SD_USER単独のため、SoD違反なし\n"
                   "・他の購買・経理関連ロールは付与しない")
    pdf.ln(5)

    pdf.h3("■ 承認経路")
    pdf.set_font("YuGoth", "B", 10)
    pdf.cell(45, 7, "役割", border=1, align="C", fill=True)
    pdf.cell(55, 7, "氏名", border=1, align="C", fill=True)
    pdf.cell(40, 7, "承認日時", border=1, align="C", fill=True)
    pdf.cell(30, 7, "承認印", border=1, align="C", fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("YuGoth", "", 10)
    approvals = [
        ("申請部門長", "田中 太郎 (SLS001)", "2025/11/10 15:32"),
        ("情シス部アプリリーダー", "加藤 洋子 (IT003)", "2025/11/11 09:15"),
    ]
    for role, name, dt in approvals:
        pdf.cell(45, 14, role, border=1, align="C")
        pdf.cell(55, 14, name, border=1, align="C")
        pdf.cell(40, 14, dt, border=1, align="C")
        x_stamp = pdf.get_x()
        y_stamp = pdf.get_y()
        pdf.cell(30, 14, "", border=1, new_x="LMARGIN", new_y="NEXT")
        pdf.stamp("承認", x=x_stamp + 15, y=y_stamp + 7)

    pdf.ln(3)
    pdf.set_font("YuGoth", "", 9)
    pdf.cell(0, 5, "SAP登録: 2025/11/11 10:30 / 登録者: IT004 西田 徹 / ユーザID: SLS006",
             new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(BASE_AC / "ITGC-AC-001_ユーザ登録申請書_USER-REG-2025-0087.pdf"))
    print("Created: ITGC-AC-001_ユーザ登録申請書_USER-REG-2025-0087.pdf")


# ============================================================
# AC-002 SAP SUIM ユーザ一覧 & アクセス権棚卸
# ============================================================
def gen_ac_002_suim_report():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SUIM_ユーザ一覧"

    ws.cell(row=1, column=1, value="SAP SUIM / トランザクション: SUIM_REPT / 有効ユーザ一覧")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=2, column=1, value="出力日時: 2025/12/05 10:15:42 (Q3定期棚卸用)")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.cell(row=3, column=1, value="出力者: IT003 加藤 洋子 / 抽出条件: ユーザステータス=有効 / Client=100 本番")
    ws.cell(row=3, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)

    headers = ["ユーザID", "氏名", "所属部門", "付与ロール", "有効期限",
               "最終ログイン", "作成日", "備考"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    users = [
        ("CEO001", "山本 健一", "代表取締役", "ALL_READ", "無期限", date(2025, 12, 4), date(2010, 4, 1), ""),
        ("CFO001", "渡辺 正博", "管理本部", "FI_MGR, CO_MGR, ALL_READ", "無期限", date(2025, 12, 4), date(2010, 4, 1), ""),
        ("ACC001", "佐藤 一郎", "経理部", "FI_MGR, CO_MGR", "無期限", date(2025, 12, 5), date(2010, 4, 1), ""),
        ("ACC002", "高橋 美咲", "経理部", "FI_SUP, GL_POST", "無期限", date(2025, 12, 5), date(2010, 4, 1), ""),
        ("ACC003", "伊藤 健太", "経理部", "CO_SUP", "無期限", date(2025, 12, 5), date(2010, 4, 1), ""),
        ("ACC004", "中村 真理", "経理部", "FI_USER, GL_POST", "無期限", date(2025, 12, 4), date(2015, 4, 1), ""),
        ("SLS001", "田中 太郎", "営業本部", "SD_MGR", "無期限", date(2025, 12, 5), date(2010, 4, 1), ""),
        ("SLS002", "斎藤 次郎", "営業本部", "SD_SUP", "無期限", date(2025, 12, 5), date(2010, 4, 1), ""),
        ("SLS004", "松本 香織", "営業本部", "SD_USER", "無期限", date(2025, 12, 5), date(2014, 4, 1), ""),
        ("SLS005", "井上 大輔", "営業本部", "SD_USER", "無期限", date(2025, 12, 4), date(2018, 4, 1), ""),
        ("PUR001", "木村 浩二", "購買部", "MM_MGR, PO_APPROVE", "無期限", date(2025, 12, 5), date(2010, 4, 1), ""),
        ("PUR002", "林 真由美", "購買部", "MM_SUP, PO_APPROVE", "無期限", date(2025, 12, 5), date(2010, 4, 1), ""),
        ("PUR003", "清水 智明", "購買部", "MM_USER, PO_CREATE", "無期限", date(2025, 12, 5), date(2013, 4, 1), ""),
        ("PUR004", "山田 純一", "購買部", "MM_USER, PO_CREATE, PO_APPROVE",
         "無期限", date(2025, 12, 5), date(2017, 4, 1),
         "※ SoD違反: PO_CREATE+PO_APPROVE (要改善)"),
        ("IT001", "岡田 宏", "情シス部", "BASIS, ALL_READ", "無期限", date(2025, 12, 5), date(2010, 4, 1), "特権ID"),
        ("IT003", "加藤 洋子", "情シス部", "DEVELOPER", "無期限", date(2025, 12, 5), date(2012, 4, 1), ""),
        # 退職者が残存（不備ケース）
        ("SLS099", "退職者A", "(退職)営業本部", "SD_USER", "無期限",
         date(2025, 9, 30), date(2018, 4, 1),
         "※ 退職2025/9/30、停止2025/10/11（11日遅延）— 不備"),
        ("PUR099", "退職者B", "(退職)購買部", "MM_USER", "無期限",
         date(2025, 11, 15), date(2017, 4, 1),
         "※ 退職2025/11/15、停止2025/12/3（18日遅延）— 不備"),
    ]
    r = 6
    for row in users:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 3, 5, 6, 7):
                cell.alignment = C_
                if c_i in (6, 7):
                    cell.number_format = "yyyy/mm/dd"
            else:
                cell.alignment = L_
        if row[7] and ("違反" in row[7] or "不備" in row[7]):
            for c_i in range(1, 9):
                ws.cell(row=r, column=c_i).fill = FILL_NG
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="※ 本一覧は四半期定期棚卸のため出力。全ユーザ245名中、経理・営業・購買・情シス部の18名を抜粋。").font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    # ------ 棚卸結果シート ------
    ws2 = wb.create_sheet("棚卸結果")
    ws2.cell(row=1, column=1, value="Q3アクセス権棚卸結果（2025/12実施）")
    ws2.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    ws2.cell(row=2, column=1, value="※ 本棚卸はQ3（2025/12）分 / 完了日: 2025/12/22").font = Font(name="Yu Gothic", size=9, italic=True)
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws2.cell(row=3, column=1,
             value="⚠ 本エビデンスだけでは抽出条件・抽出日時の完全性が不明（内部監査指摘事項）"
             ).font = Font(name="Yu Gothic", size=9, italic=True, color="C00000")
    ws2.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)

    hd = ["部門", "対象ユーザ数", "継続必要", "削除依頼", "未レビュー", "部門長承認"]
    for i, h in enumerate(hd, 1):
        c = ws2.cell(row=5, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    dept_rev = [
        ("経理部", 8, 8, 0, 0, "佐藤 一郎 [印] 2025/12/18"),
        ("営業本部", 45, 43, 2, 0, "田中 太郎 [印] 2025/12/19"),
        ("購買部", 8, 7, 1, 0, "木村 浩二 [印] 2025/12/17"),
        ("製造本部", 62, 60, 2, 0, "森 和雄 [印] 2025/12/20"),
        ("情シス部", 15, 15, 0, 0, "岡田 宏 [印] 2025/12/15"),
        ("その他", 107, 105, 2, 0, "各部門長 [印]"),
    ]
    rr = 6
    for row in dept_rev:
        for c_i, v in enumerate(row, 1):
            cell = ws2.cell(row=rr, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 5):
                cell.alignment = C_
            else:
                cell.alignment = L_
        rr += 1

    widths = [10, 14, 14, 30, 12, 14, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A6"

    wb.save(BASE_AC / "ITGC-AC-002_SAP_SUIM_有効ユーザ一覧_Q3棚卸用.xlsx")
    print("Created: ITGC-AC-002_SAP_SUIM_有効ユーザ一覧_Q3棚卸用.xlsx")


# ============================================================
# AC-003 退職者一覧と停止記録（不備ケース含む）
# ============================================================
def gen_ac_003_retiree_log():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "退職者アカウント停止記録"

    ws.cell(row=1, column=1, value="【ITGC-AC-003 統制実施記録】 FY2025 退職者アカウント停止記録")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    ws.cell(row=2, column=1, value="管理者: 情シス部 西田 徹 (IT004) / 最終更新: 2026/2/20")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    ws.cell(row=3, column=1, value="規程: 退職日の3営業日以内にSAPアカウントを停止する（R24 §5-2）")
    ws.cell(row=3, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=9)

    headers = ["社員番号", "氏名", "所属部門", "退職日", "停止依頼日",
               "実際停止日", "遅延日数", "判定", "備考"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    records = [
        ("E0095", "田中 一郎", "製造部", date(2025, 6, 30), date(2025, 6, 25),
         date(2025, 6, 30), 0, "合格", "退職当日に停止完了"),
        ("E0094", "山口 次郎", "経理部", date(2025, 7, 31), date(2025, 7, 28),
         date(2025, 7, 31), 0, "合格", "退職当日に停止完了"),
        ("E0099", "退職者A", "営業本部", date(2025, 9, 30), date(2025, 9, 28),
         date(2025, 10, 11), 11, "不合格",
         "※不備: 情シス部繁忙のため対応遅延。ログインなし（SM19確認）"),
        ("E0098", "退職者B", "購買部", date(2025, 11, 15), date(2025, 11, 12),
         date(2025, 12, 3), 18, "不合格",
         "※不備: 依頼メール見落とし。ログインなし（SM19確認）"),
        ("E0093", "井上 三郎", "人事部", date(2026, 1, 31), date(2026, 1, 28),
         date(2026, 1, 31), 0, "合格", "退職当日に停止完了"),
    ]
    r = 6
    for row in records:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 3, 4, 5, 6, 7, 8):
                cell.alignment = C_
                if c_i in (4, 5, 6):
                    cell.number_format = "yyyy/mm/dd"
            else:
                cell.alignment = L_
        if row[7] == "不合格":
            for c_i in range(1, 10):
                ws.cell(row=r, column=c_i).fill = FILL_NG
        else:
            ws.cell(row=r, column=8).fill = FILL_OK
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="■ 不備への対応").font = BBOLD
    r += 1
    for line in [
        "・退職者A (E0099): 2025/10/11 SAP停止完了。該当期間中のログイン履歴なし（SM19ログ確認済）。",
        "・退職者B (E0098): 2025/12/3 SAP停止完了。該当期間中のログイン履歴なし（SM19ログ確認済）。",
        "・再発防止策: 人事SaaS(S03)とSAPのID連携自動化を2026/Q1に実装予定（ITGC-CM経由）。",
        "・情シス部での退職者停止チケットを最優先扱いに変更（2025/12/5実施済）。",
    ]:
        ws.cell(row=r, column=1, value=line).font = BFONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        r += 1

    widths = [12, 12, 14, 12, 12, 12, 10, 10, 35]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE_AC / "ITGC-AC-003_退職者アカウント停止記録_FY2025.xlsx")
    print("Created: ITGC-AC-003_退職者アカウント停止記録_FY2025.xlsx")


# ============================================================
# AC-003 SAP SM20 ログインログ（退職者に関する期間）
# ============================================================
def gen_ac_003_sm20_log():
    path = BASE_AC / "ITGC-AC-003_SAP_SM19_SM20_退職者ログインログ抽出.csv"
    lines = [
        "# SAP SM20 Security Audit Log 抽出",
        "# 対象ユーザ: SLS099, PUR099",
        "# 対象期間: 2025/9/30 ～ 2025/12/3 (退職～停止完了まで)",
        "# 出力日時: 2026/2/15 14:22:30",
        "# 出力者: IT002 吉田 雅彦",
        "",
        "タイムスタンプ,ユーザID,イベント,クライアント,端末IP,備考",
    ]
    # 期間中、ログインなし = 該当レコード無し のログエントリ
    lines.append("# SLS099 の2025/9/30-2025/10/11 ログインレコード: 0件")
    lines.append("# PUR099 の2025/11/15-2025/12/3 ログインレコード: 0件")
    lines.append("")
    lines.append("# 結論: 退職者アカウントは停止遅延期間中、不正利用されていないことを確認")
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"Created: {path.name}")


# ============================================================
# AC-004 特権ID操作ログ
# ============================================================
def gen_ac_004_privileged_log():
    path = BASE_AC / "ITGC-AC-004_特権ID操作ログ_202511.csv"
    random.seed(4040)
    lines = [
        "# SAP SM20 - 特権ID(BASIS/ALL_READ)操作ログ",
        "# 対象ユーザ: IT001 岡田 宏、IT002 吉田 雅彦",
        "# 対象期間: 2025/11/01 ～ 2025/11/30",
        "# 出力日時: 2025/12/05 08:30 / 出力者: IT001 岡田 宏（月次レビュー）",
        "",
        "タイムスタンプ,ユーザID,トランザクション,対象,結果,承認事前申請",
    ]

    events = [
        ("2025-11-03 09:15:22", "IT002", "RZ20", "システム監視", "正常", "恒常業務・申請不要"),
        ("2025-11-05 14:22:18", "IT001", "SM37", "バッチジョブ確認", "正常", "恒常業務"),
        ("2025-11-07 10:45:33", "IT002", "SM51", "サーバ起動停止", "正常", "申請WF-IT-2025-098"),
        ("2025-11-12 16:10:08", "IT002", "DB13", "DBバックアップ確認", "正常", "恒常業務"),
        ("2025-11-15 11:22:45", "IT001", "SM59", "RFC接続テスト", "正常", "申請WF-IT-2025-102"),
        ("2025-11-18 13:05:12", "IT002", "SE38", "ABAP実行", "正常", "申請WF-IT-2025-105"),
        ("2025-11-22 09:30:28", "IT002", "RZ10", "プロファイル確認", "正常", "恒常業務"),
        ("2025-11-25 15:48:55", "IT001", "SU01", "ユーザマスタ確認", "正常", "棚卸のため"),
        ("2025-11-28 10:12:40", "IT002", "STMS", "移送管理", "正常", "変更REL-2025-067"),
    ]
    for e in events:
        lines.append(",".join(e))

    lines.append("")
    lines.append("# 件数: 9件 / すべて承認済または恒常業務")
    lines.append("# 情シス部長レビュー: 岡田 宏 [印] 2025/12/05 / 異常なし")
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"Created: {path.name}")


# ============================================================
# AC-001 SAP SU01画面のスクリーンショット
# ============================================================
def gen_ac_screenshots():
    sap_screenshot(
        "ユーザ管理",
        "SU01",
        [
            ("ユーザID", "SLS006"),
            ("氏名", "新入 社員太郎"),
            ("所属部門", "営業本部"),
            ("付与ロール", "SD_USER"),
            ("有効期限", "無期限"),
            ("作成日", "2025/11/11"),
            ("作成者", "IT004 西田 徹"),
            ("パスワード初期化", "実施済 / 初回ログイン時変更必須"),
            ("SoDチェック", "○ 違反なし"),
        ],
        status_bar="ユーザ SLS006 が作成されました。",
        output_path=str(BASE_AC / "ITGC-AC-001_SAP_SU01_ユーザ作成画面.png"),
    )

    # アクセス権マトリクス
    table_image(
        "SAPアクセス権マトリクス（販売・購買抜粋）",
        ["ユーザID", "氏名", "SD_USER", "SD_SUP", "SD_MGR", "PO_CREATE", "PO_APPROVE", "SoD違反"],
        [
            ["SLS001", "田中 太郎", "", "", "●", "", "", ""],
            ["SLS002", "斎藤 次郎", "", "●", "", "", "", ""],
            ["SLS004", "松本 香織", "●", "", "", "", "", ""],
            ["PUR001", "木村 浩二", "", "", "", "", "●", ""],
            ["PUR002", "林 真由美", "", "", "", "", "●", ""],
            ["PUR003", "清水 智明", "", "", "", "●", "", ""],
            ["PUR004", "山田 純一", "", "", "", "●", "● ⚠", "あり"],
        ],
        widths=[70, 80, 60, 60, 60, 70, 80, 60],
        caption="⚠ PUR004 にPO_CREATE と PO_APPROVE が両方付与（SoD違反、要是正）",
        output_path=str(BASE_AC / "ITGC-AC-001_SAPアクセス権マトリクス.png"),
    )
    print("Created: 2 screenshots for AC")


# ============================================================
# CM-001/002/003 変更管理一覧
# ============================================================
def gen_cm_change_log():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "変更管理一覧"

    ws.cell(row=1, column=1, value="【ITGC-CM-001 統制実施記録】 FY2025 変更管理一覧")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    ws.cell(row=2, column=1, value="出力日時: 2026/2/18 / 出力者: IT003 加藤 洋子 / 対象: FY2025期間中のSAP変更申請")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    headers = ["REL番号", "申請日", "申請者", "変更内容概要", "影響範囲",
               "テスト実施", "本番移送日", "承認者1", "承認者2", "ステータス"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[4].height = 30

    random.seed(3210)
    users_req = ["加藤 洋子 (IT003)", "業務部門担当"]
    samples = []
    for i in range(1, 43):
        rel_no = f"REL-2025-{i:03d}"
        req_date = date(2025, random.randint(4, 12), random.randint(1, 28))
        requester = random.choice(users_req)
        changes = [
            ("販売価格マスタ連携IF修正", "販売管理"),
            ("ワークフロー承認ルーティング変更", "全業務"),
            ("標準原価計算バッチ修正", "原価計算"),
            ("SUIM定期レポート出力機能追加", "情シス内"),
            ("勘定科目マスタ追加", "経理"),
            ("仕入先マスタ項目追加（銀行口座枠拡大）", "購買"),
            ("セキュリティパッチ適用", "全体"),
            ("バックアップバッチ改善", "情シス内"),
        ]
        change, scope = random.choice(changes)
        test_done = "UAT合格" if i != 37 else "UAT省略（緊急変更手順）"
        prod_date = req_date + timedelta(days=random.randint(7, 30))
        approver1 = "加藤 洋子 (IT003)"
        approver2 = "岡田 宏 (IT001)" if random.random() < 0.3 else "業務部門長"
        status = "完了"
        samples.append((rel_no, req_date, requester, change, scope, test_done,
                        prod_date, approver1, approver2, status))

    r = 5
    for row in samples[:15]:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 5, 6, 7, 10):
                cell.alignment = C_
                if c_i in (2, 7):
                    cell.number_format = "yyyy/mm/dd"
            else:
                cell.alignment = L_
        if "省略" in row[5]:
            ws.cell(row=r, column=6).fill = FILL_WARN
        r += 1

    ws.cell(row=r, column=1, value="... 以下27件省略（全42件中） ...").font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

    widths = [14, 10, 22, 38, 14, 20, 10, 22, 22, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE_CM / "ITGC-CM-001_変更管理一覧_FY2025.xlsx")
    print("Created: ITGC-CM-001_変更管理一覧_FY2025.xlsx")


# ============================================================
# CM-001 変更申請書PDF
# ============================================================
def gen_cm_change_request():
    pdf = JPPDF()
    pdf.add_page()

    pdf.h1("SAP変更申請書")
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, "REL番号: REL-2025-023 / 申請日: 2025年7月14日",
             align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.kv("件名", "販売価格マスタ連携IF修正", key_w=30)
    pdf.kv("申請者", "加藤 洋子（情シス部アプリリーダー IT003）", key_w=30)
    pdf.kv("依頼部門", "経理部 / 中村 真理（ACC004）", key_w=30)
    pdf.kv("影響範囲", "販売管理モジュール", key_w=30)
    pdf.kv("リスクレベル", "低（IF内部処理の修正のみ）", key_w=30)
    pdf.ln(5)

    pdf.h2("1. 変更理由")
    pdf.body("月末の価格マスタ一括反映バッチにおいて、顧客別価格（VKP3）の反映で"
             "稀に競合エラーが発生する事象があり、再試行ロジックを追加する。"
             "月次で3-5件発生しており、都度手動再実行で対応しているが業務負荷が高い。")
    pdf.ln(3)

    pdf.h2("2. 変更内容")
    pdf.body("・ABAPプログラム ZSD_PRICE_UPDATE に、DB更新時の排他エラー時の再試行ロジック（最大3回）を追加\n"
             "・エラー発生時のログ出力を詳細化")
    pdf.ln(3)

    pdf.h2("3. テスト計画")
    pdf.table_header(["テストフェーズ", "テスト期間", "担当", "結果"], [40, 40, 55, 45])
    pdf.table_row(["単体テスト", "2025/7/20-21", "加藤 洋子", "合格"], [40, 40, 55, 45])
    pdf.table_row(["結合テスト（開発環境）", "2025/7/22-23", "加藤 洋子", "合格"], [40, 40, 55, 45], fill=True)
    pdf.table_row(["UAT（テスト環境）", "2025/7/24-25", "中村 真理 (ACC004)", "合格"], [40, 40, 55, 45])
    pdf.ln(5)

    pdf.h3("■ 承認経路")
    pdf.set_font("YuGoth", "B", 10)
    pdf.cell(50, 7, "役割", border=1, align="C", fill=True)
    pdf.cell(55, 7, "氏名", border=1, align="C", fill=True)
    pdf.cell(40, 7, "承認日時", border=1, align="C", fill=True)
    pdf.cell(30, 7, "承認印", border=1, align="C", fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("YuGoth", "", 10)
    approvals = [
        ("情シス部アプリリーダー", "加藤 洋子 (IT003)", "2025/7/25 16:30"),
        ("業務部門責任者", "佐藤 一郎 (ACC001)", "2025/7/26 10:15"),
        ("情シス部長", "岡田 宏 (IT001)", "2025/7/26 14:45"),
    ]
    for role, name, dt in approvals:
        pdf.cell(50, 14, role, border=1, align="C")
        pdf.cell(55, 14, name, border=1, align="C")
        pdf.cell(40, 14, dt, border=1, align="C")
        x_stamp = pdf.get_x()
        y_stamp = pdf.get_y()
        pdf.cell(30, 14, "", border=1, new_x="LMARGIN", new_y="NEXT")
        pdf.stamp("承認", x=x_stamp + 15, y=y_stamp + 7)

    pdf.ln(3)
    pdf.set_font("YuGoth", "", 9)
    pdf.cell(0, 5, "本番移送日: 2025/7/28 / STMS移送番号: XXXK912345", new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(BASE_CM / "ITGC-CM-001_変更申請書_REL-2025-023.pdf"))
    print("Created: ITGC-CM-001_変更申請書_REL-2025-023.pdf")


# ============================================================
# CM-002 UAT テスト結果報告
# ============================================================
def gen_cm_uat_report():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UATテスト結果"

    ws.cell(row=1, column=1, value="【ITGC-CM-002】UATテスト結果報告書 / REL-2025-023")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    ws.cell(row=2, column=1, value="変更内容: 販売価格マスタ連携IFの再試行ロジック追加 / UAT実施期間: 2025/7/24-25")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.cell(row=3, column=1, value="実施者: 中村 真理（ACC004）/ レビュー: 佐藤 一郎（ACC001）")
    ws.cell(row=3, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)

    headers = ["ケースNo", "テストケース", "期待結果", "実施結果", "合否", "備考"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    cases = [
        (1, "正常系: 10件の価格更新", "10件とも成功", "10件とも成功", "OK", ""),
        (2, "正常系: 100件の一括更新", "100件とも成功", "100件とも成功", "OK", ""),
        (3, "異常系: 排他エラー発生 → 1回目リトライで成功", "リトライで成功", "2回目で成功", "OK", "ログに試行回数記録あり"),
        (4, "異常系: 排他エラー発生 → 3回リトライ全て失敗", "エラー通知", "エラー通知され、ログ出力", "OK", ""),
        (5, "境界系: 連続100件の同一顧客更新", "全件成功", "全件成功", "OK", ""),
        (6, "性能系: 5000件の一括実行", "5分以内", "4分32秒", "OK", ""),
    ]
    r = 6
    for row in cases:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 5):
                cell.alignment = C_
            else:
                cell.alignment = L_
        ws.cell(row=r, column=5).fill = FILL_OK
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="結論: 全ケース合格 / 本番リリース承認").font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1).fill = FILL_OK
    r += 1
    ws.cell(row=r, column=1, value="UAT合格署名: 中村 真理 [印] 2025/7/25 / 佐藤 一郎 [印] 2025/7/26")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

    widths = [8, 35, 30, 30, 8, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE_CM / "ITGC-CM-002_UATテスト結果_REL-2025-023.xlsx")
    print("Created: ITGC-CM-002_UATテスト結果_REL-2025-023.xlsx")


# ============================================================
# CM-003 SAP STMS 移送記録
# ============================================================
def gen_cm_stms_log():
    path = BASE_CM / "ITGC-CM-003_SAP_STMS_本番移送記録_FY2025Q2-Q3.csv"
    random.seed(9876)
    lines = [
        "# SAP STMS - 本番移送履歴",
        "# 出力日時: 2026/02/18 11:22:35",
        "# 出力者: IT003 加藤 洋子",
        "# 対象期間: FY2025 Q2-Q3 (2025/7/1 - 2025/12/31)",
        "",
        "タイムスタンプ,TR番号,REL番号,移送者,移送元クライアント,移送先クライアント,対象オブジェクト,結果",
    ]
    for i in range(1, 20):
        ts = datetime(2025, random.randint(7, 12), random.randint(1, 28),
                      random.randint(8, 19), random.randint(0, 59), random.randint(0, 59))
        tr_no = f"XXXK{random.randint(900000, 999999)}"
        rel_no = f"REL-2025-{random.randint(15, 40):03d}"
        lines.append(f"{ts.strftime('%Y-%m-%d %H:%M:%S')},{tr_no},{rel_no},IT003 加藤 洋子,DEV,PRD,"
                     f"ABAP/Function Module,成功")

    lines.append("")
    lines.append("# 件数: 19件 / 全件成功 / 移送者は情シス部アプリリーダーのみ")
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"Created: {path.name}")


# ============================================================
# OM-001 バックアップ実施記録
# ============================================================
def gen_om_backup_log():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "バックアップ実施記録"

    ws.cell(row=1, column=1, value="【ITGC-OM-001 統制実施記録】 SAP HANA バックアップ実施記録 (2025年11月)")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=2, column=1, value="管理者: IT002 吉田 雅彦（インフラチームリーダー）/ 承認: IT001 岡田 宏 / 更新: 日次")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headers = ["日付", "バックアップ種別", "開始時刻", "終了時刻", "サイズ(GB)",
               "保管先", "ステータス", "備考"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    random.seed(1122)
    r = 5
    for day in range(1, 31):
        bk_date = date(2025, 11, day)
        start = f"01:00:0{random.randint(0, 9)}"
        dur_min = random.randint(75, 130)
        end_h = 1 + dur_min // 60
        end_m = dur_min % 60
        end = f"0{end_h}:{end_m:02d}:00"
        size = round(random.uniform(1800, 2100), 1)
        storage = "テープ+S3クラウド" if day % 7 == 0 else "テープのみ"
        status = "成功"
        remark = ""
        data = [bk_date, "フルバックアップ", start, end, size, storage, status, remark]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 7):
                cell.alignment = C_
                if c_i == 1:
                    cell.number_format = "yyyy/mm/dd"
            elif c_i == 5:
                cell.alignment = R_
                cell.number_format = "0.0"
            else:
                cell.alignment = L_
        ws.cell(row=r, column=7).fill = FILL_OK
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="結論: 30日間すべて成功 / 週次クラウドレプリケーション4回実施済").font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    ws.cell(row=r, column=1, value="確認: 吉田 雅彦 [印] 2025/12/01 / 岡田 宏 [印] 2025/12/02")

    widths = [12, 16, 10, 10, 12, 18, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE_OM / "ITGC-OM-001_バックアップ実施記録_202511.xlsx")
    print("Created: ITGC-OM-001_バックアップ実施記録_202511.xlsx")


# ============================================================
# OM-001 DRテスト報告書
# ============================================================
def gen_om_dr_test():
    pdf = JPPDF()
    pdf.add_page()

    pdf.h1("SAP HANA DR・リストアテスト報告書")
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, "テスト番号: DR-TEST-2025-Q3 / 実施日: 2025年9月20日（土）",
             align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.h2("1. テスト目的")
    pdf.body("SAP HANAのバックアップデータからの復旧可能性を確認する（四半期に1回実施）。"
             "業務継続計画（R26 IT-BCP）の復旧目標時間（RTO=4時間）を達成できるか検証。")
    pdf.ln(3)

    pdf.h2("2. テスト構成")
    pdf.kv("テスト環境", "DR検証用サーバ（独立）")
    pdf.kv("復旧対象", "2025/9/13（土）フルバックアップ + 差分ログ")
    pdf.kv("RTO目標", "4時間以内")
    pdf.kv("RPO目標", "24時間以内")
    pdf.ln(3)

    pdf.h2("3. テスト実行結果")
    pdf.table_header(["ステップ", "開始時刻", "終了時刻", "所要時間", "判定"],
                     [55, 30, 30, 30, 20])
    pdf.table_row(["バックアップファイル読込", "09:00", "09:35", "35分", "OK"],
                  [55, 30, 30, 30, 20])
    pdf.table_row(["HANA DBリストア", "09:35", "11:20", "1時間45分", "OK"],
                  [55, 30, 30, 30, 20], fill=True)
    pdf.table_row(["差分ログ適用", "11:20", "11:55", "35分", "OK"],
                  [55, 30, 30, 30, 20])
    pdf.table_row(["整合性チェック", "11:55", "12:25", "30分", "OK"],
                  [55, 30, 30, 30, 20], fill=True)
    pdf.table_row(["アプリケーション起動確認", "12:25", "12:40", "15分", "OK"],
                  [55, 30, 30, 30, 20])
    pdf.ln(3)

    pdf.h3("合計所要時間: 3時間40分（RTO目標4時間以内 達成）")
    pdf.ln(5)

    pdf.h2("4. 結論")
    pdf.body("DRテストは成功。RTO/RPOとも目標値を達成。次回テスト予定: 2025年12月。")
    pdf.ln(5)

    pdf.set_font("YuGoth", "", 10)
    pdf.kv("テスト実施", "吉田 雅彦 (IT002)")
    pdf.kv("立会・確認", "岡田 宏 (IT001)")
    pdf.kv("承認", "渡辺 正博 (CFO001)")

    pdf.output(str(BASE_OM / "ITGC-OM-001_DRリストアテスト報告書_2025Q3.pdf"))
    print("Created: ITGC-OM-001_DRリストアテスト報告書_2025Q3.pdf")


# ============================================================
# OM-002 障害管理台帳
# ============================================================
def gen_om_incident():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "障害管理台帳"

    ws.cell(row=1, column=1, value="【ITGC-OM-002 統制実施記録】 FY2025 SAP基盤 障害管理台帳")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    ws.cell(row=2, column=1, value="管理者: IT002 吉田 雅彦 / 承認: IT001 岡田 宏")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    headers = ["障害番号", "発生日時", "重大度", "事象概要", "対応時間",
               "原因", "対応内容", "再発防止策", "承認者"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[4].height = 32

    incidents = [
        ("INC-2025-04-003", "2025/4/12 03:15", "中",
         "夜間バッチ遅延（標準原価計算）", "2時間",
         "DBロック競合", "手動でバッチ再実行、完了確認", "ロック検知ロジック追加(REL-2025-005)", "岡田"),
        ("INC-2025-06-007", "2025/6/8 14:22", "低",
         "特定ユーザでSAP GUI接続エラー", "30分",
         "クライアント側キャッシュ破損", "クライアントPCのキャッシュクリア", "ヘルプデスク手順書に追加", "吉田"),
        ("INC-2025-08-012", "2025/8/15 10:05", "中",
         "WMS-SAP間連携停止", "1時間", "ネットワーク機器の一時障害",
         "ネットワーク機器再起動、手動同期実行", "機器冗長化検討中(2026/Q1実施予定)", "岡田"),
        ("INC-2025-10-018", "2025/10/22 16:45", "低",
         "印刷キュー詰まり", "15分", "プリンタドライバの不具合",
         "プリンタ再起動", "なし", "吉田"),
        ("INC-2025-11-021", "2025/11/7 22:30", "高",
         "HANA DBメモリ不足警告", "3時間",
         "大量データ抽出クエリの負荷", "DB再起動、問題クエリを特定し停止",
         "DB監視閾値を引き下げ、大量抽出ルールを策定", "岡田"),
    ]
    r = 5
    for inc in incidents:
        for c_i, v in enumerate(inc, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 5, 9):
                cell.alignment = C_
            else:
                cell.alignment = L_
        if inc[2] == "高":
            ws.cell(row=r, column=3).fill = FILL_NG
        elif inc[2] == "中":
            ws.cell(row=r, column=3).fill = FILL_WARN
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="合計: 期中18件発生（重大度:高1件、中6件、低11件）").font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)

    widths = [16, 16, 8, 30, 10, 22, 30, 30, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE_OM / "ITGC-OM-002_障害管理台帳_FY2025.xlsx")
    print("Created: ITGC-OM-002_障害管理台帳_FY2025.xlsx")


# ============================================================
# EM-001 SOC1レポート抜粋 & 委託先評価
# ============================================================
def gen_em_soc1():
    pdf = JPPDF()
    pdf.add_page()

    pdf.h1("ITGC-EM-001 外部委託先管理")
    pdf.set_font("YuGoth", "B", 14)
    pdf.cell(0, 8, "SOC1レポート（SSAE18）評価レビューシート", align="C",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.h2("1. 評価対象委託先")
    pdf.table_header(["委託先名", "委託業務", "契約番号"], [50, 70, 60])
    pdf.table_row(["外部委託先SIer-A", "SAP カスタマイズ開発・保守", "CT-2024-SIA-001"],
                  [50, 70, 60])
    pdf.table_row(["外部委託先B", "IT インフラ保守（サーバ・NW）", "CT-2023-B-005"],
                  [50, 70, 60], fill=True)
    pdf.ln(5)

    pdf.h2("2. SOC1レポート入手状況")
    pdf.kv("SIer-A", "2024年度版 SOC1 Type II 入手済（2025/5）")
    pdf.kv("B社", "2024年度版 SOC1 Type II 入手済（2025/6）")
    pdf.ln(3)

    pdf.h2("3. 重要な統制目標と評価")
    pdf.table_header(["統制目標", "SIer-A評価", "B社評価", "当社への影響"],
                     [70, 30, 30, 55])
    pdf.table_row(["アクセス管理（論理アクセス）", "有効", "有効", "問題なし"],
                  [70, 30, 30, 55])
    pdf.table_row(["変更管理", "有効", "有効", "問題なし"],
                  [70, 30, 30, 55], fill=True)
    pdf.table_row(["バックアップ・リカバリ", "有効", "有効", "問題なし"],
                  [70, 30, 30, 55])
    pdf.table_row(["インシデント管理", "一部改善要", "有効", "SIer-Aとの協議課題"],
                  [70, 30, 30, 55], fill=True)
    pdf.table_row(["物理セキュリティ", "有効", "有効", "問題なし"],
                  [70, 30, 30, 55])
    pdf.ln(5)

    pdf.h2("4. SIer-Aの改善要事項")
    pdf.body("SIer-AのSOC1レポートで『インシデント報告のタイムリー性に一部課題あり』と指摘されている。"
             "SIer-Aは当該事項について2025年度中に改善計画を策定し、2026年度レポートで効果測定予定。"
             "当社としては、重要インシデントの通知は個別契約で迅速化を要求済（2025/7協議済）。")
    pdf.ln(5)

    pdf.h2("5. 当社評価結論")
    pdf.set_font("YuGoth", "B", 11)
    pdf.set_fill_color(230, 245, 230)
    pdf.multi_cell(0, 6, "2社とも委託業務に関連する主要統制は有効であり、当社の内部統制に重要な影響を与える"
                          "不備は発見されなかった。継続利用を承認する。", fill=True)
    pdf.set_fill_color(255, 255, 255)
    pdf.ln(5)

    pdf.set_font("YuGoth", "", 10)
    pdf.kv("評価実施", "岡田 宏 (IT001)")
    pdf.kv("レビュー", "渡辺 正博 CFO (CFO001)")
    pdf.kv("評価日", "2025/8/20")

    pdf.output(str(BASE_EM / "ITGC-EM-001_SOC1レポート評価レビューシート_2025.pdf"))
    print("Created: ITGC-EM-001_SOC1レポート評価レビューシート_2025.pdf")


def gen_em_vendor_list():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IT外部委託先一覧"

    ws.cell(row=1, column=1, value="【ITGC-EM-001 統制実施記録】 IT外部委託先管理一覧")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    ws.cell(row=2, column=1, value="管理者: 岡田 宏（情シス部長 IT001）/ 更新: 2026/2/18")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    headers = ["委託先コード", "委託先名", "委託業務", "契約期間",
               "SOC1入手状況", "最終評価日", "評価結果"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    vendors = [
        ("SIA-001", "外部委託先SIer-A", "SAP開発・保守", "2024/4-2027/3", "2024年度版入手済",
         date(2025, 8, 20), "有効（一部改善要）"),
        ("B-005", "外部委託先B社", "IT インフラ保守", "2023/4-2026/3", "2024年度版入手済",
         date(2025, 8, 20), "有効"),
    ]
    r = 5
    for row in vendors:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 4, 5, 6, 7):
                cell.alignment = C_
                if c_i == 6:
                    cell.number_format = "yyyy/mm/dd"
            else:
                cell.alignment = L_
        ws.cell(row=r, column=7).fill = FILL_OK
        r += 1

    widths = [12, 20, 20, 18, 20, 14, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE_EM / "ITGC-EM-001_IT外部委託先一覧_FY2025.xlsx")
    print("Created: ITGC-EM-001_IT外部委託先一覧_FY2025.xlsx")


if __name__ == "__main__":
    # AC
    gen_ac_001_user_application()
    gen_ac_002_suim_report()
    gen_ac_003_retiree_log()
    gen_ac_003_sm20_log()
    gen_ac_004_privileged_log()
    gen_ac_screenshots()
    # CM
    gen_cm_change_log()
    gen_cm_change_request()
    gen_cm_uat_report()
    gen_cm_stms_log()
    # OM
    gen_om_backup_log()
    gen_om_dr_test()
    gen_om_incident()
    # EM
    gen_em_soc1()
    gen_em_vendor_list()
    print("\nAll ITGC evidence generated.")
