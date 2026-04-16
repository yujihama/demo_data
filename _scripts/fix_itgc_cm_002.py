"""
ITGC-CM-002 UATテスト結果の再構成
- 既存の集約CSV 「ITGC-CM-002_25件対応_RAW_UATテスト結果ログ.csv」を削除
- 個別UAT結果Excel 25件（各RELのテスト結果報告書、実データ風のファイル名）
- テスト管理ツール（Jira/Xray）から出力した連結CSV（全ケース明細）

ファイル名方針：
- 個別：UATテスト結果_REL-2025-XXX.xlsx
- 連結：Xray_TestExecution_History_FY2025.csv
"""
import random
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from sample_gen_util import write_raw_csv

BASE = Path(r"C:\Users\nyham\work\demo_data\4.evidence\ITGC")

HF = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
BFONT = Font(name="Yu Gothic", size=10)
BBOLD = Font(name="Yu Gothic", size=10, bold=True)
C_ = Alignment(horizontal="center", vertical="center", wrap_text=True)
L_ = Alignment(horizontal="left", vertical="center", wrap_text=True)
T_ = Side("thin", color="888888")
BRD = Border(left=T_, right=T_, top=T_, bottom=T_)
FILL_OK = PatternFill("solid", fgColor="E2EFDA")
FILL_NG = PatternFill("solid", fgColor="FCE4D6")
FILL_META = PatternFill("solid", fgColor="D9E1F2")


# ============================================================
# REL × 変更内容 のマッピング（再現性のため固定シード）
# ============================================================
CHANGES = [
    "販売価格マスタ連携IF修正",
    "ワークフロー承認ルーティング変更",
    "標準原価計算バッチ修正",
    "勘定科目マスタ追加",
    "仕入先マスタ項目拡張",
    "セキュリティパッチ適用",
    "バックアップバッチ改善",
    "売上レポート機能追加",
    "購買申請画面の改善",
    "連結仕訳バリデーション強化",
]

# 変更内容 × テストケーステンプレート
CASE_TEMPLATES = {
    "販売価格マスタ連携IF修正": [
        ("正常系：価格マスタ一括更新（10件）", "F", "販売管理"),
        ("正常系：価格マスタ一括更新（100件）", "F", "販売管理"),
        ("異常系：排他エラー発生時のリトライ動作", "F", "販売管理"),
        ("異常系：3回リトライ失敗時のエラー通知", "F", "販売管理"),
        ("境界系：連続同一顧客更新（100件）", "F", "販売管理"),
        ("性能系：5000件一括実行（5分以内）", "P", "販売管理"),
        ("ログ出力内容確認（実行前後・リトライ含む）", "F", "販売管理"),
    ],
    "ワークフロー承認ルーティング変更": [
        ("正常系：¥50万以下 担当承認", "F", "ワークフロー"),
        ("正常系：¥500万以下 課長承認", "F", "ワークフロー"),
        ("正常系：¥2000万以下 部長承認", "F", "ワークフロー"),
        ("正常系：¥1億以下 CFO承認", "F", "ワークフロー"),
        ("境界系：¥500万ちょうど", "F", "ワークフロー"),
        ("異常系：承認者不在時の代行ルート", "F", "ワークフロー"),
    ],
    "標準原価計算バッチ修正": [
        ("正常系：月次原価計算の実行", "F", "管理会計"),
        ("正常系：差異計算（材料/労務/製造間接）", "F", "管理会計"),
        ("境界系：期首在庫ゼロケース", "F", "管理会計"),
        ("異常系：マスタ不整合時のエラー処理", "F", "管理会計"),
        ("性能系：3000品目処理時間", "P", "管理会計"),
    ],
    "勘定科目マスタ追加": [
        ("正常系：新規勘定科目登録", "F", "FI"),
        ("正常系：科目コード重複チェック", "F", "FI"),
        ("正常系：仕訳登録での使用確認", "F", "FI"),
    ],
    "仕入先マスタ項目拡張": [
        ("正常系：既存マスタへの新規項目追加", "F", "MM"),
        ("正常系：既存データとの互換性確認", "F", "MM"),
        ("境界系：NULL許容動作", "F", "MM"),
        ("性能系：5000件マスタ処理", "P", "MM"),
    ],
    "セキュリティパッチ適用": [
        ("正常系：SAPシステム起動", "F", "基盤"),
        ("正常系：ユーザログイン確認", "F", "基盤"),
        ("正常系：業務トランザクション実行", "F", "基盤"),
        ("非機能：認証強度向上の確認", "S", "基盤"),
    ],
    "バックアップバッチ改善": [
        ("正常系：フルバックアップ実行", "F", "基盤"),
        ("正常系：差分バックアップ実行", "F", "基盤"),
        ("異常系：ストレージ容量不足時", "F", "基盤"),
        ("性能系：バックアップ完了時間（2時間以内）", "P", "基盤"),
    ],
    "売上レポート機能追加": [
        ("正常系：月次売上レポート出力", "F", "販売管理"),
        ("正常系：顧客別集計", "F", "販売管理"),
        ("正常系：製品別集計", "F", "販売管理"),
        ("正常系：CSV/PDFエクスポート", "F", "販売管理"),
        ("性能系：1年分データ出力", "P", "販売管理"),
    ],
    "購買申請画面の改善": [
        ("正常系：購買依頼作成", "F", "MM"),
        ("正常系：添付ファイル機能", "F", "MM"),
        ("UI系：入力バリデーション", "F", "MM"),
        ("UI系：レスポンシブ表示", "F", "MM"),
    ],
    "連結仕訳バリデーション強化": [
        ("正常系：有効な連結仕訳の受入", "F", "連結"),
        ("異常系：内部取引不一致時の拒否", "F", "連結"),
        ("異常系：通貨不整合時のエラー", "F", "連結"),
        ("異常系：勘定科目マッピング誤り", "F", "連結"),
        ("ログ：バリデーションエラー詳細記録", "F", "連結"),
    ],
}


TESTERS = ["中村 真理 (ACC004)", "高橋 美咲 (ACC002)",
           "石井 健 (ACC006)", "清水 智明 (PUR003)",
           "加藤 洋子 (IT003)", "松本 香織 (SLS004)"]


def gen_rels():
    """25件のRELを再現（元のシードで）"""
    random.seed(12001)
    rels = []
    for i in range(1, 26):
        rel_no = f"REL-2025-{i * 2:03d}"
        test_date = date(2025, random.randint(4, 12), random.randint(1, 28))
        # REL の変更内容を決定
        change = CHANGES[(i - 1) % len(CHANGES)]
        tester = TESTERS[i % len(TESTERS)]
        rels.append({
            "no": i, "rel_no": rel_no, "test_date": test_date,
            "change": change, "tester": tester,
        })
    return rels


def gen_test_cases_for_rel(rel, sample_no):
    """RELに対するテストケース詳細を生成"""
    change = rel["change"]
    template = CASE_TEMPLATES.get(change, CASE_TEMPLATES["販売価格マスタ連携IF修正"])

    # サンプル18は不合格→再テスト合格のパターン（元の設計を維持）
    has_failure = sample_no == 18

    cases = []
    for idx, (case_name, case_type, module) in enumerate(template, 1):
        case_id = f"TC-{change[:3]}-{idx:03d}"
        base_time = datetime.combine(rel["test_date"], datetime.min.time()) + \
            timedelta(hours=9 + idx // 2, minutes=random.randint(0, 50))

        if has_failure and idx == 3:
            # 不合格ケース → 再テストで合格
            cases.append({
                "no": idx, "case_id": case_id, "case_name": case_name,
                "case_type": case_type, "module": module,
                "execution_ts": base_time,
                "status": "FAIL",
                "defect_id": "BUG-2025-0187",
                "comment": "値引マスタ連携で想定外の0円更新",
                "tester": rel["tester"],
            })
            # 再テスト
            cases.append({
                "no": idx + 0.5, "case_id": case_id + "-RETRY",
                "case_name": case_name + "（再テスト）",
                "case_type": case_type, "module": module,
                "execution_ts": base_time + timedelta(days=2),
                "status": "PASS",
                "defect_id": "",
                "comment": "BUG-2025-0187 修正確認",
                "tester": rel["tester"],
            })
        else:
            cases.append({
                "no": idx, "case_id": case_id, "case_name": case_name,
                "case_type": case_type, "module": module,
                "execution_ts": base_time,
                "status": "PASS",
                "defect_id": "",
                "comment": "",
                "tester": rel["tester"],
            })
    return cases


# ============================================================
# (A) 個別UAT結果Excel × 25件
# ============================================================
def gen_individual_uat_excel(rel, cases):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UATテスト結果"

    # タイトル
    ws.cell(row=1, column=1, value=f"UATテスト結果報告書 / {rel['rel_no']}")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    # メタ情報
    meta = [
        ("REL番号", rel["rel_no"]),
        ("変更件名", rel["change"]),
        ("UAT実施日", rel["test_date"].strftime("%Y年%m月%d日")),
        ("実施者", rel["tester"]),
        ("対象モジュール", cases[0]["module"] if cases else ""),
        ("総ケース数", str(len([c for c in cases if not c["case_id"].endswith("-RETRY")]))),
    ]
    for i, (k, v) in enumerate(meta):
        r = 3 + i
        ws.cell(row=r, column=1, value=k).font = BBOLD
        ws.cell(row=r, column=1).fill = FILL_META
        ws.cell(row=r, column=1).border = BRD
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.cell(row=r, column=3, value=v).font = BFONT
        ws.cell(row=r, column=3).border = BRD
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)

    # ケース明細ヘッダ
    hr = 3 + len(meta) + 2
    headers = ["№", "ケースID", "ケース名", "種別", "実施日時",
               "結果", "欠陥ID", "備考"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[hr].height = 26

    # ケース行
    for r_idx, case in enumerate(cases):
        r = hr + 1 + r_idx
        row_data = [
            case["no"], case["case_id"], case["case_name"],
            case["case_type"], case["execution_ts"].strftime("%Y-%m-%d %H:%M"),
            case["status"], case["defect_id"], case["comment"],
        ]
        for c_i, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 4, 5, 6):
                cell.alignment = C_
            else:
                cell.alignment = L_
        # 結果の色付け
        if case["status"] == "PASS":
            ws.cell(row=r, column=6).fill = FILL_OK
        elif case["status"] == "FAIL":
            ws.cell(row=r, column=6).fill = FILL_NG

    # フッタ集計
    sum_r = hr + 1 + len(cases) + 1
    exec_count = len(cases)
    pass_count = sum(1 for c in cases if c["status"] == "PASS")
    fail_count = sum(1 for c in cases if c["status"] == "FAIL")
    ws.cell(row=sum_r, column=1, value="集計").font = BBOLD
    ws.cell(row=sum_r, column=1).fill = FILL_META
    ws.merge_cells(start_row=sum_r, start_column=1, end_row=sum_r, end_column=2)
    ws.cell(row=sum_r, column=3,
            value=f"総実施: {exec_count} / 合格: {pass_count} / 不合格: {fail_count}").font = BFONT
    ws.merge_cells(start_row=sum_r, start_column=3, end_row=sum_r, end_column=8)

    widths = [5, 18, 45, 12, 16, 8, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = BASE / f"UATテスト結果_{rel['rel_no']}.xlsx"
    wb.save(out)


# ============================================================
# (B) 連結ツール出力 (Jira/Xray風CSV)
# ============================================================
def gen_consolidated_xray_export(rels):
    all_cases = []
    for rel in rels:
        cases = gen_test_cases_for_rel(rel, rel["no"])
        for c in cases:
            all_cases.append((rel, c))

    rows = []
    for idx, (rel, case) in enumerate(all_cases, 1):
        exec_id = f"TEX-{idx:05d}"
        rows.append([
            exec_id,
            rel["rel_no"],
            case["case_id"],
            case["case_name"],
            case["case_type"],
            case["module"],
            case["execution_ts"].strftime("%Y-%m-%d %H:%M:%S"),
            case["tester"],
            case["status"],
            case["defect_id"],
            case["comment"],
        ])

    write_raw_csv(
        BASE / "Xray_TestExecution_History_FY2025.csv",
        ["# Jira Xray Test Execution History Export",
         "# Project: SAP-CHG (SAP Change Management)",
         "# Filter: Test executions linked to UAT test plans in FY2025",
         "# Export: 2026-02-18 13:00:00 JST",
         "# Exporter: IT003",
         "# Case Type legend: F=Functional / P=Performance / S=Security / U=Usability"],
        "ExecutionID,REL_Number,TestCaseID,TestCaseName,CaseType,Module,ExecutedAt,Tester,Status,DefectID,Comment",
        rows,
        footer_lines=[f"# Records: {len(rows)}"]
    )


def main():
    # Step 1: 旧ファイル削除
    old = BASE / "ITGC-CM-002_25件対応_RAW_UATテスト結果ログ.csv"
    if old.exists():
        old.unlink()
        print(f"Deleted: {old.name}")

    # Step 2: RELと各ケースを生成
    rels = gen_rels()

    # Step 3: 個別UAT結果Excel 25件
    for rel in rels:
        cases = gen_test_cases_for_rel(rel, rel["no"])
        gen_individual_uat_excel(rel, cases)
    print(f"Created: 25 individual UAT result Excel files")

    # Step 4: 連結Xrayエクスポート
    gen_consolidated_xray_export(rels)
    print("Created: Xray_TestExecution_History_FY2025.csv")


if __name__ == "__main__":
    main()
