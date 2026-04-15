"""
PLC-S（販売プロセス）のエビデンス生成 v2
新方針：エビデンス＝監査人が評価を実施するために被評価部門から得る「素材」
- 母集団リスト（SAPエクスポート）
- 統制実施者（営業・経理部）が作成した実施記録
- システムログ・スクリーンショット
- マスタデータのスナップショット
- 個別取引の原本書類

会社名はマスタデータから参照（完全ダミー）
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
from pathlib import Path
import random
import sys

sys.path.insert(0, str(Path(__file__).parent))
from pdf_util import JPPDF
from image_util import sap_screenshot, workflow_screenshot

BASE = Path(r"C:\Users\nyham\work\demo_data\4.evidence\PLC-S")
BASE.mkdir(parents=True, exist_ok=True)

# スタイル
HF = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
BFONT = Font(name="Yu Gothic", size=10)
BBOLD = Font(name="Yu Gothic", size=10, bold=True)
C_ = Alignment(horizontal="center", vertical="center", wrap_text=True)
L_ = Alignment(horizontal="left", vertical="center", wrap_text=True)
R_ = Alignment(horizontal="right", vertical="center")
T_ = Side("thin", color="888888")
BRD = Border(left=T_, right=T_, top=T_, bottom=T_)
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")
FILL_NG = PatternFill("solid", fgColor="FCE4D6")

# マスタデータ（customers.xlsxから読み込む形にしてもよいが、ここでは同じ値を直接定義）
CUSTOMERS = {
    "C-10001": ("サンプル顧客A社", 500_000_000, "自動車"),
    "C-10002": ("サンプル顧客B社", 300_000_000, "自動車"),
    "C-10003": ("サンプル顧客C社", 200_000_000, "自動車"),
    "C-10004": ("サンプル顧客D社", 150_000_000, "自動車"),
    "C-10005": ("サンプル顧客E社", 100_000_000, "自動車"),
    "C-10006": ("サンプル顧客F社", 80_000_000, "自動車"),
    "C-10007": ("サンプル顧客G社", 50_000_000, "自動車"),
    "C-10011": ("サンプル顧客H社", 400_000_000, "半導体装置"),
    "C-10012": ("サンプル顧客I社", 250_000_000, "半導体装置"),
    "C-10013": ("サンプル顧客J社", 180_000_000, "半導体装置"),
    "C-10014": ("サンプル顧客K社", 120_000_000, "半導体装置"),
    "C-10015": ("サンプル顧客L社", 100_000_000, "半導体装置"),
    "C-10016": ("サンプル顧客M社", 80_000_000, "半導体装置"),
    "C-10017": ("サンプル顧客N社", 60_000_000, "半導体装置"),
    "C-10018": ("サンプル顧客O社", 40_000_000, "半導体装置"),
    "C-10021": ("サンプル顧客P社", 200_000_000, "商社"),
    "C-10022": ("サンプル顧客Q社", 150_000_000, "商社"),
    "C-10023": ("サンプル顧客R社", 80_000_000, "商社"),
    "C-10024": ("サンプル顧客S社", 50_000_000, "商社"),
    "C-10025": ("サンプル顧客T社", 30_000_000, "商社"),
}


# ============================================================
# 【PLC-S-001 受注・与信承認】用のエビデンス
# ============================================================

def gen_population_orders():
    """
    PLC-S-001 / PLC-S-002 用の母集団：
    SAP VA05（受注伝票一覧）のエクスポート
    FY2025期間（2025/4/1～2026/3/31）中の受注 150件を代表として収録
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SAP_VA05_受注伝票一覧"

    # メタ情報（SAPエクスポートの典型的なヘッダ）
    ws.cell(row=1, column=1, value="SAP S/4HANA / トランザクション: VA05 / 出力日時: 2026/02/10 10:30:12")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    ws.cell(row=2, column=1, value="出力者: ACC002 高橋 美咲 / 抽出条件: 受注日=2025/4/1-2026/3/31 / 全ステータス / キャンセル除く / 総件数: 3,247件（以下は前150件）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)

    headers = ["№", "受注番号", "受注日", "受注区分", "顧客コード", "顧客名", "営業担当",
               "受注金額(円)", "ステータス", "与信チェック結果",
               "承認要否", "承認者", "承認日", "出荷予定日"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[4].height = 28

    random.seed(2025001)
    cids = list(CUSTOMERS.keys())
    reps = ["斎藤 次郎", "藤田 修", "松本 香織", "井上 大輔"]

    # 150件生成
    r = 5
    for i in range(1, 151):
        # 日付（FY2025内に分散）
        month = random.choices(
            [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3],
            weights=[8, 10, 10, 10, 12, 12, 15, 15, 10, 10, 8, 10])[0]
        year = 2025 if month >= 4 else 2026
        day = random.randint(1, 28)
        order_date = date(year, month, day)

        cid = random.choice(cids)
        cname, credit, _industry = CUSTOMERS[cid]
        amount = random.choice([
            random.randint(500_000, 5_000_000),
            random.randint(5_000_000, 30_000_000),
            random.randint(30_000_000, 80_000_000),
        ])
        rep = random.choice(reps)

        # 与信判定
        # 仮の既存売掛金を与信枠の30~90%とする（簡略化）
        existing_ar = int(credit * random.uniform(0.3, 0.9))
        total_if_ordered = existing_ar + amount
        if total_if_ordered > credit:
            credit_check = "超過（保留）"
            approval_needed = "要"
            approval = "田中 太郎(営業本部長)"
            approval_delta = random.choice([0, 0, 0, 1])  # 主に当日承認
            approval_date = order_date + timedelta(days=approval_delta)
            approval_date_str = approval_date.strftime("%Y/%m/%d")
        else:
            credit_check = "○ 限度内"
            approval_needed = "不要（自動）"
            approval = "(SAP自動承認)"
            approval_date_str = order_date.strftime("%Y/%m/%d")

        # ステータス
        status = random.choices(
            ["完了", "処理中", "部分出荷"],
            weights=[70, 20, 10])[0]

        order_no = f"ORD-2025-{i * 22:04d}"
        ship_date = order_date + timedelta(days=random.randint(7, 45))

        data = [i, order_no, order_date,
                "通常受注" if i % 30 != 0 else "個別注文",
                cid, cname, rep, amount, status, credit_check,
                approval_needed, approval, approval_date_str, ship_date]

        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 5, 7, 9, 10, 11, 13, 14):
                cell.alignment = C_
                if c_i in (3, 14):
                    cell.number_format = "yyyy/mm/dd"
            elif c_i == 8:
                cell.alignment = R_
                cell.number_format = "#,##0"
            else:
                cell.alignment = L_

        # 与信超過を色付け
        if "超過" in credit_check:
            ws.cell(row=r, column=10).fill = FILL_WARN
            ws.cell(row=r, column=11).fill = FILL_WARN

        r += 1

    # 末尾注記
    ws.cell(row=r + 1, column=1, value="※ 本レポートはFY2025全3,247件中の前150件です。完全版はSAP上で参照可能。")
    ws.cell(row=r + 1, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=14)

    widths = [5, 15, 12, 10, 10, 18, 12, 14, 10, 12, 10, 18, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    wb.save(BASE / "PLC-S-001_SAP_VA05_受注伝票一覧_FY2025.xlsx")
    print("Created: PLC-S-001_SAP_VA05_受注伝票一覧_FY2025.xlsx")


def gen_credit_limit_master():
    """
    PLC-S-001用：与信限度マスタ（評価基準時点でのスナップショット）
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "与信限度マスタ"

    ws.cell(row=1, column=1, value="SAP FD32 / 顧客別与信限度額マスタ（FY2025期末時点スナップショット）")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=2, column=1, value="出力日時: 2026/2/10 14:22:08 / 出力者: ACC001 佐藤 一郎 / 基準日: 2026/2/10")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    headers = ["顧客コード", "顧客名", "業種", "与信限度額(円)", "最終見直日", "前回限度額(円)", "見直履歴"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    random.seed(1111)
    r = 5
    for cid, (cname, credit, industry) in CUSTOMERS.items():
        prev_credit = int(credit * random.uniform(0.85, 1.05))
        last_review = date(2025, random.choice([4, 5, 6]), random.randint(1, 28))
        history = "引上げ" if credit > prev_credit else ("引下げ" if credit < prev_credit else "据置")
        data = [cid, cname, industry, credit, last_review, prev_credit, history]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 3, 5, 7):
                cell.alignment = C_
                if c_i == 5:
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (4, 6):
                cell.alignment = R_
                cell.number_format = "#,##0"
            else:
                cell.alignment = L_
        r += 1

    # 承認記録
    r += 2
    ws.cell(row=r, column=1, value="■ マスタ更新の承認記録（FY2025期首見直し時）").font = BBOLD
    r += 1
    ws.cell(row=r, column=1, value="作成: 高橋 美咲（経理部課長）[印] 2025/5/12")
    r += 1
    ws.cell(row=r, column=1, value="承認: 佐藤 一郎（経理部長）[印] 2025/5/15 / 渡辺 正博（CFO）[印] 2025/5/20")

    widths = [12, 18, 12, 16, 12, 16, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-S-001_与信限度マスタ_SAP_FD32スナップショット.xlsx")
    print("Created: PLC-S-001_与信限度マスタ_SAP_FD32スナップショット.xlsx")


def gen_approval_authority():
    """
    PLC-S-001, S-007用：販売関連の承認権限一覧（職務権限規程R18の抜粋）
    """
    pdf = JPPDF()
    pdf.add_page()

    pdf.h1("販売関連 承認権限一覧")
    pdf.set_font("YuGoth", "", 9)
    pdf.cell(0, 5, "職務権限規程R18 抜粋 / 改訂日: 2025/4/1 / 次回見直し: 2026/3/31",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    pdf.h2("1. 受注承認権限")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
        "受注は原則として営業担当（SD_USER権限）が登録し、与信限度内であれば自動承認される。"
        "下記に該当する場合は個別承認が必要。")
    pdf.ln(2)
    pdf.table_header(["区分", "条件", "承認者", "備考"], [30, 80, 40, 40])
    pdf.table_row(["通常", "与信限度内", "自動承認", "―"], [30, 80, 40, 40])
    pdf.table_row(["与信超過", "既存残+受注額 > 与信限度", "営業本部長", "ワークフロー経由"],
                  [30, 80, 40, 40], fill=True)
    pdf.table_row(["新規顧客", "顧客コード未登録", "営業本部長+CFO", "反社チェック必要"],
                  [30, 80, 40, 40])
    pdf.table_row(["高額個別", "単一受注額 > ¥100,000,000", "代表取締役", "取締役会報告対象"],
                  [30, 80, 40, 40], fill=True)
    pdf.ln(5)

    pdf.h2("2. 価格マスタ変更権限")
    pdf.set_font("YuGoth", "", 10)
    pdf.table_header(["区分", "承認者", "稟議フォーム"], [60, 60, 60])
    pdf.table_row(["既存顧客の個別単価変更", "営業本部長", "W-YYYY-NNNN"],
                  [60, 60, 60])
    pdf.table_row(["新規品目の初期価格設定", "営業本部長+CFO", "W-YYYY-NNNN"],
                  [60, 60, 60], fill=True)
    pdf.table_row(["標準価格マスタの全体改定", "代表取締役（取締役会決議）", "別途"],
                  [60, 60, 60])
    pdf.ln(5)

    pdf.h2("3. 顧客マスタ変更権限")
    pdf.set_font("YuGoth", "", 10)
    pdf.table_header(["区分", "承認者"], [90, 90])
    pdf.table_row(["新規登録", "営業本部長（反社チェック後）"], [90, 90])
    pdf.table_row(["与信限度引上", "CFO（年次見直し）"], [90, 90], fill=True)
    pdf.table_row(["住所・連絡先等", "営業課長"], [90, 90])
    pdf.ln(10)

    pdf.set_font("YuGoth", "", 9)
    pdf.cell(0, 5, "承認: 2025/4/1施行 / 山本 健一 代表取締役社長",
             new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(BASE / "PLC-S-001_販売関連承認権限一覧_職務権限規程R18抜粋.pdf"))
    print("Created: PLC-S-001_販売関連承認権限一覧_職務権限規程R18抜粋.pdf")


def gen_credit_check_log():
    """
    PLC-S-001用：与信チェックログ（SAPの2025年11月1か月分のCSVエクスポート）
    """
    path = BASE / "PLC-S-001_SAP与信チェックログ_202511.csv"
    random.seed(12345)
    cids = list(CUSTOMERS.keys())

    lines = []
    lines.append("# SAP S/4HANA 与信管理モジュール / 与信チェックログ")
    lines.append("# 出力日時: 2025/12/01 09:15:22")
    lines.append("# 対象期間: 2025/11/01 - 2025/11/30")
    lines.append("# 出力者: ACC002 高橋 美咲")
    lines.append("")
    lines.append("タイムスタンプ,受注番号,顧客コード,受注金額,既存売掛金,与信限度額,判定,保留解除ユーザ,保留解除日時")

    for i in range(1, 120):
        ts = datetime(2025, 11, random.randint(1, 29),
                      random.randint(9, 18), random.randint(0, 59), random.randint(0, 59))
        cid = random.choice(cids)
        _name, credit, _ = CUSTOMERS[cid]
        amount = random.choice([
            random.randint(500_000, 3_000_000),
            random.randint(3_000_000, 15_000_000),
            random.randint(15_000_000, 50_000_000),
        ])
        existing_ar = int(credit * random.uniform(0.3, 0.95))
        total = existing_ar + amount
        if total > credit:
            judge = "HOLD"
            release_user = "SLS001"
            release_ts = ts + timedelta(hours=random.randint(1, 26))
            release_str = release_ts.strftime("%Y-%m-%d %H:%M:%S")
        else:
            judge = "PASS"
            release_user = ""
            release_str = ""
        lines.append(f"{ts.strftime('%Y-%m-%d %H:%M:%S')},ORD-2025-{i * 22 + 1000:04d},{cid},{amount},{existing_ar},{credit},{judge},{release_user},{release_str}")

    lines.append("")
    lines.append("# 件数: 119件 / PASS: 約95件 / HOLD: 約24件")
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"Created: {path.name}")


# ============================================================
# 個別受注のPDFエビデンス（注文書 - 顧客から受領したもの）
# ============================================================
def gen_order_pdf(order_no, order_date, customer_code, items,
                  delivery_date, rep_name, output_name,
                  customer_po_ref=None):
    pdf = JPPDF()
    pdf.add_page()

    pdf.set_font("YuGoth", "B", 20)
    pdf.cell(0, 12, "注 文 書", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, f"(顧客側) 注文番号: {customer_po_ref or order_no}",
             align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"発行日: {order_date.strftime('%Y年%m月%d日')}",
             align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.set_font("YuGoth", "B", 12)
    pdf.cell(0, 7, "株式会社テクノプレシジョン 御中", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, "営業本部 担当者殿", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    customer_name, _, _ = CUSTOMERS[customer_code]
    pdf.set_font("YuGoth", "B", 10)
    pdf.set_x(110)
    pdf.cell(90, 6, f"発注元: {customer_name}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("YuGoth", "", 9)
    pdf.set_x(110)
    pdf.cell(90, 5, f"顧客コード: {customer_code}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(110)
    pdf.cell(90, 5, "(所在地: 当社管理)", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)

    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5, "下記のとおり発注致します。ご確認のうえ、納期に間に合うよう手配をお願い致します。")
    pdf.ln(3)

    pdf.table_header(["品目コード", "品名", "数量", "単価(円)", "金額(円)"],
                     [30, 80, 20, 30, 30])
    subtotal = 0
    for code, name, qty, unit in items:
        amount = qty * unit
        subtotal += amount
        pdf.table_row([code, name, f"{qty:,}", f"{unit:,}", f"{amount:,}"],
                      [30, 80, 20, 30, 30])

    pdf.set_font("YuGoth", "B", 10)
    pdf.cell(130, 7, "小計", border=1, align="R")
    pdf.cell(60, 7, f"¥ {subtotal:,}", border=1, align="R",
             new_x="LMARGIN", new_y="NEXT")
    tax = int(subtotal * 0.1)
    pdf.cell(130, 7, "消費税 (10%)", border=1, align="R")
    pdf.cell(60, 7, f"¥ {tax:,}", border=1, align="R",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_fill_color(255, 242, 204)
    pdf.cell(130, 8, "合計", border=1, align="R", fill=True)
    total = subtotal + tax
    pdf.cell(60, 8, f"¥ {total:,}", border=1, align="R", fill=True,
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_fill_color(255, 255, 255)
    pdf.ln(5)

    pdf.h3("■ 納入条件")
    pdf.set_font("YuGoth", "", 9)
    pdf.kv("納期", delivery_date.strftime("%Y年%m月%d日"))
    pdf.kv("納入場所", "貴社指定倉庫")
    pdf.kv("支払条件", "月末締 翌月末払")
    pdf.ln(8)

    # 自社側の受領記録
    y_stamp = pdf.get_y()
    pdf.stamp("受領", x=30, y=y_stamp + 10)
    pdf.set_font("YuGoth", "", 9)
    pdf.set_xy(50, y_stamp + 5)
    pdf.cell(0, 5, f"自社受領日: {order_date.strftime('%Y/%m/%d')}",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(50)
    pdf.cell(0, 5, f"受領担当: {rep_name}", new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(50)
    pdf.cell(0, 5, f"SAP受注番号: {order_no}", new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(BASE / output_name))
    print(f"Created: {output_name}")


def gen_all_order_pdfs():
    # 代表的な注文書3件（実際の監査では多数の注文書が実在し、そこからサンプリング）
    gen_order_pdf(
        "ORD-2025-1420",
        date(2025, 11, 10),
        "C-10002",
        [("P-30006", "トランスミッションシャフト", 1000, 12500)],
        date(2025, 12, 15),
        "松本 香織",
        "PLC-S-001_注文書_ORD-2025-1420_サンプル顧客B社.pdf",
        customer_po_ref="CUST-PO-2025-8843",
    )
    gen_order_pdf(
        "ORD-2025-0412",
        date(2025, 10, 25),
        "C-10011",
        [
            ("P-30011", "ウェハー搬送ロボット用シャフト A", 80, 18500),
            ("P-30014", "エッチング装置チャンバ部品", 40, 38500),
        ],
        date(2025, 11, 30),
        "藤田 修",
        "PLC-S-001_注文書_ORD-2025-0412_サンプル顧客H社.pdf",
        customer_po_ref="CUST-PO-2025-0934",
    )
    gen_order_pdf(
        "ORD-2025-1876",
        date(2025, 12, 3),
        "C-10015",
        [("P-30015", "ウェハーチャックベース", 1500, 12800)],
        date(2026, 1, 20),
        "松本 香織",
        "PLC-S-001_注文書_ORD-2025-1876_サンプル顧客L社.pdf",
        customer_po_ref="CUST-PO-2025-4521",
    )


# ============================================================
# SAPスクリーンショット系
# ============================================================
def gen_screenshots():
    sap_screenshot(
        "受注登録 - 照会",
        "VA03",
        [
            ("受注伝票番号", "ORD-2025-1420"),
            ("受注日", "2025/11/10"),
            ("販売先", "C-10002 サンプル顧客B社"),
            ("販売担当", "SLS004 松本 香織"),
            ("与信限度額", "300,000,000 JPY"),
            ("既存売掛金残高", "242,320,000 JPY"),
            ("当該受注金額", "12,500,000 JPY"),
            ("想定与信残", "45,180,000 JPY"),
            ("与信チェック結果", "○ 限度内通過"),
            ("ステータス", "承認済（自動承認）"),
        ],
        grid_headers=["品目", "製品コード", "数量", "単価", "金額"],
        grid_rows=[
            ["トランスミッションシャフト", "P-30006", "1,000", "12,500", "12,500,000"],
        ],
        status_bar="受注伝票 ORD-2025-1420 が保存されました。",
        output_path=str(BASE / "PLC-S-001_SAP受注登録画面_ORD-2025-1420.png"),
    )

    sap_screenshot(
        "受注登録 - 与信チェック警告",
        "VA01",
        [
            ("受注伝票番号", "(新規作成中)"),
            ("販売先", "C-10007 サンプル顧客G社"),
            ("販売担当", "SLS005 井上 大輔"),
            ("与信限度額", "50,000,000 JPY"),
            ("既存売掛金残高", "48,200,000 JPY"),
            ("当該受注金額", "8,500,000 JPY"),
            ("想定与信残", "-6,700,000 JPY (超過!)"),
            ("与信チェック結果", "× 限度額超過"),
            ("ワークフロー", "営業本部長承認待ち"),
            ("ステータス", "保留中"),
        ],
        grid_headers=["品目", "製品コード", "数量", "単価", "金額"],
        grid_rows=[
            ["エンジンマウントブラケット", "P-30021", "12,500", "680", "8,500,000"],
        ],
        status_bar="⚠ 与信限度を超過しています。営業本部長の承認が必要です。",
        output_path=str(BASE / "PLC-S-001_SAP与信超過アラート画面_C-10007.png"),
    )

    workflow_screenshot(
        "WF-2025-3456",
        "受注承認（与信超過案件）",
        "井上 大輔（営業部担当）",
        [
            ("井上 大輔", "申請者（営業部）", "2025/11/25 10:15", "申請"),
            ("斎藤 次郎", "課長（営業部）", "2025/11/25 13:22", "承認"),
            ("田中 太郎", "本部長（営業本部）", "2025/11/25 16:48", "承認"),
            ("佐藤 一郎", "経理部長（与信審査）", "2025/11/26 09:30", "承認"),
        ],
        amount=8_500_000,
        comments="サンプル顧客G社 への受注、限度額超過分について本部長・経理部長承認取得",
        output_path=str(BASE / "PLC-S-001_ワークフロー承認_与信超過サンプル.png"),
    )

    sap_screenshot(
        "入金消込処理",
        "F-28",
        [
            ("入金日", "2025/11/25"),
            ("入金額", "25,480,000 JPY"),
            ("振込元", "サンプル顧客B社"),
            ("対象顧客", "C-10002"),
            ("消込方法", "SAP自動消込"),
            ("対象請求書", "INV-202510-0089 (¥25,480,000)"),
            ("差額", "0 JPY"),
            ("消込ステータス", "完了"),
            ("処理担当", "ACC006 石井 健"),
        ],
        grid_headers=["請求書番号", "発行日", "金額", "ステータス"],
        grid_rows=[
            ["INV-202510-0089", "2025/10/31", "25,480,000", "消込完了"],
        ],
        status_bar="入金消込処理が完了しました。",
        output_path=str(BASE / "PLC-S-004_SAP入金消込画面.png"),
    )

    print("Created: 4 SAP screenshots for PLC-S")


# ============================================================
# PLC-S-002 出荷売上マッチング用：出荷実績・売上明細のSAPエクスポート
# ============================================================
def gen_shipment_data():
    """WMS出荷実績エクスポート（CSV・2025年11月1ヶ月分の抜粋）"""
    path = BASE / "PLC-S-002_WMS出荷実績エクスポート_202511.csv"
    random.seed(202511)

    lines = []
    lines.append("# WMS (倉庫管理システム) / 出荷実績レポート")
    lines.append("# 出力日時: 2025/12/01 07:32:11 / 出力者: WHS001 橋本 明")
    lines.append("# 対象期間: 2025/11/01 - 2025/11/30")
    lines.append("")
    lines.append("出荷番号,出荷日時,出荷区分,受注番号,顧客コード,顧客名,製品コード,出荷数量,出荷金額,出荷倉庫,出荷担当")

    cids = list(CUSTOMERS.keys())
    products = [("P-30001", 4200), ("P-30006", 12500), ("P-30011", 18500),
                ("P-30014", 38500), ("P-30020", 42500), ("P-30022", 1280),
                ("P-30027", 4200)]

    for i in range(1, 61):
        ship_ts = datetime(2025, 11, random.randint(1, 29),
                           random.randint(8, 18), random.randint(0, 59))
        cid = random.choice(cids)
        cname = CUSTOMERS[cid][0]
        pcode, unit = random.choice(products)
        qty = random.randint(5, 500)
        amount = qty * unit
        order_no = f"ORD-2025-{i * 13 + 200:04d}"
        warehouse = random.choice(["本社倉庫A", "本社倉庫B", "東北工場倉庫"])
        user = random.choice(["WHS001", "MFG002"])
        lines.append(f"SH-202511-{i:04d},{ship_ts.strftime('%Y-%m-%d %H:%M:%S')},通常出荷,{order_no},{cid},{cname},{pcode},{qty},{amount},{warehouse},{user}")

    lines.append("")
    lines.append("# 総件数: 60件 / 総出荷金額: (集計はSAP側で実施)")
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"Created: {path.name}")


def gen_sales_posting_data():
    """SAP売上計上明細エクスポート（CSV・2025年11月分）"""
    path = BASE / "PLC-S-002_SAP売上計上明細_202511.csv"
    random.seed(20251102)

    lines = []
    lines.append("# SAP FI / 売上計上仕訳明細 (勘定科目 5100 売上高)")
    lines.append("# トランザクション: FBL3N")
    lines.append("# 出力日時: 2025/12/01 08:45:33 / 出力者: ACC004 中村 真理")
    lines.append("# 対象期間: 会計期間 8 (2025/11)")
    lines.append("")
    lines.append("仕訳番号,計上日,伝票タイプ,勘定科目,借方/貸方,金額,顧客コード,受注番号,出荷番号,摘要,起票者")

    cids = list(CUSTOMERS.keys())

    for i in range(1, 65):
        post_date = date(2025, 11, random.randint(1, 30))
        cid = random.choice(cids)
        amount = random.randint(1_000_000, 25_000_000)
        order_no = f"ORD-2025-{i * 13 + 200:04d}"
        ship_no = f"SH-202511-{i:04d}"
        jv = f"JV-202511-{i:04d}"
        # 売上計上: 借方 売掛金 / 貸方 売上高
        lines.append(f"{jv},{post_date.strftime('%Y-%m-%d')},RV,1220 売掛金,借方,{amount},{cid},{order_no},{ship_no},出荷連動自動仕訳,SAP自動")
        lines.append(f"{jv},{post_date.strftime('%Y-%m-%d')},RV,5100 売上高,貸方,{amount},{cid},{order_no},{ship_no},出荷連動自動仕訳,SAP自動")

    lines.append("")
    lines.append("# 仕訳件数: 64件（借方・貸方ペア128行）")
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"Created: {path.name}")


def gen_match_report():
    """経理部が日次で出力しているマッチング照合レポート（統制実施記録）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "出荷売上マッチング"

    ws.cell(row=1, column=1, value="【PLC-S-002 統制実施記録】 出荷-売上マッチング日次照合レポート (2025年11月)")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=2, column=1, value="作成日: 2025/12/03 / 作成者: 中村 真理（経理部主任 ACC004） / 承認: 高橋 美咲（経理部課長 ACC002）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    ws.cell(row=3, column=1, value="母集団: 2025/11の全出荷（WMS 60件）× 全売上計上（SAP 64件）")
    ws.cell(row=3, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)

    headers = ["出荷日", "出荷番号", "受注番号", "顧客", "出荷金額", "売上計上日",
               "売上金額", "マッチ結果"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    random.seed(20251103)
    cids = list(CUSTOMERS.keys())
    r = 6
    for i in range(1, 31):
        ship_date = date(2025, 11, random.randint(1, 29))
        sale_date = ship_date + timedelta(days=random.choice([0, 0, 0, 1]))
        cid = random.choice(cids)
        amount = random.randint(1_000_000, 25_000_000)
        sale_amount = amount
        result = "一致"
        if i == 17:
            sale_amount = amount - 50_000
            result = "差異¥50,000\n(値引調整・是正済)"
        data = [ship_date, f"SH-202511-{i:04d}",
                f"ORD-2025-{1000 + i * 13}", CUSTOMERS[cid][0],
                amount, sale_date, sale_amount, result]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 6, 8):
                cell.alignment = C_
                if c_i in (1, 6):
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (5, 7):
                cell.alignment = R_
                cell.number_format = "#,##0"
        if result != "一致":
            ws.cell(row=r, column=8).fill = FILL_WARN
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="突合結果サマリ: 当月30件中 一致29件、差異1件（¥50,000、値引伝票で是正済）").font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 2
    ws.cell(row=r, column=1, value="経理部課長レビュー: 高橋 美咲 [印] 2025/12/04")
    r += 1
    ws.cell(row=r, column=1, value="経理部長承認: 佐藤 一郎 [印] 2025/12/05")

    widths = [12, 15, 16, 16, 14, 12, 14, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-S-002_出荷売上マッチング照合レポート_202511.xlsx")
    print("Created: PLC-S-002_出荷売上マッチング照合レポート_202511.xlsx")


def gen_unmatch_csv():
    """未マッチ明細（SAPから定期的に出力）"""
    path = BASE / "PLC-S-002_SAP未マッチ明細リスト_202511.csv"
    lines = [
        "# SAP Query ZSD_UNMATCH / 未マッチ出荷-売上明細",
        "# 出力日時: 2025/12/03 08:15:23",
        "# 抽出条件: 出荷日=2025/11/1-11/30 AND 売上計上未実施",
        "# ユーザ: ACC004 中村 真理",
        "",
        "№,出荷番号,出荷日,受注番号,顧客コード,顧客名,出荷金額,売上計上日,マッチ結果,原因,是正記録",
        "1,SH-202511-0234,2025/11/17,ORD-2025-2468,C-10003,サンプル顧客C社,12850000,2025/11/17,金額差異,出荷時に数量訂正あり,2025/11/20値引伝票で¥50,000調整済",
        "",
        "# 件数: 1件（是正済）",
    ]
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"Created: {path.name}")


# ============================================================
# PLC-S-003 請求書発行
# ============================================================
def gen_invoice_batch_log():
    """月次請求書発行バッチのSAPログ"""
    path = BASE / "PLC-S-003_SAP請求書バッチ実行ログ_202511.txt"
    content = """================================================================
 SAP S/4HANA / SD Invoice Batch Execution Log
================================================================
バッチジョブ名: ZSD_MONTHLY_INVOICE
ジョブ番号:     JOB_20251130_2358_01
実行ユーザ:     SAP_BATCH (system)
開始日時:       2025-11-30 23:58:01
終了日時:       2025-12-01 00:12:47
実行時間:       14分46秒
ステータス:     正常終了 (RC=0)

----------------------------------------------------------------
 処理対象
----------------------------------------------------------------
 対象顧客数:        20社
 対象受注明細数:    287件
 生成請求書数:      152件
 合計請求金額:      ¥1,284,560,000 (税抜)
                    ¥1,413,016,000 (税込)

----------------------------------------------------------------
 ログ明細
----------------------------------------------------------------
2025-11-30 23:58:01 [INFO]  Start: ZSD_MONTHLY_INVOICE
2025-11-30 23:58:03 [INFO]  Reading parameters: PERIOD=202511
2025-11-30 23:58:05 [INFO]  Loading billing-due list: 287 items
2025-11-30 23:58:12 [INFO]  Consolidating by customer: 20 customers
2025-11-30 23:58:15 [INFO]  Generating billing documents...
2025-11-30 23:58:17 [INFO]  Customer C-10001: 12 invoices, ¥185,400,000
2025-11-30 23:58:18 [INFO]  Customer C-10002: 8 invoices, ¥96,800,000
2025-11-30 23:58:20 [INFO]  Customer C-10003: 11 invoices, ¥124,500,000
...
(中略: 全20顧客の生成ログ)
...
2025-11-30 23:58:40 [INFO]  Customer C-10025: 3 invoices, ¥12,400,000
2025-12-01 00:04:12 [INFO]  All documents generated: 152 items
2025-12-01 00:04:15 [INFO]  Validating against sales journal...
2025-12-01 00:04:45 [INFO]  Reconciliation: 152/152 matched. Delta = ¥0
2025-12-01 00:04:47 [INFO]  Starting PDF generation...
2025-12-01 00:11:23 [INFO]  PDF generation completed: 152 files
2025-12-01 00:11:25 [INFO]  Email dispatch for PDF-enabled customers...
2025-12-01 00:12:47 [INFO]  Email sent: 127 customers / Paper print: 25 customers
2025-12-01 00:12:47 [INFO]  End: ZSD_MONTHLY_INVOICE (SUCCESS)

----------------------------------------------------------------
 次回起動予定: 2025-12-31 23:58:00
----------------------------------------------------------------
"""
    path.write_text(content, encoding="utf-8")
    print(f"Created: {path.name}")


def gen_invoice_monthly_list():
    """月次請求書発行一覧（経理部作成）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "月次請求書発行一覧"

    ws.cell(row=1, column=1, value="【PLC-S-003 統制実施記録】 2025年11月 月次請求書発行一覧")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=2, column=1, value="発行バッチ実行日時: 2025/11/30 23:58（SAP自動発行） / 発行件数: 152件 / 合計金額: ¥1,413,016,000（税込）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headers = ["請求書番号", "発行日", "顧客コード", "顧客名", "請求金額(円)",
               "税込金額(円)", "支払期日", "送付方法"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    random.seed(3003)
    cids = list(CUSTOMERS.keys())
    r = 5
    for i in range(1, 31):
        cid = random.choice(cids)
        cname = CUSTOMERS[cid][0]
        amount = random.randint(1_000_000, 35_000_000) // 1000 * 1000
        tax = int(amount * 1.1)
        due = date(2026, 1, 31 if i % 3 == 0 else 20 if i % 3 == 1 else 15)
        send = random.choice(["PDFメール送付", "PDFメール送付", "郵送（原本）"])
        data = [f"INV-202511-{i:04d}", date(2025, 11, 30), cid, cname, amount, tax, due, send]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 7, 8):
                cell.alignment = C_
                if c_i in (2, 7):
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (5, 6):
                cell.alignment = R_
                cell.number_format = "#,##0"
        r += 1

    ws.cell(row=r, column=1, value="... 以下122件省略 ...").font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 2
    ws.cell(row=r, column=1, value="突合チェック実施: 中村 真理 [印] 2025/12/01 / 売上計上額との一致確認 OK（差額 ¥0）").font = BFONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    widths = [16, 12, 12, 18, 14, 14, 12, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-S-003_月次請求書発行一覧_202511.xlsx")
    print("Created: PLC-S-003_月次請求書発行一覧_202511.xlsx")


def gen_invoice_pdf():
    """自社発行の請求書PDF"""
    pdf = JPPDF()
    pdf.add_page()

    pdf.set_font("YuGoth", "B", 22)
    pdf.cell(0, 14, "請 求 書", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, "請求書番号: INV-202511-0234", align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "請求日: 2025年11月30日", align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    pdf.set_font("YuGoth", "B", 12)
    pdf.cell(0, 7, "サンプル顧客B社 御中", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("YuGoth", "", 9)
    pdf.cell(0, 5, "顧客コード: C-10002", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.set_x(110)
    pdf.set_font("YuGoth", "B", 11)
    pdf.cell(90, 6, "株式会社テクノプレシジョン", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("YuGoth", "", 9)
    pdf.set_x(110)
    pdf.cell(90, 5, "〒XXX-XXXX 神奈川県横浜市港北区", new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(110)
    pdf.cell(90, 5, "TEL: 045-XXX-XXXX", new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(110)
    pdf.cell(90, 5, "登録番号: T1234567890123", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)

    subtotal = 12_500_000
    tax = int(subtotal * 0.1)
    total = subtotal + tax
    pdf.set_font("YuGoth", "B", 14)
    pdf.set_fill_color(240, 245, 255)
    pdf.cell(60, 14, "ご請求金額", border=1, align="C", fill=True)
    pdf.cell(130, 14, f"¥ {total:,} -", border=1, align="R", fill=True,
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_fill_color(255, 255, 255)
    pdf.ln(6)

    pdf.table_header(["品目コード", "品名", "数量", "単価", "金額"],
                     [30, 80, 20, 30, 30])
    pdf.table_row(["P-30006", "トランスミッションシャフト", "1,000", "12,500", "12,500,000"],
                  [30, 80, 20, 30, 30])

    pdf.set_font("YuGoth", "B", 10)
    pdf.cell(130, 7, "小計", border=1, align="R")
    pdf.cell(60, 7, f"¥ {subtotal:,}", border=1, align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(130, 7, "消費税 (10%)", border=1, align="R")
    pdf.cell(60, 7, f"¥ {tax:,}", border=1, align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.set_fill_color(255, 242, 204)
    pdf.cell(130, 8, "合計", border=1, align="R", fill=True)
    pdf.cell(60, 8, f"¥ {total:,}", border=1, align="R", fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.set_fill_color(255, 255, 255)
    pdf.ln(6)

    pdf.h3("■ お支払について")
    pdf.set_font("YuGoth", "", 10)
    pdf.kv("お支払期日", "2025年12月31日")
    pdf.kv("お支払方法", "銀行振込")
    pdf.kv("振込先", "A銀行 支店X 普通 1234567")
    pdf.kv("口座名義", "カ）テクノプレシジョン")
    pdf.ln(6)

    y_stamp = pdf.get_y()
    pdf.stamp("会社印", x=170, y=y_stamp + 10)

    pdf.output(str(BASE / "PLC-S-003_請求書_INV-202511-0234.pdf"))
    print("Created: PLC-S-003_請求書_INV-202511-0234.pdf")


# ============================================================
# PLC-S-004 入金消込
# ============================================================
def gen_fb_csv():
    """銀行FBデータ"""
    path = BASE / "PLC-S-004_FB入金データ_202511.csv"
    random.seed(4004)
    cids = list(CUSTOMERS.keys())
    with open(path, "w", encoding="shift_jis", errors="replace") as f:
        f.write("データ区分,日付,口座番号,入金金額,振込依頼人名,摘要,取扱支店\n")
        # カタカナ半角: サンプルｺｷｬｸA など
        # ※SJISで書けない文字は置換されるため、ダミー名はASCIIで書く
        entries = [
            ("3", "20251102", "1234567", "15800000", "SAMPLE KOKYAKU A", "INV-202510-0012", "HONTEN"),
            ("3", "20251104", "1234567", "8420000", "SAMPLE KOKYAKU B", "INV-202510-0034", "HONTEN"),
            ("3", "20251105", "1234567", "3280000", "SAMPLE KOKYAKU C", "INV-202510-0045", "SHITEN-X"),
            ("3", "20251108", "1234567", "22150000", "SAMPLE KOKYAKU H", "INV-202510-0067", "HONTEN"),
            ("3", "20251110", "1234567", "6780000", "SAMPLE KOKYAKU I", "INV-202510-0073", "HONTEN"),
            ("3", "20251115", "1234567", "12450000", "SAMPLE KOKYAKU J", "INV-202510-0089", "SHITEN-Y"),
            ("3", "20251118", "1234567", "4620000", "SAMPLE KOKYAKU K", "INV-202510-0095", "HONTEN"),
            ("3", "20251120", "1234567", "19800000", "SAMPLE KOKYAKU L", "INV-202510-0112", "SHITEN-Z"),
            ("3", "20251122", "1234567", "2850000", "SAMPLE KOKYAKU M", "INV-202510-0125", "SHITEN-V"),
            ("3", "20251125", "1234567", "25480000", "SAMPLE KOKYAKU B", "INV-202510-0089", "HONTEN"),
            ("3", "20251125", "1234567", "7350000", "SAMPLE KOKYAKU D", "INV-202510-0134", "SHITEN-U"),
            ("3", "20251127", "1234567", "1580000", "SAMPLE KOKYAKU E", "INV-202510-0145", "SHITEN-W"),
            ("3", "20251128", "1234567", "33400000", "SAMPLE KOKYAKU A", "INV-202510-0023", "HONTEN"),
            ("3", "20251128", "1234567", "1248000", "SAMPLE KOKYAKU P", "INV-202510-0156", "SHITEN-X"),
        ]
        for e in entries:
            f.write(",".join(e) + "\n")
        total = sum(int(e[3]) for e in entries)
        f.write(f"8,,,{total},,,\n")
    print(f"Created: {path.name}")


def gen_payment_matching():
    """入金消込リスト（経理部作成の統制実施記録）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "入金消込リスト"

    ws.cell(row=1, column=1, value="【PLC-S-004 統制実施記録】 2025年11月 入金消込リスト")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    ws.cell(row=2, column=1, value="作成日: 2025/12/02 / 作成者: 石井 健（経理部担当 ACC006） / 承認: 佐藤 一郎（経理部長 ACC001）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    headers = ["№", "入金日", "入金額(円)", "銀行", "顧客コード", "顧客名",
               "消込対象\n請求書番号", "消込金額(円)", "差額(円)", "消込方法"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[4].height = 32

    random.seed(4004)
    cids = list(CUSTOMERS.keys())
    banks = ["A銀行 本店", "B銀行 支店X", "C銀行 本店", "E銀行 本店"]
    r = 5
    for i in range(1, 31):
        idate = date(2025, 11, random.randint(3, 29))
        amount = random.randint(2_000_000, 40_000_000) // 1000 * 1000
        cid = random.choice(cids)
        cname = CUSTOMERS[cid][0]
        inv = f"INV-202510-{random.randint(1, 150):04d}"
        diff = 0
        method = "SAP自動消込"
        if i in (5, 12, 24):
            method = "手動消込\n(金額部分一致)"
        if i == 9:
            diff = -1_500
            method = "手動消込\n(値引調整)"
        data = [i, idate, amount, random.choice(banks), cid, cname, inv,
                amount - (diff if diff < 0 else 0), diff, method]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 5, 7, 10):
                cell.alignment = C_
                if c_i == 2:
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (3, 8, 9):
                cell.alignment = R_
                cell.number_format = "#,##0;[Red]-#,##0"
        if diff != 0:
            ws.cell(row=r, column=9).fill = FILL_WARN
            ws.cell(row=r, column=10).fill = FILL_WARN
        r += 1

    ws.cell(row=r, column=1, value="... 以下52件省略 ...").font = Font(name="Yu Gothic", size=9, italic=True)

    widths = [5, 12, 14, 14, 12, 18, 16, 14, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-S-004_入金消込リスト_202511.xlsx")
    print("Created: PLC-S-004_入金消込リスト_202511.xlsx")


# ============================================================
# PLC-S-005 売掛金年齢分析
# ============================================================
def gen_ar_aging_xlsx():
    """売掛金年齢表（経理部作成、統制実施記録）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売掛金年齢分析"

    ws.cell(row=1, column=1, value="【PLC-S-005 統制実施記録】 2025年11月末 売掛金年齢分析表")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=2, column=1, value="基準日: 2025/11/30 / 作成: 高橋 美咲（経理部課長 ACC002） / 承認: 佐藤 一郎（経理部長 ACC001）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headers = ["顧客コード", "顧客名", "残高合計(円)", "0-30日\n(正常)", "31-60日",
               "61-90日", "91-120日", "120日超\n(要注意)"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[4].height = 32

    random.seed(5005)
    r = 5
    for cid, (cname, credit, _) in CUSTOMERS.items():
        total = random.randint(3_000_000, min(credit // 2, 200_000_000))
        normal = int(total * 0.7)
        d31 = int(total * 0.2)
        d61 = int(total * 0.06)
        d91 = int(total * 0.03)
        d120 = total - normal - d31 - d61 - d91
        if cid in ("C-10007", "C-10017", "C-10023"):
            d120 = int(total * 0.15)
            normal = total - d31 - d61 - d91 - d120
        data = [cid, cname, total, normal, d31, d61, d91, d120]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i == 1:
                cell.alignment = C_
            elif c_i == 2:
                cell.alignment = L_
            else:
                cell.alignment = R_
                cell.number_format = "#,##0"
        if d120 > 5_000_000:
            ws.cell(row=r, column=8).fill = FILL_NG
        elif d120 > 0:
            ws.cell(row=r, column=8).fill = FILL_WARN
        r += 1

    ws.cell(row=r, column=1, value="合計").font = BBOLD
    ws.cell(row=r, column=1).alignment = C_
    ws.cell(row=r, column=2, value="全20社").font = BBOLD
    for col in range(3, 9):
        ws.cell(row=r, column=col, value=f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{r-1})").font = BBOLD
        ws.cell(row=r, column=col).number_format = "#,##0"
        ws.cell(row=r, column=col).alignment = R_

    r += 2
    ws.cell(row=r, column=1, value="■ 長期滞留債権に関する営業部レビュー").font = BBOLD
    r += 1
    ws.cell(row=r, column=1, value="サンプル顧客G社 (C-10007): 120日超残高あり → 営業部より「年末までに回収見込み」")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    ws.cell(row=r, column=1, value="サンプル顧客N社 (C-10017): 120日超残高あり → 値引交渉中、一部貸倒引当計上検討")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    ws.cell(row=r, column=1, value="サンプル顧客R社 (C-10023): 120日超残高あり → 新規追加案件との相殺予定、経理部と協議中")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    r += 3
    ws.cell(row=r, column=1, value="■ 承認記録").font = BBOLD
    r += 1
    ws.cell(row=r, column=1, value="作成: 高橋 美咲（経理部課長）[印] 2025/12/08")
    r += 1
    ws.cell(row=r, column=1, value="承認: 佐藤 一郎（経理部長）[印] （※紙原本に押印後、スキャンPDFあり）")

    widths = [12, 18, 16, 14, 14, 14, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-S-005_売掛金年齢表_202511.xlsx")
    print("Created: PLC-S-005_売掛金年齢表_202511.xlsx")


def gen_ar_aging_lowres_pdf():
    """承認印が判読不能なPDF（判断保留ケース用）"""
    from PIL import Image, ImageDraw, ImageFilter, ImageFont

    img = Image.new("RGB", (1600, 2200), (255, 255, 255))
    d = ImageDraw.Draw(img)
    fh1 = ImageFont.truetype("C:/Windows/Fonts/YuGothB.ttc", 48)
    fh2 = ImageFont.truetype("C:/Windows/Fonts/YuGothB.ttc", 28)
    fb = ImageFont.truetype("C:/Windows/Fonts/YuGothM.ttc", 20)

    d.text((60, 50), "2025年11月末 売掛金年齢分析表", font=fh1, fill=(20, 20, 60))
    d.text((60, 130), "基準日: 2025/11/30 / 作成: 経理部課長 高橋 美咲", font=fb, fill=(60, 60, 60))

    headers = ["顧客コード", "顧客名", "残高合計(円)", "0-30日", "31-60日", "61-90日", "91-120日", "120日超"]
    col_x = [60, 220, 500, 720, 900, 1080, 1260, 1440]
    col_w = [160, 280, 220, 180, 180, 180, 180, 160]
    y0 = 200
    d.rectangle([60, y0, 60 + sum(col_w), y0 + 50], fill=(31, 78, 120))
    for i, h in enumerate(headers):
        d.text((col_x[i] + 10, y0 + 12), h, font=fh2, fill=(255, 255, 255))

    samples = [
        ("C-10001", "サンプル顧客A社", "128,540,000", "90,000,000", "26,000,000", "8,500,000", "3,000,000", "1,040,000"),
        ("C-10002", "サンプル顧客B社", "87,320,000", "61,200,000", "17,400,000", "5,200,000", "2,800,000", "720,000"),
        ("C-10003", "サンプル顧客C社", "56,780,000", "39,700,000", "11,300,000", "3,400,000", "1,700,000", "680,000"),
        ("C-10007", "サンプル顧客G社", "23,450,000", "12,000,000", "4,500,000", "2,100,000", "1,300,000", "3,550,000"),
        ("C-10011", "サンプル顧客H社", "156,890,000", "109,800,000", "31,400,000", "9,400,000", "4,700,000", "1,590,000"),
        ("C-10017", "サンプル顧客N社", "42,180,000", "22,000,000", "8,500,000", "3,200,000", "1,900,000", "6,580,000"),
        ("C-10023", "サンプル顧客R社", "36,720,000", "19,500,000", "7,300,000", "2,800,000", "1,600,000", "5,520,000"),
    ]
    y = y0 + 50
    for r_idx, row in enumerate(samples):
        bg = (255, 255, 255) if r_idx % 2 == 0 else (240, 245, 252)
        d.rectangle([60, y, 60 + sum(col_w), y + 45], fill=bg, outline=(200, 200, 200))
        for i, v in enumerate(row):
            d.text((col_x[i] + 10, y + 12), v, font=fb, fill=(20, 20, 20))
        y += 45

    y += 60
    d.text((60, y), "■ 承認記録", font=fh2, fill=(20, 20, 60))
    y += 50
    d.text((60, y), "作成: 高橋 美咲（経理部課長）", font=fb, fill=(40, 40, 40))
    d.ellipse([520, y - 5, 620, y + 45], outline=(200, 30, 30), width=3)
    d.text((540, y + 8), "高橋", font=fh2, fill=(200, 30, 30))
    d.text((720, y + 10), "2025/12/08", font=fb, fill=(40, 40, 40))

    y += 70
    d.text((60, y), "承認: 佐藤 一郎（経理部長）", font=fb, fill=(40, 40, 40))
    d.ellipse([520, y - 5, 620, y + 45], outline=(200, 30, 30), width=3)
    d.text((548, y + 8), "佐藤", font=fh2, fill=(200, 30, 30))
    d.text((720, y + 10), "2025/12/??", font=fb, fill=(40, 40, 40))

    # 承認印をぼかす
    box = (500, y - 20, 900, y + 60)
    crop = img.crop(box)
    crop = crop.filter(ImageFilter.GaussianBlur(radius=4))
    small = crop.resize((crop.width // 6, crop.height // 6))
    crop = small.resize(crop.size, Image.NEAREST)
    img.paste(crop, box)

    d.rectangle([1340, 40, 1550, 120], outline=(50, 50, 50), width=2)
    d.text((1360, 55), "SCAN", font=fh2, fill=(50, 50, 50))
    d.text((1360, 85), "2025/12/10", font=fb, fill=(50, 50, 50))

    img_path = BASE / "_temp_aging.png"
    img.save(img_path, "PNG")

    pdf = JPPDF(orientation="P", format="A4")
    pdf.add_page()
    pdf.set_font("YuGoth", "B", 14)
    pdf.cell(0, 8, "売掛金年齢分析表 【経理部長承認済】", align="C",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    pdf.image(str(img_path), x=10, y=20, w=190)
    pdf.set_y(240)
    pdf.set_font("YuGoth", "", 8)
    pdf.cell(0, 5, "※ 本書類は原本をPDF化したものです", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "保管: 経理部 / 文書番号: AR-AGE-202511-001", new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(BASE / "PLC-S-005_売掛金年齢表_経理部長承認PDF_低解像度.pdf"))
    img_path.unlink()
    print("Created: PLC-S-005_売掛金年齢表_経理部長承認PDF_低解像度.pdf")


# ============================================================
# PLC-S-006 期末カットオフ
# ============================================================
def gen_cutoff_test():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "期末カットオフテスト"

    ws.cell(row=1, column=1, value="【PLC-S-006 統制実施記録】 FY2025期末 売上カットオフテスト")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    ws.cell(row=2, column=1, value="実施日: 2026/4/3 / 実施者: 佐藤 一郎（経理部長 ACC001） / 対象期間: 2026/3/25～2026/4/1 の出荷41件全数")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    headers = ["№", "受注番号", "出荷日", "売上計上日", "顧客", "金額(円)",
               "計上期", "期間帰属", "判定"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[4].height = 24

    random.seed(6006)
    cids = list(CUSTOMERS.keys())
    r = 5
    for i in range(1, 42):
        dayidx = random.choice([-6, -5, -4, -3, -2, -1, 0, 1])
        ship_date = date(2026, 3, 31) + timedelta(days=dayidx)
        sale_date = ship_date + timedelta(days=random.choice([0, 0, 1]))
        cid = random.choice(cids)
        amount = random.randint(2_000_000, 20_000_000)
        fy = "FY2025" if sale_date.year < 2026 or (sale_date.year == 2026 and sale_date.month <= 3) else "FY2026"
        expected_fy = "FY2025" if ship_date.year < 2026 or (ship_date.year == 2026 and ship_date.month <= 3) else "FY2026"
        judge = "適切" if fy == expected_fy else "要調整"
        data = [i, f"ORD-2026-{3000 + i}", ship_date, sale_date,
                CUSTOMERS[cid][0], amount, fy, expected_fy, judge]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 7, 8, 9):
                cell.alignment = C_
                if c_i in (3, 4):
                    cell.number_format = "yyyy/mm/dd"
            elif c_i == 6:
                cell.alignment = R_
                cell.number_format = "#,##0"
        if judge == "適切":
            ws.cell(row=r, column=9).fill = PatternFill("solid", fgColor="E2EFDA")
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="結論: 41件すべて適切な期間に計上されており、カットオフ違反なし。").font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    ws.cell(row=r, column=1, value="実施: 佐藤 一郎 [印] 2026/4/3 / 承認: 渡辺 正博 CFO [印] 2026/4/5")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)

    widths = [5, 14, 12, 12, 18, 14, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-S-006_期末カットオフテスト.xlsx")
    print("Created: PLC-S-006_期末カットオフテスト.xlsx")


# ============================================================
# PLC-S-007 価格マスタ承認
# ============================================================
def gen_price_change_ringi():
    """価格マスタ変更の稟議書PDF"""
    pdf = JPPDF()
    pdf.add_page()

    pdf.set_font("YuGoth", "B", 16)
    pdf.cell(0, 10, "稟 議 書", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, "稟議番号: W-2025-1876", align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    pdf.kv("件名", "サンプル顧客H社向け製品P-30011の単価改定申請", key_w=30)
    pdf.kv("申請日", "2025年10月15日", key_w=30)
    pdf.kv("申請者", "松本 香織（営業部主任 SLS004）", key_w=30)
    pdf.kv("承認種別", "価格マスタ変更（顧客別単価）", key_w=30)
    pdf.ln(5)

    pdf.h2("1. 変更内容")
    pdf.table_header(["項目", "変更前", "変更後", "変更率"],
                     [50, 45, 45, 40])
    pdf.table_row(["製品 P-30011", "¥18,500", "¥19,200", "+3.8%"],
                  [50, 45, 45, 40])
    pdf.table_row(["顧客", "C-10011", "（同一）", "-"], [50, 45, 45, 40], fill=True)
    pdf.ln(5)

    pdf.h2("2. 変更理由")
    pdf.body("原材料費上昇（サンプル仕入先H社 V-20008 からの特殊合金材の価格改定通知）に対応するため、"
             "顧客サンプル顧客H社への納入単価を+3.8%（¥700）上げる旨を交渉・合意した。"
             "年間販売数量約4,800本、年間売上増加額は ¥3,360,000 の見込み。")
    pdf.ln(3)

    pdf.h2("3. 適用開始日")
    pdf.body("稟議承認後、次期受注分より適用")
    pdf.ln(5)

    pdf.h2("4. 承認経路")
    pdf.set_font("YuGoth", "B", 10)
    pdf.cell(40, 7, "役割", border=1, align="C", fill=True)
    pdf.cell(45, 7, "氏名", border=1, align="C", fill=True)
    pdf.cell(40, 7, "承認日時", border=1, align="C", fill=True)
    pdf.cell(30, 7, "承認印", border=1, align="C", fill=True, new_x="LMARGIN", new_y="NEXT")
    approvals = [
        ("営業部課長", "斎藤 次郎 (SLS002)", "2025/10/15 14:30"),
        ("営業本部長", "田中 太郎 (SLS001)", "2025/10/16 10:15"),
    ]
    for role, name, dt in approvals:
        pdf.set_font("YuGoth", "", 10)
        pdf.cell(40, 14, role, border=1, align="C")
        pdf.cell(45, 14, name, border=1, align="C")
        pdf.cell(40, 14, dt, border=1, align="C")
        x_stamp = pdf.get_x()
        y_stamp = pdf.get_y()
        pdf.cell(30, 14, "", border=1, new_x="LMARGIN", new_y="NEXT")
        pdf.stamp("承認", x=x_stamp + 15, y=y_stamp + 7)

    pdf.ln(5)
    pdf.set_font("YuGoth", "", 9)
    pdf.multi_cell(0, 5, "【添付資料】原材料費上昇根拠資料（V-20008 サンプル仕入先H社からの値上げ通知）、過去3年の販売単価推移表")

    pdf.output(str(BASE / "PLC-S-007_価格マスタ変更稟議_W-2025-1876.pdf"))
    print("Created: PLC-S-007_価格マスタ変更稟議_W-2025-1876.pdf")


def gen_price_history():
    """価格変更履歴レポート（SAP VK12から出力される経理部レビュー資料）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "価格マスタ変更履歴"

    ws.cell(row=1, column=1, value="【PLC-S-007 統制実施記録】 FY2025 Q3 価格マスタ変更履歴レポート")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    ws.cell(row=2, column=1, value="対象期間: 2025/10/1～2025/12/31 / 出力元: SAP VK12履歴 / 経理部レビュー者: 中村 真理 ACC004")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    headers = ["変更№", "変更日", "製品コード", "顧客コード", "旧単価(円)", "新単価(円)",
               "変更率", "稟議番号", "承認者"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[4].height = 28

    samples = [
        (1, date(2025, 10, 16), "P-30011", "C-10011", 18500, 19200, "+3.8%", "W-2025-1876", "田中 太郎"),
        (2, date(2025, 10, 22), "P-30014", "C-10011", 38500, 40000, "+3.9%", "W-2025-1899", "田中 太郎"),
        (3, date(2025, 11, 5), "P-30020", "C-10012", 42500, 44000, "+3.5%", "W-2025-1945", "田中 太郎"),
        (4, date(2025, 11, 12), "P-30006", "C-10002", 12500, 12800, "+2.4%", "W-2025-2012", "田中 太郎"),
        (5, date(2025, 11, 20), "P-30022", "C-10003", 1280, 1300, "+1.6%", "W-2025-2067", "田中 太郎"),
        (6, date(2025, 12, 3), "P-30008", "C-10001", 8200, 8350, "+1.8%", "W-2025-2132", "田中 太郎"),
        (7, date(2025, 12, 10), "P-30015", "C-10015", 12800, 12500, "-2.3%", "W-2025-2178", "田中 太郎"),
        (8, date(2025, 12, 18), "P-30013", "C-10012", 28500, 29200, "+2.5%", "W-2025-2224", "田中 太郎"),
    ]
    r = 5
    for d in samples:
        for c_i, v in enumerate(d, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 7, 8):
                cell.alignment = C_
                if c_i == 2:
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (5, 6):
                cell.alignment = R_
                cell.number_format = "#,##0"
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="■ レビュー結果").font = BBOLD
    r += 1
    ws.cell(row=r, column=1, value="・全8件とも稟議承認あり、承認ルート（営業課長→本部長）適切")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    ws.cell(row=r, column=1, value="・SAPマスタへの反映日とSAP登録日の一致を全件確認済")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 2
    ws.cell(row=r, column=1, value="レビュー実施: 中村 真理 [印] 2026/1/10 / 承認: 高橋 美咲 [印] 2026/1/12")

    widths = [6, 12, 12, 12, 12, 12, 10, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-S-007_価格変更履歴レポート_Q3.xlsx")
    print("Created: PLC-S-007_価格変更履歴レポート_Q3.xlsx")


# ============================================================
# 営業本部の月次会議議事録（被評価部門が保管している業務記録）
# ============================================================
def gen_meeting_minutes():
    content = """# 2025年11月 月次売上分析会議 議事録

| 項目 | 内容 |
|------|------|
| 日時 | 2025年12月8日（月）14:00〜15:30 |
| 場所 | 本社3階 第1会議室 + Teams |
| 議長 | 田中 太郎（営業本部長 SLS001） |
| 書記 | 松本 香織（営業部主任 SLS004） |
| 参加者 | 田中、斎藤、藤田、松本、井上、佐藤、高橋、中村、渡辺CFO（一部オンライン） |

---

## 1. 11月実績サマリ

| 項目 | 金額 | 前月比 | 前年同月比 |
|------|------|-------|----------|
| 売上高 | ¥3,384,560,000 | +8.2% | +5.4% |
| 売上総利益 | ¥846,140,000 | +9.1% | +6.1% |
| 売上総利益率 | 25.0% | +0.2pt | +0.2pt |
| 新規受注 | ¥3,820,000,000 | +12.5% | +8.8% |
| 受注残 | ¥18,420,000,000 | +2.4% | +15.2% |

### 部門別売上
- 自動車業界向け：¥1,845,200,000（前月比 +6.8%）
- 半導体装置向け：¥1,362,480,000（前月比 +12.2%）※サンプル顧客H社案件の立ち上がり影響
- 商社経由：¥176,880,000（前月比 +4.5%）

## 2. 主要トピック

### 2-1. サンプル顧客H社への新製品納入開始（藤田課長）
- 製品 P-30011「ウェハー搬送ロボット用シャフト A」の月間納入 80本開始
- 来年度は月間 200本以上の見込み
- 生産能力の見直しについて製造本部と調整開始

### 2-2. サンプル顧客G社 (C-10007) の長期滞留債権について（高橋課長）
- 120日超の残高 ¥3,550,000 を認識
- 営業部井上担当から、同社との定例会議にて「年末までに回収見込み、一部内入金確認」との報告
- 経理部は当面貸倒引当金計上せず、動向を注視

### 2-3. 価格改定の進捗（斎藤課長）
- Q3中に完了予定の8社への価格改定、7社合意済
- 残1社（サンプル顧客E社）は交渉継続中
- 原材料費上昇の累計転嫁率：62%（目標80%に届かず、継続交渉が課題）

## 3. 経理部からの報告（佐藤部長）

### 3-1. 入金消込状況
- 11月入金額：¥3,256,800,000
- 自動消込率：94.2%
- 手動消込：63件（主に金額一致しない値引調整等）

### 3-2. 期末カットオフに向けた注意喚起
- 3月末出荷分の売上計上タイミングは規程通り実施
- 特に3月下旬の大型受注について、出荷・納品のタイミングに注意

### 3-3. 売掛金年齢分析について（高橋課長）
- 11月末時点で120日超残高は合計 ¥18,690,000
- **注意**: 11月分の年齢分析表のPDFスキャンが一部不鮮明
- 再スキャンを実施することを確認（内部監査からの要求にも対応）

## 4. 決定事項

| No. | 決定事項 | 担当 | 期限 |
|-----|---------|------|------|
| 1 | P-30011 の生産能力見直し検討 | 藤田・製造本部 | 2026/1/31 |
| 2 | 売掛金年齢分析表の再スキャン実施 | 高橋・中村 | 2025/12/20 |
| 3 | 期末カットオフ調整会議の開始 | 斎藤・森・中村 | 2025/12/15〜 |
| 4 | サンプル顧客E社への価格改定の交渉 | 井上 | 2025/12/26 |

## 5. 次回予定

- 日時：2026年1月12日（月）14:00
- 場所：本社3階 第1会議室

---

**配布先**: 参加者全員、取締役（CC）、内部監査室（CC）

**承認**:
- 田中 太郎（営業本部長）[印] 2025/12/10
- 渡辺 正博（CFO）[印] 2025/12/11
"""
    path = BASE / "PLC-S_月次売上会議_議事録_202511.md"
    path.write_text(content, encoding="utf-8")
    print(f"Created: {path.name}")


# ============================================================
# Main
# ============================================================
if __name__ == "__main__":
    # PLC-S-001 受注・与信承認
    gen_population_orders()
    gen_credit_limit_master()
    gen_approval_authority()
    gen_credit_check_log()
    gen_all_order_pdfs()
    gen_screenshots()

    # PLC-S-002 出荷-売上マッチング
    gen_shipment_data()
    gen_sales_posting_data()
    gen_match_report()
    gen_unmatch_csv()

    # PLC-S-003 請求書発行
    gen_invoice_batch_log()
    gen_invoice_monthly_list()
    gen_invoice_pdf()

    # PLC-S-004 入金消込
    gen_fb_csv()
    gen_payment_matching()

    # PLC-S-005 売掛金年齢分析
    gen_ar_aging_xlsx()
    gen_ar_aging_lowres_pdf()

    # PLC-S-006 期末カットオフ
    gen_cutoff_test()

    # PLC-S-007 価格マスタ承認
    gen_price_change_ringi()
    gen_price_history()

    # 業務運営記録
    gen_meeting_minutes()

    print("\nAll PLC-S evidence (v2) generated.")
