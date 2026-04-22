"""会社名を明示的なダミー名に一括置換

置換対象:
株式会社テクノプレシジョン → デモA株式会社
テクノプレシジョン東北株式会社 → デモA東北株式会社
テクノプレシジョン → デモA
TP物流サービス → デモA物流サービス
TPトレーディング → デモAトレーディング
TechnoPrecision (Thailand) Co., Ltd. → Demo-A (Thailand) Co., Ltd.
TechnoPrecision Thailand → Demo-A Thailand
TechnoPrecision Inc. → Demo-A Inc.
TechnoPrecision → Demo-A
sample-tp.co.jp → sample-demoa.co.jp

子会社コード (opaque codes ですが敢えて変更):
TP-TB → DA-TB
TP-LOG → DA-LOG
TPTR → DATR
TPT → DAT
"""
import os
import sys
import io
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

ROOT = Path(r"C:\Users\nyham\work\demo_data")

# 順序重要: 長い文字列から置換
REPLACEMENTS = [
    # フル会社名 (長い順)
    ('株式会社テクノプレシジョン', 'デモA株式会社'),
    ('テクノプレシジョン東北株式会社', 'デモA東北株式会社'),
    ('テクノプレシジョン東北', 'デモA東北'),
    ('TP物流サービス株式会社', 'デモA物流サービス株式会社'),
    ('TP物流サービス', 'デモA物流サービス'),
    ('TPトレーディング株式会社', 'デモAトレーディング株式会社'),
    ('TPトレーディング', 'デモAトレーディング'),
    # 英語名
    ('TechnoPrecision (Thailand) Co., Ltd.', 'Demo-A (Thailand) Co., Ltd.'),
    ('TechnoPrecision Thailand', 'Demo-A Thailand'),
    ('TechnoPrecision Inc.', 'Demo-A Inc.'),
    ('TechnoPrecision', 'Demo-A'),
    # 残りの日本語
    ('テクノプレシジョン', 'デモA'),
    # 子会社コード (TPTR → DATR を TPT より先に実行)
    ('TPTR', 'DATR'),
    ('TP-TB', 'DA-TB'),
    ('TP-LOG', 'DA-LOG'),
    ('TPT', 'DAT'),
    # Email domain
    ('sample-tp.co.jp', 'sample-demoa.co.jp'),
]


def replace_text(content):
    new = content
    for old, rep in REPLACEMENTS:
        new = new.replace(old, rep)
    return new


changed_files = []


def process_text_file(path):
    try:
        with open(path, encoding='utf-8-sig') as f:
            content = f.read()
    except (UnicodeDecodeError, IsADirectoryError):
        return False
    new = replace_text(content)
    if new != content:
        with open(path, 'w', encoding='utf-8-sig', newline='') as f:
            f.write(new)
        return True
    return False


def process_xlsx_file(path):
    try:
        wb = load_workbook(path)
    except Exception:
        return False
    changed = False
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    new_v = replace_text(cell.value)
                    if new_v != cell.value:
                        cell.value = new_v
                        changed = True
    if changed:
        wb.save(path)
    return changed


# Walk
for root, dirs, files in os.walk(ROOT):
    # Skip git/cache/scripts
    if '.git' in root or '__pycache__' in root:
        continue
    for f in files:
        fp = Path(root) / f
        rel = fp.relative_to(ROOT)
        # Skip the rename script itself
        if str(rel).replace('\\', '/').startswith('_scripts/rename_company.py'):
            continue

        ext = f.lower()
        if ext.endswith(('.md', '.csv', '.txt', '.py', '.json')):
            if process_text_file(fp):
                changed_files.append(str(rel))
        elif ext.endswith('.xlsx'):
            if process_xlsx_file(fp):
                changed_files.append(str(rel))

print(f"=== TEXT/XLSX 置換完了: {len(changed_files)} ファイル ===")
for f in sorted(changed_files):
    print(f'  - {f}')
