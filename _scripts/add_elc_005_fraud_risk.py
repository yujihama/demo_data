"""
ELC-005 不正リスク評価シートを既存の全社リスクアセスメント xlsx に追加する。

背景:
  _scripts/gen_remaining_evidence.py::gen_elc_risk_assessment() は ELC-004 の
  全社リスクのみを生成しており、ELC-005（不正リスクファクター = 動機・機会・
  正当化、および不正シナリオ評価、内部監査室と経理部の合同実施）の記録が
  欠落していた。本スクリプトは既存ファイルを保ったまま、ELC-005 用シートを
  追補する。再現性確保のため gen_remaining_evidence.py にも同等ロジックを
  追記している。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

BASE_ELC = Path(__file__).resolve().parent.parent / "4.evidence" / "ELC"
XLSX_PATH = BASE_ELC / "全社リスクアセスメント結果_2025年度.xlsx"

HF = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
BFONT = Font(name="Yu Gothic", size=10)
BBOLD = Font(name="Yu Gothic", size=10, bold=True)
C_ = Alignment(horizontal="center", vertical="center", wrap_text=True)
L_ = Alignment(horizontal="left", vertical="center", wrap_text=True)
T_ = Side("thin", color="888888")
BRD = Border(left=T_, right=T_, top=T_, bottom=T_)
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")
FILL_NG = PatternFill("solid", fgColor="FCE4D6")


def add_fraud_risk_sheet():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(XLSX_PATH)

    wb = openpyxl.load_workbook(XLSX_PATH)
    if "不正リスク評価" in wb.sheetnames:
        del wb["不正リスク評価"]
    ws = wb.create_sheet("不正リスク評価")

    # タイトル
    ws.cell(row=1, column=1, value="【ELC-005 統制実施記録】 2025年度 不正リスク評価シート")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=2, column=1,
            value="実施日: 2025/6/18 / 実施者: 内部監査室（長谷川 剛 IA001・大塚 美穂 IA002）・"
                  "経理部（佐藤 一郎 ACC001・高橋 美咲 ACC002）合同 / 取締役会報告: 2025/6/30")
    ws.cell(row=2, column=2).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    ws.cell(row=3, column=1,
            value="根拠: R03 内部監査規程 §4 / 統制記述: 決算期前に不正リスクファクター"
                  "（動機・機会・正当化）を検討し、重要拠点・勘定について不正シナリオを評価する。")
    ws.cell(row=3, column=1).font = Font(name="Yu Gothic", size=9, italic=True, color="555555")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)

    # ===== セクション1: 不正リスクファクター（Fraud Triangle）評価 =====
    ws.cell(row=5, column=1, value="■ セクション1: 不正リスクファクター評価（Fraud Triangle）")
    ws.cell(row=5, column=1).font = BBOLD
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=8)

    headers1 = ["№", "対象プロセス/勘定", "動機（Pressure）",
                "機会（Opportunity）", "正当化（Rationalization）",
                "総合評価", "評価根拠", "リンク統制"]
    for i, h in enumerate(headers1, 1):
        c = ws.cell(row=6, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    fraud_triangle = [
        ("FT-01", "売上計上（PLC-S）",
         "中：半期業績コミットメントのプレッシャー",
         "中：与信超過受注の個別承認が属人的",
         "低：営業現場に「期ずれ調整」の慣習は限定的",
         "中",
         "ITAC-001自動統制に依存する構造のため個別承認ログをモニタリングで補完",
         "PLC-S-001 / ITAC-001"),
        ("FT-02", "購買発注（PLC-P）",
         "中：仕入先からのリベート誘因リスク",
         "高：SAPロールでSoD違反ユーザが存在する可能性",
         "中：「効率化のため一時的に権限を拡張」との正当化",
         "高",
         "PUR004に関する職務分掌違反の懸念を内部監査室が認識済。ITGC-AC-004の補完統制と合わせて重点監視",
         "PLC-P-002 / ELC-007 / ITGC-AC-004"),
        ("FT-03", "棚卸資産評価（PLC-I）",
         "低：原価変動の業績影響は限定的",
         "中：倉庫課による差異分析と経理部の連携に依存",
         "中：「軽微差異は調整仕訳で十分」との慣習",
         "中",
         "倉庫差異調整の原因分析・経理報告プロセスを重点テスト対象に設定",
         "PLC-I-001 / PLC-I-002"),
        ("FT-04", "会計上見積り（FCRP）",
         "中：四半期利益目標達成プレッシャー",
         "中：貸倒・滞留評価は経理部課長の裁量余地あり",
         "低：「保守的評価の範囲内」との正当化",
         "中",
         "見積前提の外部情報・内部情報の根拠資料を重点確認",
         "FCRP-003 / PLC-I-005"),
        ("FT-05", "連結仕訳",
         "中：連結利益調整余地の存在",
         "中：非定型連結仕訳のレビューが経理部長の単独判断",
         "低：「実務慣行」として正当化される余地",
         "中",
         "連結パッケージのバリデーションで補完。非定型仕訳のレビュー記録を重点確認",
         "FCRP-004 / ITAC-005"),
        ("FT-06", "経費精算・交際費",
         "低：個人レベルの生活資金プレッシャー",
         "中：承認者の確認が形式的になるリスク",
         "低：「業界慣行」との正当化",
         "低",
         "金額閾値超過分のみ抽出レビュー",
         "PLC-P-006"),
        ("FT-07", "固定資産計上・減損",
         "低：減損回避の業績プレッシャー（現時点では低）",
         "中：見積の主観性（将来キャッシュフロー）",
         "低：「保守的評価の過大解釈」",
         "中",
         "ITAC-003自動計算に依存。減損兆候判定の前提を重点確認",
         "FCRP-003 / ITAC-003"),
        ("FT-08", "海外子会社取引（タイTPT）",
         "中：現地目標達成プレッシャー",
         "中：親会社モニタリング頻度の低さ",
         "中：「現地商慣習」としての正当化",
         "中",
         "FCRP-002連結パッケージ検証で補完。現地内部監査を年1回実施",
         "FCRP-002 / ELC-010"),
    ]
    r = 7
    for row in fraud_triangle:
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font = BFONT; cell.border = BRD
            cell.alignment = C_ if ci in (1, 6) else L_
        # 総合評価の色
        total = row[5]
        if total == "高":
            ws.cell(row=r, column=6).fill = FILL_NG
        elif total == "中":
            ws.cell(row=r, column=6).fill = FILL_WARN
        r += 1

    r += 2

    # ===== セクション2: 重要拠点・勘定別 不正シナリオ評価 =====
    ws.cell(row=r, column=1, value="■ セクション2: 重要拠点・勘定別 不正シナリオ評価")
    ws.cell(row=r, column=1).font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1

    headers2 = ["シナリオ№", "重要拠点", "重要勘定", "不正シナリオ",
                "発生可能性", "影響度", "優先度", "対応策"]
    for i, h in enumerate(headers2, 1):
        c = ws.cell(row=r, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    r += 1

    scenarios = [
        ("F-001", "本社", "売上高・売掛金",
         "期末前倒し出荷による売上架空計上（カットオフ違反）",
         "低", "大", "中",
         "PLC-S-006 期末カットオフ統制で全数検証 / 監査人サンプル実施"),
        ("F-002", "本社", "買掛金・仕入",
         "架空仕入先への発注・送金（仕入先マスタ不正登録）",
         "低", "大", "中",
         "PLC-P-005 仕入先マスタ管理 + 反社チェック強化"),
        ("F-003", "本社倉庫A/B", "棚卸資産",
         "実地棚卸帳簿操作による在庫水増し",
         "低", "大", "中",
         "PLC-I-001 経理部立会 + 抽取検査 / 監査人立会予定"),
        ("F-004", "タイ TPT", "売上高・売掛金",
         "現地取引先への値引きを用いた裏金化",
         "中", "中", "中",
         "海外子会社内部監査（年1回）+ 連結パッケージ検証"),
        ("F-005", "本社", "販売費及び一般管理費",
         "接待交際費の私的流用・架空計上",
         "中", "中", "中",
         "金額閾値超過分のサンプルレビュー / 内部通報窓口周知"),
        ("F-006", "本社", "連結利益",
         "非定型連結仕訳による利益操作",
         "低", "大", "中",
         "FCRP-004 連結仕訳承認で2段階レビュー徹底"),
        ("F-007", "本社", "貸倒引当金",
         "回収不能債権の引当過少設定による利益操作",
         "低", "中", "中",
         "FCRP-003 + 監査法人連携で前提確認"),
        ("F-008", "東北子会社", "製造原価",
         "仕掛品評価の恣意的操作",
         "低", "中", "中",
         "PLC-I-007 月次締め + 連結パッケージ検証"),
    ]
    for row in scenarios:
        for ci, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font = BFONT; cell.border = BRD
            cell.alignment = C_ if ci in (1, 5, 6, 7) else L_
        if row[6] == "高":
            ws.cell(row=r, column=7).fill = FILL_NG
        elif row[6] == "中":
            ws.cell(row=r, column=7).fill = FILL_WARN
        r += 1

    r += 2

    # ===== セクション3: 総合評価と対応方針 =====
    ws.cell(row=r, column=1, value="■ セクション3: 総合評価と対応方針")
    ws.cell(row=r, column=1).font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1

    summary_lines = [
        "1. 最重点領域: 購買発注プロセス（FT-02）。SoD違反（PUR004）の是正状況を"
        "ITGC-AC-004・PLC-P-002と連携して継続監視する。",
        "2. 重点領域: 売上カットオフ（F-001）、棚卸在庫（FT-03/F-003）、会計上見積り"
        "（FT-04/F-007）、連結仕訳（FT-05/F-006）、海外子会社取引（FT-08/F-004）。",
        "3. 共通: 内部通報窓口の周知を四半期に1回実施し、正当化を抑制する。",
        "4. 監査計画への反映: 本評価結果を内部監査年次計画（ELC-010）および"
        "外部監査人との協議に反映済み（2025/6/25協議）。",
    ]
    for line in summary_lines:
        ws.cell(row=r, column=1, value=line).font = BFONT
        ws.cell(row=r, column=1).alignment = L_
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        ws.row_dimensions[r].height = 30
        r += 1

    r += 2
    ws.cell(row=r, column=1,
            value="承認: 内部監査室長 [印] 長谷川 剛 2025/6/20 / "
                  "経理部長 [印] 佐藤 一郎 2025/6/20 / "
                  "取締役会審議: 2025/6/30（第242回）/ "
                  "監査等委員会報告: 2025/6/27")
    ws.cell(row=r, column=1).font = BFONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    # 列幅
    widths = [9, 18, 24, 28, 28, 9, 38, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(XLSX_PATH)
    print(f"Updated: {XLSX_PATH}")


if __name__ == "__main__":
    add_fraud_risk_sheet()
