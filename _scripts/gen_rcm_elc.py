"""
ELC (全社統制) RCM生成
COSO 6要素 × 12統制
"""
import openpyxl
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent))
from rcm_common import (init_rcm_sheet, write_rcm_row, add_legend_sheet,
                         FILL_SUB_HEADER, Font, Alignment, PatternFill)

BASE = Path(r"C:\Users\nyham\work\demo_data\2.RCM")
BASE.mkdir(parents=True, exist_ok=True)

# ELC用の列構成は標準から「影響勘定科目」「アサーション」を変更し、
# 「COSO要素」を強調したシンプル版を使用
from rcm_common import (HEADER_FILL, HEADER_FONT, BODY_FONT, CENTER_WRAP,
                        LEFT_WRAP, BORDER, BORDER_HEADER, FILL_KEY, FILL_OK)
from openpyxl.utils import get_column_letter

ELC_COLUMNS = [
    ("統制ID", 10),
    ("COSO要素", 18),
    ("評価項目", 30),
    ("リスク記述", 42),
    ("統制活動", 48),
    ("統制タイプ", 10),
    ("頻度", 10),
    ("キー\nコントロール", 10),
    ("実施者/\n責任部門", 18),
    ("実施証跡\n(エビデンス)", 28),
    ("関連規程", 14),
    ("整備状況\n評価結果", 14),
    ("運用状況\n評価結果", 14),
    ("不備の\n有無", 10),
    ("最終結論", 16),
    ("評価日/評価者", 18),
]


def init_elc_sheet(ws):
    ws.sheet_view.zoomScale = 90
    ws.cell(row=1, column=1, value="【ELC】全社統制 リスク・コントロール・マトリクス（RCM）")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ELC_COLUMNS))

    ws.cell(row=2, column=1, value="評価対象: 株式会社テクノプレシジョン 親会社  /  評価期間: FY2025 (2025/4/1 - 2026/3/31)  /  作成日: 2026/04/10  /  作成者: 内部監査室")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(ELC_COLUMNS))

    for i, (name, width) in enumerate(ELC_COLUMNS, 1):
        cell = ws.cell(row=4, column=i, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        cell.border = BORDER_HEADER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[4].height = 36
    ws.freeze_panes = "C5"


def write_elc_row(ws, row_num, values, key_control=False, status="ok"):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=c, value=v)
        cell.font = BODY_FONT
        cell.border = BORDER
        if c in (1, 2, 6, 7, 8, 9, 11, 12, 13, 14, 15, 16):
            cell.alignment = CENTER_WRAP
        else:
            cell.alignment = LEFT_WRAP

    if key_control:
        ws.cell(row=row_num, column=1).fill = FILL_KEY
        ws.cell(row=row_num, column=8).fill = FILL_KEY

    if status == "ok":
        for c in (12, 13):
            ws.cell(row=row_num, column=c).fill = FILL_OK

    ws.row_dimensions[row_num].height = 80


def write_coso_divider(ws, row_num, text):
    """COSO要素ごとの区切り行"""
    ws.cell(row=row_num, column=1, value=text)
    ws.cell(row=row_num, column=1).font = Font(name="Yu Gothic", size=11, bold=True, color="FFFFFF")
    ws.cell(row=row_num, column=1).fill = PatternFill("solid", fgColor="5B9BD5")
    ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(ELC_COLUMNS))
    ws.row_dimensions[row_num].height = 24


def gen_elc():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ELC_RCM"
    init_elc_sheet(ws)

    # 評価日・評価者の共通値
    eval_info = "2026/2/15\n長谷川 剛"

    rows = []

    # ===== I. 統制環境 =====
    rows.append(("DIVIDER", "■ I. 統制環境（Control Environment）"))
    rows.append((
        ["ELC-001", "I. 統制環境", "取締役会の機能",
         "取締役会による経営監督機能の不十分",
         "取締役会は毎月1回開催され、経営成績・重要案件の審議と意思決定を行う。社外取締役3名を含む7名構成。議事録を作成し、監査等委員会へ供覧する。",
         "予防的", "月次", "Y", "取締役会\n/総務部",
         "取締役会議事録、\n上程資料、\n議案決議書",
         "R01 取締役会規則", "有効", "有効", "なし",
         "整備・運用ともに有効", eval_info],
        True, "ok"))
    rows.append((
        ["ELC-002", "I. 統制環境", "倫理綱領の浸透",
         "役職員の倫理意識欠如による不正・違反行為",
         "倫理綱領を全役職員に周知し、年1回の受領確認書を提出させる。新入社員研修および管理職昇格時研修で説明を実施する。",
         "予防的", "年次", "N", "総務部",
         "倫理綱領受領確認書、\n研修受講記録",
         "R08 倫理綱領", "有効", "有効", "なし",
         "整備・運用ともに有効", eval_info],
        False, "ok"))
    rows.append((
        ["ELC-003", "I. 統制環境", "職務権限と組織体制",
         "職務権限の曖昧さによる不正・誤謬の発生",
         "職務権限規程により各階層の承認権限（金額・業務別）を明文化し、定期的に見直す。組織変更時には遅滞なく更新する。",
         "予防的", "随時\n(年1回見直し)", "Y", "総務部",
         "職務権限規程、\n組織図、\n承認権限一覧表",
         "R18 職務権限規程", "有効", "有効", "なし",
         "整備・運用ともに有効", eval_info],
        True, "ok"))
    # ===== II. リスク評価と対応 =====
    rows.append(("DIVIDER", "■ II. リスクの評価と対応（Risk Assessment）"))
    rows.append((
        ["ELC-004", "II. リスク評価", "全社リスク評価の実施",
         "重要リスクの見落としによる対応遅延・損失発生",
         "経営企画部が全社リスクを年1回洗い出し、財務・事業・コンプライアンス等のカテゴリで評価。リスクマップを作成し取締役会へ報告する。",
         "予防的", "年次", "Y", "経営企画部",
         "リスクアセスメント結果シート、\nリスクマップ、\n取締役会議事録",
         "R05 リスク管理規程", "有効", "有効", "なし",
         "整備・運用ともに有効", eval_info],
        True, "ok"))
    rows.append((
        ["ELC-005", "II. リスク評価", "不正リスク評価",
         "財務報告における不正リスクの見落とし",
         "決算期前に不正リスクファクター（動機・機会・正当化）を検討し、重要拠点・勘定について不正シナリオを評価する。内部監査室と経理部が合同で実施。",
         "予防的", "年次", "Y", "内部監査室\n/経理部",
         "不正リスク評価シート、\n評価結果報告書",
         "R03 内部監査規程", "有効", "有効", "なし",
         "整備・運用ともに有効", eval_info],
        True, "ok"))
    # ===== III. 統制活動 =====
    rows.append(("DIVIDER", "■ III. 統制活動（Control Activities）"))
    rows.append((
        ["ELC-006", "III. 統制活動", "規程・マニュアル整備",
         "業務標準の欠如による不正・誤謬の発生",
         "業務関連規程R11-R19を整備し、業務プロセスごとのマニュアルを作成する。規程は最低年1回見直し、改訂履歴を管理する。",
         "予防的", "年次", "N", "各主管部門",
         "規程一覧、\n改訂履歴表、\n業務マニュアル",
         "R11-R19", "有効", "有効", "なし",
         "整備・運用ともに有効", eval_info],
        False, "ok"))
    rows.append((
        ["ELC-007", "III. 統制活動", "職務の分離",
         "同一人物による取引の承認・実行・記録が可能な状態",
         "発注承認・検収・買掛計上・支払の4プロセスを異なる担当者が行うよう分掌する。SAPのロール設計にSoDルールを組み込み、月次でSoD違反レポートをレビューする。",
         "予防的", "月次", "Y", "各部門長\n/情シス部",
         "職務分掌マトリクス、\nSAP SoD違反レポート",
         "R18 職務権限規程", "有効", "一部不備あり\n(要改善)", "あり\n(軽微)",
         "PUR004のSoD違反を2026/3/31までに是正予定", eval_info],
        True, None))
    # ===== IV. 情報と伝達 =====
    rows.append(("DIVIDER", "■ IV. 情報と伝達（Information & Communication）"))
    rows.append((
        ["ELC-008", "IV. 情報と伝達", "内部通報制度",
         "不正・違反行為の顕在化と是正の遅延",
         "内部通報窓口（社内・外部弁護士事務所の2ルート）を設置し、全役職員に周知する。通報は総務部長と監査等委員が受領し、調査後の結果を通報者へ報告する。",
         "予防的", "随時", "Y", "総務部\n/外部弁護士",
         "内部通報規程、\n通報受付台帳、\n調査結果報告書",
         "R06 内部通報規程", "有効", "有効", "なし",
         "整備・運用ともに有効\n(Q2に1件受領済、調査完了)", eval_info],
        True, "ok"))
    rows.append((
        ["ELC-009", "IV. 情報と伝達", "決算情報の伝達",
         "決算情報の伝達遅延による重要事項の見落とし",
         "月次決算の速報を翌月第5営業日までに経営会議へ報告する。四半期決算は取締役会・監査等委員会で説明する。",
         "予防的", "月次/四半期", "N", "経理部",
         "月次決算報告資料、\n経営会議議事録",
         "R17 決算業務規程", "有効", "有効", "なし",
         "整備・運用ともに有効", eval_info],
        False, "ok"))
    # ===== V. モニタリング =====
    rows.append(("DIVIDER", "■ V. モニタリング（Monitoring）"))
    rows.append((
        ["ELC-010", "V. モニタリング", "内部監査の実施",
         "内部統制の独立した検証機能の欠如",
         "内部監査室が年次監査計画に基づき、重要拠点・業務の監査を実施する。発見事項は経営陣および監査等委員会に報告され、是正状況をフォローアップする。",
         "発見的", "年次\n(計画ベース)", "Y", "内部監査室",
         "年次内部監査計画、\n監査報告書、\n是正フォロー記録",
         "R03 内部監査規程", "有効", "有効", "なし",
         "整備・運用ともに有効", eval_info],
        True, "ok"))
    rows.append((
        ["ELC-011", "V. モニタリング", "監査等委員会のモニタリング",
         "経営監督機能の形骸化",
         "監査等委員会（社外委員3名）は月1回開催。内部監査・会計監査・法令遵守の状況をレビューし、取締役会へ意見表明する。",
         "発見的", "月次", "Y", "監査等委員会",
         "監査等委員会議事録、\nモニタリング報告書",
         "R02 監査等委員会規則", "有効", "有効", "なし",
         "整備・運用ともに有効", eval_info],
        True, "ok"))
    # ===== VI. ITへの対応 =====
    rows.append(("DIVIDER", "■ VI. ITへの対応（IT Governance）"))
    rows.append((
        ["ELC-012", "VI. ITへの対応", "IT戦略と情報セキュリティ方針",
         "ITガバナンス不備による情報漏洩・システム障害",
         "情報セキュリティ基本方針を定め、年1回取締役会で承認。情シス部が年次IT計画を作成し、経営会議で進捗をレビューする。",
         "予防的", "年次", "Y", "情報システム部",
         "情報セキュリティ基本方針、\nIT計画書、\n取締役会議事録",
         "R21 情報セキュリティ基本方針", "有効", "有効", "なし",
         "整備・運用ともに有効", eval_info],
        True, "ok"))

    # 書き出し
    row_num = 5
    for item in rows:
        if isinstance(item, tuple) and item[0] == "DIVIDER":
            write_coso_divider(ws, row_num, item[1])
            row_num += 1
        else:
            values, key, status = item
            write_elc_row(ws, row_num, values, key_control=key, status=status)
            row_num += 1

    add_legend_sheet(wb)

    wb.save(BASE / "ELC_RCM.xlsx")
    print(f"Created: ELC_RCM.xlsx ({sum(1 for r in rows if not (isinstance(r, tuple) and r[0] == 'DIVIDER'))} controls)")


if __name__ == "__main__":
    gen_elc()
