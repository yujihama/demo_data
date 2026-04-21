"""FCRP 全統制エビデンス一括整備

FCRP-001: 月次決算チェックリスト × 12ヶ月 + 経理部長承認WF
FCRP-002: パッケージ差異往復ログ
FCRP-003: 5種見積シート × 4Q + 3階層レビューWF (Q3/Q4貸倒根拠は HOLD-2026-003 により欠落維持)
FCRP-004: Q1/Q2/Q4個別仕訳一覧 + 非定型検討書 + 承認WF (Q3既存はファイル名統一)
FCRP-005: 3段階レビューWF + 取締役会承認

Key personnel:
- 経理部課長(財務会計) E0012 ACC002 高橋 美咲
- 経理部長 E0011 ACC001 佐藤 一郎
- CFO E0002 CFO001 渡辺 正博
- CEO E0001 CEO001 山本 健一
- 経営企画部 E0003 COO001 小林 剛 (RCM記載通り)
"""
import csv
import os
import sys
import io
import random
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

ROOT = Path(r"C:\Users\nyham\work\demo_data")
FCRP_DIR = ROOT / "4.evidence" / "FCRP"

IMPLEMENTER = '高橋 美咲 (ACC002)'
IMPLEMENTER_EMP = 'E0012'
ACC_MGR = '佐藤 一郎 (ACC001)'
ACC_MGR_EMP = 'E0011'
CFO = '渡辺 正博 (CFO001)'
CFO_EMP = 'E0002'
CEO = '山本 健一 (CEO001)'
CEO_EMP = 'E0001'
PLANNING = '小林 剛 (COO001)'  # 経営企画部 per RCM E0003
PLANNING_EMP = 'E0003'

# 監査等委員会 (ダミー独立役員)
AUDIT_COMMITTEE = [
    ('吉川 博', 'AUD001', 'E0081', '監査等委員長(独立社外取締役)'),
    ('中島 久美', 'AUD002', 'E0082', '監査等委員(独立社外取締役)'),
    ('山口 貴志', 'AUD003', 'E0083', '監査等委員(独立社外取締役)'),
]


# ==============================================================
# FCRP-001: 月次決算チェックリスト × 12 + 承認WF
# ==============================================================
CHECKLIST_ITEMS = [
    # I. 売上計上関連 (8)
    ('I', '売上計上関連', 'カットオフテスト - 月末5営業日の売上伝票突合', 'FI/SD'),
    ('I', '売上計上関連', '未出荷売上計上の有無確認 (VL01N-FI連携)', 'FI/SD'),
    ('I', '売上計上関連', '返品・値引の当月確定分反映', 'FI/SD'),
    ('I', '売上計上関連', '役務提供売上の進行基準計上確認', 'FI/PS'),
    ('I', '売上計上関連', '前受金の収益認識要件確認', 'FI'),
    ('I', '売上計上関連', '海外子会社向け売上の為替換算レート確認', 'FI'),
    ('I', '売上計上関連', '売上高異常変動分析 (前月比±15%)', 'CO'),
    ('I', '売上計上関連', '売上値引・割戻引当金計上', 'FI'),
    # II. 売上原価関連 (8)
    ('II', '売上原価関連', '標準原価計算バッチ (CK40N) 完了確認', 'CO'),
    ('II', '売上原価関連', '仕掛品→製品振替 (MB1B) 確認', 'MM/CO'),
    ('II', '売上原価関連', '外注加工費の正確な期間帰属', 'MM'),
    ('II', '売上原価関連', '間接費配賦 (KSU5) の正確性確認', 'CO'),
    ('II', '売上原価関連', '原価差異分析 (材料/労務/製造間接)', 'CO'),
    ('II', '売上原価関連', '棚卸資産評価損の検討', 'CO/FI'),
    ('II', '売上原価関連', '製品在庫月次棚卸差異調整', 'MM'),
    ('II', '売上原価関連', '売上総利益率前月比分析', 'CO'),
    # III. 売掛金関連 (6)
    ('III', '売掛金関連', 'AR年齢表 (FD10N) 確認', 'FI'),
    ('III', '売掛金関連', '月末債権残高とGL総勘定照合 (S_ALR_87012168)', 'FI'),
    ('III', '売掛金関連', '滞留債権120日超の一覧抽出', 'FI'),
    ('III', '売掛金関連', '入金消込 (F-28) 未完了残高確認', 'FI'),
    ('III', '売掛金関連', '貸倒実績率更新の要否判断', 'FI'),
    ('III', '売掛金関連', '大口顧客別債権残高分析', 'FI'),
    # IV. 棚卸資産関連 (5)
    ('IV', '棚卸資産関連', 'WMS⇔SAP在庫数量整合確認', 'MM/WMS'),
    ('IV', '棚卸資産関連', '循環棚卸差異の調整伝票確認', 'MM'),
    ('IV', '棚卸資産関連', '滞留在庫 (180日超) の評価減検討', 'MM/CO'),
    ('IV', '棚卸資産関連', '在庫回転率分析 (前月比)', 'CO'),
    ('IV', '棚卸資産関連', '期末在庫金額の部門別整合', 'MM'),
    # V. 費用計上関連 (5)
    ('V', '費用計上関連', '未払費用 (MIRO未入力分) 計上確認', 'MM/FI'),
    ('V', '費用計上関連', '経費精算 (Concur) との突合', 'FI'),
    ('V', '費用計上関連', '人件費の期間帰属正確性', 'FI/HR'),
    ('V', '費用計上関連', '研究開発費計上区分の妥当性確認', 'FI/CO'),
    ('V', '費用計上関連', '賃貸料・保険料の前払費用按分', 'FI'),
    # VI. 減価償却・固定資産 (4)
    ('VI', '固定資産関連', 'AFAB減価償却バッチ完了確認', 'FI-AA'),
    ('VI', '固定資産関連', '固定資産取得・除却の仕訳レビュー', 'FI-AA'),
    ('VI', '固定資産関連', 'リース資産償却額確認 (IFRS16対応)', 'FI-AA'),
    ('VI', '固定資産関連', '減損兆候モニタリング (キャッシュ生成単位別)', 'FI-AA'),
    # VII. 引当金 (3)
    ('VII', '引当金関連', '賞与引当金月次計上', 'FI'),
    ('VII', '引当金関連', '退職給付引当金月次見直し', 'FI'),
    ('VII', '引当金関連', '製品保証引当金計上', 'FI'),
    # VIII. 期末仕訳レビュー (3)
    ('VIII', '仕訳レビュー', '期末決算整理仕訳 (F-02) 全量レビュー', 'FI'),
    ('VIII', '仕訳レビュー', '仕訳№の連番性確認 (飛び番チェック)', 'FI'),
    ('VIII', '仕訳レビュー', '締切後仕訳の有無確認 (SM37)', 'FI'),
    # IX. その他 (3)
    ('IX', 'その他', '試算表 (S_ALR_87012284) 整合性確認', 'FI'),
    ('IX', 'その他', '連結パッケージ送信準備 (貸借差異ゼロ確認)', 'FI/連結'),
    ('IX', 'その他', '経営管理資料 (損益速報) 発信', 'CO'),
]


def create_monthly_checklist(ym, close_day, approve_day):
    """月次決算チェックリストxlsx作成"""
    year, month = ym.split('-')
    wb = Workbook()
    ws = wb.active
    ws.title = f"月次決算CL_{ym}"

    ws.cell(1, 1, f'月次決算チェックリスト / 対象月: {year}年{int(month)}月度')
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(2, 1, f'作成者: {IMPLEMENTER} ({IMPLEMENTER_EMP}) 経理部課長(財務会計)')
    ws.cell(3, 1, f'締め日: {close_day} / 承認日: {approve_day}')
    ws.cell(4, 1, '承認者: ' + ACC_MGR + f' ({ACC_MGR_EMP}) 経理部長')

    # Headers
    headers = ['区分', 'カテゴリ', '№', 'チェック項目', '対象モジュール', '実施日', '実施者', '結果', '備考']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(6, c, h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='305496')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    rng = random.Random(int(year) * 100 + int(month))

    # Fill 45 items
    close_dt = datetime.strptime(close_day, '%Y-%m-%d')
    for i, (section, cat, item, module) in enumerate(CHECKLIST_ITEMS, 1):
        # 実施日 within 5 business days before close_day
        days_offset = rng.randint(-5, 0)
        impl_date = close_dt + timedelta(days=days_offset)
        # Results: almost all PASS, a few notes
        result = '完了'
        if rng.random() < 0.08:
            note_options = ['差異分析レポート別紙あり', '前月比変動±20%超、要因分析完了', '軽微な仕訳修正完了', '関連部門確認済']
            note = rng.choice(note_options)
        else:
            note = ''

        ws.cell(7 + i - 1, 1, section)
        ws.cell(7 + i - 1, 2, cat)
        ws.cell(7 + i - 1, 3, i)
        ws.cell(7 + i - 1, 4, item)
        ws.cell(7 + i - 1, 5, module)
        ws.cell(7 + i - 1, 6, impl_date.strftime('%Y-%m-%d'))
        ws.cell(7 + i - 1, 7, IMPLEMENTER)
        ws.cell(7 + i - 1, 8, result)
        ws.cell(7 + i - 1, 9, note)

    # Summary at bottom
    sum_row = 7 + 45 + 2
    ws.cell(sum_row, 1, '【集計】')
    ws.cell(sum_row, 1).font = Font(bold=True)
    ws.cell(sum_row + 1, 1, '総項目数: 45')
    ws.cell(sum_row + 2, 1, '完了項目: 45')
    ws.cell(sum_row + 3, 1, '例外項目: 0')
    ws.cell(sum_row + 4, 1, f'経理部課長締め: {close_day} {IMPLEMENTER}')
    ws.cell(sum_row + 5, 1, f'経理部長承認: {approve_day} {ACC_MGR}')

    # Column widths
    widths = [6, 16, 5, 50, 14, 12, 22, 10, 30]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+c)].width = w

    return wb


def fix_fcrp_001():
    """FCRP-001: 12ヶ月分チェックリスト + 承認WF"""
    # Months: 2025-04 through 2026-03
    months = [('2025-04', '2025-05-02', '2025-05-08'),
              ('2025-05', '2025-06-02', '2025-06-08'),
              ('2025-06', '2025-07-02', '2025-07-08'),
              ('2025-07', '2025-08-04', '2025-08-08'),
              ('2025-08', '2025-09-02', '2025-09-08'),
              ('2025-09', '2025-10-02', '2025-10-08'),
              ('2025-10', '2025-11-04', '2025-11-08'),
              ('2025-11', '2025-12-02', '2025-12-08'),
              ('2025-12', '2026-01-05', '2026-01-09'),
              ('2026-01', '2026-02-02', '2026-02-08'),
              ('2026-02', '2026-03-03', '2026-03-09'),
              ('2026-03', '2026-04-06', '2026-04-13')]

    for ym, close, approve in months:
        wb = create_monthly_checklist(ym, close, approve)
        wb.save(FCRP_DIR / f'Checklist_MonthlyClose_{ym.replace("-", "")}.xlsx')

    print(f"[Created] 月次決算チェックリスト 12ファイル (202504〜202603)")

    # WF approval log
    lines = [
        "# Workflow System (S04) - 月次決算完了承認WF",
        "# Control: FCRP-001",
        "# Export: 2026-04-20 10:00:00 JST",
        "",
        "対象月,WF番号,締め完了日時,経理部課長(起票),締め完了者社員番号,経理部長承認日時,経理部長承認者,承認者社員番号,ステータス",
    ]
    wf_idx = 1
    for ym, close, approve in months:
        wfno = f"WF-MC-2025-{wf_idx:03d}"
        wf_idx += 1
        close_ts = f"{close} 17:00:00"
        approve_ts = f"{approve} 15:30:00"
        lines.append(f"{ym},{wfno},{close_ts},{IMPLEMENTER},{IMPLEMENTER_EMP},{approve_ts},{ACC_MGR},{ACC_MGR_EMP},承認完了")

    lines.append("")
    lines.append(f"# Records: {len(months)} months approved")

    with open(FCRP_DIR / "Workflow_MonthlyClose_Approval_FY2025.csv", 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))
    print("[Created] Workflow_MonthlyClose_Approval_FY2025.csv")


# ==============================================================
# FCRP-002: 差異往復ログ
# ==============================================================
def fix_fcrp_002():
    """連結パッケージバリデーションエラーの子会社との往復ログ"""
    # PASS_AFTER_CORRECTION の件を対象: Q1 TP-LOG, TPTR; Q2 TP-TB, TP-LOG; Q3 TP-TB

    corrections = [
        {'q': 'Q1', 'sub': 'TP-LOG', 'sub_name': 'TP物流サービス', 'errs': 1,
         'upload_ts': '2025-07-08 10:00:00', 'error_items': ['内部取引売掛金バランス差異 ¥125,000'],
         'correspondence': [
             ('2025-07-08 10:15', '経理部課長→TP物流 藤本', '送信', 'バリデーションエラー: 内部取引売掛金125,000円差異。確認願います。'),
             ('2025-07-08 11:30', 'TP物流 藤本→経理部課長', '返信', 'IC-2025-Q1-018の請求書消込漏れでした。修正後再送します。'),
             ('2025-07-08 14:40', 'TP物流→S05', '再送', '修正パッケージ再アップロード'),
             ('2025-07-08 15:00', '経理部課長', '確認', '再アップロード承認 - PASS')]},
        {'q': 'Q1', 'sub': 'TPTR', 'sub_name': 'TPトレーディング', 'errs': 1,
         'upload_ts': '2025-07-08 13:00:00', 'error_items': ['為替レート不一致 (TPT向け子会社間取引)'],
         'correspondence': [
             ('2025-07-08 13:20', '経理部課長→TPTR 後藤', '送信', 'タイバーツ換算レートがグループ規定と不一致。確認願います。'),
             ('2025-07-08 14:50', 'TPTR 後藤→経理部課長', '返信', 'ご指摘ありがとうございます。4.05円/THBに修正し再送します。'),
             ('2025-07-08 16:20', 'TPTR→S05', '再送', '為替レート修正版を再アップロード'),
             ('2025-07-08 16:40', '経理部課長', '確認', 'レート整合確認 - PASS')]},
        {'q': 'Q2', 'sub': 'TP-TB', 'sub_name': 'テクノプレシジョン東北', 'errs': 2,
         'upload_ts': '2025-10-08 09:00:00', 'error_items': ['勘定科目マスタ不整合 (新規勘定)', '内部利益消去額算定誤り'],
         'correspondence': [
             ('2025-10-08 09:30', '経理部課長→TP-TB 大橋', '送信', '2件エラー: 勘定科目2211未登録 + 内部利益消去の計算方式ズレ。'),
             ('2025-10-08 11:20', 'TP-TB 大橋→経理部課長', '返信', '1件目は科目マスタ連携不具合(情シス連絡済)。2件目は計算式を原価率ベースに修正。'),
             ('2025-10-08 13:40', '経理部課長→TP-TB', '再送依頼', '両件修正の上、再送願います。'),
             ('2025-10-08 15:30', 'TP-TB→S05', '再送', '2件修正版再アップロード'),
             ('2025-10-08 15:55', '経理部課長', '確認', '修正確認 - PASS')]},
        {'q': 'Q2', 'sub': 'TP-LOG', 'sub_name': 'TP物流サービス', 'errs': 1,
         'upload_ts': '2025-10-08 09:00:00', 'error_items': ['期首残高引継エラー (Q1末残高不一致)'],
         'correspondence': [
             ('2025-10-08 09:45', '経理部課長→TP物流 藤本', '送信', 'Q1末繰越残高 vs Q2期首残高の不一致あり。確認願います。'),
             ('2025-10-08 11:00', 'TP物流 藤本→経理部課長', '返信', 'S05側繰越処理が本社で未実行でした。本社情シス経由で修正中。'),
             ('2025-10-08 13:30', '経理部課長', '本社情シスに依頼', 'S05繰越処理バッチ再実行依頼済'),
             ('2025-10-08 14:30', '情シス', '完了通知', '繰越処理完了'),
             ('2025-10-08 15:20', 'TP物流→S05', '再送', 'パッケージ再アップロード'),
             ('2025-10-08 15:35', '経理部課長', '確認', '整合確認 - PASS')]},
        {'q': 'Q2', 'sub': 'TPTR', 'sub_name': 'TPトレーディング', 'errs': 0,  # actually Q2 TPTR is PASS not PAC, skip
         'upload_ts': '', 'error_items': [], 'correspondence': []},
        {'q': 'Q3', 'sub': 'TP-TB', 'sub_name': 'テクノプレシジョン東北', 'errs': 1,
         'upload_ts': '2026-01-08 15:00:00', 'error_items': ['固定資産除却仕訳の科目誤り'],
         'correspondence': [
             ('2026-01-08 15:20', '経理部課長→TP-TB 大橋', '送信', '12月除却仕訳の科目が誤り(特別損失→営業外費用)。確認願います。'),
             ('2026-01-08 16:40', 'TP-TB 大橋→経理部課長', '返信', '判断誤りです。特別損失に修正します。'),
             ('2026-01-08 17:40', 'TP-TB→S05', '再送', '科目修正の上再アップロード'),
             ('2026-01-08 17:55', '経理部課長', '確認', 'PASS')]},
    ]

    lines = [
        "# Consolidation System S05 - バリデーションエラー差異往復ログ",
        "# Control: FCRP-002 子会社連結パッケージの受領・差異解消",
        "# Export: 2026-04-20 10:30:00 JST",
        "",
        "四半期,子会社コード,子会社名,初回アップロード日時,エラー件数,エラー内容,往復日時,往復当事者,区分,メッセージ",
    ]
    for c in corrections:
        if c['errs'] == 0:
            continue
        error_summary = ' / '.join(c['error_items'])
        for ts, actor, kind, msg in c['correspondence']:
            lines.append(f"{c['q']},{c['sub']},{c['sub_name']},{c['upload_ts']},{c['errs']},{error_summary},{ts}:00,{actor},{kind},{msg}")

    lines.append("")
    lines.append(f"# Records: 5 corrections resolved")

    with open(FCRP_DIR / "ConsolidationSystem_ValidationError_Correspondence_FY2025.csv", 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))
    print("[Created] ConsolidationSystem_ValidationError_Correspondence_FY2025.csv")


# ==============================================================
# FCRP-003: 5種見積シート × 4Q + 3階層レビューWF
# ==============================================================
def fix_fcrp_003():
    """見積シート: BadDebt / Bonus / TaxEffect / Impairment / AssetRetirement
    各4四半期分。BadDebt Q3/Q4は根拠資料欠落(HOLD-2026-003維持)"""

    def create_estimate_xlsx(estimate_type, quarterly_data, hold_quarters=None):
        """estimate_type: japanese, quarterly_data: {Q1: [...], Q2: [...], ...}
        hold_quarters: set of quarters where 根拠資料 section is empty"""
        wb = Workbook()
        wb.remove(wb.active)
        hold_quarters = hold_quarters or set()

        for q in ['Q1', 'Q2', 'Q3', 'Q4']:
            ws = wb.create_sheet(q)
            qend_date = {'Q1': '2025-06-30', 'Q2': '2025-09-30', 'Q3': '2025-12-31', 'Q4': '2026-03-31'}[q]

            ws.cell(1, 1, f'会計上の見積りシート - {estimate_type} / {q} ({qend_date})')
            ws.cell(1, 1).font = Font(bold=True, size=14)
            ws.cell(2, 1, f'作成者: {IMPLEMENTER} ({IMPLEMENTER_EMP}) 経理部課長')
            ws.cell(3, 1, f'作成日: {qend_date}')

            # Data table
            rows = quarterly_data.get(q, [])
            if rows:
                # headers from first row
                headers = list(rows[0].keys())
                for c, h in enumerate(headers, 1):
                    cell = ws.cell(5, c, h)
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill('solid', fgColor='305496')
                for r_idx, row_data in enumerate(rows, 6):
                    for c, h in enumerate(headers, 1):
                        ws.cell(r_idx, c, row_data.get(h, ''))

                # Summary
                last_row = 6 + len(rows) + 2
                ws.cell(last_row, 1, '【前提・根拠資料】')
                ws.cell(last_row, 1).font = Font(bold=True)

                if q in hold_quarters:
                    ws.cell(last_row + 1, 1, '※ 根拠資料 (顧客別信用情報・回収可能性調査) 未添付 - 追加エビデンス要求中 (REQ-2026-003)')
                    ws.cell(last_row + 1, 1).font = Font(color='CC0000', bold=True)
                else:
                    # Detailed basis
                    ws.cell(last_row + 1, 1, '根拠資料:')
                    basis = quarterly_data.get(f'{q}_basis', ['別紙根拠資料一式添付'])
                    for i, b in enumerate(basis, 2):
                        ws.cell(last_row + i, 1, f'  - {b}')

            # Column widths
            for c in range(1, 10):
                ws.column_dimensions[chr(64+c)].width = 24

        return wb

    # Bad Debt (貸倒引当金) - Q3/Q4 基礎資料 empty
    bad_debt_data = {
        'Q1': [
            {'区分': '一般債権', '対象金額(円)': 3262467324, '評価方式': '実績率法(過去3年)', '引当率': '0.18%', '引当額(円)': 5872441},
            {'区分': '個別(C-10007 サンプル顧客G社)', '対象金額(円)': 6649275, '評価方式': '財務格付けB+財産状況D', '引当率': '40%', '引当額(円)': 2659710},
            {'区分': '個別(C-10017 サンプル顧客N社)', '対象金額(円)': 5717200, '評価方式': '取引履歴分析+業績悪化', '引当率': '20%', '引当額(円)': 1143440},
            {'区分': '個別(C-10023 サンプル顧客R社)', '対象金額(円)': 3209052, '評価方式': '財務格付けC+支払遅延180日超', '引当率': '60%', '引当額(円)': 1925431},
        ],
        'Q1_basis': ['C-10007信用情報調査書(東京商工リサーチ 2025/6/20取得)', 'C-10017財産状況ヒアリング議事録(2025/6/25)', 'C-10023支払遅延履歴+弁護士受任通知'],
        'Q2': [
            {'区分': '一般債権', '対象金額(円)': 3397668432, '評価方式': '実績率法', '引当率': '0.18%', '引当額(円)': 6115803},
            {'区分': '個別(C-10007 サンプル顧客G社)', '対象金額(円)': 3591058, '評価方式': '業績改善兆候→40%→50%', '引当率': '50%', '引当額(円)': 1795529},
            {'区分': '個別(C-10017 サンプル顧客N社)', '対象金額(円)': 5910490, '評価方式': '民事再生申請検討→引当増', '引当率': '40%', '引当額(円)': 2364196},
            {'区分': '個別(C-10023 サンプル顧客R社)', '対象金額(円)': 3057757, '評価方式': '一部回収見込み→20%に下方修正', '引当率': '20%', '引当額(円)': 611551},
        ],
        'Q2_basis': ['C-10007業績改善レポート(2025/9/22)', 'C-10017民事再生検討情報(2025/9/28)', 'C-10023和解契約書写し(2025/9/15)'],
        'Q3': [
            {'区分': '一般債権', '対象金額(円)': 3432137113, '評価方式': '実績率法', '引当率': '0.18%', '引当額(円)': 6177846},
            {'区分': '個別(C-10007 サンプル顧客G社)', '対象金額(円)': 5932383, '評価方式': '財務悪化再発→50%維持', '引当率': '50%', '引当額(円)': 2966191},
            {'区分': '個別(C-10017 サンプル顧客N社)', '対象金額(円)': 6096929, '評価方式': '民事再生申立中', '引当率': '40%', '引当額(円)': 2438771},
            {'区分': '個別(C-10023 サンプル顧客R社)', '対象金額(円)': 6899338, '評価方式': '再訴訟発生', '引当率': '40%', '引当額(円)': 2759735},
        ],
        'Q4': [
            {'区分': '一般債権', '対象金額(円)': 3074260787, '評価方式': '実績率法', '引当率': '0.18%', '引当額(円)': 5533669},
            {'区分': '個別(C-10007 サンプル顧客G社)', '対象金額(円)': 3192406, '評価方式': '業況不安定→60%引き上げ', '引当率': '60%', '引当額(円)': 1915443},
            {'区分': '個別(C-10017 サンプル顧客N社)', '対象金額(円)': 5526461, '評価方式': '再生計画認可待ち', '引当率': '40%', '引当額(円)': 2210584},
            {'区分': '個別(C-10023 サンプル顧客R社)', '対象金額(円)': 4433298, '評価方式': '判決待ち+資産状況悪化', '引当率': '60%', '引当額(円)': 2659979},
        ],
    }
    wb = create_estimate_xlsx('貸倒引当金', bad_debt_data, hold_quarters={'Q3', 'Q4'})
    wb.save(FCRP_DIR / "Estimate_BadDebtAllowance_Sheets_FY2025.xlsx")
    print("[Created] Estimate_BadDebtAllowance_Sheets_FY2025.xlsx (Q1/Q2完備, Q3/Q4根拠欠落=HOLD-2026-003)")

    # Bonus Accrual (賞与引当金)
    bonus_data = {
        'Q1': [{'区分': '月次計上額', '対象人員数': 512, '計上額(円)': 145000000, '根拠': '夏期賞与支給見込額×経過月/支給対象期間6ヶ月'}],
        'Q1_basis': ['人事部賞与支給計画書2025上期(2025/6/25)', '対象者一覧(HR002 岩本提供)', '人件費予算との整合確認書'],
        'Q2': [{'区分': '月次計上額', '対象人員数': 515, '計上額(円)': 147500000, '根拠': '夏期賞与実績額+冬期賞与見込額×経過月'}],
        'Q2_basis': ['夏期賞与支給実績通知(2025/7/10)', '冬期賞与支給計画書2025下期(2025/9/25)'],
        'Q3': [{'区分': '月次計上額', '対象人員数': 517, '計上額(円)': 148200000, '根拠': '冬期賞与支給見込額×経過月'}],
        'Q3_basis': ['冬期賞与支給実績通知(2025/12/10)', '役員賞与引当金別途計上根拠書'],
        'Q4': [{'区分': '期末調整', '対象人員数': 518, '計上額(円)': 72500000, '根拠': '支給額確定反映+過不足調整'}],
        'Q4_basis': ['夏冬賞与実績確定通知', '役員賞与取締役会議事録(2026/3/15)'],
    }
    wb = create_estimate_xlsx('賞与引当金', bonus_data)
    wb.save(FCRP_DIR / "Estimate_BonusAccrual_Sheets_FY2025.xlsx")
    print("[Created] Estimate_BonusAccrual_Sheets_FY2025.xlsx")

    # Tax Effect (税効果会計)
    tax_data = {
        'Q1': [
            {'区分': '将来減算一時差異', '対象': '貸倒引当金税務否認', '金額(円)': 5872441, '税率': '30.62%', '繰延税金資産(円)': 1798106},
            {'区分': '将来減算一時差異', '対象': '賞与引当金税務否認', '金額(円)': 145000000, '税率': '30.62%', '繰延税金資産(円)': 44399000},
            {'区分': '将来加算一時差異', '対象': '固定資産償却超過', '金額(円)': 22000000, '税率': '30.62%', '繰延税金負債(円)': 6736400},
        ],
        'Q1_basis': ['税務申告書2024(別表5-1)', '将来予想年次別事業計画書2025-2027', '監査法人確認書2025年6月'],
        'Q2': [
            {'区分': '将来減算一時差異', '対象': '貸倒引当金税務否認', '金額(円)': 6115803, '税率': '30.62%', '繰延税金資産(円)': 1872659},
            {'区分': '将来減算一時差異', '対象': '賞与引当金税務否認', '金額(円)': 147500000, '税率': '30.62%', '繰延税金資産(円)': 45164500},
            {'区分': '将来加算一時差異', '対象': '固定資産償却超過', '金額(円)': 22800000, '税率': '30.62%', '繰延税金負債(円)': 6981360},
        ],
        'Q2_basis': ['中間申告書2025', '最新事業計画書2025-2027', '監査法人Q2確認書'],
        'Q3': [
            {'区分': '将来減算一時差異', '対象': '貸倒引当金税務否認', '金額(円)': 6177846, '税率': '30.62%', '繰延税金資産(円)': 1891651},
            {'区分': '将来減算一時差異', '対象': '賞与引当金税務否認', '金額(円)': 148200000, '税率': '30.62%', '繰延税金資産(円)': 45378840},
            {'区分': '将来加算一時差異', '対象': '固定資産償却超過', '金額(円)': 23500000, '税率': '30.62%', '繰延税金負債(円)': 7195700},
        ],
        'Q3_basis': ['更新事業計画書', '税制改正情報 (2026年度)', '監査法人Q3確認書'],
        'Q4': [
            {'区分': '将来減算一時差異', '対象': '貸倒引当金税務否認', '金額(円)': 5533669, '税率': '30.62%', '繰延税金資産(円)': 1694409},
            {'区分': '将来減算一時差異', '対象': '賞与引当金税務否認', '金額(円)': 72500000, '税率': '30.62%', '繰延税金資産(円)': 22199500},
            {'区分': '将来加算一時差異', '対象': '固定資産償却超過', '金額(円)': 24100000, '税率': '30.62%', '繰延税金負債(円)': 7379420},
        ],
        'Q4_basis': ['年度末税務調整書', '2026年度事業計画書', '監査法人期末確認書', '回収可能性分類書(区分1)'],
    }
    wb = create_estimate_xlsx('税効果会計', tax_data)
    wb.save(FCRP_DIR / "Estimate_TaxEffect_Sheets_FY2025.xlsx")
    print("[Created] Estimate_TaxEffect_Sheets_FY2025.xlsx")

    # Impairment (減損会計)
    impairment_data = {
        'Q1': [
            {'CGU': '量産工場A棟', '帳簿価額(円)': 2150000000, '減損兆候': 'なし', '割引前CF見積': 2850000000, '判定': '減損兆候なし'},
            {'CGU': '試作ライン B棟', '帳簿価額(円)': 380000000, '減損兆候': '受注減少の兆候あり', '割引前CF見積': 420000000, '判定': '減損兆候あり→兆候検討中'},
            {'CGU': '物流センター (東北)', '帳簿価額(円)': 890000000, '減損兆候': 'なし', '割引前CF見積': 1120000000, '判定': '減損兆候なし'},
        ],
        'Q1_basis': ['CGU別5ヶ年事業計画', '将来キャッシュフロー算定書', '割引率算定書 (WACC 6.5%)'],
        'Q2': [
            {'CGU': '量産工場A棟', '帳簿価額(円)': 2132000000, '減損兆候': 'なし', '割引前CF見積': 2810000000, '判定': '減損兆候なし'},
            {'CGU': '試作ライン B棟', '帳簿価額(円)': 373000000, '減損兆候': '継続', '割引前CF見積': 395000000, '判定': '減損の測定実施 → 減損不要'},
            {'CGU': '物流センター (東北)', '帳簿価額(円)': 870000000, '減損兆候': 'なし', '割引前CF見積': 1100000000, '判定': '減損兆候なし'},
        ],
        'Q2_basis': ['試作ライン B棟 減損測定書', 'B棟CGUの回収可能価額算定(継続使用)', '経営会議議事録2025/9/28 (B棟継続使用判断)'],
        'Q3': [
            {'CGU': '量産工場A棟', '帳簿価額(円)': 2114000000, '減損兆候': 'なし', '割引前CF見積': 2780000000, '判定': '減損兆候なし'},
            {'CGU': '試作ライン B棟', '帳簿価額(円)': 366000000, '減損兆候': '改善傾向', '割引前CF見積': 440000000, '判定': '減損兆候解消'},
            {'CGU': '物流センター (東北)', '帳簿価額(円)': 850000000, '減損兆候': 'なし', '割引前CF見積': 1090000000, '判定': '減損兆候なし'},
        ],
        'Q3_basis': ['B棟CGU受注回復レポート', '経営計画更新書'],
        'Q4': [
            {'CGU': '量産工場A棟', '帳簿価額(円)': 2096000000, '減損兆候': 'なし', '割引前CF見積': 2750000000, '判定': '減損兆候なし'},
            {'CGU': '試作ライン B棟', '帳簿価額(円)': 359000000, '減損兆候': 'なし', '割引前CF見積': 465000000, '判定': '減損兆候なし'},
            {'CGU': '物流センター (東北)', '帳簿価額(円)': 830000000, '減損兆候': 'なし', '割引前CF見積': 1080000000, '判定': '減損兆候なし'},
        ],
        'Q4_basis': ['期末減損兆候一覧表', '経営計画2026年度版', '割引率再算定書', '監査法人検討書'],
    }
    wb = create_estimate_xlsx('減損会計', impairment_data)
    wb.save(FCRP_DIR / "Estimate_ImpairmentTest_Sheets_FY2025.xlsx")
    print("[Created] Estimate_ImpairmentTest_Sheets_FY2025.xlsx")

    # Asset Retirement Obligation (資産除去債務)
    aro_data = {
        'Q1': [
            {'対象資産': '量産工場A棟土壌汚染対策', '発生原因': '法令・契約義務', '見積額(円)': 380000000, '時の経過による加算(円)': 3850000, '期末残高(円)': 180850000},
            {'対象資産': '物流センター アスベスト除去', '発生原因': '法令義務', '見積額(円)': 45000000, '時の経過による加算(円)': 470000, '期末残高(円)': 23470000},
        ],
        'Q1_basis': ['対策費見積書(専門業者 2024取得)', '割引率算定書 (国債利回り基準1.2%)', '耐用年数残存期間計算書'],
        'Q2': [
            {'対象資産': '量産工場A棟土壌汚染対策', '発生原因': '法令・契約義務', '見積額(円)': 380000000, '時の経過による加算(円)': 3850000, '期末残高(円)': 184700000},
            {'対象資産': '物流センター アスベスト除去', '発生原因': '法令義務', '見積額(円)': 45000000, '時の経過による加算(円)': 470000, '期末残高(円)': 23940000},
        ],
        'Q2_basis': ['Q1と同様 (見積更新なし)'],
        'Q3': [
            {'対象資産': '量産工場A棟土壌汚染対策', '発生原因': '法令・契約義務', '見積額(円)': 380000000, '時の経過による加算(円)': 3850000, '期末残高(円)': 188550000},
            {'対象資産': '物流センター アスベスト除去', '発生原因': '法令義務', '見積額(円)': 45000000, '時の経過による加算(円)': 470000, '期末残高(円)': 24410000},
        ],
        'Q3_basis': ['見積更新なし'],
        'Q4': [
            {'対象資産': '量産工場A棟土壌汚染対策', '発生原因': '法令・契約義務', '見積額(円)': 395000000, '時の経過による加算(円)': 3950000, '期末残高(円)': 192500000},
            {'対象資産': '物流センター アスベスト除去', '発生原因': '法令義務', '見積額(円)': 46000000, '時の経過による加算(円)': 480000, '期末残高(円)': 24890000},
        ],
        'Q4_basis': ['年度末見積再評価書 (専門業者 2026/3取得)', '為替影響調整書', '監査法人期末確認書'],
    }
    wb = create_estimate_xlsx('資産除去債務', aro_data)
    wb.save(FCRP_DIR / "Estimate_AssetRetirement_Sheets_FY2025.xlsx")
    print("[Created] Estimate_AssetRetirement_Sheets_FY2025.xlsx")

    # 3-stage review workflow: 経理部課長→経理部長→CFO→監査等委員会
    lines = [
        "# Workflow System (S04) - 会計上見積 3階層レビュー承認WF",
        "# Control: FCRP-003",
        "# Export: 2026-04-20 11:00:00 JST",
        "",
        "四半期,見積種別,WF番号,経理部課長作成日,作成者,作成者社員番号,経理部長レビュー日,経理部長,経理部長社員番号,CFOレビュー日,CFO,CFO社員番号,監査等委員会承認日,監査等委員長,監査等委員長社員番号,ステータス",
    ]
    estimates = ['貸倒引当金', '賞与引当金', '税効果会計', '減損会計', '資産除去債務']
    q_dates = {
        'Q1': ('2025-07-05', '2025-07-10', '2025-07-15', '2025-07-25'),
        'Q2': ('2025-10-05', '2025-10-10', '2025-10-15', '2025-10-25'),
        'Q3': ('2026-01-06', '2026-01-12', '2026-01-16', '2026-01-26'),
        'Q4': ('2026-04-06', '2026-04-12', '2026-04-16', '2026-04-28'),
    }
    wf_idx = 1
    chair = AUDIT_COMMITTEE[0]
    for q, (d1, d2, d3, d4) in q_dates.items():
        for est in estimates:
            # Q3/Q4 貸倒: 保留ステータス
            if est == '貸倒引当金' and q in ['Q3', 'Q4']:
                status = '審議保留(REQ-2026-003)'
                d3_actual = '-'
                d4_actual = '-'
            else:
                status = '承認完了'
                d3_actual = d3
                d4_actual = d4
            wfno = f"WF-EST-2025-{wf_idx:03d}"
            wf_idx += 1
            lines.append(f"{q},{est},{wfno},{d1},{IMPLEMENTER},{IMPLEMENTER_EMP},{d2},{ACC_MGR},{ACC_MGR_EMP},{d3_actual},{CFO if d3_actual!='-' else '-'},{CFO_EMP if d3_actual!='-' else '-'},{d4_actual},{chair[0]+' ('+chair[1]+')' if d4_actual!='-' else '-'},{chair[2] if d4_actual!='-' else '-'},{status}")

    lines.append("")
    lines.append(f"# Records: 20 (5 estimates × 4 quarters, Q3/Q4 貸倒は保留)")

    with open(FCRP_DIR / "Workflow_AccountingEstimate_3StageReview_FY2025.csv", 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))
    print("[Created] Workflow_AccountingEstimate_3StageReview_FY2025.csv")


# ==============================================================
# FCRP-004: Q1/Q2/Q4 個別xlsx + 非定型検討書 + 承認WF
# ==============================================================
def fix_fcrp_004():
    """連結仕訳一覧 4Q分 + 非定型検討 + 承認WF
    既存Q3をリネーム、Q1/Q2/Q4新規"""

    entries_by_q = {
        'Q1': [
            {'仕訳№': 'CNS-Q1-001', '区分': '投資と資本の相殺', '借方科目': '資本金', '貸方科目': '関係会社株式', '金額(円)': 300000000, '摘要': 'テクノプレシジョン東北の投資相殺', '起票': '自動'},
            {'仕訳№': 'CNS-Q1-002', '区分': '内部取引消去', '借方科目': '売上高', '貸方科目': '売上原価', '金額(円)': 1820000000, '摘要': '親会社→東北向け内部販売消去', '起票': '自動'},
            {'仕訳№': 'CNS-Q1-003', '区分': '内部取引消去', '借方科目': '売上高', '貸方科目': '売上原価', '金額(円)': 650000000, '摘要': '親会社→タイ子会社向け内部販売消去', '起票': '自動'},
            {'仕訳№': 'CNS-Q1-004', '区分': '内部取引消去', '借方科目': '売掛金', '貸方科目': '買掛金', '金額(円)': 324500000, '摘要': '内部債権債務相殺', '起票': '自動'},
            {'仕訳№': 'CNS-Q1-005', '区分': '少数株主損益', '借方科目': '少数株主損益', '貸方科目': '利益剰余金', '金額(円)': 8520000, '摘要': 'TPトレーディング 少数株主持分', '起票': '自動'},
            {'仕訳№': 'CNS-Q1-006', '区分': '内部利益消去(在庫)', '借方科目': '売上原価', '貸方科目': '棚卸資産', '金額(円)': 42800000, '摘要': '親→東北間在庫の未実現利益消去', '起票': '手動'},
        ],
        'Q2': [
            {'仕訳№': 'CNS-Q2-001', '区分': '投資と資本の相殺', '借方科目': '資本金', '貸方科目': '関係会社株式', '金額(円)': 300000000, '摘要': 'テクノプレシジョン東北の投資相殺', '起票': '自動'},
            {'仕訳№': 'CNS-Q2-002', '区分': '内部取引消去', '借方科目': '売上高', '貸方科目': '売上原価', '金額(円)': 1820000000, '摘要': '親会社→東北向け内部販売消去', '起票': '自動'},
            {'仕訳№': 'CNS-Q2-003', '区分': '内部取引消去', '借方科目': '売上高', '貸方科目': '売上原価', '金額(円)': 650000000, '摘要': '親会社→タイ子会社向け内部販売消去', '起票': '自動'},
            {'仕訳№': 'CNS-Q2-004', '区分': '内部取引消去', '借方科目': '売掛金', '貸方科目': '買掛金', '金額(円)': 324500000, '摘要': '内部債権債務相殺', '起票': '自動'},
            {'仕訳№': 'CNS-Q2-005', '区分': '少数株主損益', '借方科目': '少数株主損益', '貸方科目': '利益剰余金', '金額(円)': 8520000, '摘要': 'TPトレーディング 少数株主持分', '起票': '自動'},
            {'仕訳№': 'CNS-Q2-006', '区分': '内部利益消去(在庫)', '借方科目': '売上原価', '貸方科目': '棚卸資産', '金額(円)': 42800000, '摘要': '親→東北間在庫の未実現利益消去', '起票': '手動'},
        ],
        'Q3': [
            {'仕訳№': 'CNS-Q3-001', '区分': '投資と資本の相殺', '借方科目': '資本金', '貸方科目': '関係会社株式', '金額(円)': 300000000, '摘要': 'テクノプレシジョン東北の投資相殺', '起票': '自動'},
            {'仕訳№': 'CNS-Q3-002', '区分': '内部取引消去', '借方科目': '売上高', '貸方科目': '売上原価', '金額(円)': 1820000000, '摘要': '親会社→東北向け内部販売消去', '起票': '自動'},
            {'仕訳№': 'CNS-Q3-003', '区分': '内部取引消去', '借方科目': '売上高', '貸方科目': '売上原価', '金額(円)': 650000000, '摘要': '親会社→タイ子会社向け内部販売消去', '起票': '自動'},
            {'仕訳№': 'CNS-Q3-004', '区分': '内部取引消去', '借方科目': '売掛金', '貸方科目': '買掛金', '金額(円)': 324500000, '摘要': '内部債権債務相殺', '起票': '自動'},
            {'仕訳№': 'CNS-Q3-005', '区分': '少数株主損益', '借方科目': '少数株主損益', '貸方科目': '利益剰余金', '金額(円)': 8520000, '摘要': 'TPトレーディング 少数株主持分', '起票': '自動'},
            {'仕訳№': 'CNS-Q3-006', '区分': '内部利益消去(在庫)', '借方科目': '売上原価', '貸方科目': '棚卸資産', '金額(円)': 42800000, '摘要': '親→東北間在庫の未実現利益消去', '起票': '手動'},
        ],
        'Q4': [
            {'仕訳№': 'CNS-Q4-001', '区分': '投資と資本の相殺', '借方科目': '資本金', '貸方科目': '関係会社株式', '金額(円)': 300000000, '摘要': 'テクノプレシジョン東北の投資相殺', '起票': '自動'},
            {'仕訳№': 'CNS-Q4-002', '区分': '内部取引消去', '借方科目': '売上高', '貸方科目': '売上原価', '金額(円)': 1820000000, '摘要': '親会社→東北向け内部販売消去', '起票': '自動'},
            {'仕訳№': 'CNS-Q4-003', '区分': '内部取引消去', '借方科目': '売上高', '貸方科目': '売上原価', '金額(円)': 650000000, '摘要': '親会社→タイ子会社向け内部販売消去', '起票': '自動'},
            {'仕訳№': 'CNS-Q4-004', '区分': '内部取引消去', '借方科目': '売掛金', '貸方科目': '買掛金', '金額(円)': 324500000, '摘要': '内部債権債務相殺', '起票': '自動'},
            {'仕訳№': 'CNS-Q4-005', '区分': '少数株主損益', '借方科目': '少数株主損益', '貸方科目': '利益剰余金', '金額(円)': 8520000, '摘要': 'TPトレーディング 少数株主持分', '起票': '自動'},
            {'仕訳№': 'CNS-Q4-006', '区分': '内部利益消去(在庫)', '借方科目': '売上原価', '貸方科目': '棚卸資産', '金額(円)': 42800000, '摘要': '親→東北間在庫の未実現利益消去', '起票': '手動'},
            {'仕訳№': 'CNS-Q4-007', '区分': '期末為替換算調整', '借方科目': 'その他包括利益', '貸方科目': '為替換算調整勘定', '金額(円)': 24500000, '摘要': 'タイ子会社年度末換算差額', '起票': '自動'},
        ],
    }

    q_dates = {'Q1': '2025-06-30', 'Q2': '2025-09-30', 'Q3': '2025-12-31', 'Q4': '2026-03-31'}

    for q, entries in entries_by_q.items():
        wb = Workbook()
        ws = wb.active
        ws.title = "連結仕訳一覧"

        ws.cell(1, 1, f'連結仕訳一覧 / FY2025 {q} ({q_dates[q]})')
        ws.cell(1, 1).font = Font(bold=True, size=14)
        ws.cell(2, 1, f'出力元: 連結決算システム(S05) / 作成: {IMPLEMENTER} / レビュー承認: {ACC_MGR}')

        headers = ['仕訳№', '区分', '借方科目', '貸方科目', '金額(円)', '摘要', '起票区分']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(4, c, h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='305496')
            cell.alignment = Alignment(horizontal='center')

        for i, e in enumerate(entries, 5):
            for c, h in enumerate(['仕訳№', '区分', '借方科目', '貸方科目', '金額(円)', '摘要', '起票'], 1):
                ws.cell(i, c, e.get(h, ''))

        sum_row = 5 + len(entries) + 1
        auto_count = sum(1 for e in entries if e['起票'] == '自動')
        manual_count = sum(1 for e in entries if e['起票'] == '手動')
        ws.cell(sum_row, 1, f'合計: {len(entries)}件 (自動{auto_count}件 / 手動{manual_count}件)')
        ws.cell(sum_row, 1).font = Font(bold=True)

        widths = [14, 22, 16, 16, 16, 40, 12]
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64+c)].width = w

        wb.save(FCRP_DIR / f"ConsolidationEntries_List_2025{q}.xlsx")
        print(f"[Created] ConsolidationEntries_List_2025{q}.xlsx ({len(entries)}件)")

    # Remove old Q3 xlsx
    old_q3 = FCRP_DIR / "FCRP-004_連結仕訳一覧_2025Q3.xlsx"
    if old_q3.exists():
        old_q3.unlink()
        print(f"[Removed] FCRP-004_連結仕訳一覧_2025Q3.xlsx (命名統一のため削除)")

    # 非定型検討書 - 手動起票(CNS-Qx-006 内部利益消去, Q4のCNS-Q4-007 為替換算)
    wb = Workbook()
    ws = wb.active
    ws.title = "非定型連結仕訳検討書"

    ws.cell(1, 1, '非定型連結仕訳の検討・文書化 / FY2025')
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(2, 1, f'作成者: {IMPLEMENTER} / レビュー: {ACC_MGR}')

    headers = ['仕訳№', '四半期', '区分', '金額(円)', '起票判断根拠', '過年度整合性', 'レビュー結論', '承認者']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='305496')

    nonstd = [
        ('CNS-Q1-006', 'Q1', '内部利益消去(在庫)', 42800000, '親→東北向け在庫期末残高 × 連結利益率10.5%', '前四半期同額', '適正 - 算定ロジック変更なし', ACC_MGR),
        ('CNS-Q2-006', 'Q2', '内部利益消去(在庫)', 42800000, '親→東北向け在庫期末残高 × 連結利益率10.5%', 'Q1同額', '適正', ACC_MGR),
        ('CNS-Q3-006', 'Q3', '内部利益消去(在庫)', 42800000, '親→東北向け在庫期末残高 × 連結利益率10.5%', 'Q2同額', '適正', ACC_MGR),
        ('CNS-Q4-006', 'Q4', '内部利益消去(在庫)', 42800000, '親→東北向け在庫期末残高 × 連結利益率10.5%', 'Q3同額', '適正', ACC_MGR),
        ('CNS-Q4-007', 'Q4', '期末為替換算調整', 24500000, 'TPT(タイ)の期末為替レート改定影響(3.80→4.05円/THB)', 'Q4のみ発生', '適正 - 換算レート変更反映', ACC_MGR),
    ]
    for i, row in enumerate(nonstd, 5):
        for c, v in enumerate(row, 1):
            ws.cell(i, c, v)

    widths = [14, 8, 20, 14, 40, 14, 30, 24]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+c)].width = w

    wb.save(FCRP_DIR / "ConsolidationEntries_NonStandard_DeliberationMemo_FY2025.xlsx")
    print("[Created] ConsolidationEntries_NonStandard_DeliberationMemo_FY2025.xlsx")

    # Approval WF
    lines = [
        "# Workflow System (S04) - 連結仕訳承認WF",
        "# Control: FCRP-004",
        "# Export: 2026-04-20 11:30:00 JST",
        "",
        "四半期,WF番号,経理部課長レビュー日,レビュー者,レビュー者社員番号,経理部長承認日,承認者,承認者社員番号,仕訳件数,うち手動,ステータス",
    ]
    q_review = {
        'Q1': ('2025-07-08', '2025-07-12', 6, 1),
        'Q2': ('2025-10-08', '2025-10-12', 6, 1),
        'Q3': ('2026-01-10', '2026-01-16', 6, 1),
        'Q4': ('2026-04-08', '2026-04-14', 7, 1),
    }
    wf_idx = 1
    for q, (rev_d, appr_d, cnt, manual) in q_review.items():
        wfno = f"WF-CNS-2025-{wf_idx:03d}"
        wf_idx += 1
        lines.append(f"{q},{wfno},{rev_d},{IMPLEMENTER},{IMPLEMENTER_EMP},{appr_d},{ACC_MGR},{ACC_MGR_EMP},{cnt},{manual},承認完了")

    lines.append("")
    lines.append(f"# Records: 4 quarters approved")

    with open(FCRP_DIR / "Workflow_ConsolidationEntries_Approval_FY2025.csv", 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))
    print("[Created] Workflow_ConsolidationEntries_Approval_FY2025.csv")


# ==============================================================
# FCRP-005: 3段階レビューWF + 取締役会承認
# ==============================================================
def fix_fcrp_005():
    """開示書類3段階レビュー: 経営企画部作成 → 経理部長レビュー → CFOレビュー → 監査等委員会 → 取締役会承認"""

    lines = [
        "# Workflow System (S04) - 開示書類3段階レビューWF",
        "# Control: FCRP-005",
        "# Export: 2026-05-20 10:00:00 JST",
        "",
        "四半期,書類ID,WF番号,経営企画部起票日,起票者,経理部長レビュー日,経理部長,CFOレビュー日,CFO,監査等委員会審議日,監査等委員長,ステータス",
    ]
    chair = AUDIT_COMMITTEE[0]

    q_timeline = {
        'Q1': ('EDINET-XBRL-202508', '2025-08-01', '2025-08-05', '2025-08-08', '2025-08-12'),
        'Q2': ('EDINET-XBRL-202511', '2025-11-01', '2025-11-05', '2025-11-08', '2025-11-12'),
        'Q3': ('EDINET-XBRL-202602', '2026-02-01', '2026-02-05', '2026-02-08', '2026-02-12'),
        'Q4': ('EDINET-XBRL-202605', '2026-04-25', '2026-04-28', '2026-05-03', '2026-05-07'),
    }
    wf_idx = 1
    for q, (doc, d1, d2, d3, d4) in q_timeline.items():
        wfno = f"WF-DISC-2025-{wf_idx:03d}"
        wf_idx += 1
        lines.append(f"{q},{doc},{wfno},{d1},{PLANNING} ({PLANNING_EMP}),{d2},{ACC_MGR} ({ACC_MGR_EMP}),{d3},{CFO} ({CFO_EMP}),{d4},{chair[0]} ({chair[1]}/{chair[2]}),承認完了")

    lines.append("")
    lines.append(f"# Records: 4 quarters approved")

    with open(FCRP_DIR / "Workflow_Disclosure_3StageReview_FY2025.csv", 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))
    print("[Created] Workflow_Disclosure_3StageReview_FY2025.csv")

    # BOD (取締役会) Meeting Minutes for 開示書類
    lines = [
        "# 取締役会議事録 - FY2025 開示書類承認",
        "# Control: FCRP-005",
        "# Export: 2026-05-20 10:30:00 JST",
        "",
        "四半期,書類ID,開催日,議案番号,議案,審議結果,出席取締役数,可決数,議長,議事録署名",
    ]
    bod_meetings = {
        'Q1': ('EDINET-XBRL-202508', '2025-08-13', '第3号議案', '2026年3月期 第1四半期報告書の承認'),
        'Q2': ('EDINET-XBRL-202511', '2025-11-13', '第3号議案', '2026年3月期 第2四半期報告書の承認'),
        'Q3': ('EDINET-XBRL-202602', '2026-02-13', '第3号議案', '2026年3月期 第3四半期報告書の承認'),
        'Q4': ('EDINET-XBRL-202605', '2026-05-08', '第1号議案', '2026年3月期 有価証券報告書の承認'),
    }
    for q, (doc, date, agenda_no, agenda) in bod_meetings.items():
        # Assume 7 directors (3 EDs + 3 independent + CEO), all vote yes
        lines.append(f"{q},{doc},{date},{agenda_no},{agenda},全会一致承認,7,7,{CEO} ({CEO_EMP}),{CEO} 議長 + 取締役全員")

    lines.append("")
    lines.append(f"# Records: 4 BOD meetings")

    with open(FCRP_DIR / "BOD_Meeting_Disclosure_Approval_FY2025.csv", 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))
    print("[Created] BOD_Meeting_Disclosure_Approval_FY2025.csv")


# ==============================================================
# Update Evidence_Mapping_FCRP.csv
# ==============================================================
def update_mapping():
    # Build mapping from scratch based on actual files + control assignment
    import os
    actual_files = sorted(os.listdir(FCRP_DIR))

    mapping = []
    for fn in actual_files:
        # Assignment rules
        if 'MonthlyClose' in fn or 'PeriodClose' in fn or 'Checklist_MonthlyClose' in fn:
            mapping.append(('FCRP-001', '1', fn))
        elif 'PackageUpload' in fn or 'ValidationError_Correspondence' in fn:
            mapping.append(('FCRP-002', '1', fn))
        elif 'Estimate_' in fn or 'BadDebt' in fn or ('Workflow_AccountingEstimate' in fn):
            mapping.append(('FCRP-003', '1', fn))
        elif 'ConsolidationEntries' in fn or 'Workflow_ConsolidationEntries' in fn or 'ConsolidationSystem_Entries' in fn:
            mapping.append(('FCRP-004', '1', fn))
        elif 'Disclosure' in fn or 'XBRL' in fn or 'BOD_Meeting' in fn:
            mapping.append(('FCRP-005', '1', fn))

    # Sort by control ID order
    order = ['FCRP-001', 'FCRP-002', 'FCRP-003', 'FCRP-004', 'FCRP-005']
    mapping.sort(key=lambda x: (order.index(x[0]) if x[0] in order else 99, x[2]))

    path = ROOT / "2.RCM" / "Evidence_Mapping_FCRP.csv"
    with open(path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL)
        writer.writerow(['key', 'sample_no', 'filename'])
        for row in mapping:
            writer.writerow(row)

    print(f"[Fixed] Evidence_Mapping_FCRP.csv: {len(mapping)} entries")


if __name__ == '__main__':
    fix_fcrp_001()
    fix_fcrp_002()
    fix_fcrp_003()
    fix_fcrp_004()
    fix_fcrp_005()
    update_mapping()
    print("\n=== FCRP ALL FIXES COMPLETED ===")
