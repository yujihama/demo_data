"""
PLC-S-002 エビデンス簡素化（新方針適用）

新方針：エビデンス = 監査人が受領するRAWデータのみ。
統制実施者（経理部）の確認・レビュー・承認の記録は削除。
監査人はRAWデータから自ら検証・評価を行う。

実施内容：
1. 2ファイル削除
   - 出荷売上マッチング照合レポート（経理部の突合記録）
   - 25件対応_日次未マッチレビュー記録（経理部のレビュー記録）
2. 4ファイル簡素化
   - SAP未マッチ明細リスト (経理部コメント削除)
   - 例外サンプル9 .txt (分析コメント削除、純粋なSAP VA02出力のみ)
   - 例外サンプル14 .txt (分析コメント削除、純粋なSAPバッチログのみ)
   - 25件サンプル対応エビデンス.xlsx → 監査対象25件サンプルリスト.xlsxに変更
     （提出者・承認者・例外対応記録コラム削除、純粋な取引リストに）
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
from pathlib import Path

BASE = Path(r"C:\Users\nyham\work\demo_data\4.evidence\PLC-S")

HF = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
BFONT = Font(name="Yu Gothic", size=10)
BBOLD = Font(name="Yu Gothic", size=10, bold=True)
C_ = Alignment(horizontal="center", vertical="center", wrap_text=True)
L_ = Alignment(horizontal="left", vertical="center", wrap_text=True)
R_ = Alignment(horizontal="right", vertical="center")
T_ = Side("thin", color="888888")
BRD = Border(left=T_, right=T_, top=T_, bottom=T_)
FILL_META = PatternFill("solid", fgColor="D9E1F2")


# ==============================================================
# 1. 削除対象
# ==============================================================
def delete_files():
    targets = [
        BASE / "PLC-S-002_出荷売上マッチング照合レポート_202511.xlsx",
        BASE / "PLC-S-002_25件対応_日次未マッチレビュー記録_FY2025抜粋.xlsx",
    ]
    for t in targets:
        if t.exists():
            t.unlink()
            print(f"Deleted: {t.name}")


# ==============================================================
# 2. SAP未マッチ明細リスト（簡素化）
# ==============================================================
def simplify_unmatch_csv():
    path = BASE / "PLC-S-002_SAP未マッチ明細リスト_202511.csv"
    lines = [
        "# SAP Query ZSD_UNMATCH / 未マッチ出荷-売上明細",
        "# Output:    2025-12-03 08:15:23 JST",
        "# Filter:    Shipment date = 2025/11/01 - 2025/11/30 AND no matching sales journal posting within tolerance",
        "# Tolerance: +/- JPY 10,000 OR +/- 5.0%",
        "#",
        "No,出荷番号,出荷日,受注番号,顧客コード,顧客名,出荷金額,対応売上仕訳,売上計上額,システム検知原因,検知タイムスタンプ",
        "1,SH-202511-0234,2025-11-17,ORD-2025-2468,C-10003,サンプル顧客C社,12850000,JV-202511-0234,12800000,AMOUNT_DIFF_OVER_TOLERANCE,2025-11-18 01:48:01",
        "",
        "# Records: 1",
    ]
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"Simplified: {path.name}")


# ==============================================================
# 3. 例外サンプル9 (SAP VA02変更履歴のみ、分析コメント削除)
# ==============================================================
def simplify_sample9_txt():
    # 該当サンプルの情報を既存ファイルから取得
    wb = openpyxl.load_workbook(BASE / "PLC-S-002_25件サンプル対応エビデンス.xlsx")
    ws = wb.active
    s9 = None
    for r in range(17, 42):
        if ws.cell(row=r, column=1).value == 9:
            s9 = {
                "ship_no": ws.cell(row=r, column=2).value,
                "ord_no": ws.cell(row=r, column=4).value,
                "ship_date": ws.cell(row=r, column=3).value,
                "sale_amount": ws.cell(row=r, column=12).value,
                "qty": int(str(ws.cell(row=r, column=8).value).split()[0].replace(",", "")),
            }
            break

    # 旧ファイル削除
    for old in BASE.glob("PLC-S-002_25件対応_RAW_例外サンプル9_*.txt"):
        old.unlink()
        print(f"Deleted old: {old.name}")

    orig_qty = s9["qty"] + 2
    new_qty = s9["qty"]
    orig_amount = s9["sale_amount"] * orig_qty // new_qty
    new_amount = s9["sale_amount"]
    change_date = s9["ship_date"] + timedelta(days=1)

    content = f"""================================================================
 SAP VA02 - Document Change History
================================================================
Report:        Document Changes Overview (Table: CDHDR / CDPOS)
Export Time:   2026-02-11 09:45:30 JST
Filter:        Document = {s9['ord_no']}

--------------------------------------------------------------
Document:      {s9['ord_no']}
Document Type: Sales Order (OR)
Related Ship:  {s9['ship_no']}
--------------------------------------------------------------

[CHG-000 Document Creation]
Timestamp:     {s9['ship_date'].strftime('%Y-%m-%d')} 09:15:22
User:          SLS004
Action:        CREATE
Item 10:
  Material:    (via product master)
  Quantity:    {orig_qty:>10,} EA
  Net value:   {orig_amount:>12,} JPY

[CHG-001 Field Change]
Timestamp:     {change_date.strftime('%Y-%m-%d')} 14:22:18
User:          SLS004
Action:        UPDATE
Approval WF:   WF-2025-SLS-2341 (approver: SLS002)
Item 10:
  Field:       Quantity (VBAP-KWMENG)
    Before:    {orig_qty:>10,} EA
    After:     {new_qty:>10,} EA
  Field:       Net value (VBAP-NETWR)
    Before:    {orig_amount:>12,} JPY
    After:     {new_amount:>12,} JPY

--------------------------------------------------------------
END OF CHANGE HISTORY
--------------------------------------------------------------
"""
    path = BASE / f"PLC-S-002_25件対応_RAW_例外サンプル9_SAP_VA02変更履歴_{s9['ship_no']}.txt"
    path.write_text(content, encoding="utf-8")
    print(f"Simplified: {path.name}")


# ==============================================================
# 4. 例外サンプル14 (SAPバッチログのみ、分析コメント削除)
# ==============================================================
def simplify_sample14_txt():
    wb = openpyxl.load_workbook(BASE / "PLC-S-002_25件サンプル対応エビデンス.xlsx")
    ws = wb.active
    s14 = None
    for r in range(17, 42):
        if ws.cell(row=r, column=1).value == 14:
            s14 = {
                "ship_no": ws.cell(row=r, column=2).value,
                "ord_no": ws.cell(row=r, column=4).value,
                "jv_no": ws.cell(row=r, column=10).value,
                "ship_date": ws.cell(row=r, column=3).value,
                "sale_date": ws.cell(row=r, column=11).value,
                "ship_amount": ws.cell(row=r, column=9).value,
                "sale_amount": ws.cell(row=r, column=12).value,
            }
            break

    for old in BASE.glob("PLC-S-002_25件対応_RAW_例外サンプル14_*.txt"):
        old.unlink()
        print(f"Deleted old: {old.name}")

    content = f"""================================================================
 SAP Background Job Log - ZSD_SHIP_SALES_MATCH
================================================================
Job Name:      ZSD_SHIP_SALES_MATCH
Schedule:      Daily nightly batch (starts 01:00 JST)
Export Time:   2026-02-11 09:52:10 JST
Filter:        Target shipment = {s14['ship_no']}

--------------------------------------------------------------
[Run 1 - Scheduled]
--------------------------------------------------------------
Start time:    {s14['ship_date'].strftime('%Y-%m-%d')} 23:58:00
End time:      {s14['ship_date'].strftime('%Y-%m-%d')} 23:58:04
Batch ID:      ZSD-{s14['ship_date'].strftime('%Y%m%d')}-014
User:          SAP_BATCH
Target:        {s14['ship_no']} / shipment posted {s14['ship_date'].strftime('%Y-%m-%d')}

Processing:
  WMS source:      Record found (shipment completed)
                   Amount: {s14['ship_amount']:,} JPY
  SAP FI source:   No corresponding sales journal within search range
  Tolerance check: N/A (no counter-record to compare)

Result:          EXCEPTION
Reason code:     SALES_JOURNAL_NOT_FOUND
Action:          Transferred to ZSD_UNMATCH queue for retry on next run

--------------------------------------------------------------
[Run 2 - Retry]
--------------------------------------------------------------
Start time:    {(s14['sale_date'] + timedelta(days=1)).strftime('%Y-%m-%d')} 01:15:00
End time:      {(s14['sale_date'] + timedelta(days=1)).strftime('%Y-%m-%d')} 01:15:03
Batch ID:      ZSD-{(s14['sale_date'] + timedelta(days=1)).strftime('%Y%m%d')}-RETRY-014
User:          SAP_BATCH
Target:        {s14['ship_no']} / retry from ZSD_UNMATCH queue

Processing:
  WMS source:      Record found
                   Amount: {s14['ship_amount']:,} JPY
  SAP FI source:   Sales journal found: {s14['jv_no']} posted {s14['sale_date'].strftime('%Y-%m-%d')} 17:30:12
                   Amount: {s14['sale_amount']:,} JPY
  Tolerance check: 0 JPY diff, 0.0% (within tolerance +/- 10,000 JPY / +/- 5.0%)

Result:          OK
Match type:      AMOUNT_MATCH_EXACT
Action:          Record closed in ZSD_UNMATCH queue

--------------------------------------------------------------
END OF JOB LOG
--------------------------------------------------------------
"""
    path = BASE / f"PLC-S-002_25件対応_RAW_例外サンプル14_SAPバッチログ_{s14['ship_no']}.txt"
    path.write_text(content, encoding="utf-8")
    print(f"Simplified: {path.name}")


# ==============================================================
# 5. 25件サンプル対応エビデンス.xlsx を「監査対象25件サンプルリスト.xlsx」に簡素化
# ==============================================================
def simplify_sample_list():
    old_path = BASE / "PLC-S-002_25件サンプル対応エビデンス.xlsx"

    # 既存データを読み込み
    wb_old = openpyxl.load_workbook(old_path)
    ws_old = wb_old.active
    samples = []
    for r in range(17, 42):
        if ws_old.cell(row=r, column=1).value is None:
            continue
        samples.append({
            "no": ws_old.cell(row=r, column=1).value,
            "ship_no": ws_old.cell(row=r, column=2).value,
            "ship_date": ws_old.cell(row=r, column=3).value,
            "ord_no": ws_old.cell(row=r, column=4).value,
            "cid": ws_old.cell(row=r, column=5).value,
            "cname": ws_old.cell(row=r, column=6).value,
            "pcode": ws_old.cell(row=r, column=7).value,
            "qty": ws_old.cell(row=r, column=8).value,
            "ship_amount": ws_old.cell(row=r, column=9).value,
            "jv_no": ws_old.cell(row=r, column=10).value,
            "sale_date": ws_old.cell(row=r, column=11).value,
            "sale_amount": ws_old.cell(row=r, column=12).value,
        })

    # 旧ファイル削除
    old_path.unlink()
    print(f"Deleted old: {old_path.name}")

    # 新しい簡素化ファイル
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "25件サンプルリスト"

    # タイトル
    ws.cell(row=1, column=1, value="【PLC-S-002】 監査対象25件サンプルリスト")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    ws.cell(row=2, column=1, value="（RAWデータをナビゲートするための取引リスト。分析・判定は含まない）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=10, italic=True, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    # メタ情報（純粋に抽出条件のみ）
    meta = [
        ("母集団", "FY2025 出荷実績 3,158件（SAP VA05 + WMS出荷実績）"),
        ("抽出方法", "系統抽出 / 間隔126件 / 開始位置57（無作為決定）"),
        ("抽出日時", "2026-02-10 11:15 JST"),
        ("関連RAWデータ", "PLC-S-002_25件対応_RAW_*.csv (WMS/SAP FI/バッチログ)"),
    ]
    for i, (k, v) in enumerate(meta):
        r = 4 + i
        ws.cell(row=r, column=1, value=k).font = BBOLD
        ws.cell(row=r, column=1).fill = FILL_META
        ws.cell(row=r, column=1).border = BRD
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.cell(row=r, column=4, value=v).font = BFONT
        ws.cell(row=r, column=4).border = BRD
        ws.cell(row=r, column=4).alignment = L_
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=12)

    # ヘッダ（12列のみ・純粋な取引情報）
    headers = [
        "サンプル\n№", "出荷番号", "出荷日", "受注番号",
        "顧客\nコード", "顧客名", "製品コード", "数量",
        "出荷金額\n(円)", "売上仕訳\n番号", "売上計上日", "売上金額\n(円)"
    ]
    hr = 10
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[hr].height = 36

    # 明細
    for idx, s in enumerate(samples):
        r = hr + 1 + idx
        row_data = [
            s["no"], s["ship_no"], s["ship_date"], s["ord_no"],
            s["cid"], s["cname"], s["pcode"], s["qty"],
            s["ship_amount"], s["jv_no"], s["sale_date"], s["sale_amount"]
        ]
        for c_i, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 5, 7, 8, 10, 11):
                cell.alignment = C_
                if c_i in (3, 11):
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (9, 12):
                cell.alignment = R_
                cell.number_format = "#,##0"
            else:
                cell.alignment = L_
        ws.row_dimensions[r].height = 22

    # 列幅
    widths = [6, 16, 11, 15, 10, 16, 12, 10, 14, 16, 11, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hr + 1}"

    new_path = BASE / "PLC-S-002_監査対象25件サンプルリスト.xlsx"
    wb.save(new_path)
    print(f"Created: {new_path.name}")


if __name__ == "__main__":
    delete_files()
    simplify_unmatch_csv()
    simplify_sample9_txt()
    simplify_sample14_txt()
    simplify_sample_list()
    print("\nPLC-S-002 evidence simplified under new policy.")
