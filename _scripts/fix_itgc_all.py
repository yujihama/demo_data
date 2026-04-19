"""Fix ITGC evidence inconsistencies.

Intentional deficiencies to preserve:
- ITGC-AC-003 DEF-2026-001: 退職者5名中2名(SLS099/PUR099)の停止遅延(11日/18日)
- ITGC-AC-002 REQ-2026-002: Q3/Q4 SUIM抽出条件が判断保留 (keep Q3/Q4 SUIM header ambiguous)
- ITGC-CM-002: 軽微例外1件 (keep one UAT with minor defect)

Fixes applied:
1. ITGC-AC-001 SU01 timestamps (15/25 were BEFORE WF final approval) → after approval
2. ITGC-AC-001 CSV Sample 11 parse issue (MM_USER,PO_CREATE two-col) → single role
3. ITGC-AC-001 Add 社員番号 column to SU01 & WF CSVs
4. ITGC-AC-003 Expand retired users 2 → 5 (keep 2 delays, add 3 timely stops)
5. ITGC-AC-003 Remove reviewer conclusion embed
6. ITGC-AC-004 Expand 1 month → 12 months of data
7. ITGC-AC-004 Remove leader review approval embed
8. ITGC-CM-001 Register xlsx ← rewrite to match Detailed CSV & SampleSubmission
9. ITGC-CM-001 SampleSubmission: remove "_25件対応_" naming
10. ITGC-CM-002 UAT 21/25 content mismatches → regenerate to match Register
11. ITGC-CM-002 UAT dates: after 申請日, before 本番移送日
12. ITGC-CM-003 STMS dates ← match Register 本番移送日 exactly
13. ITGC-CM-003 Q2-Q3 STMS file consolidate (rename to Population, make consistent subset)
14. ITGC-OM-001 10 FY2026 dates → shift within FY2025
15. ITGC-OM-002 Timeline inversion (INVESTIGATING > RESOLVED) → fix ordering
"""

import csv
import os
import random
import re
import sys
import io
from datetime import datetime, timedelta
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

ROOT = Path(r"C:\Users\nyham\work\demo_data\4.evidence\ITGC")

# ==============================================================
# Fix 1+2+3: ITGC-AC-001 SU01 CSV
# ==============================================================
def fix_su01_csv():
    """Fix SU01 CSV:
    - SU01 CREATE timestamp must be AFTER WF final approval (WF=13:00, so 14:00-16:00)
    - Sample 11 had extra field; make single role
    - Add 社員番号 column (IT004 → E0054)
    """
    path = ROOT / "SAP_SU01_UserMaster_ChangeHistory_FY2025.csv"

    # Read WF final approval timestamps to use as baseline
    wf_final = {}  # sno -> datetime
    wf_path = ROOT / "Workflow_UserRegistration_ApprovalHistory_FY2025.csv"
    with open(wf_path, encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            if not row or row[0].startswith('#') or row[0] == 'タイムスタンプ':
                continue
            if len(row) < 6: continue
            ts, wfno, sno, reqno, actor, action = row[:6]
            if 'IT003' in actor:
                wf_final[sno] = datetime.strptime(ts, '%Y-%m-%d %H:%M:%S')

    # Rewrite SU01
    lines = [
        "# SAP S/4HANA - Transaction SU01",
        "# Report:   User Master Change History (Table USR02 / USH02)",
        "# Export:   2026-02-15 09:20:30 JST",
        "# Filter:   FY2025 User Master Create/Modify entries (population + sample rows)",
        "",
        "タイムスタンプ,サンプル№,申請番号,ユーザID,アクション,部門,付与ロール,実行ユーザ(SAP),実行ユーザ(社員番号),ステータス",
    ]

    # Sample rows from original (25 samples)
    # Parse with fix for sample 11
    sample_rows = []
    with open(path, encoding='utf-8') as f:
        for line in f:
            if line.startswith('#') or line.startswith('タイムスタンプ') or not line.strip():
                continue
            parts = line.strip().split(',')
            # normal: ts, sno, reqno, uid, action, dept, role, exec, status = 9 fields
            # sample 11: has 10 fields because MM_USER,PO_CREATE
            if len(parts) == 10:
                # merge 6,7 into single role
                parts = parts[:6] + [f"{parts[6]}+{parts[7]}"] + parts[8:]
            if len(parts) >= 9:
                sample_rows.append(parts[:9])

    # Build new sample rows with corrected timestamps
    # Map role col: position 6. Fix to single role per business design.
    # For sample 11 (was dual-role), choose single role (MM_USER).
    role_fix = {'11': 'MM_USER'}

    for parts in sample_rows:
        ts_orig = parts[0]
        sno = parts[1]
        reqno = parts[2]
        uid = parts[3]
        action = parts[4]
        dept = parts[5]
        role = parts[6]
        exec_sap = parts[7]
        status = parts[8]

        if sno in role_fix:
            role = role_fix[sno]

        # Fix timestamp: AFTER WF final approval (13:00 + 1-3h)
        if sno in wf_final:
            base = wf_final[sno]
            # deterministic offset based on sno
            h_off = (int(sno) * 7) % 4 + 1  # 1-4h after 13:00
            m_off = (int(sno) * 13) % 60
            new_ts = base + timedelta(hours=h_off, minutes=m_off)
            ts_new = new_ts.strftime('%Y-%m-%d %H:%M:%S')
        else:
            ts_new = ts_orig

        # Map IT004 → 社員番号 E0054
        emp_no = 'E0054' if 'IT004' in exec_sap else ''

        lines.append(f"{ts_new},{sno},{reqno},{uid},{action},{dept},{role},{exec_sap},{emp_no},{status}")

    lines.append("")
    lines.append(f"# Records: {len(sample_rows)}")

    with open(path, 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))

    print(f"[Fixed] SU01 CSV: {len(sample_rows)} rows, timestamps after WF approval")


def fix_workflow_csv():
    """Add 社員番号 column to Workflow CSV.
    Also convert '部門長' generic to specific 部門長 with SAP_UID + 社員番号.
    """
    path = ROOT / "Workflow_UserRegistration_ApprovalHistory_FY2025.csv"

    # Dept → head mapping
    head_map = {
        '営業本部': ('田中 太郎 (SLS001)', 'E0021'),
        '購買部': ('木村 浩二 (PUR001)', 'E0031'),
        '製造本部': ('森 和雄 (MFG001)', 'E0041'),
        '経理部': ('佐藤 一郎 (ACC001)', 'E0011'),
        '人事部': ('近藤 文子 (HR001)', 'E0061'),
        '情報システム部': ('岡田 宏 (IT001)', 'E0051'),
        '品質保証部': ('森 和雄 (MFG001)', 'E0041'),  # quality assurance is under mfg for this co.
        '総務部': ('前田 美香 (GA001)', 'E0071'),
    }

    # Read SU01 CSV to get dept per sno
    su01_path = ROOT / "SAP_SU01_UserMaster_ChangeHistory_FY2025.csv"
    sno_dept = {}
    sno_apluser = {}
    with open(su01_path, encoding='utf-8') as f:
        for line in f:
            if line.startswith('#') or line.startswith('タイムスタンプ') or not line.strip():
                continue
            parts = line.strip().split(',')
            if len(parts) >= 6:
                sno_dept[parts[1]] = parts[5]

    # Rewrite workflow CSV
    new_lines = [
        "# Workflow System - User Registration Approval History",
        "# Export:   2026-02-15 09:25:12 JST",
        "",
        "タイムスタンプ,ワークフロー番号,サンプル№,申請番号,アクター(氏名SAP_UID),アクター(社員番号),アクション",
    ]
    with open(path, encoding='utf-8') as f:
        for line in f:
            if line.startswith('#') or line.startswith('タイムスタンプ') or not line.strip():
                continue
            parts = line.strip().split(',')
            if len(parts) < 6: continue
            ts, wfno, sno, reqno, actor, action = parts[:6]
            dept = sno_dept.get(sno, '')

            if '申請者' in actor:
                actor_new = '申請者(所属部門担当)'
                emp_no = '-'
            elif '部門長' in actor:
                if dept in head_map:
                    actor_new, emp_no = head_map[dept]
                else:
                    actor_new, emp_no = actor, '-'
            elif 'IT003' in actor:
                actor_new = '加藤 洋子 (IT003)'
                emp_no = 'E0053'
            else:
                actor_new = actor
                emp_no = '-'

            new_lines.append(f"{ts},{wfno},{sno},{reqno},{actor_new},{emp_no},{action}")

    new_lines.append("")
    new_lines.append(f"# Records: 75")

    with open(path, 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(new_lines))

    print("[Fixed] Workflow CSV: added 社員番号 column, expanded 部門長 to specific head")


# ==============================================================
# Fix 4+5: ITGC-AC-003 SM20 retired users
# ==============================================================
def fix_retired_users():
    """Expand 2 retired users (SLS099, PUR099) to 5 (add 3 timely).
    Keep SLS099/PUR099 delays as DEF-2026-001.
    Remove reviewer conclusion embed.
    """
    path = ROOT / "SAP_SM20_SecurityAuditLog_RetiredUsers.csv"

    # New 5 retired users:
    # SLS099 退職 9/30, 停止 10/11 (11日遅延 - INTENTIONAL) → delayed
    # PUR099 退職 11/15, 停止 12/3 (18日遅延 - INTENTIONAL) → delayed
    # MFG099 退職 2025/6/30, 停止 2025/6/30 (当日) → timely
    # ACC099 退職 2025/8/31, 停止 2025/9/1 (1日後) → within 3日
    # HR099  退職 2026/1/31, 停止 2026/2/2 (2日後) → within 3日

    lines = [
        "# SAP SM20 Security Audit Log",
        "# Report:   SAL (Security Audit Log) Login Event Records",
        "# Target:   FY2025退職者全数 (5名)",
        "# Period:   各退職者の退職前2週間～停止完了日",
        "# Export:   2026-02-15 14:22:30 JST",
        "",
        "ユーザID,退職日,停止日,対象期間,ログイン件数,備考",
        "MFG099,2025-06-30,2025-06-30,2025-06-16 ～ 2025-06-30,0,退職日当日停止",
        "ACC099,2025-08-31,2025-09-01,2025-08-17 ～ 2025-09-01,0,退職翌営業日停止",
        "SLS099,2025-09-30,2025-10-11,2025-09-16 ～ 2025-10-11,0,停止遅延11日(DEF-2026-001)",
        "PUR099,2025-11-15,2025-12-03,2025-11-01 ～ 2025-12-03,0,停止遅延18日(DEF-2026-001)",
        "HR099,2026-01-31,2026-02-02,2026-01-17 ～ 2026-02-02,0,退職翌々日停止",
        "",
        "# Records: 5 / ログイン件数集計結果",
        "# 詳細ログイン明細は対象期間でログイン件数ゼロのため本レポートに明細行なし",
    ]

    with open(path, 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))

    print("[Fixed] SM20 Retired Users: 2→5 users, removed reviewer conclusion")


# ==============================================================
# Fix 6+7: ITGC-AC-004 SM20 Privileged User
# ==============================================================
def fix_privileged_log():
    """Expand 1 month to 12 months of privileged ID operation log.
    Remove leader review approval embed.
    """
    path = ROOT / "SAP_SM20_PrivilegedUser_OperationLog_202511.csv"
    new_path = ROOT / "SAP_SM20_PrivilegedUser_OperationLog_FY2025.csv"

    # 12 months of data: Apr 2025 - Mar 2026
    months = [
        ('2025-04', 6), ('2025-05', 8), ('2025-06', 7), ('2025-07', 9),
        ('2025-08', 7), ('2025-09', 8), ('2025-10', 8), ('2025-11', 9),
        ('2025-12', 7), ('2026-01', 8), ('2026-02', 9), ('2026-03', 10),
    ]

    tx_patterns = [
        ('RZ20', 'システム監視', '正常', '恒常業務'),
        ('SM37', 'バッチジョブ確認', '正常', '恒常業務'),
        ('SM51', 'サーバ起動停止', '正常', '事前申請WF-IT-{wf}'),
        ('DB13', 'DBバックアップ確認', '正常', '恒常業務'),
        ('SM59', 'RFC接続テスト', '正常', '事前申請WF-IT-{wf}'),
        ('SE38', 'ABAP実行', '正常', '事前申請WF-IT-{wf}'),
        ('RZ10', 'プロファイル確認', '正常', '恒常業務'),
        ('SU01', 'ユーザマスタ確認', '正常', '棚卸支援'),
        ('STMS', '移送管理', '正常', '変更REL-2025-{rel}'),
        ('ST22', 'ダンプ分析', '正常', '恒常業務'),
        ('SM21', 'システムログ', '正常', '恒常業務'),
    ]

    rng = random.Random(40001)
    lines = [
        "# SAP S/4HANA - Transaction SM20",
        "# Report:   Privileged User Operation Log (BASIS/ALL_READ)",
        "# Target:   IT001(E0051)岡田 宏, IT002(E0052)吉田 雅彦",
        "# Period:   FY2025 (2025/04/01 - 2026/03/31) 12ヶ月分",
        "# Export:   2026-04-05 08:30:00 JST",
        "",
        "タイムスタンプ,ユーザID(SAP),ユーザID(社員番号),トランザクション,対象,結果,事前申請/区分",
    ]

    wf_counter = 90
    rel_counter = 1
    total = 0
    for ym, cnt in months:
        y, m = ym.split('-')
        for i in range(cnt):
            day = rng.randint(1, 28)
            hh = rng.randint(8, 19)
            mm = rng.randint(0, 59)
            ss = rng.randint(0, 59)
            ts = f"{y}-{m}-{day:02d} {hh:02d}:{mm:02d}:{ss:02d}"
            user_sap = rng.choice(['IT001', 'IT002'])
            user_emp = 'E0051' if user_sap == 'IT001' else 'E0052'
            tx_name, tx_desc, result, ref_pat = rng.choice(tx_patterns)
            if '{wf}' in ref_pat:
                wf_counter += 1
                ref = ref_pat.format(wf=f'2025-{wf_counter:03d}')
            elif '{rel}' in ref_pat:
                rel_counter += 2
                ref = ref_pat.format(rel=f'{rel_counter:03d}')
            else:
                ref = ref_pat
            lines.append(f"{ts},{user_sap},{user_emp},{tx_name},{tx_desc},{result},{ref}")
            total += 1

    lines.append("")
    lines.append(f"# Records: {total}")

    with open(new_path, 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))

    # Delete old 1-month file
    if path.exists():
        path.unlink()

    print(f"[Fixed] SM20 Privileged: expanded 1 month → 12 months ({total} records), removed leader review embed, renamed to FY2025")
    return new_path.name, path.name


# ==============================================================
# Fix 8+9: ITGC-CM-001 Register consistency
# ==============================================================
def fix_change_register():
    """Rebuild Change Register xlsx to be consistent with Detailed CSV.
    Remove '_25件対応_' from Sample Submission List.
    """
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Read Detailed CSV (authoritative for 25 samples)
    detailed_path = ROOT / "ChangeManagement_Register_Detailed_FY2025.csv"
    sample_info = {}
    with open(detailed_path, encoding='utf-8') as f:
        for line in f:
            if line.startswith('#') or line.startswith('サンプル') or not line.strip():
                continue
            parts = line.strip().split(',')
            if len(parts) < 11: continue
            sno, rel, apl, applicant, content, risk, test, mv, app_it, app_biz, status = parts[:11]
            sample_info[rel] = {
                'sno': sno, 'apl': apl, 'applicant': applicant, 'content': content,
                'risk': risk, 'test': test, 'mv': mv, 'app_it': app_it, 'app_biz': app_biz, 'status': status
            }

    # Build full population (42 records): 25 sample (even REL 002-050) + 17 non-sample
    # Keep even REL consistent with Detailed CSV, generate odd REL + extras (051..059) as non-sample
    rng = random.Random(50001)
    content_pool = [
        '売上レポート機能追加', '購買申請画面の改善', '販売価格マスタ連携IF修正',
        '標準原価計算バッチ修正', 'セキュリティパッチ適用', 'バックアップバッチ改善',
        '仕入先マスタ項目拡張', 'ワークフロー承認ルーティング変更',
        '連結仕訳バリデーション強化', '勘定科目マスタ追加', 'SUIM定期レポート出力機能追加',
    ]

    # 42 population records
    population = []
    # 25 samples (even REL 002-050)
    for i in range(1, 26):
        rel = f'REL-2025-{i*2:03d}'
        if rel in sample_info:
            info = sample_info[rel]
            population.append({
                'rel': rel,
                'apl': info['apl'],
                'applicant': '業務部門担当' if i % 3 == 0 else info['applicant'],
                'content': info['content'],
                'scope': content_to_scope(info['content']),
                'test': info['test'],
                'mv': info['mv'],
                'app_it': info['app_it'],
                'app_biz': info['app_biz'],
                'status': info['status'],
            })

    # 17 non-sample records (odd REL 001-033 + extras 051-059)
    rels_non = [f'REL-2025-{n:03d}' for n in [1,3,5,7,9,11,13,15,17,19,21,23,25,27,29,31,33]]
    for rel in rels_non:
        content = rng.choice(content_pool)
        # random date in FY2025
        m = rng.randint(4, 15)  # 4-15 → month 4-12 of 2025, 1-3 of 2026
        if m > 12:
            apl_d = datetime(2026, m-12, rng.randint(1, 28))
        else:
            apl_d = datetime(2025, m, rng.randint(1, 28))
        mv_d = apl_d + timedelta(days=rng.randint(14, 30))
        population.append({
            'rel': rel,
            'apl': apl_d.strftime('%Y-%m-%d'),
            'applicant': rng.choice(['業務部門担当', '加藤 洋子 (IT003)']),
            'content': content,
            'scope': content_to_scope(content),
            'test': 'UAT合格',
            'mv': mv_d.strftime('%Y-%m-%d'),
            'app_it': '加藤 洋子 (IT003)',
            'app_biz': rng.choice(['業務部門長', '岡田 宏 (IT001)']),
            'status': '完了',
        })

    # Sort by REL number
    population.sort(key=lambda x: int(x['rel'].replace('REL-2025-', '')))

    # Write xlsx
    reg_xlsx = ROOT / "ChangeManagement_Register_FY2025.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "変更管理一覧"

    ws.cell(1, 1, "【ITGC-CM-001 統制実施記録】 FY2025 変更管理一覧")
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(2, 1, "出力日時: 2026/2/18 / 出力者: IT003 加藤 洋子 (E0053) / 対象: FY2025期間中のSAP変更申請 全42件")

    headers = ['REL番号', '申請日', '申請者', '変更内容概要', '影響範囲', 'テスト実施', '本番移送日', '承認者1(情シス)', '承認者2(業務)', 'ステータス']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(4, c, h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='305496')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for i, rec in enumerate(population, 5):
        ws.cell(i, 1, rec['rel'])
        apl_d = rec['apl']
        if isinstance(apl_d, str):
            try:
                apl_d = datetime.strptime(apl_d, '%Y-%m-%d')
            except:
                pass
        ws.cell(i, 2, apl_d)
        ws.cell(i, 3, rec['applicant'])
        ws.cell(i, 4, rec['content'])
        ws.cell(i, 5, rec['scope'])
        ws.cell(i, 6, rec['test'])
        mv_d = rec['mv']
        if isinstance(mv_d, str):
            try:
                mv_d = datetime.strptime(mv_d, '%Y-%m-%d')
            except:
                pass
        ws.cell(i, 7, mv_d)
        ws.cell(i, 8, rec['app_it'])
        ws.cell(i, 9, rec['app_biz'])
        ws.cell(i, 10, rec['status'])

    # Column widths
    widths = [16, 12, 22, 30, 14, 12, 14, 24, 22, 10]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+c)].width = w

    wb.save(reg_xlsx)
    print(f"[Fixed] ChangeManagement_Register_FY2025.xlsx: rebuilt with 42 records consistent with Detailed CSV")

    # Fix SampleSubmissionList: remove _25件対応_ naming
    sample_xlsx = ROOT / "ChangeManagement_SampleSubmissionList_FY2025.xlsx"
    wb = load_workbook(sample_xlsx)
    ws = wb.active
    for r in range(1, ws.max_row+1):
        for c in range(1, ws.max_column+1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and '_25件対応_' in v:
                new_v = v.replace('ITGC-CM-001_25件対応_RAW_*.csv', 'ChangeManagement_Register_Detailed_FY2025.csv')
                new_v = new_v.replace('_25件対応_', '_')
                ws.cell(r, c, new_v)
                print(f"  Replaced: {v} → {new_v}")
    wb.save(sample_xlsx)
    print(f"[Fixed] SampleSubmissionList: removed '_25件対応_' references")


def content_to_scope(content):
    """Map 変更内容 to 影響範囲"""
    if '売上' in content or '販売価格' in content:
        return '販売管理'
    if '購買' in content or '仕入先' in content:
        return '購買/MM'
    if '標準原価' in content:
        return '原価計算'
    if 'ワークフロー' in content:
        return '全業務'
    if 'セキュリティ' in content or 'バックアップ' in content:
        return '基盤'
    if '勘定科目' in content or '連結' in content:
        return '経理/FI'
    if 'SUIM' in content:
        return '情シス内'
    return '業務共通'


# ==============================================================
# Fix 10+11: ITGC-CM-002 UAT files
# ==============================================================
def fix_uat_files():
    """Rewrite each UAT file so 変更件名 matches Register 変更内容.
    Ensure UAT date is after 申請日 and before 本番移送日.
    Keep ONE UAT with a minor defect (軽微例外).
    """
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    # Read Detailed CSV
    detailed_path = ROOT / "ChangeManagement_Register_Detailed_FY2025.csv"
    sample_info = {}
    with open(detailed_path, encoding='utf-8') as f:
        for line in f:
            if line.startswith('#') or line.startswith('サンプル') or not line.strip():
                continue
            parts = line.strip().split(',')
            if len(parts) < 11: continue
            sno, rel, apl, applicant, content, risk, test, mv, app_it, app_biz, status = parts[:11]
            sample_info[rel] = {
                'sno': sno, 'apl': apl, 'content': content, 'mv': mv
            }

    # Content → test case pack
    tc_packs = {
        '売上レポート機能追加': [
            ('TC-売上レ-001', '正常系：月次売上レポート出力', 'F', '販売管理'),
            ('TC-売上レ-002', '正常系：顧客別集計', 'F', '販売管理'),
            ('TC-売上レ-003', '正常系：製品別集計', 'F', '販売管理'),
            ('TC-売上レ-004', '正常系：CSV/PDFエクスポート', 'F', '販売管理'),
            ('TC-売上レ-005', '性能系：1年分データ出力', 'P', '販売管理'),
        ],
        '購買申請画面の改善': [
            ('TC-購買申-001', '正常系：購買依頼作成', 'F', 'MM'),
            ('TC-購買申-002', '正常系：添付ファイル機能', 'F', 'MM'),
            ('TC-購買申-003', 'UI系：入力バリデーション', 'F', 'MM'),
            ('TC-購買申-004', 'UI系：レスポンシブ表示', 'F', 'MM'),
        ],
        '販売価格マスタ連携IF修正': [
            ('TC-販売価-001', '正常系：価格マスタ一括更新（10件）', 'F', '販売管理'),
            ('TC-販売価-002', '正常系：価格マスタ一括更新（100件）', 'F', '販売管理'),
            ('TC-販売価-003', '異常系：排他エラー発生時のリトライ動作', 'F', '販売管理'),
            ('TC-販売価-004', '異常系：3回リトライ失敗時のエラー通知', 'F', '販売管理'),
            ('TC-販売価-005', '境界系：連続同一顧客更新（100件）', 'F', '販売管理'),
            ('TC-販売価-006', '性能系：5000件一括実行（5分以内）', 'P', '販売管理'),
            ('TC-販売価-007', 'ログ出力内容確認（実行前後・リトライ含む）', 'F', '販売管理'),
        ],
        '標準原価計算バッチ修正': [
            ('TC-標準原-001', '正常系：月次原価計算の実行', 'F', '管理会計'),
            ('TC-標準原-002', '正常系：差異計算（材料/労務/製造間接）', 'F', '管理会計'),
            ('TC-標準原-003', '境界系：期首在庫ゼロケース', 'F', '管理会計'),
            ('TC-標準原-004', '異常系：マスタ不整合時のエラー処理', 'F', '管理会計'),
            ('TC-標準原-005', '性能系：3000品目処理時間', 'P', '管理会計'),
        ],
        'セキュリティパッチ適用': [
            ('TC-セキュ-001', '正常系：SAPシステム起動', 'F', '基盤'),
            ('TC-セキュ-002', '正常系：ユーザログイン確認', 'F', '基盤'),
            ('TC-セキュ-003', '正常系：業務トランザクション実行', 'F', '基盤'),
            ('TC-セキュ-004', '非機能：認証強度向上の確認', 'S', '基盤'),
        ],
        'バックアップバッチ改善': [
            ('TC-バック-001', '正常系：フルバックアップ実行', 'F', '基盤'),
            ('TC-バック-002', '正常系：差分バックアップ実行', 'F', '基盤'),
            ('TC-バック-003', '異常系：ストレージ容量不足時', 'F', '基盤'),
            ('TC-バック-004', '性能系：バックアップ完了時間（2時間以内）', 'P', '基盤'),
        ],
        '仕入先マスタ項目拡張': [
            ('TC-仕入先-001', '正常系：既存マスタへの新規項目追加', 'F', 'MM'),
            ('TC-仕入先-002', '正常系：既存データとの互換性確認', 'F', 'MM'),
            ('TC-仕入先-003', '境界系：NULL許容動作', 'F', 'MM'),
            ('TC-仕入先-004', '性能系：5000件マスタ処理', 'P', 'MM'),
        ],
        'ワークフロー承認ルーティング変更': [
            ('TC-ワーク-001', '正常系：¥50万以下 担当承認', 'F', 'ワークフロー'),
            ('TC-ワーク-002', '正常系：¥500万以下 課長承認', 'F', 'ワークフロー'),
            ('TC-ワーク-003', '正常系：¥2000万以下 部長承認', 'F', 'ワークフロー'),
            ('TC-ワーク-004', '正常系：¥1億以下 CFO承認', 'F', 'ワークフロー'),
            ('TC-ワーク-005', '境界系：¥500万ちょうど', 'F', 'ワークフロー'),
            ('TC-ワーク-006', '異常系：承認者不在時の代行ルート', 'F', 'ワークフロー'),
        ],
        '連結仕訳バリデーション強化': [
            ('TC-連結-001', '正常系：消去仕訳生成', 'F', 'FI/連結'),
            ('TC-連結-002', '正常系：通貨換算レート適用', 'F', 'FI/連結'),
            ('TC-連結-003', '境界系：為替変動大時の整合性', 'F', 'FI/連結'),
            ('TC-連結-004', '異常系：未消去残の検出通知', 'F', 'FI/連結'),
        ],
        '勘定科目マスタ追加': [
            ('TC-勘定科-001', '正常系：新規勘定科目登録', 'F', 'FI'),
            ('TC-勘定科-002', '正常系：科目コード重複チェック', 'F', 'FI'),
            ('TC-勘定科-003', '正常系：仕訳登録での使用確認', 'F', 'FI'),
        ],
    }

    # Tester assignment by scope
    testers = {
        '販売管理': ['高橋 美咲 (ACC002)', '石井 健 (ACC006)'],
        'MM': ['清水 智明 (PUR003)', '松本 香織 (SLS004)'],
        '管理会計': ['中村 真理 (ACC004)', '高橋 美咲 (ACC002)'],
        '基盤': ['吉田 雅彦 (IT002)', '加藤 洋子 (IT003)'],
        'ワークフロー': ['石井 健 (ACC006)', '清水 智明 (PUR003)'],
        'FI/連結': ['中村 真理 (ACC004)', '佐藤 一郎 (ACC001)'],
        'FI': ['高橋 美咲 (ACC002)', '中村 真理 (ACC004)'],
    }

    # Choose 1 sample with minor defect (軽微例外) - sample 16 (REL-2025-032 セキュリティパッチ適用)
    MINOR_EXCEPTION_SNO = '16'

    rng = random.Random(60001)
    xray_rows = []  # for Xray CSV regen

    uat_files = sorted([f for f in os.listdir(ROOT) if f.startswith('UATテスト結果_REL') and f.endswith('.xlsx')])

    for fn in uat_files:
        rel = fn.replace('UATテスト結果_', '').replace('.xlsx', '')
        if rel not in sample_info:
            continue
        info = sample_info[rel]
        content = info['content']
        apl_d = datetime.strptime(info['apl'], '%Y-%m-%d')
        mv_d = datetime.strptime(info['mv'], '%Y-%m-%d')
        # UAT day: midpoint between apl and mv, at least 3 days before mv
        span = (mv_d - apl_d).days
        offset = max(1, min(span - 3, span // 2))
        uat_d = apl_d + timedelta(days=offset)

        # Test cases for this content type
        if content in tc_packs:
            cases = tc_packs[content]
        else:
            # fallback generic
            cases = [
                ('TC-汎用-001', '正常系：基本動作', 'F', '業務共通'),
                ('TC-汎用-002', '正常系：データ整合性', 'F', '業務共通'),
                ('TC-汎用-003', '異常系：エラーハンドリング', 'F', '業務共通'),
            ]

        scope = cases[0][3]
        tester = rng.choice(testers.get(scope, ['中村 真理 (ACC004)']))

        # Build xlsx
        wb = Workbook()
        ws = wb.active
        ws.title = 'UATテスト結果'

        ws.cell(1, 1, f'UATテスト結果報告書 / {rel}')
        ws.cell(1, 1).font = Font(bold=True, size=14)

        ws.cell(3, 1, 'REL番号'); ws.cell(3, 3, rel)
        ws.cell(4, 1, '変更件名'); ws.cell(4, 3, content)
        ws.cell(5, 1, 'UAT実施日'); ws.cell(5, 3, uat_d.strftime('%Y年%m月%d日'))
        ws.cell(6, 1, '実施者'); ws.cell(6, 3, tester)
        ws.cell(7, 1, '対象モジュール'); ws.cell(7, 3, scope)
        ws.cell(8, 1, '総ケース数'); ws.cell(8, 3, len(cases))

        # Bold labels
        for r in [3, 4, 5, 6, 7, 8]:
            ws.cell(r, 1).font = Font(bold=True)

        # Header
        headers = ['№', 'ケースID', 'ケース名', '種別', '実施日時', '結果', '欠陥ID', '備考']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(11, c, h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='305496')

        is_exception = (info['sno'] == MINOR_EXCEPTION_SNO)

        for idx, (tc_id, tc_name, tc_type, _) in enumerate(cases, 1):
            # Time within UAT day
            hh = 9 + (idx - 1) // 2
            mm = (idx * 17) % 60
            exec_ts = uat_d.replace(hour=hh, minute=mm).strftime('%Y-%m-%d %H:%M')

            # First pass
            result = 'PASS'
            defect = ''
            memo = ''

            # Minor exception: one case fails on first pass then retest passes
            if is_exception and idx == len(cases):
                result = 'PASS'
                defect = f'DEF-UAT-{rel[-3:]}-01'
                memo = '初回NG→欠陥修正後リテスト合格(軽微例外)'

            ws.cell(12 + idx - 1, 1, idx)
            ws.cell(12 + idx - 1, 2, tc_id)
            ws.cell(12 + idx - 1, 3, tc_name)
            ws.cell(12 + idx - 1, 4, tc_type)
            ws.cell(12 + idx - 1, 5, exec_ts)
            ws.cell(12 + idx - 1, 6, result)
            ws.cell(12 + idx - 1, 7, defect)
            ws.cell(12 + idx - 1, 8, memo)

            # Collect for Xray
            xray_rows.append({
                'rel': rel, 'tc_id': tc_id, 'tc_name': tc_name, 'type': tc_type,
                'scope': scope, 'ts': exec_ts + ':00', 'tester': tester,
                'status': result, 'defect': defect, 'memo': memo
            })

        # Column widths
        widths = [6, 18, 36, 6, 18, 8, 16, 40]
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64+c)].width = w

        wb.save(ROOT / fn)

    print(f"[Fixed] UAT xlsx: {len(uat_files)} files regenerated with matching Register content + 1 minor exception (sno=16)")

    # Regen Xray CSV
    xray_path = ROOT / "Xray_TestExecution_History_FY2025.csv"
    lines = [
        "# Jira Xray Test Execution History Export",
        "# Project: SAP-CHG (SAP Change Management)",
        "# Filter: Test executions linked to UAT test plans in FY2025",
        "# Export: 2026-02-18 13:00:00 JST",
        "# Exporter: IT003 加藤 洋子 (E0053)",
        "# Case Type legend: F=Functional / P=Performance / S=Security / U=Usability",
        "",
        "ExecutionID,REL_Number,TestCaseID,TestCaseName,CaseType,Module,ExecutedAt,Tester,Status,DefectID,Comment",
    ]
    for i, r in enumerate(xray_rows, 1):
        lines.append(f"TEX-{i:05d},{r['rel']},{r['tc_id']},{r['tc_name']},{r['type']},{r['scope']},{r['ts']},{r['tester']},{r['status']},{r['defect']},{r['memo']}")

    with open(xray_path, 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))

    print(f"[Fixed] Xray CSV: regenerated with {len(xray_rows)} executions matching UAT files")


# ==============================================================
# Fix 12+13: ITGC-CM-003 STMS
# ==============================================================
def fix_stms():
    """Make STMS dates match Register 本番移送日 exactly.
    Consolidate two STMS files; keep one as sample file, one as population subset.
    """
    detailed_path = ROOT / "ChangeManagement_Register_Detailed_FY2025.csv"
    sample_info = {}
    with open(detailed_path, encoding='utf-8') as f:
        for line in f:
            if line.startswith('#') or line.startswith('サンプル') or not line.strip():
                continue
            parts = line.strip().split(',')
            if len(parts) < 11: continue
            sno, rel, apl, applicant, content, risk, test, mv, app_it, app_biz, status = parts[:11]
            sample_info[sno] = (rel, mv)

    rng = random.Random(70001)

    # Main STMS FY2025: sample 25 records with EXACT date match
    lines = [
        "# SAP S/4HANA - Transaction STMS",
        "# Report:   Transport Management System History",
        "# Export:   2026-02-18 14:00:00 JST",
        "",
        "タイムスタンプ,サンプル№,TR番号,REL番号,移送者(SAP),移送者(社員番号),移送元,移送先,対象オブジェクト,結果",
    ]
    for sno in sorted(sample_info.keys(), key=int):
        rel, mv = sample_info[sno]
        # Use mv date with time 02:00-05:59 (overnight transport window)
        hh = rng.randint(2, 5)
        mm = rng.randint(0, 59)
        ss = rng.randint(0, 59)
        ts = f"{mv} {hh:02d}:{mm:02d}:{ss:02d}"
        tr = f"XXXK9{rng.randint(0, 99999):05d}"
        lines.append(f"{ts},{sno},{tr},{rel},IT003 加藤 洋子,E0053,DEV,PRD,ABAP + Function Module,成功")

    lines.append("")
    lines.append(f"# Records: 25 (FY2025 監査25サンプル対応)")

    with open(ROOT / "SAP_STMS_ProductionTransport_History_FY2025.csv", 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))

    # Population file: FY2025 全42件. Rename Q2-Q3 → Population
    # Build 42 records - get from rebuilt register xlsx
    from openpyxl import load_workbook
    reg_xlsx = ROOT / "ChangeManagement_Register_FY2025.xlsx"
    wb = load_workbook(reg_xlsx)
    ws = wb.active

    pop_lines = [
        "# SAP STMS - 本番移送履歴 (FY2025 全母集団)",
        "# 出力日時: 2026/02/18 11:22:35",
        "# 出力者: IT003 加藤 洋子 (E0053)",
        "# 対象期間: FY2025 (2025/04/01 - 2026/03/31) 全42件",
        "",
        "タイムスタンプ,TR番号,REL番号,移送者(SAP),移送者(社員番号),移送元,移送先,対象オブジェクト,結果",
    ]

    pop_count = 0
    for r in range(5, ws.max_row + 1):
        rel = ws.cell(r, 1).value
        mv = ws.cell(r, 7).value
        if rel and mv:
            if isinstance(mv, datetime):
                mv_s = mv.strftime('%Y-%m-%d')
            else:
                mv_s = str(mv)[:10]
            hh = rng.randint(2, 5)
            mm = rng.randint(0, 59)
            ss = rng.randint(0, 59)
            ts = f"{mv_s} {hh:02d}:{mm:02d}:{ss:02d}"
            tr = f"XXXK9{rng.randint(0, 99999):05d}"
            pop_lines.append(f"{ts},{tr},{rel},IT003 加藤 洋子,E0053,DEV,PRD,ABAP + Function Module,成功")
            pop_count += 1

    pop_lines.append("")
    pop_lines.append(f"# Records: {pop_count} / 全件成功 / 移送者は情シス部アプリリーダー限定")

    # Write new population file, delete Q2-Q3
    pop_path = ROOT / "SAP_STMS_ProductionTransport_History_FY2025_Population.csv"
    with open(pop_path, 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(pop_lines))

    old_q_path = ROOT / "SAP_STMS_ProductionTransport_History_FY2025Q2-Q3.csv"
    if old_q_path.exists():
        old_q_path.unlink()

    print(f"[Fixed] STMS: sample file with 25 exact-match dates + population file ({pop_count} records), old Q2-Q3 file removed")
    return pop_path.name, old_q_path.name


# ==============================================================
# Fix 14: ITGC-OM-001 DB13 backup
# ==============================================================
def fix_db13_dates():
    """Shift FY2026 dates to fit within FY2025 (2025-04-01 ~ 2026-03-31).
    25 unique days.
    """
    path = ROOT / "SAP_DB13_DatabaseBackup_Log_FY2025.csv"

    # Read all lines
    with open(path, encoding='utf-8') as f:
        lines = f.readlines()

    # Collect unique days
    from collections import OrderedDict
    day_rows = OrderedDict()  # day -> list of row strings
    header_lines = []
    trailer_lines = []
    in_data = False

    for line in lines:
        if line.startswith('#'):
            if in_data:
                trailer_lines.append(line)
            else:
                header_lines.append(line)
        elif line.startswith('タイムスタンプ'):
            header_lines.append(line)
            in_data = True
        elif line.strip():
            m = re.match(r'(\d{4}-\d{2}-\d{2})', line)
            if m:
                day = m.group(1)
                day_rows.setdefault(day, []).append(line)

    # Target 25 unique days within FY2025 (2025-04-01 ~ 2026-03-31)
    # Spread evenly: one per ~14 days
    import datetime as dt
    start = dt.date(2025, 4, 1)
    end = dt.date(2026, 3, 31)
    span = (end - start).days  # 364 days

    target_days = []
    for i in range(25):
        offset = int(span * i / 24)
        d = start + dt.timedelta(days=offset)
        target_days.append(d.strftime('%Y-%m-%d'))

    # Deduplicate & sort
    target_days = sorted(set(target_days))

    # If fewer than 25 after dedup, add additional
    rng = random.Random(80001)
    while len(target_days) < 25:
        d = start + dt.timedelta(days=rng.randint(0, span))
        if d.strftime('%Y-%m-%d') not in target_days:
            target_days.append(d.strftime('%Y-%m-%d'))
    target_days = sorted(target_days)[:25]

    # Map old days to new days (preserve order)
    old_days = list(day_rows.keys())
    day_map = dict(zip(old_days, target_days))

    # Rewrite
    new_lines = []
    new_lines.extend(header_lines)
    for old_day in old_days:
        new_day = day_map[old_day]
        for row in day_rows[old_day]:
            # Replace old date in BK_YYYYMMDD_xxx too
            old_bk = f"BK_{old_day.replace('-', '')}_"
            new_bk = f"BK_{new_day.replace('-', '')}_"
            new_row = row.replace(old_day, new_day).replace(old_bk, new_bk)
            new_lines.append(new_row)
    new_lines.extend(trailer_lines)

    with open(path, 'w', encoding='utf-8', newline='') as f:
        f.writelines(new_lines)

    print(f"[Fixed] DB13: {len(target_days)} backup days all within FY2025 ({target_days[0]} ~ {target_days[-1]})")


# ==============================================================
# Fix 15: ITGC-OM-002 Zabbix timeline
# ==============================================================
def fix_zabbix_timeline():
    """Fix INVESTIGATING > RESOLVED ordering.
    Order: DETECTED < NOTIFIED < ACKNOWLEDGED < INVESTIGATING < RESOLVED < CLOSED
    """
    path = ROOT / "Zabbix_IncidentDetection_Log_FY2025.csv"

    # Read all and group by incident
    from collections import defaultdict
    header_lines = []
    trailer_lines = []
    header_end = False

    rows = []
    with open(path, encoding='utf-8') as f:
        for line in f:
            if line.startswith('#') and not header_end:
                header_lines.append(line)
            elif line.startswith('タイムスタンプ'):
                header_lines.append(line)
                header_end = True
            elif line.startswith('#') and header_end:
                trailer_lines.append(line)
            elif line.strip():
                parts = line.strip().split(',')
                if len(parts) >= 8:
                    rows.append(parts)

    # Group by incident
    inc = defaultdict(dict)
    for r in rows:
        ts, iid, ev = r[0], r[1], r[2]
        inc[iid][ev] = r

    order_events = ['DETECTED', 'NOTIFIED', 'ACKNOWLEDGED', 'INVESTIGATING', 'RESOLVED', 'CLOSED']

    new_rows = []
    for iid in inc:
        events = inc[iid]
        # Collect present timestamps in canonical order
        present = [(ev, events[ev]) for ev in order_events if ev in events]
        # Reassign timestamps in ascending order using existing timestamps as pool (sorted)
        ts_pool = sorted([datetime.strptime(events[ev][0], '%Y-%m-%d %H:%M:%S') for ev in events])
        # For each event in canonical order, assign next timestamp from pool
        for (ev, row), ts in zip(present, ts_pool):
            new_row = row.copy()
            new_row[0] = ts.strftime('%Y-%m-%d %H:%M:%S')
            new_rows.append(new_row)

    # Sort new rows by timestamp
    new_rows.sort(key=lambda r: r[0])

    with open(path, 'w', encoding='utf-8', newline='') as f:
        f.writelines(header_lines)
        for r in new_rows:
            f.write(','.join(r) + '\n')
        f.writelines(trailer_lines)

    print(f"[Fixed] Zabbix: {len(inc)} incidents reordered - INVESTIGATING always before RESOLVED")


# ==============================================================
# Fix 16: Update Evidence_Mapping_ITGC.csv
# ==============================================================
def fix_evidence_mapping(new_priv_file, old_priv_file, new_pop_file, old_q_file):
    """Update mapping CSV with renamed/added files"""
    path = Path(r"C:\Users\nyham\work\demo_data\2.RCM\Evidence_Mapping_ITGC.csv")

    lines = []
    with open(path, encoding='utf-8') as f:
        for line in f:
            s = line
            if old_priv_file and old_priv_file in s:
                s = s.replace(old_priv_file, new_priv_file)
            if old_q_file and old_q_file in s:
                s = s.replace(old_q_file, new_pop_file)
            lines.append(s)

    with open(path, 'w', encoding='utf-8', newline='') as f:
        f.writelines(lines)

    print(f"[Fixed] Evidence_Mapping_ITGC.csv: renamed entries ({old_priv_file} → {new_priv_file}, {old_q_file} → {new_pop_file})")


# ==============================================================
# Main
# ==============================================================
if __name__ == '__main__':
    # 1-3: ITGC-AC-001
    fix_workflow_csv()  # fix WF first (SU01 reads from it)
    fix_su01_csv()

    # 4-5: ITGC-AC-003
    fix_retired_users()

    # 6-7: ITGC-AC-004
    new_priv, old_priv = fix_privileged_log()

    # 8-9: ITGC-CM-001
    fix_change_register()

    # 10-11: ITGC-CM-002
    fix_uat_files()

    # 12-13: ITGC-CM-003
    new_pop, old_q = fix_stms()

    # 14: ITGC-OM-001
    fix_db13_dates()

    # 15: ITGC-OM-002
    fix_zabbix_timeline()

    # 16: Update evidence mapping
    fix_evidence_mapping(new_priv, old_priv, new_pop, old_q)

    print("\n=== ALL ITGC FIXES COMPLETED ===")
