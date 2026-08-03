# -*- coding: utf-8 -*-
"""連結決算ダミーデータ生成 パートB: 02_intercompany〜06_reports"""
import os
from openpyxl import Workbook
from consolidation_model import (
    COMPANIES, SUBS, ALL, FX, BS_COA, PL_COA, BS_NAME, PL_NAME, PL_SIGN,
    ASSET_CODES, LIAB_EQ_CODES,
    IC_TRANSACTIONS, IC_BALANCES, IC_LOAN, TRANSIT, UNREALIZED, DIVIDENDS,
    UNREALIZED_CLOSE, UNREALIZED_OPEN, URP_DTA_CLOSE, URP_DTA_OPEN,
    GOODWILL, NCI, net_income, consolidate,
)
from xlsx_style import (
    set_widths, title, header_row, data_row, kv_block, note,
    FILL_SECTION, FILL_TOTAL, FILL_WARN,
)

BASE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "6.consolidation")
R = consolidate()
T = R["translate"]
PERIOD = "FY2025（自 2025年4月1日 至 2026年3月31日）"
Q4DATE = "2026年3月31日"


# ===========================================================================
# 02_intercompany 内部取引照合表
# ===========================================================================

def gen_ic_matching():
    wb = Workbook()
    # ---- 債権債務照合 ----
    ws = wb.active
    ws.title = "債権債務照合表"
    set_widths(ws, [12, 24, 12, 24, 16, 16, 12, 44, 14])
    r = title(ws, 1, "内部取引照合表（債権・債務残高） FY2025 年度末",
              "基準日: 2026/3/31 / 作成: 経理部 連結G 西村 拓海 (2026-04-09) / レビュー: 高橋 美咲 (2026-04-10) / （単位：千円）")
    r = header_row(ws, r, ["債権側\n会社コード", "債権側計上額の科目・会社", "債務側\n会社コード", "債務側計上額の科目・会社",
                           "債権側計上額", "債務側計上額", "差異", "差異原因・解消方法", "ステータス"])
    for seller, buyer, close, _ in IC_BALANCES:
        ar = close
        ap = TRANSIT["buyer_booked_ap"] if (seller, buyer) == ("DA-HQ", "DATR") else close
        diff = ar - ap
        cause = ""
        status = "一致"
        if diff:
            cause = ("未達商品（2026/3/30出荷 SH-260330-018・2026/4/1着荷）。"
                     "連結上、未達整理仕訳（CNS-FY25-004）で商品・買掛金を計上のうえ相殺消去。")
            status = "差異解消済"
        memo_thb = ""
        if seller == "DAT" or buyer == "DAT":
            memo_thb = "JPY建て取引（DAT帳簿はTHB換算計上・円貨ベースで照合）"
            cause = (cause + " " + memo_thb).strip()
        r = data_row(ws, r, [seller, f"売掛金 / {COMPANIES[seller]['name']}", buyer,
                             f"買掛金 / {COMPANIES[buyer]['name']}", ar, ap, diff, cause, status],
                     numcols=(5, 6, 7), fill=FILL_WARN if diff else None)
    r = data_row(ws, r, ["DA-HQ", f"関係会社長期貸付金 / {COMPANIES['DA-HQ']['name']}", "DAT",
                         f"長期借入金 / {COMPANIES['DAT']['name']}", IC_LOAN["jpy"], IC_LOAN["jpy"], 0,
                         "JPY建て210,000千円（DAT帳簿 THB50,000 @CR4.20）", "一致"], numcols=(5, 6, 7))
    ar_sum = sum(x[2] for x in IC_BALANCES)
    ap_sum = sum(TRANSIT["buyer_booked_ap"] if (x[0], x[1]) == ("DA-HQ", "DATR") else x[2] for x in IC_BALANCES)
    r = data_row(ws, r, ["", "売掛金・買掛金 合計", "", "", ar_sum, ap_sum, ar_sum - ap_sum,
                         "差異6,200千円＝未達商品。未達整理後の消去額 324,500千円（S05自動仕訳と一致）", ""],
                 numcols=(5, 6, 7), bold=True, fill=FILL_TOTAL)
    r += 1
    r = note(ws, r, "※ 各社パッケージの「グループ債権債務明細」と相手先残高確認書（2026-04-03交換）に基づき照合。")
    r = note(ws, r, "※ 差異の許容基準：経理規程(R16) 100千円未満は僅少差異として当四半期中に売り手基準で調整。当期末の僅少差異はなし。")

    # ---- 取引高照合 ----
    ws2 = wb.create_sheet("取引高照合表")
    set_widths(ws2, [12, 24, 12, 24, 30, 16, 16, 12, 30])
    r = title(ws2, 1, "内部取引照合表（取引高） FY2025 年度累計",
              "対象期間: 2025/4/1-2026/3/31 / 作成: 経理部 連結G (2026-04-09) / （単位：千円）")
    r = header_row(ws2, r, ["売り手\nコード", "売り手会社", "買い手\nコード", "買い手会社", "取引内容",
                            "売り手計上額", "買い手計上額", "差異", "備考"])
    for seller, buyer, desc, amt, kind in IC_TRANSACTIONS:
        memo = ""
        if seller == "DAT" or buyer == "DAT":
            memo = "JPY建て取引（円貨ベースで照合）"
        r = data_row(ws2, r, [seller, COMPANIES[seller]["name"], buyer, COMPANIES[buyer]["name"],
                              f"{desc}（買い手: {kind}）", amt, amt, 0, memo], numcols=(6, 7, 8))
    tot = sum(x[3] for x in IC_TRANSACTIONS)
    r = data_row(ws2, r, ["", "", "", "", "合計", tot, tot, 0,
                          "四半期消去累計 (1,820,000+650,000)×4 と一致"], numcols=(6, 7, 8), bold=True, fill=FILL_TOTAL)
    r += 1
    r = data_row(ws2, r, ["", "", "", "", "【資金・資本取引】", "", "", "", ""], bold=True, fill=FILL_SECTION)
    r = data_row(ws2, r, ["DA-HQ", "デモA株式会社", "DAT", COMPANIES["DAT"]["name"], "受取利息／支払利息（長期貸付金）",
                          4920, 4920, 0, "DAT帳簿 THB1,200×AR4.10"], numcols=(6, 7, 8))
    r = data_row(ws2, r, ["DA-TB", COMPANIES["DA-TB"]["name"], "DA-HQ", "デモA株式会社", "支払配当金／受取配当金",
                          120000, 120000, 0, "2025/6/25"], numcols=(6, 7, 8))
    r = data_row(ws2, r, ["DATR", COMPANIES["DATR"]["name"], "DA-HQ", "デモA株式会社", "支払配当金／受取配当金（80%）",
                          40000, 40000, 0, "2025/6/20"], numcols=(6, 7, 8))

    # ---- 未達整理 ----
    ws3 = wb.create_sheet("未達取引整理表")
    set_widths(ws3, [18, 90])
    r = title(ws3, 1, "未達取引整理表 FY2025 年度末", "作成: 経理部 連結G 西村 拓海 / 承認: 高橋 美咲 (2026-04-10)")
    r = kv_block(ws3, r, [
        ("対象取引", "DA-HQ → DATR 製品（商品）供給"),
        ("出荷情報", "2026/3/30 出荷（出荷番号 SH-260330-018、送状 INV-26-03-1184）"),
        ("着荷・検収", "2026/4/1 DATR横浜倉庫にて検収（検収番号 GR-26-04-0007）"),
        ("金額", "6,200千円（売り手売価）"),
        ("両社の計上状況", "DA-HQ: 3月度売上・売掛金計上済 / DATR: 未検収のため未計上"),
        ("連結上の処理", "未達整理仕訳（CNS-FY25-004）で商品6,200・買掛金6,200を計上のうえ、債権債務324,500千円を相殺消去。"
                     "未達商品に含まれる未実現利益1,240千円（利益率20%）は未実現利益消去（CNS-FY25-007）に含めて消去。"),
        ("再発防止", "期末月は出荷基準日を3営業日前倒しする運用を継続（販売管理規程(R11)但書）。"),
    ])
    wb.save(os.path.join(BASE, "02_intercompany", "内部取引照合表_FY2025Q4.xlsx"))


# ===========================================================================
# 03_translation 在外子会社換算
# ===========================================================================

def gen_translation():
    c = COMPANIES["DAT"]
    wb = Workbook()
    ws = wb.active
    ws.title = "換算ワークシート"
    set_widths(ws, [10, 32, 16, 10, 10, 16, 16, 30])
    r = title(ws, 1, "在外子会社 財務諸表換算ワークシート（DAT） FY2025",
              "Demo-A (Thailand) Co., Ltd. / 作成: 経理部 連結G (2026-04-10) / レビュー: 高橋 美咲 / THB→JPY")
    r = header_row(ws, r, ["科目コード", "科目", "THB残高\n(千THB)", "レート\n区分", "適用\nレート", "換算後JPY\n(千円)", "前期末JPY\n(千円)", "備考"])

    def rate_for(code):
        if code == "E100":
            return "HR", FX["HR_capital"]
        if code in ("E120", "E140"):
            return "―", None
        return "CR", FX["CR_close"]

    r = data_row(ws, r, ["【貸借対照表】", "", "", "", "", "", "", ""], bold=True, fill=FILL_SECTION)
    for code, name, sec in BS_COA:
        thb = c["bs_close"].get(code)
        if thb is None and code not in ("E140",):
            continue
        if code == "E140":
            r = data_row(ws, r, [code, name + "（プラグ）", "―", "―", "―", T["bs_close"]["E140"], T["bs_open"]["E140"],
                                 "換算差額（貸借差額として算定）"], numcols=(6, 7), bold=True, fill=FILL_WARN)
            continue
        rk, rv = rate_for(code)
        memo = ""
        if code == "E120":
            memo = "期首換算後残高＋当期純利益(AR換算)の積上げ"
            r = data_row(ws, r, [code, name, thb, "積上", "―", T["bs_close"]["E120"], T["bs_open"]["E120"], memo], numcols=(3, 6, 7))
            continue
        if code == "L200":
            memo = "JPY建て借入210,000千円（換算差なし）"
        jpy_c = T["bs_close"].get(code, 0)
        jpy_o = T["bs_open"].get(code, 0)
        r = data_row(ws, r, [code, name, thb, rk, rv, jpy_c, jpy_o, memo], numcols=(3, 6, 7))
        if rv:
            ws.cell(row=r - 1, column=5).number_format = "0.00"
    a_thb = sum(c["bs_close"].get(k, 0) for k in ASSET_CODES)
    a_jpy = sum(T["bs_close"].get(k, 0) for k in ASSET_CODES)
    le_jpy = sum(T["bs_close"].get(k, 0) for k in LIAB_EQ_CODES)
    r = data_row(ws, r, ["", "資産合計 / 負債・純資産合計", a_thb, "", "", a_jpy, "", f"貸借一致確認: {a_jpy == le_jpy}"],
                 numcols=(3, 6), bold=True, fill=FILL_TOTAL)
    r += 1
    r = data_row(ws, r, ["【損益計算書】", "", "", "", "", "", "", ""], bold=True, fill=FILL_SECTION)
    for code, name, _ in PL_COA:
        thb = c["pl"].get(code, 0)
        if not thb:
            continue
        r = data_row(ws, r, [code, name, thb, "AR", FX["AR"], T["pl"][code], "", ""], numcols=(3, 6))
        ws.cell(row=r - 1, column=5).number_format = "0.00"
    r = data_row(ws, r, ["", "当期純利益", net_income(c["pl"]), "AR", FX["AR"], T["ni_jpy"], "", ""],
                 numcols=(3, 6), bold=True, fill=FILL_TOTAL)

    ws2 = wb.create_sheet("為替換算調整勘定の推移")
    set_widths(ws2, [52, 18, 40])
    r = title(ws2, 1, "為替換算調整勘定（CTA）の算定", "（単位：千円）")
    r = header_row(ws2, r, ["項目", "金額", "算定根拠"])
    rows = [
        ("期首 為替換算調整勘定", T["cta_open"], "前期連結精算表より繰越"),
        ("期末純資産の期末レート換算額", round((c["bs_close"]["E100"] + c["bs_close"]["E120"]) * FX["CR_close"]),
         f"THB{c['bs_close']['E100'] + c['bs_close']['E120']:,}千 × CR{FX['CR_close']:.2f}"),
        ("　うち資本金（取得時レート換算）", -round(c["bs_close"]["E100"] * FX["HR_capital"]),
         f"THB{c['bs_close']['E100']:,}千 × HR{FX['HR_capital']:.2f}（控除）"),
        ("　うち利益剰余金（歴史的換算積上額）", -T["re_hist_close"], "期首換算後残高＋当期純利益(AR換算)（控除）"),
        ("期末 為替換算調整勘定", T["cta_close"], "貸借差額"),
        ("当期増減（連結包括利益計算書 その他の包括利益）", T["cta_delta"], "円安進行（CR 4.00→4.20）による換算差益"),
    ]
    for k, v, m in rows:
        r = data_row(ws2, r, [k, v, m], numcols=(2,), bold=k.startswith("期末 為替") or k.startswith("当期増減"))
    r += 1
    r = note(ws2, r, "※ 換算方法：資産・負債は決算日レート(CR)、資本金は取得時レート(HR)、収益・費用は期中平均レート(AR)。差額は為替換算調整勘定として純資産の部に計上（外貨建取引等会計処理基準）。")

    ws3 = wb.create_sheet("適用レート")
    set_widths(ws3, [22, 14, 44])
    r = title(ws3, 1, "適用為替レート一覧", "00_master 為替レートマスタ_FY2025 と同一（グループ統一レート）")
    r = header_row(ws3, r, ["区分", "円/THB", "適用対象"])
    for a, b, cc in [("期末CR(2026/3/31)", FX["CR_close"], "BS項目"), ("前期末CR(2025/3/31)", FX["CR_open"], "前期末BS項目"),
                     ("期中平均AR", FX["AR"], "PL項目"), ("取得時HR", FX["HR_capital"], "資本金")]:
        r = data_row(ws3, r, [a, b, cc])
        ws3.cell(row=r - 1, column=2).number_format = "0.00"
    wb.save(os.path.join(BASE, "03_translation", "在外子会社換算ワークシート_DAT_FY2025.xlsx"))


# ===========================================================================
# 04_eliminations 連結仕訳帳・資本連結・未実現利益
# ===========================================================================

def _acct_name(code):
    if code == "PNCI":
        return "非支配株主に帰属する当期純利益"
    if code == "SS_DIV":
        return "剰余金の配当（連結S/S）"
    return BS_NAME.get(code) or PL_NAME.get(code) or code


def gen_journal():
    wb = Workbook()
    ws = wb.active
    ws.title = "連結仕訳帳"
    set_widths(ws, [14, 30, 8, 10, 30, 16, 10, 30, 16, 40, 10])
    r = title(ws, 1, "連結仕訳帳（FY2025 年度・累計ベース）",
              "連結決算システム(S05) / 起票: 経理部 連結G / レビュー: 高橋 美咲 (ACC002) 2026-04-08〜10 / 承認: 佐藤 一郎 (ACC001) 2026-04-14 (WF-CNS-2025-004) / （単位：千円）")
    r = header_row(ws, r, ["仕訳No", "区分", "行", "借/貸", "借方科目", "借方金額", "", "貸方科目", "貸方金額", "摘要", "起票区分"])
    for j in R["journal"]:
        first = True
        ln = 0
        dr_total = cr_total = 0
        for drcr, code, desc, amt in j["lines"]:
            ln += 1
            if drcr == "Dr":
                dr_total += amt
                vals = [j["no"] if first else "", j["category"] if first else "", ln, "借",
                        f"{code} {_acct_name(code)}", amt, "", "", "", desc, j["auto"] if first else ""]
            else:
                cr_total += amt
                vals = [j["no"] if first else "", j["category"] if first else "", ln, "貸",
                        "", "", "", f"{code} {_acct_name(code)}", amt, desc, j["auto"] if first else ""]
            r = data_row(ws, r, vals, numcols=(6, 9),
                         fill=FILL_WARN if j["auto"] == "MANUAL" and first else None)
            first = False
        assert dr_total == cr_total, f"{j['no']} 貸借不一致 {dr_total} vs {cr_total}"
        r = data_row(ws, r, ["", "", "", "", "（仕訳計）", dr_total, "", "", cr_total, "", ""],
                     numcols=(6, 9), bold=True, fill=FILL_TOTAL)
        if j.get("note"):
            r = note(ws, r, "※ " + j["note"])
        r += 1
    r = note(ws, r, "※ 仕訳件数 7件（うち手動起票 1件: CNS-FY25-007）。承認WF番号 WF-CNS-2025-004（S04ワークフローログ参照）。")
    r = note(ws, r, "※ S05の四半期別自動仕訳（ConsolidationSystem_Entries_FY2025.csv）は本仕訳帳の四半期スライスを要約表示したもの。")
    wb.save(os.path.join(BASE, "04_eliminations", "連結仕訳帳_FY2025.xlsx"))


def gen_capital_consol():
    wb = Workbook()
    ws = wb.active
    ws.title = "資本連結ワークシート"
    set_widths(ws, [34, 16, 16, 16, 16, 16, 40])
    r = title(ws, 1, "資本連結ワークシート（投資と資本の相殺消去） FY2025",
              "作成: 経理部 連結G / レビュー: 高橋 美咲 (2026-04-10) / （単位：千円）")
    r = header_row(ws, r, ["項目", "DA-TB", "DA-LOG", "DAT", "DATR", "合計", "備考"])
    rows = [
        ("持分比率", "100%", "100%", "100%", "80%", "", ""),
        ("取得・設立日", "2005/10/1", "2010/4/1", "2012/7/1", "2019/4/1", "", "DATRのみ株式取得（他は設立時出資）"),
        ("親会社投資簿価（関係会社株式）", 100000, 30000, 90000, 560000, 780000, "DAT: THB30,000千×HR3.00"),
        ("支配獲得時 資本金", 90000, 30000, 90000, 90000, 300000, ""),
        ("支配獲得時 資本剰余金", 10000, 0, 0, 10000, 20000, ""),
        ("支配獲得時 利益剰余金", 0, 0, 0, 400000, 400000, "設立出資の3社は0"),
        ("支配獲得時 純資産合計", 100000, 30000, 90000, 500000, 720000, ""),
        ("親会社持分相当額", 100000, 30000, 90000, 400000, 620000, ""),
        ("非支配株主持分（取得時）", 0, 0, 0, 100000, 100000, "20%"),
        ("のれん発生額", 0, 0, 0, 160000, 160000, "560,000－400,000"),
    ]
    for row in rows:
        numc = tuple(i for i in (2, 3, 4, 5, 6) if isinstance(row[i - 1], (int, float)))
        r = data_row(ws, r, list(row), numcols=numc)
    r += 1
    r = header_row(ws, r, ["のれんの推移", "金額", "", "", "", "", "備考"])
    g = GOODWILL
    for k, v, m in [("取得原価（2019/4/1）", g["cost"], "10年定額償却"),
                    ("過年度償却累計（FY2019-FY2024）", g["accum_open"], "6年経過"),
                    ("期首帳簿価額", g["nbv_open"], ""),
                    ("当期償却額（CNS-FY25-002）", g["annual_amort"], "販管費「のれん償却額」"),
                    ("期末帳簿価額", g["nbv_close"], "残存償却期間3年"),
                    ("減損の兆候", "なし", "DATRの収益性は取得時計画を上回って推移")]:
        r = data_row(ws, r, [k, v, "", "", "", "", m], numcols=(2,) if isinstance(v, int) else ())

    ws2 = wb.create_sheet("非支配株主持分の推移")
    set_widths(ws2, [44, 18, 44])
    r = title(ws2, 1, "非支配株主持分（DATR 20%）の推移", "非支配株主: 商栄興産株式会社 / （単位：千円）")
    r = header_row(ws2, r, ["項目", "金額", "算定根拠"])
    for k, v, m in [("期首残高", NCI["open"], "DATR期首純資産589,600×20%"),
                    ("非支配株主に帰属する当期純利益", NCI["pl"], "DATR当期純利益170,400×20%（四半期8,520×4）"),
                    ("剰余金の配当（非支配株主分）", -NCI["dividend"], "配当50,000×20%（2025/6/20）"),
                    ("期末残高", NCI["close"], "DATR期末純資産710,000×20% と一致")]:
        r = data_row(ws2, r, [k, v, m], numcols=(2,), bold=k.startswith("期末"))
    wb.save(os.path.join(BASE, "04_eliminations", "資本連結・のれん・非支配株主持分_FY2025.xlsx"))


def gen_urp():
    wb = Workbook()
    ws = wb.active
    ws.title = "未実現利益計算"
    set_widths(ws, [14, 14, 30, 16, 14, 16, 16, 40])
    r = title(ws, 1, "棚卸資産未実現利益 計算シート（S07 EUC） FY2025",
              "作成: 経理部 連結G 西村 拓海 (2026-04-09) / レビュー: 高橋 美咲 (2026-04-10) / 承認: 佐藤 一郎 (2026-04-14) / （単位：千円）"
              )
    r = header_row(ws, r, ["売り手", "保有会社", "対象在庫", "期末在庫額", "売り手\n利益率", "未実現利益\n（期末）", "未実現利益\n（期首）", "備考"])
    for (seller, holder, inv_c, margin, urp_c) in UNREALIZED["close"]:
        inv_o, urp_o = next((x[2], x[4]) for x in UNREALIZED["open"] if x[0] == seller and x[1] == holder)
        memo = ""
        if holder == "DA-HQ":
            memo = "DA-TB申告の売上総利益率11.0%（FY2025実績）を適用"
        if holder == "DATR":
            memo = "手許在庫53,800＋未達商品6,200。親会社の対DATR売上総利益率20.0%を適用"
        r = data_row(ws, r, [seller, holder, ("材料・仕掛品（部品）" if holder == "DA-HQ" else "商品"),
                             inv_c, f"{margin*100:.1f}%", urp_c, urp_o, memo], numcols=(4, 6, 7))
    r = data_row(ws, r, ["", "", "合計", 340000, "", UNREALIZED_CLOSE, UNREALIZED_OPEN,
                         "期末42,800（S05手動仕訳CNS-FY25-007）／期首38,600（開始仕訳で実現）"],
                 numcols=(4, 6, 7), bold=True, fill=FILL_TOTAL)
    r += 1
    r = header_row(ws, r, ["税効果", "", "", "", "", "期末", "期首", ""])
    r = data_row(ws, r, ["繰延税金資産（未実現利益×法定実効税率30.6%）", "", "", "", "", URP_DTA_CLOSE, URP_DTA_OPEN,
                         "連結固有の一時差異。売却元の税率を使用"], numcols=(6, 7))
    r += 1
    r = note(ws, r, "※ DA-HQ→DA-TB の有償支給材（期末在庫55,000千円）は原価売買（利益率0%）のため未実現利益は発生しない。")
    r = note(ws, r, "※ DAT保有在庫にグループ仕入品はない（原材料の親会社仕入分は当期中に全量費消）。")
    r = note(ws, r, "※ 本シートはEUC管理規程(R27)の対象。計算式保護・変更履歴管理を実施。")
    wb.save(os.path.join(BASE, "04_eliminations", "未実現利益計算シート_FY2025.xlsx"))


# ===========================================================================
# 05_worksheet 連結精算表
# ===========================================================================

def gen_worksheet():
    wb = Workbook()
    cols = R["cols"]
    ws = wb.active
    ws.title = "連結精算表BS"
    set_widths(ws, [10, 30, 14, 13, 13, 13, 13, 15, 14, 14, 15])
    r = title(ws, 1, "連結精算表（貸借対照表） FY2025 年度末",
              "作成: 経理部 連結G (2026-04-15) / レビュー: 高橋 美咲 / 承認: 佐藤 一郎 / DATは円換算後 / （単位：千円）")
    r = header_row(ws, r, ["科目コード", "科目", "DA-HQ\n親会社", "DA-TB", "DA-LOG", "DAT\n(換算後)", "DATR",
                           "単純合算", "消去修正\n借方", "消去修正\n貸方", "連結"])
    # 消去のDr/Cr集計（BS科目）
    dr = {}
    cr = {}
    for j in R["journal"]:
        for drcr, code, desc, amt in j["lines"]:
            if code in ("PNCI", "SS_DIV") or code.startswith("P"):
                continue
            if drcr == "Dr":
                dr[code] = dr.get(code, 0) + amt
            else:
                cr[code] = cr.get(code, 0) + amt
    # E120: PL消去純額・配当戻し・非支配株主振替を貸借に反映
    pl_net = R["cons_bs"]["E120"] - (R["simple_bs"]["E120"] + cr.get("E120", 0) - dr.get("E120", 0))
    if pl_net >= 0:
        cr["E120"] = cr.get("E120", 0) + pl_net
    else:
        dr["E120"] = dr.get("E120", 0) - pl_net

    sec_label = {"CA": "【流動資産】", "FA": "【固定資産】", "CL": "【流動負債】", "NCL": "【固定負債】", "EQ": "【純資産】"}
    cur = None
    for code, name, sec in BS_COA:
        vals = [cols[co]["bs"].get(code, 0) for co in ALL]
        simple = R["simple_bs"].get(code, 0)
        d, c_ = dr.get(code, 0), cr.get(code, 0)
        cons = R["cons_bs"].get(code, 0)
        if not any(vals) and not d and not c_ and not cons:
            continue
        if sec != cur:
            r = data_row(ws, r, [sec_label[sec]] + [""] * 10, bold=True, fill=FILL_SECTION)
            cur = sec
        memo_fill = FILL_WARN if code in ("A215", "E140", "E150") else None
        r = data_row(ws, r, [code, name] + vals + [simple, d or "", c_ or "", cons],
                     numcols=(3, 4, 5, 6, 7, 8, 9, 10, 11), fill=memo_fill)
    a_cols = [sum(cols[co]["bs"].get(k, 0) for k in ASSET_CODES) for co in ALL]
    a_simple = sum(R["simple_bs"].get(k, 0) for k in ASSET_CODES)
    a_cons = sum(R["cons_bs"].get(k, 0) for k in ASSET_CODES)
    le_cons = sum(R["cons_bs"].get(k, 0) for k in LIAB_EQ_CODES)
    dr_tot = sum(dr.values())
    cr_tot = sum(cr.values())
    r = data_row(ws, r, ["", "資産合計"] + a_cols + [a_simple, "", "", a_cons], numcols=tuple(range(3, 12)), bold=True, fill=FILL_TOTAL)
    le_cols = [sum(cols[co]["bs"].get(k, 0) for k in LIAB_EQ_CODES) for co in ALL]
    r = data_row(ws, r, ["", "負債・純資産合計"] + le_cols + [sum(R["simple_bs"].get(k, 0) for k in LIAB_EQ_CODES), "", "", le_cons],
                 numcols=tuple(range(3, 12)), bold=True, fill=FILL_TOTAL)
    assert a_cons == le_cons
    r += 1
    r = note(ws, r, "※ 消去修正列は連結仕訳帳（04_eliminations）7仕訳のBS科目部分。利益剰余金にはPL消去の純利益影響・配当戻し・非支配株主振替を含めて表示。")
    r = note(ws, r, f"※ 貸借一致確認：連結資産合計 {a_cons:,} ＝ 連結負債・純資産合計 {le_cons:,}")

    ws2 = wb.create_sheet("連結精算表PL")
    set_widths(ws2, [10, 32, 14, 13, 13, 13, 13, 15, 14, 14, 15])
    r = title(ws2, 1, "連結精算表（損益計算書） FY2025",
              "DATは期中平均レート4.10で円換算 / （単位：千円）")
    r = header_row(ws2, r, ["科目コード", "科目", "DA-HQ", "DA-TB", "DA-LOG", "DAT\n(換算後)", "DATR",
                            "単純合算", "消去修正\n借方", "消去修正\n貸方", "連結"])
    pdr = {}
    pcr = {}
    for j in R["journal"]:
        for drcr, code, desc, amt in j["lines"]:
            if not (code.startswith("P") and code != "PNCI"):
                continue
            if drcr == "Dr":
                pdr[code] = pdr.get(code, 0) + amt
            else:
                pcr[code] = pcr.get(code, 0) + amt
    for code, name, kind in PL_COA:
        vals = [cols[co]["pl"].get(code, 0) for co in ALL]
        simple = R["simple_pl"].get(code, 0)
        d, c_ = pdr.get(code, 0), pcr.get(code, 0)
        cons = R["cons_pl"].get(code, 0)
        if not any(vals) and not d and not c_:
            continue
        r = data_row(ws2, r, [code, name] + vals + [simple, d or "", c_ or "", cons], numcols=tuple(range(3, 12)))
    ni_cols = [net_income(cols[co]["pl"]) for co in ALL]
    r = data_row(ws2, r, ["", "当期純利益"] + ni_cols + [net_income(R["simple_pl"]), "", "", R["cons_ni"]],
                 numcols=tuple(range(3, 12)), bold=True, fill=FILL_TOTAL)
    r = data_row(ws2, r, ["", "非支配株主に帰属する当期純利益", "", "", "", "", "", "", R["pnci"], "", R["pnci"]],
                 numcols=(9, 11))
    r = data_row(ws2, r, ["", "親会社株主に帰属する当期純利益", "", "", "", "", "", "", "", "", R["ni_parent"]],
                 numcols=(11,), bold=True, fill=FILL_TOTAL)
    wb.save(os.path.join(BASE, "05_worksheet", "連結精算表_FY2025.xlsx"))


# ===========================================================================
# 06_reports 連結財務諸表・回収状況管理表
# ===========================================================================

def _sum(bs, codes):
    return sum(bs.get(k, 0) for k in codes)


def gen_consol_fs():
    wb = Workbook()
    cb, ob = R["cons_bs"], R["cons_bs_open"]
    cp = R["cons_pl"]

    ws = wb.active
    ws.title = "連結貸借対照表"
    set_widths(ws, [40, 18, 18, 16])
    r = title(ws, 1, "連結貸借対照表", f"デモA株式会社 / {Q4DATE}現在 / （単位：千円）")
    r = header_row(ws, r, ["科目", "前連結会計年度末\n(2025/3/31)", "当連結会計年度末\n(2026/3/31)", "増減"])
    ca = [c for c, _, s in BS_COA if s == "CA"]
    fa = [c for c, _, s in BS_COA if s == "FA"]
    cl = [c for c, _, s in BS_COA if s == "CL"]
    ncl = [c for c, _, s in BS_COA if s == "NCL"]
    eq = [c for c, _, s in BS_COA if s == "EQ"]

    def block(label, codes, total_label):
        nonlocal r
        r = data_row(ws, r, [label, "", "", ""], bold=True, fill=FILL_SECTION)
        for code in codes:
            vo, vc = ob.get(code, 0), cb.get(code, 0)
            if vo == 0 and vc == 0:
                continue
            r = data_row(ws, r, ["　" + BS_NAME[code], vo, vc, vc - vo], numcols=(2, 3, 4))
        r = data_row(ws, r, [total_label, _sum(ob, codes), _sum(cb, codes), _sum(cb, codes) - _sum(ob, codes)],
                     numcols=(2, 3, 4), bold=True, fill=FILL_TOTAL)

    block("【資産の部：流動資産】", ca, "流動資産合計")
    block("【資産の部：固定資産】", fa, "固定資産合計")
    r = data_row(ws, r, ["資産合計", _sum(ob, ASSET_CODES), _sum(cb, ASSET_CODES),
                         _sum(cb, ASSET_CODES) - _sum(ob, ASSET_CODES)], numcols=(2, 3, 4), bold=True, fill=FILL_TOTAL)
    block("【負債の部：流動負債】", cl, "流動負債合計")
    block("【負債の部：固定負債】", ncl, "固定負債合計")
    r = data_row(ws, r, ["負債合計", _sum(ob, cl + ncl), _sum(cb, cl + ncl),
                         _sum(cb, cl + ncl) - _sum(ob, cl + ncl)], numcols=(2, 3, 4), bold=True, fill=FILL_TOTAL)
    block("【純資産の部】", eq, "純資産合計")
    r = data_row(ws, r, ["負債純資産合計", _sum(ob, LIAB_EQ_CODES), _sum(cb, LIAB_EQ_CODES),
                         _sum(cb, LIAB_EQ_CODES) - _sum(ob, LIAB_EQ_CODES)], numcols=(2, 3, 4), bold=True, fill=FILL_TOTAL)

    # ---- PL ----
    ws2 = wb.create_sheet("連結損益計算書")
    set_widths(ws2, [40, 18, 20])
    r = title(ws2, 1, "連結損益計算書", f"デモA株式会社 / {PERIOD} / （単位：千円）")
    r = header_row(ws2, r, ["科目", "当連結会計年度", "備考"])
    gross = cp["P100"] - cp["P200"]
    op = gross - cp["P300"]
    ord_inc = op + cp["P400a"] + cp["P400b"] + cp["P400c"] + cp["P400d"] - cp["P410a"] - cp["P410b"]
    pt = ord_inc + cp["P500"] - cp["P510"]
    rows = [
        ("売上高", cp["P100"], ""),
        ("売上原価", cp["P200"], "未実現利益消去後"),
        ("売上総利益", gross, ""),
        ("販売費及び一般管理費", cp["P300"], "のれん償却額16,000を含む"),
        ("営業利益", op, ""),
        ("営業外収益（受取利息・為替差益等）", cp["P400a"] + cp["P400b"] + cp["P400c"] + cp["P400d"], ""),
        ("営業外費用（支払利息等）", cp["P410a"] + cp["P410b"], ""),
        ("経常利益", ord_inc, ""),
        ("特別利益", cp["P500"], "投資有価証券売却益"),
        ("特別損失", cp["P510"], "固定資産除却損"),
        ("税金等調整前当期純利益", pt, ""),
        ("法人税等（法人税等調整額含む）", cp["P600"], ""),
        ("当期純利益", R["cons_ni"], ""),
        ("非支配株主に帰属する当期純利益", R["pnci"], "DATR 20%"),
        ("親会社株主に帰属する当期純利益", R["ni_parent"], ""),
    ]
    for k, v, m in rows:
        bold = k in ("売上総利益", "営業利益", "経常利益", "税金等調整前当期純利益", "当期純利益", "親会社株主に帰属する当期純利益")
        r = data_row(ws2, r, [k, v, m], numcols=(2,), bold=bold, fill=FILL_TOTAL if bold else None)

    # ---- 包括利益 ----
    ws3 = wb.create_sheet("連結包括利益計算書")
    set_widths(ws3, [44, 18, 30])
    r = title(ws3, 1, "連結包括利益計算書", f"デモA株式会社 / {PERIOD} / （単位：千円）")
    r = header_row(ws3, r, ["科目", "当連結会計年度", "備考"])
    ci = R["cons_ni"] + T["cta_delta"]
    for k, v, m in [
        ("当期純利益", R["cons_ni"], ""),
        ("その他の包括利益：為替換算調整勘定", T["cta_delta"], "円安進行（CR4.00→4.20）"),
        ("包括利益", ci, ""),
        ("（内訳）親会社株主に係る包括利益", ci - R["pnci"], ""),
        ("（内訳）非支配株主に係る包括利益", R["pnci"], "DATはNCIなしのためCTAは全額親会社株主分"),
    ]:
        r = data_row(ws3, r, [k, v, m], numcols=(2,), bold=(k == "包括利益"), fill=FILL_TOTAL if k == "包括利益" else None)

    # ---- 株主資本等変動計算書 ----
    ws4 = wb.create_sheet("連結株主資本等変動計算書")
    set_widths(ws4, [34, 13, 13, 15, 12, 14, 14, 15])
    r = title(ws4, 1, "連結株主資本等変動計算書", f"デモA株式会社 / {PERIOD} / （単位：千円）")
    r = header_row(ws4, r, ["", "資本金", "資本剰余金", "利益剰余金", "自己株式", "為替換算\n調整勘定", "非支配株主\n持分", "純資産合計"])
    o = [ob["E100"], ob["E110"], ob["E120"], ob["E130"], ob["E140"], ob["E150"]]
    div = 500000
    ni = R["ni_parent"]
    cta_d = T["cta_delta"]
    nci_d = NCI["pl"] - NCI["dividend"]
    c_ = [cb["E100"], cb["E110"], cb["E120"], cb["E130"], cb["E140"], cb["E150"]]
    rows = [
        ("当期首残高", o[0], o[1], o[2], o[3], o[4], o[5], sum(o)),
        ("剰余金の配当", 0, 0, -div, 0, 0, 0, -div),
        ("親会社株主に帰属する当期純利益", 0, 0, ni, 0, 0, 0, ni),
        ("株主資本以外の項目の当期変動額（純額）", 0, 0, 0, 0, cta_d, nci_d, cta_d + nci_d),
        ("当期末残高", c_[0], c_[1], c_[2], c_[3], c_[4], c_[5], sum(c_)),
    ]
    for i, row in enumerate(rows):
        r = data_row(ws4, r, list(row), numcols=(2, 3, 4, 5, 6, 7, 8), bold=(i in (0, 4)), fill=FILL_TOTAL if i == 4 else None)
    assert o[2] - div + ni == c_[2], "連結利益剰余金ロール不一致"
    assert o[4] + cta_d == c_[4] and o[5] + nci_d == c_[5]
    r += 1
    r = note(ws4, r, "※ 非支配株主持分の変動：非支配株主に帰属する当期純利益 +34,080 / 非支配株主への配当 △10,000")

    # ---- CF ----
    ws5 = wb.create_sheet("連結キャッシュ・フロー計算書")
    set_widths(ws5, [52, 18, 36])
    r = title(ws5, 1, "連結キャッシュ・フロー計算書（間接法）", f"デモA株式会社 / {PERIOD} / （単位：千円）")
    r = header_row(ws5, r, ["科目", "金額", "備考"])

    dep = 895000 + 80000 + 173000 + 12000 + 65000 + 4000 + round(35000 * FX["AR"]) + 14000
    tax_paid = -(cp["P600"] + 1285 + _sum(ob, ["L140"]) - _sum(cb, ["L140"]))  # 法調除きの支払額
    int_div_recv = cp["P400a"] + cp["P400b"]
    int_paid = -cp["P410a"]
    delta = lambda codes: _sum(cb, codes) - _sum(ob, codes)
    wc_ar = -delta(["A110", "A115"])
    wc_inv = -delta(["A120"])
    wc_oth_a = -delta(["A130", "A140"])
    wc_ap = delta(["L100"])
    wc_oth_l = delta(["L130", "L220"])
    prov = delta(["L150", "L210"])
    pt_row = pt
    gw = GOODWILL["annual_amort"]
    disposal_loss = 55000 + 12000
    gain_sec = -15000
    fx_gain = -cp["P400c"]
    op_items = [
        ("税金等調整前当期純利益", pt_row, ""),
        ("減価償却費", dep, "有形・無形（製造原価分含む）"),
        ("のれん償却額", gw, ""),
        ("引当金の増減額（△は減少）", prov, "賞与・退職給付"),
        ("受取利息及び受取配当金", -int_div_recv, ""),
        ("支払利息", cp["P410a"], ""),
        ("為替差損益（△は益）", fx_gain, ""),
        ("固定資産除却損", disposal_loss, ""),
        ("投資有価証券売却損益（△は益）", gain_sec, ""),
        ("売上債権の増減額（△は増加）", wc_ar, ""),
        ("棚卸資産の増減額（△は増加）", wc_inv, ""),
        ("仕入債務の増減額（△は減少）", wc_ap, ""),
        ("その他の資産・負債の増減額", wc_oth_a + wc_oth_l, ""),
    ]
    subtotal = sum(v for _, v, _ in op_items)
    op_items2 = [
        ("小計", subtotal, ""),
        ("利息及び配当金の受取額", int_div_recv, ""),
        ("利息の支払額", int_paid, ""),
        ("法人税等の支払額", tax_paid, ""),
    ]
    op_cf = subtotal + int_div_recv + int_paid + tax_paid
    inv_items = [
        ("有形固定資産の取得による支出", -(800000 + 130000 + 45000 + round(28000 * FX["AR"]) + 8000), ""),
        ("無形固定資産の取得による支出", -(60000 + 6000 + 2000), ""),
        ("投資有価証券の取得による支出", -65000, ""),
        ("投資有価証券の売却による収入", 50000, ""),
        ("その他", -delta(["A260", "A250"]) - (-(URP_DTA_CLOSE - URP_DTA_OPEN)) - 5000, "敷金・保証金の差入れ等"),
    ]
    inv_cf = sum(v for _, v, _ in inv_items)
    fin_items = [
        ("短期借入金の純増減額（△は減少）", delta(["L110"]), ""),
        ("長期借入金の返済による支出", delta(["L200"]), "約定弁済"),
        ("リース債務の返済による支出", delta(["L120"]), ""),
        ("自己株式の取得による支出", 0, ""),
        ("配当金の支払額", -500000, ""),
        ("非支配株主への配当金の支払額", -10000, ""),
    ]
    fin_cf = sum(v for _, v, _ in fin_items)
    cash_open = _sum(ob, ["A100"])
    cash_close = _sum(cb, ["A100"])
    # 現金の換算差額（在外子会社保有現金）: 期首残高×レート変動 + 期中増加×(CR-AR)
    fx_effect = round(25700 * (FX["CR_close"] - FX["CR_open"]) + (69500 - 25700) * (FX["CR_close"] - FX["AR"]))
    # 在外子会社の営業資産負債等に含まれる換算差額（非資金項目）を営業CFで調整
    op_fx_adj = cash_close - cash_open - op_cf - inv_cf - fin_cf - fx_effect
    op_items.append(("在外子会社の資産・負債に係る換算調整額（非資金項目）", op_fx_adj,
                     "CTA変動のうち営業資産・負債見合い分"))
    subtotal += op_fx_adj
    op_items2[0] = ("小計", subtotal, "")
    op_cf += op_fx_adj
    for k, v, m in op_items + op_items2:
        bold = k == "小計"
        r = data_row(ws5, r, [k, v, m], numcols=(2,), bold=bold, fill=FILL_TOTAL if bold else None)
    r = data_row(ws5, r, ["営業活動によるキャッシュ・フロー", op_cf, ""], numcols=(2,), bold=True, fill=FILL_TOTAL)
    for k, v, m in inv_items:
        r = data_row(ws5, r, [k, v, m], numcols=(2,))
    r = data_row(ws5, r, ["投資活動によるキャッシュ・フロー", inv_cf, ""], numcols=(2,), bold=True, fill=FILL_TOTAL)
    for k, v, m in fin_items:
        r = data_row(ws5, r, [k, v, m], numcols=(2,))
    r = data_row(ws5, r, ["財務活動によるキャッシュ・フロー", fin_cf, ""], numcols=(2,), bold=True, fill=FILL_TOTAL)
    r = data_row(ws5, r, ["現金及び現金同等物に係る換算差額", fx_effect, "在外子会社保有現金の換算影響等"], numcols=(2,))
    r = data_row(ws5, r, ["現金及び現金同等物の増減額", cash_close - cash_open, ""], numcols=(2,), bold=True)
    r = data_row(ws5, r, ["現金及び現金同等物の期首残高", cash_open, ""], numcols=(2,))
    r = data_row(ws5, r, ["現金及び現金同等物の期末残高", cash_close, ""], numcols=(2,), bold=True, fill=FILL_TOTAL)
    assert cash_open + op_cf + inv_cf + fin_cf + fx_effect == cash_close
    print(f"  CF check: 営業{op_cf:,} 投資{inv_cf:,} 財務{fin_cf:,} 換算差額{fx_effect:,}")

    # ---- 主要指標 ----
    ws6 = wb.create_sheet("連結ハイライト")
    set_widths(ws6, [40, 20, 40])
    r = title(ws6, 1, "連結決算ハイライト FY2025", "決算短信サマリー整合用 / （単位：千円）")
    ta = _sum(cb, ASSET_CODES)
    eq_owner = _sum(cb, ["E100", "E110", "E120", "E130", "E140"])
    for k, v, m in [
        ("売上高", cp["P100"], "前期 約38,000,000（横ばい）"),
        ("営業利益", op, "営業利益率 " + f"{op/cp['P100']*100:.1f}%"),
        ("経常利益", ord_inc, ""),
        ("親会社株主に帰属する当期純利益", R["ni_parent"], ""),
        ("総資産", ta, ""),
        ("純資産", _sum(cb, eq), ""),
        ("自己資本比率", f"{eq_owner/ta*100:.1f}%", "自己資本＝純資産－非支配株主持分"),
        ("連結従業員数", "850名", "単体580名・子会社270名（就業人員）"),
    ]:
        r = data_row(ws6, r, [k, v, m], numcols=(2,) if isinstance(v, int) else ())
    wb.save(os.path.join(BASE, "06_reports", "連結財務諸表_FY2025.xlsx"))


def gen_collection_status():
    wb = Workbook()
    ws = wb.active
    ws.title = "回収状況管理表"
    set_widths(ws, [8, 12, 26, 18, 18, 10, 12, 22, 44])
    r = title(ws, 1, "連結パッケージ回収状況管理表 FY2025（Q1〜Q4）",
              "管理: 経理部 連結G / S05アップロードログと連動 / 期限: 四半期末+6営業日 17:30")
    r = header_row(ws, r, ["四半期", "会社コード", "会社名", "提出期限", "初回アップロード", "エラー\n件数", "最終版", "結果", "エラー内容・対応（S05往復ログ参照）"])
    data = [
        ("Q1", "DA-TB", "2025-07-08 17:30", "2025-07-08 16:00", 0, "v1.0", "PASS", ""),
        ("Q1", "DA-LOG", "2025-07-08 17:30", "2025-07-08 10:00", 1, "v2.0", "PASS_AFTER_CORRECTION", "内部取引売掛金バランス差異125千円（IC-2025-Q1-018消込漏れ）→同日再送"),
        ("Q1", "DAT", "2025-07-08 17:30", "2025-07-08 17:00", 0, "v1.0", "PASS", ""),
        ("Q1", "DATR", "2025-07-08 17:30", "2025-07-08 13:00", 1, "v2.0", "PASS_AFTER_CORRECTION", "対DAT取引の換算レート不一致→グループ統一レート4.05に修正・再送"),
        ("Q2", "DA-TB", "2025-10-08 17:30", "2025-10-08 09:00", 2, "v2.0", "PASS_AFTER_CORRECTION", "科目マスタ不整合（新規科目2211）＋内部利益計算方式ズレ→修正・再送"),
        ("Q2", "DA-LOG", "2025-10-08 17:30", "2025-10-08 09:00", 1, "v2.0", "PASS_AFTER_CORRECTION", "期首残高引継エラー（S05繰越バッチ未実行）→情シス対応後再送"),
        ("Q2", "DAT", "2025-10-08 17:30", "2025-10-08 11:00", 0, "v1.0", "PASS", ""),
        ("Q2", "DATR", "2025-10-08 17:30", "2025-10-08 09:00", 0, "v1.0", "PASS", ""),
        ("Q3", "DA-TB", "2026-01-08 17:30", "2026-01-08 15:00", 1, "v2.0", "PASS_AFTER_CORRECTION", "固定資産除却仕訳の科目誤り（営業外→特別損失に訂正）"),
        ("Q3", "DA-LOG", "2026-01-08 17:30", "2026-01-08 12:00", 0, "v1.0", "PASS", ""),
        ("Q3", "DAT", "2026-01-08 17:30", "2026-01-08 17:00", 0, "v1.0", "PASS", ""),
        ("Q3", "DATR", "2026-01-08 17:30", "2026-01-08 10:00", 0, "v1.0", "PASS", ""),
        ("Q4", "DA-TB", "2026-04-08 17:30", "2026-04-08 14:00", 0, "v1.0", "PASS", ""),
        ("Q4", "DA-LOG", "2026-04-08 17:30", "2026-04-08 17:00", 0, "v1.0", "PASS", ""),
        ("Q4", "DAT", "2026-04-08 17:30", "2026-04-08 17:00", 0, "v1.0", "PASS", ""),
        ("Q4", "DATR", "2026-04-08 17:30", "2026-04-08 11:00", 0, "v1.0", "PASS", ""),
    ]
    for q, code, due, up, err, ver, res, memo in data:
        fill = FILL_WARN if err else None
        r = data_row(ws, r, [q, code, COMPANIES[code]["name"], due, up, err, ver, res, memo], fill=fill)
    r += 1
    r = note(ws, r, "※ 全四半期・全社が期限内提出。FY2025のバリデーションエラーは計5件・全件当日中に解消（FCRP-002統制の運用証跡）。")
    r = note(ws, r, "※ Q4（年度決算）は4社とも初回PASS。年度パッケージは01_packagesに保管。")
    wb.save(os.path.join(BASE, "06_reports", "連結パッケージ回収状況管理表_FY2025.xlsx"))


def main():
    for sub in ("02_intercompany", "03_translation", "04_eliminations", "05_worksheet", "06_reports"):
        os.makedirs(os.path.join(BASE, sub), exist_ok=True)
    gen_ic_matching()
    gen_translation()
    gen_journal()
    gen_capital_consol()
    gen_urp()
    gen_worksheet()
    gen_consol_fs()
    gen_collection_status()
    print("パートB生成完了")


if __name__ == "__main__":
    main()
