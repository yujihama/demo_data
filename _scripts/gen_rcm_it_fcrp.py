"""
ITGC, ITAC, FCRP, RCM Summary の生成
"""
import openpyxl
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent))
from rcm_common import (write_rcm_row, add_legend_sheet, STANDARD_COLUMNS,
                         HEADER_FILL, HEADER_FONT, CENTER_WRAP, BORDER_HEADER,
                         Font, Alignment, PatternFill, BODY_FONT, BORDER, LEFT_WRAP,
                         FILL_KEY, FILL_DEFICIENCY, FILL_HOLD, FILL_OK, FILL_SUB_HEADER)
from openpyxl.utils import get_column_letter

BASE = Path(r"C:\Users\nyham\work\demo_data\2.RCM")
EVAL_INFO = "2026/2/20\n長谷川 剛"


# ==================== ITGC ====================
def gen_itgc():
    # ITGC用の列構成
    ITGC_COLUMNS = [
        ("統制ID", 12),
        ("IT統制領域", 16),
        ("サブプロセス", 16),
        ("リスク記述", 42),
        ("統制活動", 48),
        ("統制タイプ", 10),
        ("手作業/IT自動", 10),
        ("頻度", 10),
        ("キー\nコントロール", 10),
        ("実施者", 18),
        ("実施証跡", 30),
        ("関連規程", 14),
        ("対象システム", 14),
        ("整備状況\n評価結果", 14),
        ("運用状況\n評価結果", 14),
        ("不備の\n有無", 10),
        ("最終結論", 18),
        ("評価日/評価者", 18),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ITGC_RCM"
    ws.sheet_view.zoomScale = 85

    # タイトル
    ws.cell(row=1, column=1, value="【ITGC】IT全般統制 リスク・コントロール・マトリクス（RCM）")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ITGC_COLUMNS))
    ws.cell(row=2, column=1, value="評価対象システム: SAP S/4HANA(S01), WMS(S02)  /  評価期間: FY2025  /  作成日: 2026/04/12  /  作成者: 内部監査室")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(ITGC_COLUMNS))

    # ヘッダ
    for i, (name, width) in enumerate(ITGC_COLUMNS, 1):
        cell = ws.cell(row=4, column=i, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        cell.border = BORDER_HEADER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[4].height = 36
    ws.freeze_panes = "B5"

    def write_itgc(row_num, values, key=False, status="ok"):
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            if c in (1, 2, 6, 7, 8, 9, 12, 13, 14, 15, 16, 18):
                cell.alignment = CENTER_WRAP
            else:
                cell.alignment = LEFT_WRAP
        if key:
            ws.cell(row=row_num, column=1).fill = FILL_KEY
            ws.cell(row=row_num, column=9).fill = FILL_KEY
        if status == "deficiency":
            for c in (14, 15, 16):
                ws.cell(row=row_num, column=c).fill = FILL_DEFICIENCY
        elif status == "hold":
            for c in (14, 15, 16):
                ws.cell(row=row_num, column=c).fill = FILL_HOLD
        elif status == "ok":
            for c in (14, 15):
                ws.cell(row=row_num, column=c).fill = FILL_OK
        ws.row_dimensions[row_num].height = 80

    def write_divider(row_num, text):
        ws.cell(row=row_num, column=1, value=text)
        ws.cell(row=row_num, column=1).font = Font(name="Yu Gothic", size=11, bold=True, color="FFFFFF")
        ws.cell(row=row_num, column=1).fill = PatternFill("solid", fgColor="5B9BD5")
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(ITGC_COLUMNS))
        ws.row_dimensions[row_num].height = 24

    r = 5
    write_divider(r, "■ AC. アクセス管理（Access Management）"); r += 1
    write_itgc(r, [
        "ITGC-AC-001", "アクセス管理", "ユーザ登録",
        "承認なしでのユーザ登録による職務分掌違反・不正アクセス。",
        "新規ユーザ登録は、所属部門長の申請→情シス部アプリチームリーダー（E0053）承認→情シス部担当（E0054）の登録という3段階プロセスをワークフロー（S04）で実施。ロール付与はSoDマトリクスに基づき判定される。",
        "予防的", "手作業+IT", "都度", "Y", "情シス部\nアプリチームリーダー",
        "ユーザ登録申請書（ワークフロー）、\n登録実行ログ、\n月次ユーザ追加レポート",
        "R24 アクセス管理規程", "S01 SAP",
        "有効", "有効", "なし",
        "整備・運用ともに有効", EVAL_INFO
    ], key=True, status="ok"); r += 1

    # 判断保留：ITGC-AC-002 定期棚卸
    write_itgc(r, [
        "ITGC-AC-002", "アクセス管理", "定期棚卸",
        "退職・異動後の不要権限の放置による職務分掌違反・不正アクセスの温床。",
        "情シス部が四半期ごとにSAPユーザ権限一覧（SUIMレポート）を出力し、各部門長に配布。部門長が必要性を確認し、不要権限の削除を申請する。情シス部長（E0051）が棚卸完了を承認する。",
        "発見的", "手作業+IT", "四半期", "Y", "各部門長\n情シス部長",
        "SUIMユーザ一覧、\n部門長承認書、\n削除申請記録",
        "R24 アクセス管理規程", "S01 SAP",
        "有効", "判断保留\n(追加エビデンス要求中)", "調査中",
        "SUIM出力時のタイムスタンプ・抽出条件が不明。完全性の追加エビデンス要求中。", EVAL_INFO
    ], key=True, status="hold"); r += 1

    # 真の不備：ITGC-AC-003 退職者アカウント停止
    write_itgc(r, [
        "ITGC-AC-003", "アクセス管理", "退職者対応",
        "退職者のアカウントが停止されず、不正アクセス・情報漏洩が発生するリスク。",
        "人事部（E0061）が退職決定時に、退職日の3営業日前までに情シス部（E0054）へアカウント停止依頼を送付。情シス部が退職日当日にSAPユーザをロックし、90日後に削除する。",
        "予防的", "手作業", "都度", "Y", "人事部\n情シス部",
        "退職者リスト、\nアカウント停止申請、\nSAPユーザロック履歴",
        "R24 アクセス管理規程", "S01 SAP",
        "有効", "不備あり\n(2件違反)", "あり\n(重要)",
        "退職者5名中2名でSAP停止が11日/18日遅延。職務分掌違反リスク顕在化。是正計画策定中。", EVAL_INFO
    ], key=True, status="deficiency"); r += 1

    write_itgc(r, [
        "ITGC-AC-004", "アクセス管理", "特権ID管理",
        "特権IDの濫用によるシステム改竄・データ漏洩。",
        "特権ID（BASIS等）の使用は申請・承認制とし、全操作ログを情シス部長（E0051）が月次レビュー。特権ID保有者は情シス部2名（E0051/E0052）に限定する。",
        "予防的", "手作業+IT", "月次", "Y", "情シス部長",
        "特権ID利用申請書、\n特権ID操作ログ、\nレビュー記録",
        "R24 アクセス管理規程", "S01 SAP",
        "有効", "有効", "なし",
        "整備・運用ともに有効", EVAL_INFO
    ], key=True, status="ok"); r += 1

    write_divider(r, "■ CM. 変更管理（Change Management）"); r += 1
    write_itgc(r, [
        "ITGC-CM-001", "変更管理", "申請・承認",
        "未承認のプログラム変更による計算誤り・データ破壊。",
        "SAPのプログラム変更は、変更申請書（REL番号）をワークフロー（S04）で起票し、情シス部アプリチームリーダー（E0053）と業務部門責任者の両方の承認を得る。承認なしでの開発・移送は禁止。",
        "予防的", "手作業+IT", "都度", "Y", "アプリチームリーダー\n業務部門責任者",
        "変更申請書(REL-NNN)、\n承認ワークフロー履歴",
        "R22 システム開発・変更管理規程", "S01 SAP",
        "有効", "有効", "なし",
        "整備・運用ともに有効", EVAL_INFO
    ], key=True, status="ok"); r += 1

    write_itgc(r, [
        "ITGC-CM-002", "変更管理", "テスト実施",
        "テスト不備のあるプログラムの本番リリースによる業務停止・データ不整合。",
        "変更のリリース前に、開発環境→テスト環境での単体テスト・結合テスト・UAT（ユーザ受入テスト）の3段階テストを実施し、ユーザ部門の合格署名を得る。緊急変更時は別途緊急変更手順に従う。",
        "予防的", "手作業", "都度", "Y", "アプリチーム\nユーザ部門",
        "テスト計画書、\nテスト結果報告書、\nUAT合格署名",
        "R22 システム開発・変更管理規程", "S01 SAP",
        "有効", "一部例外あり\n(緊急変更1件)", "なし\n(例外許容)",
        "緊急変更1件が事前UAT省略、緊急変更手順に則り事後承認取得済。例外として許容。", EVAL_INFO
    ], key=True, status="ok"); r += 1

    write_itgc(r, [
        "ITGC-CM-003", "変更管理", "本番移送",
        "不正な本番移送による計算誤り・データ改竄。",
        "本番移送（Transport）の実行権限は情シス部アプリチームリーダー（E0053）のみに付与し、移送実行時は変更申請書の承認を確認する。移送ログはSAP STMSで保持。",
        "予防的", "手作業+IT", "都度", "Y", "アプリチームリーダー",
        "本番移送記録(STMS)、\n変更申請書との突合",
        "R22 システム開発・変更管理規程", "S01 SAP",
        "有効", "有効", "なし",
        "整備・運用ともに有効", EVAL_INFO
    ], key=True, status="ok"); r += 1

    write_divider(r, "■ OM. 運用管理（Operations Management）"); r += 1
    write_itgc(r, [
        "ITGC-OM-001", "運用管理", "バックアップ",
        "データ喪失時の復旧不能、業務停止の長期化。",
        "毎日深夜1:00にSAP HANAのフルバックアップを実施（テープ＋クラウド二重化）。情シス部インフラチーム（E0052）が翌朝ジョブ実行ログを確認し、異常時はエスカレーション。四半期に1回復旧テストを実施。",
        "予防的", "IT自動+手作業", "日次/四半期", "Y", "インフラチーム",
        "バックアップジョブ実行ログ、\nリストアテスト結果報告書",
        "R23 システム運用管理規程", "S01 SAP",
        "有効", "有効", "なし",
        "整備・運用ともに有効", EVAL_INFO
    ], key=True, status="ok"); r += 1

    write_itgc(r, [
        "ITGC-OM-002", "運用管理", "障害管理",
        "システム障害の未検知・対応遅延による業務停止・データ不整合。",
        "監視ツール（Zabbix相当）がサーバ・DB・アプリの異常を検知すると、情シス部インフラチーム（E0052）にアラート。障害は障害管理台帳に記録し、原因究明・再発防止策を情シス部長（E0051）がレビューする。",
        "発見的", "IT自動+手作業", "随時", "N", "インフラチーム",
        "障害管理台帳、\n監視アラート履歴、\n再発防止策レポート",
        "R23 システム運用管理規程", "S01 SAP",
        "有効", "有効", "なし",
        "整備・運用ともに有効", EVAL_INFO
    ], key=False, status="ok"); r += 1

    write_divider(r, "■ EM. 外部委託管理（External Management）"); r += 1
    write_itgc(r, [
        "ITGC-EM-001", "外部委託管理", "委託先管理",
        "外部委託先の内部統制不備に起因するデータ漏洩・品質不良。",
        "重要な外部委託先（ERP保守SIer A社、インフラ保守B社）について、年1回SOC1レポート（SSAE18）を入手し、情シス部長（E0051）がコントロール有効性を評価。課題発見時は委託先と改善計画を協議する。",
        "発見的", "手作業", "年次", "Y", "情シス部長",
        "SOC1レポート（SSAE18）、\n評価レビューシート、\n改善計画書",
        "R25 外部委託管理規程", "S01 SAP",
        "有効", "有効", "なし",
        "整備・運用ともに有効", EVAL_INFO
    ], key=True, status="ok"); r += 1

    add_legend_sheet(wb)
    wb.save(BASE / "ITGC_RCM.xlsx")
    print(f"Created: ITGC_RCM.xlsx (10 controls)")


# ==================== ITAC ====================
def gen_itac():
    ITAC_COLUMNS = [
        ("統制ID", 12),
        ("業務領域", 14),
        ("連動PLC統制", 14),
        ("リスク記述", 40),
        ("影響勘定", 14),
        ("アサーション", 12),
        ("統制活動（IT自動統制の内容）", 50),
        ("検証方法", 28),
        ("頻度", 10),
        ("キー", 8),
        ("実施者\n(設計・保守)", 16),
        ("実施証跡", 28),
        ("対象システム", 14),
        ("整備状況\n評価結果", 14),
        ("運用状況\n評価結果", 14),
        ("不備の有無", 10),
        ("最終結論", 16),
        ("評価日/評価者", 18),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ITAC_RCM"
    ws.sheet_view.zoomScale = 85

    ws.cell(row=1, column=1, value="【ITAC】IT業務処理統制 リスク・コントロール・マトリクス（RCM）")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ITAC_COLUMNS))
    ws.cell(row=2, column=1, value="評価対象: SAP S/4HANA(S01)に組み込まれたIT自動統制  /  評価期間: FY2025  /  作成日: 2026/04/14")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(ITAC_COLUMNS))

    for i, (name, width) in enumerate(ITAC_COLUMNS, 1):
        cell = ws.cell(row=4, column=i, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        cell.border = BORDER_HEADER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[4].height = 36
    ws.freeze_panes = "B5"

    def wr(row, values, key=False, status="ok"):
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            if c in (1, 2, 3, 6, 9, 10, 11, 13, 14, 15, 16, 18):
                cell.alignment = CENTER_WRAP
            else:
                cell.alignment = LEFT_WRAP
        if key:
            ws.cell(row=row, column=1).fill = FILL_KEY
            ws.cell(row=row, column=10).fill = FILL_KEY
        if status == "ok":
            for c in (14, 15):
                ws.cell(row=row, column=c).fill = FILL_OK
        ws.row_dimensions[row].height = 85

    rows = [
        ["ITAC-001", "販売", "PLC-S-001",
         "与信限度額を超過する受注が営業担当により登録される。",
         "売上高\n売掛金", "E, A, V",
         "SAPでは受注登録時に顧客マスタの与信限度額と既存売掛金+新規受注合計を自動比較し、超過時は受注が保留となり、営業本部長（E0021）のワークフロー承認がない限り出荷指示に進めない仕様。",
         "SAP設定スクリーンショット、\nITGC-CM統制の健全性確認、\nテストデータによる動作検証",
         "都度", "Y", "情シス部\nアプリチーム\n(設計)",
         "SAP設定画面、\nエラーログ、\nITAC動作検証結果",
         "S01 SAP", "有効", "有効", "なし", "整備・運用ともに有効", EVAL_INFO],
        ["ITAC-002", "購買", "PLC-P-004",
         "発注・検収・請求書の3点が一致しない取引の買掛金計上。",
         "買掛金\n仕入高", "E, C, V",
         "SAPは請求書入力時に発注書(PO)と検収伝票(GR)を自動参照し、金額・数量差異が公差（±5%または¥10,000）を超える場合は計上保留とする。差異は経理部担当（E0016）が原因調査。",
         "SAP MIRO設定の確認、\n25件サンプルによる再実施テスト",
         "都度", "Y", "情シス部\nアプリチーム",
         "SAP MIRO設定画面、\n3wayマッチング結果ログ",
         "S01 SAP", "有効", "有効", "なし", "整備・運用ともに有効", EVAL_INFO],
        ["ITAC-003", "固定資産", "固定資産管理",
         "固定資産の減価償却計算誤り（計算基準・耐用年数・残存価額）。",
         "減価償却費\n減価償却累計額", "V",
         "SAP固定資産モジュール（AA）が月次バッチ（AFAB）で資産ごとに定額法/定率法により自動計算し、仕訳を自動起票。資産取得時に耐用年数・償却方法が自動セットされる（勘定科目クラスルール）。",
         "SAP AA設定画面の確認、\n月次償却ログの閲覧、\n2-3資産の手計算再実施",
         "月次", "Y", "情シス部\n経理部",
         "AA設定画面、\nAFABバッチログ、\n再計算シート",
         "S01 SAP", "有効", "有効", "なし", "整備・運用ともに有効", EVAL_INFO],
        ["ITAC-004", "全プロセス", "PLC-P-002",
         "発注・稟議の承認ルーティングが金額に応じて適切に判定されない。",
         "全勘定科目", "E, A",
         "SAPワークフローは申請金額を自動判定し、承認者ルーティング（50万/500万/2,000万/1億円の閾値）を自動設定する。閾値設定の変更は変更管理プロセス（ITGC-CM）経由のみ。",
         "ワークフロールーティング設定の閲覧、\nテストデータによる動作確認",
         "都度", "Y", "情シス部",
         "ワークフロー設定画面、\nルーティングログ",
         "S01 SAP\nS04 WF", "有効", "有効", "なし", "整備・運用ともに有効", EVAL_INFO],
        ["ITAC-005", "決算", "FCRP-002",
         "子会社からの連結パッケージ取込時のデータ整合性エラー。",
         "連結財務諸表\n全般", "C, V",
         "連結決算システム（S05）が子会社アップロード時にフォーマット・勘定科目マスタ・内部取引額の整合性を自動検証し、エラー時はアップロード拒否。エラーログを経理部（E0012）が確認。",
         "S05設定画面の閲覧、\nエラーログの確認",
         "四半期", "Y", "情シス部\n経理部",
         "S05バリデーションルール、\n取込エラーログ",
         "S05 連結", "有効", "有効", "なし", "整備・運用ともに有効", EVAL_INFO],
    ]

    for i, row in enumerate(rows):
        wr(5 + i, row, key=True, status="ok")

    add_legend_sheet(wb)
    wb.save(BASE / "ITAC_RCM.xlsx")
    print(f"Created: ITAC_RCM.xlsx ({len(rows)} controls)")


# ==================== FCRP ====================
def gen_fcrp():
    FCRP_COLUMNS = [
        ("統制ID", 12),
        ("プロセス", 16),
        ("サブプロセス", 16),
        ("リスク記述", 42),
        ("影響勘定科目", 16),
        ("アサーション", 12),
        ("統制活動", 48),
        ("統制タイプ", 10),
        ("手作業/IT", 10),
        ("頻度", 10),
        ("キー", 8),
        ("実施者", 16),
        ("実施証跡", 28),
        ("関連規程", 14),
        ("対象システム", 12),
        ("整備状況\n評価結果", 14),
        ("運用状況\n評価結果", 14),
        ("不備の\n有無", 10),
        ("最終結論", 18),
        ("評価日/評価者", 18),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FCRP_RCM"
    ws.sheet_view.zoomScale = 85

    ws.cell(row=1, column=1, value="【FCRP】決算財務報告プロセス リスク・コントロール・マトリクス（RCM）")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(FCRP_COLUMNS))
    ws.cell(row=2, column=1, value="評価対象: 親会社経理部・連結決算プロセス  /  評価期間: FY2025  /  作成日: 2026/04/14")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(FCRP_COLUMNS))

    for i, (name, width) in enumerate(FCRP_COLUMNS, 1):
        cell = ws.cell(row=4, column=i, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        cell.border = BORDER_HEADER
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[4].height = 36
    ws.freeze_panes = "B5"

    def wr(row, values, key=False, status="ok"):
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            if c in (1, 2, 3, 6, 8, 9, 10, 11, 14, 15, 16, 17, 18, 20):
                cell.alignment = CENTER_WRAP
            else:
                cell.alignment = LEFT_WRAP
        if key:
            ws.cell(row=row, column=1).fill = FILL_KEY
            ws.cell(row=row, column=11).fill = FILL_KEY
        if status == "deficiency":
            for c in (16, 17, 18):
                ws.cell(row=row, column=c).fill = FILL_DEFICIENCY
        elif status == "hold":
            for c in (16, 17, 18):
                ws.cell(row=row, column=c).fill = FILL_HOLD
        elif status == "ok":
            for c in (16, 17):
                ws.cell(row=row, column=c).fill = FILL_OK
        ws.row_dimensions[row].height = 85

    rows = [
        ["FCRP-001", "月次決算", "月次締め",
         "月次決算のスケジュール遅延・ステップ漏れによる月次財務報告の誤り。",
         "全勘定科目", "C, V, P",
         "経理部課長（E0012）が月次決算チェックリスト（45項目）を使い、期日管理・承認・レビューを実施。経理部長（E0011）が月次決算完了を承認する。",
         "予防的", "手作業", "月次", "Y", "経理部課長\n経理部長",
         "月次決算チェックリスト、\n完了承認記録",
         "R17 決算業務規程", "S01 SAP",
         "有効", "有効", "なし", "整備・運用ともに有効", EVAL_INFO],
        ["FCRP-002", "連結決算", "パッケージ検証",
         "子会社からの連結パッケージの入力誤り・遅延による連結財務諸表の誤り。",
         "連結財務諸表\n全般", "E, C, V",
         "経理部課長（E0012）が各子会社（東北/物流/タイ/TPT）からアップロードされた連結パッケージを連結決算システム（S05）のバリデーションレポートでチェックし、差異について子会社担当者と往復確認する。",
         "発見的", "IT自動+手作業", "四半期", "Y", "経理部課長",
         "連結パッケージ受領台帳、\nバリデーションエラーログ、\n子会社往復確認記録",
         "R17 決算業務規程", "S05 連結",
         "有効", "有効", "なし", "整備・運用ともに有効", EVAL_INFO],
        ["FCRP-003", "決算", "見積レビュー",
         "重要な会計上の見積り（貸倒引当金・棚卸評価損・減損・税効果・退給）の仮定の不適切な設定による利益操作の懸念。",
         "引当金各種\n棚卸資産\n固定資産\n繰延税金資産", "V",
         "四半期末に経理部課長（E0012）が5種の会計上見積りシートを作成し、経理部長（E0011）→CFO（E0002）→監査等委員会でレビュー。前提となる外部情報・内部情報の根拠を添付。",
         "発見的", "手作業", "四半期", "Y", "経理部長\nCFO\n監査等委員会",
         "見積計算シート、\n根拠資料（顧客信用情報等）、\nレビュー議事録",
         "R17 決算業務規程", "S01 SAP\nS07 Excel",
         "有効", "判断保留\n(根拠資料不足)", "調査中",
         "貸倒引当金の個別評価の根拠資料（顧客別信用情報）が未添付。追加エビデンス要求中。", EVAL_INFO],
        ["FCRP-004", "連結決算", "連結仕訳",
         "内部取引消去・投資と資本の相殺などの連結仕訳の誤り・漏れ。",
         "連結財務諸表\n全般", "C, V, R",
         "連結決算システム（S05）で自動起票された連結仕訳について、経理部課長（E0012）がレビューし、経理部長（E0011）が承認。非定型の連結仕訳は個別に検討・文書化する。",
         "予防的", "IT自動+手作業", "四半期", "Y", "経理部課長\n経理部長",
         "連結仕訳一覧、\n承認記録、\n非定型仕訳の個別検討書",
         "R17 決算業務規程", "S05 連結",
         "有効", "有効", "なし", "整備・運用ともに有効", EVAL_INFO],
        ["FCRP-005", "開示", "書類作成",
         "有価証券報告書・短信・適時開示書類の記載誤り・添付資料の不備。",
         "開示書類全般", "R",
         "経営企画部（E0003）が開示システム（S06）で書類を作成し、経理部長（E0011）→CFO（E0002）→監査等委員会→取締役会と3段階レビューを実施。前期比較やXBRLタグの整合性を確認する。",
         "発見的", "IT自動+手作業", "四半期", "Y", "経営企画部\n経理部長\nCFO",
         "開示書類ドラフト、\nレビュー指摘事項一覧、\n取締役会議事録",
         "R17 決算業務規程", "S06 開示",
         "有効", "有効", "なし", "整備・運用ともに有効", EVAL_INFO],
    ]

    for i, row in enumerate(rows):
        status = row[-4].split()[0] if "保留" in str(row[16]) else ("deficiency" if "不備" in str(row[17]) else "ok")
        wr(5 + i, row, key=True, status="hold" if i == 2 else "ok")

    add_legend_sheet(wb)
    wb.save(BASE / "FCRP_RCM.xlsx")
    print(f"Created: FCRP_RCM.xlsx ({len(rows)} controls)")


# ==================== RCM Summary ====================
def gen_summary():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RCM_Summary"
    ws.sheet_view.zoomScale = 100

    ws.cell(row=1, column=1, value="J-SOX 内部統制 RCM サマリ（53統制クロスリファレンス）")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    headers = ["RCM区分", "統制ID", "統制名(概要)", "プロセス/領域", "キー", "運用評価結果", "不備", "リンク先ファイル"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        cell.border = BORDER_HEADER
    ws.row_dimensions[3].height = 30

    data = [
        # ELC (12)
        ("ELC", "ELC-001", "取締役会の機能", "I.統制環境", "Y", "有効", "", "ELC_RCM.xlsx"),
        ("ELC", "ELC-002", "倫理綱領の浸透", "I.統制環境", "N", "有効", "", "ELC_RCM.xlsx"),
        ("ELC", "ELC-003", "職務権限と組織体制", "I.統制環境", "Y", "有効", "", "ELC_RCM.xlsx"),
        ("ELC", "ELC-004", "全社リスク評価の実施", "II.リスク評価", "Y", "有効", "", "ELC_RCM.xlsx"),
        ("ELC", "ELC-005", "不正リスク評価", "II.リスク評価", "Y", "有効", "", "ELC_RCM.xlsx"),
        ("ELC", "ELC-006", "規程・マニュアル整備", "III.統制活動", "N", "有効", "", "ELC_RCM.xlsx"),
        ("ELC", "ELC-007", "職務の分離", "III.統制活動", "Y", "一部不備あり", "軽微", "ELC_RCM.xlsx"),
        ("ELC", "ELC-008", "内部通報制度", "IV.情報と伝達", "Y", "有効", "", "ELC_RCM.xlsx"),
        ("ELC", "ELC-009", "決算情報の伝達", "IV.情報と伝達", "N", "有効", "", "ELC_RCM.xlsx"),
        ("ELC", "ELC-010", "内部監査の実施", "V.モニタリング", "Y", "有効", "", "ELC_RCM.xlsx"),
        ("ELC", "ELC-011", "監査等委員会のモニタリング", "V.モニタリング", "Y", "有効", "", "ELC_RCM.xlsx"),
        ("ELC", "ELC-012", "IT戦略と情報セキュリティ方針", "VI.IT対応", "Y", "有効", "", "ELC_RCM.xlsx"),
        # PLC-S (7)
        ("PLC-S", "PLC-S-001", "受注・与信承認", "販売", "Y", "一部例外(許容)", "", "PLC_Sales_RCM.xlsx"),
        ("PLC-S", "PLC-S-002", "出荷-売上マッチング", "販売", "Y", "有効", "", "PLC_Sales_RCM.xlsx"),
        ("PLC-S", "PLC-S-003", "請求書発行", "販売", "Y", "有効", "", "PLC_Sales_RCM.xlsx"),
        ("PLC-S", "PLC-S-004", "入金消込", "販売", "Y", "有効", "", "PLC_Sales_RCM.xlsx"),
        ("PLC-S", "PLC-S-005", "売掛金年齢分析", "販売", "N", "判断保留", "調査中", "PLC_Sales_RCM.xlsx"),
        ("PLC-S", "PLC-S-006", "期末カットオフ", "販売", "Y", "有効", "", "PLC_Sales_RCM.xlsx"),
        ("PLC-S", "PLC-S-007", "価格マスタ承認", "販売", "N", "有効", "", "PLC_Sales_RCM.xlsx"),
        # PLC-P (7)
        ("PLC-P", "PLC-P-001", "購買依頼承認", "購買", "N", "有効", "", "PLC_Purchasing_RCM.xlsx"),
        ("PLC-P", "PLC-P-002", "発注承認(金額別)", "購買", "Y", "不備あり", "重要", "PLC_Purchasing_RCM.xlsx"),
        ("PLC-P", "PLC-P-003", "検収", "購買", "Y", "有効", "", "PLC_Purchasing_RCM.xlsx"),
        ("PLC-P", "PLC-P-004", "3-wayマッチング", "購買", "Y", "有効", "", "PLC_Purchasing_RCM.xlsx"),
        ("PLC-P", "PLC-P-005", "仕入先マスタ管理", "購買", "N", "有効", "", "PLC_Purchasing_RCM.xlsx"),
        ("PLC-P", "PLC-P-006", "支払承認", "購買", "Y", "有効", "", "PLC_Purchasing_RCM.xlsx"),
        ("PLC-P", "PLC-P-007", "期末未払計上", "購買", "Y", "有効", "", "PLC_Purchasing_RCM.xlsx"),
        # PLC-I (7)
        ("PLC-I", "PLC-I-001", "実地棚卸", "在庫", "Y", "有効", "", "PLC_Inventory_Cost_RCM.xlsx"),
        ("PLC-I", "PLC-I-002", "棚卸差異調整", "在庫", "Y", "不備あり", "軽微", "PLC_Inventory_Cost_RCM.xlsx"),
        ("PLC-I", "PLC-I-003", "標準原価更新承認", "原価", "Y", "有効", "", "PLC_Inventory_Cost_RCM.xlsx"),
        ("PLC-I", "PLC-I-004", "原価差異分析", "原価", "N", "有効", "", "PLC_Inventory_Cost_RCM.xlsx"),
        ("PLC-I", "PLC-I-005", "滞留在庫評価損", "在庫", "Y", "有効", "", "PLC_Inventory_Cost_RCM.xlsx"),
        ("PLC-I", "PLC-I-006", "WMS-ERP在庫一致", "在庫", "N", "有効", "", "PLC_Inventory_Cost_RCM.xlsx"),
        ("PLC-I", "PLC-I-007", "原価計算月次締め", "原価", "Y", "有効", "", "PLC_Inventory_Cost_RCM.xlsx"),
        # ITGC (10)
        ("ITGC", "ITGC-AC-001", "新規ユーザ登録承認", "アクセス管理", "Y", "有効", "", "ITGC_RCM.xlsx"),
        ("ITGC", "ITGC-AC-002", "アクセス権定期棚卸", "アクセス管理", "Y", "判断保留", "調査中", "ITGC_RCM.xlsx"),
        ("ITGC", "ITGC-AC-003", "退職者アカウント停止", "アクセス管理", "Y", "不備あり", "重要", "ITGC_RCM.xlsx"),
        ("ITGC", "ITGC-AC-004", "特権ID管理", "アクセス管理", "Y", "有効", "", "ITGC_RCM.xlsx"),
        ("ITGC", "ITGC-CM-001", "変更申請・承認", "変更管理", "Y", "有効", "", "ITGC_RCM.xlsx"),
        ("ITGC", "ITGC-CM-002", "テスト実施", "変更管理", "Y", "一部例外(許容)", "", "ITGC_RCM.xlsx"),
        ("ITGC", "ITGC-CM-003", "本番移送", "変更管理", "Y", "有効", "", "ITGC_RCM.xlsx"),
        ("ITGC", "ITGC-OM-001", "バックアップ", "運用管理", "Y", "有効", "", "ITGC_RCM.xlsx"),
        ("ITGC", "ITGC-OM-002", "障害管理", "運用管理", "N", "有効", "", "ITGC_RCM.xlsx"),
        ("ITGC", "ITGC-EM-001", "委託先管理(SOC1)", "外部委託", "Y", "有効", "", "ITGC_RCM.xlsx"),
        # ITAC (5)
        ("ITAC", "ITAC-001", "与信限度自動チェック", "販売(連動)", "Y", "有効", "", "ITAC_RCM.xlsx"),
        ("ITAC", "ITAC-002", "3-way自動マッチング", "購買(連動)", "Y", "有効", "", "ITAC_RCM.xlsx"),
        ("ITAC", "ITAC-003", "減価償却自動計算", "固定資産", "Y", "有効", "", "ITAC_RCM.xlsx"),
        ("ITAC", "ITAC-004", "承認ルーティング判定", "全プロセス", "Y", "有効", "", "ITAC_RCM.xlsx"),
        ("ITAC", "ITAC-005", "連結パッケージ検証", "決算(連動)", "Y", "有効", "", "ITAC_RCM.xlsx"),
        # FCRP (5)
        ("FCRP", "FCRP-001", "月次決算チェックリスト", "月次決算", "Y", "有効", "", "FCRP_RCM.xlsx"),
        ("FCRP", "FCRP-002", "連結パッケージ検証", "連結決算", "Y", "有効", "", "FCRP_RCM.xlsx"),
        ("FCRP", "FCRP-003", "会計上の見積レビュー", "決算", "Y", "判断保留", "調査中", "FCRP_RCM.xlsx"),
        ("FCRP", "FCRP-004", "連結仕訳承認", "連結決算", "Y", "有効", "", "FCRP_RCM.xlsx"),
        ("FCRP", "FCRP-005", "開示書類レビュー", "開示", "Y", "有効", "", "FCRP_RCM.xlsx"),
    ]

    for r, row in enumerate(data, start=4):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = BORDER
            if c in (1, 2, 4, 5, 6, 7):
                cell.alignment = CENTER_WRAP
            else:
                cell.alignment = LEFT_WRAP
        # 不備の色
        status = row[5]
        if "不備あり" in status:
            ws.cell(row=r, column=6).fill = FILL_DEFICIENCY
            ws.cell(row=r, column=7).fill = FILL_DEFICIENCY
        elif "判断保留" in status:
            ws.cell(row=r, column=6).fill = FILL_HOLD
            ws.cell(row=r, column=7).fill = FILL_HOLD
        elif status == "有効":
            ws.cell(row=r, column=6).fill = FILL_OK
        # キー強調
        if row[4] == "Y":
            ws.cell(row=r, column=5).fill = FILL_KEY

    widths = [12, 16, 30, 16, 8, 18, 10, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    # 集計シート
    ws2 = wb.create_sheet("統計")
    summary = [
        ["", "統制数", "キー統制数", "不備あり", "判断保留", "有効"],
        ["ELC", 12, 9, 1, 0, 11],
        ["PLC-Sales", 7, 5, 0, 1, 6],
        ["PLC-Purchasing", 7, 5, 1, 0, 6],
        ["PLC-Inventory", 7, 5, 1, 0, 6],
        ["ITGC", 10, 9, 1, 1, 8],
        ["ITAC", 5, 5, 0, 0, 5],
        ["FCRP", 5, 5, 0, 1, 4],
        ["合計", 53, 43, 4, 3, 46],
    ]
    for r, row in enumerate(summary, 1):
        for c, v in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT if r > 1 and r < 9 else Font(name="Yu Gothic", size=10, bold=True)
            cell.border = BORDER
            cell.alignment = CENTER_WRAP
            if r == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            if r == 9:  # 合計行
                cell.fill = PatternFill("solid", fgColor="D9E1F2")
    for i, w in enumerate([16, 10, 12, 12, 12, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "RCM_Summary.xlsx")
    print(f"Created: RCM_Summary.xlsx ({len(data)} controls indexed)")


if __name__ == "__main__":
    gen_itgc()
    gen_itac()
    gen_fcrp()
    gen_summary()
