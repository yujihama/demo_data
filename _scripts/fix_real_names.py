"""実在する会社名・サービス名をダミー名称に置換

検出された実在名:
- Microsoft (company_profile.md 2箇所) → 表計算ソフト
- Concur (FCRP Checklist 12ファイル) → 経費精算SaaS
- 帝国データバンク (audit procedures) → 外部信用調査機関
- 東京商工リサーチ (FCRP Estimate) → 外部信用調査機関A
"""
import os
import sys
import io
from pathlib import Path
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

ROOT = Path(r"C:\Users\nyham\work\demo_data")

# 置換マップ
REPLACEMENTS = {
    'Microsoft': '表計算ソフト (商用)',
    'Concur': '経費精算SaaS',
    '帝国データバンク': '外部信用調査機関',
    '東京商工リサーチ': '外部信用調査機関A',
}


def replace_in_markdown(path):
    """Markdown/CSV/TXT textual replacement"""
    with open(path, encoding='utf-8-sig') as f:
        content = f.read()
    original = content
    for real, dummy in REPLACEMENTS.items():
        content = content.replace(real, dummy)
    if content != original:
        with open(path, 'w', encoding='utf-8-sig', newline='') as f:
            f.write(content)
        return True
    return False


def replace_in_xlsx(path):
    """xlsx cell value replacement"""
    wb = load_workbook(path)
    changed = False
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    v = cell.value
                    new_v = v
                    for real, dummy in REPLACEMENTS.items():
                        new_v = new_v.replace(real, dummy)
                    if new_v != v:
                        cell.value = new_v
                        changed = True
    if changed:
        wb.save(path)
    return changed


# Scan and replace
changed_files = []
for root, dirs, files in os.walk(ROOT):
    if '.git' in root or '__pycache__' in root or '_scripts' in root: continue
    for f in files:
        fp = Path(root) / f
        if f.endswith(('.md', '.csv', '.txt')):
            try:
                if replace_in_markdown(fp):
                    changed_files.append(str(fp))
            except Exception as e:
                print(f'[Err] {fp}: {e}')
        elif f.endswith('.xlsx'):
            try:
                if replace_in_xlsx(fp):
                    changed_files.append(str(fp))
            except Exception as e:
                print(f'[Err] {fp}: {e}')

print(f'\n=== 置換完了: {len(changed_files)} ファイル ===')
for f in changed_files:
    print(f'  - {Path(f).relative_to(ROOT)}')
