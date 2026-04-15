"""
PLC-P（購買プロセス）エビデンス生成
【真の不備ケース：PLC-P-002 発注承認の権限超過3件】を含む
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
from pathlib import Path
import random
import sys

sys.path.insert(0, str(Path(__file__).parent))
from pdf_util import JPPDF
from image_util import sap_screenshot, workflow_screenshot

BASE = Path(r"C:\Users\nyham\work\demo_data\4.evidence\PLC-P")
BASE.mkdir(parents=True, exist_ok=True)

HF = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
BFONT = Font(name="Yu Gothic", size=10)
BBOLD = Font(name="Yu Gothic", size=10, bold=True)
C_ = Alignment(horizontal="center", vertical="center", wrap_text=True)
L_ = Alignment(horizontal="left", vertical="center", wrap_text=True)
R_ = Alignment(horizontal="right", vertical="center")
T_ = Side("thin", color="888888")
BRD = Border(left=T_, right=T_, top=T_, bottom=T_)
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")
FILL_NG = PatternFill("solid", fgColor="FCE4D6")
FILL_OK = PatternFill("solid", fgColor="E2EFDA")

# 仕入先マスタ（ダミー）
VENDORS = {
    "V-20001": ("サンプル仕入先A社", "原材料(鋼材)"),
    "V-20002": ("サンプル仕入先B社", "原材料(鋼材)"),
    "V-20003": ("サンプル仕入先C社", "原材料(鋼材)"),
    "V-20004": ("サンプル仕入先D社", "原材料(鋼材)"),
    "V-20005": ("サンプル仕入先E社", "原材料(鋼材)"),
    "V-20006": ("サンプル仕入先F社", "原材料(銅材)"),
    "V-20007": ("サンプル仕入先G社", "原材料(銅材)"),
    "V-20008": ("サンプル仕入先H社", "原材料(特殊合金)"),
    "V-20009": ("サンプル仕入先I社", "原材料(鉛材)"),
    "V-20010": ("サンプル仕入先J社", "原材料(アルミ)"),
    "V-20011": ("サンプル仕入先K社", "原材料(亜鉛)"),
    "V-20012": ("サンプル仕入先L社", "原材料(チタン)"),
    "V-20013": ("サンプル仕入先M社", "原材料(ワイヤ)"),
    "V-20014": ("サンプル仕入先N社", "原材料(電線)"),
    "V-20015": ("サンプル仕入先O社", "原材料(鋼材)"),
    "V-20021": ("サンプル仕入先P社", "外注加工(金型)"),
    "V-20022": ("サンプル仕入先Q社", "外注加工(切削)"),
    "V-20023": ("サンプル仕入先R社", "外注加工(プレス)"),
    "V-20024": ("サンプル仕入先S社", "外注加工(研削)"),
    "V-20025": ("サンプル仕入先T社", "外注加工(熱処理)"),
    "V-20026": ("サンプル仕入先U社", "外注加工(表面処理)"),
    "V-20027": ("サンプル仕入先V社", "外注加工(溶接)"),
    "V-20028": ("サンプル仕入先W社", "外注加工(研磨)"),
    "V-20031": ("サンプル仕入先X社", "消耗品"),
    "V-20032": ("サンプル仕入先Y社", "消耗品・治工具"),
}


# ============================================================
# PLC-P-002 母集団: SAP ME2N 発注伝票一覧
# ============================================================
def gen_population_po():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SAP_ME2N_発注伝票一覧"

    ws.cell(row=1, column=1, value="SAP S/4HANA / トランザクション: ME2N / 出力日時: 2026/02/12 14:20:08")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    ws.cell(row=2, column=1, value="出力者: ACC001 佐藤 一郎 / 抽出条件: 発注日=2025/4/1-2026/3/31 / 発注タイプ=通常発注 / 総件数: 3,874件（以下は前120件）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13)

    headers = ["№", "発注番号", "発注日", "発注区分", "仕入先コード", "仕入先名", "品目分類",
               "発注金額(円)", "ステータス", "承認者", "承認者上限(円)",
               "承認ルート", "備考"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[4].height = 28

    # 承認者定義
    APPROVERS = [
        ("清水 智明 (PUR003)", 1_000_000),      # 購買部主任（ただしPOは通常 linkのみ）
        ("林 真由美 (PUR002)", 5_000_000),      # 購買部課長
        ("木村 浩二 (PUR001)", 20_000_000),     # 購買部長
        ("渡辺 正博 (CFO001)", 100_000_000),    # CFO
        ("山本 健一 (CEO001)", 999_999_999),    # 代表取締役
    ]

    def select_approver(amount):
        for name, limit in APPROVERS:
            if amount <= limit:
                return name, limit
        return APPROVERS[-1]

    random.seed(2025002)
    vids = list(VENDORS.keys())
    r = 5

    # 意図的な不備ケース：3件の権限超過
    deficiency_rows = {
        17: {
            "po_no": "PO-2025-0234",
            "date": date(2025, 9, 12),
            "vid": "V-20002",
            "amount": 680_000,
            "approver": "清水 智明 (PUR003)",  # 本来PO_APPROVE権限なし
            "limit": 500_000,
            "remark": "※ 承認者PUR003にPO_APPROVE権限なし（SoD違反）",
        },
        42: {
            "po_no": "PO-2025-0789",
            "date": date(2025, 10, 3),
            "vid": "V-20008",
            "amount": 1_250_000,
            "approver": "山田 純一 (PUR004)",  # 上限50万円だが承認
            "limit": 500_000,
            "remark": "※ 購買部担当PUR004が上限超の発注を承認",
        },
        78: {
            "po_no": "PO-2025-1456",
            "date": date(2025, 11, 8),
            "vid": "V-20004",
            "amount": 7_850_000,
            "approver": "林 真由美 (PUR002)",  # 上限500万円
            "limit": 5_000_000,
            "remark": "※ 課長上限¥5M超の¥7.85Mを承認",
        },
    }

    for i in range(1, 121):
        # 日付（FY2025内）
        month = random.choices(
            [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3],
            weights=[10, 10, 10, 10, 10, 10, 12, 12, 8, 8, 6, 4])[0]
        year = 2025 if month >= 4 else 2026
        day = random.randint(1, 28)
        po_date = date(year, month, day)

        vid = random.choice(vids)
        vname, vcat = VENDORS[vid]

        # 金額
        amount = random.choice([
            random.randint(100_000, 500_000),
            random.randint(500_000, 3_000_000),
            random.randint(3_000_000, 10_000_000),
            random.randint(10_000_000, 30_000_000),
        ])
        approver, limit = select_approver(amount)
        remark = ""
        po_no = f"PO-2025-{i * 32:04d}"

        # 不備ケースの上書き
        if i in deficiency_rows:
            d = deficiency_rows[i]
            po_no = d["po_no"]
            po_date = d["date"]
            vid = d["vid"]
            vname, vcat = VENDORS[vid]
            amount = d["amount"]
            approver = d["approver"]
            limit = d["limit"]
            remark = d["remark"]

        status = random.choices(["完了", "処理中"], weights=[85, 15])[0]
        route = "課長→部長" if amount > 5_000_000 else "課長単独"

        data = [i, po_no, po_date, "通常発注", vid, vname, vcat, amount,
                status, approver, limit, route, remark]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 5, 7, 9, 12):
                cell.alignment = C_
                if c_i == 3:
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (8, 11):
                cell.alignment = R_
                cell.number_format = "#,##0"
            else:
                cell.alignment = L_
        # 不備行を赤く強調
        if i in deficiency_rows:
            for c_i in range(1, 14):
                ws.cell(row=r, column=c_i).fill = FILL_NG
        r += 1

    ws.cell(row=r + 1, column=1, value="※ 本レポートはFY2025全3,874件中の前120件です。完全版はSAP上で参照可能。")
    ws.cell(row=r + 1, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=13)

    widths = [5, 15, 12, 10, 10, 18, 14, 14, 10, 22, 14, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    wb.save(BASE / "PLC-P-002_SAP_ME2N_発注伝票一覧_FY2025.xlsx")
    print("Created: PLC-P-002_SAP_ME2N_発注伝票一覧_FY2025.xlsx")


# ============================================================
# 承認権限規程抜粋（PDF）
# ============================================================
def gen_purchase_authority_pdf():
    pdf = JPPDF()
    pdf.add_page()

    pdf.h1("購買関連 承認権限一覧")
    pdf.set_font("YuGoth", "", 9)
    pdf.cell(0, 5, "職務権限規程R18 抜粋（購買部分）/ 改訂日: 2025/4/1",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    pdf.h2("1. 発注承認権限（金額別）")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
                   "発注金額に応じて承認者が自動判定される。SAPワークフロー（S04）により、"
                   "金額超の場合は上位者の追加承認が必要。")
    pdf.ln(2)
    pdf.table_header(["金額区分", "承認者", "承認上限(円)", "SAPロール"],
                     [40, 50, 40, 60])
    pdf.table_row(["〜¥500,000", "購買部担当(PUR003)", "¥500,000", "PO_CREATE のみ"],
                  [40, 50, 40, 60])
    pdf.table_row(["〜¥5,000,000", "購買部課長(PUR002)", "¥5,000,000", "PO_APPROVE"],
                  [40, 50, 40, 60], fill=True)
    pdf.table_row(["〜¥20,000,000", "購買部長(PUR001)", "¥20,000,000", "PO_APPROVE"],
                  [40, 50, 40, 60])
    pdf.table_row(["〜¥100,000,000", "管理本部長（CFO）", "¥100,000,000", "PO_APPROVE"],
                  [40, 50, 40, 60], fill=True)
    pdf.table_row(["¥100,000,000超", "代表取締役社長", "制限なし", "PO_APPROVE"],
                  [40, 50, 40, 60])
    pdf.ln(5)

    pdf.h2("2. 例外手続")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
                   "・緊急案件で上位承認者が不在の場合は、課長経由で仮承認後、"
                   "翌営業日までに正式承認を取得する。\n"
                   "・承認者不在時のみ、取締役会で指名された代理人が承認可能。\n"
                   "・承認権限を超えた発注が事後的に発覚した場合は、必ず上位承認を事後取得し、"
                   "再発防止策を文書化する。")
    pdf.ln(5)

    pdf.h2("3. 職務分掌の原則")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
                   "・発注作成（PO_CREATE）と発注承認（PO_APPROVE）は同一人物が兼任してはならない。\n"
                   "・発注承認と検収は異なる担当者が行う。\n"
                   "・検収と買掛計上は異なる担当者が行う。\n"
                   "・買掛計上と支払は異なる担当者が行う。")
    pdf.ln(10)

    pdf.set_font("YuGoth", "", 9)
    pdf.cell(0, 5, "承認: 2025/4/1施行 / 山本 健一 代表取締役社長",
             new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(BASE / "PLC-P-002_購買関連承認権限一覧_職務権限規程R18抜粋.pdf"))
    print("Created: PLC-P-002_購買関連承認権限一覧_職務権限規程R18抜粋.pdf")


# ============================================================
# PLC-P-001 購買依頼一覧
# ============================================================
def gen_purchase_requisition():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "購買依頼一覧"

    ws.cell(row=1, column=1, value="SAP ME5A / 購買依頼一覧 (2025年11月分)")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.cell(row=2, column=1, value="出力日時: 2025/12/02 10:15 / 出力者: PUR003 清水 智明")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    headers = ["依頼番号", "依頼日", "起案部門", "起案者", "品目名", "数量",
               "予算額(円)", "予算コード", "部門長承認", "承認日"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    random.seed(2025001)
    depts = [("製造本部", "森 和雄 (MFG001)"), ("製造本部", "池田 昌夫 (MFG002)"),
             ("技術本部", "山田 技術長"), ("情報システム部", "岡田 宏 (IT001)"),
             ("品質保証部", "品証部長")]
    items = ["SUS304鋼材 φ20×1000mm", "特殊合金材 A種 100kg",
             "プレス金型用材料", "切削油 ML-2", "ベアリング部品一式",
             "精密測定器具", "安全手袋 1000組"]

    r = 5
    for i in range(1, 31):
        req_date = date(2025, 11, random.randint(1, 28))
        dept, person = random.choice(depts)
        item = random.choice(items)
        qty = random.randint(1, 50)
        amount = random.choice([
            random.randint(50_000, 300_000),
            random.randint(300_000, 2_000_000),
            random.randint(2_000_000, 10_000_000),
        ])
        budget = f"BGT-{dept[:2]}-{random.randint(1, 20):02d}"
        approval = "森 和雄 [印]" if dept == "製造本部" else "部門長 [印]"
        appr_date = req_date + timedelta(days=random.randint(0, 2))
        data = [f"PR-2025-{i * 3 + 100:04d}", req_date, dept, person,
                item, qty, amount, budget, approval, appr_date]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 6, 8, 9, 10):
                cell.alignment = C_
                if c_i in (2, 10):
                    cell.number_format = "yyyy/mm/dd"
            elif c_i == 7:
                cell.alignment = R_
                cell.number_format = "#,##0"
            else:
                cell.alignment = L_
        r += 1

    widths = [14, 12, 14, 20, 25, 8, 14, 14, 15, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-P-001_SAP購買依頼一覧_202511.xlsx")
    print("Created: PLC-P-001_SAP購買依頼一覧_202511.xlsx")


# ============================================================
# 発注書PDF（通常ケース）
# ============================================================
def gen_po_pdf(po_no, po_date, vendor_code, items, delivery_date,
               approver_role, approver_name, approver_date,
               output_name, is_deficiency=False):
    pdf = JPPDF()
    pdf.add_page()

    pdf.set_font("YuGoth", "B", 20)
    pdf.cell(0, 12, "発 注 書", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, f"発注番号: {po_no}", align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"発注日: {po_date.strftime('%Y年%m月%d日')}",
             align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    # 発注元（自社）
    pdf.set_font("YuGoth", "B", 11)
    pdf.cell(0, 6, "株式会社テクノプレシジョン 購買部",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("YuGoth", "", 9)
    pdf.cell(0, 5, "〒XXX-XXXX 神奈川県横浜市港北区", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "TEL: 045-XXX-XXXX", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    # 発注先
    vname, vcat = VENDORS[vendor_code]
    pdf.set_font("YuGoth", "B", 12)
    pdf.cell(0, 7, f"{vname} 御中", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("YuGoth", "", 9)
    pdf.cell(0, 5, f"仕入先コード: {vendor_code} / 品目分類: {vcat}",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5, "下記のとおり発注申し上げます。納期厳守のうえ、ご納入ください。")
    pdf.ln(3)

    # 明細
    pdf.table_header(["品目コード", "品名", "数量", "単価(円)", "金額(円)"],
                     [30, 80, 20, 30, 30])
    subtotal = 0
    for code, name, qty, unit in items:
        amount = qty * unit
        subtotal += amount
        pdf.table_row([code, name, f"{qty:,}", f"{unit:,}", f"{amount:,}"],
                      [30, 80, 20, 30, 30])

    pdf.set_font("YuGoth", "B", 10)
    pdf.cell(130, 7, "小計", border=1, align="R")
    pdf.cell(60, 7, f"¥ {subtotal:,}", border=1, align="R",
             new_x="LMARGIN", new_y="NEXT")
    tax = int(subtotal * 0.1)
    pdf.cell(130, 7, "消費税 (10%)", border=1, align="R")
    pdf.cell(60, 7, f"¥ {tax:,}", border=1, align="R",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_fill_color(255, 242, 204)
    total = subtotal + tax
    pdf.cell(130, 8, "合計", border=1, align="R", fill=True)
    pdf.cell(60, 8, f"¥ {total:,}", border=1, align="R", fill=True,
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_fill_color(255, 255, 255)
    pdf.ln(5)

    pdf.h3("■ 納入条件")
    pdf.set_font("YuGoth", "", 9)
    pdf.kv("納期", delivery_date.strftime("%Y年%m月%d日"))
    pdf.kv("納入場所", "当社指定倉庫")
    pdf.kv("支払条件", "月末締 翌月末払")
    pdf.ln(5)

    # 承認欄
    pdf.h3("■ 社内承認記録")
    pdf.set_font("YuGoth", "B", 10)
    pdf.cell(50, 7, "承認者役割", border=1, align="C", fill=True)
    pdf.cell(60, 7, "氏名", border=1, align="C", fill=True)
    pdf.cell(40, 7, "承認日", border=1, align="C", fill=True)
    pdf.cell(30, 7, "承認印", border=1, align="C", fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(50, 14, approver_role, border=1, align="C")
    pdf.cell(60, 14, approver_name, border=1, align="C")
    pdf.cell(40, 14, approver_date.strftime("%Y/%m/%d"), border=1, align="C")
    x_stamp = pdf.get_x()
    y_stamp = pdf.get_y()
    pdf.cell(30, 14, "", border=1, new_x="LMARGIN", new_y="NEXT")
    if is_deficiency:
        # 不備の場合、赤色で「要検討」スタンプ
        pdf.set_text_color(200, 30, 30)
        pdf.set_draw_color(200, 30, 30)
        pdf.set_line_width(0.5)
        pdf.circle(x_stamp + 15, y_stamp + 7, 8)
        pdf.set_font("YuGoth", "B", 8)
        pdf.text(x_stamp + 9, y_stamp + 8, "要検討")
        pdf.set_text_color(0, 0, 0)
        pdf.set_draw_color(0, 0, 0)
    else:
        pdf.stamp("承認", x=x_stamp + 15, y=y_stamp + 7)

    if is_deficiency:
        pdf.ln(8)
        pdf.set_text_color(200, 30, 30)
        pdf.set_font("YuGoth", "B", 10)
        pdf.multi_cell(0, 5, f"※ 本発注書は承認権限を超過する金額が承認されている可能性あり。"
                             f"内部監査の発見事項。是正措置検討中。")
        pdf.set_text_color(0, 0, 0)

    pdf.output(str(BASE / output_name))
    print(f"Created: {output_name}")


def gen_po_pdfs():
    # 正常ケース
    gen_po_pdf(
        "PO-2025-2560",
        date(2025, 10, 18),
        "V-20001",
        [
            ("RAW-001", "SUS304鋼材 φ30×3000mm", 50, 28500),
            ("RAW-002", "SUS304鋼材 φ20×1500mm", 30, 15800),
        ],
        date(2025, 11, 15),
        "購買部課長",
        "林 真由美 (PUR002)",
        date(2025, 10, 18),
        "PLC-P-002_発注書_PO-2025-2560_通常承認.pdf",
    )
    # 通常：大型案件（部長承認）
    gen_po_pdf(
        "PO-2025-3072",
        date(2025, 11, 5),
        "V-20008",
        [("RAW-H01", "特殊合金材インコネル718", 20, 385000)],
        date(2025, 12, 20),
        "購買部長",
        "木村 浩二 (PUR001)",
        date(2025, 11, 5),
        "PLC-P-002_発注書_PO-2025-3072_通常承認.pdf",
    )
    # 不備ケース1
    gen_po_pdf(
        "PO-2025-0234",
        date(2025, 9, 12),
        "V-20002",
        [("RAW-002", "SUS304鋼材 φ20×1500mm", 40, 17000)],
        date(2025, 9, 25),
        "購買部主任（※権限なし）",
        "清水 智明 (PUR003)",
        date(2025, 9, 12),
        "PLC-P-002_発注書_PO-2025-0234_不備ケース1_権限外承認.pdf",
        is_deficiency=True,
    )
    # 不備ケース2
    gen_po_pdf(
        "PO-2025-0789",
        date(2025, 10, 3),
        "V-20008",
        [("RAW-H02", "特殊合金材 B種", 5, 250000)],
        date(2025, 10, 20),
        "購買部担当（※上限50万円を超過）",
        "山田 純一 (PUR004)",
        date(2025, 10, 3),
        "PLC-P-002_発注書_PO-2025-0789_不備ケース2_担当者承認.pdf",
        is_deficiency=True,
    )
    # 不備ケース3
    gen_po_pdf(
        "PO-2025-1456",
        date(2025, 11, 8),
        "V-20004",
        [("RAW-D01", "特殊鋼材大型", 100, 78500)],
        date(2025, 12, 10),
        "購買部課長（※上限500万円を超過）",
        "林 真由美 (PUR002)",
        date(2025, 11, 8),
        "PLC-P-002_発注書_PO-2025-1456_不備ケース3_課長上限超過.pdf",
        is_deficiency=True,
    )


# ============================================================
# SAP承認ワークフロー履歴ログ
# ============================================================
def gen_workflow_log():
    path = BASE / "PLC-P-002_SAPワークフロー承認履歴ログ_FY2025抜粋.csv"
    lines = [
        "# SAP Business Workflow / 発注承認履歴",
        "# 出力日時: 2026/02/12 15:10:08",
        "# 出力者: IT003 加藤 洋子（情シス部アプリチームリーダー）",
        "# 抽出条件: FY2025 期間中の発注承認ワークフロー",
        "",
        "ワークフロー番号,発注番号,起票日時,起票者,金額,承認ルート,承認者1,承認日時1,承認者2,承認日時2,最終ステータス",
    ]

    random.seed(5555)
    import datetime as dt

    # サンプル30件 + 不備3件
    normal_samples = []
    for i in range(1, 31):
        po_no = f"PO-2025-{random.randint(100, 3000):04d}"
        amt = random.choice([random.randint(100_000, 5_000_000),
                             random.randint(5_000_000, 20_000_000)])
        ts_start = datetime(2025, random.randint(4, 12), random.randint(1, 28),
                            random.randint(9, 17), random.randint(0, 59))
        if amt <= 500_000:
            approver1 = "PUR003 清水 智明"; limit1 = 1_000_000
            route = "単独承認"
            approver2 = ""; ts2 = ""
        elif amt <= 5_000_000:
            approver1 = "PUR002 林 真由美"; limit1 = 5_000_000
            route = "課長単独"
            approver2 = ""; ts2 = ""
        elif amt <= 20_000_000:
            approver1 = "PUR002 林 真由美"
            approver2 = "PUR001 木村 浩二"
            route = "課長→部長"
            ts_appr1 = ts_start + timedelta(hours=random.randint(1, 5))
            ts_appr2 = ts_appr1 + timedelta(hours=random.randint(1, 8))
            ts1 = ts_appr1.strftime("%Y-%m-%d %H:%M:%S")
            ts2 = ts_appr2.strftime("%Y-%m-%d %H:%M:%S")
            normal_samples.append((po_no, ts_start, amt, route, approver1, ts1, approver2, ts2))
            continue
        else:
            approver1 = "PUR001 木村 浩二"
            approver2 = "CFO001 渡辺 正博"
            route = "部長→CFO"
            ts_appr1 = ts_start + timedelta(hours=random.randint(1, 5))
            ts_appr2 = ts_appr1 + timedelta(hours=random.randint(1, 24))
            ts1 = ts_appr1.strftime("%Y-%m-%d %H:%M:%S")
            ts2 = ts_appr2.strftime("%Y-%m-%d %H:%M:%S")
            normal_samples.append((po_no, ts_start, amt, route, approver1, ts1, approver2, ts2))
            continue

        ts_appr1 = ts_start + timedelta(hours=random.randint(1, 8))
        ts1 = ts_appr1.strftime("%Y-%m-%d %H:%M:%S")
        normal_samples.append((po_no, ts_start, amt, route, approver1, ts1, approver2, ts2))

    # 不備ケース
    deficiency_samples = [
        ("PO-2025-0234", datetime(2025, 9, 12, 10, 15), 680_000, "※権限外承認",
         "PUR003 清水 智明(権限外)", "2025-09-12 10:45:12", "", ""),
        ("PO-2025-0789", datetime(2025, 10, 3, 14, 20), 1_250_000, "※上限超過",
         "PUR004 山田 純一(上限50万円超)", "2025-10-03 14:55:38", "", ""),
        ("PO-2025-1456", datetime(2025, 11, 8, 11, 5), 7_850_000, "※上限超過",
         "PUR002 林 真由美(上限500万円超)", "2025-11-08 11:42:21", "", ""),
    ]

    i = 1
    for (po_no, ts_start, amt, route, app1, ts1, app2, ts2) in normal_samples[:15]:
        wf_no = f"WF-2025-{i * 23 + 1000:05d}"
        lines.append(f"{wf_no},{po_no},{ts_start.strftime('%Y-%m-%d %H:%M:%S')},PUR003/PUR004,{amt},{route},{app1},{ts1},{app2},{ts2},承認完了")
        i += 1

    # 不備行をマーカー行に
    lines.append("# ↓↓ 以下、内部監査の検出により「要検討」とマークされた案件 ↓↓")
    for (po_no, ts_start, amt, route, app1, ts1, app2, ts2) in deficiency_samples:
        wf_no = f"WF-2025-{i * 23 + 1000:05d}"
        lines.append(f"{wf_no},{po_no},{ts_start.strftime('%Y-%m-%d %H:%M:%S')},PUR003/PUR004,{amt},{route},{app1},{ts1},{app2},{ts2},要検討")
        i += 1

    for (po_no, ts_start, amt, route, app1, ts1, app2, ts2) in normal_samples[15:]:
        wf_no = f"WF-2025-{i * 23 + 1000:05d}"
        lines.append(f"{wf_no},{po_no},{ts_start.strftime('%Y-%m-%d %H:%M:%S')},PUR003/PUR004,{amt},{route},{app1},{ts1},{app2},{ts2},承認完了")
        i += 1

    lines.append("")
    lines.append("# 件数: 33件（うち3件は権限超過疑い、要検討）")
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"Created: {path.name}")


# ============================================================
# 検収記録（PLC-P-003）
# ============================================================
def gen_grn_pdf():
    """検収報告書PDF"""
    pdf = JPPDF()
    pdf.add_page()

    pdf.set_font("YuGoth", "B", 20)
    pdf.cell(0, 12, "検 収 報 告 書", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, "検収番号: REC-2025-5678", align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "検収日: 2025年11月18日", align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.kv("発注番号", "PO-2025-2560", key_w=30)
    pdf.kv("仕入先", "V-20001 サンプル仕入先A社", key_w=30)
    pdf.kv("納入場所", "本社倉庫A", key_w=30)
    pdf.kv("検収担当", "橋本 明（倉庫課長 WHS001）", key_w=30)
    pdf.kv("品質保証確認", "品質保証部 検査員 [印]", key_w=30)
    pdf.ln(5)

    pdf.h2("検収明細")
    pdf.table_header(["品目コード", "品名", "発注数量", "納入数量", "検収判定"],
                     [30, 80, 25, 25, 30])
    pdf.table_row(["RAW-001", "SUS304鋼材 φ30×3000mm", "50", "50", "合格"],
                  [30, 80, 25, 25, 30])
    pdf.table_row(["RAW-002", "SUS304鋼材 φ20×1500mm", "30", "30", "合格"],
                  [30, 80, 25, 25, 30], fill=True)
    pdf.ln(5)

    pdf.h2("品質検査結果")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
                   "・材料証明書: 添付あり（ミルシート確認済）\n"
                   "・外観検査: 良好\n"
                   "・寸法検査: 抜取検査（各品目3本）、規格内\n"
                   "・材質検査: ミルシート記載の材質と一致")
    pdf.ln(5)

    pdf.h3("■ 検収判定")
    pdf.set_font("YuGoth", "B", 12)
    pdf.set_fill_color(220, 240, 220)
    pdf.cell(0, 10, "検収合格 / SAPに登録済（数量差異なし）", align="C", fill=True,
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_fill_color(255, 255, 255)
    pdf.ln(5)

    # 承認欄
    pdf.set_font("YuGoth", "B", 10)
    pdf.cell(45, 7, "検収担当者", border=1, align="C", fill=True)
    pdf.cell(45, 7, "倉庫課長", border=1, align="C", fill=True)
    pdf.cell(45, 7, "品質保証部", border=1, align="C", fill=True)
    pdf.cell(45, 7, "購買部確認", border=1, align="C", fill=True,
             new_x="LMARGIN", new_y="NEXT")

    pdf.set_font("YuGoth", "", 10)
    pdf.cell(45, 16, "", border=1)
    pdf.cell(45, 16, "", border=1)
    pdf.cell(45, 16, "", border=1)
    pdf.cell(45, 16, "", border=1, new_x="LMARGIN", new_y="NEXT")
    # 複数のスタンプ
    y = pdf.get_y() - 12
    pdf.stamp("検収", x=22, y=y)
    pdf.stamp("確認", x=67, y=y)
    pdf.stamp("合格", x=112, y=y)
    pdf.stamp("確認", x=157, y=y)

    pdf.output(str(BASE / "PLC-P-003_検収報告書_REC-2025-5678.pdf"))
    print("Created: PLC-P-003_検収報告書_REC-2025-5678.pdf")


# ============================================================
# 検収差異報告書（例外）
# ============================================================
def gen_grn_diff_pdf():
    pdf = JPPDF()
    pdf.add_page()

    pdf.set_font("YuGoth", "B", 18)
    pdf.cell(0, 12, "検収差異報告書", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, "差異報告書番号: DIF-2025-0019", align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "発行日: 2025年10月28日", align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.kv("発注番号", "PO-2025-1892")
    pdf.kv("仕入先", "V-20022 サンプル仕入先Q社")
    pdf.kv("検収日", "2025/10/27")
    pdf.kv("検収担当", "橋本 明（倉庫課長 WHS001）")
    pdf.ln(5)

    pdf.h2("差異内容")
    pdf.table_header(["品目コード", "発注数", "納入数", "差異", "原因"],
                     [25, 25, 25, 25, 80])
    pdf.table_row(["MCH-002", "100", "95", "-5", "仕入先側の出荷漏れ"],
                  [25, 25, 25, 25, 80])
    pdf.ln(5)

    pdf.h2("対応")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
                   "1. 95個について検収完了・SAPに数量95で登録（金額もそれに従い調整）\n"
                   "2. 残5個については仕入先Q社に2025/10/27電話連絡、翌週中に追加納入予定\n"
                   "3. 追加納入時に別の検収番号REC-2025-5721で検収する\n"
                   "4. 購買部清水主任に差異発生を連絡済")
    pdf.ln(5)

    pdf.h3("■ 報告先・承認")
    pdf.set_font("YuGoth", "", 10)
    pdf.kv("報告先", "購買部 清水 主任 → 林 課長")
    pdf.kv("承認者", "木村 浩二 購買部長 [印]")
    pdf.kv("是正フォロー", "2025/11/4 追加納入完了")
    pdf.ln(5)

    pdf.output(str(BASE / "PLC-P-003_検収差異報告書_DIF-2025-0019.pdf"))
    print("Created: PLC-P-003_検収差異報告書_DIF-2025-0019.pdf")


# ============================================================
# PLC-P-004 3-wayマッチング結果
# ============================================================
def gen_3way_match():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "3wayマッチング結果"

    ws.cell(row=1, column=1, value="【PLC-P-004 統制実施記録】 3-wayマッチング結果レポート (2025年11月)")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    ws.cell(row=2, column=1, value="出力元: SAP MIRO自動マッチング / 出力日: 2025/12/03 / 確認: 石井 健（ACC006）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    headers = ["№", "請求書番号", "請求日", "発注番号", "検収番号", "仕入先",
               "PO金額", "検収金額", "請求金額", "マッチ結果"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    random.seed(4004)
    vids = list(VENDORS.keys())
    r = 5
    for i in range(1, 31):
        idate = date(2025, 11, random.randint(1, 30))
        amount = random.randint(200_000, 10_000_000)
        vid = random.choice(vids)
        vname = VENDORS[vid][0]
        po_amount = amount
        rec_amount = amount
        inv_amount = amount
        match = "一致"
        if i == 19:
            inv_amount = amount - 3_000
            match = "差異¥3,000\n(端数調整・許容)"
        elif i == 25:
            inv_amount = amount + 150_000
            match = "差異¥150,000\n(公差超過・保留)"
        data = [i, f"INV-V-202511-{i:04d}", idate, f"PO-2025-{random.randint(100, 3000):04d}",
                f"REC-2025-{random.randint(1000, 9000):04d}", vname,
                po_amount, rec_amount, inv_amount, match]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 5, 10):
                cell.alignment = C_
                if c_i == 3:
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (7, 8, 9):
                cell.alignment = R_
                cell.number_format = "#,##0"
            else:
                cell.alignment = L_
        if match == "一致":
            ws.cell(row=r, column=10).fill = FILL_OK
        elif "許容" in match:
            ws.cell(row=r, column=10).fill = FILL_WARN
        else:
            ws.cell(row=r, column=10).fill = FILL_NG
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="マッチング結果: 一致28件、公差内差異1件（許容）、保留1件（原因調査中）").font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

    widths = [5, 16, 12, 15, 15, 18, 12, 12, 12, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-P-004_3wayマッチング結果_202511.xlsx")
    print("Created: PLC-P-004_3wayマッチング結果_202511.xlsx")


# ============================================================
# PLC-P-005 仕入先マスタ登録申請
# ============================================================
def gen_vendor_reg_application():
    pdf = JPPDF()
    pdf.add_page()

    pdf.h1("仕入先マスタ登録申請書")
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, "申請番号: VEND-REG-2025-0015 / 申請日: 2025年10月12日",
             align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.kv("申請者", "清水 智明（購買部主任 PUR003）")
    pdf.kv("申請理由", "新規切削加工外注先として登録")
    pdf.ln(3)

    pdf.h2("登録内容")
    pdf.kv("仕入先コード(予)", "V-20029")
    pdf.kv("仕入先名", "サンプル仕入先Z社")
    pdf.kv("品目分類", "外注加工（研磨）")
    pdf.kv("所在地", "関東地区")
    pdf.kv("支払条件", "月末締 翌月末払")
    pdf.kv("銀行口座", "A銀行 支店P 普通 XXXXXXX")
    pdf.ln(5)

    pdf.h2("反社会的勢力チェック")
    pdf.kv("チェック実施日", "2025/10/8")
    pdf.kv("チェック担当", "総務部 前田 美香 (GA001)")
    pdf.kv("チェック結果", "○ 問題なし")
    pdf.kv("参照したデータベース", "公的DB（警察庁、都道府県公安委員会）、民間DB X")
    pdf.ln(5)

    pdf.h2("信用調査結果")
    pdf.kv("調査日", "2025/10/10")
    pdf.kv("調査機関", "外部信用情報会社 Y")
    pdf.kv("評点", "58点（業界平均52点、問題なし）")
    pdf.kv("財務状況", "直近3期とも黒字、自己資本比率45%")
    pdf.ln(5)

    pdf.h3("■ 承認")
    pdf.set_font("YuGoth", "B", 10)
    pdf.cell(60, 7, "役割", border=1, align="C", fill=True)
    pdf.cell(60, 7, "氏名", border=1, align="C", fill=True)
    pdf.cell(40, 7, "日付", border=1, align="C", fill=True)
    pdf.cell(30, 7, "承認印", border=1, align="C", fill=True, new_x="LMARGIN", new_y="NEXT")

    approvals = [
        ("購買部課長", "林 真由美 (PUR002)", "2025/10/13"),
        ("購買部長", "木村 浩二 (PUR001)", "2025/10/14"),
    ]
    pdf.set_font("YuGoth", "", 10)
    for role, name, dt in approvals:
        pdf.cell(60, 14, role, border=1, align="C")
        pdf.cell(60, 14, name, border=1, align="C")
        pdf.cell(40, 14, dt, border=1, align="C")
        x_stamp = pdf.get_x()
        y_stamp = pdf.get_y()
        pdf.cell(30, 14, "", border=1, new_x="LMARGIN", new_y="NEXT")
        pdf.stamp("承認", x=x_stamp + 15, y=y_stamp + 7)

    pdf.ln(3)
    pdf.set_font("YuGoth", "", 9)
    pdf.cell(0, 5, "SAP登録日: 2025/10/15 / 登録者: IT004 西田 徹 / 四半期レビュー対象",
             new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(BASE / "PLC-P-005_仕入先マスタ登録申請書_V-20029.pdf"))
    print("Created: PLC-P-005_仕入先マスタ登録申請書_V-20029.pdf")


# ============================================================
# PLC-P-006 支払予定一覧
# ============================================================
def gen_payment_schedule():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "支払予定一覧"

    ws.cell(row=1, column=1, value="【PLC-P-006 統制実施記録】 2025年11月末 支払予定一覧")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    ws.cell(row=2, column=1, value="作成日: 2025/11/25 / 作成者: 石井 健（経理部担当 ACC006） / 承認: 佐藤 一郎（経理部長 ACC001）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    headers = ["№", "支払予定日", "仕入先コード", "仕入先名", "買掛金残高(円)",
               "支払額(円)", "振込銀行", "振込口座", "備考"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    random.seed(4046)
    vids = list(VENDORS.keys())
    r = 5
    total = 0
    for i in range(1, 26):
        vid = random.choice(vids)
        vname = VENDORS[vid][0]
        amount = random.randint(500_000, 25_000_000) // 1000 * 1000
        total += amount
        bank_info = f"A銀行 支店X / 普通 XXXX{random.randint(100, 999)}"
        data = [i, date(2025, 11, 30), vid, vname, amount, amount,
                bank_info[:5], bank_info[6:], ""]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 7, 8, 9):
                cell.alignment = C_
                if c_i == 2:
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (5, 6):
                cell.alignment = R_
                cell.number_format = "#,##0"
            else:
                cell.alignment = L_
        r += 1

    # 合計行
    ws.cell(row=r, column=1, value="合計").font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1).alignment = C_
    ws.cell(row=r, column=6, value=total).font = BBOLD
    ws.cell(row=r, column=6).number_format = "#,##0"
    ws.cell(row=r, column=6).alignment = R_

    r += 2
    ws.cell(row=r, column=1, value="作成: 石井 健 [印] 2025/11/25 / 承認: 佐藤 一郎 [印] 2025/11/26")
    r += 1
    ws.cell(row=r, column=1, value="振込実行者: 小川 由紀 (ACC005) / 実行日: 2025/11/30（職務分離確認）")

    widths = [5, 12, 12, 18, 14, 14, 10, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-P-006_支払予定一覧_202511.xlsx")
    print("Created: PLC-P-006_支払予定一覧_202511.xlsx")


# ============================================================
# PLC-P-007 期末未払計上リスト
# ============================================================
def gen_accrual_list():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "期末未払計上"

    ws.cell(row=1, column=1, value="【PLC-P-007 統制実施記録】 FY2025期末 未払計上リスト")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=2, column=1, value="作成日: 2026/4/5 / 作成: 高橋 美咲（経理部課長 ACC002） / 承認: 佐藤 一郎（経理部長 ACC001）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headers = ["№", "検収番号", "検収日", "仕入先コード", "仕入先名",
               "検収金額(円)", "請求書受領", "未払計上額(円)"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    random.seed(7007)
    vids = list(VENDORS.keys())
    r = 5
    for i in range(1, 21):
        rec_date = date(2026, 3, random.randint(20, 31))
        vid = random.choice(vids)
        amount = random.randint(500_000, 15_000_000)
        data = [i, f"REC-2026-{random.randint(5000, 9000):04d}",
                rec_date, vid, VENDORS[vid][0],
                amount, "未着", amount]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 7):
                cell.alignment = C_
                if c_i == 3:
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (6, 8):
                cell.alignment = R_
                cell.number_format = "#,##0"
            else:
                cell.alignment = L_
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="件数: 20件 / 合計未払計上額: 業者照会後確定").font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    ws.cell(row=r, column=1, value="業者照会記録: 全件対応済（2026/4/3〜4/4）/ 経理部長承認: 佐藤 一郎 [印] 2026/4/5")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    widths = [5, 16, 12, 12, 18, 14, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-P-007_期末未払計上リスト.xlsx")
    print("Created: PLC-P-007_期末未払計上リスト.xlsx")


# ============================================================
# SAP/WFスクリーンショット
# ============================================================
def gen_screenshots():
    sap_screenshot(
        "発注登録 - 照会",
        "ME23N",
        [
            ("発注番号", "PO-2025-2560"),
            ("発注日", "2025/10/18"),
            ("仕入先", "V-20001 サンプル仕入先A社"),
            ("発注担当", "PUR003 清水 智明"),
            ("承認者", "PUR002 林 真由美（購買部課長）"),
            ("承認日", "2025/10/18 15:23"),
            ("発注金額合計", "1,898,000 JPY"),
            ("ステータス", "承認済 - 発注書送付済"),
            ("納期", "2025/11/15"),
        ],
        grid_headers=["品目", "製品コード", "数量", "単価", "金額"],
        grid_rows=[
            ["SUS304鋼材 φ30×3000mm", "RAW-001", "50", "28,500", "1,425,000"],
            ["SUS304鋼材 φ20×1500mm", "RAW-002", "30", "15,800", "474,000"],
        ],
        status_bar="発注伝票 PO-2025-2560 が承認されました。",
        output_path=str(BASE / "PLC-P-002_SAP発注登録画面_PO-2025-2560.png"),
    )

    # 承認権限超過の発注画面（意図的不備を示す）
    sap_screenshot(
        "発注登録 - 承認権限警告",
        "ME29N",
        [
            ("発注番号", "PO-2025-1456"),
            ("発注日", "2025/11/08"),
            ("仕入先", "V-20004 サンプル仕入先D社"),
            ("発注担当", "PUR003 清水 智明"),
            ("発注金額", "7,850,000 JPY"),
            ("承認者", "PUR002 林 真由美 (※上限5,000,000)"),
            ("承認上限判定", "× 上限超過（要上位承認）"),
            ("上位承認取得", "未取得（内部監査指摘事項）"),
            ("ステータス", "承認済 - 要検討"),
        ],
        grid_headers=["品目", "製品コード", "数量", "単価", "金額"],
        grid_rows=[
            ["特殊鋼材大型", "RAW-D01", "100", "78,500", "7,850,000"],
        ],
        status_bar="⚠ 承認権限超過の疑いあり。内部監査による追跡中。",
        output_path=str(BASE / "PLC-P-002_SAP発注画面_不備ケースPO-2025-1456.png"),
    )

    workflow_screenshot(
        "WF-2025-4789",
        "発注承認（通常案件）",
        "清水 智明（購買部主任）",
        [
            ("清水 智明", "起票者（購買部主任）", "2025/10/18 11:22", "申請"),
            ("林 真由美", "課長（購買部）", "2025/10/18 15:23", "承認"),
        ],
        amount=1_898_000,
        comments="サンプル仕入先A社 へのSUS304鋼材発注（原材料調達）",
        output_path=str(BASE / "PLC-P-002_ワークフロー承認画面_通常案件.png"),
    )

    print("Created: 3 screenshots for PLC-P")


# ============================================================
# Main
# ============================================================
if __name__ == "__main__":
    gen_population_po()
    gen_purchase_authority_pdf()
    gen_purchase_requisition()
    gen_po_pdfs()
    gen_workflow_log()
    gen_grn_pdf()
    gen_grn_diff_pdf()
    gen_3way_match()
    gen_vendor_reg_application()
    gen_payment_schedule()
    gen_accrual_list()
    gen_screenshots()
    print("\nAll PLC-P evidence generated.")
