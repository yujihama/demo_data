"""
RCM生成の共通ユーティリティ
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Yu Gothic", size=10)
BODY_BOLD = Font(name="Yu Gothic", size=10, bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="888888")
MED = Side(style="medium", color="1F4E78")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_HEADER = Border(left=MED, right=MED, top=MED, bottom=MED)

# キーコントロール色
FILL_KEY = PatternFill("solid", fgColor="FFF2CC")  # 薄い黄色
FILL_DEFICIENCY = PatternFill("solid", fgColor="FCE4D6")  # 薄い赤
FILL_HOLD = PatternFill("solid", fgColor="DEEBF7")  # 薄い青（判断保留）
FILL_OK = PatternFill("solid", fgColor="E2EFDA")  # 薄い緑
FILL_SUB_HEADER = PatternFill("solid", fgColor="BDD7EE")


STANDARD_COLUMNS = [
    ("統制ID", 12),
    ("プロセス\n(COSO要素)", 14),
    ("サブプロセス", 14),
    ("リスク記述", 40),
    ("影響勘定科目", 16),
    ("アサーション", 12),
    ("統制活動", 46),
    ("統制タイプ\n(予防/発見)", 10),
    ("手作業/\nIT自動", 10),
    ("頻度", 10),
    ("キー\nコントロール", 10),
    ("実施者(役割)", 18),
    ("実施証跡(エビデンス)", 28),
    ("関連規程", 14),
    ("関連システム", 14),
    ("整備状況\n評価結果", 14),
    ("運用状況\n評価結果", 14),
    ("不備の\n有無", 10),
    ("最終結論", 16),
    ("評価日/評価者", 18),
]


def init_rcm_sheet(ws, title_line, columns=STANDARD_COLUMNS):
    """RCMシートの初期化（タイトル行 + ヘッダ行）"""
    ws.sheet_view.zoomScale = 90
    # Title row
    ws.cell(row=1, column=1, value=title_line)
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))

    # Header
    for i, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=3, column=i, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        cell.border = BORDER_HEADER
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[3].height = 36
    ws.freeze_panes = "B4"


def write_rcm_row(ws, row_num, values, key_control=False, status=None):
    """1行のRCMデータを書き込み。statusは 'deficiency' 'hold' 'ok' または None"""
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=c, value=v)
        cell.font = BODY_FONT
        cell.border = BORDER
        # 列ごとの寄せ
        if c in (1, 2, 3, 6, 8, 9, 10, 11, 14, 15, 16, 17, 18, 20):
            cell.alignment = CENTER_WRAP
        else:
            cell.alignment = LEFT_WRAP

    # キーコントロールなら列1-3を強調
    if key_control:
        for c in (1, 11):
            ws.cell(row=row_num, column=c).fill = FILL_KEY
            ws.cell(row=row_num, column=c).font = BODY_BOLD

    # 不備・保留の色付け（16,17,18,19列）
    if status == "deficiency":
        for c in (17, 18, 19):
            ws.cell(row=row_num, column=c).fill = FILL_DEFICIENCY
    elif status == "hold":
        for c in (17, 18, 19):
            ws.cell(row=row_num, column=c).fill = FILL_HOLD
    elif status == "ok":
        for c in (16, 17):
            ws.cell(row=row_num, column=c).fill = FILL_OK


def add_legend_sheet(wb):
    """凡例シートを追加"""
    ws = wb.create_sheet("凡例")
    ws["A1"] = "RCM凡例"
    ws["A1"].font = Font(name="Yu Gothic", size=14, bold=True)

    legend = [
        ("アサーション", ""),
        ("E", "実在性 (Existence)"),
        ("C", "網羅性 (Completeness)"),
        ("A", "権利と義務の帰属 (Rights and Obligations)"),
        ("V", "評価の妥当性 (Valuation)"),
        ("P", "期間帰属の適切性 (Period Allocation / Cut-off)"),
        ("R", "表示の妥当性 (Presentation)"),
        ("", ""),
        ("統制タイプ", ""),
        ("予防的", "誤謬・不正の発生を未然に防ぐ統制"),
        ("発見的", "発生した誤謬・不正を事後的に発見する統制"),
        ("", ""),
        ("セル背景色", ""),
        ("黄色（キーコントロール列）", "重要統制（Key Control）"),
        ("緑（運用評価列）", "有効"),
        ("赤（運用評価列）", "不備あり（不合格）"),
        ("青（運用評価列）", "判断保留（追加エビデンス要求中）"),
        ("", ""),
        ("関連システム略号", ""),
        ("S01", "SAP S/4HANA（基幹ERP）"),
        ("S02", "WMS（倉庫管理システム）"),
        ("S03", "給与計算SaaS"),
        ("S04", "稟議ワークフロー"),
        ("S05", "連結決算システム"),
        ("S06", "開示システム"),
        ("S07", "Excel(EUC)"),
    ]

    for r, (k, v) in enumerate(legend, start=3):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
        if k and not v:
            ws.cell(row=r, column=1).font = Font(name="Yu Gothic", size=11, bold=True)
            ws.cell(row=r, column=1).fill = FILL_SUB_HEADER
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        else:
            ws.cell(row=r, column=1).font = BODY_FONT
            ws.cell(row=r, column=2).font = BODY_FONT

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
