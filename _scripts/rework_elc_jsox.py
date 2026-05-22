# -*- coding: utf-8 -*-
"""Rebuild ELC RCM, audit procedures, and aligned evidence for FY2025 J-SOX demo data."""

from __future__ import annotations

import csv
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import PageBreak, Paragraph, Preformatted, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet


BASE = Path(__file__).resolve().parents[1]
RCM_DIR = BASE / "2.RCM"
PROC_DIR = BASE / "3.audit_procedures"
ELC_DIR = BASE / "4.evidence" / "ELC"

FSA_BASELINE_URL = "https://www.fsa.go.jp/singi/singi_kigyou/kijun/20230407_naibutousei_kansa.pdf"
FSA_RELEASE_URL = "https://www.fsa.go.jp/news/r4/sonota/20230407/20230407.html"
FSA_QA_URL = "https://www.fsa.go.jp/news/r5/sonota/20230831-2/20230831-2.html"


CONTROLS = [
    {
        "id": "ELC-001",
        "coso": "統制環境",
        "item": "取締役会・監査等委員会による監督",
        "risk": "経営者による財務報告内部統制の整備・運用に対する監督が不足し、重要な不備や経営者による統制無視が看過される。",
        "control": "取締役会は内部統制基本方針、J-SOX評価範囲、重要リスク、不備是正状況を年次又は四半期で審議する。監査等委員会は内部監査室・外部監査人から報告を受け、財務報告内部統制に関する指摘事項を取締役会へ共有する。",
        "type": "発見的",
        "frequency": "四半期",
        "key": "Y",
        "owner": "取締役会 / 監査等委員会 / 総務部",
        "evidence": "取締役会・監査等委員会_財務報告内部統制モニタリング_FY2025.csv; 取締役会議事録_第245回_2025年9月.pdf",
        "policy": "R01, R02, R03",
        "sample": "4",
        "design_proc": "取締役会規則、監査等委員会規則、内部監査規程を閲覧し、財務報告内部統制の監督責任、報告経路、議題化の要件が定義されていることを確認する。",
        "oper_proc": "四半期のモニタリング台帳及び議事録を閲覧し、評価範囲、重要リスク、監査結果、不備是正が取締役会又は監査等委員会で報告・審議されていることを確認する。",
    },
    {
        "id": "ELC-002",
        "coso": "統制環境",
        "item": "倫理・コンプライアンスと不正防止",
        "risk": "役職員の倫理意識や不正防止意識が不足し、財務報告に影響する不正・法令違反が発生する。",
        "control": "会社は倫理綱領及びコンプライアンス規程を全役職員に周知し、年1回のeラーニング受講及び受領確認を求める。未了者は人事部が部門長へ督促し、完了状況を総務部へ報告する。",
        "type": "予防的",
        "frequency": "年次",
        "key": "Y",
        "owner": "人事部 / 総務部",
        "evidence": "HRIS_CodeOfEthics_AcknowledgmentLog_FY2025.csv",
        "policy": "R04, R08",
        "sample": "1",
        "design_proc": "倫理綱領、コンプライアンス規程、研修運用手順を閲覧し、対象者、頻度、未了者フォロー、報告責任が定義されていることを確認する。",
        "oper_proc": "HRISの対象者別受領確認ログを母集団として入手し、集計行の対象者数522名と個票明細522件、受領確認済件数522件が一致していることを確認する。",
    },
    {
        "id": "ELC-003",
        "coso": "統制環境",
        "item": "組織構造・権限責任・財務報告人材",
        "risk": "組織構造、職務権限、財務報告責任が曖昧なため、承認権限逸脱、職務分掌不備、決算レビュー不足が発生する。",
        "control": "総務部は年1回及び組織変更時に組織図、職務権限規程、財務報告関連部門の責任者一覧を見直し、経理・情報システム・事業部門の責任分担を更新する。重要な改訂は取締役会又は管理本部長が承認する。",
        "type": "予防的",
        "frequency": "年次 / 随時",
        "key": "Y",
        "owner": "総務部 / 管理本部",
        "evidence": "財務報告責任者一覧_承認証跡_FY2025.csv; 職務権限・組織体制_年次レビュー_FY2025.csv; 規程_職務権限規程_R18.pdf; employees.xlsx; company_profile.pdf",
        "policy": "R07, R18, R19",
        "sample": "1",
        "design_proc": "職務権限規程、組織図、責任者一覧を閲覧し、財務報告に関係する承認権限と責任部門が明文化されていることを確認する。",
        "oper_proc": "年次レビュー記録を閲覧し、組織変更・人事異動・権限変更が識別され、必要な規程更新と承認が実施されたことを確認する。",
    },
    {
        "id": "ELC-004",
        "coso": "リスクの評価と対応",
        "item": "財務報告リスク評価と評価範囲",
        "risk": "財務報告に重要な影響を及ぼす拠点、勘定、業務プロセス、ITリスクが評価範囲から漏れる。",
        "control": "経営企画部及び内部監査室は年次で財務報告リスク評価を実施し、売上、在庫、購買、決算、IT等の重要リスクを識別する。評価結果はJ-SOX評価範囲、キーコントロール、監査計画に反映し、取締役会へ報告する。",
        "type": "予防的",
        "frequency": "年次 / 重要変化時",
        "key": "Y",
        "owner": "経営企画部 / 内部監査室",
        "evidence": "全社リスクアセスメント結果_2025年度.xlsx",
        "policy": "R05, R03",
        "sample": "1",
        "design_proc": "リスク管理規程及びJ-SOX評価計画を閲覧し、財務報告リスクの識別、評価、範囲決定、重要変化時の再評価プロセスが定義されていることを確認する。",
        "oper_proc": "全社リスクアセスメント結果を閲覧し、財務報告リスク、評価範囲、対応統制、取締役会報告が記録されていることを確認する。",
    },
    {
        "id": "ELC-005",
        "coso": "リスクの評価と対応",
        "item": "不正リスク評価と経営者による統制無視への対応",
        "risk": "不正リスク、経営者による統制無視、会計上の見積り・仕訳・開示の恣意性が十分に評価されない。",
        "control": "内部監査室、経理部、総務部は年次で不正リスク評価を実施し、動機・機会・正当化の観点から売上前倒し、架空仕入、在庫評価、見積り、手入力仕訳、開示遅延を評価する。リスク対応はPLC、FCRP、ITGCのキー統制及び監査手続に反映する。",
        "type": "予防的",
        "frequency": "年次",
        "key": "Y",
        "owner": "内部監査室 / 経理部 / 総務部",
        "evidence": "全社リスクアセスメント結果_2025年度.xlsx",
        "policy": "R03, R04, R05",
        "sample": "1",
        "design_proc": "不正リスク評価手順を閲覧し、動機・機会・正当化、経営者による統制無視、見積り・仕訳・開示を評価対象としていることを確認する。",
        "oper_proc": "不正リスク評価シートを閲覧し、重要な不正シナリオ、対応統制、残余リスク、監査計画への反映が記録されていることを確認する。",
    },
    {
        "id": "ELC-006",
        "coso": "統制活動",
        "item": "規程・業務手順の整備と年次見直し",
        "risk": "財務報告に関連する業務規程や手順が陳腐化し、現行業務・システム・承認権限と不一致になる。",
        "control": "各主管部門は財務報告に関連する主要規程及び業務手順を年1回見直し、法令・会計基準・組織・システム変更の影響を反映する。改訂要否、承認者、施行日、周知状況を規程レビュー台帳で管理する。",
        "type": "予防的",
        "frequency": "年次 / 随時",
        "key": "Y",
        "owner": "総務部 / 各主管部門",
        "evidence": "規程・業務手順_年次レビュー台帳_FY2025.xlsx; company_profile.pdf",
        "policy": "R01-R27",
        "sample": "1",
        "design_proc": "規程管理ルールを閲覧し、主要規程の主管部門、見直し頻度、承認、改訂履歴、周知の要件が定義されていることを確認する。",
        "oper_proc": "年次レビュー台帳を閲覧し、財務報告関連規程の見直し結果、改訂要否、承認、周知が記録されていることを確認する。",
    },
    {
        "id": "ELC-007",
        "coso": "情報と伝達",
        "item": "決算・開示情報の伝達と締め管理",
        "risk": "決算進捗、未処理事項、重要な会計論点が経営者・関係部門に適時に伝達されず、財務報告の誤謬又は遅延が発生する。",
        "control": "経理部は月次決算締めカレンダーを作成し、SAP締め処理、経理部長レビュー、経営会議報告の完了状況を管理する。四半期及び年度末は財務報告連絡会で重要論点、見積り、開示、子会社報告状況を共有する。",
        "type": "発見的",
        "frequency": "月次 / 四半期",
        "key": "Y",
        "owner": "経理部",
        "evidence": "SAP_PeriodClose_JobLog_FY2025.csv; 財務報告連絡会_議事メモ_2025年度Q4.pdf",
        "policy": "R16, R17",
        "sample": "12",
        "design_proc": "決算業務規程、月次締めカレンダー、財務報告連絡会の運営ルールを閲覧し、締め、レビュー、報告の責任と期限が定義されていることを確認する。",
        "oper_proc": "12か月分の締めログとレビュー記録を閲覧し、締め処理、経理部長レビュー、経営会議報告が期限内に完了していることを確認する。",
    },
    {
        "id": "ELC-008",
        "coso": "情報と伝達",
        "item": "内部通報・苦情処理と監査等委員会へのエスカレーション",
        "risk": "不正・法令違反・財務報告上の懸念が通常の報告経路に埋没し、適切な調査及び是正に結び付かない。",
        "control": "会社は社内窓口及び外部弁護士窓口を設置し、通報受付、利益相反判定、調査責任者決定、調査結果、是正、監査等委員会報告を台帳で管理する。財務報告に影響する可能性がある通報はCFO及び監査等委員会へ速やかに報告する。",
        "type": "発見的",
        "frequency": "随時 / 四半期報告",
        "key": "Y",
        "owner": "総務部 / 外部弁護士 / 監査等委員会",
        "evidence": "内部通報受付台帳_FY2025.xlsx",
        "policy": "R06",
        "sample": "1",
        "design_proc": "内部通報規程を閲覧し、独立した通報経路、匿名性、調査、是正、監査等委員会報告の要件が定義されていることを確認する。",
        "oper_proc": "通報台帳を閲覧し、FY2025受付案件について受付から調査完了、財務報告影響評価、監査等委員会報告まで記録されていることを確認する。",
    },
    {
        "id": "ELC-009",
        "coso": "情報と伝達",
        "item": "会計方針・見積り・開示論点の部門横断共有",
        "risk": "会計基準変更、重要な見積り、開示論点、子会社情報が経理部、経営企画部、事業部門、子会社間で共有されず、開示誤りが発生する。",
        "control": "経理部は四半期ごとに財務報告連絡会を開催し、会計方針、会計上の見積り、開示スケジュール、子会社パッケージ、外部監査人指摘、内部統制不備を共有する。重要事項は取締役会・監査等委員会へ報告する。",
        "type": "予防的",
        "frequency": "四半期",
        "key": "Y",
        "owner": "経理部 / 経営企画部",
        "evidence": "財務報告連絡会_開催実績_FY2025.csv; 財務報告連絡会_議事メモ_2025年度Q4.pdf",
        "policy": "R16, R17",
        "sample": "4",
        "design_proc": "財務報告連絡会の開催目的、参加部門、議題、エスカレーション基準を確認し、財務報告上の重要情報の伝達経路が定義されていることを確認する。",
        "oper_proc": "四半期の議事メモを閲覧し、会計論点、見積り、開示、子会社情報、内部統制不備が共有され、必要事項が経営者へ報告されていることを確認する。",
    },
    {
        "id": "ELC-010",
        "coso": "モニタリング",
        "item": "内部監査による独立評価",
        "risk": "日常的モニタリングだけに依存し、財務報告内部統制の不備が独立的に識別されない。",
        "control": "内部監査室は財務報告リスク評価に基づき年次内部監査計画を策定し、ELC、PLC、ITGC、ITAC、FCRPの評価対象、実施時期、担当者を定めて整備・運用評価を実施する。発見事項はCFO、代表取締役、監査等委員会へ報告し、是正をフォローする。",
        "type": "発見的",
        "frequency": "年次 / 監査計画ベース",
        "key": "Y",
        "owner": "内部監査室",
        "evidence": "2025年度内部監査計画書.pdf",
        "policy": "R03",
        "sample": "1",
        "design_proc": "内部監査規程と年次監査計画を閲覧し、内部監査室の独立性、評価範囲、報告先、是正フォローが定義されていることを確認する。",
        "oper_proc": "FY2025内部監査計画を閲覧し、財務報告リスクに基づく評価対象、実施時期、担当者、報告先が計画されていることを確認する。",
    },
    {
        "id": "ELC-011",
        "coso": "モニタリング",
        "item": "内部統制不備の評価・是正措置モニタリング",
        "risk": "識別された不備が重要性評価、根本原因分析、是正期限、責任者管理を経ずに放置される。",
        "control": "内部監査室は内部統制不備・是正措置管理台帳を管理し、不備の影響度、発生可能性、補完統制、是正責任者、期限、完了証跡を記録する。重要な不備及び判断保留事項はCFO及び監査等委員会へ四半期報告する。",
        "type": "発見的",
        "frequency": "四半期 / 随時",
        "key": "Y",
        "owner": "内部監査室 / CFO",
        "evidence": "内部統制不備・是正措置管理台帳_FY2025.xlsx",
        "policy": "R03, R05",
        "sample": "4",
        "design_proc": "不備評価フレームワークと是正措置管理手順を閲覧し、不備区分、重要性評価、報告先、是正フォロー要件が定義されていることを確認する。",
        "oper_proc": "不備・是正措置管理台帳を閲覧し、FY2025の不備及び判断保留事項について評価、責任者、期限、ステータス、報告状況が記録されていることを確認する。",
    },
    {
        "id": "ELC-012",
        "coso": "ITへの対応",
        "item": "ITガバナンスと外部委託先管理",
        "risk": "財務報告に係るIT環境、アクセス、変更、運用、外部委託先のリスクが経営レベルで把握されず、ITGC/ITACの不備が財務報告へ波及する。",
        "control": "情報システム部は年次IT計画、情報セキュリティ方針、ITGC評価範囲、外部委託先SOC1報告書及びCUEC対応状況をレビューする。CUECごとに自社側統制、責任者、対応状況、証跡、報告記録を台帳で管理し、重要なITリスク及び委託先不備を経営会議又は取締役会へ報告する。",
        "type": "予防的",
        "frequency": "年次 / 四半期",
        "key": "Y",
        "owner": "情報システム部 / 管理本部",
        "evidence": "ITガバナンス・外部委託先レビュー_FY2025.xlsx; SOC1_TypeII_Report_SIerA_FY2024.pdf",
        "policy": "R21-R27",
        "sample": "1",
        "design_proc": "情報セキュリティ基本方針、IT関連規程、外部委託管理規程を閲覧し、IT計画、ITリスク、委託先報告書、CUEC対応のレビュー要件及び経営報告要件が定義されていることを確認する。",
        "oper_proc": "ITガバナンス・外部委託先レビュー台帳とSOC1報告書を閲覧し、財務報告関連ITリスク、SOC1例外、CUECごとの自社側統制ID・責任者・対応状況・証跡、及び経営報告ログが追跡可能に記録されていることを確認する。",
    },
]


THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9EAF7")
KEY_FILL = PatternFill("solid", fgColor="FFF2CC")
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
BODY_FONT = Font(name="Yu Gothic", size=10)
HEADER_FONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Yu Gothic", size=14, bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, quoting=csv.QUOTE_ALL)
        writer.writeheader()
        writer.writerows(rows)


def style_worksheet(ws, freeze: str | None = None) -> None:
    ws.sheet_view.zoomScale = 90
    if freeze:
        ws.freeze_panes = freeze
    for row in ws.iter_rows():
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = LEFT
            cell.border = BORDER
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, values_only=True):
            for value in cell:
                if value is not None:
                    parts = str(value).splitlines() or [""]
                    max_len = max(max_len, max(len(part) for part in parts))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 44)


def add_title(ws, title: str, subtitle: str, max_col: int) -> None:
    ws.insert_rows(1, 3)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Yu Gothic", size=9, italic=True, color="666666")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    for cell in ws[4]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
    ws.freeze_panes = "A5"


def rebuild_rcm_csv() -> None:
    rows = [
        {
            "key": c["id"],
            "procedure": c["oper_proc"],
            "control": c["control"],
            "risk": c["risk"],
            "sample_num": c["sample"],
        }
        for c in CONTROLS
    ]
    write_csv(RCM_DIR / "RCM_ELC.csv", rows, ["key", "procedure", "control", "risk", "sample_num"])


def rebuild_mapping_csv() -> None:
    rows: list[dict[str, str]] = []
    for control in CONTROLS:
        filenames = [name.strip() for name in control["evidence"].split(";")]
        for filename in filenames:
            rows.append({"key": control["id"], "sample_no": "1", "filename": filename})
    write_csv(RCM_DIR / "Evidence_Mapping_ELC.csv", rows, ["key", "sample_no", "filename"])


def rebuild_rcm_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "ELC_RCM"
    headers = [
        "統制ID",
        "基本的要素",
        "評価観点",
        "財務報告リスク",
        "統制活動",
        "統制タイプ",
        "頻度",
        "キー",
        "主担当",
        "整備状況評価手続",
        "運用状況評価手続",
        "主なEvidence",
        "関連規程",
        "整備評価",
        "運用評価",
        "不備",
        "結論",
    ]
    ws.append(headers)
    for c in CONTROLS:
        ws.append(
            [
                c["id"],
                c["coso"],
                c["item"],
                c["risk"],
                c["control"],
                c["type"],
                c["frequency"],
                c["key"],
                c["owner"],
                c["design_proc"],
                c["oper_proc"],
                c["evidence"],
                c["policy"],
                "有効",
                "有効",
                "なし",
                "整備・運用ともに有効",
            ]
        )
    style_worksheet(ws, "A2")
    add_title(
        ws,
        "【ELC】全社統制 リスク・コントロール・マトリクス（RCM）",
        "評価対象: デモA株式会社 連結ベース / 評価期間: FY2025 (2025/4/1 - 2026/3/31) / 参照: 金融庁 令和5年改訂 内部統制基準・実施基準",
        len(headers),
    )
    widths = [12, 16, 26, 44, 56, 12, 14, 8, 22, 48, 52, 42, 14, 12, 12, 10, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in range(5, ws.max_row + 1):
        ws.row_dimensions[row].height = 86
        if ws.cell(row=row, column=8).value == "Y":
            ws.cell(row=row, column=1).fill = KEY_FILL
            ws.cell(row=row, column=8).fill = KEY_FILL
        ws.cell(row=row, column=14).fill = OK_FILL
        ws.cell(row=row, column=15).fill = OK_FILL
        for col in [1, 2, 6, 7, 8, 13, 14, 15, 16]:
            ws.cell(row=row, column=col).alignment = CENTER

    ref = wb.create_sheet("参照基準")
    ref_rows = [
        ["区分", "内容"],
        ["参照資料", "金融庁・企業会計審議会「財務報告に係る内部統制の評価及び監査の基準」及び実施基準（令和5年4月7日改訂）"],
        ["URL", FSA_BASELINE_URL],
        ["設計方針", "6つの基本的要素に沿いつつ、財務報告に重要な影響を及ぼす会社レベルの方針、監督、リスク評価、情報伝達、モニタリング、IT対応に絞った。"],
        ["評価単位", "ELCは個別取引テストではなく、統制頻度に応じて年次1件、四半期4件、月次12件又は全件を確認する。"],
        ["Evidence設計", "取引単票ではなく、会議体、台帳、レビュー記録、リスク評価、規程レビュー、締めログ、委託先レビューなど、統制実施を直接示す証跡を対応付けた。"],
    ]
    for row in ref_rows:
        ref.append(row)
    style_worksheet(ref, "A2")
    ref.column_dimensions["A"].width = 18
    ref.column_dimensions["B"].width = 110

    wb.save(RCM_DIR / "ELC_RCM.xlsx")


def rebuild_board_monitoring_csv() -> None:
    rows = [
        ["会議体", "開催日", "対象期間", "主な財務報告内部統制議題", "出席状況", "決議/指示", "関連ELC"],
        ["取締役会", "2025-04-15", "FY2025計画", "内部統制基本方針、J-SOX評価計画、内部監査計画の承認", "取締役7/7、監査等委員4/4", "FY2025評価計画を承認", "ELC-001, ELC-010"],
        ["監査等委員会", "2025-05-22", "Q1", "外部監査人監査計画、内部監査室独立性、重点監査領域", "委員4/4", "販売・購買・ITGCを重点領域に指定", "ELC-001, ELC-010"],
        ["取締役会", "2025-06-30", "Q1", "全社リスクアセスメント結果、J-SOX評価範囲", "取締役7/7、監査等委員4/4", "評価範囲とキー統制を承認", "ELC-001, ELC-004"],
        ["財務報告連絡会", "2025-07-18", "Q1", "会計上の見積り、四半期開示、子会社報告パッケージ", "経理/経企/内部監査/子会社経理", "見積り根拠資料の添付標準を改訂", "ELC-007, ELC-009"],
        ["取締役会", "2025-09-25", "上期", "上期決算、リスク管理規程改訂、内部監査進捗", "取締役7/7、監査等委員4/4", "R05改訂を承認、監査頻度見直しを次年度反映", "ELC-001, ELC-004"],
        ["監査等委員会", "2025-10-24", "Q2", "内部通報1件の調査結果、内部監査中間報告", "委員4/4", "財務報告影響なし、接待規程の周知強化", "ELC-001, ELC-008"],
        ["財務報告連絡会", "2025-12-12", "Q3", "年末決算体制、開示スケジュール、ITGC進捗", "経理/経企/情シス/内部監査", "退職者アカウント停止遅延の是正状況をCFOへ報告", "ELC-009, ELC-011, ELC-012"],
        ["取締役会", "2026-01-28", "Q3", "Q3決算、内部統制不備評価の中間報告", "取締役7/7、監査等委員4/4", "重要な不備候補を監査等委員会へ継続報告", "ELC-001, ELC-011"],
        ["監査等委員会", "2026-02-20", "年度末前", "内部統制不備・判断保留事項、外部監査人コミュニケーション", "委員4/4", "追加Evidence取得期限を2026-03-15に設定", "ELC-001, ELC-011"],
        ["財務報告連絡会", "2026-03-18", "年度末", "年度末決算、見積り、開示、連結パッケージ、SOC1/CUEC", "経理/経企/情シス/内部監査/子会社経理", "未完了タスクを期末締めチェックリストへ反映", "ELC-007, ELC-009, ELC-012"],
        ["取締役会", "2026-03-25", "FY2025総括", "内部統制評価結果、是正措置、内部統制報告書ドラフト", "取締役7/7、監査等委員4/4", "内部統制報告書ドラフトをレビュー、是正計画を承認", "ELC-001, ELC-011"],
    ]
    with (ELC_DIR / "取締役会・監査等委員会_財務報告内部統制モニタリング_FY2025.csv").open(
        "w", encoding="utf-8-sig", newline=""
    ) as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def rebuild_ack_log() -> None:
    departments = [
        ("経営陣", 4, ""),
        ("内部監査室", 3, ""),
        ("経理部", 20, ""),
        ("営業本部", 45, ""),
        ("製造本部", 280, "2025-06-12"),
        ("技術本部", 60, ""),
        ("管理本部(除く経理)", 60, "2025-06-20"),
        ("品質保証部", 25, ""),
        ("情報システム部", 15, ""),
        ("経営企画部", 10, ""),
    ]
    summary_rows = [["区分", "部門", "対象者数", "受領確認済", "個票ログ件数", "未了", "完了率", "最終督促日"]]
    total = sum(count for _, count, _ in departments)
    summary_rows.append(["集計", "全社", total, total, total, 0, "100.0%", "2025-06-20"])
    for dept, count, reminder in departments:
        summary_rows.append(["部門別", dept, count, count, count, 0, "100.0%", reminder])

    detail_rows = [
        ["確認タイムスタンプ", "社員番号", "所属部門", "対象文書", "確認ステータス", "HRISログID", "備考"]
    ]
    employee_seq = 1
    for dept_idx, (dept, count, _) in enumerate(departments, start=1):
        for dept_seq in range(1, count + 1):
            day = 1 + ((employee_seq * 7 + dept_idx) % 28)
            hour = 9 + (employee_seq % 8)
            minute = (employee_seq * 11) % 60
            timestamp = f"2025-06-{day:02d} {hour:02d}:{minute:02d}"
            detail_rows.append(
                [
                    timestamp,
                    f"E{employee_seq:04d}",
                    dept,
                    "倫理綱領2025年度版",
                    "受領確認済",
                    f"ACK-2025-{employee_seq:04d}",
                    "HRIS個票ログ",
                ]
            )
            employee_seq += 1

    rows = summary_rows + [[]] + detail_rows
    with (ELC_DIR / "HRIS_CodeOfEthics_AcknowledgmentLog_FY2025.csv").open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(
            [
                ["# HRIS Code of Ethics Acknowledgment Log FY2025"],
                ["# Export: 2025-06-30 18:00 JST / Population: active officers and employees as of 2025-06-30"],
                ["# Control: ELC-002 / Completeness check: target population 522 = individual log rows 522 = acknowledged 522"],
                [],
                *rows,
            ]
        )


def rebuild_authority_review_csv() -> None:
    rows = [
        ["レビュー日", "レビュー対象", "確認観点", "結果", "対応", "承認者", "関連Evidence"],
        ["2025-04-10", "組織図・職務権限規程R18", "FY2025組織変更、管理本部/経理部/情シス部の責任分担", "改訂要", "R18を2025-04-01改訂版へ更新済", "管理本部長(CFO)", "規程_職務権限規程_R18.pdf"],
        ["2025-04-10", "財務報告責任者一覧", "CFO、経理部長、内部監査室長、情シス部長、子会社経理責任者の明確化", "有効", "company_profile.pdf 第3-1章及び改訂履歴へ反映。承認証跡を別紙化。", "CFO", "財務報告責任者一覧_承認証跡_FY2025.csv; company_profile.pdf"],
        ["2025-04-12", "承認権限金額と社員マスタ", "職位別承認限度と社員マスタの整合", "有効", "差異なし", "総務部長", "employees.xlsx"],
        ["2025-04-12", "SAPロール設計との連携", "権限規程とSAPロール付与方針の整合", "一部要フォロー", "SoD例外はITGC-AC-002/003及びELC-011で管理", "情シス部長", "user_roles_matrix.xlsx"],
    ]
    with (ELC_DIR / "職務権限・組織体制_年次レビュー_FY2025.csv").open("w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerows(rows)


def rebuild_financial_reporting_responsibility_approval_csv() -> None:
    rows = [
        ["証跡ID", "承認日", "対象文書", "改訂箇所", "改訂内容", "作成者", "レビュー者", "承認者", "反映確認"],
        ["FRR-2025-001", "2025-04-10", "company_profile.pdf", "第3-1章 財務報告責任者一覧 / 改訂履歴", "CFO、経理部長、内部監査室長、情報システム部長、子会社経理責任者の役割と報告先を明確化", "総務部長 前田 美香", "内部監査室長 長谷川 剛", "取締役CFO 渡辺 正博", "2025-04-10反映済"],
        ["FRR-2025-002", "2025-04-12", "職務権限・組織体制_年次レビュー_FY2025.csv", "レビュー結果2行目", "company_profile.pdfへの反映箇所と承認証跡IDを追記", "総務部長 前田 美香", "経理部長 佐藤 一郎", "取締役CFO 渡辺 正博", "2025-04-12反映済"],
    ]
    with (ELC_DIR / "財務報告責任者一覧_承認証跡_FY2025.csv").open("w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerows(rows)


def company_profile_markdown() -> str:
    source = ELC_DIR / "company_profile.md"
    if source.exists():
        return source.read_text(encoding="utf-8")

    text = (BASE / "0.profile" / "company_profile.md").read_text(encoding="utf-8")
    revision = """## 改訂履歴

| 版 | 改訂日 | 改訂内容 | 作成/レビュー | 承認 | 関連証跡 |
|----|--------|----------|---------------|------|----------|
| 2025.1 | 2025年4月10日 | FY2025 J-SOX評価範囲、財務報告責任者一覧、組織体制を更新 | 総務部長 / 内部監査室長 | 取締役CFO | 財務報告責任者一覧_承認証跡_FY2025.csv |

---

"""
    text = text.replace("---\n\n## 1. 会社概要", f"---\n\n{revision}## 1. 会社概要", 1)
    responsibility = """### 3-1. 財務報告責任者一覧（FY2025）

| 領域 | 責任者 | 所属/役職 | 主な責任 | 報告先 | 代替責任者 |
|------|--------|-----------|----------|--------|------------|
| 内部統制評価の統括 | 渡辺 正博 | 取締役CFO | J-SOX評価範囲、評価結果、不備評価、内部統制報告書の承認 | 代表取締役、取締役会、監査等委員会 | 経理部長 |
| 決算・財務報告 | 佐藤 一郎 | 経理部長 | 月次・四半期・年度決算、会計上の見積り、連結パッケージ、開示数値のレビュー | CFO | 経理部課長(財務会計) |
| 内部監査・独立評価 | 長谷川 剛 | 内部監査室長 | ELC、PLC、ITGC、ITAC、FCRPの整備・運用評価、不備・是正措置のフォロー | 代表取締役、監査等委員会 | 内部監査担当 |
| IT統制・外部委託先管理 | 情報システム部長 | 情報システム部 | 財務報告関連ITのアクセス、変更、運用、SOC1/CUEC対応のレビュー | CFO | アプリケーションチームリーダー |
| 子会社財務報告 | 各子会社経理責任者 | デモA東北、デモA物流、Demo-A Thailand、デモAトレーディング | 子会社決算、連結パッケージ、重要な会計論点の報告 | 経理部長 | 各子会社管理部長 |

上記の責任者一覧は、2025年4月10日に総務部長が更新し、内部監査室長のレビューを経て、取締役CFOが承認した。承認証跡は `財務報告責任者一覧_承認証跡_FY2025.csv` に保管する。

---

"""
    text = text.replace("---\n\n## 4. 主要業務プロセス（PLC対象）", f"{responsibility}## 4. 主要業務プロセス（PLC対象）", 1)
    return text


def markdown_to_pdf(md_text: str, output_path: Path) -> None:
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        rightMargin=28,
        leftMargin=28,
        topMargin=28,
        bottomMargin=28,
    )
    styles = getSampleStyleSheet()
    body = ParagraphStyle("jp-body", parent=styles["BodyText"], fontName="HeiseiKakuGo-W5", fontSize=8.5, leading=12)
    title = ParagraphStyle("jp-title", parent=body, fontSize=15, leading=20, spaceAfter=8)
    h2 = ParagraphStyle("jp-h2", parent=body, fontSize=12, leading=16, spaceBefore=8, spaceAfter=4)
    h3 = ParagraphStyle("jp-h3", parent=body, fontSize=10, leading=14, spaceBefore=6, spaceAfter=3)
    code = ParagraphStyle("jp-code", parent=body, fontName="HeiseiKakuGo-W5", fontSize=7, leading=9, leftIndent=8)

    story = []
    lines = md_text.splitlines()
    idx = 0
    in_code = False
    code_lines: list[str] = []
    paragraph_lines: list[str] = []

    def flush_paragraph() -> None:
        if not paragraph_lines:
            return
        text = "<br/>".join(line.strip() for line in paragraph_lines if line.strip())
        if text:
            story.append(Paragraph(text, body))
            story.append(Spacer(1, 3))
        paragraph_lines.clear()

    def add_table(table_lines: list[str]) -> None:
        data = []
        for raw in table_lines:
            cells = [cell.strip() for cell in raw.strip().strip("|").split("|")]
            if all(set(cell) <= {"-", ":"} for cell in cells):
                continue
            data.append([Paragraph(cell.replace("`", ""), body) for cell in cells])
        if not data:
            return
        available = A4[0] - 56
        col_width = available / max(len(data[0]), 1)
        table = Table(data, colWidths=[col_width] * len(data[0]), repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("FONT", (0, 0), (-1, -1), "HeiseiKakuGo-W5"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 5))

    while idx < len(lines):
        line = lines[idx]
        if line.startswith("```"):
            flush_paragraph()
            if in_code:
                story.append(Preformatted("\n".join(code_lines), code))
                story.append(Spacer(1, 5))
                code_lines = []
                in_code = False
            else:
                in_code = True
            idx += 1
            continue
        if in_code:
            code_lines.append(line)
            idx += 1
            continue
        if line.startswith("|") and "|" in line[1:]:
            flush_paragraph()
            table_lines = []
            while idx < len(lines) and lines[idx].startswith("|"):
                table_lines.append(lines[idx])
                idx += 1
            add_table(table_lines)
            continue
        stripped = line.strip()
        if not stripped:
            flush_paragraph()
        elif stripped == "---":
            flush_paragraph()
            story.append(Spacer(1, 8))
        elif stripped.startswith("# "):
            flush_paragraph()
            story.append(Paragraph(stripped[2:], title))
        elif stripped.startswith("## "):
            flush_paragraph()
            if story:
                story.append(PageBreak())
            story.append(Paragraph(stripped[3:], h2))
        elif stripped.startswith("### "):
            flush_paragraph()
            story.append(Paragraph(stripped[4:], h3))
        elif stripped.startswith("- "):
            paragraph_lines.append("・" + stripped[2:])
        else:
            paragraph_lines.append(stripped)
        idx += 1
    flush_paragraph()
    doc.build(story)


def rebuild_company_profile_pdf() -> None:
    output = ELC_DIR / "company_profile.pdf"
    markdown_to_pdf(company_profile_markdown(), output)
    source = ELC_DIR / "company_profile.md"
    if source.exists():
        source.unlink()


def rebuild_risk_assessment_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["項目", "内容"])
    rows = [
        ["評価名", "FY2025 財務報告に係る全社リスクアセスメント"],
        ["実施日", "2025-06-15"],
        ["実施者", "経営企画部、内部監査室、経理部、情報システム部"],
        ["承認/報告", "2025-06-30 取締役会報告、J-SOX評価範囲承認"],
        ["対象", "親会社及び連結子会社4社、重要な業務プロセス、決算・財務報告プロセス、財務報告関連IT"],
        ["結論", "売上、在庫、購買、決算、ITアクセス/変更、外部委託をFY2025の重点評価領域とする。"],
    ]
    for row in rows:
        ws.append(row)
    style_worksheet(ws, "A2")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 95

    rr = wb.create_sheet("FinancialReportingRisks")
    rr.append(["リスクID", "領域", "財務報告リスク", "影響勘定/開示", "関連アサーション", "発生可能性", "影響度", "評価", "対応統制/評価範囲", "責任部門"])
    risk_rows = [
        ["FR-001", "販売", "出荷基準の不適切適用又は押込み出荷", "売上高、売掛金", "実在性、期間帰属", "中", "大", "高", "PLC-S-001/002/006、FCRP-005", "営業本部/経理部"],
        ["FR-002", "購買", "架空発注又は承認権限逸脱による不正支出", "仕入高、買掛金、原材料", "実在性、権利義務", "中", "大", "高", "PLC-P-002/004、ITAC-004", "購買部/経理部"],
        ["FR-003", "在庫", "実地棚卸差異及び滞留在庫評価の見落とし", "棚卸資産、売上原価", "実在性、評価", "中", "大", "高", "PLC-I-001/002/005", "製造本部/経理部"],
        ["FR-004", "決算", "会計上の見積りの根拠不足又は恣意的判断", "貸倒引当金、減損、税効果", "評価、表示", "中", "大", "高", "FCRP-003、ELC-005/009", "経理部"],
        ["FR-005", "IT", "不適切なアクセス権又は変更により財務データが改竄される", "全勘定", "全般", "中", "大", "高", "ITGC-AC、ITGC-CM、ELC-012", "情報システム部"],
        ["FR-006", "連結/開示", "子会社パッケージ誤り、連結消去漏れ、開示遅延", "連結財務諸表、注記", "網羅性、表示", "中", "大", "高", "FCRP-002/005、ELC-007/009", "経理部/経営企画部"],
    ]
    for row in risk_rows:
        rr.append(row)
    style_worksheet(rr, "A2")

    fraud = wb.create_sheet("FraudRisk")
    fraud.append(["不正リスクID", "シナリオ", "動機/機会/正当化", "影響", "対応統制", "残余リスク", "監査計画への反映"])
    fraud_rows = [
        ["F-001", "売上前倒し又は押込み出荷", "業績目標未達の圧力、期末出荷判断の裁量", "売上高過大", "PLC-S-002/006、FCRP-005、ELC-009", "中", "期末カットオフと請求/出荷突合を重点化"],
        ["F-002", "発注承認権限逸脱又はキックバック", "購買担当者と仕入先の癒着機会", "仕入高/買掛金過大、資産流出", "PLC-P-002/004、内部通報ELC-008", "中", "発注承認例外3件を不備台帳でフォロー"],
        ["F-003", "棚卸差異・評価損の隠蔽", "評価損計上回避の動機、Excel見積りの裁量", "棚卸資産過大", "PLC-I-002/005、FCRP-003", "中", "棚卸差異と評価損計算の追加検証"],
        ["F-004", "手入力仕訳による見積り調整", "経営者による統制無視の機会", "期間損益操作", "FCRP-001/004、ELC-001/011", "中", "非定型仕訳と承認ログを重点確認"],
    ]
    for row in fraud_rows:
        fraud.append(row)
    style_worksheet(fraud, "A2")

    change = wb.create_sheet("ChangeTriggers")
    change.append(["発生日", "変化事象", "財務報告への影響", "再評価結果", "対応"])
    change_rows = [
        ["2025-04-01", "職務権限規程R18改訂", "承認権限とSAPロール設計に影響", "評価範囲変更なし", "ELC-003/006でレビュー"],
        ["2025-06-30", "リスク管理規程R05改訂", "全社リスク評価プロセスを更新", "評価プロセス更新", "取締役会承認済"],
        ["2025-10-01", "タイ子会社の受注増加", "連結パッケージ及び在庫評価への影響", "子会社パッケージ確認を強化", "FCRP-002/ELC-009へ反映"],
        ["2025-12-12", "退職者アカウント停止遅延", "財務データ改竄リスク", "ITGC不備として評価", "ITGC-AC-003/ELC-011で是正管理"],
    ]
    for row in change_rows:
        change.append(row)
    style_worksheet(change, "A2")
    wb.save(ELC_DIR / "全社リスクアセスメント結果_2025年度.xlsx")


def rebuild_policy_review_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "PolicyReview"
    ws.append(["規程番号", "規程名", "主管部門", "財務報告関連性", "最終レビュー日", "改訂要否", "対応/改訂内容", "承認者", "周知日"])
    rows = [
        ["R01", "取締役会規則", "総務部", "ELC-001", "2025-04-10", "不要", "FY2025運用で変更なし", "総務部長", "2025-04-15"],
        ["R02", "監査等委員会規則", "総務部", "ELC-001", "2025-04-10", "不要", "内部監査・外部監査人との連携を確認", "監査等委員会委員長", "2025-04-15"],
        ["R03", "内部監査規程", "内部監査室", "ELC-010/011", "2025-04-10", "不要", "不備フォロー報告先を確認", "代表取締役", "2025-04-15"],
        ["R04", "コンプライアンス規程", "総務部", "ELC-002/008", "2025-05-20", "不要", "内部通報教育資料と整合", "総務部長", "2025-06-01"],
        ["R05", "リスク管理規程", "経営企画部", "ELC-004/005", "2025-06-15", "要", "財務報告リスク及び不正リスク評価項目を追加", "取締役会", "2025-10-01"],
        ["R06", "内部通報規程", "総務部", "ELC-008", "2025-05-20", "不要", "財務報告影響あり案件のCFO報告基準を確認", "総務部長", "2025-06-01"],
        ["R16", "経理規程", "経理部", "ELC-007/009", "2025-07-01", "不要", "月次締め・レビュー責任と整合", "CFO", "2025-07-10"],
        ["R17", "決算業務規程", "経理部", "ELC-007/009", "2025-07-01", "要", "財務報告連絡会と見積り根拠添付ルールを追記", "CFO", "2025-08-01"],
        ["R18", "職務権限規程", "総務部", "ELC-003/006", "2025-04-01", "要", "組織改編と購買承認権限表を更新", "取締役会", "2025-04-01"],
        ["R21-R27", "IT関連規程", "情報システム部", "ELC-012", "2025-09-30", "不要", "外部委託SOC1/CUECレビュー手順と整合", "情報システム部長", "2025-10-15"],
    ]
    for row in rows:
        ws.append(row)
    style_worksheet(ws, "A2")
    wb.save(ELC_DIR / "規程・業務手順_年次レビュー台帳_FY2025.xlsx")


def fifth_business_day(year: int, month: int) -> date:
    current = date(year, month, 1)
    count = 0
    while True:
        if current.weekday() < 5:
            count += 1
            if count == 5:
                return current
        current += timedelta(days=1)


def rebuild_period_close_log() -> None:
    rows = [["対象月", "締め期限(翌月第5営業日)", "SAP締め実行日時", "ステータス", "経理部長レビュー日時", "経営会議報告日", "例外", "証跡ID"]]
    months = [(2025, m) for m in range(4, 13)] + [(2026, m) for m in range(1, 4)]
    for year, month in months:
        next_year, next_month = (year + 1, 1) if month == 12 else (year, month + 1)
        due = fifth_business_day(next_year, next_month)
        exec_dt = datetime.combine(due, datetime.min.time()).replace(hour=18)
        review_dt = exec_dt + timedelta(hours=2)
        meeting_dt = due + timedelta(days=2)
        while meeting_dt.weekday() >= 5:
            meeting_dt += timedelta(days=1)
        period = f"{year}-{month:02d}"
        rows.append(
            [
                period,
                due.isoformat(),
                exec_dt.strftime("%Y-%m-%d %H:%M"),
                "PERIOD_CLOSE_COMPLETE",
                review_dt.strftime("%Y-%m-%d %H:%M"),
                meeting_dt.isoformat(),
                "なし",
                f"FCLOSE_{year}{month:02d}",
            ]
        )
    with (ELC_DIR / "SAP_PeriodClose_JobLog_FY2025.csv").open("w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerows(
            [
                ["# SAP FI Period Close and Management Reporting Log"],
                ["# Control: ELC-007 / Population: FY2025 monthly close, 12 records"],
                [],
                *rows,
            ]
        )


def rebuild_whistleblower_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "WhistleblowingLog"
    ws.append(["通報番号", "受付日", "受付ルート", "概要", "利益相反確認", "調査責任者", "財務報告影響評価", "調査結果", "是正/周知", "監査等委員会報告日", "完了日", "ステータス"])
    ws.append(["WB-2025-001", "2025-07-15", "外部弁護士", "特定取引先への過剰接待疑い", "総務部長に利害関係なし", "総務部長 / 外部弁護士", "直接影響なし。購買不正兆候なし。", "接待規程の範囲内。ただし記録の粒度不足を確認。", "接待申請時の相手先・目的記載を再周知", "2025-09-25", "2025-08-30", "完了"])
    ws.append(["FY2025合計", "", "", "受付1件、全件調査完了", "", "", "財務報告へ直接影響する案件なし", "", "", "監査等委員会へ四半期報告済", "", ""])
    style_worksheet(ws, "A2")
    wb.save(ELC_DIR / "内部通報受付台帳_FY2025.xlsx")


def rebuild_financial_reporting_minutes() -> None:
    content = """# 財務報告連絡会 議事メモ（2025年度Q4）

| 項目 | 内容 |
|---|---|
| 開催日 | 2026年3月18日 |
| 参加者 | CFO、経理部長、経営企画部長、情報システム部長、内部監査室長、子会社経理責任者 |
| 対象統制 | ELC-007, ELC-009, ELC-012 |

## 議題と確認結果

| 議題 | 確認内容 | 結論/対応 |
|---|---|---|
| 年度末決算スケジュール | 連結パッケージ提出、SAP締め、開示ドラフト、監査人レビュー日程 | 全タスクを決算チェックリストへ登録。期限超過なし。 |
| 会計上の見積り | 貸倒引当金、棚卸評価損、減損、税効果の根拠資料添付状況 | FCRP-003の追加Evidenceを2026-03-25までに経理部が回収。 |
| 開示論点 | 重要な後発事象、関連当事者、セグメント情報、継続企業前提 | 重要な未解決論点なし。 |
| 内部統制不備 | PLC-P-002、ITGC-AC-003、PLC-I-002、判断保留3件の状況 | ELC-011の是正台帳へ集約し、監査等委員会へ報告。 |
| IT/外部委託 | SIer-A SOC1 Type II、CUEC対応、年度末変更凍結 | SOC1に重大例外なし。CUECはITGC手続と整合。 |

## エスカレーション

- CFOは、重要な不備候補及び判断保留事項を2026年3月25日取締役会へ報告する。
- 内部監査室は、追加Evidenceの入手状況を2026年3月31日までに更新する。
"""
    markdown_to_pdf(content, ELC_DIR / "財務報告連絡会_議事メモ_2025年度Q4.pdf")
    legacy_md = ELC_DIR / "財務報告連絡会_議事メモ_2025年度Q4.md"
    if legacy_md.exists():
        legacy_md.unlink()


def rebuild_financial_reporting_meeting_log() -> None:
    rows = [
        ["回次", "開催日", "対象期間", "参加部門", "主要議題", "経営者報告/エスカレーション", "関連統制"],
        ["Q1", "2025-07-18", "2025年4月-6月", "経理、経営企画、内部監査、子会社経理", "Q1決算、会計上の見積り、子会社パッケージ、内部統制評価計画", "J-SOX評価範囲を2025-06-30取締役会へ報告済", "ELC-007, ELC-009"],
        ["Q2", "2025-10-17", "2025年7月-9月", "経理、経営企画、内部監査、情報システム、子会社経理", "上期決算、リスク管理規程改訂、内部通報1件の財務報告影響", "内部通報の財務報告影響なしを監査等委員会へ報告", "ELC-008, ELC-009"],
        ["Q3", "2025-12-12", "2025年10月-12月", "経理、経営企画、内部監査、情報システム、子会社経理", "Q3決算、ITGC進捗、退職者アカウント停止遅延", "ITGC-AC-003を不備候補としてCFOへ報告", "ELC-009, ELC-011, ELC-012"],
        ["Q4", "2026-03-18", "2026年1月-3月", "CFO、経理、経営企画、情報システム、内部監査、子会社経理", "年度末決算、見積り、開示、SOC1/CUEC、内部統制不備", "2026-03-25取締役会へ不備及び判断保留事項を報告", "ELC-007, ELC-009, ELC-011, ELC-012"],
    ]
    with (ELC_DIR / "財務報告連絡会_開催実績_FY2025.csv").open("w", encoding="utf-8-sig", newline="") as f:
        csv.writer(f).writerows(rows)


def rebuild_deficiency_tracker_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "DeficiencyTracker"
    ws.append(["管理番号", "関連統制", "区分", "内容", "影響度", "補完統制", "責任者", "期限", "ステータス", "監査等委員会報告"])
    rows = [
        ["DEF-2026-001", "ITGC-AC-003", "重要な不備", "退職者5名中2名のSAP停止が11日/18日遅延", "財務データ改竄リスク", "SM20ログ確認、対象者ログインなし", "情報システム部長", "2026-03-31", "是正実施中", "2026-02-20"],
        ["DEF-2026-002", "PLC-P-002", "重要な不備の可能性", "発注承認25件中3件で承認権限超過", "不正支出リスク", "事後CFOレビュー、支払前3-way match", "購買部長", "2026-04-30", "是正計画承認済", "2026-03-25"],
        ["DEF-2026-003", "PLC-I-002", "軽微な不備", "棚卸差異850千円の原因分析・経理報告なし", "棚卸資産評価誤り", "期末棚卸再集計、経理レビュー", "製造本部長", "2026-03-31", "完了確認中", "2026-03-25"],
        ["HOLD-2026-001", "PLC-S-005", "判断保留", "売掛金年齢分析承認印PDFが低解像度", "回収可能性判断のEvidence不足", "メール承認ログを追加依頼", "経理部長", "2026-03-15", "追加入手中", "2026-02-20"],
        ["HOLD-2026-002", "ITGC-AC-002", "判断保留", "アクセス権棚卸の抽出条件・タイムスタンプ不足", "母集団完全性不足", "SUIM抽出ログ再取得", "情報システム部長", "2026-03-15", "追加入手中", "2026-02-20"],
        ["HOLD-2026-003", "FCRP-003", "判断保留", "貸倒引当金の信用情報等が未添付", "見積り根拠不足", "信用調査資料と回収交渉記録を追加依頼", "経理部長", "2026-03-25", "追加入手中", "2026-03-18"],
    ]
    for row in rows:
        ws.append(row)
    style_worksheet(ws, "A2")

    q = wb.create_sheet("QuarterlyReportLog")
    q.append(["回次", "報告日", "報告先", "報告内容", "指示/結論"])
    q_rows = [
        ["Q1", "2025-07-18", "CFO", "FY2025評価計画、現時点の不備なし", "計画通り評価を継続"],
        ["Q2", "2025-10-24", "監査等委員会", "内部通報1件、PLC-P-002暫定例外、追加Evidence依頼", "通報は財務報告影響なし。購買承認例外を重点フォロー"],
        ["Q3", "2026-01-28", "取締役会", "ITGC-AC-003、PLC-P-002、PLC-I-002の不備候補", "不備重要性評価と是正期限を更新"],
        ["Q4", "2026-03-25", "取締役会/監査等委員会", "FY2025不備評価、判断保留事項、是正計画", "ELC自体に開示すべき重要な不備なし。個別不備は是正管理を継続"],
    ]
    for row in q_rows:
        q.append(row)
    style_worksheet(q, "A2")
    wb.save(ELC_DIR / "内部統制不備・是正措置管理台帳_FY2025.xlsx")


def rebuild_it_governance_xlsx() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "ITGovernance"
    ws.append(["レビュー日", "領域", "確認事項", "結果", "対応", "報告先", "報告記録ID"])
    rows = [
        ["2025-04-20", "IT計画", "FY2025 IT投資、ERP保守、セキュリティ施策、年度末変更凍結方針", "承認済", "経営会議で四半期進捗報告", "経営会議", "RPT-IT-2025-01"],
        ["2025-06-30", "ITGC評価範囲", "SAP、WMS、連結決算、開示システム、給与SaaSの財務報告関連性", "評価範囲に反映", "ITGC/ITAC RCMへ連携", "取締役会", "RPT-IT-2025-02"],
        ["2025-09-30", "SOC1/CUEC・情報セキュリティ", "SOC1例外、CUEC割当、アクセス管理、ログ監視、特権ID、退職者ID停止", "CUEC割当済。退職者停止遅延をフォロー", "CUECMapping及びITGC-AC-003/ELC-011で管理", "CFO/監査等委員会", "RPT-IT-2025-03"],
        ["2026-03-18", "年度末変更凍結", "年度末決算期間の本番変更凍結、緊急変更時のCFO承認、CUEC対応状況の期末確認", "有効", "変更管理ログ、CUECMapping、ReportLogと照合済", "財務報告連絡会", "RPT-IT-2025-04"],
    ]
    for row in rows:
        ws.append(row)
    style_worksheet(ws, "A2")

    soc = wb.create_sheet("SOC1Review")
    soc.append(["委託先", "報告書", "対象期間", "統制目的", "SOC1例外", "CUEC総数", "CUEC割当結果", "結論", "報告記録ID"])
    soc.append(["SIer-A", "SOC1 Type II Report SIerA FY2024", "2024-04-01 - 2025-03-31", "ERP変更管理、運用、アクセス、バックアップ", "重大例外なし", "5件すべて自社側統制へ割当済", "CUECMapping参照。1件は退職者停止遅延としてELC-011/ITGC-AC-003で是正管理中", "未割当CUECなし。財務報告への未対応リスクなし", "RPT-IT-2025-03"])
    style_worksheet(soc, "A2")

    cuec = wb.create_sheet("CUECMapping")
    cuec.append([
        "CUEC ID",
        "SOC1該当箇所",
        "CUEC要求事項",
        "自社側統制ID",
        "自社側統制名",
        "統制責任者",
        "対応状況",
        "確認日",
        "確認証跡",
        "関連不備/フォロー",
        "報告記録ID",
    ])
    cuec_rows = [
        ["CUEC-01", "User Access - UCC1", "利用者登録・権限変更はユーザ企業が承認し、委託先へ正確に依頼する。", "ITGC-AC-001", "SAPユーザ登録・変更承認", "情報システム部長", "対応済", "2025-09-30", "Workflow_UserRegistration_ApprovalHistory_FY2025.csv; UserRegistration_SampleTransactionList_FY2025.xlsx", "なし", "RPT-IT-2025-03"],
        ["CUEC-02", "User Access - UCC2", "ユーザ企業は定期的に利用者権限をレビューし、不要権限を削除する。", "ITGC-AC-002; ITGC-AC-003", "四半期ユーザアクセスレビュー / 退職者ID停止", "情報システム部長", "対応中", "2025-12-12", "user_roles_matrix.xlsx; 内部統制不備・是正措置管理台帳_FY2025.xlsx", "ITGC-AC-003としてELC-011で是正管理", "RPT-IT-2025-03; RPT-IT-2025-04"],
        ["CUEC-03", "Change Management - UCC1", "本番変更はユーザ企業の変更依頼・UAT承認に基づき実施する。", "ITGC-CM-001; ITGC-CM-002", "変更要求承認 / UAT承認", "アプリケーションチームリーダー", "対応済", "2025-09-30", "change_management_log.xlsx; UAT_Approval_Evidence_FY2025.xlsx", "なし", "RPT-IT-2025-03"],
        ["CUEC-04", "Operations - UCC1", "障害通知先及びエスカレーション先をユーザ企業が維持する。", "ITGC-OM-002", "ジョブ・障害監視とエスカレーション", "インフラチームリーダー", "対応済", "2026-03-18", "Zabbix_IncidentLog_FY2025.csv; 財務報告連絡会_開催実績_FY2025.csv", "なし", "RPT-IT-2025-04"],
        ["CUEC-05", "Backup - UCC1", "バックアップ対象、保管期間、リストア確認結果をユーザ企業が確認する。", "ITGC-OM-001", "バックアップ監視・リストア確認", "インフラチームリーダー", "対応済", "2026-03-18", "DB13_BackupJobLog_FY2025.csv; DR_RestoreTest_Result_FY2025.xlsx", "なし", "RPT-IT-2025-04"],
    ]
    for row in cuec_rows:
        cuec.append(row)
    style_worksheet(cuec, "A2")

    reports = wb.create_sheet("ReportLog")
    reports.append(["報告記録ID", "報告日", "会議体/報告先", "報告者", "報告内容", "関連CUEC ID", "フォローID", "結論/指示"])
    report_rows = [
        ["RPT-IT-2025-01", "2025-04-20", "経営会議", "情報システム部長", "FY2025 IT計画、ERP保守、セキュリティ施策、年度末変更凍結方針", "N/A", "なし", "四半期進捗を経営会議へ報告する"],
        ["RPT-IT-2025-02", "2025-06-30", "取締役会", "CFO / 情報システム部長", "財務報告関連IT範囲、ITGC/ITAC評価範囲、外部委託SOC1入手予定", "N/A", "なし", "SAP、WMS、連結決算、開示システムを評価範囲に含める"],
        ["RPT-IT-2025-03", "2025-09-30", "CFO/監査等委員会", "情報システム部長", "SOC1レビュー結果、CUEC別自社側統制割当、退職者ID停止遅延", "CUEC-01,CUEC-02,CUEC-03", "ITGC-AC-003", "CUEC-02関連の退職者停止遅延をELC-011で是正管理する"],
        ["RPT-IT-2025-04", "2026-03-18", "財務報告連絡会", "情報システム部長", "年度末変更凍結、CUEC対応状況、運用・バックアップ証跡確認", "CUEC-02,CUEC-04,CUEC-05", "ITGC-AC-003", "未割当CUECなし。ITGC-AC-003はQ4不備評価へ反映する"],
    ]
    for row in report_rows:
        reports.append(row)
    style_worksheet(reports, "A2")
    wb.save(ELC_DIR / "ITガバナンス・外部委託先レビュー_FY2025.xlsx")


def rebuild_internal_audit_plan_pdf() -> None:
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    path = ELC_DIR / "2025年度内部監査計画書.pdf"
    doc = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("jp-title", parent=styles["Title"], fontName="HeiseiKakuGo-W5", fontSize=16, leading=20)
    heading = ParagraphStyle("jp-heading", parent=styles["Heading2"], fontName="HeiseiKakuGo-W5", fontSize=12, leading=16, spaceBefore=8)
    body = ParagraphStyle("jp-body", parent=styles["BodyText"], fontName="HeiseiKakuGo-W5", fontSize=9, leading=13)
    table_body = ParagraphStyle("jp-table", parent=body, fontSize=7.2, leading=9)

    def cell(text: str) -> Paragraph:
        return Paragraph(str(text).replace("\n", "<br/>"), table_body)

    story = [
        Paragraph("2025年度 内部監査計画書", title),
        Paragraph("作成: デモA株式会社 内部監査室 / 作成日: 2025年4月10日 / 承認: 取締役会 2025年4月15日", body),
        Spacer(1, 10),
        Paragraph("1. 監査目的", heading),
        Paragraph("金融商品取引法に基づく財務報告に係る内部統制の整備状況及び運用状況について、全社統制、業務プロセス統制、IT統制、決算・財務報告プロセスをリスクベースで評価する。", body),
        Paragraph("2. 評価範囲", heading),
    ]
    scope_data = [
        ["区分", "対象"],
        ["ELC", "12統制（6つの基本的要素に対応）"],
        ["PLC", "販売・購買・在庫/原価の21統制"],
        ["ITGC/ITAC", "SAP、WMS、連結決算、開示システム等"],
        ["FCRP", "月次/四半期/年度決算、連結、見積り、開示"],
    ]
    table = Table(scope_data, colWidths=[90, 360])
    table.setStyle(TableStyle([("FONT", (0, 0), (-1, -1), "HeiseiKakuGo-W5"), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.5, colors.grey), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story.extend([table, Spacer(1, 10), Paragraph("3. 実施時期・担当者計画", heading)])
    schedule_data = [
        [cell("対象領域"), cell("実施時期"), cell("主担当者"), cell("レビュー者"), cell("主な手続/成果物")],
        [cell("ELC"), cell("2025年6月-7月"), cell("内部監査担当 大塚 美穂 (IA002)"), cell("内部監査室長 長谷川 剛 (IA001)"), cell("全社統制12項目の整備評価、運用証跡入手、取締役会・監査等委員会報告状況の確認")],
        [cell("PLC 販売"), cell("2025年8月-10月"), cell("内部監査担当 山口 健 (IA003)"), cell("内部監査室長 長谷川 剛 (IA001)"), cell("販売プロセスRCM、受注・出荷・請求・入金サンプル、例外一覧")],
        [cell("PLC 購買"), cell("2025年9月-11月"), cell("内部監査担当 大塚 美穂 (IA002)"), cell("内部監査室長 長谷川 剛 (IA001)"), cell("購買承認、発注・検収・請求照合、PLC-P-002例外フォロー")],
        [cell("PLC 在庫/原価"), cell("2025年9月-12月"), cell("内部監査担当 山口 健 (IA003)"), cell("内部監査室長 長谷川 剛 (IA001)"), cell("実地棚卸、棚卸差異、標準原価改訂、評価損計算の検証")],
        [cell("ITGC/ITAC"), cell("2025年8月-2026年1月"), cell("内部監査担当 大塚 美穂 (IA002)"), cell("内部監査室長 長谷川 剛 (IA001)"), cell("アクセス、変更、運用、SOC1/CUEC対応、ITAC自動統制の整備・運用評価")],
        [cell("FCRP"), cell("2025年12月-2026年3月"), cell("内部監査担当 山口 健 (IA003)"), cell("内部監査室長 長谷川 剛 (IA001)"), cell("月次・四半期・年度決算、連結、見積り、開示チェックリストの検証")],
        [cell("不備評価・是正フォロー"), cell("2026年1月-3月"), cell("内部監査室長 長谷川 剛 (IA001)"), cell("CFO 渡辺 正博"), cell("不備重要性評価、是正責任者・期限確認、CFO/監査等委員会/取締役会への報告")],
    ]
    schedule_table = Table(schedule_data, colWidths=[58, 92, 100, 100, 116], repeatRows=1)
    schedule_table.setStyle(TableStyle([("FONT", (0, 0), (-1, -1), "HeiseiKakuGo-W5"), ("FONTSIZE", (0, 0), (-1, -1), 7.2), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.4, colors.grey), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3)]))
    story.extend([schedule_table, Spacer(1, 10), Paragraph("4. ELC重点評価項目", heading)])
    elc_data = [["統制ID", "重点評価項目", "主なEvidence"]]
    for control in CONTROLS:
        elc_data.append([control["id"], control["item"], control["evidence"].split(";")[0]])
    elc_table = Table(elc_data, colWidths=[55, 155, 240], repeatRows=1)
    elc_table.setStyle(TableStyle([("FONT", (0, 0), (-1, -1), "HeiseiKakuGo-W5"), ("FONTSIZE", (0, 0), (-1, -1), 7.5), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.4, colors.grey), ("VALIGN", (0, 0), (-1, -1), "TOP")]))
    story.extend([elc_table, Spacer(1, 10), Paragraph("5. 実施体制・報告", heading)])
    team_data = [
        [cell("役割"), cell("担当者"), cell("責任")],
        [cell("統括/最終レビュー"), cell("内部監査室長 長谷川 剛 (IA001)"), cell("内部監査計画の統括、調書レビュー、不備重要性評価、CFO/代表取締役/監査等委員会への報告")],
        [cell("ELC/購買/IT担当"), cell("内部監査担当 大塚 美穂 (IA002)"), cell("ELC、購買、ITGC/ITAC、SOC1/CUEC対応の評価実施")],
        [cell("販売/在庫/FCRP担当"), cell("内部監査担当 山口 健 (IA003)"), cell("販売、在庫/原価、決算・財務報告プロセスの評価実施")],
    ]
    team_table = Table(team_data, colWidths=[95, 145, 226], repeatRows=1)
    team_table.setStyle(TableStyle([("FONT", (0, 0), (-1, -1), "HeiseiKakuGo-W5"), ("FONTSIZE", (0, 0), (-1, -1), 7.2), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("GRID", (0, 0), (-1, -1), 0.4, colors.grey), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 3), ("RIGHTPADDING", (0, 0), (-1, -1), 3)]))
    story.extend([team_table, Spacer(1, 6)])
    story.append(Paragraph("評価結果及び不備・是正措置はCFO、代表取締役、監査等委員会、取締役会へ報告する。重要な不備候補及び判断保留事項は四半期ごとに更新する。", body))
    doc.build(story)


def rebuild_audit_procedure_md() -> None:
    rows = "\n".join(
        f"| {c['id']} | {c['coso']} | {c['item']} | {c['sample']} | {c['evidence']} |" for c in CONTROLS
    )
    procedures = "\n".join(
        f"""### {c['id']} {c['item']}

**整備状況評価**
: {c['design_proc']}

**運用状況評価**
: {c['oper_proc']}

**判定観点**
: 統制頻度、実施者の権限、レビュー証跡、例外時対応、経営者又は監査等委員会への報告がRCMの統制活動と整合していることを確認する。
"""
        for c in CONTROLS
    )
    content = f"""# ELC 全社統制 監査手続書

**対象RCM**: [../2.RCM/ELC_RCM.xlsx](../2.RCM/ELC_RCM.xlsx)
**評価期間**: FY2025（2025年4月1日 - 2026年3月31日）
**作成者**: 内部監査室
**作成日**: 2026年4月10日

## 1. 評価方針

本手続は、金融庁・企業会計審議会が公表した令和5年改訂の「財務報告に係る内部統制の評価及び監査の基準」及び実施基準を参照し、6つの基本的要素に沿って、財務報告に重要な影響を及ぼす会社レベルの統制を評価する。

- 参照: [金融庁 令和5年4月7日公表ページ]({FSA_RELEASE_URL})
- 参照: [内部統制基準・実施基準 PDF]({FSA_BASELINE_URL})
- 参照: [内部統制報告制度Q&A等の改訂]({FSA_QA_URL})

ELCは個別取引の承認サンプルを多数抽出する統制ではなく、経営監督、方針、リスク評価、情報伝達、モニタリング、ITガバナンスが有効に設計・運用されているかを、統制頻度に応じた証跡で確認する。年次統制は1件、四半期統制は4件、月次統制は12件又は母集団全件を基本とする。

## 2. 対象統制とEvidence

| 統制ID | 基本的要素 | 評価観点 | サンプル数 | 主なEvidence |
|---|---|---|---:|---|
{rows}

## 3. 共通手続

1. RCM上の統制活動、リスク、頻度、主担当、Evidenceが実態と整合していることを確認する。
2. Evidenceは被評価部門から入手した原資料又はシステム出力であることを確認し、作成者、作成日、対象期間、母集団、承認/レビュー証跡を確認する。
3. 経営者レビュー又は監査等委員会報告が必要な統制では、議事録、台帳、報告記録により報告先と対応状況を確認する。
4. 例外又は不備がある場合は、[内部統制不備・是正措置管理台帳_FY2025.xlsx](../4.evidence/ELC/内部統制不備・是正措置管理台帳_FY2025.xlsx)に集約し、影響度、補完統制、是正期限、責任者を評価する。

## 4. 個別評価手続

{procedures}

## 5. 評価結果の取りまとめ

全12統制について、整備状況及び運用状況の評価結果をRCMへ反映する。FY2025のELC自体には開示すべき重要な不備は識別していない。ただし、PLC-P-002、ITGC-AC-003、PLC-I-002及び判断保留事項はELC-011のモニタリング対象として継続管理する。
"""
    (PROC_DIR / "01_ELC_監査手続.md").write_text(content, encoding="utf-8")


def remove_non_elc_duplicate_evidence() -> None:
    for filename in ["発注書_PO-2025-0234.pdf", "発注書_PO-2025-0789.pdf", "発注書_PO-2025-1456.pdf"]:
        path = ELC_DIR / filename
        if path.exists():
            path.unlink()


def main() -> None:
    RCM_DIR.mkdir(parents=True, exist_ok=True)
    PROC_DIR.mkdir(parents=True, exist_ok=True)
    ELC_DIR.mkdir(parents=True, exist_ok=True)
    rebuild_rcm_csv()
    rebuild_mapping_csv()
    rebuild_rcm_xlsx()
    rebuild_board_monitoring_csv()
    rebuild_ack_log()
    rebuild_authority_review_csv()
    rebuild_financial_reporting_responsibility_approval_csv()
    rebuild_company_profile_pdf()
    rebuild_risk_assessment_xlsx()
    rebuild_policy_review_xlsx()
    rebuild_period_close_log()
    rebuild_whistleblower_xlsx()
    rebuild_financial_reporting_meeting_log()
    rebuild_financial_reporting_minutes()
    rebuild_deficiency_tracker_xlsx()
    rebuild_it_governance_xlsx()
    rebuild_internal_audit_plan_pdf()
    rebuild_audit_procedure_md()
    remove_non_elc_duplicate_evidence()
    print("Rebuilt ELC RCM, audit procedures, mapping, and evidence set.")


if __name__ == "__main__":
    main()
