"""
PLC-S-002 出荷-売上マッチング 25件サンプル対応エビデンス 試作
新方針：監査人が抽出した25件について、被評価部門（経理部）が提供する詳細トレース資料
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
from pathlib import Path
import random

BASE = Path(r"C:\Users\nyham\work\demo_data\4.evidence\PLC-S")

HF = PatternFill("solid", fgColor="1F4E78")
HF2 = PatternFill("solid", fgColor="305496")
HFONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
BFONT = Font(name="Yu Gothic", size=10)
BBOLD = Font(name="Yu Gothic", size=10, bold=True)
C_ = Alignment(horizontal="center", vertical="center", wrap_text=True)
L_ = Alignment(horizontal="left", vertical="center", wrap_text=True)
R_ = Alignment(horizontal="right", vertical="center")
T_ = Side("thin", color="888888")
BRD = Border(left=T_, right=T_, top=T_, bottom=T_)
FILL_META = PatternFill("solid", fgColor="D9E1F2")
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")  # 例外
FILL_OK = PatternFill("solid", fgColor="E2EFDA")

# 顧客マスタから抜粋
CUSTOMERS = [
    ("C-10001", "サンプル顧客A社"), ("C-10002", "サンプル顧客B社"),
    ("C-10003", "サンプル顧客C社"), ("C-10004", "サンプル顧客D社"),
    ("C-10005", "サンプル顧客E社"), ("C-10006", "サンプル顧客F社"),
    ("C-10011", "サンプル顧客H社"), ("C-10012", "サンプル顧客I社"),
    ("C-10013", "サンプル顧客J社"), ("C-10014", "サンプル顧客K社"),
    ("C-10015", "サンプル顧客L社"), ("C-10021", "サンプル顧客P社"),
    ("C-10022", "サンプル顧客Q社"), ("C-10023", "サンプル顧客R社"),
]

PRODUCTS = [
    ("P-30001", "エンジンピストンピン A型", 4200, "個"),
    ("P-30006", "トランスミッションシャフト", 12500, "個"),
    ("P-30011", "ウェハー搬送ロボット用シャフト A", 18500, "個"),
    ("P-30014", "エッチング装置チャンバ部品", 38500, "個"),
    ("P-30015", "ウェハーチャックベース", 12800, "個"),
    ("P-30020", "検査装置ステージベース", 42500, "個"),
    ("P-30022", "サスペンションアーム A型", 1280, "個"),
    ("P-30027", "ロボットアーム外装パネル A", 4200, "個"),
]


def gen():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "25件サンプル対応"
    ws.sheet_view.zoomScale = 90

    # ========== メタ情報エリア ==========
    ws.cell(row=1, column=1, value="【PLC-S-002】 出荷-売上マッチング 25件サンプル対応エビデンス")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)

    ws.cell(row=2, column=1, value="（内部監査室からの25件サンプル個別トレース要求に対する経理部からの提出書類）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=10, italic=True, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=16)

    meta = [
        ("提出日", "2026年2月15日"),
        ("提出元", "経理部 中村 真理（ACC004 主任）"),
        ("承認者", "経理部 高橋 美咲（ACC002 課長）"),
        ("提出先", "内部監査室 長谷川 剛（IA001 室長）"),
        ("要求根拠", "監査依頼書 IA-REQ-2026-002 『PLC-S-002 運用状況評価のための25件個別エビデンス要求』"),
        ("対象統制", "PLC-S-002 出荷-売上マッチング（日次自動マッチ＋経理部担当による日次未マッチレビュー）"),
        ("母集団", "FY2025 出荷実績 3,158件（SAP VA05 + WMS出荷実績データ）"),
        ("抽出方法", "系統抽出 / 間隔126件 / 開始位置57（監査人による無作為決定）"),
        ("抽出日時", "2026/02/10 11:15（内部監査室にて実行）"),
        ("本書類の位置づけ", "監査人が抽出した25件サンプルについて、経理部が該当取引のSAP/WMS記録から個別詳細を抽出し提示する。マッチング実行タイムスタンプ、差異発生時の是正記録等を網羅。"),
    ]
    for i, (k, v) in enumerate(meta):
        r = 4 + i
        ws.cell(row=r, column=1, value=k).font = BBOLD
        ws.cell(row=r, column=1).fill = FILL_META
        ws.cell(row=r, column=1).border = BRD
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.cell(row=r, column=4, value=v).font = BFONT
        ws.cell(row=r, column=4).border = BRD
        ws.cell(row=r, column=4).alignment = L_
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=16)

    # ========== 明細ヘッダ ==========
    headers = [
        "サンプル\n№", "出荷番号", "出荷日", "受注番号", "顧客\nコード", "顧客名",
        "製品コード", "数量", "出荷金額\n(円)", "売上仕訳\n番号", "売上計上日", "売上金額\n(円)",
        "金額差異\n(円)", "マッチング\n判定", "マッチ実行\nタイムスタンプ", "例外対応記録"
    ]
    hr = 16
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[hr].height = 38

    # ========== 25件サンプル ==========
    random.seed(20252002)

    # FY2025の月を均等に分散
    # 25件 ≈ 12ヶ月、2件/月ペース + 一部は3件/月
    samples_config = [
        # (sample_no, year, month, day_hint, exception_type)
        (1, 2025, 4, 8, None), (2, 2025, 4, 23, None),
        (3, 2025, 5, 12, None), (4, 2025, 5, 27, None),
        (5, 2025, 6, 10, None), (6, 2025, 6, 25, None),
        (7, 2025, 7, 9, None), (8, 2025, 7, 24, None),
        (9, 2025, 8, 11, "quantity_correction"),  # 例外1: 数量訂正
        (10, 2025, 8, 26, None),
        (11, 2025, 9, 13, None), (12, 2025, 9, 28, None),
        (13, 2025, 10, 14, None),
        (14, 2025, 10, 29, "timing_delay"),  # 例外2: 売上計上1日遅延
        (15, 2025, 11, 6, None),
        (16, 2025, 11, 17, "amount_discount"),  # 例外3: ¥50,000値引調整（既存の11月分）
        (17, 2025, 11, 28, None),
        (18, 2025, 12, 12, None), (19, 2025, 12, 24, None),
        (20, 2026, 1, 14, None), (21, 2026, 1, 29, None),
        (22, 2026, 2, 10, None), (23, 2026, 2, 25, None),
        (24, 2026, 3, 10, None), (25, 2026, 3, 24, None),
    ]

    data_rows = []
    for sn, yr, mo, day, exc in samples_config:
        cid, cname = random.choice(CUSTOMERS)
        pcode, pname, unit_price, pu = random.choice(PRODUCTS)
        qty = random.choice([50, 100, 200, 300, 500, 800, 1000, 1500])
        ship_amount = qty * unit_price
        ship_dt = date(yr, mo, day)
        # 売上計上は通常同日または翌営業日
        sale_dt = ship_dt + timedelta(days=0)

        # 受注番号は出荷日の数週間前
        ord_no = f"ORD-2025-{random.randint(100, 3200):04d}"
        ship_no = f"SH-{yr}{mo:02d}-{random.randint(20, 290):04d}"
        jv_no = f"JV-{yr}{mo:02d}-{random.randint(100, 290):04d}"

        # デフォルト
        sale_amount = ship_amount
        diff = 0
        judgment = "SAP自動マッチ成功"
        # マッチングTS: 夜間バッチ 0:30-2:30頃
        match_ts = datetime(yr, mo, day) + timedelta(days=1) + timedelta(
            hours=random.randint(0, 2), minutes=random.randint(5, 55),
            seconds=random.randint(0, 59))
        exception_note = "N/A"

        # 例外処理
        if exc == "quantity_correction":
            # 出荷時に数量訂正 → 売上も訂正
            original_qty = qty + 2
            sale_amount = ship_amount  # 訂正後は一致
            judgment = "例外対応（数量訂正）"
            exception_note = (f"出荷直後に{original_qty}個→{qty}個への数量訂正。"
                              f"経理部担当中村主任が検知（{ship_dt + timedelta(days=1)}）し、"
                              f"SAP出荷伝票と売上仕訳の両方を訂正。訂正後の金額で一致を確認。")
        elif exc == "timing_delay":
            # 売上計上が翌日
            sale_dt = ship_dt + timedelta(days=1)
            match_ts = datetime(sale_dt.year, sale_dt.month, sale_dt.day) + timedelta(
                days=1, hours=1, minutes=random.randint(10, 50))
            judgment = "例外対応（計上日1日遅延）"
            exception_note = (f"出荷{ship_dt}に対し売上計上{sale_dt}（翌営業日）。"
                              f"SAPバッチの処理タイミング差。マッチング自体は翌日成功、"
                              f"期間帰属の問題なし（同一月内）。経理部主任の確認済。")
        elif exc == "amount_discount":
            # ¥50,000値引調整（既存の11月分と整合）
            ship_no = "SH-202511-0234"
            ord_no = "ORD-2025-2468"
            cid, cname = ("C-10003", "サンプル顧客C社")
            ship_amount = 12_850_000
            sale_amount = 12_800_000
            diff = -50_000
            jv_no = "JV-202511-0234"
            judgment = "例外対応（値引調整）"
            exception_note = (
                "出荷後の値引合意により売上計上額を¥50,000減額。"
                "2025/11/20に値引伝票（DR-202511-0012）で調整済み。"
                "経理部主任中村が検知、高橋課長承認のうえ是正仕訳計上。"
                "PLC-S-002_SAP未マッチ明細リスト_202511.csvにも記録あり。")

        data_rows.append([
            sn, ship_no, ship_dt, ord_no, cid, cname, pcode, f"{qty:,} {pu}",
            ship_amount, jv_no, sale_dt, sale_amount, diff, judgment,
            match_ts.strftime("%Y-%m-%d %H:%M:%S"), exception_note
        ])

    # 明細書き込み
    for idx, row in enumerate(data_rows):
        r = hr + 1 + idx
        is_exception = row[13] != "SAP自動マッチ成功"
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 5, 7, 8, 10, 11, 14, 15):
                cell.alignment = C_
                if c_i in (3, 11):
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (9, 12, 13):
                cell.alignment = R_
                cell.number_format = "#,##0;[Red]-#,##0"
            else:
                cell.alignment = L_
        # 例外行は黄色、判定と対応記録列は強調
        if is_exception:
            for c_i in (14, 16):
                ws.cell(row=r, column=c_i).fill = FILL_WARN
            if row[12] != 0:
                ws.cell(row=r, column=13).fill = FILL_WARN
        else:
            ws.cell(row=r, column=14).fill = FILL_OK
        ws.row_dimensions[r].height = 28

    # ========== 集計 ==========
    sum_r = hr + 26 + 1
    ws.cell(row=sum_r, column=1, value="■ 集計").font = BBOLD
    ws.cell(row=sum_r, column=1).fill = FILL_META
    ws.merge_cells(start_row=sum_r, start_column=1, end_row=sum_r, end_column=3)

    normal_count = sum(1 for r in data_rows if r[13] == "SAP自動マッチ成功")
    exception_count = 25 - normal_count
    ws.cell(row=sum_r, column=4, value=f"自動マッチ成功: {normal_count}件 / 例外対応（すべて是正済）: {exception_count}件").font = BFONT
    ws.merge_cells(start_row=sum_r, start_column=4, end_row=sum_r, end_column=16)

    # ========== 例外ケース詳細補足 ==========
    detail_r = sum_r + 2
    ws.cell(row=detail_r, column=1, value="■ 例外ケース補足説明（参照可能な関連エビデンス）").font = BBOLD
    ws.cell(row=detail_r, column=1).fill = FILL_META
    ws.merge_cells(start_row=detail_r, start_column=1, end_row=detail_r, end_column=16)

    exc_headers = ["サンプル№", "例外種別", "金額影響", "発見者", "是正者", "是正日", "関連エビデンス"]
    for i, h in enumerate(exc_headers, 1):
        c = ws.cell(row=detail_r + 1, column=i, value=h)
        c.fill = HF2; c.font = HFONT; c.alignment = C_; c.border = BRD

    exc_details = [
        ("9", "数量訂正", "金額影響なし（訂正後一致）", "中村 真理（ACC004）",
         "中村 真理", "2025/8/12",
         "PLC-S-002_SAP売上計上明細_202511.csv の訂正仕訳記録 / SAP VA02 訂正履歴"),
        ("14", "売上計上1日遅延", "金額影響なし（期間帰属OK）", "中村 真理",
         "SAP翌日バッチで自動計上", "2025/10/30",
         "PLC-S-002_SAP売上計上明細_202511.csv / 日次未マッチ明細（翌日解消）"),
        ("16", "値引調整", "¥-50,000（是正仕訳済）", "中村 真理",
         "高橋 美咲承認のうえ中村", "2025/11/20",
         "PLC-S-002_SAP未マッチ明細リスト_202511.csv / 値引伝票DR-202511-0012"),
    ]
    for i, row in enumerate(exc_details):
        r = detail_r + 2 + i
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 6):
                cell.alignment = C_
            else:
                cell.alignment = L_
        ws.row_dimensions[r].height = 35

    # ========== 作成/承認欄 ==========
    sign_r = detail_r + 2 + len(exc_details) + 2
    ws.cell(row=sign_r, column=1, value="■ 本書類の作成・承認").font = BBOLD
    ws.cell(row=sign_r, column=1).fill = FILL_META
    ws.merge_cells(start_row=sign_r, start_column=1, end_row=sign_r, end_column=16)

    ws.cell(row=sign_r + 1, column=1, value="作成者: 中村 真理（経理部主任 ACC004）[印] 2026/2/15")
    ws.cell(row=sign_r + 1, column=1).font = BFONT
    ws.merge_cells(start_row=sign_r + 1, start_column=1, end_row=sign_r + 1, end_column=16)

    ws.cell(row=sign_r + 2, column=1, value="承認者: 高橋 美咲（経理部課長 ACC002）[印] 2026/2/15")
    ws.cell(row=sign_r + 2, column=1).font = BFONT
    ws.merge_cells(start_row=sign_r + 2, start_column=1, end_row=sign_r + 2, end_column=16)

    ws.cell(row=sign_r + 3, column=1,
            value="※ 上記25件の詳細データは、SAP S/4HANA販売管理モジュール、WMS出荷実績、"
                  "SAP FI売上計上仕訳から経理部が抽出・整理したものである。"
                  "内部監査室の要求日（2026/2/10）時点のシステム状態に基づく。")
    ws.cell(row=sign_r + 3, column=1).font = Font(name="Yu Gothic", size=9, italic=True, color="555555")
    ws.merge_cells(start_row=sign_r + 3, start_column=1, end_row=sign_r + 3, end_column=16)

    # 列幅
    widths = [6, 16, 11, 15, 10, 16, 12, 10, 14, 16, 11, 14, 12, 18, 19, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hr + 1}"

    out = BASE / "PLC-S-002_25件サンプル対応エビデンス.xlsx"
    wb.save(out)
    print(f"Created: {out.name}")


if __name__ == "__main__":
    gen()
