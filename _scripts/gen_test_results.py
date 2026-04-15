"""
5.test_results/ 配下の監査人の調書類を生成
（エビデンスではない：監査人が評価を実施した結果の記録）
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
from pathlib import Path
import random

BASE = Path(r"C:\Users\nyham\work\demo_data\5.test_results")
BASE.mkdir(parents=True, exist_ok=True)

HF = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
BFONT = Font(name="Yu Gothic", size=10)
BBOLD = Font(name="Yu Gothic", size=10, bold=True)
C_ = Alignment(horizontal="center", vertical="center", wrap_text=True)
L_ = Alignment(horizontal="left", vertical="center", wrap_text=True)
R_ = Alignment(horizontal="right", vertical="center")
T_ = Side("thin", color="888888")
BRD = Border(left=T_, right=T_, top=T_, bottom=T_)
FILL_OK = PatternFill("solid", fgColor="E2EFDA")
FILL_NG = PatternFill("solid", fgColor="FCE4D6")
FILL_HOLD = PatternFill("solid", fgColor="DEEBF7")
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")

CUSTOMERS = {
    "C-10001": "サンプル顧客A社", "C-10002": "サンプル顧客B社", "C-10003": "サンプル顧客C社",
    "C-10004": "サンプル顧客D社", "C-10005": "サンプル顧客E社", "C-10006": "サンプル顧客F社",
    "C-10007": "サンプル顧客G社", "C-10011": "サンプル顧客H社", "C-10012": "サンプル顧客I社",
    "C-10013": "サンプル顧客J社", "C-10014": "サンプル顧客K社", "C-10015": "サンプル顧客L社",
    "C-10016": "サンプル顧客M社", "C-10017": "サンプル顧客N社", "C-10018": "サンプル顧客O社",
    "C-10021": "サンプル顧客P社", "C-10022": "サンプル顧客Q社", "C-10023": "サンプル顧客R社",
    "C-10024": "サンプル顧客S社", "C-10025": "サンプル顧客T社",
}

CREDITS = {
    "C-10001": 500_000_000, "C-10002": 300_000_000, "C-10003": 200_000_000,
    "C-10004": 150_000_000, "C-10005": 100_000_000, "C-10006": 80_000_000,
    "C-10007": 50_000_000, "C-10011": 400_000_000, "C-10012": 250_000_000,
    "C-10013": 180_000_000, "C-10014": 120_000_000, "C-10015": 100_000_000,
    "C-10016": 80_000_000, "C-10017": 60_000_000, "C-10018": 40_000_000,
    "C-10021": 200_000_000, "C-10022": 150_000_000, "C-10023": 80_000_000,
    "C-10024": 50_000_000, "C-10025": 30_000_000,
}


# ============================================================
# PLC-S-001 受注承認 25件サンプル判定シート（監査人の調書）
# ============================================================
def gen_plc_s_001_judgment():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PLC-S-001サンプル25件判定"

    ws.cell(row=1, column=1, value="【PLC-S-001 運用評価】 受注・与信承認 サンプルテスト結果（監査調書）")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

    meta = [
        ("対象統制", "PLC-S-001 受注・与信承認"),
        ("評価対象期間", "FY2025 (2025/4/1 - 2026/3/31)"),
        ("母集団", "SAP VA05 受注伝票一覧 3,247件"),
        ("参照エビデンス", "4.evidence/PLC-S/PLC-S-001_SAP_VA05_受注伝票一覧_FY2025.xlsx"),
        ("抽出方法", "系統抽出（間隔130件、開始位置57）"),
        ("サンプル数", "25件"),
        ("テスト実施者", "長谷川 剛（IA001 内部監査室）"),
        ("テスト期間", "2026/2/10 〜 2026/2/13"),
        ("レビュー者", "大塚 美穂（IA002）"),
    ]
    for i, (k, v) in enumerate(meta):
        ws.cell(row=2 + i, column=1, value=k)
        ws.cell(row=2 + i, column=1).font = BBOLD
        ws.cell(row=2 + i, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(row=2 + i, column=1).border = BRD
        ws.cell(row=2 + i, column=2, value=v)
        ws.cell(row=2 + i, column=2).font = BFONT
        ws.cell(row=2 + i, column=2).border = BRD
        ws.merge_cells(start_row=2 + i, start_column=2, end_row=2 + i, end_column=11)

    headers = ["№", "受注番号", "受注日", "顧客コード", "顧客名", "受注金額(円)",
               "営業担当", "与信限度\nチェック", "与信超過時\n承認", "判定", "備考"]
    hr = 12
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[hr].height = 32

    random.seed(1001)
    cids = list(CUSTOMERS.keys())
    reps = ["斎藤 次郎", "藤田 修", "松本 香織", "井上 大輔"]
    samples = []
    for i in range(1, 26):
        cid = random.choice(cids)
        month = random.choice([5, 6, 7, 8, 9, 10, 11, 12, 1, 2])
        year = 2026 if month <= 3 else 2025
        day = random.randint(1, 28)
        order_date = date(year, month, day)
        order_no = f"ORD-2025-{100 + i * 130:04d}"
        amount = random.choice([
            random.randint(500_000, 5_000_000),
            random.randint(5_000_000, 30_000_000),
            random.randint(30_000_000, 80_000_000),
        ])
        rep = random.choice(reps)
        samples.append({
            "n": i, "order_no": order_no, "date": order_date,
            "cid": cid, "cname": CUSTOMERS[cid], "amount": amount, "rep": rep,
        })

    rows_out = []
    for s in samples:
        credit_ok = "○（自動チェック通過）"
        approval = "（不要／限度内）"
        result = "合格"
        remark = ""
        if s["n"] in (3, 11, 19):
            credit_ok = "超過検知"
            approval = f"○ 営業本部長承認済\n(2025/{s['date'].month:02d}/{s['date'].day:02d})"
            remark = "与信限度超過のためワークフロー承認を確認"
        if s["n"] == 14:
            credit_ok = "超過検知"
            s["amount"] = 15_200_000
            approval = "△ 承認日が1日遅れ\n(受注日+1日)"
            result = "軽微例外\n(許容)"
            remark = "承認者出張中、翌日復帰後承認。業務影響なし。"
        rows_out.append({**s, "credit_ok": credit_ok, "approval": approval,
                         "result": result, "remark": remark})

    for idx, s in enumerate(rows_out):
        r = hr + 1 + idx
        row_data = [
            s["n"], s["order_no"], s["date"], s["cid"], s["cname"],
            s["amount"], s["rep"], s["credit_ok"], s["approval"],
            s["result"], s["remark"]
        ]
        for c_i, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 7, 8, 9, 10):
                cell.alignment = C_
            elif c_i == 6:
                cell.alignment = R_
                cell.number_format = "#,##0"
            elif c_i == 3:
                cell.number_format = "yyyy/mm/dd"
            else:
                cell.alignment = L_
        ws.row_dimensions[r].height = 35

        if s["result"] == "合格":
            ws.cell(row=r, column=10).fill = FILL_OK
        elif "例外" in s["result"]:
            for c_i in (8, 9, 10):
                ws.cell(row=r, column=c_i).fill = FILL_WARN

    sum_row = hr + 26 + 1
    ws.cell(row=sum_row, column=1, value="集計")
    ws.cell(row=sum_row, column=1).font = BBOLD
    ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=3)
    pass_count = sum(1 for s in rows_out if s["result"] == "合格")
    exc_count = sum(1 for s in rows_out if "例外" in s["result"])
    ws.cell(row=sum_row, column=4, value=f"合格: {pass_count}件 / 不合格: 0件 / 軽微例外: {exc_count}件（許容）")
    ws.cell(row=sum_row, column=4).font = BFONT
    ws.merge_cells(start_row=sum_row, start_column=4, end_row=sum_row, end_column=11)

    con_row = sum_row + 2
    ws.cell(row=con_row, column=1, value="結論")
    ws.cell(row=con_row, column=1).font = BBOLD
    ws.cell(row=con_row, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
    ws.merge_cells(start_row=con_row, start_column=1, end_row=con_row, end_column=3)
    ws.cell(row=con_row, column=4, value="運用評価：有効（軽微例外1件は業務影響なしとして許容）")
    ws.cell(row=con_row, column=4).font = BBOLD
    ws.cell(row=con_row, column=4).fill = FILL_OK
    ws.merge_cells(start_row=con_row, start_column=4, end_row=con_row, end_column=11)

    widths = [6, 18, 12, 12, 20, 14, 14, 18, 22, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A13"

    wb.save(BASE / "運用評価_PLC-S-001_受注承認サンプル25件判定.xlsx")
    print("Created: 運用評価_PLC-S-001_受注承認サンプル25件判定.xlsx")


# ============================================================
# 監査人の例外事例メモ
# ============================================================
def gen_exception_memo():
    content = """# 例外事例検討メモ — PLC-S-001 受注承認 サンプル14

**作成日**: 2026/2/13
**作成者**: 長谷川 剛（内部監査室長 IA001）
**調書区分**: 監査人の例外事例判定メモ
**統制ID**: PLC-S-001
**サンプル番号**: 14（受注番号 ORD-2025-1876）

---

## 1. 事実関係（エビデンスに基づく）

参照エビデンス：
- `4.evidence/PLC-S/PLC-S-001_SAP_VA05_受注伝票一覧_FY2025.xlsx`
- `4.evidence/PLC-S/PLC-S-001_注文書_ORD-2025-1876_サンプル顧客L社.pdf`
- `4.evidence/PLC-S/PLC-S-001_与信限度マスタ_SAP_FD32スナップショット.xlsx`

| 項目 | 内容 |
|------|------|
| 受注日 | 2025/12/3（水） |
| 受注金額 | ¥15,200,000 |
| 顧客 | C-10015 サンプル顧客L社 |
| 受注担当 | 松本 香織（SLS004） |
| 与信限度額 | ¥100,000,000 |
| 既存売掛金残高（想定） | ¥92,400,000 |
| 与信残（受注前） | ¥7,600,000 |
| **与信超過** | ¥7,600,000（超過額） |
| 承認必要者 | 田中 太郎（営業本部長 SLS001） |
| 承認期待日 | 2025/12/3（受注日同日） |
| **実際の承認日** | **2025/12/4**（翌日 14:32） |

## 2. 承認遅延の原因（ヒアリングによる）

営業本部長 田中 太郎氏は2025/12/3に顧客訪問（関西エリア・日帰り出張）のため、
終日SAPワークフローへのアクセスが困難であった。翌12/4午後に帰社後、速やかに承認を実施した。

## 3. 業務影響の評価

| 観点 | 評価 |
|------|------|
| 出荷への影響 | なし（出荷指示は12/10予定、承認は余裕をもって取得） |
| 顧客への影響 | なし（顧客側の納期に影響なし） |
| 財務諸表への影響 | なし（期間帰属問題なし、売上計上は出荷連動） |
| 統制の本来目的 | 達成（承認者による適切な審査あり） |

## 4. 規程との照合

| 規程 | 要求事項 | 実態 | 評価 |
|------|---------|------|------|
| R11 販売管理規程 §4-2 | 与信超過時は営業本部長承認 | 承認取得済 | ○ |
| R11 販売管理規程 §4-3 | 承認は受注日当日中に取得すべし | 1日遅延 | △ |
| R11 販売管理規程 §4-4 | 遅延時は代替承認者（営業副本部長）による対応可 | 代替承認者未活用 | △ |

## 5. 判定

**判定：軽微な例外として許容（合格扱い）**

**理由**:
1. 統制の本来目的（権限のない受注の防止）は達成されている
2. 業務・財務への実害はなし
3. 遅延は1日のみで、かつ出荷実施前に承認完了している
4. 不正の意図も見受けられない

## 6. 改善提案

代替承認者（営業副本部長）制度を実質的に機能させるため、以下を営業本部に提案：
1. 代替承認者ルートの明示（ワークフロー設定強化）
2. 本部長スケジュール共有（1週間以上の不在予定を営業部員全員に共有）
3. 当日承認困難時の代替記録（課長ルートでの事前記録）

---

**報告先**: 監査等委員会（2026/2/18 定例会）
**備考**: 本例外事象はRCM PLC-S-001 運用評価結果欄に「軽微例外（許容）」として反映済
"""
    path = BASE / "監査人調書_PLC-S-001_例外事例メモ_サンプル14.md"
    path.write_text(content, encoding="utf-8")
    print(f"Created: {path.name}")


# ============================================================
# 追加エビデンス要求記録（判断保留案件追跡）
# ============================================================
def gen_evidence_request_log():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "追加エビデンス要求一覧"

    ws.cell(row=1, column=1, value="追加エビデンス要求一覧（判断保留案件の追跡）")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    ws.cell(row=2, column=1, value="管理者: 長谷川 剛（内部監査室長 IA001） / 更新: 2026/3/1時点")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    headers = ["要求№", "対象統制", "不足エビデンスの内容", "要求先", "要求日",
               "期限", "入手状況", "入手日", "現時点の運用評価"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[4].height = 32

    rows = [
        ("REQ-2026-001", "PLC-S-005",
         "2025年10-11-12月・2026年1月の売掛金年齢表の経理部長承認印が低解像度PDFで判読不能。再スキャンまたはメール承認記録。",
         "経理部 高橋課長", "2026/2/12", "2026/3/10",
         "対応中（スキャナ修理待ち）", "", "判断保留"),
        ("REQ-2026-002", "ITGC-AC-002",
         "Q3・Q4のアクセス権棚卸で、SUIMレポートの抽出日時・抽出条件のエビデンスが不明。レポート再出力時のスクリーンショット。",
         "情シス部 加藤リーダー", "2026/2/14", "2026/3/15",
         "対応中", "", "判断保留"),
        ("REQ-2026-003", "FCRP-003",
         "Q3・Q4の貸倒引当金の個別評価シートに、顧客別信用情報・回収可能性調査資料が未添付。",
         "経理部 佐藤部長", "2026/2/20", "2026/3/20",
         "対応中（外部信用情報の再取得中）", "", "判断保留"),
    ]

    r = 5
    for row in rows:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 5, 6, 8, 9):
                cell.alignment = C_
            else:
                cell.alignment = L_
        ws.cell(row=r, column=9).fill = FILL_HOLD
        ws.row_dimensions[r].height = 60
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="■ 期限までに入手できない場合の対応").font = BBOLD
    r += 1
    ws.cell(row=r, column=1,
            value="追加エビデンスが期限内に入手できない場合、運用状況評価は『不備』として認定する。"
                  "監査法人と協議のうえ、代替手続（ヒアリング文書化等）の適用可否を判断する。")
    ws.cell(row=r, column=1).font = BFONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)

    widths = [14, 14, 50, 16, 12, 12, 22, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    wb.save(BASE / "追加エビデンス要求一覧_判断保留案件追跡.xlsx")
    print("Created: 追加エビデンス要求一覧_判断保留案件追跡.xlsx")


# ============================================================
# ヒアリング記録（監査人の調書）
# ============================================================
def gen_interview_record():
    content = """# ウォークスルー・ヒアリング記録（PLC-S 販売プロセス）

| 項目 | 内容 |
|------|------|
| 日時 | 2026年2月5日（木）10:00～12:30 |
| 場所 | 営業本部 会議スペース |
| ヒアリング実施者 | 長谷川 剛（内部監査室長 IA001） |
| 調書区分 | 監査人のウォークスルー記録 |
| ヒアリング対象者 | 松本 香織（営業部主任 SLS004）、斎藤 次郎（営業部課長 SLS002）、中村 真理（経理部主任 ACC004） |
| 目的 | PLC-S-001～007 統制の整備状況評価（ウォークスルー） |

---

## 1. 受注プロセス全体（松本氏ヒアリング）

### Q1. 受注はどのように登録されますか？
**A**: 顧客からの注文書（PDFメール or FAX）を受領後、SAP VA01 で受注登録します。
顧客コード、製品コード、数量、単価、納期を入力すると、システムが自動的に与信限度額をチェックします。

### Q2. 与信限度を超えた場合はどうなりますか？
**A**: SAP上で「限度額超過」のワーニングが表示され、受注ステータスが「保留」となります。
ワークフロー（S04）が自動で起票され、営業本部長の承認がないと出荷指示に進めません。
先月も数件（5件程度）の超過案件があり、すべて本部長承認を経て進めました。

### Q3. 承認が翌日以降になるケースはありますか？
**A**: 本部長が出張等で当日承認できないケースが月に1-2件あります。
その場合は翌日に承認を取得しています。出荷までに時間的余裕があるため、業務影響はありません。
（※これが PLC-S-001 サンプル14 の例外ケース）

### Q4. 顧客マスタの新規登録・変更は？
**A**: 新規登録は営業本部長承認が必要です。変更は営業課長の承認で可能。
反社チェックは総務部で実施してもらいます。

---

## 2. 出荷・売上計上（松本・中村氏）

### Q5. 出荷指示から売上計上まではどう流れますか？
**A（松本）**: 営業部がSAPで出荷指示を作成 → WMS（S02）に連携 → 倉庫が出荷処理。
WMSで出荷完了を入力すると、SAP側に連携されます。
**A（中村）**: SAP側で出荷実績と受注がマッチされ、自動的に売上計上されます。
夜間バッチで処理されるため、翌営業日に結果が確認できます。

### Q6. 未マッチになることはありますか？
**A（中村）**: 月に数件はあります。原因は出荷時の数量訂正、システム連携タイミングのズレなど。
私が日次で未マッチ明細を確認し、原因調査して是正します。11月は1件あり、値引伝票で調整しました。

### Q7. 請求書の発行は？
**A（中村）**: 月末日にSAPが自動で請求書を生成・PDF化します。
発行件数と売上計上件数を突合し、私が確認印を押します。

---

## 3. 入金消込（中村氏）

### Q8. 入金消込の流れは？
**A**: 銀行のFBデータ（全銀形式）をSAPに日次で取込。
請求書番号がFBデータの摘要欄にあれば自動消込、なければ石井（ACC006）が手動消込します。

### Q9. 振込手数料の差額は？
**A**: 顧客側で振込手数料を差し引かれるケースがあり、少額（数百円～数千円）の差異が発生します。
規程に基づき当方負担として費用処理します。

---

## 4. 売掛金年齢分析（中村氏）

### Q10. 年齢表は毎月作成しますか？
**A**: 月末締め後、翌月第5営業日までに作成します。
60日超の債権はリストアップし、営業部に回収状況をコメントしてもらいます。
高橋課長が確認・承認印を押し、佐藤部長が最終承認します。

### Q11. 11月分の承認印が判読できない件について
**A**: 承知しております。紙の原本に佐藤部長が押印したものをスキャンしましたが、
スキャナの設定が低解像度になっていたようです。先日の内部監査室からの要求を受け、再スキャン中です。
スキャナが故障しているので、3月初旬までにコピー機経由で対応予定です。

---

## 5. 期末カットオフ（中村氏）

### Q12. 期末カットオフはどう実施されますか？
**A**: 期末日前後5営業日（3/25～4/1）の全出荷を佐藤部長が確認します。
出荷日と売上計上日の整合性、期間帰属の正確性をレビューします。
問題があれば修正仕訳を起票します。過去3年不整合は発生していません。

---

## 6. 価格マスタ（松本氏）

### Q13. 価格マスタ変更の承認フローは？
**A**: 営業担当が起案、営業課長レビュー、営業本部長承認。
稟議システム（S04）で記録されます。SAPマスタへの反映日は承認日以降です。

---

## 7. 総括（長谷川）

### 整備状況の評価
- すべての統制について、担当者が規程に基づき適切に業務を遂行していることを確認
- ウォークスルー・ヒアリング結果から、統制の設計は有効と判断
- **特記事項**: PLC-S-005のスキャン品質問題は整備面ではなく運用面の問題。追加エビデンス要求で対応中。

### 次工程
- 各統制について運用評価（サンプルテスト）を実施する
- 統制責任者への報告は 2026/2/20 予定

---

**作成**: 長谷川 剛 / 2026/2/5
**レビュー**: 大塚 美穂 / 2026/2/6
"""
    path = BASE / "監査人調書_PLC-S_ウォークスルーヒアリング記録_20260205.md"
    path.write_text(content, encoding="utf-8")
    print(f"Created: {path.name}")


if __name__ == "__main__":
    gen_plc_s_001_judgment()
    gen_exception_memo()
    gen_evidence_request_log()
    gen_interview_record()
    print("\n5.test_results/ 配下の監査調書を生成完了")
