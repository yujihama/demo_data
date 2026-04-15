"""
PLC-S-002 25件サンプルに対応するRAWエビデンス生成
- WMS出荷記録RAW (SAP/WMSから直接エクスポートされたもの)
- SAP売上計上仕訳RAW (SAP FI FBL3N)
- SAPマッチングログRAW (夜間バッチZSD_SHIP_SALES_MATCHのログ)
- 例外3件の個別是正記録（SAP画面コピー相当のテキスト記録）

これらが本来のエビデンス。Excel（25件サンプル対応エビデンス）はインデックス。
監査人は各サンプルをこれらRAWデータから再実施・検証する。
"""
import openpyxl
from pathlib import Path
from datetime import datetime, timedelta
import sys

sys.path.insert(0, str(Path(__file__).parent))
from pdf_util import JPPDF

BASE = Path(r"C:\Users\nyham\work\demo_data\4.evidence\PLC-S")

PRODUCT_NAMES = {
    "P-30001": "エンジンピストンピン A型",
    "P-30006": "トランスミッションシャフト",
    "P-30011": "ウェハー搬送ロボット用シャフト A",
    "P-30014": "エッチング装置チャンバ部品",
    "P-30015": "ウェハーチャックベース",
    "P-30020": "検査装置ステージベース",
    "P-30022": "サスペンションアーム A型",
    "P-30027": "ロボットアーム外装パネル A",
}


def read_samples_from_summary():
    """インデックスExcelから25件のサンプルデータを読み込む"""
    wb = openpyxl.load_workbook(BASE / "PLC-S-002_25件サンプル対応エビデンス.xlsx")
    ws = wb.active

    samples = []
    for r in range(17, 42):
        sample_no = ws.cell(row=r, column=1).value
        if sample_no is None:
            continue
        # 数量列は "300 個" のような書式。最初の数値部分のみ取り出し
        qty_str = str(ws.cell(row=r, column=8).value or "")
        try:
            qty = int(qty_str.split()[0].replace(",", ""))
        except (ValueError, IndexError):
            qty = 0
        sample = {
            "no": sample_no,
            "ship_no": ws.cell(row=r, column=2).value,
            "ship_date": ws.cell(row=r, column=3).value,
            "ord_no": ws.cell(row=r, column=4).value,
            "cid": ws.cell(row=r, column=5).value,
            "cname": ws.cell(row=r, column=6).value,
            "pcode": ws.cell(row=r, column=7).value,
            "qty": qty,
            "ship_amount": ws.cell(row=r, column=9).value,
            "jv_no": ws.cell(row=r, column=10).value,
            "sale_date": ws.cell(row=r, column=11).value,
            "sale_amount": ws.cell(row=r, column=12).value,
            "diff": ws.cell(row=r, column=13).value or 0,
            "judgment": ws.cell(row=r, column=14).value,
            "match_ts": ws.cell(row=r, column=15).value,
            "exception_note": ws.cell(row=r, column=16).value,
        }
        samples.append(sample)
    return samples


def gen_wms_raw(samples):
    """WMS出荷実績RAW（25件抽出版）"""
    path = BASE / "PLC-S-002_25件対応_RAW_WMS出荷実績エクスポート.csv"
    lines = [
        "# WMS (倉庫管理システム) / 出荷実績RAW抽出",
        "# 出力日時: 2026/02/11 09:20:15 JST",
        "# 出力者: WHS001 橋本 明（倉庫課長）",
        "# 抽出条件: 監査依頼IA-REQ-2026-002 指定の25件出荷番号",
        "# 抽出件数: 25件 / 当ファイルはFY2025全出荷3,158件のうち監査対象25件のみ",
        "#",
        "# 【監査人注】各行は監査人が選定したサンプルの原RAWデータ。",
        "# 以下の行それぞれについて、SAP売上計上仕訳RAW・マッチングログRAWとの突合可能。",
        "",
        ",".join(["サンプル№", "出荷番号", "出荷日時", "出荷区分", "受注番号",
                  "顧客コード", "顧客名", "製品コード", "製品名",
                  "出荷数量", "出荷金額", "出荷倉庫", "出荷担当者ID", "WMSレコード更新TS"]),
    ]
    for s in samples:
        # 出荷時刻はランダムだが固定化（sample_noをシードに使う）
        ship_hour = 8 + (s["no"] * 7) % 10
        ship_min = (s["no"] * 13) % 60
        ship_ts = datetime.combine(s["ship_date"], datetime.min.time()) + timedelta(
            hours=ship_hour, minutes=ship_min)
        warehouse = ["本社倉庫A", "本社倉庫B", "東北工場倉庫"][s["no"] % 3]
        user = ["WHS001", "MFG002", "WHS001"][s["no"] % 3]
        # WMSの更新TSは出荷TSとほぼ同時刻
        wms_update = ship_ts + timedelta(minutes=3)
        pname = PRODUCT_NAMES.get(s["pcode"], "(不明)")
        lines.append(",".join([
            str(s["no"]), s["ship_no"], ship_ts.strftime("%Y-%m-%d %H:%M:%S"),
            "通常出荷", s["ord_no"], s["cid"], s["cname"],
            s["pcode"], pname, str(s["qty"]), str(s["ship_amount"]),
            warehouse, user, wms_update.strftime("%Y-%m-%d %H:%M:%S")
        ]))
    lines.append("")
    lines.append("# 出力終了 / レコード数: 25 / CRC32: (WMS側で自動付与)")
    lines.append("# 本CSVはWMS管理画面 Tools > Data Export から生成")
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"Created: {path.name}")


def gen_sap_fi_raw(samples):
    """SAP FI 売上計上仕訳RAW（25件x2行＝50行、借方・貸方ペア）"""
    path = BASE / "PLC-S-002_25件対応_RAW_SAP売上計上仕訳_FBL3N.csv"
    lines = [
        "# SAP FI / 売上計上仕訳RAW明細",
        "# トランザクション: FBL3N (勘定科目別明細)",
        "# 対象勘定: 1220 売掛金 / 5100 売上高",
        "# 出力日時: 2026/02/11 09:25:42 JST",
        "# 出力者: ACC004 中村 真理（経理部主任）",
        "# 抽出条件: 監査依頼IA-REQ-2026-002 指定の25件売上仕訳",
        "#",
        "# 【監査人注】売上計上は「借方 売掛金 / 貸方 売上高」の2行で構成される。",
        "# SAP自動連動仕訳（出荷実績から起動）。Field Status Group: ZSDxx",
        "",
        ",".join(["サンプル№", "仕訳番号", "計上日", "転記日", "伝票タイプ",
                  "勘定科目コード", "勘定科目名", "借方/貸方", "金額(円)",
                  "顧客コード", "受注番号", "出荷番号", "摘要", "起票者"]),
    ]
    for s in samples:
        # 借方: 売掛金
        lines.append(",".join([
            str(s["no"]), s["jv_no"], s["sale_date"].strftime("%Y-%m-%d"),
            s["sale_date"].strftime("%Y-%m-%d"), "RV",
            "1220", "売掛金", "借方", str(s["sale_amount"]),
            s["cid"], s["ord_no"], s["ship_no"],
            f"出荷連動自動仕訳 {s['ship_no']}", "SAP_BATCH"
        ]))
        # 貸方: 売上高
        lines.append(",".join([
            str(s["no"]), s["jv_no"], s["sale_date"].strftime("%Y-%m-%d"),
            s["sale_date"].strftime("%Y-%m-%d"), "RV",
            "5100", "売上高", "貸方", str(s["sale_amount"]),
            s["cid"], s["ord_no"], s["ship_no"],
            f"出荷連動自動仕訳 {s['ship_no']}", "SAP_BATCH"
        ]))
        # 例外：値引調整の場合は是正仕訳を追加
        if "値引" in (s["exception_note"] or ""):
            diff_jv = s["jv_no"].replace("JV-", "JV-ADJ-")
            adj_date = s["sale_date"] + timedelta(days=3)
            # 借方: 売上高(取消)
            lines.append(",".join([
                str(s["no"]), diff_jv, adj_date.strftime("%Y-%m-%d"),
                adj_date.strftime("%Y-%m-%d"), "RV",
                "5100", "売上高", "借方", "50000",
                s["cid"], s["ord_no"], s["ship_no"],
                f"値引伝票DR-202511-0012 {s['ship_no']}", "ACC004"
            ]))
            # 貸方: 売掛金(減額)
            lines.append(",".join([
                str(s["no"]), diff_jv, adj_date.strftime("%Y-%m-%d"),
                adj_date.strftime("%Y-%m-%d"), "RV",
                "1220", "売掛金", "貸方", "50000",
                s["cid"], s["ord_no"], s["ship_no"],
                f"値引伝票DR-202511-0012 {s['ship_no']}", "ACC004"
            ]))
    lines.append("")
    lines.append("# 出力終了 / レコード数: 52（通常50行＋是正仕訳2行）")
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"Created: {path.name}")


def gen_match_log_raw(samples):
    """SAP夜間バッチのマッチングログRAW"""
    path = BASE / "PLC-S-002_25件対応_RAW_SAPマッチングバッチログ.csv"
    lines = [
        "# SAP Background Job Log",
        "# ジョブ名: ZSD_SHIP_SALES_MATCH",
        "# ジョブ種別: 日次夜間バッチ (毎日01:00JST起動)",
        "# 出力日時: 2026/02/11 09:30:08 JST",
        "# 出力者: IT003 加藤 洋子（情シス部アプリリーダー）",
        "# 抽出条件: 監査依頼IA-REQ-2026-002 指定の25件に対応するバッチ実行ログ",
        "#",
        "# 【監査人注】本ログは出荷-売上マッチングの自動判定記録。",
        "# MATCH_RESULT = OK: 自動マッチ成功（公差内）",
        "# MATCH_RESULT = EXCEPTION: 公差超過または不整合（未マッチ明細として経理部レビューへ）",
        "",
        ",".join(["実行タイムスタンプ", "バッチ実行ID", "サンプル№",
                  "出荷番号", "受注番号", "売上仕訳番号",
                  "出荷金額(円)", "売上金額(円)", "差異(円)",
                  "MATCH_RESULT", "詳細メッセージ", "実行ユーザ"]),
    ]
    for s in samples:
        exec_id = f"ZSD-{s['match_ts'][:10].replace('-', '') if isinstance(s['match_ts'], str) else s['ship_date'].strftime('%Y%m%d')}-{s['no']:03d}"
        match_ts = s["match_ts"] if isinstance(s["match_ts"], str) else \
                   datetime.combine(s["ship_date"] + timedelta(days=1), datetime.min.time()).strftime("%Y-%m-%d %H:%M:%S")

        if "値引" in (s["exception_note"] or ""):
            result = "EXCEPTION"
            msg = "公差超過:差異50000円(+0.39%) → 未マッチ明細リストへ転送"
        elif "数量訂正" in (s["exception_note"] or ""):
            result = "EXCEPTION"
            msg = "出荷金額更新検知:当初値と売上計上値に差異 → 未マッチ明細リストへ転送"
        elif "計上日1日遅延" in (s["judgment"] or ""):
            result = "EXCEPTION"
            msg = "対応売上仕訳未検出 → 翌日バッチで再試行"
        else:
            result = "OK"
            msg = "金額一致・受注号一致・顧客一致"

        lines.append(",".join([
            match_ts, exec_id, str(s["no"]),
            s["ship_no"], s["ord_no"], s["jv_no"],
            str(s["ship_amount"]), str(s["sale_amount"]),
            str(s["diff"]), result, msg, "SAP_BATCH"
        ]))

    # 例外3件について、再試行ログも追加
    lines.append("")
    lines.append("# --- 例外サンプルの翌日再試行ログ ---")
    for s in samples:
        if "計上日1日遅延" in (s["judgment"] or ""):
            # 翌日01:15頃に再実行して成功
            retry_ts = datetime.combine(s["sale_date"] + timedelta(days=1),
                                         datetime.min.time()) + timedelta(hours=1, minutes=15)
            exec_id = f"ZSD-{(s['sale_date'] + timedelta(days=1)).strftime('%Y%m%d')}-RETRY-{s['no']:03d}"
            lines.append(",".join([
                retry_ts.strftime("%Y-%m-%d %H:%M:%S"), exec_id, str(s["no"]),
                s["ship_no"], s["ord_no"], s["jv_no"],
                str(s["ship_amount"]), str(s["sale_amount"]),
                "0", "OK", "翌日再試行で売上仕訳マッチ成功", "SAP_BATCH"
            ]))
    lines.append("")
    lines.append("# 出力終了 / 通常25行 + 再試行1行 = 26行")
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"Created: {path.name}")


def gen_exception_correction_records(samples):
    """例外3件の個別是正記録（担当者メモ/システム画面記録）"""

    # サンプル9: 数量訂正
    s9 = next(s for s in samples if "数量訂正" in (s["exception_note"] or ""))
    s14 = next(s for s in samples
               if "計上日1日遅延" in (s["judgment"] or "")
               or "翌営業日" in (s["exception_note"] or ""))
    # s16は既存の値引調整（既にPLC-S-002_SAP未マッチ明細リスト_202511.csvに記録あり）

    # 数量訂正のSAP画面キャプチャメモ（テキスト形式で詳細記録）
    path_s9 = BASE / f"PLC-S-002_25件対応_RAW_例外サンプル9_数量訂正SAP変更履歴_{s9['ship_no']}.txt"
    content = f"""================================================================
 SAP VA02 / 受注伝票変更履歴レポート
================================================================
出力日時:     2026-02-11 09:45:30 JST
出力者:       ACC004 中村 真理（経理部主任）
抽出対象:     {s9['ord_no']} (関連出荷 {s9['ship_no']})

----------------------------------------------------------------
 変更履歴
----------------------------------------------------------------
レコード作成:
  日時:       {s9['ship_date'].strftime('%Y-%m-%d')} 09:15:22
  ユーザ:     SLS004 松本 香織
  受注数量:   {s9['qty'] + 2}個
  受注金額:   (当初金額 = 単価 × {s9['qty'] + 2}個)

変更1（数量訂正）:
  日時:       {(s9['ship_date'] + timedelta(days=1)).strftime('%Y-%m-%d')} 14:22:18
  ユーザ:     SLS004 松本 香織
  変更理由:   顧客からの数量変更要請（出荷予定削減）
  訂正前数量: {s9['qty'] + 2}個
  訂正後数量: {s9['qty']}個
  承認:       SLS002 斎藤 次郎 課長（同日 14:35 承認）

----------------------------------------------------------------
 SAP FI 連動仕訳の訂正
----------------------------------------------------------------
当初売上仕訳: {s9['jv_no']} (取消処理は実施せず、金額が直接更新された)
訂正後金額:   {s9['sale_amount']:,}円

経理部主任確認（中村）:
  - 日時:   {(s9['ship_date'] + timedelta(days=2)).strftime('%Y-%m-%d')} 10:12
  - 内容:   日次未マッチレビューで検知。SAP変更履歴を確認のうえ、
            訂正経緯が規程（販売管理R11）に整合することを確認。
            金額の最終一致を確認済。経理部長への月次報告対象。

----------------------------------------------------------------
 マッチングへの影響
----------------------------------------------------------------
変更直後: 一時的にWMS出荷金額とSAP売上金額に差異が発生した可能性
変更翌日のマッチングバッチ: 訂正後の金額同士で自動マッチ成功
  → PLC-S-002_25件対応_RAW_SAPマッチングバッチログ.csv サンプル№9 参照

================================================================
"""
    path_s9.write_text(content, encoding="utf-8")
    print(f"Created: {path_s9.name}")

    # サンプル14: 売上計上日遅延
    path_s14 = BASE / f"PLC-S-002_25件対応_RAW_例外サンプル14_計上日遅延ログ_{s14['ship_no']}.txt"
    content = f"""================================================================
 SAP Batch Job Execution Detail
================================================================
出力日時:     2026-02-11 09:52:10 JST
出力者:       IT003 加藤 洋子（情シス部アプリリーダー）
抽出対象:     夜間バッチZSD_SHIP_SALES_MATCH の {s14['ship_no']} 関連実行履歴

----------------------------------------------------------------
 1回目バッチ実行（出荷日当日深夜）
----------------------------------------------------------------
実行日時:     {s14['ship_date'].strftime('%Y-%m-%d')} 23:58:00 (夜間バッチ定時起動)
対象:         {s14['ship_no']} (出荷日: {s14['ship_date'].strftime('%Y-%m-%d')})
処理結果:     EXCEPTION - 対応売上仕訳が未検出
詳細:         WMSからは出荷確定を受信済。しかしSAP FIへの自動売上計上が
              未実施（SAP連動IFのタイミング差 or 人手確認待ち）
アクション:   未マッチ明細リストへ転送、翌日バッチで再試行設定

----------------------------------------------------------------
 2回目バッチ実行（翌日深夜）
----------------------------------------------------------------
実行日時:     {(s14['ship_date'] + timedelta(days=1)).strftime('%Y-%m-%d')} 01:15:00 (RETRY)
対象:         {s14['ship_no']}
前日夜に処理: 売上仕訳 {s14['jv_no']} が当日{s14['sale_date'].strftime('%Y-%m-%d')} 17:30に計上済
処理結果:     OK - 金額一致 {s14['ship_amount']:,}円 = {s14['sale_amount']:,}円
マッチング完了

----------------------------------------------------------------
 経理部主任（中村）による確認
----------------------------------------------------------------
確認日時:     {(s14['ship_date'] + timedelta(days=2)).strftime('%Y-%m-%d')} 09:35
確認内容:
  - 遅延の原因: SAP FI連動IFで当日バッチのタイミング差によるもの
  - 翌営業日に自動マッチ成功を確認
  - 期間帰属: 同一会計月内のため財務諸表への影響なし
  - R17決算業務規程への抵触なし
判定:         例外として許容（翌日解消、金額・期間帰属とも問題なし）

================================================================
"""
    path_s14.write_text(content, encoding="utf-8")
    print(f"Created: {path_s14.name}")


def gen_daily_review_log():
    """経理部担当による日次未マッチレビュー記録（統制実施記録の一部）"""
    path = BASE / "PLC-S-002_25件対応_日次未マッチレビュー記録_FY2025抜粋.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "日次未マッチレビュー記録"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HF = PatternFill("solid", fgColor="1F4E78")
    HFONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
    BFONT = Font(name="Yu Gothic", size=10)
    C_ = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L_ = Alignment(horizontal="left", vertical="center", wrap_text=True)
    T_ = Side("thin", color="888888")
    BRD = Border(left=T_, right=T_, top=T_, bottom=T_)
    FILL_WARN = PatternFill("solid", fgColor="FFF2CC")

    ws.cell(row=1, column=1, value="PLC-S-002 日次未マッチレビュー記録 (25件サンプル対応)")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    ws.cell(row=2, column=1, value="実施者: 中村 真理 (ACC004) / 抽出: 25件サンプル対象日のみ / 作成日: 2026/2/12")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    headers = ["レビュー日", "対象出荷番号", "バッチ結果",
               "未マッチの有無", "中村主任の確認", "対応実施日", "備考"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[4].height = 28

    wb_idx = openpyxl.load_workbook(BASE / "PLC-S-002_25件サンプル対応エビデンス.xlsx")
    ws_idx = wb_idx.active
    samples = []
    for r in range(17, 42):
        no = ws_idx.cell(row=r, column=1).value
        if no is None:
            continue
        samples.append({
            "no": no,
            "ship_no": ws_idx.cell(row=r, column=2).value,
            "ship_date": ws_idx.cell(row=r, column=3).value,
            "sale_date": ws_idx.cell(row=r, column=11).value,
            "judgment": ws_idx.cell(row=r, column=14).value,
            "exception_note": ws_idx.cell(row=r, column=16).value or "",
        })

    r = 5
    for s in samples:
        review_date = s["ship_date"] + timedelta(days=1)
        is_exception = "例外" in s["judgment"]
        if is_exception:
            batch_result = "EXCEPTION"
            unmatched = "あり"
            if "値引" in s["exception_note"]:
                confirmation = "○ 金額差異を検知、高橋課長報告、値引伝票で是正"
                action_date = s["ship_date"] + timedelta(days=3)
                remark = "DR-202511-0012 値引伝票計上済"
            elif "数量訂正" in s["exception_note"]:
                confirmation = "○ 数量訂正を検知、SAP変更履歴確認、金額一致を再確認"
                action_date = s["ship_date"] + timedelta(days=2)
                remark = "訂正経緯は販売管理規程R11に整合"
            elif "計上日1日遅延" in (s["judgment"] or ""):
                confirmation = "○ 翌日再試行でマッチ成功、期間帰属OK"
                action_date = s["ship_date"] + timedelta(days=2)
                remark = "SAP FI連動IFタイミング差、業務影響なし"
            else:
                confirmation = "○ 問題なし"
                action_date = review_date
                remark = ""
        else:
            batch_result = "OK"
            unmatched = "なし"
            confirmation = "○ 確認済"
            action_date = review_date
            remark = ""

        data = [review_date, s["ship_no"], batch_result, unmatched,
                confirmation, action_date, remark]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 6):
                cell.alignment = C_
                if c_i in (1, 6):
                    cell.number_format = "yyyy/mm/dd"
            else:
                cell.alignment = L_
        if is_exception:
            for c_i in (3, 4):
                ws.cell(row=r, column=c_i).fill = FILL_WARN
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="■ 本記録の位置づけ").font = Font(name="Yu Gothic", size=10, bold=True)
    r += 1
    ws.cell(row=r, column=1, value=(
        "経理部担当（中村主任 ACC004）が実施する日次未マッチレビューの記録から、"
        "25件サンプルの対象日に関連する行を抽出した抜粋。"
        "本来は年間240日分の日次記録があるが、監査目的のため該当日のみ提示。")).font = BFONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 2
    ws.cell(row=r, column=1, value="承認: 高橋 美咲（経理部課長 ACC002）[印] 2026/2/12")

    widths = [12, 16, 12, 12, 36, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    print(f"Created: {path.name}")


if __name__ == "__main__":
    samples = read_samples_from_summary()
    print(f"Loaded {len(samples)} samples from summary Excel")
    gen_wms_raw(samples)
    gen_sap_fi_raw(samples)
    gen_match_log_raw(samples)
    gen_exception_correction_records(samples)
    gen_daily_review_log()
    print("\nAll RAW evidence for PLC-S-002 25 samples generated.")
