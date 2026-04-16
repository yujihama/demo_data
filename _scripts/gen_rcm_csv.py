"""
全RCMから統一フォーマットのCSVを生成

カラム:
- key: 統制ID
- procedure: 監査手続（どのように評価するか）
- control: 統制活動（何をどう行うか）
- risk: リスク記述
- sample_num: サンプル件数
"""
import csv
import openpyxl
from pathlib import Path

BASE = Path(r"C:\Users\nyham\work\demo_data")
OUTPUT = BASE / "2.RCM" / "RCM.csv"

# 各RCMファイルの列位置定義
# (file, header_row, data_start_row, col_id, col_control, col_risk, col_frequency)
RCM_DEFS = [
    {
        "file": "ELC_RCM.xlsx",
        "header_row": 4,
        "data_start": 5,
        "col_id": 1,        # 統制ID
        "col_control": 5,   # 統制活動
        "col_risk": 4,      # リスク記述
        "col_frequency": 7, # 頻度
        "divider_check": True,  # COSO区切り行あり
    },
    {
        "file": "PLC_Sales_RCM.xlsx",
        "header_row": 4,
        "data_start": 5,
        "col_id": 1,
        "col_control": 7,   # 統制活動
        "col_risk": 4,      # リスク記述
        "col_frequency": 10, # 頻度
    },
    {
        "file": "PLC_Purchasing_RCM.xlsx",
        "header_row": 4,
        "data_start": 5,
        "col_id": 1,
        "col_control": 7,
        "col_risk": 4,
        "col_frequency": 10,
    },
    {
        "file": "PLC_Inventory_Cost_RCM.xlsx",
        "header_row": 4,
        "data_start": 5,
        "col_id": 1,
        "col_control": 7,
        "col_risk": 4,
        "col_frequency": 10,
    },
    {
        "file": "ITGC_RCM.xlsx",
        "header_row": 4,
        "data_start": 5,
        "col_id": 1,
        "col_control": 5,   # 統制活動
        "col_risk": 4,
        "col_frequency": 8,  # 頻度
        "divider_check": True,
    },
    {
        "file": "ITAC_RCM.xlsx",
        "header_row": 4,
        "data_start": 5,
        "col_id": 1,
        "col_control": 7,   # 統制活動（IT自動統制の内容）
        "col_risk": 4,
        "col_frequency": 9,  # 頻度
    },
    {
        "file": "FCRP_RCM.xlsx",
        "header_row": 4,
        "data_start": 5,
        "col_id": 1,
        "col_control": 7,
        "col_risk": 4,
        "col_frequency": 10,
    },
]


# 頻度 → サンプル数 の変換
# 監査手続書の設計に基づく
FREQUENCY_TO_SAMPLE = {
    "都度": 25,
    "日次": 25,
    "日次/週次": 25,
    "週次": 15,
    "月次": 12,        # 全数
    "月次/四半期": 12,
    "四半期": 4,        # 全数
    "半期": 2,          # 全数
    "年次": 1,          # 全数
    "年次\n(計画ベース)": 1,
    "随時": 25,
    "随時\n(年1回見直し)": 1,
}

# 統制別オーバーライド（監査手続書に明記されたサンプル数）
SAMPLE_OVERRIDE = {
    "PLC-S-001": 25,  # 全受注3,247件→25件
    "PLC-S-002": 25,  # 全出荷3,158件→25件
    "PLC-S-003": 12,  # 月次バッチ12回→全数
    "PLC-S-004": 25,  # 入金2,847件→25件
    "PLC-S-005": 12,  # 月次年齢表12回→全数
    "PLC-S-006": 41,  # 期末カットオフ41件→全数
    "PLC-S-007": 25,  # マスタ変更36件→25件
    "PLC-P-001": 25,
    "PLC-P-002": 25,
    "PLC-P-003": 25,
    "PLC-P-004": 25,
    "PLC-P-005": 25,
    "PLC-P-006": 12,
    "PLC-P-007": 87,  # 全数
    "PLC-I-001": 2,   # 半期2回
    "PLC-I-002": 24,  # 全数
    "PLC-I-003": 1,   # 年1回
    "PLC-I-004": 3,   # 四半期代表3
    "PLC-I-005": 4,   # 四半期全数
    "PLC-I-006": 25,
    "PLC-I-007": 12,  # 月次全数
    "ITGC-AC-001": 25,
    "ITGC-AC-002": 4,  # 四半期全数
    "ITGC-AC-003": 5,  # 退職者全数
    "ITGC-AC-004": 12, # 月次全数
    "ITGC-CM-001": 25,
    "ITGC-CM-002": 25,
    "ITGC-CM-003": 25,
    "ITGC-OM-001": 25,
    "ITGC-OM-002": 18, # 全数
    "ITGC-EM-001": 2,  # 委託先2社全数
    "ITAC-001": 5,     # 動作検証テストケース
    "ITAC-002": 3,
    "ITAC-003": 3,
    "ITAC-004": 5,
    "ITAC-005": 2,
    "FCRP-001": 12,
    "FCRP-002": 4,
    "FCRP-003": 4,
    "FCRP-004": 4,
    "FCRP-005": 4,
}

# 統制別の監査手続（procedure）定義
PROCEDURES = {
    # ELC
    "ELC-001": "取締役会議事録を閲覧し、月次開催実績・出席者・決議事項を確認する。",
    "ELC-002": "HRシステムから倫理綱領受領確認ログを取得し、全役職員の確認完了を検証する。",
    "ELC-003": "職務権限規程R18の原本を閲覧し、承認権限体系・組織構造が定義されていることを確認する。従業員マスタ・SAPロールマトリクスと照合する。",
    "ELC-004": "年次リスクアセスメント結果を閲覧し、財務・事業・ITリスクの網羅的評価と対策が記録されていることを確認する。",
    "ELC-005": "全社リスクアセスメント結果から不正リスクファクター（動機・機会・正当化）の評価が含まれていることを確認する。",
    "ELC-006": "会社プロファイルの規程体系一覧（R01-R27）を参照し、主要規程が制定・改訂されていることを確認する。",
    "ELC-007": "SAPロールマトリクスを閲覧し、職務分掌（SoD）違反がないか検証する。PLC-P-002の発注書（不備ケース）を確認する。",
    "ELC-008": "内部通報受付台帳を閲覧し、通報の受付・調査・完了プロセスが機能していることを確認する。",
    "ELC-009": "月次決算ジョブログを閲覧し、期限内に決算処理が完了していることを確認する。",
    "ELC-010": "年次内部監査計画書を閲覧し、リスクベースの監査テーマ・実施体制が計画されていることを確認する。",
    "ELC-011": "取締役会議事録および内部通報台帳から、監査等委員会への報告が行われていることを確認する。",
    "ELC-012": "会社プロファイルのIT構成情報およびSOC1レポートから、IT戦略・セキュリティ方針が策定・評価されていることを確認する。",
    # PLC-S
    "PLC-S-001": "SAP VA05受注一覧から25件を系統抽出し、各受注の与信チェックログ・ワークフロー承認履歴・注文書原本を突合する。与信超過案件は営業本部長承認の有無を検証する。",
    "PLC-S-002": "WMS出荷実績・SAP売上計上仕訳・マッチングバッチログの3点を25件分突合し、出荷と売上計上の金額・日付一致を検証する。例外（未マッチ）ケースの是正経緯を個別ログで確認する。",
    "PLC-S-003": "12ヶ月分のSAP請求書バッチログを確認し、各月の正常終了を検証する。請求書原本と売上計上額の一致をサンプル確認する。",
    "PLC-S-004": "銀行FBデータとSAP入金消込履歴を25件分突合し、入金額・消込対象請求書・差額処理の適切性を検証する。",
    "PLC-S-005": "12ヶ月分のSAP FB10N売掛金年齢表を閲覧し、滞留債権（60日超）の認識・分析が各月実施されていることを確認する。",
    "PLC-S-006": "期末前後5営業日（41件）のSAP出荷・売上明細を全数検証し、出荷日と売上計上日の期間帰属が適切であることを確認する。",
    "PLC-S-007": "SAP VK12価格変更履歴から25件を抽出し、各変更の稟議書原本（承認経路含む）との整合性を検証する。",
    # PLC-P
    "PLC-P-001": "SAP ME5A購買依頼一覧から25件を系統抽出し、各依頼の起案部門・部門長承認・予算コードを検証する。",
    "PLC-P-002": "SAP ME2N発注一覧から25件を系統抽出し、発注書原本・ワークフロー承認履歴を突合する。各発注の承認者が職務権限規程R18の金額区分に準拠しているか検証する。",
    "PLC-P-003": "SAP MIGO検収一覧から25件を系統抽出し、各検収の報告書原本（品目・数量・品質判定）と発注書を突合する。数量差異がある場合は差異報告書を確認する。",
    "PLC-P-004": "SAP MIRO請求書計上から25件を系統抽出し、PO金額・検収金額・請求金額の3点一致（公差内）を検証する。公差超過案件の保留処理を確認する。",
    "PLC-P-005": "SAP XK01/XK02仕入先マスタ変更履歴から25件を抽出し、各変更の申請書原本（新規登録時は反社チェック・信用調査含む）を検証する。",
    "PLC-P-006": "12ヶ月分のSAP F110支払実行ログを閲覧し、各月の支払バッチが正常実行され、ベンダー別支払明細が適切であることを検証する。",
    "PLC-P-007": "期末時点のSAP未払計上明細（全87件）を検証し、検収済・請求書未着取引の未払計上が網羅的に行われていることを確認する。",
    # PLC-I
    "PLC-I-001": "SAP MI07棚卸差異一覧を閲覧し、半期2回の実地棚卸が倉庫別に実施され、差異が記録されていることを確認する。棚卸写真で実施状況を補完確認する。",
    "PLC-I-002": "SAP MIGO棚卸差異調整仕訳（全24件）を検証し、各差異について調整仕訳が適切に起票されていることを確認する。",
    "PLC-I-003": "標準原価更新稟議（年1回）の原本を閲覧し、承認経路（経理部長→CFO）が適切であることを確認する。",
    "PLC-I-004": "SAP CO88原価差異レポートを3四半期分閲覧し、材料費・労務費・製造間接費・外注加工費の差異分析が実施されていることを確認する。",
    "PLC-I-005": "SAP MB52滞留在庫評価損計算結果を4四半期分閲覧し、回転期間基準（12/18/24ヶ月超）に基づく評価損計算の正確性を検証する。",
    "PLC-I-006": "WMS-SAP在庫照合レポートから25日分を抽出し、日次の在庫数量一致（または差異是正）を検証する。",
    "PLC-I-007": "SAP MMPV期末処理ログを12ヶ月分閲覧し、MM/CO/FIの各モジュール別ステップが毎月完了していることを確認する。",
    # ITGC
    "ITGC-AC-001": "SAP SU01ユーザ作成履歴から25件を抽出し、各登録のワークフロー承認履歴・申請書原本を突合する。SoDチェック結果を確認する。",
    "ITGC-AC-002": "SAP SUIMアクティブユーザ一覧を4四半期分取得し、各四半期に棚卸が実施されアクセス権の見直しが行われたことを確認する。",
    "ITGC-AC-003": "退職者5名（全数）のSAP SM20ログインログを閲覧し、退職日以降のログインがないことを確認する。停止処理の適時性を検証する。",
    "ITGC-AC-004": "SAP SM20特権IDの操作ログを12ヶ月分閲覧し、不正な操作がないことを確認する。",
    "ITGC-CM-001": "変更管理台帳から25件を抽出し、各変更の申請書原本・承認記録を検証する。",
    "ITGC-CM-002": "25件の変更に対応するUATテスト結果（個別Excel）およびXrayテスト実行履歴を閲覧し、テストケースが実施・合格していることを検証する。",
    "ITGC-CM-003": "SAP STMS本番移送履歴から25件を抽出し、各移送が承認済みの変更申請に紐づいていることを検証する。",
    "ITGC-OM-001": "SAP DB13バックアップログから25日分を抽出し、各コンポーネント（DATA/LOG/CATALOG）のバックアップが正常完了していることを検証する。DRテスト報告書でリストア可能性を確認する。",
    "ITGC-OM-002": "Zabbixインシデントログ（全18件）のタイムラインを検証し、検知→通知→対応→解決→クローズの各フェーズが適切に記録されていることを確認する。",
    "ITGC-EM-001": "外部委託先2社のSOC1 Type IIレポート原本を閲覧し、統制目標の有効性評価結果および例外事項を確認する。",
    # ITAC
    "ITAC-001": "SAP OVAK与信チェック設定画面を確認し、与信限度超過時の自動保留が設計通りであることを検証する。ITGCの変更管理（CM）が有効であることを前提とする。",
    "ITAC-002": "SAP OMRK 3-wayマッチング設定および月次実行ログを閲覧し、公差設定と自動保留機能が設計通りであることを検証する。",
    "ITAC-003": "SAP AFAB減価償却バッチの実行画面および結果を確認し、自動計算が設計通りであることを検証する。手計算との再実施による突合を行う。",
    "ITAC-004": "ワークフロー承認履歴（購買・変更管理）を閲覧し、金額に応じた承認ルーティングが自動判定されていることを検証する。",
    "ITAC-005": "連結システムのパッケージアップロードログを閲覧し、バリデーションエラー検知機能が機能していることを確認する。",
    # FCRP
    "FCRP-001": "SAP期末処理ジョブログを12ヶ月分閲覧し、FI/CO/MM各モジュールの月次締めステップが完了していることを検証する。",
    "FCRP-002": "連結システムのパッケージ受信ログを4四半期分閲覧し、各子会社パッケージの受領・バリデーション通過を確認する。",
    "FCRP-003": "SAP FB10N売掛金データおよび引当金算定データを4四半期分閲覧し、一般債権（実績率法）・個別債権の引当計算の正確性を検証する。根拠資料の有無を確認する。",
    "FCRP-004": "連結システムの仕訳一覧を4四半期分閲覧し、投資相殺・内部取引消去・少数株主損益等の連結仕訳が適切に計上されていることを検証する。",
    "FCRP-005": "開示システムのXBRL検証ログを4四半期分閲覧し、開示書類のバリデーションが正常通過していることを確認する。",
}


def clean_text(v):
    """セル値をクリーンなテキストに変換（改行→空白、前後空白除去）"""
    if v is None:
        return ""
    return str(v).replace("\n", " ").replace("\r", "").strip()


def extract_from_rcm(rcm_def):
    """1つのRCMファイルから統制データを抽出"""
    wb = openpyxl.load_workbook(BASE / "2.RCM" / rcm_def["file"], data_only=True)
    ws = wb.active
    results = []

    for r in range(rcm_def["data_start"], ws.max_row + 1):
        cid = clean_text(ws.cell(row=r, column=rcm_def["col_id"]).value)

        # 区切り行（COSO要素の見出し等）をスキップ
        if not cid or not any(cid.startswith(p) for p in
                              ["ELC-", "PLC-", "ITGC-", "ITAC-", "FCRP-"]):
            continue

        control = clean_text(ws.cell(row=r, column=rcm_def["col_control"]).value)
        risk = clean_text(ws.cell(row=r, column=rcm_def["col_risk"]).value)
        freq = clean_text(ws.cell(row=r, column=rcm_def["col_frequency"]).value)

        # サンプル数の決定
        sample_num = SAMPLE_OVERRIDE.get(cid)
        if sample_num is None:
            for freq_key, num in FREQUENCY_TO_SAMPLE.items():
                if freq_key in freq:
                    sample_num = num
                    break
        if sample_num is None:
            sample_num = 25  # デフォルト

        # 監査手続（procedure）
        procedure = PROCEDURES.get(cid, "")

        results.append({
            "key": cid,
            "procedure": procedure,
            "control": control,
            "risk": risk,
            "sample_num": sample_num,
        })

    return results


def main():
    all_rows = []
    for rcm_def in RCM_DEFS:
        rows = extract_from_rcm(rcm_def)
        all_rows.extend(rows)
        print(f"  {rcm_def['file']}: {len(rows)} controls extracted")

    # ソート
    def sort_key(row):
        cid = row["key"]
        if cid.startswith("ELC"):
            return (1, cid)
        elif cid.startswith("PLC-S"):
            return (2, cid)
        elif cid.startswith("PLC-P"):
            return (3, cid)
        elif cid.startswith("PLC-I"):
            return (4, cid)
        elif cid.startswith("ITGC"):
            return (5, cid)
        elif cid.startswith("ITAC"):
            return (6, cid)
        elif cid.startswith("FCRP"):
            return (7, cid)
        return (99, cid)

    all_rows.sort(key=sort_key)

    # CSV出力
    with open(OUTPUT, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["key", "procedure", "control", "risk", "sample_num"],
            quoting=csv.QUOTE_ALL,
        )
        writer.writeheader()
        for row in all_rows:
            writer.writerow(row)

    print(f"\nCreated: {OUTPUT.name}")
    print(f"  Total controls: {len(all_rows)}")

    # 漏れチェック
    expected = set(PROCEDURES.keys())
    actual = set(r["key"] for r in all_rows)
    missing = expected - actual
    extra = actual - expected
    if missing:
        print(f"  WARNING: Missing from RCM: {missing}")
    if extra:
        print(f"  WARNING: No procedure defined: {extra}")


if __name__ == "__main__":
    main()
