"""
PLC-I（在庫・原価プロセス）エビデンス生成
【真の不備ケース：PLC-I-002 棚卸差異調整 ¥850,000の原因分析・報告欠如】を含む
棚卸写真JPG含む
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
from pathlib import Path
import random
import sys

sys.path.insert(0, str(Path(__file__).parent))
from pdf_util import JPPDF
from image_util import sap_screenshot, warehouse_photo

BASE = Path(r"C:\Users\nyham\work\demo_data\4.evidence\PLC-I")
BASE.mkdir(parents=True, exist_ok=True)

HF = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
BFONT = Font(name="Yu Gothic", size=10)
BBOLD = Font(name="Yu Gothic", size=10, bold=True)
C_ = Alignment(horizontal="center", vertical="center", wrap_text=True)
L_ = Alignment(horizontal="left", vertical="center", wrap_text=True)
R_ = Alignment(horizontal="right", vertical="center")
T_ = Side("thin", color="888888")
BRD = Border(left=T_, right=T_, top=T_, bottom=T_)
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")
FILL_NG = PatternFill("solid", fgColor="FCE4D6")
FILL_OK = PatternFill("solid", fgColor="E2EFDA")

# 製品コード一覧（マスタから抜粋）
PRODUCTS = {
    "P-30001": ("エンジンピストンピン A型", 2850, 4200),
    "P-30002": ("エンジンピストンピン B型", 3150, 4650),
    "P-30003": ("バルブステム A型", 1850, 2750),
    "P-30006": ("トランスミッションシャフト", 8200, 12500),
    "P-30008": ("燃料インジェクタノズル", 5500, 8200),
    "P-30011": ("ウェハー搬送ロボット用シャフト A", 12500, 18500),
    "P-30014": ("エッチング装置チャンバ部品", 25800, 38500),
    "P-30015": ("ウェハーチャックベース", 8500, 12800),
    "P-30020": ("検査装置ステージベース", 28500, 42500),
    "P-30021": ("エンジンマウントブラケット", 450, 680),
    "P-30022": ("サスペンションアーム A型", 850, 1280),
    "P-30027": ("ロボットアーム外装パネル A", 2800, 4200),
}


# ============================================================
# PLC-I-001 棚卸計画書
# ============================================================
def gen_inventory_plan():
    pdf = JPPDF()
    pdf.add_page()

    pdf.h1("2025年度下期 実地棚卸計画書")
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, "作成日: 2025年8月28日 / 作成者: 佐藤 一郎（経理部長 ACC001）",
             align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.h2("1. 実施目的")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5, "財務諸表に計上される棚卸資産の実在性・網羅性・評価の検証のため、"
                         "在庫管理規程R14に基づき全数実地棚卸を実施する。")
    pdf.ln(3)

    pdf.h2("2. 実施日程")
    pdf.table_header(["日程", "対象倉庫", "内容"], [40, 55, 90])
    pdf.table_row(["2025/9/26(金) 午後", "本社倉庫A", "休業日、全数棚卸"], [40, 55, 90])
    pdf.table_row(["2025/9/27(土) 全日", "本社倉庫B", "休業日、全数棚卸"], [40, 55, 90], fill=True)
    pdf.table_row(["2025/9/27(土) 全日", "東北工場倉庫", "休業日、全数棚卸"], [40, 55, 90])
    pdf.table_row(["2025/9/28(日)", "差異調整・報告書作成", "各倉庫で実施"], [40, 55, 90], fill=True)
    pdf.ln(5)

    pdf.h2("3. 実施体制")
    pdf.table_header(["倉庫", "リーダー", "計数者(2名体制)", "立会者"], [35, 40, 65, 45])
    pdf.table_row(["本社倉庫A", "橋本 明 (WHS001)",
                   "池田 昌夫、森 和雄、外部応援3名",
                   "中村 真理 (ACC004)"], [35, 40, 65, 45])
    pdf.table_row(["本社倉庫B", "橋本 明 (WHS001)",
                   "ムソクA社派遣5名",
                   "高橋 美咲 (ACC002)"], [35, 40, 65, 45], fill=True)
    pdf.table_row(["東北工場倉庫", "東北工場長",
                   "東北工場員10名",
                   "佐藤 一郎 (ACC001)"], [35, 40, 65, 45])
    pdf.ln(5)

    pdf.h2("4. 手順")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
                   "① 棚卸票（SAP MI01で出力、製品コード順）を各区画ごとに配布\n"
                   "② 計数者2名体制で実地数量を記入、双方が署名\n"
                   "③ 差異発生時は即座に倉庫課長（橋本）に報告\n"
                   "④ 経理部立会者が無作為抽出で再計数（10区画/倉庫）\n"
                   "⑤ 全区画完了後、SAP MI07で確定入力・差異レポート出力\n"
                   "⑥ 差異調整仕訳を起票、経理部長承認を得る")
    pdf.ln(5)

    pdf.h2("5. 差異発生時の処理")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
                   "在庫管理規程R14 §6に従い、以下の基準で原因分析を実施：\n"
                   "・数量差異±1%以上 または 金額差異10万円以上：原因分析書を作成\n"
                   "・経理部長への報告、調整仕訳の起票\n"
                   "・必要に応じて業務プロセスの見直し")
    pdf.ln(10)

    pdf.set_font("YuGoth", "", 9)
    pdf.cell(0, 5, "承認: 2025/8/28 / 渡辺 正博 CFO [印]", new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(BASE / "PLC-I-001_実地棚卸計画書_2025下期.pdf"))
    print("Created: PLC-I-001_実地棚卸計画書_2025下期.pdf")


# ============================================================
# PLC-I-001 棚卸票・棚卸報告書（Excel）
# ============================================================
def gen_inventory_report():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "実地棚卸報告"

    ws.cell(row=1, column=1, value="【PLC-I-001 統制実施記録】 2025年9月 実地棚卸報告書")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    ws.cell(row=2, column=1, value="棚卸実施日: 2025/9/26-27 / 報告書作成日: 2025/10/2 / 作成: 橋本 明（倉庫課長 WHS001）/ 承認: 佐藤 一郎（経理部長）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    # 倉庫別サマリ
    r = 4
    ws.cell(row=r, column=1, value="■ 倉庫別サマリ").font = BBOLD
    r += 1
    headers = ["倉庫", "区画数", "品目数", "帳簿金額(円)", "実地金額(円)",
               "差異金額(円)", "差異率", "差異件数", "備考"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    r += 1

    summary = [
        ("本社倉庫A", 45, 120, 1_245_800_000, 1_245_650_000, -150_000, "-0.012%", 3,
         "軽微差異、原因調査完了"),
        ("本社倉庫B", 38, 95, 982_500_000, 983_350_000, 850_000, "+0.087%", 7,
         "※内1件 ¥850,000 の原因分析未実施"),
        ("東北工場倉庫", 52, 185, 1_824_000_000, 1_823_600_000, -400_000, "-0.022%", 5,
         "原因調査完了、調整仕訳起票済"),
    ]
    for row in summary:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 7, 8):
                cell.alignment = C_
            elif c_i in (4, 5, 6):
                cell.alignment = R_
                cell.number_format = "#,##0;[Red]-#,##0"
            else:
                cell.alignment = L_
        # 本社倉庫Bの差異を強調
        if row[0] == "本社倉庫B":
            for c_i in (6, 9):
                ws.cell(row=r, column=c_i).fill = FILL_NG
        r += 1

    # 合計
    ws.cell(row=r, column=1, value="合計").font = BBOLD
    ws.cell(row=r, column=1).alignment = C_
    ws.cell(row=r, column=2, value=135).font = BBOLD
    ws.cell(row=r, column=3, value=400).font = BBOLD
    ws.cell(row=r, column=4, value=4_052_300_000).font = BBOLD
    ws.cell(row=r, column=4).number_format = "#,##0"
    ws.cell(row=r, column=4).alignment = R_
    ws.cell(row=r, column=5, value=4_052_600_000).font = BBOLD
    ws.cell(row=r, column=5).number_format = "#,##0"
    ws.cell(row=r, column=5).alignment = R_
    ws.cell(row=r, column=6, value=300_000).font = BBOLD
    ws.cell(row=r, column=6).number_format = "#,##0;[Red]-#,##0"
    ws.cell(row=r, column=6).alignment = R_
    ws.cell(row=r, column=7, value="+0.007%").font = BBOLD
    ws.cell(row=r, column=7).alignment = C_
    ws.cell(row=r, column=8, value=15).font = BBOLD
    ws.cell(row=r, column=8).alignment = C_

    # 差異明細（本社倉庫B）
    r += 3
    ws.cell(row=r, column=1, value="■ 本社倉庫B 差異明細（抜粋）").font = BBOLD
    r += 1
    headers2 = ["差異№", "区画", "製品コード", "品名", "帳簿数", "実地数",
                "差異数", "差異金額(円)", "原因分析の有無"]
    for i, h in enumerate(headers2, 1):
        c = ws.cell(row=r, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    r += 1

    random.seed(888)
    details = [
        ("INV-DIFF-2025-09-012", "B-1", "P-30001", "エンジンピストンピン A型",
         1200, 1195, -5, -14_250, "有：出庫処理遅延と判明"),
        ("INV-DIFF-2025-09-013", "B-2", "P-30022", "サスペンションアーム A型",
         3500, 3498, -2, -2_560, "有：軽微、出荷時カウント誤差"),
        ("INV-DIFF-2025-09-014", "B-2", "P-30008", "燃料インジェクタノズル",
         80, 83, 3, 16_500, "有：生産完了の反映遅延"),
        ("INV-DIFF-2025-09-015", "B-3", "P-30006", "トランスミッションシャフト",
         450, 518, 68, 850_000,
         "※ 未実施（不備検出）"),
        ("INV-DIFF-2025-09-016", "B-4", "P-30011", "ウェハー搬送ロボット用シャフト A",
         200, 199, -1, -12_500, "有：出荷伝票との整合確認済"),
        ("INV-DIFF-2025-09-017", "B-5", "P-30015", "ウェハーチャックベース",
         150, 152, 2, 17_000, "有：製造指図完了の反映遅延"),
        ("INV-DIFF-2025-09-018", "B-6", "P-30027", "ロボットアーム外装パネル A",
         80, 79, -1, -2_800, "有：出庫時の端数処理"),
    ]
    for d in details:
        for c_i, v in enumerate(d, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 5, 6, 7):
                cell.alignment = C_
            elif c_i == 8:
                cell.alignment = R_
                cell.number_format = "#,##0;[Red]-#,##0"
            else:
                cell.alignment = L_
        # 不備ケースを強調
        if "未実施" in d[8]:
            for c_i in range(1, 10):
                ws.cell(row=r, column=c_i).fill = FILL_NG
        r += 1

    # 不備箇所のコメント
    r += 2
    ws.cell(row=r, column=1, value="■ 内部監査指摘事項").font = BBOLD
    r += 1
    ws.cell(row=r, column=1,
            value="倉庫B 区画B-3 で発見された差異 +68個（¥850,000）について、原因分析書が作成されておらず、"
                  "経理部への報告も行われていない。在庫管理規程R14 §6 違反。"
                  "内部監査により軽微な不備として認定、是正指示中。")
    ws.cell(row=r, column=1).font = BFONT
    ws.cell(row=r, column=1).fill = FILL_NG
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)

    widths = [22, 8, 12, 32, 10, 10, 10, 14, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-I-001_実地棚卸報告書_2025年9月.xlsx")
    print("Created: PLC-I-001_実地棚卸報告書_2025年9月.xlsx")


# ============================================================
# PLC-I-002 棚卸差異分析書（存在する分）
# ============================================================
def gen_diff_analysis_pdf():
    pdf = JPPDF()
    pdf.add_page()

    pdf.h1("棚卸差異分析書")
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, "差異№: INV-DIFF-2025-09-012 / 作成日: 2025/9/29",
             align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.kv("対象倉庫・区画", "本社倉庫A / 区画A-3")
    pdf.kv("製品コード", "P-30001 エンジンピストンピン A型")
    pdf.kv("発生日", "2025/9/26 (実地棚卸日)")
    pdf.ln(3)

    pdf.h2("1. 差異内容")
    pdf.table_header(["項目", "数値"], [50, 100])
    pdf.table_row(["帳簿数量", "1,200個"], [50, 100])
    pdf.table_row(["実地数量", "1,195個"], [50, 100], fill=True)
    pdf.table_row(["差異", "-5個"], [50, 100])
    pdf.table_row(["標準原価", "¥2,850/個"], [50, 100], fill=True)
    pdf.table_row(["差異金額", "-¥14,250"], [50, 100])
    pdf.ln(3)

    pdf.h2("2. 原因調査")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
                   "■ 調査方法：\n"
                   "・直近30日の出荷伝票、入庫伝票をSAPから抽出（MB51）\n"
                   "・WMSの出庫ログと突合\n"
                   "・区画A-3の担当者2名にヒアリング\n\n"
                   "■ 判明した事実：\n"
                   "9月20日の出庫処理（出庫伝票 OUT-2025-9-234）で、"
                   "WMS側では出庫完了だが SAP側に連携が遅延。"
                   "担当者が手動で後日同期したが、5個分が二重計上された状態となっていた。\n\n"
                   "■ 結論：\n"
                   "システム連携の一時的な遅延と、手動修正時のダブルカウントミスが原因。"
                   "実地棚卸の実数が正しい。")
    pdf.ln(3)

    pdf.h2("3. 是正措置")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
                   "① SAPで数量調整仕訳を起票（2025/9/30）\n"
                   "② WMS-SAP連携の監視強化を情シス部に依頼\n"
                   "③ 手動同期手順の見直し（チェックリスト化）")
    pdf.ln(5)

    pdf.h3("■ 承認")
    pdf.set_font("YuGoth", "B", 10)
    pdf.cell(60, 7, "役割", border=1, align="C", fill=True)
    pdf.cell(60, 7, "氏名", border=1, align="C", fill=True)
    pdf.cell(40, 7, "承認日", border=1, align="C", fill=True)
    pdf.cell(30, 7, "承認印", border=1, align="C", fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("YuGoth", "", 10)
    approvals = [
        ("倉庫課長", "橋本 明 (WHS001)", "2025/9/29"),
        ("経理部課長", "高橋 美咲 (ACC002)", "2025/9/30"),
        ("経理部長", "佐藤 一郎 (ACC001)", "2025/10/1"),
    ]
    for role, name, dt in approvals:
        pdf.cell(60, 14, role, border=1, align="C")
        pdf.cell(60, 14, name, border=1, align="C")
        pdf.cell(40, 14, dt, border=1, align="C")
        x_stamp = pdf.get_x()
        y_stamp = pdf.get_y()
        pdf.cell(30, 14, "", border=1, new_x="LMARGIN", new_y="NEXT")
        pdf.stamp("承認", x=x_stamp + 15, y=y_stamp + 7)

    pdf.output(str(BASE / "PLC-I-002_棚卸差異分析書_INV-DIFF-2025-09-012.pdf"))
    print("Created: PLC-I-002_棚卸差異分析書_INV-DIFF-2025-09-012.pdf")


# ============================================================
# 棚卸写真JPG（複数）
# ============================================================
def gen_warehouse_photos():
    warehouse_photo(
        "本社倉庫A",
        "A-3",
        "2025/09/26 14:32",
        "橋本 明（倉庫課長 WHS001）",
        str(BASE / "PLC-I-001_棚卸写真_本社倉庫A_区画A-3.jpg"),
        scene_type="rack",
    )
    warehouse_photo(
        "本社倉庫B",
        "B-3",
        "2025/09/27 10:15",
        "橋本 明（倉庫課長 WHS001）",
        str(BASE / "PLC-I-001_棚卸写真_本社倉庫B_区画B-3_差異発生区画.jpg"),
        scene_type="rack",
    )
    warehouse_photo(
        "東北工場倉庫",
        "T-1",
        "2025/09/27 13:50",
        "東北工場長",
        str(BASE / "PLC-I-001_棚卸写真_東北工場倉庫_区画T-1.jpg"),
        scene_type="floor",
    )
    warehouse_photo(
        "本社倉庫A",
        "A-7",
        "2025/09/26 16:08",
        "中村 真理（経理部立会）",
        str(BASE / "PLC-I-001_棚卸写真_本社倉庫A_区画A-7_立会.jpg"),
        scene_type="rack",
    )
    print("Created: 4 warehouse photos")


# ============================================================
# PLC-I-003 標準原価更新稟議
# ============================================================
def gen_cost_update_ringi():
    pdf = JPPDF()
    pdf.add_page()

    pdf.h1("稟議書 — 標準原価更新")
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, "稟議番号: W-2025-0089 / 申請日: 2025年4月1日",
             align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    pdf.kv("件名", "FY2025期首 標準原価更新", key_w=30)
    pdf.kv("申請者", "伊藤 健太（経理部課長 ACC003）", key_w=30)
    pdf.kv("対象期間", "FY2025 (2025/4/1 - 2026/3/31)", key_w=30)
    pdf.ln(5)

    pdf.h2("1. 更新の必要性")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
                   "会計方針および在庫管理規程R14に基づき、標準原価は年1回（期首）見直しを行う。"
                   "FY2024実績を基に、以下の変動要因を反映した標準原価を設定する。")
    pdf.ln(3)

    pdf.h2("2. 主要変動要因")
    pdf.table_header(["要因", "変動率", "影響金額"], [60, 40, 60])
    pdf.table_row(["特殊合金材（仕入先H社）の値上げ", "+8%", "約+¥42,000,000/年"],
                  [60, 40, 60])
    pdf.table_row(["労務費（賃上げ2.8%反映）", "+2.8%", "約+¥28,000,000/年"],
                  [60, 40, 60], fill=True)
    pdf.table_row(["製造間接費配賦率の見直し", "-1.2%", "約-¥12,000,000/年"],
                  [60, 40, 60])
    pdf.table_row(["エネルギー費の上昇", "+4.5%", "約+¥8,000,000/年"],
                  [60, 40, 60], fill=True)
    pdf.ln(5)

    pdf.h2("3. 主要製品の標準原価変動")
    pdf.table_header(["製品コード", "品名", "旧標準原価", "新標準原価", "変動率"],
                     [25, 65, 25, 25, 20])
    pdf.table_row(["P-30011", "ウェハー搬送シャフト A", "12,000", "12,500", "+4.2%"],
                  [25, 65, 25, 25, 20])
    pdf.table_row(["P-30014", "エッチング装置部品", "24,500", "25,800", "+5.3%"],
                  [25, 65, 25, 25, 20], fill=True)
    pdf.table_row(["P-30006", "トランスミッションシャフト", "7,900", "8,200", "+3.8%"],
                  [25, 65, 25, 25, 20])
    pdf.table_row(["P-30001", "エンジンピストンピン A", "2,780", "2,850", "+2.5%"],
                  [25, 65, 25, 25, 20], fill=True)
    pdf.ln(3)
    pdf.set_font("YuGoth", "", 9)
    pdf.cell(0, 5, "※ 全30品目の詳細は添付「新旧標準原価比較表」を参照",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.h2("4. SAP反映予定")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
                   "・反映日: 2025/4/1 0:00\n"
                   "・反映方法: SAP CK11N/CK24 にて登録\n"
                   "・実施担当: 伊藤 健太、IT003 加藤 洋子が補助")
    pdf.ln(5)

    pdf.h3("■ 承認経路")
    pdf.set_font("YuGoth", "B", 10)
    pdf.cell(40, 7, "役割", border=1, align="C", fill=True)
    pdf.cell(50, 7, "氏名", border=1, align="C", fill=True)
    pdf.cell(40, 7, "承認日", border=1, align="C", fill=True)
    pdf.cell(30, 7, "承認印", border=1, align="C", fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("YuGoth", "", 10)
    approvals = [
        ("経理部長", "佐藤 一郎 (ACC001)", "2025/3/28"),
        ("CFO", "渡辺 正博 (CFO001)", "2025/3/30"),
    ]
    for role, name, dt in approvals:
        pdf.cell(40, 14, role, border=1, align="C")
        pdf.cell(50, 14, name, border=1, align="C")
        pdf.cell(40, 14, dt, border=1, align="C")
        x_stamp = pdf.get_x()
        y_stamp = pdf.get_y()
        pdf.cell(30, 14, "", border=1, new_x="LMARGIN", new_y="NEXT")
        pdf.stamp("承認", x=x_stamp + 15, y=y_stamp + 7)

    pdf.output(str(BASE / "PLC-I-003_標準原価更新稟議_W-2025-0089.pdf"))
    print("Created: PLC-I-003_標準原価更新稟議_W-2025-0089.pdf")


# ============================================================
# PLC-I-004 原価差異分析表
# ============================================================
def gen_cost_variance():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "原価差異分析"

    ws.cell(row=1, column=1, value="【PLC-I-004 統制実施記録】 2025年11月 原価差異分析表")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=2, column=1, value="作成: 伊藤 健太（経理部課長 管理会計 ACC003）/ レビュー: 森 和雄 製造本部長 / 承認: 佐藤 一郎 経理部長")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headers = ["分類", "標準原価(円)", "実際原価(円)", "差異(円)", "差異率",
               "要因分析", "重要性", "翌月標準原価反映"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[4].height = 32

    rows = [
        ("材料費(直接)", 125_800_000, 128_200_000, 2_400_000, "+1.9%",
         "V-20008の材料単価改定分のみ反映遅延", "軽微", "翌月以降反映"),
        ("労務費(直接)", 48_500_000, 49_100_000, 600_000, "+1.2%",
         "残業増加（生産計画の変動）", "軽微", "不要"),
        ("製造間接費", 62_000_000, 63_800_000, 1_800_000, "+2.9%",
         "エネルギー費・保守費の上昇", "軽微", "検討中"),
        ("外注加工費", 18_500_000, 19_700_000, 1_200_000, "+6.5%",
         "V-20023 プレス加工の単価改定", "重要(要レビュー)", "要反映"),
        ("能率差異", 0, 850_000, 850_000, "-",
         "製造ライン切替ロス", "軽微", "不要"),
        ("配賦差異", 0, -320_000, -320_000, "-",
         "稼働時間の差", "軽微", "不要"),
    ]
    r = 5
    total_std = 0
    total_act = 0
    for row in rows:
        total_std += row[1]
        total_act += row[2]
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 5, 7, 8):
                cell.alignment = C_
            elif c_i in (2, 3, 4):
                cell.alignment = R_
                cell.number_format = "#,##0;[Red]-#,##0"
            else:
                cell.alignment = L_
        if row[6] == "重要(要レビュー)":
            ws.cell(row=r, column=7).fill = FILL_WARN
        r += 1

    ws.cell(row=r, column=1, value="合計").font = BBOLD
    ws.cell(row=r, column=1).alignment = C_
    ws.cell(row=r, column=2, value=total_std).font = BBOLD
    ws.cell(row=r, column=2).number_format = "#,##0"
    ws.cell(row=r, column=2).alignment = R_
    ws.cell(row=r, column=3, value=total_act).font = BBOLD
    ws.cell(row=r, column=3).number_format = "#,##0"
    ws.cell(row=r, column=3).alignment = R_
    ws.cell(row=r, column=4, value=total_act - total_std).font = BBOLD
    ws.cell(row=r, column=4).number_format = "#,##0;[Red]-#,##0"
    ws.cell(row=r, column=4).alignment = R_

    # コメント
    r += 2
    ws.cell(row=r, column=1, value="■ 経理部・製造本部協議結果（2025/12/5 開催）").font = BBOLD
    r += 1
    ws.cell(row=r, column=1, value="・外注加工費（V-20023プレス加工）の上昇は今後継続見込み、標準原価に反映")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    ws.cell(row=r, column=1, value="・製造間接費は当面現行水準を維持、四半期ごとにモニタリング")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    widths = [16, 14, 14, 14, 10, 32, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-I-004_原価差異分析表_202511.xlsx")
    print("Created: PLC-I-004_原価差異分析表_202511.xlsx")


# ============================================================
# PLC-I-005 滞留在庫評価損計算
# ============================================================
def gen_obsolete_stock():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "滞留在庫評価損計算"

    ws.cell(row=1, column=1, value="【PLC-I-005 統制実施記録】 2025年12月末 滞留在庫評価損計算")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=2, column=1, value="基準日: 2025/12/31 / 作成: 高橋 美咲（経理部課長）/ 承認: 佐藤 一郎（経理部長）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headers = ["製品コード", "品名", "在庫数", "帳簿残高(円)", "最終出庫日",
               "回転期間", "評価率", "評価損(円)"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[4].height = 28

    random.seed(5050)
    rows = [
        ("P-30003-旧型", "バルブステム A型（旧規格）", 450, 832_500, date(2024, 8, 15), "16ヶ月", "80%", 666_000),
        ("P-30017-廃", "旧型イオン注入シャッター", 120, 816_000, date(2023, 12, 20), "24ヶ月超", "100%", 816_000),
        ("P-30001-B", "エンジンピストンA（特注品）", 80, 228_000, date(2024, 3, 10), "21ヶ月", "100%", 228_000),
        ("P-30027-V2", "ロボットパネル（V2廃番）", 200, 560_000, date(2023, 6, 5), "24ヶ月超", "100%", 560_000),
        ("P-30013-旧", "旧露光装置リング", 35, 647_500, date(2024, 10, 2), "14ヶ月", "50%", 323_750),
        ("P-30024-旧", "旧ドアヒンジ", 1500, 420_000, date(2024, 7, 1), "17ヶ月", "80%", 336_000),
        ("P-30005-限", "限定版カムシャフト", 25, 112_500, date(2024, 2, 8), "22ヶ月", "100%", 112_500),
    ]
    r = 5
    total = 0
    for row in rows:
        total += row[7]
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 3, 5, 6, 7):
                cell.alignment = C_
                if c_i == 5:
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (4, 8):
                cell.alignment = R_
                cell.number_format = "#,##0"
            else:
                cell.alignment = L_
        r += 1

    # 合計
    ws.cell(row=r, column=1, value="合計").font = BBOLD
    ws.cell(row=r, column=1).alignment = C_
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=8, value=total).font = BBOLD
    ws.cell(row=r, column=8).number_format = "#,##0"
    ws.cell(row=r, column=8).alignment = R_

    # 評価率の説明
    r += 2
    ws.cell(row=r, column=1, value="■ 評価率基準（在庫管理規程R14）").font = BBOLD
    r += 1
    ws.cell(row=r, column=1, value="回転期間 12〜18ヶ月：50% / 18〜24ヶ月：80% / 24ヶ月超：100%")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 2
    ws.cell(row=r, column=1, value="仕訳: 借方 評価損(5300) / 貸方 製品・棚卸資産評価引当").font = BBOLD
    r += 1
    ws.cell(row=r, column=1, value="承認: 高橋 美咲 [印] 2026/1/5 / 佐藤 一郎 [印] 2026/1/8")

    widths = [18, 30, 10, 14, 12, 10, 10, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-I-005_滞留在庫評価損計算_2025年12月末.xlsx")
    print("Created: PLC-I-005_滞留在庫評価損計算_2025年12月末.xlsx")


# ============================================================
# PLC-I-006 WMS-ERP照合レポート
# ============================================================
def gen_wms_erp_reconciliation():
    path = BASE / "PLC-I-006_WMS-ERP在庫照合レポート_202511月次サンプル.csv"
    lines = [
        "# WMS-ERP在庫データ照合バッチ結果",
        "# バッチ名: ZMM_STOCK_RECON",
        "# 実行日時: 2025/11/30 03:15:22",
        "# 比較対象: WMS在庫テーブル vs SAP MB52",
        "",
        "製品コード,倉庫コード,ロケーション,WMS在庫数,SAP在庫数,差異,判定,差異原因",
    ]

    random.seed(6006)
    codes = list(PRODUCTS.keys())
    for i in range(30):
        pcode = random.choice(codes)
        wh = random.choice(["WH-A", "WH-B", "WH-T"])
        loc = f"{wh}-{random.randint(1, 8)}-{random.choice('ABCDE')}"
        sap_qty = random.randint(50, 3000)
        wms_qty = sap_qty
        judge = "一致"
        cause = ""
        if random.random() < 0.1:
            wms_qty = sap_qty + random.choice([-1, 1])
            judge = "差異あり"
            cause = "同期タイミング差異（日次処理で吸収）"
        lines.append(f"{pcode},{wh},{loc},{wms_qty},{sap_qty},{wms_qty - sap_qty},{judge},{cause}")

    lines.append("")
    lines.append("# 件数: 30件 / 一致: 27件 / 差異: 3件（すべて軽微）")
    lines.append("# 確認者: 橋本 明 (WHS001) / 確認日: 2025/11/30 08:30")
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"Created: {path.name}")


# ============================================================
# PLC-I-007 月次原価計算締め
# ============================================================
def gen_monthly_close_checklist():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "月次原価計算締め"

    ws.cell(row=1, column=1, value="【PLC-I-007 統制実施記録】 2025年11月 月次原価計算締めチェックリスト")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    ws.cell(row=2, column=1, value="実施日: 2025/12/5 / 実施者: 伊藤 健太（経理部課長 ACC003）/ 承認: 佐藤 一郎（経理部長 ACC001）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    headers = ["№", "チェック項目", "実施日", "実施者", "結果", "備考"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    items = [
        (1, "製造指図の完了確認（SAP CO02）", "2025/12/2", "伊藤", "完了", "全指図完了確認済"),
        (2, "材料投入実績の確定（MB51）", "2025/12/3", "伊藤", "完了", ""),
        (3, "労務費の配賦（KKAO）", "2025/12/3", "伊藤", "完了", ""),
        (4, "製造間接費の配賦（KSU5）", "2025/12/4", "伊藤", "完了", ""),
        (5, "原価差異の計算（KKS1）", "2025/12/4", "伊藤", "完了", ""),
        (6, "差異配賦実行（CO88）", "2025/12/4", "伊藤", "完了", ""),
        (7, "棚卸資産評価の確認（MB5B）", "2025/12/4", "中村", "完了", ""),
        (8, "材料費の確定仕訳", "2025/12/5", "伊藤", "完了", ""),
        (9, "仕掛品残高の確定（MB51）", "2025/12/5", "中村", "完了", ""),
        (10, "製品残高の確定（MB52）", "2025/12/5", "中村", "完了", ""),
        # ...続きは省略
    ]
    r = 5
    for row in items:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 3, 4, 5):
                cell.alignment = C_
            else:
                cell.alignment = L_
        ws.cell(row=r, column=5).fill = FILL_OK
        r += 1

    ws.cell(row=r, column=1, value="... 以下35項目省略（全45項目中）").font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 2
    ws.cell(row=r, column=1, value="全45項目完了 / 経理部長承認: 佐藤 一郎 [印] 2025/12/6").font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

    widths = [5, 32, 12, 12, 10, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-I-007_月次原価計算締めチェックリスト_202511.xlsx")
    print("Created: PLC-I-007_月次原価計算締めチェックリスト_202511.xlsx")


# ============================================================
# SAP画面
# ============================================================
def gen_screenshots():
    sap_screenshot(
        "在庫数量一覧",
        "MB52",
        [
            ("出力日時", "2025/11/30 18:22"),
            ("プラント", "1000 本社"),
            ("倉庫", "WH-A 本社倉庫A"),
            ("出力者", "MFG001 森 和雄"),
            ("表示範囲", "全品目（0評価額を含まず）"),
            ("総品目数", "120品目"),
            ("総金額", "1,245,650,000 JPY"),
        ],
        grid_headers=["製品コード", "品名", "在庫数", "標準原価", "残高"],
        grid_rows=[
            ["P-30001", "エンジンピストンピン A型", "1,195", "2,850", "3,405,750"],
            ["P-30006", "トランスミッションシャフト", "450", "8,200", "3,690,000"],
            ["P-30011", "ウェハー搬送ロボット用シャフト A", "200", "12,500", "2,500,000"],
        ],
        status_bar="120品目表示完了",
        output_path=str(BASE / "PLC-I-001_SAP在庫数量一覧_MB52.png"),
    )
    print("Created: SAP screenshots for PLC-I")


# ============================================================
# Main
# ============================================================
if __name__ == "__main__":
    gen_inventory_plan()
    gen_inventory_report()
    gen_diff_analysis_pdf()
    gen_warehouse_photos()
    gen_cost_update_ringi()
    gen_cost_variance()
    gen_obsolete_stock()
    gen_wms_erp_reconciliation()
    gen_monthly_close_checklist()
    gen_screenshots()
    print("\nAll PLC-I evidence generated.")
