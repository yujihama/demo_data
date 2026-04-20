"""ITGC第3弾修正

指摘と対応:
1. ITGC-AC-003: SM20生ログが未整備 → 正式なSAL loginイベントログ生成 + 退職者フィルタ結果
2. ITGC-CM-001: 変更申請書PDF22件不足 + IT003役職紐づけ不可 → PDF生成 + 役職列追加
3. ITGC-CM-003: TR-REL紐づけ情報不足 → TR説明/対象オブジェクト/release承認者追加
4. ITGC-EM-001: SIer-A SOC1の不備は意図的でない → clean opinion版に再生成
"""

import csv
import io
import os
import random
import sys
from datetime import datetime, timedelta
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

ROOT = Path(r"C:\Users\nyham\work\demo_data\4.evidence\ITGC")

# ==============================================================
# Fix 1: ITGC-AC-003 Raw SM20 Security Audit Log
# ==============================================================
def fix_sm20_raw_log():
    """Generate proper SAP SM20 SAL raw log:
    - Daily login events from various non-retired users during period
    - Shows coverage of retention period
    - Final query result sections for each retired user (zero hits)
    """
    path = ROOT / "SAP_SM20_SecurityAuditLog_RetiredUsers.csv"

    # Active users that would be logging in during the audit period
    active_users = [
        ('SLS001', '田中 太郎'), ('SLS002', '斎藤 次郎'), ('SLS003', '鈴木 花子'), ('SLS004', '松本 香織'),
        ('PUR001', '木村 浩二'), ('PUR002', '小林 浩太'), ('PUR003', '清水 智明'),
        ('ACC001', '佐藤 一郎'), ('ACC002', '高橋 美咲'), ('ACC003', '渡辺 俊介'), ('ACC004', '中村 真理'), ('ACC006', '石井 健'),
        ('MFG001', '森 和雄'), ('MFG002', '池田 直樹'), ('MFG003', '山田 拓也'),
        ('HR001', '近藤 文子'), ('HR002', '岩本 涼子'),
        ('IT001', '岡田 宏'), ('IT002', '吉田 雅彦'), ('IT003', '加藤 洋子'), ('IT004', '西田 徹'),
    ]

    # Retired users (5人): show they appear in recent periods then stop
    retired_users = [
        {'uid': 'MFG099', 'name': '山崎 龍一', 'ret_date': '2025-06-30', 'stop_date': '2025-06-30',
         'period_start': '2025-06-16', 'period_end': '2025-06-30', 'delay': False},
        {'uid': 'ACC099', 'name': '北川 昭子', 'ret_date': '2025-08-31', 'stop_date': '2025-09-01',
         'period_start': '2025-08-17', 'period_end': '2025-09-01', 'delay': False},
        {'uid': 'SLS099', 'name': '藤井 修', 'ret_date': '2025-09-30', 'stop_date': '2025-10-11',
         'period_start': '2025-09-16', 'period_end': '2025-10-11', 'delay': True},
        {'uid': 'PUR099', 'name': '菅原 美奈子', 'ret_date': '2025-11-15', 'stop_date': '2025-12-03',
         'period_start': '2025-11-01', 'period_end': '2025-12-03', 'delay': True},
        {'uid': 'HR099', 'name': '大野 健介', 'ret_date': '2026-01-31', 'stop_date': '2026-02-02',
         'period_start': '2026-01-17', 'period_end': '2026-02-02', 'delay': False},
    ]

    lines = [
        "# SAP S/4HANA - Transaction SM20",
        "# Report:   Security Audit Log (SAL) - Login Events",
        "# Client:   100 (Production)",
        "# Audit Class: AU1 (Dialog Logon), AU3 (Logoff)",
        "# Retention: 730 days (Profile parameter rsau/max_file_size confirmed)",
        "# Export:   2026-02-15 14:22:30 JST by IT002 吉田 雅彦 (E0052)",
        "# Scope:    退職者5名の離任期間±2週間の全ログオンイベント",
        "",
        "# --- Section A: 期間内全ログオンイベント (母集団抜粋) ---",
        "# Filter: ログオン成功 / 期間: 各退職者離任前後",
        "",
        "タイムスタンプ,ユーザID,イベント,メッセージ,端末,ターミナル,クライアント,結果",
    ]

    rng = random.Random(51001)

    # Generate login events during each retired user's period
    all_events = []
    for ru in retired_users:
        start = datetime.strptime(ru['period_start'], '%Y-%m-%d')
        end = datetime.strptime(ru['period_end'], '%Y-%m-%d')
        cur = start
        while cur <= end:
            # Skip weekends roughly
            if cur.weekday() < 5:
                # Generate 25-40 login events per business day (other users)
                n_events = rng.randint(25, 40)
                for _ in range(n_events):
                    uid, name = rng.choice(active_users)
                    hh = rng.randint(7, 20)
                    mm = rng.randint(0, 59)
                    ss = rng.randint(0, 59)
                    ts = cur.replace(hour=hh, minute=mm, second=ss)
                    ip = f"10.20.{rng.randint(1, 30)}.{rng.randint(10, 250)}"
                    tcode = rng.choice(['SAPLSMTR_NAVIGATION', 'SESSION_MANAGER', 'SAPMSSY0'])
                    event = 'AU1'
                    msg = 'Logon successful (type=A)'
                    all_events.append((ts, uid, event, msg, ip, tcode, '100', 'SUCCESS'))
            cur += timedelta(days=1)

    # Sort by timestamp and take representative subset (too many would bloat file)
    all_events.sort(key=lambda x: x[0])
    # Sample every 15th event to keep file manageable
    sampled = all_events[::15]

    for ev in sampled:
        ts, uid, event, msg, ip, tcode, client, result = ev
        lines.append(f"{ts.strftime('%Y-%m-%d %H:%M:%S')},{uid},{event},{msg},{ip},{tcode},{client},{result}")

    lines.append("")
    lines.append(f"# Section A Records: {len(sampled)} (期間内全イベント {len(all_events)}件から15件間隔で抽出)")
    lines.append("")

    # Section B: Per-user filter query results
    lines.append("# --- Section B: 退職者個別フィルタ結果 ---")
    lines.append("# Query Syntax: SM20 → Dynamic Filters: USER = <retired_uid> / DATE BETWEEN <period>")
    lines.append("")

    for ru in retired_users:
        lines.append(f"# --- Retired User: {ru['uid']} ({ru['name']}) ---")
        lines.append(f"# 退職日: {ru['ret_date']} / 停止日: {ru['stop_date']}")
        lines.append(f"# フィルタ期間: {ru['period_start']} ～ {ru['period_end']}")
        lines.append(f"# Query: USER = '{ru['uid']}' AND DATE BETWEEN '{ru['period_start']}' AND '{ru['period_end']}'")
        # Show the retired user's login events BEFORE retirement date (to prove they existed and log was capturing them)
        pre_start = datetime.strptime(ru['period_start'], '%Y-%m-%d')
        ret_date = datetime.strptime(ru['ret_date'], '%Y-%m-%d')
        # Login events before retirement: 3-5 events
        pre_events = []
        for _ in range(rng.randint(3, 5)):
            day_offset = rng.randint(0, max(1, (ret_date - pre_start).days - 1))
            day = pre_start + timedelta(days=day_offset)
            if day.weekday() < 5:
                hh = rng.randint(8, 18)
                mm = rng.randint(0, 59)
                ss = rng.randint(0, 59)
                ip = f"10.20.{rng.randint(1, 30)}.{rng.randint(10, 250)}"
                pre_events.append((day.replace(hour=hh, minute=mm, second=ss), ip))
        pre_events.sort()

        if pre_events:
            lines.append(f"# 退職日前ログオン (参考: 通常勤務実績確認):")
            for ts, ip in pre_events:
                lines.append(f"{ts.strftime('%Y-%m-%d %H:%M:%S')},{ru['uid']},AU1,Logon successful (type=A),{ip},SESSION_MANAGER,100,SUCCESS")

        lines.append(f"# 退職日以降 ({ru['ret_date']} 翌日 ～ {ru['period_end']}):")
        lines.append(f"# → Query Result: 0 records found")
        lines.append("")

    # Final summary
    lines.append("# --- Section C: 集計 ---")
    for ru in retired_users:
        delay_note = f"[停止遅延{(datetime.strptime(ru['stop_date'],'%Y-%m-%d')-datetime.strptime(ru['ret_date'],'%Y-%m-%d')).days}日 DEF-2026-001]" if ru['delay'] else "[適時停止]"
        lines.append(f"# {ru['uid']}: 退職日 {ru['ret_date']} / 停止日 {ru['stop_date']} / 退職後ログオン件数=0 {delay_note}")

    with open(path, 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))

    print(f"[Fixed] SM20 SAL raw log: generated with {len(sampled)} login events + per-user filter results")


# ==============================================================
# Fix 2A: ITGC-CM-001 Generate 22 missing 変更申請書 PDFs
# ==============================================================
def generate_change_request_pdfs():
    """Generate 22 missing 変更申請書 PDFs for samples 4-25 (REL-008 to REL-050 even numbers).
    Remove REL-023 PDF (non-sample)."""
    from fpdf import FPDF
    import warnings
    warnings.filterwarnings('ignore', category=DeprecationWarning)

    # Read Detailed CSV
    detailed_path = ROOT / "ChangeManagement_Register_Detailed_FY2025.csv"
    sample_info = {}
    with open(detailed_path, encoding='utf-8') as f:
        for line in f:
            if line.startswith('#') or line.startswith('サンプル') or not line.strip():
                continue
            parts = line.strip().split(',')
            if len(parts) < 11: continue
            sno, rel, apl, applicant, content, risk, test, mv, app_it, app_biz, status = parts[:11]
            sample_info[rel] = {
                'sno': sno, 'apl': apl, 'applicant': applicant, 'content': content,
                'test': test, 'mv': mv
            }

    FONT_REG = r"C:\Windows\Fonts\YuGothM.ttc"
    FONT_BLD = r"C:\Windows\Fonts\YuGothB.ttc"

    # Scope mapping
    def scope_of(content):
        if '売上' in content or '販売価格' in content: return '販売管理'
        if '購買' in content or '仕入先' in content: return '購買/MM'
        if '標準原価' in content: return '原価計算'
        if 'ワークフロー' in content: return '全業務'
        if 'セキュリティ' in content or 'バックアップ' in content: return '基盤'
        if '勘定科目' in content or '連結' in content: return '経理/FI'
        return '業務共通'

    # Test result mapping
    def test_cases_of(content):
        return {
            '売上レポート機能追加': 5,
            '購買申請画面の改善': 4,
            '販売価格マスタ連携IF修正': 7,
            '標準原価計算バッチ修正': 5,
            'セキュリティパッチ適用': 4,
            'バックアップバッチ改善': 4,
            '仕入先マスタ項目拡張': 4,
            'ワークフロー承認ルーティング変更': 6,
            '連結仕訳バリデーション強化': 4,
            '勘定科目マスタ追加': 3,
        }.get(content, 4)

    generated = 0
    for i in range(4, 26):  # sample 4-25 → REL-008 to REL-050 (even)
        rel = f"REL-2025-{i*2:03d}"
        if rel not in sample_info:
            continue
        info = sample_info[rel]
        pdf_path = ROOT / f"変更申請書_{rel}.pdf"

        if pdf_path.exists():
            continue  # skip existing

        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.add_page()
        pdf.add_font('YG', '', FONT_REG, uni=True)
        pdf.add_font('YGB', '', FONT_BLD, uni=True)

        # Title
        pdf.set_font('YGB', '', 16)
        pdf.cell(0, 10, 'SAP変更申請書 (Change Request)', ln=1)

        pdf.set_font('YG', '', 10)
        pdf.cell(0, 6, f'REL番号: {rel} / 申請日: {info["apl"]}', ln=1, align='R')
        pdf.ln(3)

        # Basic info
        pdf.set_font('YG', '', 10)
        rows = [
            ('REL番号', rel),
            ('申請日', info['apl']),
            ('申請者', info['applicant']),
            ('変更内容', info['content']),
            ('影響範囲', scope_of(info['content'])),
            ('リスク評価', '低リスク'),
            ('テスト結果', info['test']),
            ('本番移送予定日', info['mv']),
        ]
        for label, val in rows:
            pdf.set_fill_color(240, 240, 240)
            pdf.cell(40, 8, label, border=1, align='C', fill=True)
            pdf.cell(135, 8, str(val), border=1)
            pdf.ln()
        pdf.ln(4)

        # Change details
        pdf.set_font('YGB', '', 12)
        pdf.cell(0, 8, '1. 変更理由・背景', ln=1)
        pdf.set_font('YG', '', 10)
        pdf.multi_cell(0, 6, f'業務要件に基づき、{info["content"]}を実施する。 業務部門からの改善要望を受け、計画的に実装・テスト・本番移送を実施する。')
        pdf.ln(3)

        pdf.set_font('YGB', '', 12)
        pdf.cell(0, 8, '2. テスト計画', ln=1)
        pdf.set_font('YG', '', 10)
        n_cases = test_cases_of(info['content'])
        pdf.multi_cell(0, 6, f'UATテスト {n_cases}ケースを計画。DEV環境 → QAS環境 → 単体テスト → 結合テスト → UAT → 本番移送の順で実施。')
        pdf.ln(3)

        # Approval
        pdf.set_font('YGB', '', 12)
        pdf.cell(0, 8, '■ 承認経路', ln=1)
        pdf.set_font('YG', '', 10)
        pdf.set_fill_color(48, 84, 150)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(55, 8, '役割', border=1, align='C', fill=True)
        pdf.cell(65, 8, '氏名', border=1, align='C', fill=True)
        pdf.cell(35, 8, '承認日', border=1, align='C', fill=True)
        pdf.cell(20, 8, '承認印', border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)

        # 申請者
        pdf.cell(55, 10, '申請者', border=1, align='C')
        pdf.cell(65, 10, info['applicant'], border=1, align='C')
        pdf.cell(35, 10, info['apl'], border=1, align='C')
        pdf.set_text_color(200, 0, 0)
        pdf.cell(20, 10, '申請', border=1, align='C')
        pdf.set_text_color(0, 0, 0)
        pdf.ln()

        # 情シス部アプリチームリーダー E0053 加藤 洋子 (IT003)
        apl_d = datetime.strptime(info['apl'], '%Y-%m-%d')
        it_appr_d = apl_d + timedelta(days=3)
        pdf.cell(55, 10, '情シス部アプリチームリーダー', border=1, align='C')
        pdf.cell(65, 10, '加藤 洋子 (IT003 / E0053)', border=1, align='C')
        pdf.cell(35, 10, it_appr_d.strftime('%Y-%m-%d'), border=1, align='C')
        pdf.set_text_color(200, 0, 0)
        pdf.cell(20, 10, '承認', border=1, align='C')
        pdf.set_text_color(0, 0, 0)
        pdf.ln()

        # 業務部門長承認
        biz_appr_d = apl_d + timedelta(days=5)
        pdf.cell(55, 10, '業務部門長', border=1, align='C')
        pdf.cell(65, 10, '業務部門長', border=1, align='C')
        pdf.cell(35, 10, biz_appr_d.strftime('%Y-%m-%d'), border=1, align='C')
        pdf.set_text_color(200, 0, 0)
        pdf.cell(20, 10, '承認', border=1, align='C')
        pdf.set_text_color(0, 0, 0)
        pdf.ln()

        pdf.output(str(pdf_path))
        generated += 1

    # Also regenerate REL-002, 004, 006 to include new approval format (IT003 / E0053)
    for sno in [1, 2, 3]:
        rel = f"REL-2025-{sno*2:03d}"
        if rel not in sample_info:
            continue
        info = sample_info[rel]
        pdf_path = ROOT / f"変更申請書_{rel}.pdf"

        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.add_page()
        pdf.add_font('YG', '', FONT_REG, uni=True)
        pdf.add_font('YGB', '', FONT_BLD, uni=True)

        pdf.set_font('YGB', '', 16)
        pdf.cell(0, 10, 'SAP変更申請書 (Change Request)', ln=1)

        pdf.set_font('YG', '', 10)
        pdf.cell(0, 6, f'REL番号: {rel} / 申請日: {info["apl"]}', ln=1, align='R')
        pdf.ln(3)

        rows = [
            ('REL番号', rel),
            ('申請日', info['apl']),
            ('申請者', info['applicant']),
            ('変更内容', info['content']),
            ('影響範囲', scope_of(info['content'])),
            ('リスク評価', '低リスク'),
            ('テスト結果', info['test']),
            ('本番移送予定日', info['mv']),
        ]
        for label, val in rows:
            pdf.set_fill_color(240, 240, 240)
            pdf.cell(40, 8, label, border=1, align='C', fill=True)
            pdf.cell(135, 8, str(val), border=1)
            pdf.ln()
        pdf.ln(4)

        pdf.set_font('YGB', '', 12)
        pdf.cell(0, 8, '1. 変更理由・背景', ln=1)
        pdf.set_font('YG', '', 10)
        pdf.multi_cell(0, 6, f'業務要件に基づき、{info["content"]}を実施する。業務部門からの改善要望を受け、計画的に実装・テスト・本番移送を実施する。')
        pdf.ln(3)

        pdf.set_font('YGB', '', 12)
        pdf.cell(0, 8, '2. テスト計画', ln=1)
        pdf.set_font('YG', '', 10)
        n_cases = test_cases_of(info['content'])
        pdf.multi_cell(0, 6, f'UATテスト {n_cases}ケースを計画。DEV環境 → QAS環境 → 単体テスト → 結合テスト → UAT → 本番移送の順で実施。')
        pdf.ln(3)

        pdf.set_font('YGB', '', 12)
        pdf.cell(0, 8, '■ 承認経路', ln=1)
        pdf.set_font('YG', '', 10)
        pdf.set_fill_color(48, 84, 150)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(55, 8, '役割', border=1, align='C', fill=True)
        pdf.cell(65, 8, '氏名', border=1, align='C', fill=True)
        pdf.cell(35, 8, '承認日', border=1, align='C', fill=True)
        pdf.cell(20, 8, '承認印', border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)

        pdf.cell(55, 10, '申請者', border=1, align='C')
        pdf.cell(65, 10, info['applicant'], border=1, align='C')
        pdf.cell(35, 10, info['apl'], border=1, align='C')
        pdf.set_text_color(200, 0, 0); pdf.cell(20, 10, '申請', border=1, align='C')
        pdf.set_text_color(0, 0, 0); pdf.ln()

        apl_d = datetime.strptime(info['apl'], '%Y-%m-%d')
        it_appr_d = apl_d + timedelta(days=3)
        pdf.cell(55, 10, '情シス部アプリチームリーダー', border=1, align='C')
        pdf.cell(65, 10, '加藤 洋子 (IT003 / E0053)', border=1, align='C')
        pdf.cell(35, 10, it_appr_d.strftime('%Y-%m-%d'), border=1, align='C')
        pdf.set_text_color(200, 0, 0); pdf.cell(20, 10, '承認', border=1, align='C')
        pdf.set_text_color(0, 0, 0); pdf.ln()

        biz_appr_d = apl_d + timedelta(days=5)
        pdf.cell(55, 10, '業務部門長', border=1, align='C')
        pdf.cell(65, 10, '業務部門長', border=1, align='C')
        pdf.cell(35, 10, biz_appr_d.strftime('%Y-%m-%d'), border=1, align='C')
        pdf.set_text_color(200, 0, 0); pdf.cell(20, 10, '承認', border=1, align='C')
        pdf.set_text_color(0, 0, 0); pdf.ln()

        pdf.output(str(pdf_path))
        generated += 1

    # Remove non-sample PDF (REL-023)
    non_sample = ROOT / "変更申請書_REL-2025-023.pdf"
    if non_sample.exists():
        non_sample.unlink()
        print(f"[Removed] 変更申請書_REL-2025-023.pdf (non-sample)")

    print(f"[Fixed] Change request PDFs: generated/regenerated {generated} with IT003/E0053 role shown")


# ==============================================================
# Fix 2B: ITGC-CM-001 Add 役職 column to Register
# ==============================================================
def add_role_to_register():
    """Add 役職 columns to Change Register (xlsx) and Detailed CSV"""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    # Detailed CSV
    detailed_path = ROOT / "ChangeManagement_Register_Detailed_FY2025.csv"
    lines = []
    with open(detailed_path, encoding='utf-8') as f:
        for line in f:
            if line.startswith('#') or not line.strip():
                lines.append(line.rstrip('\r\n'))
                continue
            if line.startswith('サンプル'):
                # header row - add 役職 columns
                parts = line.strip().split(',')
                # Insert after 承認者(IT) and 承認者(業務)
                new_parts = parts[:8] + ['承認者(IT)役職'] + [parts[8]] + ['承認者(業務)役職'] + [parts[9]] + parts[10:]
                lines.append(','.join(new_parts))
            else:
                parts = line.strip().split(',')
                if len(parts) >= 11:
                    # position 8 = app_it, position 9 = app_biz
                    app_it_role = 'アプリチームリーダー'
                    app_biz_role = '業務部門長'
                    new_parts = parts[:8] + [app_it_role] + [parts[8]] + [app_biz_role] + [parts[9]] + parts[10:]
                    lines.append(','.join(new_parts))
                else:
                    lines.append(line.rstrip('\r\n'))

    with open(detailed_path, 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))

    print(f"[Fixed] ChangeManagement_Register_Detailed_FY2025.csv: added 役職 columns")

    # xlsx Register
    reg_xlsx = ROOT / "ChangeManagement_Register_FY2025.xlsx"
    wb = load_workbook(reg_xlsx)
    ws = wb.active

    # Find header row (row 4)
    # Current cols: REL番号, 申請日, 申請者, 変更内容概要, 影響範囲, テスト実施, 本番移送日, 承認者1(情シス), 承認者2(業務), ステータス
    # Insert 承認者1役職 after col 8, 承認者2役職 after col 9 (shifted)

    # Easier approach: overwrite cells with 役職 info included
    # Col 8 = 承認者1(情シス) currently "加藤 洋子 (IT003)" → change to "加藤 洋子 (IT003 / E0053) / アプリチームリーダー"
    # Col 9 = 承認者2(業務) → append 役職

    for r in range(5, ws.max_row + 1):
        c8 = ws.cell(r, 8).value
        if c8 and '加藤' in str(c8):
            ws.cell(r, 8, '加藤 洋子 (IT003/E0053) アプリチームリーダー')
        c9 = ws.cell(r, 9).value
        if c9:
            s = str(c9)
            if '部門長' in s and '/' not in s:
                # generic
                ws.cell(r, 9, f'{s} / 業務部門長')
            elif '岡田' in s:
                ws.cell(r, 9, '岡田 宏 (IT001/E0051) 情報システム部長')

    # Update header
    ws.cell(4, 8, '承認者1(情シス:役職付)')
    ws.cell(4, 9, '承認者2(業務:役職付)')

    # Adjust widths
    ws.column_dimensions['H'].width = 36
    ws.column_dimensions['I'].width = 32

    wb.save(reg_xlsx)

    print(f"[Fixed] ChangeManagement_Register_FY2025.xlsx: role info embedded in approver columns")


# ==============================================================
# Fix 3: ITGC-CM-003 Add TR description + object list + release approver
# ==============================================================
def enhance_stms():
    """Add TR description, specific object list, release approver to STMS files"""
    # Object list per content type
    obj_map = {
        '売上レポート機能追加': 'ZSALES_REPORT / ZTBL_SALES_AGG / ZFM_SALES_SUMM',
        '購買申請画面の改善': 'ZPUR_REQ_UI / ZTBL_PUR_REQ_EXT / ZFM_PUR_REQ_SUBMIT',
        '販売価格マスタ連携IF修正': 'ZPRICE_IF_BATCH / ZFM_VK12_UPD / KONV (table)',
        '標準原価計算バッチ修正': 'ZCOST_CALC_BATCH / ZFM_CK40N_WRAPPER / MBEW',
        'セキュリティパッチ適用': 'Note 3452108 / Note 3461234 / RFC basis libs',
        'バックアップバッチ改善': 'ZBACKUP_JOB / RSBDCSUB / DB13 definitions',
        '仕入先マスタ項目拡張': 'LFA1 / LFB1 / ZMM_VENDOR_EXT / ZFM_VENDOR_UPD',
        'ワークフロー承認ルーティング変更': 'WS20000001 / WS20000015 / ZTBL_WF_ROUTE',
        '連結仕訳バリデーション強化': 'ZFM_CONSOL_VALIDATE / ZTBL_CONSOL_RULE',
        '勘定科目マスタ追加': 'SKA1 / SKB1 / SKAT',
    }

    # Process both STMS files
    for fname, has_sample_col in [
        ('SAP_STMS_ProductionTransport_History_FY2025.csv', True),
        ('SAP_STMS_ProductionTransport_History_FY2025_Population.csv', False),
    ]:
        path = ROOT / fname
        if not path.exists():
            continue

        # Read detailed CSV for content mapping
        detailed_path = ROOT / "ChangeManagement_Register_Detailed_FY2025.csv"
        rel_content = {}
        with open(detailed_path, encoding='utf-8') as f:
            for line in f:
                if line.startswith('#') or line.startswith('サンプル') or not line.strip():
                    continue
                parts = line.strip().split(',')
                if len(parts) >= 11:
                    rel_content[parts[1]] = parts[4]

        # Also read Register xlsx for population REL mapping
        from openpyxl import load_workbook
        reg_xlsx = ROOT / "ChangeManagement_Register_FY2025.xlsx"
        if reg_xlsx.exists():
            wb = load_workbook(reg_xlsx)
            ws = wb.active
            for r in range(5, ws.max_row + 1):
                rel = ws.cell(r, 1).value
                content = ws.cell(r, 4).value
                if rel and content:
                    rel_content.setdefault(str(rel), str(content))

        # Read current STMS
        with open(path, encoding='utf-8') as f:
            lines = f.readlines()

        header_end = False
        new_lines = []
        for line in lines:
            if line.startswith('#'):
                new_lines.append(line.rstrip('\r\n'))
            elif line.startswith('タイムスタンプ'):
                # Enhance header
                parts = line.strip().split(',')
                if has_sample_col:
                    # タイムスタンプ,サンプル№,TR番号,REL番号,移送者(SAP),移社員番号,移送元,移送先,対象オブジェクト,結果
                    new_header = 'タイムスタンプ,サンプル№,TR番号,TR説明(REL紐づけ),REL番号,対象オブジェクト,移送者(SAP),移送者(社員番号),移送承認者(SAP),移送承認者(社員番号),移送元,移送先,結果'
                else:
                    new_header = 'タイムスタンプ,TR番号,TR説明(REL紐づけ),REL番号,対象オブジェクト,移送者(SAP),移送者(社員番号),移送承認者(SAP),移送承認者(社員番号),移送元,移送先,結果'
                new_lines.append(new_header)
                header_end = True
            elif line.strip() and header_end:
                parts = line.strip().split(',')
                if has_sample_col and len(parts) >= 10:
                    ts, sno, tr, rel, mover_sap, mover_emp, src, dst, obj_old, result = parts[:10]
                    content = rel_content.get(rel, '-')
                    tr_desc = f"{rel}:{content[:20]}"
                    obj_new = obj_map.get(content, obj_old)
                    # Release approver: IT001/E0051 (情シス部長) for critical changes OR IT003 self-release for low risk
                    approver_sap = 'IT001 岡田 宏'
                    approver_emp = 'E0051'
                    new_lines.append(f"{ts},{sno},{tr},{tr_desc},{rel},{obj_new},{mover_sap},{mover_emp},{approver_sap},{approver_emp},{src},{dst},{result}")
                elif not has_sample_col and len(parts) >= 9:
                    ts, tr, rel, mover_sap, mover_emp, src, dst, obj_old, result = parts[:9]
                    content = rel_content.get(rel, '-')
                    tr_desc = f"{rel}:{content[:20]}"
                    obj_new = obj_map.get(content, obj_old)
                    approver_sap = 'IT001 岡田 宏'
                    approver_emp = 'E0051'
                    new_lines.append(f"{ts},{tr},{tr_desc},{rel},{obj_new},{mover_sap},{mover_emp},{approver_sap},{approver_emp},{src},{dst},{result}")
                else:
                    new_lines.append(line.rstrip('\r\n'))
            else:
                if line.strip():
                    new_lines.append(line.rstrip('\r\n'))

        with open(path, 'w', encoding='utf-8', newline='') as f:
            f.write('\n'.join(new_lines))

        print(f"[Fixed] {fname}: added TR説明/対象オブジェクト詳細/移送承認者 columns")


# ==============================================================
# Fix 4: ITGC-EM-001 SIer-A SOC1 → clean opinion
# ==============================================================
def fix_siera_soc1():
    """Regenerate SIer-A SOC1 with unqualified (clean) opinion - align with 運用評価=有効/例外0"""
    from fpdf import FPDF
    import warnings
    warnings.filterwarnings('ignore', category=DeprecationWarning)

    FONT_REG = r"C:\Windows\Fonts\YuGothM.ttc"
    FONT_BLD = r"C:\Windows\Fonts\YuGothB.ttc"

    pdf_path = ROOT / "SOC1_TypeII_Report_SIerA_FY2024.pdf"

    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.add_font('YG', '', FONT_REG, uni=True)
    pdf.add_font('YGB', '', FONT_BLD, uni=True)

    # Cover page
    pdf.add_page()
    pdf.set_font('YGB', '', 20)
    pdf.ln(40)
    pdf.cell(0, 12, 'SOC 1 Type II Report', ln=1, align='C')
    pdf.set_font('YG', '', 12)
    pdf.ln(4)
    pdf.multi_cell(0, 7, 'Report on Management\'s Description of a Service Organization\'s System and the Suitability of the Design and Operating Effectiveness of Controls', align='C')
    pdf.ln(20)
    pdf.set_font('YGB', '', 14)
    pdf.cell(0, 10, 'Service Organization: 外部委託先SIer-A', ln=1, align='C')
    pdf.ln(10)
    pdf.set_font('YG', '', 11)
    pdf.cell(0, 7, 'Report Period: April 1, 2024 - March 31, 2025', ln=1, align='C')
    pdf.cell(0, 7, 'Issued: May 20, 2025', ln=1, align='C')
    pdf.ln(60)
    pdf.cell(0, 7, 'Prepared in accordance with SSAE No. 18 (AT-C 320)', ln=1, align='C')
    pdf.cell(0, 7, 'By: Independent Service Auditor XYZ CPA Firm', ln=1, align='C')

    # TOC
    pdf.add_page()
    pdf.set_font('YGB', '', 16)
    pdf.cell(0, 10, 'Table of Contents', ln=1, align='C')
    pdf.ln(4)
    pdf.set_font('YG', '', 11)
    toc = [
        ("Section I   Independent Service Auditor's Report", "3"),
        ("Section II  Management's Assertion", "5"),
        ("Section III Description of the Service Organization's System", "7"),
        ("   III-1 Company Overview", "7"),
        ("   III-2 Scope of Services", "8"),
        ("   III-3 System Components", "9"),
        ("   III-4 Control Environment", "10"),
        ("Section IV  Control Objectives, Related Controls, and Test Results", "12"),
        ("   IV-1 Logical Access", "12"),
        ("   IV-2 Change Management", "14"),
        ("   IV-3 Backup and Recovery", "16"),
        ("   IV-4 Incident Management", "18"),
        ("   IV-5 Physical Security", "20"),
        ("Section V   Complementary User Entity Controls (CUECs)", "22"),
    ]
    for entry, page in toc:
        pdf.cell(150, 7, entry)
        pdf.cell(0, 7, f'... {page}', ln=1, align='R')

    # Section I - Independent Service Auditor's Report (CLEAN OPINION)
    pdf.add_page()
    pdf.set_font('YGB', '', 14)
    pdf.cell(0, 10, "Section I. Independent Service Auditor's Report", ln=1)
    pdf.ln(2)
    pdf.set_font('YG', '', 10)
    pdf.multi_cell(0, 6, "To the Management of 外部委託先SIer-A and other specified parties:")
    pdf.ln(3)
    pdf.multi_cell(0, 6,
        "Scope: We have examined 外部委託先SIer-A's description of its SAP ERP Operations and Development Services "
        "(the System) throughout the period April 1, 2024 to March 31, 2025, and the suitability of the design and operating "
        "effectiveness of the controls stated in the description to achieve the related control objectives.")
    pdf.ln(3)
    pdf.multi_cell(0, 6,
        "Opinion: In our opinion, in all material respects, (1) the description fairly presents the System; (2) the controls were "
        "suitably designed; and (3) the controls operated effectively throughout the period to provide reasonable assurance that the "
        "control objectives were achieved.")
    pdf.ln(3)
    pdf.multi_cell(0, 6,
        "Restriction on Use: This report is intended solely for the information and use of management of 外部委託先SIer-A, "
        "user entities of the System, and their auditors.")
    pdf.ln(10)
    pdf.cell(0, 7, 'XYZ CPA Firm', ln=1, align='R')
    pdf.cell(0, 7, 'Tokyo, Japan / May 20, 2025', ln=1, align='R')

    # Section II - Management's Assertion
    pdf.add_page()
    pdf.set_font('YGB', '', 14)
    pdf.cell(0, 10, "Section II. Management's Assertion", ln=1)
    pdf.ln(2)
    pdf.set_font('YG', '', 10)
    pdf.multi_cell(0, 6,
        "We, the management of 外部委託先SIer-A, have prepared the accompanying description of the SAP ERP Operations "
        "and Development Services System (the System) throughout the period April 1, 2024 to March 31, 2025. We confirm that:")
    pdf.ln(3)
    pdf.multi_cell(0, 6, "a. The description fairly presents the System.")
    pdf.multi_cell(0, 6, "b. The controls related to the control objectives stated in the description were suitably designed.")
    pdf.multi_cell(0, 6, "c. The controls operated effectively throughout the period.")
    pdf.multi_cell(0, 6, "d. Complementary user entity controls (CUECs) are described in Section V.")
    pdf.ln(10)
    pdf.cell(0, 7, '外部委託先SIer-A 代表取締役 [Chief Executive Officer]', ln=1, align='R')

    # Section III
    pdf.add_page()
    pdf.set_font('YGB', '', 14)
    pdf.cell(0, 10, "Section III. Description of the Service Organization's System", ln=1)
    pdf.ln(2)
    pdf.set_font('YGB', '', 12)
    pdf.cell(0, 8, "III-1. Company Overview", ln=1)
    pdf.set_font('YG', '', 10)
    pdf.multi_cell(0, 6,
        "外部委託先SIer-A is a system integration services provider specializing in SAP ERP implementation, operations, and "
        "enhancement services for manufacturing and service industries in Japan. Established in 1985, the company "
        "operates from headquarters in Tokyo with delivery centers in Osaka and Fukuoka. As of December 2024, the "
        "company employs approximately 1,500 consultants and engineers.")
    pdf.ln(3)
    pdf.set_font('YGB', '', 12)
    pdf.cell(0, 8, "III-2. Scope of Services Provided to 株式会社テクノプレシジョン", ln=1)
    pdf.set_font('YG', '', 10)
    pdf.multi_cell(0, 6,
        "外部委託先SIer-A provides the following services to the user entity:\n"
        "(1) SAP S/4HANA application support (L2/L3 support)\n"
        "(2) Custom development and enhancements (ABAP programming)\n"
        "(3) Change management and deployment coordination\n"
        "(4) Testing support (UAT coordination)\n"
        "(5) Periodic SAP upgrade management")
    pdf.ln(2)
    pdf.multi_cell(0, 6,
        "Note: Service scope does NOT include: infrastructure management, database administration, and backup "
        "operations (handled by 外部委託先B社 separately).")
    pdf.ln(3)
    pdf.set_font('YGB', '', 12)
    pdf.cell(0, 8, "III-3. System Components", ln=1)
    pdf.set_font('YG', '', 10)
    pdf.multi_cell(0, 6,
        "The System includes:\n"
        "- SAP development environment (DEV client 100)\n"
        "- SAP quality assurance environment (QAS client 200)\n"
        "- Change request management tool (ServiceNow-equivalent)\n"
        "- Source code repository (Git-based)\n"
        "- Incident management system")
    pdf.ln(3)
    pdf.set_font('YGB', '', 12)
    pdf.cell(0, 8, "III-4. Control Environment", ln=1)
    pdf.set_font('YG', '', 10)
    pdf.multi_cell(0, 6,
        "外部委託先SIer-A maintains the following control environment elements:\n"
        "- ISO 27001:2022 certification (latest recertification: 2024-09)\n"
        "- Annual employee code of conduct training\n"
        "- Segregation of duties between development and production access\n"
        "- Background check for all consultants handling client systems")

    # Section IV - All with "No exceptions"
    pdf.add_page()
    pdf.set_font('YGB', '', 14)
    pdf.cell(0, 10, "Section IV. Control Objectives, Related Controls, and Test Results", ln=1)
    pdf.ln(2)

    def iv_section(title, objective, rows):
        pdf.set_font('YGB', '', 12)
        pdf.cell(0, 8, title, ln=1)
        pdf.set_font('YG', '', 10)
        pdf.multi_cell(0, 6, f'Control Objective: {objective}')
        pdf.ln(2)
        pdf.set_fill_color(48, 84, 150)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(70, 8, 'Control Activity', border=1, align='C', fill=True)
        pdf.cell(70, 8, 'Test Performed', border=1, align='C', fill=True)
        pdf.cell(35, 8, 'Result', border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)
        for activity, test, result in rows:
            pdf.cell(70, 8, activity, border=1)
            pdf.cell(70, 8, test, border=1)
            pdf.cell(35, 8, result, border=1)
            pdf.ln()
        pdf.ln(3)

    iv_section("IV-1. Logical Access",
        "Controls provide reasonable assurance that logical access to the System is restricted to authorized individuals.",
        [
            ('New user access requires manager approval', 'Inspected 25 new user requests', 'No exceptions'),
            ('Quarterly access reviews performed', 'Inspected 4 quarterly reviews', 'No exceptions'),
            ('Terminated user access removed within 1 day', 'Tested 10 terminations', 'No exceptions'),
        ])

    iv_section("IV-2. Change Management",
        "Controls provide reasonable assurance that changes are authorized, tested, and approved before production implementation.",
        [
            ('All changes require written approval', 'Inspected 30 change requests', 'No exceptions'),
            ('Changes tested in QAS before PRD', 'Inspected 30 test records', 'No exceptions'),
        ])

    pdf.add_page()

    pdf.set_font('YGB', '', 12)
    pdf.cell(0, 8, "IV-3. Backup and Recovery (Reference only - performed by B社)", ln=1)
    pdf.set_font('YG', '', 10)
    pdf.multi_cell(0, 6,
        "This control area is primarily the responsibility of 外部委託先B社 (the infrastructure provider). "
        "外部委託先SIer-A coordinates restore testing with B社.")
    pdf.ln(3)

    iv_section("IV-4. Incident Management",
        "Controls provide reasonable assurance that incidents are identified, tracked, and resolved timely.",
        [
            ('All incidents logged within 1 hour', 'Inspected 45 incidents', 'No exceptions'),
            ('Root cause analysis for severity 1', 'Inspected 5 severity-1 incidents', 'No exceptions'),
        ])

    iv_section("IV-5. Physical Security",
        "Access to development/QAS facilities is restricted to authorized personnel.",
        [
            ('Badge access to development areas', 'Inspected access logs (3 months)', 'No exceptions'),
        ])

    # Section V
    pdf.add_page()
    pdf.set_font('YGB', '', 14)
    pdf.cell(0, 10, "Section V. Complementary User Entity Controls (CUECs)", ln=1)
    pdf.ln(2)
    pdf.set_font('YG', '', 10)
    pdf.multi_cell(0, 6,
        "外部委託先SIer-A assumes that user entities (including 株式会社テクノプレシジョン) will implement the following controls:")
    pdf.ln(3)
    pdf.multi_cell(0, 6, "CUEC-1: User entity management will review and authorize all change requests before SIer-A implements them.")
    pdf.ln(2)
    pdf.multi_cell(0, 6, "CUEC-2: User entity will perform User Acceptance Testing (UAT) for all changes.")
    pdf.ln(2)
    pdf.multi_cell(0, 6, "CUEC-3: User entity will maintain its own access review process and notify SIer-A promptly of user access changes.")
    pdf.ln(2)
    pdf.multi_cell(0, 6, "CUEC-4: User entity will monitor incident resolution and escalate as needed.")

    pdf.output(str(pdf_path))
    print(f"[Fixed] SOC1 SIer-A: regenerated with unqualified (clean) opinion, no Section VI")


# ==============================================================
# Main
# ==============================================================
if __name__ == '__main__':
    fix_sm20_raw_log()
    generate_change_request_pdfs()
    add_role_to_register()
    enhance_stms()
    fix_siera_soc1()
    print("\n=== ITGC ROUND 3 FIXES COMPLETED ===")
