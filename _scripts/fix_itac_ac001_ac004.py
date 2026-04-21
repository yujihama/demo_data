"""ITAC-001 / ITAC-004 エビデンス補強

ITAC-004:
- 母集団ファイル(FY2025.csv)をサンプル25件を含む42件に再生成
- WF番号/PO番号をSamples.csvと整合させる

ITAC-001:
- 5件分の与信チェックサンプル作成
- 受注→与信超過検知→WF承認→出荷指示 の完全フロー証跡
"""
import csv
import os
import random
import sys
import io
from pathlib import Path
from datetime import datetime, timedelta

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

ROOT = Path(r"C:\Users\nyham\work\demo_data")
ITAC_DIR = ROOT / "4.evidence" / "ITAC"


# ==============================================================
# ITAC-004 Fix: Regenerate FY2025.csv population to include samples
# ==============================================================
def fix_itac004_population():
    """母集団CSVを再生成: 全25サンプルWF + 17件追加 = 42件"""
    # Read sample file to extract actual sample WFs
    sample_records = {}  # wf -> {ts, po, amount, limit, approver}
    with open(ITAC_DIR / "SAP_WF_PurchaseOrder_ApprovalHistory_FY2025Samples.csv", encoding='utf-8') as f:
        current_sno = None
        current_wf = None
        for line in f:
            if line.startswith('#') or line.startswith('タイムスタンプ') or not line.strip():
                continue
            parts = line.strip().split(',')
            if len(parts) < 7: continue
            ts, wf, sno, po, actor, action, comment = parts[:7]
            if action == '起票':
                sample_records[wf] = {
                    'init_ts': ts, 'po': po, 'initiator': actor, 'sno': sno
                }
            elif action == '承認':
                if wf in sample_records:
                    # Parse amount from comment: "金額XXX,XXX円 / 上限YYY,YYY円"
                    import re
                    amt_m = re.search(r'金額([\d,]+)円', comment)
                    lim_m = re.search(r'上限([\d,]+)円', comment)
                    amount = int(amt_m.group(1).replace(',', '')) if amt_m else 0
                    limit = int(lim_m.group(1).replace(',', '')) if lim_m else 0
                    sample_records[wf]['approve_ts'] = ts
                    sample_records[wf]['approver'] = actor
                    sample_records[wf]['amount'] = amount
                    sample_records[wf]['limit'] = limit

    # Determine route based on amount and approver
    def determine_route(amount, approver):
        # R18 thresholds: 担当≤50万 / 課長≤500万 / 部長≤2000万 / 管理本部長(CFO)≤1億 / 代表取締役超過
        if amount <= 500000:
            expected = '担当単独'
            expected_approver = 'PUR003'
        elif amount <= 5000000:
            expected = '課長単独'
            expected_approver = 'PUR002'
        elif amount <= 20000000:
            expected = '課長→部長'
            expected_approver = 'PUR001'
        elif amount <= 100000000:
            expected = '課長→部長→CFO'
            expected_approver = 'CFO001'
        else:
            expected = '代表取締役'
            expected_approver = 'CEO001'
        return expected, expected_approver

    # Build population with all 25 samples + 17 additional
    rng = random.Random(77001)
    records = []

    # Add all 25 samples to population
    for wf, rec in sample_records.items():
        amount = rec['amount']
        route, _ = determine_route(amount, rec['approver'])
        records.append({
            'wf': wf,
            'po': rec['po'],
            'init_ts': rec['init_ts'],
            'amount': amount,
            'route': route,
            'approver': rec['approver'],
            'approve_ts': rec.get('approve_ts', ''),
            'is_sample': True,
        })

    # Add 17 additional non-sample records for population realism
    used_wf_ids = {int(wf.split('-')[-1]) for wf in sample_records.keys()}
    used_po_ids = {rec['po'] for rec in sample_records.values()}

    for i in range(17):
        # Generate new WF and PO numbers that don't conflict
        while True:
            wf_id = rng.randint(4000, 4999)
            if wf_id not in used_wf_ids:
                used_wf_ids.add(wf_id)
                break
        wf = f"WF-2025-{wf_id:05d}"

        while True:
            po_id = rng.randint(100, 4999)
            po = f"PO-2025-{po_id:04d}"
            if po not in used_po_ids:
                used_po_ids.add(po)
                break

        # Random amount across thresholds
        tier = rng.choices([1, 2, 3, 4], weights=[4, 5, 4, 2])[0]
        if tier == 1:
            amount = rng.randint(100000, 490000)
            approver = 'PUR003 清水 智明'
            route = '担当単独'
        elif tier == 2:
            amount = rng.randint(600000, 4900000)
            approver = 'PUR002 林 真由美'
            route = '課長単独'
        elif tier == 3:
            amount = rng.randint(5100000, 19900000)
            approver = 'PUR001 木村 浩二'
            route = '課長→部長'
        else:
            amount = rng.randint(21000000, 80000000)
            approver = 'CFO001 渡辺 正博'
            route = '課長→部長→CFO'

        # Random date in FY2025
        m = rng.randint(4, 15)
        if m > 12:
            d = datetime(2026, m - 12, rng.randint(1, 28), rng.randint(9, 17), rng.randint(0, 59))
        else:
            d = datetime(2025, m, rng.randint(1, 28), rng.randint(9, 17), rng.randint(0, 59))
        init_ts = d.strftime('%Y-%m-%d %H:%M:%S')
        approve_ts = (d + timedelta(hours=rng.randint(2, 8))).strftime('%Y-%m-%d %H:%M:%S')

        records.append({
            'wf': wf,
            'po': po,
            'init_ts': init_ts,
            'amount': amount,
            'route': route,
            'approver': approver,
            'approve_ts': approve_ts,
            'is_sample': False,
        })

    # Sort by init_ts
    records.sort(key=lambda r: r['init_ts'])

    # Write new population file
    lines = [
        "# SAP Business Workflow / 発注承認履歴 (FY2025 母集団)",
        "# 出力日時: 2026/02/12 15:10:08 JST",
        "# 出力者: IT003 加藤 洋子 (E0053 情シス部アプリチームリーダー)",
        "# 抽出条件: FY2025 期間中の発注承認ワークフロー全件",
        f"# レコード数: {len(records)}件 (うちPLC-P-002/ITAC-004監査サンプル 25件を含む)",
        "",
        "ワークフロー番号,発注番号,起票日時,起票者,金額,承認ルート,承認者,承認日時,最終ステータス,サンプル対象",
    ]

    # Approver resolution
    def approver_name(code):
        m = {
            'PUR003 清水 智明': 'PUR003 清水 智明',
            'PUR002 林 真由美': 'PUR002 林 真由美',
            'PUR001 木村 浩二': 'PUR001 木村 浩二',
            'CFO001 渡辺 正博': 'CFO001 渡辺 正博',
            'CEO001 山本 健一': 'CEO001 山本 健一',
            'PUR003': 'PUR003 清水 智明',
            'PUR002': 'PUR002 林 真由美',
            'PUR001': 'PUR001 木村 浩二',
            'CFO001': 'CFO001 渡辺 正博',
            'PUR004': 'PUR004 山田 純一',
        }
        for k, v in m.items():
            if k in code:
                return v
        return code

    for r in records:
        marker = 'Y' if r['is_sample'] else ''
        appr = approver_name(r['approver']) if r['approver'] else ''
        lines.append(f"{r['wf']},{r['po']},{r['init_ts']},PUR003 清水 智明,{r['amount']},{r['route']},{appr},{r['approve_ts']},承認完了,{marker}")

    lines.append("")
    lines.append(f"# 件数: {len(records)}件（サンプル25件+非サンプル17件）")

    with open(ITAC_DIR / "SAP_WF_PurchaseOrder_ApprovalHistory_FY2025.csv", 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))

    print(f"[Fixed] ITAC-004 FY2025 population: {len(records)} records (25 sample + 17 additional)")


# ==============================================================
# ITAC-001 Fix: Create 5-sample credit check evidence chain
# ==============================================================
def fix_itac001_evidence():
    """5件の与信チェックサンプル + WF承認 + 出荷指示のフルチェーン作成"""
    # 5 samples: 2 PASS (within limit) + 3 HOLD (exceeded → WF approved → shipped)
    samples = [
        {
            'sno': 1, 'order': 'ORD-2025-0623', 'customer_code': 'C-10008',
            'customer_name': 'サンプル顧客H社', 'order_date': '2025-06-12',
            'amount': 3200000, 'ar_balance': 45000000, 'limit': 60000000,
            'result': 'PASS', 'reason': 'WITHIN_LIMIT',
            'wf_ref': None,
            'delivery_no': 'DLV-2025-0812', 'delivery_date': '2025-06-25',
        },
        {
            'sno': 2, 'order': 'ORD-2025-1010', 'customer_code': 'C-10007',
            'customer_name': 'サンプル顧客G社', 'order_date': '2025-06-28',
            'amount': 4500000, 'ar_balance': 48000000, 'limit': 50000000,
            'result': 'HOLD', 'reason': 'CREDIT_LIMIT_EXCEEDED',
            'wf_ref': 'WF-CRD-2025-0018',
            'delivery_no': 'DLV-2025-0934', 'delivery_date': '2025-07-10',
        },
        {
            'sno': 3, 'order': 'ORD-2025-1470', 'customer_code': 'C-10003',
            'customer_name': 'サンプル顧客C社', 'order_date': '2025-08-05',
            'amount': 7800000, 'ar_balance': 120000000, 'limit': 180000000,
            'result': 'PASS', 'reason': 'WITHIN_LIMIT',
            'wf_ref': None,
            'delivery_no': 'DLV-2025-1456', 'delivery_date': '2025-08-18',
        },
        {
            'sno': 4, 'order': 'ORD-2025-1920', 'customer_code': 'C-10015',
            'customer_name': 'サンプル顧客O社', 'order_date': '2025-09-22',
            'amount': 2800000, 'ar_balance': 19000000, 'limit': 20000000,
            'result': 'HOLD', 'reason': 'CREDIT_LIMIT_EXCEEDED',
            'wf_ref': 'WF-CRD-2025-0031',
            'delivery_no': 'DLV-2025-1821', 'delivery_date': '2025-10-05',
        },
        {
            'sno': 5, 'order': 'ORD-2025-2960', 'customer_code': 'C-10011',
            'customer_name': 'サンプル顧客K社', 'order_date': '2025-12-18',
            'amount': 5500000, 'ar_balance': 38000000, 'limit': 40000000,
            'result': 'HOLD', 'reason': 'CREDIT_LIMIT_EXCEEDED',
            'wf_ref': 'WF-CRD-2025-0049',
            'delivery_no': 'DLV-2026-0234', 'delivery_date': '2025-12-30',
        },
    ]

    # --- File 1: 5-sample credit check log ---
    lines = [
        "# SAP S/4HANA - Credit Management Check Log",
        "# Module:      FD32 (Credit master) + OVAK (Credit Check Automation)",
        "# Report:      ITAC-001 監査サンプル5件の自動与信チェック結果",
        "# Export:      2026-02-10 11:22:05 JST by IT003 加藤 洋子 (E0053)",
        "",
        "サンプル№,チェック日時,受注番号,顧客コード,顧客名,受注金額(円),既存売掛金(円),与信限度額(円),チェック結果,理由コード,WF参照,出荷指示後処理",
    ]
    for s in samples:
        check_ts = f"{s['order_date']} " + ('10:15:00' if s['result'] == 'PASS' else '14:32:15')
        wf = s['wf_ref'] or '-'
        if s['result'] == 'PASS':
            post = '自動解放→出荷指示生成'
        else:
            post = 'SO_HOLD フラグ付与→承認後解放'
        lines.append(f"{s['sno']},{check_ts},{s['order']},{s['customer_code']},{s['customer_name']},{s['amount']},{s['ar_balance']},{s['limit']},{s['result']},{s['reason']},{wf},{post}")

    lines.append("")
    lines.append("# Records: 5 samples (2 PASS / 3 HOLD+承認解放)")

    with open(ITAC_DIR / "SAP_VA05_CreditCheck_SampleResults_FY2025.csv", 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))
    print("[Created] SAP_VA05_CreditCheck_SampleResults_FY2025.csv (5 samples)")

    # --- File 2: WF approval for HOLD samples (WF=営業本部長 E0021 田中 太郎) ---
    lines = [
        "# Workflow System (S04) - 与信超過例外承認履歴 (ITAC-001 サンプル)",
        "# Export:      2026-02-10 11:30:00 JST by IT003 加藤 洋子 (E0053)",
        "",
        "ワークフロー番号,サンプル№,受注番号,起票日時,起票者,承認日時,承認者,承認コメント,ステータス",
    ]
    for s in samples:
        if s['result'] == 'HOLD':
            init_ts = f"{s['order_date']} 14:40:00"
            appr_ts = f"{s['order_date']} 16:15:00"
            exceed = s['amount'] + s['ar_balance'] - s['limit']
            lines.append(f"{s['wf_ref']},{s['sno']},{s['order']},{init_ts},SLS003 鈴木 花子 (E0023),{appr_ts},田中 太郎 (SLS001/E0021 営業本部長),与信限度超過額¥{exceed:,}の発生・取引継続判断につき承認,承認完了")

    lines.append("")
    lines.append("# Records: 3 (HOLD→WF承認解放)")

    with open(ITAC_DIR / "Workflow_CreditException_SampleApproval_FY2025.csv", 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))
    print("[Created] Workflow_CreditException_SampleApproval_FY2025.csv (3 HOLD samples)")

    # --- File 3: Delivery (出荷指示) log post-approval ---
    lines = [
        "# SAP S/4HANA - VL01N Outbound Delivery Creation Log (ITAC-001 サンプル)",
        "# Export:      2026-02-10 11:45:00 JST by IT003 加藤 洋子 (E0053)",
        "",
        "サンプル№,受注番号,出荷指示番号,出荷指示日,顧客コード,WF前提,処理結果",
    ]
    for s in samples:
        prereq = '与信PASS自動解放' if s['result'] == 'PASS' else f"{s['wf_ref']} 承認完了後"
        lines.append(f"{s['sno']},{s['order']},{s['delivery_no']},{s['delivery_date']},{s['customer_code']},{prereq},出荷指示生成成功")

    lines.append("")
    lines.append("# Records: 5 (全サンプル出荷指示到達を確認)")

    with open(ITAC_DIR / "SAP_VL01N_DeliveryCreation_SampleLog_FY2025.csv", 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))
    print("[Created] SAP_VL01N_DeliveryCreation_SampleLog_FY2025.csv (5 delivery records)")

    # --- File 4: Sample list / selection basis ---
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "与信チェック5件サンプル"

    ws.cell(1, 1, "【ITGC-ITAC-001】 自動与信チェック監査対象サンプル一覧")
    ws.cell(1, 1).font = Font(bold=True, size=14)
    ws.cell(2, 1, "抽出基準: FY2025中の与信チェック実施受注から系統抽出 (PASS 2件+HOLD 3件)")
    ws.cell(3, 1, "抽出日: 2026-02-08 / 抽出者: 内部監査室 大塚 美穂 (IA002)")

    headers = ['サンプル№', '受注番号', '顧客', '受注日', '受注金額', '与信限度', '期待される挙動']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(5, c, h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='305496')

    for i, s in enumerate(samples, 6):
        ws.cell(i, 1, s['sno'])
        ws.cell(i, 2, s['order'])
        ws.cell(i, 3, s['customer_name'])
        ws.cell(i, 4, s['order_date'])
        ws.cell(i, 5, s['amount'])
        ws.cell(i, 6, s['limit'])
        if s['result'] == 'PASS':
            ws.cell(i, 7, '自動PASS→出荷指示')
        else:
            ws.cell(i, 7, 'HOLD→営業本部長承認→出荷指示')

    widths = [10, 18, 20, 12, 14, 14, 32]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+c)].width = w

    wb.save(ITAC_DIR / "CreditCheck_SampleTransactionList_FY2025.xlsx")
    print("[Created] CreditCheck_SampleTransactionList_FY2025.xlsx (sample list)")


# ==============================================================
# Update Evidence_Mapping_ITAC.csv
# ==============================================================
def update_mapping():
    path = ROOT / "2.RCM" / "Evidence_Mapping_ITAC.csv"

    new_entries = [
        ('ITAC-001', '1', 'CreditCheck_SampleTransactionList_FY2025.xlsx'),
        ('ITAC-001', '1', 'SAP_OVAK_CreditCheckConfig_Screen.png'),
        ('ITAC-001', '1', 'SAP_VA05_CreditCheck_SampleResults_FY2025.csv'),
        ('ITAC-001', '1', 'SAP_VL01N_DeliveryCreation_SampleLog_FY2025.csv'),
        ('ITAC-001', '1', 'Workflow_CreditException_SampleApproval_FY2025.csv'),
        ('ITAC-002', '1', 'SAP_MIRO_3WayMatch_ResultLog_202511.csv'),
        ('ITAC-002', '1', 'SAP_OMRK_InvoiceMatchingConfig_Screen.png'),
        ('ITAC-003', '1', 'SAP_AFAB_DepreciationRun_Screen.png'),
        ('ITAC-004', '1', 'ChangeManagement_Register_Detailed_FY2025.csv'),
        ('ITAC-004', '1', 'SAP_WF_PurchaseOrder_ApprovalHistory_FY2025.csv'),
        ('ITAC-004', '1', 'SAP_WF_PurchaseOrder_ApprovalHistory_FY2025Samples.csv'),
        ('ITAC-005', '1', 'ConsolidationSystem_PackageUpload_Log_FY2025.csv'),
    ]

    with open(path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL)
        writer.writerow(['key', 'sample_no', 'filename'])
        for row in new_entries:
            writer.writerow(row)

    print(f"[Fixed] Evidence_Mapping_ITAC.csv: {len(new_entries)} entries")


if __name__ == '__main__':
    fix_itac004_population()
    fix_itac001_evidence()
    update_mapping()
    print("\n=== ITAC-001/004 FIXES COMPLETED ===")
