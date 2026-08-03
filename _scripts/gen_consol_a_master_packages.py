# -*- coding: utf-8 -*-
"""連結決算ダミーデータ生成 パートA: 00_master / 01_packages"""
import os
from openpyxl import Workbook
from consolidation_model import (
    COMPANIES, SUBS, ALL, FX, BS_COA, PL_COA, BS_NAME, PL_NAME, PL_SIGN,
    IC_TRANSACTIONS, IC_BALANCES, IC_LOAN, TRANSIT, UNREALIZED, DIVIDENDS,
    net_income, operating_income, consolidate,
)
from xlsx_style import (
    set_widths, title, header_row, data_row, kv_block, note,
    F_BOLD, F_BODY, FILL_SECTION, FILL_TOTAL, FILL_WARN, NUM,
)

BASE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "6.consolidation")
R = consolidate()

PERIOD = "FY2025（自 2025年4月1日 至 2026年3月31日）"
Q4DATE = "2026年3月31日"


def unit_of(code):
    return "千THB" if COMPANIES[code]["currency"] == "THB" else "千円"


# ===========================================================================
# 00_master
# ===========================================================================

def gen_group_master():
    wb = Workbook()
    ws = wb.active
    ws.title = "連結グループ会社マスタ"
    set_widths(ws, [10, 30, 14, 26, 10, 9, 9, 12, 26, 30, 24])
    r = title(ws, 1, "連結グループ会社マスタ", f"デモA株式会社グループ / 基準日: {Q4DATE} / 連結決算システム(S05)会社マスタ準拠")
    r = header_row(ws, r, ["会社コード", "会社名", "区分", "所在地", "決算日", "通貨", "持分比率", "連結方法", "設立・取得", "主要事業", "経理責任者(パッケージ作成)"])
    for code in ALL:
        c = COMPANIES[code]
        own = "―" if c["ownership"] is None else f"{c['ownership']*100:.0f}%"
        method = "―（親会社）" if code == "DA-HQ" else "全部連結"
        r = data_row(ws, r, [code, c["name"], c["role"], c["location"], c["fye"], c["currency"], own, method, c["established"], c["business"], c.get("preparer", "―")])
    r += 1
    r = note(ws, r, "※ 会社コードは連結決算システム(S05)のアップロードログ・バリデーションログと共通。")
    r = note(ws, r, "※ Demo-A (Thailand) はJPY建て取引（対親会社輸出入・借入金）を有する。換算は「03_translation」参照。")

    ws2 = wb.create_sheet("資本構成・株式情報")
    set_widths(ws2, [10, 30, 14, 14, 14, 16, 30])
    r = title(ws2, 1, "子会社 資本構成", "（単位：千円）※DATはTHB建て資本金を取得時レート3.00円/THBで換算")
    r = header_row(ws2, r, ["会社コード", "会社名", "資本金", "資本剰余金", "親会社投資簿価", "非支配株主持分", "備考"])
    inv = {"DA-TB": 100000, "DA-LOG": 30000, "DAT": 90000, "DATR": 560000}
    for code in SUBS:
        c = COMPANIES[code]
        cap = c["bs_close"]["E100"] if code != "DAT" else 90000
        surp = c["bs_close"].get("E110", 0)
        nci = 142000 if code == "DATR" else 0
        memo = {"DA-TB": "設立時出資", "DA-LOG": "設立時出資",
                "DAT": "資本金 THB30,000千（登記上）", "DATR": "2019/4/1 80%取得・のれん160,000千円計上"}[code]
        r = data_row(ws2, r, [code, c["name"], cap, surp, inv[code], nci or "―", memo], numcols=(3, 4, 5, 6))
    r = data_row(ws2, r, ["", "合計", 300000, 20000, 780000, 142000, "資本金消去合計＝S05資本連結仕訳と一致"], numcols=(3, 4, 5, 6), bold=True, fill=FILL_TOTAL)
    wb.save(os.path.join(BASE, "00_master", "連結グループ会社マスタ.xlsx"))


def gen_coa_mapping():
    wb = Workbook()
    ws = wb.active
    ws.title = "連結科目マスタ"
    set_widths(ws, [10, 30, 12, 40])
    r = title(ws, 1, "連結勘定科目マスタ（連結パッケージ共通科目）", "連結決算システム(S05) 科目体系")
    r = header_row(ws, r, ["連結科目コード", "科目名", "区分", "備考"])
    sec_name = {"CA": "流動資産", "FA": "固定資産", "CL": "流動負債", "NCL": "固定負債", "EQ": "純資産"}
    for code, name, sec in BS_COA:
        memo = ""
        if code == "A215":
            memo = "連結手続上のみ発生（子会社パッケージでは使用しない）"
        if code == "A115":
            memo = "控除項目（マイナス表示）"
        if code in ("E140", "E150"):
            memo = "連結手続上のみ発生"
        r = data_row(ws, r, [code, name, "BS/" + sec_name[sec], memo])
    for code, name, kind in PL_COA:
        r = data_row(ws, r, [code, name, "PL/" + ("収益" if kind == "rev" else "費用"), ""])

    ws2 = wb.create_sheet("各社科目マッピング")
    set_widths(ws2, [10, 26, 22, 22, 24, 22])
    r = title(ws2, 1, "各社勘定科目 → 連結科目 マッピング表",
              "DA-HQ/DA-TB: SAP S/4HANA(S01) 同一科目体系 / DAT: 現地会計システム / DATR: 販売管理パッケージ")
    r = header_row(ws2, r, ["連結科目", "科目名", "DA-HQ・DA-TB (SAP)", "DAT (現地GL)", "DATR (国内パッケージ)", "備考"])
    mapping = [
        ("A100", "1110/1120/1130 現金・当座・普通預金", "1000 Cash and deposits", "100/101 現金・預金", ""),
        ("A110", "1210 受取手形 / 1220 売掛金", "1100 Trade receivables", "110 売掛金", ""),
        ("A115", "1300 貸倒引当金(流動)", "1190 Allowance for doubtful accounts", "119 貸倒引当金", "控除項目"),
        ("A120", "1410 製品 / 1420 仕掛品 / 1430 原材料 / 1440 貯蔵品", "1200 Inventories", "120 商品", ""),
        ("A130", "1500 前払費用 / 1510 仮払金", "1300 Prepaid & other current", "130 前払費用ほか", ""),
        ("A140", "1230 未収入金", "1310 Other receivables", "123 未収入金", ""),
        ("A200", "2110-2190 有形固定資産 / 2200 減価償却累計額", "1500 PP&E (net)", "150 有形固定資産", "純額表示"),
        ("A210", "2310 ソフトウェア", "―", "―", ""),
        ("A220", "2410 投資有価証券", "―", "―", ""),
        ("A230", "2420 関係会社株式", "―", "―", "連結上相殺消去"),
        ("A240", "2430 関係会社長期貸付金", "―", "―", "連結上相殺消去"),
        ("A250", "2500 繰延税金資産", "1700 Deferred tax assets", "170 繰延税金資産", ""),
        ("A260", "2600 その他投資等", "1800 Other non-current assets", "180 敷金保証金ほか", ""),
        ("L100", "3110 支払手形 / 3120 買掛金", "2000 Trade payables", "200 買掛金", ""),
        ("L110", "3130 短期借入金", "2100 Short-term borrowings", "―", ""),
        ("L120", "3135 リース債務", "―", "―", ""),
        ("L130", "3140 未払金 / 3150 未払費用 / 3170 未払消費税等 / 3180 預り金", "2200 Accrued expenses & others", "220 未払金ほか", ""),
        ("L140", "3160 未払法人税等", "2300 Income tax payable", "230 未払法人税等", ""),
        ("L150", "3200 賞与引当金", "―", "232 賞与引当金", "タイは賞与引当なし（年俸制）"),
        ("L200", "3310 長期借入金", "2500 Long-term borrowings", "―", "DATは親会社借入を含む"),
        ("L210", "3320 退職給付引当金", "―", "―", ""),
        ("L220", "3400 その他固定負債", "2600 Other non-current liabilities", "260 長期預り金ほか", ""),
        ("E100", "4110 資本金", "3000 Share capital", "300 資本金", ""),
        ("E110", "4120 資本準備金", "―", "310 資本準備金", ""),
        ("E120", "4210 利益準備金 / 4220 繰越利益剰余金", "3200 Retained earnings", "320 繰越利益剰余金", ""),
        ("P100", "5100 売上高", "4000 Sales", "500 売上高", ""),
        ("P200", "5200-5300 売上原価・棚卸資産評価損", "5000 Cost of sales", "510 売上原価", ""),
        ("P300", "6100-6500 販売費及び一般管理費", "6000 SG&A", "600 販売費及び一般管理費", ""),
        ("P400a", "7100 受取利息", "7000 Interest income", "700 受取利息", ""),
        ("P410a", "7200 支払利息", "7500 Interest expense", "750 支払利息", ""),
    ]
    for code, sap, dat, datr, memo in mapping:
        r = data_row(ws2, r, [code, BS_NAME.get(code) or PL_NAME.get(code), sap, dat, datr, memo])
    r += 1
    r = note(ws2, r, "※ FY2025 Q2において、DA-TBの新規科目（2211 建物附属設備〔賃借設備改修〕）のマスタ未連携によりバリデーションエラーが発生（S05往復ログ参照）。同月中にマスタ登録済。")
    wb.save(os.path.join(BASE, "00_master", "連結科目マスタ_科目マッピング.xlsx"))


def gen_fx_master():
    wb = Workbook()
    ws = wb.active
    ws.title = "為替レートマスタ"
    set_widths(ws, [14, 16, 16, 40])
    r = title(ws, 1, "為替レートマスタ（THB/JPY） FY2025", "経理規程(R16)に基づくグループ統一レート / 出所: A銀行公表TTMレート")
    r = header_row(ws, r, ["区分", "レート(円/THB)", "適用対象", "備考"])
    rows = [
        ("前期末日レート(CR)", FX["CR_open"], "前期末BS換算", "2025/3/31 TTM"),
        ("当期末日レート(CR)", FX["CR_close"], "当期末BS換算", "2026/3/31 TTM"),
        ("期中平均レート(AR)", FX["AR"], "PL・当期純利益換算", "FY2025 月末TTM単純平均"),
        ("取得時レート(HR)", FX["HR_capital"], "資本金換算", "2012/7 DAT設立時"),
    ]
    for a, b, c, d in rows:
        r = data_row(ws, r, [a, b, c, d], numcols=(2,))
        ws.cell(row=r - 1, column=2).number_format = "0.00"
    r += 1
    r = header_row(ws, r, ["四半期", "期中平均AR", "期末CR", "備考"])
    for q, ar, cr in FX["quarterly"]:
        memo = "Q1パッケージのDATR側取引換算誤りを4.05に統一修正（S05往復ログ参照）" if q == "Q1" else ""
        r = data_row(ws, r, [q, ar, cr, memo])
        ws.cell(row=r - 1, column=2).number_format = "0.00"
        ws.cell(row=r - 1, column=3).number_format = "0.00"
    r += 1
    r = header_row(ws, r, ["年月", "月末TTM", "", ""])
    for ym, rate in FX["monthly_ttm"]:
        r = data_row(ws, r, [ym, rate, "", ""])
        ws.cell(row=r - 1, column=2).number_format = "0.00"
    wb.save(os.path.join(BASE, "00_master", "為替レートマスタ_FY2025.xlsx"))


def gen_schedule():
    wb = Workbook()
    ws = wb.active
    ws.title = "連結決算スケジュール"
    set_widths(ws, [8, 44, 16, 16, 22, 22, 14])
    r = title(ws, 1, "FY2025 年度決算 連結決算スケジュール（Q4）",
              "決算業務規程(R17) 別表 / 作成: 経理部 高橋 美咲 / 承認: 経理部長 佐藤 一郎 (2026-03-10)")
    r = header_row(ws, r, ["No", "タスク", "開始", "期限", "担当", "関連統制", "実績"])
    rows = [
        (1, "決算スケジュール・連結パッケージ様式の各社展開", "2026-03-02", "2026-03-06", "経理部 連結G", "FCRP-001", "3/4 展開済"),
        (2, "各社単体決算確定（親会社・国内子会社）", "2026-04-01", "2026-04-06", "各社経理", "FCRP-001", "4/6 完了"),
        (3, "DAT単体決算確定・円換算前パッケージ作成", "2026-04-01", "2026-04-07", "DAT Accounting", "FCRP-002", "4/7 完了"),
        (4, "連結パッケージ提出（S05アップロード）", "2026-04-07", "2026-04-08", "各子会社経理", "FCRP-002", "4/8 全社PASS"),
        (5, "パッケージ検証（バリデーション・異常値レビュー）", "2026-04-08", "2026-04-09", "経理部 連結G", "FCRP-002", "4/9 完了"),
        (6, "内部取引照合（債権債務・取引高）・差異解消", "2026-04-08", "2026-04-10", "経理部 連結G", "FCRP-002", "4/10 未達1件整理"),
        (7, "在外子会社財務諸表の換算（CTA算定）", "2026-04-09", "2026-04-10", "経理部 連結G", "FCRP-004", "4/10 完了"),
        (8, "連結仕訳起票（S05自動仕訳＋手動仕訳）", "2026-04-08", "2026-04-10", "経理部 連結G", "FCRP-004", "4/10 起票7件"),
        (9, "連結仕訳レビュー（経理部課長）", "2026-04-08", "2026-04-10", "高橋 美咲", "FCRP-004", "4/8 レビュー"),
        (10, "未実現利益計算（S07 EUCシート）・手動仕訳", "2026-04-09", "2026-04-10", "経理部 連結G", "FCRP-003", "4/10 完了"),
        (11, "連結精算表作成・連結財務諸表ドラフト", "2026-04-13", "2026-04-15", "経理部 連結G", "FCRP-004", "4/15 完了"),
        (12, "連結仕訳承認（経理部長）", "2026-04-13", "2026-04-14", "佐藤 一郎", "FCRP-004", "4/14 承認"),
        (13, "会計上の見積りレビュー（3段階レビュー）", "2026-04-13", "2026-04-17", "経理部長・CFO", "FCRP-003", "4/17 完了"),
        (14, "決算短信ドラフト作成・開示システム(S06)入力", "2026-04-16", "2026-04-22", "経理部・経営企画部", "FCRP-005", "4/22 完了"),
        (15, "決算短信 3段階レビュー・適時開示", "2026-04-23", "2026-04-28", "CFO・社長", "FCRP-005", "4/28 開示"),
        (16, "有価証券報告書ドラフト・XBRL検証", "2026-04-20", "2026-05-01", "経理部・経営企画部", "FCRP-005", "5/1 完了"),
        (17, "取締役会承認（有価証券報告書）", "2026-05-08", "2026-05-08", "取締役会", "FCRP-005", "5/8 承認"),
        (18, "外部監査人による連結財務諸表監査対応", "2026-04-15", "2026-05-07", "経理部", "―", "対応完了"),
    ]
    for row in rows:
        r = data_row(ws, r, list(row))
    r += 1
    r = note(ws, r, "※ 四半期決算は月次+2営業日で単体締め、+6営業日でパッケージ提出（S05アップロードログ参照）。")
    r = note(ws, r, "※ 提出期限 4/8 17:30。Q4は全社期限内提出・初回PASS（再提出なし）。")
    wb.save(os.path.join(BASE, "00_master", "連結決算スケジュール_FY2025Q4.xlsx"))


# ===========================================================================
# 01_packages
# ===========================================================================

def _cover(wb, code, extra_pairs=None, checklist=None):
    c = COMPANIES[code]
    ws = wb.active
    ws.title = "表紙"
    set_widths(ws, [26, 64])
    r = title(ws, 1, "連結パッケージ（年度決算）", f"デモA株式会社グループ 連結決算システム(S05) 所定様式 PKG-FY2025-Q4")
    pairs = [
        ("会社コード", code),
        ("会社名", c["name"]),
        ("報告対象期間", PERIOD),
        ("決算基準日", Q4DATE),
        ("報告通貨", c["currency"] + ("（千THB）" if c["currency"] == "THB" else "（千円）")),
        ("持分比率（親会社）", f"{c['ownership']*100:.0f}%"),
        ("作成者", c["preparer"]),
        ("社内レビュー", c["reviewer"]),
        ("承認者", c["approver"]),
        ("S05アップロード日時", c["submitted"]),
        ("バージョン", c["version"] + "（初回提出・バリデーションPASS）"),
        ("提出期限", "2026-04-08 17:30"),
    ]
    if extra_pairs:
        pairs += extra_pairs
    r = kv_block(ws, r, pairs)
    r += 1
    r = note(ws, r, "本パッケージは連結決算システム(S05)所定の年度決算様式です。全シートの数値整合はS05バリデーション（貸借一致・内部取引突合・期首残高引継）で機械的に検証されます。")

    ws2 = wb.create_sheet("提出前チェックリスト")
    set_widths(ws2, [6, 58, 10, 46])
    r = title(ws2, 1, "提出前セルフチェックリスト", f"{c['name']} / 実施者: {c['preparer']} / 実施日: 2026-04-07")
    r = header_row(ws2, r, ["No", "チェック項目", "結果", "コメント"])
    base_items = [
        ("貸借対照表の貸借一致（期首・期末）", "○", ""),
        ("利益剰余金ロールフォワード一致（期首＋当期純利益－配当＝期末）", "○", ""),
        ("株主資本等変動計算書とBS純資産の整合", "○", ""),
        ("グループ会社別債権債務残高の相手先への事前照会実施", "○", "2026-04-03 相手先残高確認書を交換"),
        ("グループ会社別取引高（売上・仕入・役務・利息・配当）の集計", "○", ""),
        ("棚卸資産明細とBS棚卸資産の一致", "○", ""),
        ("固定資産増減表とBS・減価償却費の整合", "○", ""),
        ("引当金増減明細の作成", "○", ""),
        ("借入金・リース明細の作成（うちグループ内残高の区分）", "○", ""),
        ("税金・税効果情報の記載（実効税率差異の説明）", "○", ""),
        ("後発事象・偶発債務・担保提供の有無の確認", "○", ""),
        ("前期末残高とS05繰越残高の一致", "○", ""),
        ("CF作成用データの記載", "○", ""),
        ("経理責任者・承認者の承認取得", "○", ""),
    ]
    items = checklist if checklist else base_items
    for i, (item, res, com) in enumerate(items, 1):
        fill = FILL_WARN if res != "○" else None
        r = data_row(ws2, r, [i, item, res, com], fill=fill)
    return ws


def _bs_sheet(wb, code):
    c = COMPANIES[code]
    unit = unit_of(code)
    ws = wb.create_sheet("BS（貸借対照表）")
    set_widths(ws, [10, 34, 16, 16, 16, 34])
    ttl = "貸借対照表 / Balance Sheet" if code == "DAT" else "貸借対照表"
    r = title(ws, 1, ttl, f"{c['name']} / （単位：{unit}）")
    r = header_row(ws, r, ["科目コード", "科目", "前期末\n(2025/3/31)", "当期末\n(2026/3/31)", "増減", "主な増減理由"])
    sec_rows = {"CA": "【流動資産】", "FA": "【固定資産】", "CL": "【流動負債】", "NCL": "【固定負債】", "EQ": "【純資産】"}
    reasons = {
        ("DA-TB", "A200"): "切削ライン更新投資130,000・除却12,000",
        ("DA-TB", "L110"): "短期借入金の約定返済",
        ("DA-LOG", "L120"): "リース債務の約定弁済",
        ("DAT", "A100"): "営業CFの積み上がり（無配のため）",
        ("DAT", "L200"): "親会社借入JPY210,000千円の期末レート換算差",
        ("DATR", "A115"): "サンプル顧客R社 個別引当8,000計上",
        ("DATR", "A120"): "期末商品在庫の積み増し（うち親会社仕入分53,800）",
        ("DA-HQ", "A110"): "3月度売上増に伴う売掛金増",
    }
    cur_sec = None
    subtotal = {"assets_o": 0, "assets_c": 0}
    for bcode, name, sec in BS_COA:
        if bcode in ("A215", "E140", "E150"):
            continue
        vo = c["bs_open"].get(bcode)
        vc = c["bs_close"].get(bcode)
        if vo is None and vc is None:
            continue
        vo = vo or 0
        vc = vc or 0
        if vo == 0 and vc == 0 and bcode not in ("E110",):
            continue
        if sec != cur_sec:
            r = data_row(ws, r, [sec_rows[sec], "", "", "", "", ""], bold=True, fill=FILL_SECTION)
            cur_sec = sec
        r = data_row(ws, r, [bcode, name, vo, vc, vc - vo, reasons.get((code, bcode), "")], numcols=(3, 4, 5))
    ao = sum(c["bs_open"].get(k, 0) for k in [x for x, _, s in BS_COA if s in ("CA", "FA")])
    ac = sum(c["bs_close"].get(k, 0) for k in [x for x, _, s in BS_COA if s in ("CA", "FA")])
    lo = sum(c["bs_open"].get(k, 0) for k in [x for x, _, s in BS_COA if s in ("CL", "NCL", "EQ")])
    lc = sum(c["bs_close"].get(k, 0) for k in [x for x, _, s in BS_COA if s in ("CL", "NCL", "EQ")])
    r = data_row(ws, r, ["", "資産合計", ao, ac, ac - ao, ""], numcols=(3, 4, 5), bold=True, fill=FILL_TOTAL)
    r = data_row(ws, r, ["", "負債・純資産合計", lo, lc, lc - lo, ""], numcols=(3, 4, 5), bold=True, fill=FILL_TOTAL)
    assert ao == lo and ac == lc


def _pl_sheet(wb, code):
    c = COMPANIES[code]
    unit = unit_of(code)
    ws = wb.create_sheet("PL（損益計算書）")
    set_widths(ws, [10, 34, 16, 18, 40])
    ttl = "損益計算書 / Income Statement" if code == "DAT" else "損益計算書"
    r = title(ws, 1, ttl, f"{c['name']} / {PERIOD} / （単位：{unit}）")
    r = header_row(ws, r, ["科目コード", "科目", "当期実績", "うちグループ内\n取引額", "備考"])
    ic_sales = sum(a for s, b, _, a, _ in IC_TRANSACTIONS if s == code)
    ic_pur_cogs = sum(a for s, b, _, a, k in IC_TRANSACTIONS if b == code and k.startswith("売上原価"))
    ic_pur_sga = sum(a for s, b, _, a, k in IC_TRANSACTIONS if b == code and k.startswith("販管費"))
    if code == "DAT":
        # DAT帳簿はTHB。JPY建て取引をTHB換算表示
        ic_sales = 117073
        ic_pur_cogs = 185366
    pl = c["pl"]
    ni = net_income(pl)
    rows = [
        ("P100", "売上高", pl["P100"], ic_sales or None, ""),
        ("P200", "売上原価", pl["P200"], ic_pur_cogs or None, ""),
        (None, "売上総利益", pl["P100"] - pl["P200"], None, ""),
        ("P300", "販売費及び一般管理費", pl["P300"], ic_pur_sga or None, ""),
        (None, "営業利益", operating_income(pl), None, ""),
        ("P400a", "受取利息", pl["P400a"], None, ""),
        ("P400b", "受取配当金", pl["P400b"], 160000 if code == "DA-HQ" else None,
         "グループ内: DA-TB 120,000 / DATR 40,000" if code == "DA-HQ" else ""),
        ("P400c", "為替差益", pl["P400c"], None,
         "JPY建て借入金の換算益を含む" if code == "DAT" else ""),
        ("P400d", "その他営業外収益", pl["P400d"], None, ""),
        ("P410a", "支払利息", pl["P410a"], 1200 if code == "DAT" else None,
         "親会社借入に係る利息（THB1,200＝JPY4,920千円）" if code == "DAT" else ""),
        ("P410b", "その他営業外費用", pl["P410b"], None, ""),
        (None, "経常利益", operating_income(pl) + pl["P400a"] + pl["P400b"] + pl["P400c"] + pl["P400d"] - pl["P410a"] - pl["P410b"], None, ""),
        ("P500", "特別利益", pl["P500"], None, "投資有価証券売却益" if code == "DA-HQ" and pl["P500"] else ""),
        ("P510", "特別損失", pl["P510"], None, c.get("special_loss_note", "") if pl["P510"] else ""),
        (None, "税引前当期純利益", ni + pl["P600"], None, ""),
        ("P600", "法人税等", pl["P600"], None, c.get("tax_note", "")),
        (None, "当期純利益", ni, None, ""),
    ]
    if code == "DA-HQ":
        rows[1] = ("P200", "売上原価", pl["P200"], 4000000 + 480000 + 620000,
                   "グループ内仕入: DA-TB 4,000,000 / DAT 480,000 / 有償支給戻り 620,000")
        rows[3] = ("P300", "販売費及び一般管理費", pl["P300"], 1140000, "グループ内: DA-LOG物流委託 1,140,000")
        rows[5] = ("P400a", "受取利息", pl["P400a"], 4920, "グループ内: DAT貸付金利息 4,920")
    for bcode, name, v, icv, memo in rows:
        bold = bcode is None
        r = data_row(ws, r, [bcode or "", name, v, icv if icv else "", memo],
                     numcols=(3, 4), bold=bold, fill=FILL_TOTAL if bold else None)
    if code == "DAT":
        r += 1
        r = note(ws, r, "※ 対親会社取引はJPY建て。上表「うちグループ内取引額」は期中平均レート等によるTHB換算参考値（売上 JPY480,000千円 / 材料仕入 JPY760,000千円）。")


def _ss_sheet(wb, code):
    c = COMPANIES[code]
    unit = unit_of(code)
    ws = wb.create_sheet("株主資本等変動計算書")
    set_widths(ws, [30, 14, 14, 16, 14, 16])
    r = title(ws, 1, "株主資本等変動計算書", f"{c['name']} / {PERIOD} / （単位：{unit}）")
    r = header_row(ws, r, ["", "資本金", "資本剰余金", "利益剰余金", "自己株式", "純資産合計"])
    e100o, e110o, e120o = c["bs_open"]["E100"], c["bs_open"].get("E110", 0), c["bs_open"]["E120"]
    e130o = c["bs_open"].get("E130", 0)
    ni = net_income(c["pl"])
    div = c["dividend_paid"]
    rows = [
        ("当期首残高", e100o, e110o, e120o, e130o, e100o + e110o + e120o + e130o),
        ("剰余金の配当", 0, 0, -div, 0, -div),
        ("当期純利益", 0, 0, ni, 0, ni),
        ("当期末残高", e100o, e110o, e120o - div + ni, e130o, e100o + e110o + e120o + e130o - div + ni),
    ]
    for i, row in enumerate(rows):
        r = data_row(ws, r, list(row), numcols=(2, 3, 4, 5, 6), bold=(i in (0, 3)), fill=FILL_TOTAL if i == 3 else None)
    if code in DIVIDENDS:
        d = DIVIDENDS[code]
        r += 1
        r = note(ws, r, f"※ 配当支払日 {d['date']}：親会社 {d['to_parent']:,}千円" + (f" / 非支配株主 {d['to_nci']:,}千円" if d["to_nci"] else "（全額親会社）"))
    assert c["bs_close"]["E120"] == e120o - div + ni


def _ic_balance_sheet(wb, code):
    c = COMPANIES[code]
    unit = unit_of(code)
    ws = wb.create_sheet("グループ債権債務明細")
    set_widths(ws, [12, 26, 22, 16, 16, 44])
    r = title(ws, 1, "グループ会社別 債権債務残高明細", f"{c['name']} / 基準日: {Q4DATE} / （単位：{unit}）※JPY建て取引は円貨額を併記")
    r = header_row(ws, r, ["相手先コード", "相手先会社名", "科目", "前期末残高", "当期末残高", "備考"])
    fx_c, fx_o = FX["CR_close"], FX["CR_open"]

    def conv(jpy_c, jpy_o):
        if code == "DAT":
            return round(jpy_o / fx_o), round(jpy_c / fx_c)
        return jpy_o, jpy_c

    rows = []
    for seller, buyer, close, open_ in IC_BALANCES:
        if seller == code:
            o, cl = conv(close, open_) if code == "DAT" else (open_, close)
            memo = ""
            if code == "DAT":
                memo = f"JPY建て請求（円貨 期末{close:,}／期首{open_:,}千円をCRで換算）"
            if seller == "DA-HQ" and buyer == "DATR":
                memo = "未達商品6,200千円を含む（2026/3/30出荷・4/1相手方着荷）"
            rows.append((buyer, COMPANIES[buyer]["name"], "売掛金", o, cl, memo))
        if buyer == code:
            o, cl = conv(close, open_) if code == "DAT" else (open_, close)
            memo = ""
            if code == "DAT":
                memo = f"JPY建て請求（円貨 期末{close:,}／期首{open_:,}千円をCRで換算）"
            if buyer == "DATR" and seller == "DA-HQ":
                cl, o = TRANSIT["buyer_booked_ap"], open_
                memo = "帳簿残高。親会社側売掛金95,500千円との差異6,200千円は未達商品（4/1検収・照合済）"
            rows.append((seller, COMPANIES[seller]["name"], "買掛金", o, cl, memo))
    if code == "DA-HQ":
        rows.append(("DAT", COMPANIES["DAT"]["name"], "関係会社長期貸付金", IC_LOAN["jpy"], IC_LOAN["jpy"], IC_LOAN["note"]))
        rows.append(("DA-TB", COMPANIES["DA-TB"]["name"], "未収配当金", 0, 0, "配当120,000千円は2025/6/25受領済（期末残なし）"))
    if code == "DAT":
        rows.append(("DA-HQ", COMPANIES["DA-HQ"]["name"], "長期借入金", 52500, 50000, "JPY建て210,000千円（年2.35%・2028/3/31期限）を各期末CRで換算"))
    for row in rows:
        r = data_row(ws, r, list(row), numcols=(4, 5))
    ar_o = sum(v for _, _, k, v, _, _ in rows if k == "売掛金")
    ar_c = sum(v for _, _, k, _, v, _ in rows if k == "売掛金")
    ap_o = sum(v for _, _, k, v, _, _ in rows if k in ("買掛金",))
    ap_c = sum(v for _, _, k, _, v, _ in rows if k in ("買掛金",))
    r = data_row(ws, r, ["", "売掛金 合計", "", ar_o, ar_c, ""], numcols=(4, 5), bold=True, fill=FILL_TOTAL)
    r = data_row(ws, r, ["", "買掛金 合計", "", ap_o, ap_c, ""], numcols=(4, 5), bold=True, fill=FILL_TOTAL)
    r += 1
    r = note(ws, r, "※ グループ内債権には経理規程(R16)により貸倒引当金を設定しない。")
    r = note(ws, r, "※ 残高確認書は2026-04-03に相手先と交換済（差異は上記備考のとおり）。")


def _ic_volume_sheet(wb, code):
    c = COMPANIES[code]
    unit = unit_of(code)
    ws = wb.create_sheet("グループ取引高明細")
    set_widths(ws, [12, 26, 26, 18, 44])
    r = title(ws, 1, "グループ会社別 取引高明細", f"{c['name']} / {PERIOD} / （単位：{unit}）")
    r = header_row(ws, r, ["相手先コード", "相手先会社名", "取引内容", "当期取引高", "備考"])
    rows = []
    for seller, buyer, desc, amt, kind in IC_TRANSACTIONS:
        if seller == code:
            v = amt if code != "DAT" else 117073
            memo = "JPY建て（円貨480,000千円・期中平均換算）" if code == "DAT" else ""
            rows.append((buyer, COMPANIES[buyer]["name"], f"売上高：{desc}", v, memo))
        if buyer == code:
            v = amt if code != "DAT" else 185366
            memo = "JPY建て（円貨760,000千円・期中平均換算）" if code == "DAT" else ""
            rows.append((seller, COMPANIES[seller]["name"], f"仕入・費用：{desc}（{kind}）", v, memo))
    if code == "DA-HQ":
        rows.append(("DAT", COMPANIES["DAT"]["name"], "受取利息（長期貸付金）", 4920, "年2.35%"))
        rows.append(("DA-TB", COMPANIES["DA-TB"]["name"], "受取配当金", 120000, "2025/6/25"))
        rows.append(("DATR", COMPANIES["DATR"]["name"], "受取配当金", 40000, "2025/6/20（持分80%相当）"))
    if code == "DAT":
        rows.append(("DA-HQ", COMPANIES["DA-HQ"]["name"], "支払利息（長期借入金）", 1200, "JPY4,920千円・年2.35%"))
    if code == "DA-TB":
        rows.append(("DA-HQ", COMPANIES["DA-HQ"]["name"], "支払配当金", 120000, "2025/6/25"))
    if code == "DATR":
        rows.append(("DA-HQ", COMPANIES["DA-HQ"]["name"], "支払配当金", 40000, "2025/6/20"))
        rows.append(("（非支配株主）", "商栄興産株式会社（20%）", "支払配当金", 10000, "2025/6/20"))
    for row in rows:
        r = data_row(ws, r, list(row), numcols=(4,))
    r += 1
    r = note(ws, r, "※ 取引高は相手先と四半期ごとに照合済（02_intercompany 内部取引照合表参照）。")


def _inventory_sheet(wb, code):
    c = COMPANIES[code]
    unit = unit_of(code)
    ws = wb.create_sheet("棚卸資産明細")
    set_widths(ws, [26, 16, 16, 48])
    r = title(ws, 1, "棚卸資産明細（未実現利益計算用データ含む）", f"{c['name']} / （単位：{unit}）")
    if not c["inventory_detail"]:
        r = note(ws, r, "該当なし（当社は物流サービス業のため棚卸資産を保有していない）。")
        return
    r = header_row(ws, r, ["区分", "前期末", "当期末", "備考"])
    for k, v in c["inventory_detail"].items():
        vo = c["inventory_detail_open"].get(k, 0)
        r = data_row(ws, r, [k, vo, v, ""], numcols=(2, 3))
    r = data_row(ws, r, ["合計", sum(c["inventory_detail_open"].values()), sum(c["inventory_detail"].values()), ""],
                 numcols=(2, 3), bold=True, fill=FILL_TOTAL)
    r += 1
    r = data_row(ws, r, ["【うちグループ会社仕入分（未実現利益消去用）】", "", "", ""], bold=True, fill=FILL_SECTION)
    r = header_row(ws, r, ["仕入元", "前期末残高", "当期末残高", "売り手売上総利益率（売り手申告）"])
    unre_rows = []
    for seller, holder, inv_c, margin, urp in UNREALIZED["close"]:
        if holder == code:
            inv_o = next(x[2] for x in UNREALIZED["open"] if x[0] == seller and x[1] == holder)
            if code == "DATR":
                unre_rows.append((COMPANIES[seller]["name"], inv_o, inv_c - TRANSIT["amount"],
                                  f"{margin*100:.0f}%（手許在庫のみ。ほかに未達商品6,200千円あり）"))
            else:
                unre_rows.append((COMPANIES[seller]["name"], inv_o, inv_c, f"{margin*100:.0f}%"))
    if code == "DA-TB":
        unre_rows.append((COMPANIES["DA-HQ"]["name"] + "（有償支給材）", 48000, 55000, "0%（原価売買のため未実現利益なし）"))
    if not unre_rows:
        r = note(ws, r, "該当なし（グループ会社からの仕入在庫なし）。")
    for row in unre_rows:
        r = data_row(ws, r, list(row), numcols=(2, 3))
    if code == "DA-HQ":
        r += 1
        r = note(ws, r, "※ DA-TB仕入分は材料・仕掛品に含まれる。売り手利益率はDA-TB申告値11.0%（FY2025実績）。")


def _fixed_asset_sheet(wb, code):
    c = COMPANIES[code]
    unit = unit_of(code)
    ws = wb.create_sheet("固定資産増減表")
    set_widths(ws, [26, 14, 14, 14, 14, 14, 36])
    r = title(ws, 1, "固定資産増減表", f"{c['name']} / {PERIOD} / （単位：{unit}・純額ベース）")
    r = header_row(ws, r, ["区分", "期首帳簿価額", "取得", "減価償却費", "除売却（簿価）", "期末帳簿価額", "備考"])
    fa, ia = c["fixed_assets"], c["intangibles"]
    memo_fa = {"DA-TB": "取得: 切削ライン更新130,000 / 除却: 旧ライン設備（12月・特別損失計上）",
               "DA-HQ": "取得: 金型・検査装置等 / 除却: 旧塗装設備（簿価55,000）",
               "DAT": "取得: プレス金型 THB28,000",
               "DA-LOG": "取得: 配送車両5台",
               "DATR": "取得: 倉庫内設備"}.get(code, "")
    r = data_row(ws, r, ["有形固定資産", fa["open"], fa["capex"], fa["dep"], fa["disposal"],
                         fa["open"] + fa["capex"] - fa["dep"] - fa["disposal"], memo_fa], numcols=(2, 3, 4, 5, 6))
    if ia["open"] or ia["capex"]:
        r = data_row(ws, r, ["無形固定資産（ソフトウェア等）", ia["open"], ia["capex"], ia["amort"], 0,
                             ia["open"] + ia["capex"] - ia["amort"], ""], numcols=(2, 3, 4, 5, 6))
    tot_o = fa["open"] + ia["open"]
    tot_c = fa["open"] + fa["capex"] - fa["dep"] - fa["disposal"] + ia["open"] + ia["capex"] - ia["amort"]
    r = data_row(ws, r, ["合計", tot_o, fa["capex"] + ia["capex"], fa["dep"] + ia["amort"], fa["disposal"], tot_c, ""],
                 numcols=(2, 3, 4, 5, 6), bold=True, fill=FILL_TOTAL)
    if code == "DA-HQ":
        r += 1
        r = note(ws, r, "※ 減価償却費895,000の内訳：売上原価715,000／販管費180,000。無形固定資産償却80,000は販管費。")
    if code == "DA-TB":
        r += 1
        r = note(ws, r, "※ 減価償却費173,000の内訳：製造原価153,000／販管費20,000。")


def _provision_sheet(wb, code):
    c = COMPANIES[code]
    unit = unit_of(code)
    ws = wb.create_sheet("引当金明細")
    set_widths(ws, [30, 14, 14, 14, 14, 40])
    r = title(ws, 1, "引当金増減明細", f"{c['name']} / {PERIOD} / （単位：{unit}）")
    r = header_row(ws, r, ["区分", "期首残高", "繰入", "取崩・目的使用", "期末残高", "備考"])
    rows = []
    ado, adc = -c["bs_open"].get("A115", 0), -c["bs_close"].get("A115", 0)
    if code == "DATR":
        rows.append(("貸倒引当金（一般・0.5%）", 4900, 0, 0, 4900, "一般債権期末残高に実績繰入率0.5%"))
        rows.append(("貸倒引当金（個別）", 4675, 8000, 4675, 8000,
                     "期首: サンプル顧客N社（債権償却により取崩）/ 期末: サンプル顧客R社 民事再生（債権16,000×50%）"))
    elif adc or ado:
        rows.append(("貸倒引当金", ado, max(adc - ado, 0), max(ado - adc, 0), adc, "一般債権に対する実績率繰入"))
    bo, bc = c["bs_open"].get("L150", 0), c["bs_close"].get("L150", 0)
    if bo or bc:
        rows.append(("賞与引当金", bo, bc, bo, bc, "夏季賞与支給見込額（6月支給分）"))
    ro, rc = c["bs_open"].get("L210", 0), c["bs_close"].get("L210", 0)
    if ro or rc:
        rows.append(("退職給付引当金", ro, rc - ro if rc > ro else 0, ro - rc if ro > rc else 0, rc, "簡便法（自己都合要支給額基準）" if code != "DA-HQ" else "原則法（数理計算・割引率0.8%）"))
    for row in rows:
        r = data_row(ws, r, list(row), numcols=(2, 3, 4, 5))
    if code == "DATR":
        r += 1
        r = note(ws, r, "※ " + c["baddebt_note"])
    if code == "DAT":
        r = note(ws, r, "※ タイは年俸制のため賞与引当金なし。退職給付はLegal Severance Pay（その他固定負債に含む）。")


def _loan_sheet(wb, code):
    c = COMPANIES[code]
    unit = unit_of(code)
    ws = wb.create_sheet("借入金・リース明細")
    set_widths(ws, [26, 18, 14, 14, 12, 14, 36])
    r = title(ws, 1, "借入金・リース債務明細", f"{c['name']} / 基準日: {Q4DATE} / （単位：{unit}）")
    r = header_row(ws, r, ["借入先", "区分", "前期末残高", "当期末残高", "利率", "返済期限", "備考"])
    rows = {
        "DA-HQ": [
            ("A銀行", "短期借入金", 700000, 600000, "0.85%", "2026-09-30", "運転資金"),
            ("B銀行", "短期借入金", 500000, 400000, "0.90%", "2026-06-30", "運転資金"),
            ("A銀行", "長期借入金", 1400000, 1300000, "1.10%", "2029-03-31", "設備資金"),
            ("C銀行", "長期借入金", 1000000, 900000, "1.05%", "2028-09-30", "設備資金"),
        ],
        "DA-TB": [
            ("D銀行（仙台支店）", "短期借入金", 180000, 150000, "0.95%", "2026-08-31", "運転資金"),
            ("D銀行（仙台支店）", "長期借入金", 280000, 250000, "1.20%", "2029-06-30", "切削ライン設備資金"),
        ],
        "DA-LOG": [
            ("E リース株式会社", "リース債務", 145000, 130000, "―", "2029-03-31", "配送車両・マテハン機器（ファイナンス・リース）"),
        ],
        "DAT": [
            ("デモA株式会社（親会社）", "長期借入金", 52500, 50000, "2.35%", "2028-03-31",
             "JPY建て210,000千円。期末CR4.20（前期末4.00）で換算。グループ内取引。"),
        ],
        "DATR": [],
    }[code]
    if not rows:
        r = note(ws, r, "該当なし（無借金）。")
    for row in rows:
        r = data_row(ws, r, list(row), numcols=(3, 4))
    if code == "DAT":
        r += 1
        r = note(ws, r, "※ 換算差により為替差益THB2,500千を営業外収益（為替差益）に計上（他の外貨建取引差損益と純額表示）。")


def _tax_sheet(wb, code):
    c = COMPANIES[code]
    unit = unit_of(code)
    pl = c["pl"]
    ni = net_income(pl)
    pt = ni + pl["P600"]
    ws = wb.create_sheet("税金・税効果")
    set_widths(ws, [40, 18, 44])
    r = title(ws, 1, "税金・税効果情報", f"{c['name']} / {PERIOD} / （単位：{unit}）")
    r = header_row(ws, r, ["項目", "金額・率", "備考"])
    statutory = "20.0%" if code == "DAT" else "30.6%"
    rows = [
        ("税引前当期純利益", pt, ""),
        ("法人税、住民税及び事業税（法人税等調整額含む）", pl["P600"], ""),
        ("税負担率", f"{pl['P600']/pt*100:.1f}%", ""),
        ("法定実効税率", statutory, "タイ法人税率20%" if code == "DAT" else "法人税・住民税・事業税の合算"),
    ]
    for k, v, m in rows:
        r = data_row(ws, r, [k, v, m], numcols=(2,) if not isinstance(v, str) else ())
    r += 1
    diff = {
        "DA-HQ": "税率差異の主因：受取配当金益金不算入△2.6pt、試験研究費税額控除△0.9pt、交際費等加算+0.5pt、住民税均等割+0.2pt 等",
        "DA-TB": "税率差異の主因：交際費等加算、住民税均等割",
        "DA-LOG": "税率差異の主因：住民税均等割、軽減税率適用",
        "DAT": "タイBOI恩典は適用期間終了（FY2023）。当期は標準税率20%を適用。",
        "DATR": "税率差異の主因：交際費等加算、住民税均等割",
    }[code]
    r = note(ws, r, "※ " + diff)
    r += 1
    r = header_row(ws, r, ["繰延税金資産・負債の主な内訳", "期末残高", "備考"])
    dta_rows = {
        "DA-HQ": [("賞与引当金", 159100, ""), ("退職給付引当金", 443700, ""), ("棚卸資産評価損", 45000, ""),
                  ("減損損失・除却損否認", 38000, ""), ("その他（評価性引当額控除後）", -365800, "繰延税金資産純額 320,000")],
        "DA-TB": [("賞与引当金", 29100, ""), ("退職給付引当金", 79600, ""), ("その他", 12000, "繰延税金資産はその他投資等に含む")],
        "DA-LOG": [("賞与引当金", 10700, ""), ("その他", 4000, "")],
        "DAT": [("減価償却超過額ほか", 3200, "THB。回収可能性ありと判断")],
        "DATR": [("貸倒引当金（個別）", 2400, ""), ("賞与引当金", 13800, ""), ("その他", 3000, "")],
    }[code]
    for row in dta_rows:
        r = data_row(ws, r, list(row), numcols=(2,))


def _hc_other_sheet(wb, code):
    c = COMPANIES[code]
    ws = wb.create_sheet("従業員数・その他")
    set_widths(ws, [36, 22, 50])
    r = title(ws, 1, "従業員数・担保提供・偶発債務等", f"{c['name']} / 基準日: {Q4DATE}")
    r = header_row(ws, r, ["項目", "内容", "備考"])
    hc = {"DA-HQ": "580名（ほか平均臨時雇用 92名）", "DA-TB": "120名（ほか平均臨時雇用 35名）",
          "DA-LOG": "40名（ほか平均臨時雇用 18名）", "DAT": "80名（ほか平均臨時雇用 22名）", "DATR": "30名"}[code]
    rows = [
        ("期末従業員数", hc, "就業人員ベース"),
        ("担保提供資産", {"DA-TB": "工場財団（土地・建物）簿価480,000千円をD銀行借入の担保に供している", }.get(code, "該当なし"), ""),
        ("偶発債務・保証債務", {"DA-HQ": "DAT の現地銀行取引に係る経営指導念書の差入れ（保証予約等の債務保証には該当しない）"}.get(code, "該当なし"), ""),
        ("係争事件等", {"DATR": "サンプル顧客R社の民事再生手続に係る債権届出を実施（届出額16,000千円）"}.get(code, "該当なし"), ""),
        ("会計方針の変更", "該当なし", ""),
        ("表示方法の変更", "該当なし", ""),
    ]
    for row in rows:
        r = data_row(ws, r, list(row))


def _subsequent_sheet(wb, code):
    c = COMPANIES[code]
    ws = wb.create_sheet("後発事象・特記事項")
    set_widths(ws, [22, 88])
    r = title(ws, 1, "後発事象・特記事項", f"{c['name']} / 作成日: 2026-04-07")
    r = header_row(ws, r, ["区分", "内容"])
    items = {
        "DA-HQ": [("後発事象", "該当なし"), ("特記事項", "2026年4月開催予定の取締役会において自己株式取得枠（上限100,000千円）を審議予定（決議前のため開示後発事象には該当しない）。")],
        "DA-TB": [("後発事象", "該当なし"), ("特記事項", "2025年12月に旧切削ライン設備を除却（特別損失12,000千円）。Q3パッケージで科目誤りを訂正済（S05往復ログ参照）。")],
        "DA-LOG": [("後発事象", "該当なし"), ("特記事項", "2026年5月に配送車両8台のリース更改を予定（年間リース料増加見込 約6,000千円）。")],
        "DAT": [("後発事象", "該当なし"),
                ("特記事項", "雨季の洪水リスクに対しBCP保険を付保（保険期間2026/1-2026/12）。バーツ高進行時は輸出採算が悪化するため、JPY建て取引比率の見直しを親会社と協議中。")],
        "DATR": [("後発事象", "サンプル顧客R社の民事再生計画案が2026年4月15日に提出予定であり、弁済率次第で個別引当金8,000千円の見直しが必要となる可能性がある。"),
                 ("特記事項", "期末日直前（2026/3/30）に親会社から出荷された商品6,200千円は当社未着のため当期末在庫・買掛金に計上していない（4/1検収済・親会社と照合済）。")],
    }[code]
    for row in items:
        r = data_row(ws, r, list(row))


def _cf_data_sheet(wb, code):
    c = COMPANIES[code]
    unit = unit_of(code)
    fa, ia = c["fixed_assets"], c["intangibles"]
    pl = c["pl"]
    tax_paid = pl["P600"] + c["bs_open"].get("L140", 0) - c["bs_close"].get("L140", 0)
    ws = wb.create_sheet("CF作成用データ")
    set_widths(ws, [44, 18, 40])
    r = title(ws, 1, "連結キャッシュ・フロー計算書 作成用データ", f"{c['name']} / {PERIOD} / （単位：{unit}）")
    r = header_row(ws, r, ["項目", "金額", "備考"])
    rows = [
        ("減価償却費（有形・製造原価分含む）", fa["dep"], ""),
        ("無形固定資産償却費", ia["amort"], ""),
        ("有形固定資産の取得による支出", fa["capex"], ""),
        ("無形固定資産の取得による支出", ia["capex"], ""),
        ("固定資産除売却（簿価）", fa["disposal"], ""),
        ("利息の受取額", pl["P400a"], ""),
        ("利息の支払額", pl["P410a"], ""),
        ("配当金の受取額", pl["P400b"], ""),
        ("配当金の支払額", c["dividend_paid"], ""),
        ("法人税等の支払額（概算）", tax_paid, "P600±未払法人税等増減"),
    ]
    if code == "DA-HQ":
        rows += [("投資有価証券の取得による支出", 65000, ""),
                 ("投資有価証券の売却による収入", 50000, "簿価35,000＋売却益15,000")]
    for row in rows:
        r = data_row(ws, r, list(row), numcols=(2,))


def _cogs_sheet(wb, code):
    c = COMPANIES[code]
    if "cogs_detail" not in c:
        return
    unit = unit_of(code)
    ws = wb.create_sheet("製造原価報告書")
    set_widths(ws, [56, 18, 30])
    r = title(ws, 1, "製造原価報告書", f"{c['name']} / {PERIOD} / （単位：{unit}）")
    r = header_row(ws, r, ["区分", "金額", "備考"])
    for name, v in c["cogs_detail"]:
        bold = name in ("当期総製造費用", "売上原価") or name.startswith("当期総") or name.startswith("売上原価")
        r = data_row(ws, r, [name, v, ""], numcols=(2,), bold=bold, fill=FILL_TOTAL if bold else None)


def _sga_sheet(wb, code):
    c = COMPANIES[code]
    unit = unit_of(code)
    ws = wb.create_sheet("販管費内訳")
    set_widths(ws, [56, 18, 30])
    r = title(ws, 1, "販売費及び一般管理費 内訳", f"{c['name']} / {PERIOD} / （単位：{unit}）")
    r = header_row(ws, r, ["区分", "金額", "備考"])
    for name, v in c["sga_detail"]:
        r = data_row(ws, r, [name, v, ""], numcols=(2,))
    total = sum(v for _, v in c["sga_detail"])
    r = data_row(ws, r, ["合計", total, ""], numcols=(2,), bold=True, fill=FILL_TOTAL)
    assert total == c["pl"]["P300"], f"{code} 販管費内訳不一致 {total} vs {c['pl']['P300']}"


def gen_package(code):
    c = COMPANIES[code]
    wb = Workbook()
    extra = None
    checklist = None
    if code == "DAT":
        extra = [
            ("使用為替レート", f"BS: 期末CR {FX['CR_close']:.2f}円/THB（前期末 {FX['CR_open']:.2f}） / PL: 期中平均AR {FX['AR']:.2f} / 資本金: 取得時HR {FX['HR_capital']:.2f}"),
            ("換算実施者", "親会社 経理部 連結G（03_translation 換算ワークシート参照）"),
        ]
    if code == "DATR":
        checklist_base_replace = ("グループ会社別債権債務残高の相手先への事前照会実施", "○",
                                  "親会社宛買掛金に未達差異6,200千円あり（3/30出荷分・4/1検収）。親会社経理部と照合し差異内容確定済。")
        checklist = [
            ("貸借対照表の貸借一致（期首・期末）", "○", ""),
            ("利益剰余金ロールフォワード一致（期首＋当期純利益－配当＝期末）", "○", ""),
            ("株主資本等変動計算書とBS純資産の整合", "○", ""),
            checklist_base_replace,
            ("グループ会社別取引高（売上・仕入・役務・利息・配当）の集計", "○", ""),
            ("棚卸資産明細とBS棚卸資産の一致", "○", ""),
            ("固定資産増減表とBS・減価償却費の整合", "○", ""),
            ("引当金増減明細の作成", "○", "サンプル顧客R社 個別引当の根拠資料（民事再生手続開始決定書写し）添付"),
            ("借入金・リース明細の作成（うちグループ内残高の区分）", "○", "該当なし（無借金）"),
            ("税金・税効果情報の記載（実効税率差異の説明）", "○", ""),
            ("後発事象・偶発債務・担保提供の有無の確認", "○", "R社民事再生の進捗を後発事象に記載"),
            ("前期末残高とS05繰越残高の一致", "○", ""),
            ("CF作成用データの記載", "○", ""),
            ("経理責任者・承認者の承認取得", "○", ""),
        ]
    _cover(wb, code, extra_pairs=extra, checklist=checklist)
    _bs_sheet(wb, code)
    _pl_sheet(wb, code)
    _cogs_sheet(wb, code)
    _sga_sheet(wb, code)
    _ss_sheet(wb, code)
    _ic_balance_sheet(wb, code)
    _ic_volume_sheet(wb, code)
    _inventory_sheet(wb, code)
    _fixed_asset_sheet(wb, code)
    _provision_sheet(wb, code)
    _loan_sheet(wb, code)
    _tax_sheet(wb, code)
    _hc_other_sheet(wb, code)
    _subsequent_sheet(wb, code)
    _cf_data_sheet(wb, code)
    fname = {
        "DA-TB": "連結パッケージ_FY2025Q4_DA-TB_デモA東北.xlsx",
        "DA-LOG": "連結パッケージ_FY2025Q4_DA-LOG_デモA物流サービス.xlsx",
        "DAT": "連結パッケージ_FY2025Q4_DAT_DemoA_Thailand.xlsx",
        "DATR": "連結パッケージ_FY2025Q4_DATR_デモAトレーディング.xlsx",
    }[code]
    wb.save(os.path.join(BASE, "01_packages", fname))


def gen_parent_fs():
    code = "DA-HQ"
    c = COMPANIES[code]
    wb = Workbook()
    ws = wb.active
    ws.title = "表紙"
    set_widths(ws, [26, 64])
    r = title(ws, 1, "親会社 個別財務諸表・連結用内訳データ", "デモA株式会社 / 連結決算システム(S05)取込用 PKG-FY2025-Q4-HQ")
    r = kv_block(ws, r, [
        ("会社コード", code),
        ("会社名", c["name"]),
        ("報告対象期間", PERIOD),
        ("決算基準日", Q4DATE),
        ("報告通貨", "JPY（千円）"),
        ("作成者", c["preparer"]),
        ("レビュー", c["reviewer"]),
        ("承認者", c["approver"]),
        ("S05取込日時", "2026-04-07 18:30（SAP S/4HANA(S01) 試算表I/F自動連携）"),
        ("備考", "単体決算は2026-04-06確定（月次決算チェックリスト202603参照）"),
    ])
    _bs_sheet(wb, code)
    _pl_sheet(wb, code)
    _sga_sheet(wb, code)
    _ss_sheet(wb, code)
    _ic_balance_sheet(wb, code)
    _ic_volume_sheet(wb, code)
    _inventory_sheet(wb, code)
    _fixed_asset_sheet(wb, code)
    _provision_sheet(wb, code)
    _loan_sheet(wb, code)
    _tax_sheet(wb, code)
    _hc_other_sheet(wb, code)
    _subsequent_sheet(wb, code)
    _cf_data_sheet(wb, code)
    wb.save(os.path.join(BASE, "01_packages", "親会社個別財務諸表_FY2025Q4_DA-HQ.xlsx"))


def main():
    for sub in ("00_master", "01_packages"):
        os.makedirs(os.path.join(BASE, sub), exist_ok=True)
    gen_group_master()
    gen_coa_mapping()
    gen_fx_master()
    gen_schedule()
    for code in SUBS:
        gen_package(code)
    gen_parent_fs()
    print("パートA生成完了")


if __name__ == "__main__":
    main()
