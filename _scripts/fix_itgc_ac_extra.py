"""ITGC-AC-001 / AC-002 追加修正

AC-001:
- WF CSVに「登録」アクション追加 (E0054 IT004 全25件, IT003承認後1-2h)
- SU01タイムスタンプをWF登録アクションと一致
- ユーザ登録申請書PDFを20枚追加生成 (USER-REG-2025-0006〜0025)

AC-002:
- 四半期棚卸ワークフロー履歴CSV新規 (Q1-Q4各サイクル)
  - SUIM出力 → 各部門長配布 → 部門長レビュー → 削除申請 → 情シス部長承認
- 部門長レビュー結果xlsx (部門別必要性確認一覧)
- 権限削除申請・実施記録CSV
- 情シス部長完了承認ログ (Q1/Q2完了、Q3/Q4は部門長レビューまで完了、情シス部長承認待ち)
- Q3/Q4 SUIM xlsx ヘッダの出力日時/抽出条件を削除 (HOLD-2026-002理由維持)
"""

import csv
import io
import os
import random
import sys
from datetime import datetime, timedelta
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

ROOT = Path(r"C:\Users\nyham\work\demo_data\4.evidence\ITGC")

# ==============================================================
# AC-001 Fix A: Add 登録 action to WF CSV + align SU01 timestamp
# ==============================================================
def fix_wf_add_registration():
    """Add 登録 action (E0054 IT004 西田 徹) to workflow for all 25 samples.
    Insert 1-2 hours after IT003 approval.
    """
    wf_path = ROOT / "Workflow_UserRegistration_ApprovalHistory_FY2025.csv"
    su01_path = ROOT / "SAP_SU01_UserMaster_ChangeHistory_FY2025.csv"

    # Read WF
    wf_lines = []
    wf_rows = []
    wf_header_end = False
    with open(wf_path, encoding='utf-8') as f:
        for line in f:
            if line.startswith('#') and not wf_header_end:
                wf_lines.append(line.rstrip('\r\n'))
            elif line.startswith('タイムスタンプ'):
                wf_lines.append(line.rstrip('\r\n'))
                wf_header_end = True
            elif line.startswith('#') and wf_header_end:
                # trailer
                continue
            elif line.strip():
                parts = line.strip().split(',')
                if len(parts) >= 7:
                    wf_rows.append(parts)

    # Read SU01 to get existing timestamps & keep aligned
    su01_rows = {}
    su01_lines_header = []
    su01_header_end = False
    with open(su01_path, encoding='utf-8') as f:
        for line in f:
            if line.startswith('#') and not su01_header_end:
                su01_lines_header.append(line.rstrip('\r\n'))
            elif line.startswith('タイムスタンプ'):
                su01_lines_header.append(line.rstrip('\r\n'))
                su01_header_end = True
            elif line.startswith('#') and su01_header_end:
                continue
            elif line.strip():
                parts = line.strip().split(',')
                if len(parts) >= 10:
                    su01_rows[parts[1]] = parts  # sno -> parts

    # For each sample, find IT003 approval ts and add 登録 row with ts+90min
    new_wf_rows = []
    # Group WF rows by sno
    from collections import defaultdict
    by_sno = defaultdict(list)
    for r in wf_rows:
        by_sno[r[2]].append(r)

    rng = random.Random(12345)
    new_su01_ts = {}  # sno -> new timestamp

    for sno in sorted(by_sno.keys(), key=int):
        rows = by_sno[sno]
        new_wf_rows.extend(rows)
        # find IT003 approval
        it003_ts = None
        wfno = None
        reqno = None
        for r in rows:
            if 'IT003' in r[4]:
                it003_ts = datetime.strptime(r[0], '%Y-%m-%d %H:%M:%S')
                wfno = r[1]
                reqno = r[3]
        if it003_ts:
            # Add 登録 action 1-2 hours after
            offset_min = 60 + (int(sno) * 13) % 60
            reg_ts = it003_ts + timedelta(minutes=offset_min)
            new_wf_rows.append([
                reg_ts.strftime('%Y-%m-%d %H:%M:%S'),
                wfno,
                sno,
                reqno,
                '西田 徹 (IT004)',
                'E0054',
                '登録実行'
            ])
            new_su01_ts[sno] = reg_ts.strftime('%Y-%m-%d %H:%M:%S')

    # Sort all WF rows by timestamp
    new_wf_rows.sort(key=lambda r: r[0])

    # Write WF
    out = wf_lines + [','.join(r) for r in new_wf_rows]
    out.append('')
    out.append(f"# Records: {len(new_wf_rows)}")
    with open(wf_path, 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(out))

    # Update SU01 timestamps to match 登録実行 timestamp
    out_su01 = su01_lines_header[:]
    for sno in sorted(su01_rows.keys(), key=int):
        parts = su01_rows[sno][:]
        if sno in new_su01_ts:
            parts[0] = new_su01_ts[sno]
        out_su01.append(','.join(parts))
    out_su01.append('')
    out_su01.append(f"# Records: {len(su01_rows)}")
    with open(su01_path, 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(out_su01))

    print(f"[Fixed] WF: added 登録実行 (E0054/IT004) action for {len(new_su01_ts)} samples, total rows={len(new_wf_rows)}")
    print(f"[Fixed] SU01: aligned timestamps with WF 登録実行 action")


# ==============================================================
# AC-001 Fix B: Generate 20 more USER-REG PDFs (samples 6-25)
# ==============================================================
def generate_user_reg_pdfs():
    """Generate USER-REG-2025-0006 through 0025 PDFs based on SU01/WF content."""
    from fpdf import FPDF

    # Gather sample info
    su01_info = {}
    with open(ROOT / "SAP_SU01_UserMaster_ChangeHistory_FY2025.csv", encoding='utf-8') as f:
        for line in f:
            if line.startswith('#') or line.startswith('タイムスタンプ') or not line.strip():
                continue
            parts = line.strip().split(',')
            if len(parts) >= 10:
                # sno, reqno, uid, action, dept, role
                sno = parts[1]
                su01_info[sno] = {
                    'reqno': parts[2], 'uid': parts[3], 'dept': parts[5], 'role': parts[6]
                }

    wf_info = {}
    with open(ROOT / "Workflow_UserRegistration_ApprovalHistory_FY2025.csv", encoding='utf-8') as f:
        for line in f:
            if line.startswith('#') or line.startswith('タイムスタンプ') or not line.strip():
                continue
            parts = line.strip().split(',')
            if len(parts) < 7: continue
            sno = parts[2]
            if sno not in wf_info:
                wf_info[sno] = {'head': None, 'head_date': None, 'it': None, 'it_date': None, 'apl_date': None}
            ts = parts[0]
            actor = parts[4]
            action = parts[6]
            if action == '起票':
                wf_info[sno]['apl_date'] = ts[:10]
            elif action == '承認' and 'IT003' not in actor:
                wf_info[sno]['head'] = actor
                wf_info[sno]['head_date'] = ts[:10]
            elif action == '承認' and 'IT003' in actor:
                wf_info[sno]['it'] = actor
                wf_info[sno]['it_date'] = ts[:10]

    role_desc = {
        'SD_USER': '販売管理標準ユーザ',
        'PP_SUP': '生産管理スーパーバイザ',
        'MM_USER': '購買管理標準ユーザ',
        'FI_USER': '会計標準ユーザ',
        'HR_USER': '人事標準ユーザ',
        'BASIS': '基盤管理者ロール',
    }

    FONT_REG = r"C:\Windows\Fonts\YuGothM.ttc"
    FONT_BLD = r"C:\Windows\Fonts\YuGothB.ttc"

    # Generate PDFs for samples 6-25 (skip 1-5 already exist)
    for sno in range(6, 26):
        sno_s = str(sno)
        if sno_s not in su01_info or sno_s not in wf_info:
            continue
        s = su01_info[sno_s]
        w = wf_info[sno_s]

        reqno = s['reqno']
        pdf_path = ROOT / f"ユーザ登録申請書_{reqno}.pdf"

        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.add_page()
        pdf.add_font('YG', '', FONT_REG, uni=True)
        pdf.add_font('YGB', '', FONT_BLD, uni=True)

        # Title
        pdf.set_font('YGB', '', 18)
        pdf.cell(0, 12, 'SAPユーザ登録申請書', ln=1)

        pdf.set_font('YG', '', 10)
        pdf.cell(0, 6, f'申請番号: {reqno} / 申請日: {w["apl_date"] or "-"}', ln=1, align='R')
        pdf.ln(3)

        # Section: 基本情報 table
        pdf.set_font('YG', '', 10)
        pdf.cell(35, 8, '申請部門', border=1, align='C')
        pdf.cell(140, 8, s['dept'], border=1)
        pdf.ln()
        pdf.cell(35, 8, '申請者', border=1, align='C')
        pdf.cell(140, 8, '申請者（部門）', border=1)
        pdf.ln()
        pdf.cell(35, 8, '登録対象者', border=1, align='C')
        pdf.cell(140, 8, f'{s["uid"]} (新入社員/異動)', border=1)
        pdf.ln()
        pdf.cell(35, 8, '申請理由', border=1, align='C')
        pdf.cell(140, 8, '業務遂行に必要なアクセス権付与', border=1)
        pdf.ln()
        pdf.ln(5)

        # Roles
        pdf.set_font('YGB', '', 12)
        pdf.cell(0, 8, '1. 付与希望ロール', ln=1)
        pdf.set_font('YG', '', 10)
        pdf.set_fill_color(48, 84, 150)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(40, 8, 'ロール名', border=1, align='C', fill=True)
        pdf.cell(60, 8, '内容', border=1, align='C', fill=True)
        pdf.cell(75, 8, '業務上の必要性', border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)
        role = s['role']
        pdf.cell(40, 8, role, border=1)
        pdf.cell(60, 8, role_desc.get(role.split('+')[0], '業務機能'), border=1)
        pdf.cell(75, 8, '日常業務のため', border=1)
        pdf.ln()
        pdf.ln(5)

        # SoD
        pdf.set_font('YGB', '', 12)
        pdf.cell(0, 8, '2. 職務分掌(SoD)チェック', ln=1)
        pdf.set_font('YG', '', 10)
        pdf.cell(0, 6, '付与予定ロールの組合せについてSoD違反なし。', ln=1)
        pdf.ln(3)

        # Approval
        pdf.set_font('YGB', '', 12)
        pdf.cell(0, 8, '■ 承認経路', ln=1)
        pdf.set_font('YG', '', 10)
        pdf.set_fill_color(48, 84, 150)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(50, 8, '役割', border=1, align='C', fill=True)
        pdf.cell(60, 8, '氏名', border=1, align='C', fill=True)
        pdf.cell(40, 8, '承認日時', border=1, align='C', fill=True)
        pdf.cell(25, 8, '承認印', border=1, align='C', fill=True)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)
        pdf.cell(50, 10, '申請部門長', border=1, align='C')
        pdf.cell(60, 10, w['head'] or '部門長', border=1, align='C')
        pdf.cell(40, 10, w['head_date'] or '-', border=1, align='C')
        pdf.set_text_color(200, 0, 0)
        pdf.cell(25, 10, '承認', border=1, align='C')
        pdf.set_text_color(0, 0, 0)
        pdf.ln()
        pdf.cell(50, 10, '情シス部アプリリーダー', border=1, align='C')
        pdf.cell(60, 10, w['it'] or '加藤 洋子 (IT003)', border=1, align='C')
        pdf.cell(40, 10, w['it_date'] or '-', border=1, align='C')
        pdf.set_text_color(200, 0, 0)
        pdf.cell(25, 10, '承認', border=1, align='C')
        pdf.set_text_color(0, 0, 0)
        pdf.ln()

        pdf.output(str(pdf_path))

    print(f"[Fixed] Generated 20 USER-REG PDFs (samples 6-25)")


# ==============================================================
# AC-002 Fix: Create quarterly review evidence
# ==============================================================
def create_ac002_evidence():
    """Create evidence for quarterly access review:
    - Workflow log (4 cycles × steps)
    - 部門長レビュー結果 xlsx (per quarter)
    - 削除申請・実施記録 CSV
    - 情シス部長完了承認ログ
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    # Dept heads
    dept_heads = [
        ('営業本部', '田中 太郎 (SLS001)', 'E0021'),
        ('購買部', '木村 浩二 (PUR001)', 'E0031'),
        ('製造本部', '森 和雄 (MFG001)', 'E0041'),
        ('経理部', '佐藤 一郎 (ACC001)', 'E0011'),
        ('人事部', '近藤 文子 (HR001)', 'E0061'),
        ('情報システム部', '岡田 宏 (IT001)', 'E0051'),
        ('総務部', '前田 美香 (GA001)', 'E0071'),
    ]

    # Quarterly cycle timeline
    # Q1 (Apr-Jun SUIM出力 2025/6/30 → review 7/1-7/20 → 承認 7/25)
    # Q2 (Jul-Sep SUIM 2025/9/30 → review 10/1-10/20 → 承認 10/25)
    # Q3 (Oct-Dec SUIM 2025/12/31 → review 1/5-1/22 → 承認 1/27) ← HOLD
    # Q4 (Jan-Mar SUIM 2026/3/31 → review 4/1-4/22 → 承認 4/27) ← HOLD
    cycles = [
        {'q': 'Q1', 'suim_date': '2025-06-30', 'dist_date': '2025-07-01', 'review_end': '2025-07-20',
         'apply_date': '2025-07-22', 'approve_date': '2025-07-25', 'approve_status': 'COMPLETED', 'note': '棚卸完了'},
        {'q': 'Q2', 'suim_date': '2025-09-30', 'dist_date': '2025-10-01', 'review_end': '2025-10-20',
         'apply_date': '2025-10-22', 'approve_date': '2025-10-25', 'approve_status': 'COMPLETED', 'note': '棚卸完了'},
        {'q': 'Q3', 'suim_date': '2025-12-31', 'dist_date': '2026-01-05', 'review_end': '2026-01-22',
         'apply_date': '2026-01-26', 'approve_date': None, 'approve_status': 'PENDING',
         'note': 'SUIM抽出条件のエビデンス再取得依頼中(REQ-2026-002)により情シス部長承認保留'},
        {'q': 'Q4', 'suim_date': '2026-03-31', 'dist_date': '2026-04-01', 'review_end': '2026-04-22',
         'apply_date': '2026-04-26', 'approve_date': None, 'approve_status': 'PENDING',
         'note': 'SUIM抽出条件のエビデンス再取得依頼中(REQ-2026-002)により情シス部長承認保留'},
    ]

    # ============================
    # 1. Workflow log CSV
    # ============================
    wf_lines = [
        "# Workflow System (S04) - Quarterly Access Review Approval History",
        "# Export:   2026-04-20 10:00:00 JST",
        "# Export by: IT003 加藤 洋子 (E0053)",
        "",
        "タイムスタンプ,ワークフロー番号,棚卸サイクル,アクター(氏名SAP_UID),アクター(社員番号),アクション,備考",
    ]
    wf_idx = 1
    for c in cycles:
        wfno = f"WF-REV-2025-{wf_idx:03d}"
        wf_idx += 1
        # Step 1: SUIM出力 (IT003)
        wf_lines.append(f"{c['suim_date']} 14:15:00,{wfno},{c['q']},加藤 洋子 (IT003),E0053,SUIM出力_添付,抽出条件: Status=Active / Client=100")
        # Step 2: 配布 (IT003)
        wf_lines.append(f"{c['dist_date']} 09:00:00,{wfno},{c['q']},加藤 洋子 (IT003),E0053,部門長配布,7部門長宛に配布")
        # Step 3: 各部門長レビュー
        for i, (dept, head, emp) in enumerate(dept_heads):
            rev_date = datetime.strptime(c['dist_date'], '%Y-%m-%d') + timedelta(days=5 + i)
            wf_lines.append(f"{rev_date.strftime('%Y-%m-%d')} {(10+i):02d}:30:00,{wfno},{c['q']},{head},{emp},部門長レビュー完了,必要性確認・不要権限リスト提出")
        # Step 4: 削除申請 (IT003受領)
        wf_lines.append(f"{c['apply_date']} 14:00:00,{wfno},{c['q']},加藤 洋子 (IT003),E0053,削除申請受領,情シス部へ削除実行依頼")
        # Step 5: 削除実行 (IT004)
        apply_dt = datetime.strptime(c['apply_date'], '%Y-%m-%d')
        wf_lines.append(f"{(apply_dt + timedelta(days=1)).strftime('%Y-%m-%d')} 10:00:00,{wfno},{c['q']},西田 徹 (IT004),E0054,権限削除実行,対象ユーザ権限をSU01で削除")
        # Step 6: 情シス部長承認
        if c['approve_date']:
            wf_lines.append(f"{c['approve_date']} 16:30:00,{wfno},{c['q']},岡田 宏 (IT001),E0051,棚卸完了承認,{c['note']}")
        else:
            wf_lines.append(f"-,{wfno},{c['q']},岡田 宏 (IT001),E0051,承認待ち,{c['note']}")

    wf_lines.append("")
    wf_lines.append("# Records: above cycles × 12 steps each (一部情シス部長承認待ち)")

    with open(ROOT / "Workflow_QuarterlyAccessReview_ApprovalHistory_FY2025.csv", 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(wf_lines))

    print("[Created] Workflow_QuarterlyAccessReview_ApprovalHistory_FY2025.csv")

    # ============================
    # 2. 部門長レビュー結果 xlsx (for each quarter)
    # ============================
    def create_review_xlsx(cycle, dept_reviews):
        """dept_reviews: list of (dept, head, active_users, users_to_remove, reason)"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"棚卸結果_{cycle['q']}"

        ws.cell(1, 1, f"アクセス権棚卸 部門長レビュー結果 / {cycle['q']}")
        ws.cell(1, 1).font = Font(bold=True, size=14)
        ws.cell(2, 1, f"SUIM出力日: {cycle['suim_date']} / 配布日: {cycle['dist_date']} / レビュー期限: {cycle['review_end']}")

        # Headers
        headers = ['部門', '部門長(氏名SAP_UID)', '社員番号', 'レビュー実施日', 'アクティブユーザ数', '削除対象', '削除理由', 'レビューコメント']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(4, c, h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='305496')

        for i, rec in enumerate(dept_reviews, 5):
            for c, v in enumerate(rec, 1):
                ws.cell(i, c, v)

        widths = [16, 22, 12, 14, 18, 14, 30, 30]
        for c, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64+c)].width = w

        return wb

    rng = random.Random(65432)
    for c in cycles:
        dept_reviews = []
        dist_dt = datetime.strptime(c['dist_date'], '%Y-%m-%d')
        for i, (dept, head, emp) in enumerate(dept_heads):
            rev_date = (dist_dt + timedelta(days=5 + i)).strftime('%Y-%m-%d')
            active_n = rng.randint(5, 12)
            remove_n = rng.randint(0, 2) if c['q'] in ['Q1', 'Q2'] else rng.randint(0, 1)
            if remove_n == 0:
                reason = '不要権限なし'
                comment = '全ユーザ必要性確認済'
            else:
                reasons = ['退職に伴う権限削除', '異動による業務変更', '職務変更']
                reason = rng.choice(reasons)
                comment = f'{remove_n}名の権限削除申請'
            dept_reviews.append([dept, head, emp, rev_date, active_n, remove_n, reason, comment])
        wb = create_review_xlsx(c, dept_reviews)
        wb.save(ROOT / f"部門長レビュー結果_AccessReview_{c['q']}_FY2025.xlsx")
        print(f"[Created] 部門長レビュー結果_AccessReview_{c['q']}_FY2025.xlsx")

    # ============================
    # 3. 権限削除申請・実施記録 CSV
    # ============================
    removal_lines = [
        "# SAP SU01 - 権限削除申請・実施記録 FY2025",
        "# Export: 2026-04-20 10:30:00 JST",
        "",
        "棚卸サイクル,申請日,対象ユーザ,対象ロール,削除理由,削除実施日,実施者(SAP),実施者(社員番号),ステータス",
    ]
    # Generate some removal records
    rng2 = random.Random(98765)
    sample_users = [f'U-2024-{n:04d}' for n in range(100, 250, 7)]
    all_roles = ['SD_USER', 'MM_USER', 'FI_USER', 'PP_SUP', 'HR_USER', 'BASIS']
    for c in cycles:
        if c['q'] in ['Q1', 'Q2']:
            n_remove = rng2.randint(4, 7)
        else:
            n_remove = rng2.randint(2, 4)
        apl_dt = datetime.strptime(c['apply_date'], '%Y-%m-%d')
        exec_dt = apl_dt + timedelta(days=1)
        for _ in range(n_remove):
            user = rng2.choice(sample_users)
            role = rng2.choice(all_roles)
            reason = rng2.choice(['退職', '異動', '業務変更', '職務変更'])
            status = '完了' if c['q'] in ['Q1', 'Q2'] else '完了'
            removal_lines.append(f"{c['q']},{c['apply_date']},{user},{role},{reason},{exec_dt.strftime('%Y-%m-%d')},IT004 西田 徹,E0054,{status}")

    with open(ROOT / "SAP_SU01_AccessRemoval_FY2025.csv", 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(removal_lines))

    print("[Created] SAP_SU01_AccessRemoval_FY2025.csv")

    # ============================
    # 4. 情シス部長承認ログ CSV
    # ============================
    approval_lines = [
        "# Workflow System (S04) - 情シス部長 棚卸完了承認ログ",
        "# Target: FY2025 四半期アクセス権棚卸",
        "# Export: 2026-04-20 10:45:00 JST",
        "",
        "承認日時,ワークフロー番号,棚卸サイクル,承認者(氏名SAP_UID),承認者(社員番号),判定,備考",
    ]
    wf_idx = 1
    for c in cycles:
        wfno = f"WF-REV-2025-{wf_idx:03d}"
        wf_idx += 1
        if c['approve_date']:
            approval_lines.append(f"{c['approve_date']} 16:30:00,{wfno},{c['q']},岡田 宏 (IT001),E0051,承認,{c['note']}")
        else:
            approval_lines.append(f"-,{wfno},{c['q']},岡田 宏 (IT001),E0051,承認待ち,{c['note']}")

    with open(ROOT / "Workflow_ITDirector_ReviewCompletion_Approval_FY2025.csv", 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(approval_lines))

    print("[Created] Workflow_ITDirector_ReviewCompletion_Approval_FY2025.csv")


# ==============================================================
# AC-002 Fix: Remove Q3/Q4 SUIM xlsx ヘッダ 出力日時/抽出条件 (HOLD-2026-002維持)
# ==============================================================
def degrade_q3q4_suim():
    """Q3/Q4 SUIM xlsxから 出力日時 と 抽出条件 を削除 (HOLD-2026-002判断保留理由維持)"""
    from openpyxl import load_workbook

    for q in ['Q3', 'Q4']:
        path = ROOT / f"SAP_SUIM_ActiveUserList_2025{q}.xlsx"
        wb = load_workbook(path)
        ws = wb.active
        # 出力日時 row 2 - clear value
        ws.cell(2, 3, '(判別不能 - 再出力依頼中)')
        # 抽出条件 row 5 - clear value
        ws.cell(5, 3, '(条件指定の証跡なし - 再出力依頼中)')
        # Add note at top
        ws.cell(1, 1, f'SAP SUIM / Active User List Export - 2025{q} [※ヘッダ情報一部欠落]')
        wb.save(path)

    print("[Fixed] Q3/Q4 SUIM xlsx: removed 出力日時/抽出条件 (HOLD-2026-002 rationale)")


# ==============================================================
# Update evidence mapping
# ==============================================================
def update_mapping():
    """Add new evidence files to Evidence_Mapping_ITGC.csv"""
    path = Path(r"C:\Users\nyham\work\demo_data\2.RCM\Evidence_Mapping_ITGC.csv")

    # Existing mapping
    with open(path, encoding='utf-8') as f:
        content = f.read()

    # Add new entries for AC-001 PDFs (already in directory, need mapping)
    new_entries = []
    for sno in range(6, 26):
        pdf_name = f"ユーザ登録申請書_USER-REG-2025-{sno:04d}.pdf"
        if pdf_name not in content:
            new_entries.append(f'"ITGC-AC-001","{pdf_name}"')

    # Add new entries for AC-002
    ac002_files = [
        "Workflow_QuarterlyAccessReview_ApprovalHistory_FY2025.csv",
        "部門長レビュー結果_AccessReview_Q1_FY2025.xlsx",
        "部門長レビュー結果_AccessReview_Q2_FY2025.xlsx",
        "部門長レビュー結果_AccessReview_Q3_FY2025.xlsx",
        "部門長レビュー結果_AccessReview_Q4_FY2025.xlsx",
        "SAP_SU01_AccessRemoval_FY2025.csv",
        "Workflow_ITDirector_ReviewCompletion_Approval_FY2025.csv",
    ]
    for fn in ac002_files:
        if fn not in content:
            new_entries.append(f'"ITGC-AC-002","{fn}"')

    # Insert new entries before ITGC-AC-003
    lines = content.split('\n')
    out_lines = []
    inserted_ac1 = False
    inserted_ac2 = False
    for i, line in enumerate(lines):
        out_lines.append(line)
        # Insert AC-001 PDFs after last AC-001 entry, before AC-002
        if not inserted_ac1 and '"ITGC-AC-002"' in line:
            # Re-check: insert AC-001 new entries just before this line
            out_lines.pop()
            for ent in [e for e in new_entries if '"ITGC-AC-001"' in e]:
                out_lines.append(ent)
            out_lines.append(line)
            inserted_ac1 = True
        # Insert AC-002 new entries after last AC-002 line
        if not inserted_ac2 and '"ITGC-AC-003"' in line:
            out_lines.pop()
            for ent in [e for e in new_entries if '"ITGC-AC-002"' in e]:
                out_lines.append(ent)
            out_lines.append(line)
            inserted_ac2 = True

    with open(path, 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(out_lines))

    print(f"[Fixed] Evidence_Mapping_ITGC.csv: added {len(new_entries)} new entries")


if __name__ == '__main__':
    fix_wf_add_registration()
    generate_user_reg_pdfs()
    create_ac002_evidence()
    degrade_q3q4_suim()
    update_mapping()

    print("\n=== ITGC-AC-001/002 ADDITIONAL FIXES COMPLETED ===")
