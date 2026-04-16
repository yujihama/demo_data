"""
ITAC, ELC, FCRP エビデンス生成
【真の不備】【判断保留】ケースを含む
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
from pathlib import Path
import random
import sys

sys.path.insert(0, str(Path(__file__).parent))
from pdf_util import JPPDF
from image_util import sap_screenshot

BASE_ITAC = Path(r"C:\Users\nyham\work\demo_data\4.evidence\ITAC")
BASE_ELC = Path(r"C:\Users\nyham\work\demo_data\4.evidence\ELC")
BASE_FCRP = Path(r"C:\Users\nyham\work\demo_data\4.evidence\FCRP")
for p in [BASE_ITAC, BASE_ELC, BASE_FCRP]:
    p.mkdir(parents=True, exist_ok=True)

HF = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
BFONT = Font(name="Yu Gothic", size=10)
BBOLD = Font(name="Yu Gothic", size=10, bold=True)
C_ = Alignment(horizontal="center", vertical="center", wrap_text=True)
L_ = Alignment(horizontal="left", vertical="center", wrap_text=True)
R_ = Alignment(horizontal="right", vertical="center")
T_ = Side("thin", color="888888")
BRD = Border(left=T_, right=T_, top=T_, bottom=T_)
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")
FILL_NG = PatternFill("solid", fgColor="FCE4D6")
FILL_OK = PatternFill("solid", fgColor="E2EFDA")
FILL_HOLD = PatternFill("solid", fgColor="DEEBF7")


# ============================================================
# ITAC エビデンス
# ============================================================
def gen_itac_screenshots():
    sap_screenshot(
        "信用管理マスタ / 自動チェック設定",
        "OVAK",
        [
            ("与信チェック区分", "B: 基本チェック + 動的チェック"),
            ("更新グループ", "000012"),
            ("反応", "A: エラーで拒否 / B: 警告のみ / C: チェックなし"),
            ("現在設定", "A: エラーで拒否（受注保留）"),
            ("承認ワークフロー連携", "有効（WF_CREDIT_APPROVAL）"),
            ("最終変更日", "2024/4/1"),
            ("変更番号", "REL-2024-008"),
            ("変更者", "IT003 加藤 洋子"),
        ],
        status_bar="設定は保存されています。変更はITGC-CM経由のみ。",
        output_path=str(BASE_ITAC / "ITAC-001_SAP与信限度自動チェック設定画面_OVAK.png"),
    )

    sap_screenshot(
        "請求書計上 / 3-wayマッチング設定",
        "OMRK",
        [
            ("マッチング区分", "3-way (PO/GR/IR)"),
            ("金額公差", "±¥10,000"),
            ("数量公差", "±5%"),
            ("公差超過時の動作", "計上保留（ブロック）"),
            ("ブロック解除権限", "経理部課長以上"),
            ("最終変更日", "2024/10/15"),
            ("変更番号", "REL-2024-023"),
        ],
        status_bar="3-way マッチング設定は有効です。",
        output_path=str(BASE_ITAC / "ITAC-002_SAP3wayマッチング設定画面_OMRK.png"),
    )

    sap_screenshot(
        "固定資産 / 減価償却実行",
        "AFAB",
        [
            ("会社コード", "1000"),
            ("減価償却領域", "01 帳簿"),
            ("会計期間", "2025/11"),
            ("実行タイプ", "計画実行（本番）"),
            ("対象資産数", "2,847件"),
            ("減価償却費合計", "38,285,420 JPY"),
            ("ステータス", "正常終了"),
            ("実行者", "SAP_BATCH (夜間自動)"),
            ("実行日時", "2025/11/30 23:15:08"),
        ],
        status_bar="AFAB が正常終了しました。仕訳は自動計上されます。",
        output_path=str(BASE_ITAC / "ITAC-003_SAP減価償却実行画面_AFAB.png"),
    )

    print("Created: ITAC screenshots")


def gen_itac_test_excel():
    wb = openpyxl.Workbook()

    # ITAC-001
    ws = wb.active
    ws.title = "ITAC-001_与信自動チェック"
    ws.cell(row=1, column=1, value="【ITAC-001】与信限度自動チェック 動作検証テスト結果")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    ws.cell(row=2, column=1, value="実施日: 2026/2/10 / 実施者: 加藤 洋子 (IT003) + 中村 真理 (ACC004)")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    headers = ["ケースNo", "テストデータ", "期待動作", "実際の動作", "判定", "備考"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    cases = [
        (1, "テスト顧客CT-001（与信¥10M）に¥5M受注",
         "自動承認", "自動承認、ワークフロー起票なし", "OK", ""),
        (2, "テスト顧客CT-001（与信¥10M）に¥12M受注",
         "エラー・保留、WF起票", "エラー表示、保留、WF-TEST-001起票", "OK", ""),
        (3, "WF-TEST-001に本部長承認付与",
         "受注進行可能", "受注ステータス「承認済」、出荷指示可能", "OK", ""),
        (4, "与信限度ちょうど（±0）", "自動承認", "自動承認", "OK", "境界値"),
        (5, "既存売掛金＋受注が与信超過（¥8.5M + ¥2M = ¥10.5M）",
         "エラー・保留", "エラー、保留", "OK", "動的チェック"),
    ]
    r = 5
    for row in cases:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 5):
                cell.alignment = C_
            else:
                cell.alignment = L_
        ws.cell(row=r, column=5).fill = FILL_OK
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="結論: 全5ケース合格。ITAC-001は設計通り機能している。").font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)

    widths = [10, 35, 25, 35, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE_ITAC / "ITAC-001_与信限度自動チェック_動作検証.xlsx")
    print("Created: ITAC-001_与信限度自動チェック_動作検証.xlsx")


def gen_itac_003_calc_verify():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "減価償却再計算"

    ws.cell(row=1, column=1, value="【ITAC-003】減価償却自動計算 再実施テスト結果")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=2, column=1, value="実施日: 2026/2/15 / 実施者: 伊藤 健太 (ACC003)")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headers = ["資産番号", "資産名", "取得原価", "耐用年数",
               "償却方法", "SAP計算(11月)", "手計算", "差異"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    samples = [
        ("AS-1000-0123", "CNC切削加工機 M-3000", 48_000_000, 10, "定額法", 400_000, 400_000, 0),
        ("AS-1000-0089", "プレス加工機 P-5000", 72_000_000, 10, "定額法", 600_000, 600_000, 0),
        ("AS-2000-0015", "ソフトウェア SAP",
         125_000_000, 5, "定額法", 2_083_333, 2_083_333, 0),
    ]
    r = 5
    for s in samples:
        for c_i, v in enumerate(s, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 4, 5):
                cell.alignment = C_
            elif c_i in (3, 6, 7, 8):
                cell.alignment = R_
                cell.number_format = "#,##0;[Red]-#,##0"
            else:
                cell.alignment = L_
        ws.cell(row=r, column=8).fill = FILL_OK
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="結論: 3件すべて手計算とSAP計算が一致。ITAC-003は有効。").font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    widths = [14, 25, 14, 10, 10, 14, 14, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE_ITAC / "ITAC-003_減価償却手計算検証.xlsx")
    print("Created: ITAC-003_減価償却手計算検証.xlsx")


def gen_itac_002_log():
    path = BASE_ITAC / "ITAC-002_3wayマッチング結果ログ_202511.csv"
    random.seed(5050)
    lines = [
        "# SAP MIRO - 3-wayマッチング実行ログ",
        "# 対象期間: 2025/11/1 - 2025/11/30",
        "# 出力日時: 2025/12/2 09:15 / 出力者: ACC006 石井 健",
        "",
        "実行時刻,請求書番号,PO番号,GR番号,PO金額,GR金額,IR金額,公差内判定,アクション",
    ]
    for i in range(20):
        ts = datetime(2025, 11, random.randint(1, 30), random.randint(9, 17), random.randint(0, 59))
        amt = random.randint(500_000, 10_000_000)
        po = amt
        gr = amt
        ir = amt
        action = "自動計上"
        judge = "OK"
        if i == 12:
            ir = amt + 150_000
            action = "保留"
            judge = "NG (公差超過)"
        elif i == 8:
            ir = amt + 3_000
            action = "自動計上"
            judge = "OK (公差内)"
        lines.append(f"{ts.strftime('%Y-%m-%d %H:%M:%S')},INV-V-{i:04d},"
                     f"PO-2025-{i * 7:04d},GR-{i * 3:04d},{po},{gr},{ir},{judge},{action}")

    lines.append("")
    lines.append("# 件数: 20件 / 自動計上: 19件 / 保留: 1件")
    path.write_text("\n".join(lines), encoding="utf-8-sig")
    print(f"Created: {path.name}")


# ============================================================
# ELC エビデンス
# ============================================================
def gen_elc_board_minutes():
    pdf = JPPDF()
    pdf.add_page()

    pdf.h1("取締役会議事録 (抜粋)")
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, "第245回定時取締役会 / 2025年9月25日(木) 14:00～17:30",
             align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.h2("1. 出席者")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
                   "代表取締役社長 山本 健一\n"
                   "取締役CFO 渡辺 正博\n"
                   "取締役COO 小林 剛\n"
                   "社外取締役 A氏（監査等委員）\n"
                   "社外取締役 B氏（監査等委員）\n"
                   "社外取締役 C氏（監査等委員）\n"
                   "監査等委員 D氏\n"
                   "（以下、陪席）内部監査室長 長谷川 剛、総務部長 前田 美香")
    pdf.ln(3)

    pdf.h2("2. 議案・決議事項")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
                   "第1号議案: 2025年度上期決算について\n"
                   "  → 全会一致で承認\n\n"
                   "第2号議案: 2025年度下期業績見通しについて\n"
                   "  → 全会一致で承認\n\n"
                   "第3号議案: リスク管理規程の改訂について（R05）\n"
                   "  → 全会一致で承認 / 適用日 2025/10/1\n\n"
                   "第4号議案: 内部監査計画の進捗報告（内部監査室）\n"
                   "  → 報告を受領")
    pdf.ln(5)

    pdf.h2("3. 主要な議論点")
    pdf.body("・半導体装置向け売上の成長と、それに伴う与信リスクの管理強化について\n"
             "・製造現場の人員確保と自動化投資のバランス\n"
             "・グローバル展開（タイ子会社）の業績状況")
    pdf.ln(5)

    pdf.h2("4. 監査等委員からの意見")
    pdf.body("社外取締役A氏より、『内部監査室の独立性は十分確保されているが、"
             "テーマ別監査の頻度を上げることを提案』との意見あり。"
             "内部監査室長より『次年度計画で反映する』と回答。")
    pdf.ln(10)

    pdf.set_font("YuGoth", "", 9)
    pdf.cell(0, 5, "議事録作成: 総務部 前田 美香 [印] 2025/9/26", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "議長承認: 山本 健一 [印] 2025/9/26", new_x="LMARGIN", new_y="NEXT")

    pdf.output(str(BASE_ELC / "ELC-001_取締役会議事録_第245回_2025年9月.pdf"))
    print("Created: ELC-001_取締役会議事録_第245回_2025年9月.pdf")


def gen_elc_ethics_ack():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "倫理綱領受領確認"

    ws.cell(row=1, column=1, value="【ELC-002 統制実施記録】 2025年度 倫理綱領受領確認書 提出状況")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

    ws.cell(row=2, column=1, value="管理者: 総務部 前田 美香 / 集計日: 2025/6/30")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    headers = ["部門", "対象者数", "提出済", "提出率", "未提出者対応"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    data = [
        ("取締役・役員", 7, 7, "100%", ""),
        ("経理部", 20, 20, "100%", ""),
        ("営業本部", 45, 45, "100%", ""),
        ("製造本部", 280, 280, "100%", ""),
        ("技術本部", 60, 60, "100%", ""),
        ("管理本部（経理除く）", 60, 60, "100%", ""),
        ("品質保証部", 25, 25, "100%", ""),
        ("情報システム部", 15, 15, "100%", ""),
        ("経営企画部", 10, 10, "100%", ""),
        ("合計", 522, 522, "100%", "全員提出済"),
    ]
    r = 5
    for row in data:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT if row[0] != "合計" else BBOLD
            cell.border = BRD
            if c_i in (1, 2, 3, 4):
                cell.alignment = C_
            else:
                cell.alignment = L_
        if row[0] == "合計":
            for c_i in range(1, 6):
                ws.cell(row=r, column=c_i).fill = PatternFill("solid", fgColor="D9E1F2")
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="管理者承認: 前田 美香（総務部長）[印] 2025/6/30")

    widths = [22, 12, 12, 10, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE_ELC / "ELC-002_倫理綱領受領確認書提出状況_2025年度.xlsx")
    print("Created: ELC-002_倫理綱領受領確認書提出状況_2025年度.xlsx")


def gen_elc_risk_assessment():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "全社リスク評価"

    ws.cell(row=1, column=1, value="【ELC-004 統制実施記録】 2025年度 全社リスクアセスメント結果")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    ws.cell(row=2, column=1, value="実施日: 2025/6/15 / 実施者: 経営企画部 / 取締役会報告: 2025/6/30")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    headers = ["リスク№", "カテゴリ", "リスク記述", "発生可能性",
               "影響度", "優先度", "対応策"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    risks = [
        ("R-001", "事業環境", "半導体市況の急落による販売減少", "中", "大", "高", "多角化戦略の継続、自動車分野強化"),
        ("R-002", "サプライチェーン", "特殊合金の調達難・価格高騰", "中", "大", "高", "複数調達先の確保、長期契約"),
        ("R-003", "為替", "円安による輸入原材料コスト増", "高", "中", "中", "為替ヘッジの検討"),
        ("R-004", "人材", "熟練工の高齢化・後継者不足", "高", "中", "中", "教育訓練プログラム強化、自動化推進"),
        ("R-005", "IT/セキュリティ", "サイバー攻撃による業務停止", "中", "大", "高", "セキュリティ投資継続、演習実施"),
        ("R-006", "コンプライアンス", "輸出規制違反（米中対立）", "低", "大", "中", "輸出管理体制の強化、専門部署化"),
        ("R-007", "財務報告", "会計上見積りの不適切な設定", "低", "大", "中", "3段階レビュー、外部監査連携"),
        ("R-008", "環境", "カーボンニュートラル規制対応遅延", "中", "中", "中", "CO2削減計画の策定・実行"),
    ]
    r = 5
    for row in risks:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 4, 5, 6):
                cell.alignment = C_
            else:
                cell.alignment = L_
        if row[5] == "高":
            ws.cell(row=r, column=6).fill = FILL_NG
        elif row[5] == "中":
            ws.cell(row=r, column=6).fill = FILL_WARN
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="承認: 経営企画部長 [印] 2025/6/20 / 取締役会審議: 2025/6/30")

    widths = [8, 14, 40, 12, 10, 10, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ===== ELC-005 用 不正リスク評価シート =====
    # 統制記述: 決算期前に不正リスクファクター（動機・機会・正当化）を検討し、
    # 重要拠点・勘定について不正シナリオを評価する。内部監査室と経理部が合同で実施。
    ws2 = wb.create_sheet("不正リスク評価")
    ws2.cell(row=1, column=1,
             value="【ELC-005 統制実施記録】 2025年度 不正リスク評価シート")
    ws2.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws2.cell(row=2, column=1,
             value="実施日: 2025/6/18 / 実施者: 内部監査室（長谷川 剛 IA001・大塚 美穂 IA002）・"
                   "経理部（佐藤 一郎 ACC001・高橋 美咲 ACC002）合同 / 取締役会報告: 2025/6/30")
    ws2.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    ws2.cell(row=3, column=1,
             value="根拠: R03 内部監査規程 §4 / 統制記述: 決算期前に不正リスクファクター"
                   "（動機・機会・正当化）を検討し、重要拠点・勘定について不正シナリオを評価する。")
    ws2.cell(row=3, column=1).font = Font(name="Yu Gothic", size=9, italic=True, color="555555")
    ws2.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)

    # セクション1: Fraud Triangle
    ws2.cell(row=5, column=1, value="■ セクション1: 不正リスクファクター評価（Fraud Triangle）").font = BBOLD
    ws2.merge_cells(start_row=5, start_column=1, end_row=5, end_column=8)
    ft_headers = ["№", "対象プロセス/勘定", "動機（Pressure）",
                  "機会（Opportunity）", "正当化（Rationalization）",
                  "総合評価", "評価根拠", "リンク統制"]
    for i, h in enumerate(ft_headers, 1):
        c = ws2.cell(row=6, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    fraud_triangle = [
        ("FT-01", "売上計上（PLC-S）",
         "中：半期業績コミットメントのプレッシャー",
         "中：与信超過受注の個別承認が属人的",
         "低：営業現場に「期ずれ調整」の慣習は限定的",
         "中",
         "ITAC-001自動統制に依存する構造のため個別承認ログをモニタリングで補完",
         "PLC-S-001 / ITAC-001"),
        ("FT-02", "購買発注（PLC-P）",
         "中：仕入先からのリベート誘因リスク",
         "高：SAPロールでSoD違反ユーザが存在する可能性",
         "中：「効率化のため一時的に権限を拡張」との正当化",
         "高",
         "PUR004に関する職務分掌違反の懸念を内部監査室が認識済。ITGC-AC-004の補完統制と合わせて重点監視",
         "PLC-P-002 / ELC-007 / ITGC-AC-004"),
        ("FT-03", "棚卸資産評価（PLC-I）",
         "低：原価変動の業績影響は限定的",
         "中：倉庫課による差異分析と経理部の連携に依存",
         "中：「軽微差異は調整仕訳で十分」との慣習",
         "中",
         "倉庫差異調整の原因分析・経理報告プロセスを重点テスト対象に設定",
         "PLC-I-001 / PLC-I-002"),
        ("FT-04", "会計上見積り（FCRP）",
         "中：四半期利益目標達成プレッシャー",
         "中：貸倒・滞留評価は経理部課長の裁量余地あり",
         "低：「保守的評価の範囲内」との正当化",
         "中",
         "見積前提の外部情報・内部情報の根拠資料を重点確認",
         "FCRP-003 / PLC-I-005"),
        ("FT-05", "連結仕訳",
         "中：連結利益調整余地の存在",
         "中：非定型連結仕訳のレビューが経理部長の単独判断",
         "低：「実務慣行」として正当化される余地",
         "中",
         "連結パッケージのバリデーションで補完。非定型仕訳のレビュー記録を重点確認",
         "FCRP-004 / ITAC-005"),
        ("FT-06", "経費精算・交際費",
         "低：個人レベルの生活資金プレッシャー",
         "中：承認者の確認が形式的になるリスク",
         "低：「業界慣行」との正当化",
         "低",
         "金額閾値超過分のみ抽出レビュー",
         "PLC-P-006"),
        ("FT-07", "固定資産計上・減損",
         "低：減損回避の業績プレッシャー（現時点では低）",
         "中：見積の主観性（将来キャッシュフロー）",
         "低：「保守的評価の過大解釈」",
         "中",
         "ITAC-003自動計算に依存。減損兆候判定の前提を重点確認",
         "FCRP-003 / ITAC-003"),
        ("FT-08", "海外子会社取引（タイTPT）",
         "中：現地目標達成プレッシャー",
         "中：親会社モニタリング頻度の低さ",
         "中：「現地商慣習」としての正当化",
         "中",
         "FCRP-002連結パッケージ検証で補完。現地内部監査を年1回実施",
         "FCRP-002 / ELC-010"),
    ]
    rr = 7
    for row in fraud_triangle:
        for ci, v in enumerate(row, 1):
            cell = ws2.cell(row=rr, column=ci, value=v)
            cell.font = BFONT; cell.border = BRD
            cell.alignment = C_ if ci in (1, 6) else L_
        if row[5] == "高":
            ws2.cell(row=rr, column=6).fill = FILL_NG
        elif row[5] == "中":
            ws2.cell(row=rr, column=6).fill = FILL_WARN
        rr += 1

    rr += 2
    ws2.cell(row=rr, column=1,
             value="■ セクション2: 重要拠点・勘定別 不正シナリオ評価").font = BBOLD
    ws2.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
    rr += 1

    sc_headers = ["シナリオ№", "重要拠点", "重要勘定", "不正シナリオ",
                  "発生可能性", "影響度", "優先度", "対応策"]
    for i, h in enumerate(sc_headers, 1):
        c = ws2.cell(row=rr, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    rr += 1
    scenarios = [
        ("F-001", "本社", "売上高・売掛金",
         "期末前倒し出荷による売上架空計上（カットオフ違反）",
         "低", "大", "中",
         "PLC-S-006 期末カットオフ統制で全数検証 / 監査人サンプル実施"),
        ("F-002", "本社", "買掛金・仕入",
         "架空仕入先への発注・送金（仕入先マスタ不正登録）",
         "低", "大", "中",
         "PLC-P-005 仕入先マスタ管理 + 反社チェック強化"),
        ("F-003", "本社倉庫A/B", "棚卸資産",
         "実地棚卸帳簿操作による在庫水増し",
         "低", "大", "中",
         "PLC-I-001 経理部立会 + 抽取検査 / 監査人立会予定"),
        ("F-004", "タイ TPT", "売上高・売掛金",
         "現地取引先への値引きを用いた裏金化",
         "中", "中", "中",
         "海外子会社内部監査（年1回）+ 連結パッケージ検証"),
        ("F-005", "本社", "販売費及び一般管理費",
         "接待交際費の私的流用・架空計上",
         "中", "中", "中",
         "金額閾値超過分のサンプルレビュー / 内部通報窓口周知"),
        ("F-006", "本社", "連結利益",
         "非定型連結仕訳による利益操作",
         "低", "大", "中",
         "FCRP-004 連結仕訳承認で2段階レビュー徹底"),
        ("F-007", "本社", "貸倒引当金",
         "回収不能債権の引当過少設定による利益操作",
         "低", "中", "中",
         "FCRP-003 + 監査法人連携で前提確認"),
        ("F-008", "東北子会社", "製造原価",
         "仕掛品評価の恣意的操作",
         "低", "中", "中",
         "PLC-I-007 月次締め + 連結パッケージ検証"),
    ]
    for row in scenarios:
        for ci, v in enumerate(row, 1):
            cell = ws2.cell(row=rr, column=ci, value=v)
            cell.font = BFONT; cell.border = BRD
            cell.alignment = C_ if ci in (1, 5, 6, 7) else L_
        if row[6] == "高":
            ws2.cell(row=rr, column=7).fill = FILL_NG
        elif row[6] == "中":
            ws2.cell(row=rr, column=7).fill = FILL_WARN
        rr += 1

    rr += 2
    ws2.cell(row=rr, column=1,
             value="■ セクション3: 総合評価と対応方針").font = BBOLD
    ws2.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
    rr += 1
    for line in [
        "1. 最重点領域: 購買発注プロセス（FT-02）。SoD違反（PUR004）の是正状況を"
        "ITGC-AC-004・PLC-P-002と連携して継続監視する。",
        "2. 重点領域: 売上カットオフ（F-001）、棚卸在庫（FT-03/F-003）、会計上見積り"
        "（FT-04/F-007）、連結仕訳（FT-05/F-006）、海外子会社取引（FT-08/F-004）。",
        "3. 共通: 内部通報窓口の周知を四半期に1回実施し、正当化を抑制する。",
        "4. 監査計画への反映: 本評価結果を内部監査年次計画（ELC-010）および"
        "外部監査人との協議に反映済み（2025/6/25協議）。",
    ]:
        c = ws2.cell(row=rr, column=1, value=line)
        c.font = BFONT; c.alignment = L_
        ws2.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)
        ws2.row_dimensions[rr].height = 30
        rr += 1

    rr += 2
    ws2.cell(row=rr, column=1,
             value="承認: 内部監査室長 [印] 長谷川 剛 2025/6/20 / "
                   "経理部長 [印] 佐藤 一郎 2025/6/20 / "
                   "取締役会審議: 2025/6/30（第242回）/ "
                   "監査等委員会報告: 2025/6/27").font = BFONT
    ws2.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=8)

    ws2_widths = [9, 18, 24, 28, 28, 9, 38, 22]
    for i, w in enumerate(ws2_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE_ELC / "ELC-004_全社リスクアセスメント結果_2025年度.xlsx")
    print("Created: ELC-004_全社リスクアセスメント結果_2025年度.xlsx (with ELC-005 sheet)")


def gen_elc_whistleblower():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "内部通報受付台帳"

    ws.cell(row=1, column=1, value="【ELC-008 統制実施記録】 FY2025 内部通報受付台帳")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    ws.cell(row=2, column=1, value="管理者: 前田 美香（総務部長 GA001）/ 守秘義務あり（関係者限定閲覧）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    headers = ["通報番号", "受付日", "受付ルート", "通報内容(要約)",
               "調査開始", "調査結果", "完了日"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    reports = [
        ("WB-2025-001", date(2025, 7, 15), "外部弁護士経由",
         "特定の取引先への過剰接待疑い（匿名）", date(2025, 7, 20),
         "事実を確認、接待規程の範囲内と判断。注意喚起にとどめる。",
         date(2025, 8, 30)),
    ]
    r = 5
    for row in reports:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 5, 7):
                cell.alignment = C_
                if c_i in (2, 5, 7):
                    cell.number_format = "yyyy/mm/dd"
            else:
                cell.alignment = L_
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="FY2025 受付件数: 1件 / 全件調査完了 / 監査等委員会報告済（2025/9/25）").font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

    widths = [14, 12, 16, 40, 12, 50, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE_ELC / "ELC-008_内部通報受付台帳_FY2025.xlsx")
    print("Created: ELC-008_内部通報受付台帳_FY2025.xlsx")


def gen_elc_internal_audit_plan():
    pdf = JPPDF()
    pdf.add_page()

    pdf.h1("2025年度 内部監査計画")
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, "作成日: 2025年4月10日 / 作成: 内部監査室 長谷川 剛",
             align="R", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.h2("1. 監査の目的")
    pdf.body("内部監査規程（R03）に基づき、2025年度の内部統制および業務運営の監査を計画的に実施する。")
    pdf.ln(3)

    pdf.h2("2. 監査テーマ")
    pdf.table_header(["テーマ", "対象", "実施時期"], [60, 60, 60])
    pdf.table_row(["J-SOX 整備状況評価", "全社", "6月-7月"], [60, 60, 60])
    pdf.table_row(["J-SOX 運用状況評価", "全社", "10月-12月"], [60, 60, 60], fill=True)
    pdf.table_row(["販売プロセス 重点監査", "営業本部", "8月"], [60, 60, 60])
    pdf.table_row(["購買プロセス 重点監査", "購買部", "9月"], [60, 60, 60], fill=True)
    pdf.table_row(["情シス部監査（ITGC深掘）", "情シス部", "11月"], [60, 60, 60])
    pdf.table_row(["子会社監査（東北）", "テクノプレシジョン東北", "10月"], [60, 60, 60], fill=True)
    pdf.table_row(["子会社監査（タイ）", "TechnoPrecision (Thailand)", "2月"], [60, 60, 60])
    pdf.ln(5)

    pdf.h2("3. 体制")
    pdf.body("内部監査室長: 長谷川 剛（IA001）/ 担当: 大塚 美穂（IA002）/ 必要に応じて外部専門家")
    pdf.ln(5)

    pdf.h2("4. 報告")
    pdf.body("各監査完了後、経営陣および監査等委員会へ個別報告。年度末に年次報告書を提出。")
    pdf.ln(10)

    pdf.set_font("YuGoth", "", 10)
    pdf.kv("承認", "長谷川 剛 [印] 2025/4/10")
    pdf.kv("取締役会審議", "2025/4/15")
    pdf.kv("監査等委員会審議", "2025/4/20")

    pdf.output(str(BASE_ELC / "ELC-010_2025年度内部監査計画書.pdf"))
    print("Created: ELC-010_2025年度内部監査計画書.pdf")


# ============================================================
# FCRP エビデンス
# ============================================================
def gen_fcrp_monthly_close():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "月次決算チェックリスト"

    ws.cell(row=1, column=1, value="【FCRP-001 統制実施記録】 2025年11月 月次決算チェックリスト")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    ws.cell(row=2, column=1, value="実施者: 高橋 美咲 (ACC002) / 承認: 佐藤 一郎 (ACC001)")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

    headers = ["№", "カテゴリ", "チェック項目", "実施日", "担当", "結果"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    items = [
        (1, "売上", "売上計上のカットオフ確認", "2025/12/2", "中村", "完了"),
        (2, "売上", "売掛金残高の完全性確認", "2025/12/2", "中村", "完了"),
        (3, "売上", "売上金額のサブシステム突合", "2025/12/3", "中村", "完了"),
        (4, "仕入", "買掛金残高の完全性確認", "2025/12/3", "石井", "完了"),
        (5, "仕入", "支払手形残高確認", "2025/12/3", "石井", "完了"),
        (6, "在庫", "月次原価計算の実行", "2025/12/4", "伊藤", "完了"),
        (7, "在庫", "在庫残高の突合（SAP vs WMS）", "2025/12/4", "中村", "完了"),
        (8, "固定資産", "減価償却自動計算の実行（AFAB）", "2025/12/1", "SAP自動", "完了"),
        (9, "固定資産", "固定資産取得・除却の確認", "2025/12/4", "小川", "完了"),
        (10, "人件費", "給与計算データの取込", "2025/12/2", "石井", "完了"),
        (11, "引当金", "賞与引当金の月次計上", "2025/12/3", "高橋", "完了"),
        (12, "経費", "未払費用の計上", "2025/12/4", "石井", "完了"),
        (13, "経費", "前払費用の月割按分", "2025/12/4", "小川", "完了"),
        (14, "税金", "消費税の計算・仕訳", "2025/12/5", "高橋", "完了"),
        (15, "決算", "試算表の出力・確認", "2025/12/5", "高橋", "完了"),
    ]
    r = 5
    for row in items:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 4, 5, 6):
                cell.alignment = C_
            else:
                cell.alignment = L_
        ws.cell(row=r, column=6).fill = FILL_OK
        r += 1

    ws.cell(row=r, column=1, value="... 以下30項目省略（全45項目） ...").font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 2
    ws.cell(row=r, column=1, value="全45項目完了 / 経理部長承認: 佐藤 一郎 [印] 2025/12/7").font = BBOLD

    widths = [5, 12, 35, 12, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE_FCRP / "FCRP-001_月次決算チェックリスト_202511.xlsx")
    print("Created: FCRP-001_月次決算チェックリスト_202511.xlsx")


def gen_fcrp_consolidation():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "連結パッケージ受領管理"

    ws.cell(row=1, column=1, value="【FCRP-002 統制実施記録】 FY2025 Q3 連結パッケージ受領・検証管理")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    ws.cell(row=2, column=1, value="管理者: 高橋 美咲 (ACC002) / 更新: 2026/1/20")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    headers = ["子会社", "提出期限", "提出日", "バリデーション結果",
               "往復確認の有無", "最終受領日", "ステータス"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    subs = [
        ("テクノプレシジョン東北", date(2026, 1, 10), date(2026, 1, 9),
         "エラーなし", "なし", date(2026, 1, 9), "完了"),
        ("TP物流サービス", date(2026, 1, 10), date(2026, 1, 10),
         "エラー1件（科目コード誤り）", "あり(往復1回)", date(2026, 1, 12), "完了"),
        ("TechnoPrecision Thailand", date(2026, 1, 12), date(2026, 1, 11),
         "エラーなし", "なし", date(2026, 1, 11), "完了"),
        ("TPトレーディング", date(2026, 1, 10), date(2026, 1, 10),
         "エラー2件（内部取引相殺）", "あり(往復2回)", date(2026, 1, 15), "完了"),
    ]
    r = 5
    for row in subs:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 6, 7):
                cell.alignment = C_
                if c_i in (2, 3, 6):
                    cell.number_format = "yyyy/mm/dd"
            else:
                cell.alignment = L_
        if row[6] == "完了":
            ws.cell(row=r, column=7).fill = FILL_OK
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="経理部課長承認: 高橋 美咲 [印] 2026/1/18 / 経理部長承認: 佐藤 一郎 [印] 2026/1/20").font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

    widths = [22, 12, 12, 26, 18, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE_FCRP / "FCRP-002_連結パッケージ受領管理_2025Q3.xlsx")
    print("Created: FCRP-002_連結パッケージ受領管理_2025Q3.xlsx")


def gen_fcrp_estimate():
    """FCRP-003 見積レビュー（貸倒引当金計算シート - 判断保留ケース）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "貸倒引当金計算"

    ws.cell(row=1, column=1, value="【FCRP-003 統制実施記録】 2025年12月末 貸倒引当金計算シート")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    ws.cell(row=2, column=1, value="基準日: 2025/12/31 / 作成: 高橋 美咲（経理部課長）/ レビュー: 佐藤 一郎（経理部長）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws.cell(row=3, column=1, value="⚠ 内部監査指摘: 個別評価の根拠資料（顧客信用情報）が未添付。追加エビデンス要求中。").font = Font(name="Yu Gothic", size=9, bold=True, color="C00000")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=7)

    # 一般債権
    r = 5
    ws.cell(row=r, column=1, value="■ 一般債権（実績率法）").font = BBOLD
    r += 1
    headers = ["項目", "金額(円)", "実績率", "引当額(円)", "", "", ""]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=i, value=h)
        if h:
            c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    r += 1
    general = [
        ("売掛金残高（一般）", 3_524_820_000, "0.18%", 6_344_676),
        ("受取手形残高", 1_245_600_000, "0.12%", 1_494_720),
    ]
    for row in general:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 3):
                cell.alignment = L_
            else:
                cell.alignment = R_
                cell.number_format = "#,##0"
        r += 1
    ws.cell(row=r, column=1, value="一般債権合計").font = BBOLD
    ws.cell(row=r, column=1).border = BRD
    ws.cell(row=r, column=4, value=7_839_396).font = BBOLD
    ws.cell(row=r, column=4).number_format = "#,##0"
    ws.cell(row=r, column=4).alignment = R_
    ws.cell(row=r, column=4).border = BRD
    r += 2

    # 個別評価
    ws.cell(row=r, column=1, value="■ 個別評価（破産更生債権等）").font = BBOLD
    r += 1
    headers2 = ["顧客", "債権額(円)", "担保等", "回収可能額(円)", "引当額(円)", "評価根拠", "根拠資料"]
    for i, h in enumerate(headers2, 1):
        c = ws.cell(row=r, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    r += 1

    individual = [
        ("C-10007 サンプル顧客G社", 3_550_000, "なし", 2_130_000, 1_420_000,
         "同社は直近2年赤字、信用情報Cクラス", "未添付(要追加)"),
        ("C-10017 サンプル顧客N社", 6_580_000, "なし", 3_290_000, 3_290_000,
         "支払い延滞5ヶ月以上", "未添付(要追加)"),
        ("C-10023 サンプル顧客R社", 5_520_000, "商品在庫", 4_416_000, 1_104_000,
         "相殺予定、信用情報は未取得", "未添付(要追加)"),
    ]
    for row in individual:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (3, 6, 7):
                cell.alignment = L_
            elif c_i == 1:
                cell.alignment = L_
            else:
                cell.alignment = R_
                cell.number_format = "#,##0"
        # 「未添付」を強調
        if row[6].startswith("未添付"):
            ws.cell(row=r, column=7).fill = FILL_HOLD
        r += 1

    ws.cell(row=r, column=1, value="個別評価合計").font = BBOLD
    ws.cell(row=r, column=5, value=5_814_000).font = BBOLD
    ws.cell(row=r, column=5).number_format = "#,##0"
    ws.cell(row=r, column=5).alignment = R_
    r += 2

    ws.cell(row=r, column=1, value="■ 貸倒引当金 合計").font = BBOLD
    r += 1
    ws.cell(row=r, column=1, value="一般債権 + 個別評価").font = BFONT
    ws.cell(row=r, column=4, value=13_653_396).font = BBOLD
    ws.cell(row=r, column=4).number_format = "#,##0"
    ws.cell(row=r, column=4).alignment = R_
    ws.cell(row=r, column=4).fill = FILL_WARN

    r += 3
    ws.cell(row=r, column=1, value="■ 個別評価の根拠資料について（監査対応）").font = BBOLD
    r += 1
    for line in [
        "個別評価3社について、外部信用情報機関からの調査レポートを根拠資料として添付すべきだが、",
        "Q3・Q4（2025/12末、2026/3末）分について未取得のまま決算処理に使用した。",
        "監査等委員会・内部監査室からの指摘を受け、2026年3月までに追加取得予定。",
        "入手後、再度計算見直し可能性あり。現時点は暫定値として処理。",
    ]:
        ws.cell(row=r, column=1, value=line).font = BFONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        r += 1

    widths = [28, 18, 14, 18, 18, 30, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE_FCRP / "FCRP-003_貸倒引当金計算シート_2025年12月末.xlsx")
    print("Created: FCRP-003_貸倒引当金計算シート_2025年12月末.xlsx")


def gen_fcrp_consol_je():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "連結仕訳一覧"

    ws.cell(row=1, column=1, value="【FCRP-004 統制実施記録】 FY2025 Q3 連結仕訳一覧")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    ws.cell(row=2, column=1, value="出力元: 連結決算システム(S05) / 作成: 高橋 美咲 / 承認: 佐藤 一郎")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    headers = ["仕訳№", "区分", "借方科目", "貸方科目", "金額(円)", "摘要", "承認"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    entries = [
        ("CNS-Q3-001", "投資と資本の相殺", "資本金", "関係会社株式", 300_000_000,
         "テクノプレシジョン東北の投資相殺", "自動"),
        ("CNS-Q3-002", "内部取引消去", "売上高", "売上原価", 1_820_000_000,
         "親会社→東北向け内部販売消去", "自動"),
        ("CNS-Q3-003", "内部取引消去", "売上高", "売上原価", 650_000_000,
         "親会社→タイ子会社向け内部販売消去", "自動"),
        ("CNS-Q3-004", "内部取引消去", "売掛金", "買掛金", 324_500_000,
         "内部債権債務相殺", "自動"),
        ("CNS-Q3-005", "少数株主損益", "少数株主損益", "利益剰余金", 8_520_000,
         "TPトレーディング 少数株主持分", "自動"),
        ("CNS-Q3-006", "内部利益消去(在庫)", "売上原価", "棚卸資産", 42_800_000,
         "親→東北間在庫の未実現利益消去", "手動 → 伊藤作成、佐藤承認"),
    ]
    r = 5
    for row in entries:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 7):
                cell.alignment = C_
            elif c_i == 5:
                cell.alignment = R_
                cell.number_format = "#,##0"
            else:
                cell.alignment = L_
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="連結仕訳合計: 6件 / 承認: 佐藤 一郎 [印] 2026/1/20").font = BBOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

    widths = [14, 20, 18, 18, 14, 38, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE_FCRP / "FCRP-004_連結仕訳一覧_2025Q3.xlsx")
    print("Created: FCRP-004_連結仕訳一覧_2025Q3.xlsx")


def gen_fcrp_disclosure_review():
    pdf = JPPDF()
    pdf.add_page()

    pdf.h1("開示書類レビューシート (2026年3月期 第3四半期)")
    pdf.set_font("YuGoth", "", 10)
    pdf.cell(0, 5, "対象: 四半期報告書 / 作成: 2026年2月10日", align="R",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    pdf.h2("1. 対象書類")
    pdf.set_font("YuGoth", "", 10)
    pdf.multi_cell(0, 5,
                   "・四半期報告書（XBRL含む）\n"
                   "・決算短信\n"
                   "・適時開示資料")
    pdf.ln(3)

    pdf.h2("2. レビュー体制（3段階）")
    pdf.table_header(["レビュー者", "役職", "担当範囲", "完了日"],
                     [40, 40, 60, 30])
    pdf.table_row(["高橋 美咲", "経理部課長", "数値・注記・前期比較", "2026/1/28"],
                  [40, 40, 60, 30])
    pdf.table_row(["佐藤 一郎", "経理部長", "会計方針・重要な見積り", "2026/1/30"],
                  [40, 40, 60, 30], fill=True)
    pdf.table_row(["渡辺 正博", "CFO", "全体的妥当性", "2026/2/3"],
                  [40, 40, 60, 30])
    pdf.table_row(["監査等委員会", "委員会", "最終確認", "2026/2/8"],
                  [40, 40, 60, 30], fill=True)
    pdf.ln(5)

    pdf.h2("3. レビューで指摘された主要事項")
    pdf.body("・貸倒引当金の個別評価について、根拠資料の添付が不足している旨を指摘\n"
             "  → 経理部にて追加資料を取得中（2026/3/20までに対応）\n"
             "・前期比較表示の一部誤記1件 → 修正済み\n"
             "・重要な会計方針の文言改善1件 → 修正済み")
    pdf.ln(5)

    pdf.h2("4. 最終結論")
    pdf.set_font("YuGoth", "B", 11)
    pdf.set_fill_color(230, 245, 230)
    pdf.multi_cell(0, 6, "修正対応後、開示書類は適切と判断。取締役会承認を得て提出。", fill=True)
    pdf.set_fill_color(255, 255, 255)
    pdf.ln(10)

    pdf.set_font("YuGoth", "", 10)
    pdf.kv("取締役会決議", "2026/2/10")
    pdf.kv("四半期報告書提出", "2026/2/12（EDINET）")

    pdf.output(str(BASE_FCRP / "FCRP-005_開示書類レビューシート_2026年3月期Q3.pdf"))
    print("Created: FCRP-005_開示書類レビューシート_2026年3月期Q3.pdf")


if __name__ == "__main__":
    # ITAC
    gen_itac_screenshots()
    gen_itac_test_excel()
    gen_itac_003_calc_verify()
    gen_itac_002_log()
    # ELC
    gen_elc_board_minutes()
    gen_elc_ethics_ack()
    gen_elc_risk_assessment()
    gen_elc_whistleblower()
    gen_elc_internal_audit_plan()
    # FCRP
    gen_fcrp_monthly_close()
    gen_fcrp_consolidation()
    gen_fcrp_estimate()
    gen_fcrp_consol_je()
    gen_fcrp_disclosure_review()
    print("\nAll ITAC/ELC/FCRP evidence generated.")
