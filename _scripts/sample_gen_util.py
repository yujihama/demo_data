"""
サンプル生成共通ユーティリティ
- マスタデータ (顧客/仕入先/製品/従業員)
- サンプル選定ヘルパ
- PDF生成ヘルパ (注文書/発注書/検収書/稟議書等)
- Excel生成ヘルパ (25件サンプルリスト、RAWログ)
"""
import random
from datetime import date, datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============ マスタデータ ============
CUSTOMERS = [
    ("C-10001", "サンプル顧客A社", 500_000_000, "自動車"),
    ("C-10002", "サンプル顧客B社", 300_000_000, "自動車"),
    ("C-10003", "サンプル顧客C社", 200_000_000, "自動車"),
    ("C-10004", "サンプル顧客D社", 150_000_000, "自動車"),
    ("C-10005", "サンプル顧客E社", 100_000_000, "自動車"),
    ("C-10006", "サンプル顧客F社", 80_000_000, "自動車"),
    ("C-10007", "サンプル顧客G社", 50_000_000, "自動車"),
    ("C-10011", "サンプル顧客H社", 400_000_000, "半導体装置"),
    ("C-10012", "サンプル顧客I社", 250_000_000, "半導体装置"),
    ("C-10013", "サンプル顧客J社", 180_000_000, "半導体装置"),
    ("C-10014", "サンプル顧客K社", 120_000_000, "半導体装置"),
    ("C-10015", "サンプル顧客L社", 100_000_000, "半導体装置"),
    ("C-10016", "サンプル顧客M社", 80_000_000, "半導体装置"),
    ("C-10017", "サンプル顧客N社", 60_000_000, "半導体装置"),
    ("C-10018", "サンプル顧客O社", 40_000_000, "半導体装置"),
    ("C-10021", "サンプル顧客P社", 200_000_000, "商社"),
    ("C-10022", "サンプル顧客Q社", 150_000_000, "商社"),
    ("C-10023", "サンプル顧客R社", 80_000_000, "商社"),
    ("C-10024", "サンプル顧客S社", 50_000_000, "商社"),
    ("C-10025", "サンプル顧客T社", 30_000_000, "商社"),
]

VENDORS = [
    ("V-20001", "サンプル仕入先A社", "原材料(鋼材)"),
    ("V-20002", "サンプル仕入先B社", "原材料(鋼材)"),
    ("V-20003", "サンプル仕入先C社", "原材料(鋼材)"),
    ("V-20004", "サンプル仕入先D社", "原材料(鋼材)"),
    ("V-20005", "サンプル仕入先E社", "原材料(鋼材)"),
    ("V-20006", "サンプル仕入先F社", "原材料(銅材)"),
    ("V-20007", "サンプル仕入先G社", "原材料(銅材)"),
    ("V-20008", "サンプル仕入先H社", "原材料(特殊合金)"),
    ("V-20009", "サンプル仕入先I社", "原材料(鉛材)"),
    ("V-20010", "サンプル仕入先J社", "原材料(アルミ)"),
    ("V-20011", "サンプル仕入先K社", "原材料(亜鉛)"),
    ("V-20012", "サンプル仕入先L社", "原材料(チタン)"),
    ("V-20015", "サンプル仕入先O社", "原材料(鋼材)"),
    ("V-20021", "サンプル仕入先P社", "外注加工(金型)"),
    ("V-20022", "サンプル仕入先Q社", "外注加工(切削)"),
    ("V-20023", "サンプル仕入先R社", "外注加工(プレス)"),
    ("V-20024", "サンプル仕入先S社", "外注加工(研削)"),
    ("V-20025", "サンプル仕入先T社", "外注加工(熱処理)"),
]

PRODUCTS = [
    ("P-30001", "エンジンピストンピン A型", 2850, 4200),
    ("P-30003", "バルブステム A型", 1850, 2750),
    ("P-30006", "トランスミッションシャフト", 8200, 12500),
    ("P-30008", "燃料インジェクタノズル", 5500, 8200),
    ("P-30011", "ウェハー搬送ロボット用シャフト A", 12500, 18500),
    ("P-30014", "エッチング装置チャンバ部品", 25800, 38500),
    ("P-30015", "ウェハーチャックベース", 8500, 12800),
    ("P-30020", "検査装置ステージベース", 28500, 42500),
    ("P-30022", "サスペンションアーム A型", 850, 1280),
    ("P-30027", "ロボットアーム外装パネル A", 2800, 4200),
]

RAW_MATERIALS = [
    ("RAW-001", "SUS304鋼材 φ30×3000mm", 28500),
    ("RAW-002", "SUS304鋼材 φ20×1500mm", 15800),
    ("RAW-003", "SUS316鋼材 φ25×2000mm", 22400),
    ("RAW-H01", "特殊合金材インコネル718", 385000),
    ("RAW-H02", "特殊合金材 B種", 250000),
    ("RAW-T01", "チタン合金 TC6", 182000),
    ("RAW-A01", "アルミ材 A7075", 45800),
    ("RAW-C01", "銅材 C1020", 38600),
]


# ============ スタイル ============
HF = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
BFONT = Font(name="Yu Gothic", size=10)
BBOLD = Font(name="Yu Gothic", size=10, bold=True)
C_ = Alignment(horizontal="center", vertical="center", wrap_text=True)
L_ = Alignment(horizontal="left", vertical="center", wrap_text=True)
R_ = Alignment(horizontal="right", vertical="center")
T_ = Side("thin", color="888888")
BRD = Border(left=T_, right=T_, top=T_, bottom=T_)
FILL_META = PatternFill("solid", fgColor="D9E1F2")


# ============ サンプル選定ヘルパ ============
def generate_systematic_samples(
    count,
    year_start=2025,
    month_start=4,
    year_end=2026,
    month_end=3,
    seed=None,
):
    """FY2025内に均等分散するサンプル日付リストを生成"""
    if seed is not None:
        random.seed(seed)
    # 12ヶ月に分散
    months = []
    y, m = year_start, month_start
    while (y, m) <= (year_end, month_end):
        months.append((y, m))
        m += 1
        if m > 12:
            m = 1
            y += 1
    dates = []
    per_month = count // len(months)
    extra = count % len(months)
    for i, (y, m) in enumerate(months):
        n = per_month + (1 if i < extra else 0)
        used_days = set()
        for _ in range(n):
            for _attempt in range(10):
                d = random.randint(1, 28)
                if d not in used_days:
                    used_days.add(d)
                    dates.append(date(y, m, d))
                    break
    dates.sort()
    return dates[:count]


# ============ 25件サンプルリスト Excel 生成 ============
def create_sample_list_excel(
    path,
    title,
    note,
    meta_items,  # list of (key, value) tuples
    headers,  # list of header names
    rows,  # list of row data (matching headers)
    col_widths=None,
    col_center=(0, 1, 2, 3, 6, 8),  # 0-indexed columns that should be center-aligned
    col_right=(),  # 0-indexed columns that should be right-aligned (numeric)
    col_date=(),  # 0-indexed columns with date format
):
    """純粋なサンプルリストExcelを作成（分析・判定なし）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "サンプルリスト"

    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    ws.cell(row=2, column=1, value=note)
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=10, italic=True, color="555555")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    # メタ情報
    for i, (k, v) in enumerate(meta_items):
        r = 4 + i
        ws.cell(row=r, column=1, value=k).font = BBOLD
        ws.cell(row=r, column=1).fill = FILL_META
        ws.cell(row=r, column=1).border = BRD
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.cell(row=r, column=4, value=v).font = BFONT
        ws.cell(row=r, column=4).border = BRD
        ws.cell(row=r, column=4).alignment = L_
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=len(headers))

    hr = 4 + len(meta_items) + 1
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[hr].height = 32

    for idx, row in enumerate(rows):
        r = hr + 1 + idx
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i - 1 in col_center:
                cell.alignment = C_
            elif c_i - 1 in col_right:
                cell.alignment = R_
                cell.number_format = "#,##0"
            else:
                cell.alignment = L_
            if c_i - 1 in col_date:
                cell.number_format = "yyyy/mm/dd"
                cell.alignment = C_
        ws.row_dimensions[r].height = 22

    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"A{hr + 1}"

    wb.save(path)


# ============ RAW CSV出力ヘルパ ============
def write_raw_csv(path, header_lines, column_header, rows, footer_lines=None):
    """CSV形式のRAW出力ファイル生成"""
    lines = list(header_lines)
    lines.append("")
    lines.append(column_header)
    for row in rows:
        lines.append(",".join(str(v) for v in row))
    if footer_lines:
        lines.append("")
        lines.extend(footer_lines)
    path.write_text("\n".join(lines), encoding="utf-8-sig")
