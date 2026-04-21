"""ITGC-AC-003 退職者停止日修正
- ACC099, HR099: 1-2日遅延を「当日停止」に是正
- ACC099, PUR099, HR099: 土日退職を平日(月末に近い金曜)に変更
- SLS099: 停止日が土曜→月曜に変更 (遅延11日→13日)
- DEF-2026-001の遅延日数も更新 (11日/18日 → 13日/18日)
"""
import re
import sys, io
from pathlib import Path
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

ROOT = Path(r"C:\Users\nyham\work\demo_data")

# New dates
NEW_DATES = {
    'MFG099': {'ret': '2025-06-30', 'stop': '2025-06-30', 'delay_note': '退職日当日停止'},  # unchanged
    'ACC099': {'ret': '2025-08-29', 'stop': '2025-08-29', 'delay_note': '退職日当日停止'},  # was 8/31→9/1
    'SLS099': {'ret': '2025-09-30', 'stop': '2025-10-13', 'delay_note': '停止遅延13日(DEF-2026-001)'},  # was stop 10/11 Sat → 10/13 Mon
    'PUR099': {'ret': '2025-11-14', 'stop': '2025-12-02', 'delay_note': '停止遅延18日(DEF-2026-001)'},  # was 11/15 Sat→12/3
    'HR099':  {'ret': '2026-01-30', 'stop': '2026-01-30', 'delay_note': '退職日当日停止'},  # was 1/31 Sat→2/2
}

# ==============================================================
# Fix 1: Regenerate SM20 SAL log with corrected dates
# ==============================================================
def regen_sm20():
    import random
    path = ROOT / "4.evidence" / "ITGC" / "SAP_SM20_SecurityAuditLog_RetiredUsers.csv"

    active_users = [
        ('SLS001', '田中 太郎'), ('SLS002', '斎藤 次郎'), ('SLS003', '鈴木 花子'), ('SLS004', '松本 香織'),
        ('PUR001', '木村 浩二'), ('PUR002', '小林 浩太'), ('PUR003', '清水 智明'),
        ('ACC001', '佐藤 一郎'), ('ACC002', '高橋 美咲'), ('ACC003', '渡辺 俊介'), ('ACC004', '中村 真理'), ('ACC006', '石井 健'),
        ('MFG001', '森 和雄'), ('MFG002', '池田 直樹'), ('MFG003', '山田 拓也'),
        ('HR001', '近藤 文子'), ('HR002', '岩本 涼子'),
        ('IT001', '岡田 宏'), ('IT002', '吉田 雅彦'), ('IT003', '加藤 洋子'), ('IT004', '西田 徹'),
    ]

    retired_users = [
        {'uid': 'MFG099', 'name': '山崎 龍一', 'ret_date': NEW_DATES['MFG099']['ret'],
         'stop_date': NEW_DATES['MFG099']['stop'],
         'period_start': '2025-06-16', 'period_end': NEW_DATES['MFG099']['stop'], 'delay': False},
        {'uid': 'ACC099', 'name': '北川 昭子', 'ret_date': NEW_DATES['ACC099']['ret'],
         'stop_date': NEW_DATES['ACC099']['stop'],
         'period_start': '2025-08-15', 'period_end': NEW_DATES['ACC099']['stop'], 'delay': False},
        {'uid': 'SLS099', 'name': '藤井 修', 'ret_date': NEW_DATES['SLS099']['ret'],
         'stop_date': NEW_DATES['SLS099']['stop'],
         'period_start': '2025-09-16', 'period_end': NEW_DATES['SLS099']['stop'], 'delay': True, 'delay_days': 13},
        {'uid': 'PUR099', 'name': '菅原 美奈子', 'ret_date': NEW_DATES['PUR099']['ret'],
         'stop_date': NEW_DATES['PUR099']['stop'],
         'period_start': '2025-10-31', 'period_end': NEW_DATES['PUR099']['stop'], 'delay': True, 'delay_days': 18},
        {'uid': 'HR099', 'name': '大野 健介', 'ret_date': NEW_DATES['HR099']['ret'],
         'stop_date': NEW_DATES['HR099']['stop'],
         'period_start': '2026-01-16', 'period_end': NEW_DATES['HR099']['stop'], 'delay': False},
    ]

    lines = [
        "# SAP S/4HANA - Transaction SM20",
        "# Report:   Security Audit Log (SAL) - Login Events",
        "# Client:   100 (Production)",
        "# Audit Class: AU1 (Dialog Logon), AU3 (Logoff)",
        "# Retention: 730 days (Profile parameter rsau/max_file_size confirmed)",
        "# Export:   2026-02-15 14:22:30 JST by IT002 吉田 雅彦 (E0052)",
        "# Scope:    退職者5名の離任期間±2週間の全ログオンイベント",
        "",
        "# --- Section A: 期間内全ログオンイベント (母集団抜粋) ---",
        "# Filter: ログオン成功 / 期間: 各退職者離任前後",
        "",
        "タイムスタンプ,ユーザID,イベント,メッセージ,端末,ターミナル,クライアント,結果",
    ]

    from datetime import timedelta
    rng = random.Random(51001)

    all_events = []
    for ru in retired_users:
        start = datetime.strptime(ru['period_start'], '%Y-%m-%d')
        end = datetime.strptime(ru['period_end'], '%Y-%m-%d')
        cur = start
        while cur <= end:
            if cur.weekday() < 5:
                n_events = rng.randint(25, 40)
                for _ in range(n_events):
                    uid, name = rng.choice(active_users)
                    hh = rng.randint(7, 20)
                    mm = rng.randint(0, 59)
                    ss = rng.randint(0, 59)
                    ts = cur.replace(hour=hh, minute=mm, second=ss)
                    ip = f"10.20.{rng.randint(1, 30)}.{rng.randint(10, 250)}"
                    tcode = rng.choice(['SAPLSMTR_NAVIGATION', 'SESSION_MANAGER', 'SAPMSSY0'])
                    all_events.append((ts, uid, 'AU1', 'Logon successful (type=A)', ip, tcode, '100', 'SUCCESS'))
            cur += timedelta(days=1)

    all_events.sort(key=lambda x: x[0])
    sampled = all_events[::15]

    for ev in sampled:
        ts, uid, event, msg, ip, tcode, client, result = ev
        lines.append(f"{ts.strftime('%Y-%m-%d %H:%M:%S')},{uid},{event},{msg},{ip},{tcode},{client},{result}")

    lines.append("")
    lines.append(f"# Section A Records: {len(sampled)} (期間内全イベント {len(all_events)}件から15件間隔で抽出)")
    lines.append("")
    lines.append("# --- Section B: 退職者個別フィルタ結果 ---")
    lines.append("# Query Syntax: SM20 → Dynamic Filters: USER = <retired_uid> / DATE BETWEEN <period>")
    lines.append("")

    for ru in retired_users:
        lines.append(f"# --- Retired User: {ru['uid']} ({ru['name']}) ---")
        lines.append(f"# 退職日: {ru['ret_date']} / 停止日: {ru['stop_date']}")
        lines.append(f"# フィルタ期間: {ru['period_start']} ～ {ru['period_end']}")
        lines.append(f"# Query: USER = '{ru['uid']}' AND DATE BETWEEN '{ru['period_start']}' AND '{ru['period_end']}'")

        pre_start = datetime.strptime(ru['period_start'], '%Y-%m-%d')
        ret_date = datetime.strptime(ru['ret_date'], '%Y-%m-%d')
        pre_events = []
        for _ in range(rng.randint(3, 5)):
            max_days = max(1, (ret_date - pre_start).days - 1)
            day_offset = rng.randint(0, max_days)
            day = pre_start + timedelta(days=day_offset)
            if day.weekday() < 5:
                hh = rng.randint(8, 18)
                mm = rng.randint(0, 59)
                ss = rng.randint(0, 59)
                ip = f"10.20.{rng.randint(1, 30)}.{rng.randint(10, 250)}"
                pre_events.append((day.replace(hour=hh, minute=mm, second=ss), ip))
        pre_events.sort()

        if pre_events:
            lines.append(f"# 退職日前ログオン (参考: 通常勤務実績確認):")
            for ts, ip in pre_events:
                lines.append(f"{ts.strftime('%Y-%m-%d %H:%M:%S')},{ru['uid']},AU1,Logon successful (type=A),{ip},SESSION_MANAGER,100,SUCCESS")

        lines.append(f"# 退職日以降 ({ru['ret_date']} 翌日 ～ {ru['period_end']}):")
        lines.append(f"# → Query Result: 0 records found")
        lines.append("")

    lines.append("# --- Section C: 集計 ---")
    for ru in retired_users:
        if ru['delay']:
            delay_note = f"[停止遅延{ru['delay_days']}日 DEF-2026-001]"
        else:
            delay_note = "[退職日当日停止]"
        lines.append(f"# {ru['uid']}: 退職日 {ru['ret_date']} / 停止日 {ru['stop_date']} / 退職後ログオン件数=0 {delay_note}")

    with open(path, 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))

    print(f"[Fixed] SM20 SAL log: corrected retire/stop dates to business days + 0-day stops")


# ==============================================================
# Fix 2: Update DEF-2026-001 text in 不備管理台帳
# ==============================================================
def update_def_register():
    from openpyxl import load_workbook
    path = ROOT / "5.test_results" / "不備管理台帳_FY2025.xlsx"
    wb = load_workbook(path)
    ws = wb.active

    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v == 'DEF-2026-001':
            old = ws.cell(r, 3).value
            new = '退職者5名中2名のSAPアカウント停止が規程の3営業日を大きく超過(13日/18日)'
            ws.cell(r, 3, new)
            print(f"[Fixed] 不備管理台帳 DEF-2026-001: '{old}' → '{new}'")
            break

    wb.save(path)


if __name__ == '__main__':
    regen_sm20()
    update_def_register()
    print("\n=== ITGC-AC-003 DELAY FIX COMPLETED ===")
