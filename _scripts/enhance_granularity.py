"""
サマリのみのRAWファイルを本来の詳細粒度に拡充

対象:
1. FCRP-001 SAP_PeriodClose_JobLog_FY2025.csv → モジュール別ステップログ
2. PLC-I-007 SAP_MMPV_PeriodClose_Log_FY2025.csv → MM/CO/FIモジュール別締めステップ
3. PLC-P-006 SAP_F110_AutomaticPayment_RunLog_FY2025.csv → ベンダー別支払明細
4. ITGC-OM-001 SAP_DB13_DatabaseBackup_Log_FY2025.csv → コンポーネント×フェーズ別ログ
5. ITGC-OM-002 Zabbix_IncidentDetection_Log_FY2025.csv → インシデントタイムライン
6. ITGC-AC-002 SAP_SUIM_ActiveUserList_2025Q3.xlsx → 4四半期分に拡張
"""
import random
import calendar
from datetime import date, datetime, timedelta
from pathlib import Path
import sys
sys.path.insert(0, str(Path(__file__).parent))
from sample_gen_util import VENDORS, write_raw_csv

BASE = Path(r"C:\Users\nyham\work\demo_data\4.evidence")


# ============================================================
# 1. FCRP-001 SAP期末処理ジョブログ（モジュール別）
# ============================================================
def gen_fcrp_001_detailed():
    path = BASE / "FCRP" / "SAP_PeriodClose_JobLog_FY2025.csv"
    random.seed(101)

    # SAP期末処理の主要ステップ（モジュール別）
    MODULE_STEPS = {
        "FI_GL": [
            ("T_F.15", "入出金伝票のクロージング", 180, 300),
            ("T_F.19", "有価証券残高評価", 120, 240),
            ("T_FAGLF101", "再分類・繰越処理", 300, 600),
            ("T_F.01", "試算表作成", 60, 120),
        ],
        "FI_AR": [
            ("T_FBL5N", "売掛金残高確定", 120, 180),
            ("T_F.13", "債権自動クリアリング", 180, 300),
            ("T_F104", "売掛金評価換え", 240, 360),
        ],
        "FI_AP": [
            ("T_FBL1N", "買掛金残高確定", 120, 180),
            ("T_F.13_AP", "債務自動クリアリング", 180, 300),
            ("T_F.05", "外貨評価換え", 240, 360),
        ],
        "FI_AA": [
            ("T_AFAB", "減価償却費計算", 600, 900),
            ("T_ASKB", "固定資産残高転送", 180, 300),
            ("T_AR01", "固定資産一覧表作成", 120, 240),
        ],
        "CO": [
            ("T_KKS1", "原価差異計算", 360, 540),
            ("T_CO88", "製品原価配賦", 420, 720),
            ("T_KSU5", "製造間接費配賦", 300, 480),
        ],
        "MM": [
            ("T_MMPV", "在庫管理期間クローズ", 60, 120),
            ("T_MB5B", "在庫評価換算", 240, 360),
            ("T_MB52", "在庫残高確定", 120, 180),
        ],
    }

    rows = []
    for month_offset in range(12):
        y = 2025 if month_offset < 9 else 2026
        m = 4 + month_offset if month_offset < 9 else month_offset - 8
        last_day = calendar.monthrange(y, m)[1]
        close_base = date(y, m, last_day) + timedelta(days=3)  # 締め日+3営業日

        hour_offset = 0
        for module, steps in MODULE_STEPS.items():
            for tx_code, step_name, min_dur, max_dur in steps:
                start_ts = datetime.combine(close_base, datetime.min.time()) + \
                    timedelta(hours=18 + hour_offset, minutes=random.randint(0, 30))
                dur_sec = random.randint(min_dur, max_dur)
                end_ts = start_ts + timedelta(seconds=dur_sec)
                status = "SUCCESS"
                # たまにWARNING
                if random.random() < 0.03:
                    status = "SUCCESS_WITH_WARNING"
                    warning = "少数の警告レコードあり（監査対象外）"
                else:
                    warning = ""
                rows.append([
                    f"{y}-{m:02d}", module, tx_code, step_name,
                    start_ts.strftime("%Y-%m-%d %H:%M:%S"),
                    end_ts.strftime("%Y-%m-%d %H:%M:%S"),
                    dur_sec, status, "SAP_BATCH", warning
                ])
                hour_offset += 0.5

    write_raw_csv(
        path,
        ["# SAP S/4HANA - Period Close Job Log",
         "# Report:   FI/CO/MM Period-End Close Execution Log",
         "# Period:   FY2025 (2025/4 - 2026/3)",
         "# Export:   2026-04-10 09:00:00 JST",
         "# Note:     Each row represents one transaction execution within the module close"],
        "対象月,モジュール,トランザクション,ステップ名,開始時刻,終了時刻,所要秒数,結果,実行ユーザ,警告内容",
        rows,
        footer_lines=[f"# Records: {len(rows)} (12 months x 6 modules x 3-4 steps)"]
    )
    print(f"Created: {path.name} ({len(rows)} rows)")


# ============================================================
# 2. PLC-I-007 SAP MMPV 期間クローズログ（モジュール別）
# ============================================================
def gen_plc_i_007_detailed():
    path = BASE / "PLC-I" / "SAP_MMPV_PeriodClose_Log_FY2025.csv"
    random.seed(202)

    # MM/CO期末処理の主要ステップ
    STEPS = [
        ("MM_01", "MMPV", "在庫評価期間の切替", 30, 120),
        ("MM_02", "MMRV", "在庫評価期間確認", 15, 60),
        ("MM_03", "MB5B", "期末在庫集計", 180, 360),
        ("MM_04", "MB52", "在庫残高検証", 120, 240),
        ("CO_01", "KKS1", "製品別差異計算", 360, 720),
        ("CO_02", "CO88", "原価差異配賦（単一レベル）", 300, 540),
        ("CO_03", "CON2", "連結原価評価", 180, 360),
        ("CO_04", "KSU5", "間接費配賦実行", 240, 420),
        ("FI_01", "F.01", "試算表作成", 120, 240),
        ("FI_02", "F.13", "自動クリアリング", 180, 360),
    ]

    rows = []
    for month_offset in range(12):
        y = 2025 if month_offset < 9 else 2026
        m = 4 + month_offset if month_offset < 9 else month_offset - 8
        last_day = calendar.monthrange(y, m)[1]
        exec_date = date(y, m, last_day) + timedelta(days=3)

        for step_id, tx_code, step_name, min_dur, max_dur in STEPS:
            start_ts = datetime.combine(exec_date, datetime.min.time()) + \
                timedelta(hours=19 + int(step_id.split("_")[1]) * 0.3,
                         minutes=random.randint(0, 30))
            dur_sec = random.randint(min_dur, max_dur)
            end_ts = start_ts + timedelta(seconds=dur_sec)
            rows.append([
                f"{y}-{m:02d}", step_id, tx_code, step_name,
                start_ts.strftime("%Y-%m-%d %H:%M:%S"),
                end_ts.strftime("%Y-%m-%d %H:%M:%S"),
                dur_sec, "SUCCESS", "SAP_BATCH",
                f"JOB_{y}{m:02d}_{step_id}"
            ])

    write_raw_csv(
        path,
        ["# SAP S/4HANA - MM/CO Period Close Log",
         "# Report:   Transaction MMPV + Related Period Close Steps",
         "# Period:   FY2025 (2025/4 - 2026/3)",
         "# Export:   2026-04-10 10:00:00 JST"],
        "対象月,ステップID,トランザクション,ステップ名,開始時刻,終了時刻,所要秒数,結果,実行ユーザ,ジョブ番号",
        rows,
        footer_lines=[f"# Records: {len(rows)} (12 months x 10 steps)"]
    )
    print(f"Created: {path.name} ({len(rows)} rows)")


# ============================================================
# 3. PLC-P-006 SAP F110 ベンダー別支払明細
# ============================================================
def gen_plc_p_006_detailed():
    path = BASE / "PLC-P" / "SAP_F110_AutomaticPayment_RunLog_FY2025.csv"
    random.seed(303)

    rows = []
    for month_offset in range(12):
        y = 2025 if month_offset < 9 else 2026
        m = 4 + month_offset if month_offset < 9 else month_offset - 8
        pay_date = date(y, m, calendar.monthrange(y, m)[1])
        run_id = f"RUN_{y}{m:02d}01"

        # その月に支払いが発生する仕入先 (約15-22社)
        active_vendors = random.sample(VENDORS, random.randint(12, min(18, len(VENDORS))))
        for vid, vname, vcat in active_vendors:
            invoice_count = random.randint(1, 8)
            total_amount = random.randint(500_000, 50_000_000) // 1000 * 1000
            payment_method = "BANK_TRANSFER" if random.random() > 0.05 else "BILL"
            bank_code = random.choice(["A銀行", "B銀行", "C銀行", "D銀行", "E銀行"])
            bank_branch = f"支店{random.choice(['X', 'Y', 'Z', 'V', 'W'])}"
            account_no = f"普通 XXXX{random.randint(100, 999)}"
            status = "PAID" if random.random() > 0.02 else "PENDING_EXCEPTION"
            execution_ts = datetime.combine(pay_date, datetime.min.time()) + \
                timedelta(hours=14, minutes=random.randint(0, 59))

            rows.append([
                run_id, pay_date.strftime("%Y-%m-%d"),
                execution_ts.strftime("%Y-%m-%d %H:%M:%S"),
                vid, vname, vcat, invoice_count,
                total_amount, payment_method,
                f"{bank_code} {bank_branch}", account_no,
                status, "小川 由紀 (ACC005)"
            ])

    write_raw_csv(
        path,
        ["# SAP S/4HANA - Transaction F110",
         "# Report:   Automatic Payment Program - Payment Run Detail (per vendor)",
         "# Period:   FY2025 (12 monthly runs, exhaustive)",
         "# Export:   2026-04-10 11:00:00 JST"],
        "支払ID,支払日,実行時刻,仕入先コード,仕入先名,品目分類,請求書件数,支払額合計,支払方法,振込銀行,振込口座,ステータス,実行ユーザ",
        rows,
        footer_lines=[f"# Records: {len(rows)} (12 months x ~18 vendors/month)"]
    )
    print(f"Created: {path.name} ({len(rows)} rows)")


# ============================================================
# 4. ITGC-OM-001 SAP DB13 バックアップ（コンポーネント×フェーズ）
# ============================================================
def gen_itgc_om_001_detailed():
    path = BASE / "ITGC" / "SAP_DB13_DatabaseBackup_Log_FY2025.csv"
    random.seed(404)

    COMPONENTS = [
        ("DATA", "Data Volume", 1400, 1800),  # GB範囲
        ("LOG", "Log Volume", 150, 280),
        ("CATALOG", "Backup Catalog", 1, 5),
    ]

    rows = []
    # 25日分（系統抽出）
    sample_days = []
    y, m = 2025, 4
    while len(sample_days) < 25:
        day = random.randint(1, calendar.monthrange(y, m)[1])
        sample_days.append(date(y, m, day))
        m += 1
        if m > 12:
            m = 1
            y = 2026

    backup_id_counter = 1
    for bk_date in sample_days:
        backup_id = f"BK_{bk_date.strftime('%Y%m%d')}_001"
        bk_start = datetime.combine(bk_date, datetime.min.time()) + \
            timedelta(hours=1, minutes=random.randint(0, 5))

        # バックアップ全体の開始
        rows.append([
            bk_start.strftime("%Y-%m-%d %H:%M:%S"),
            backup_id, "FULL", "SYSTEM", "SAPPROD",
            "ALL", "INITIATED", 0, 0,
            "Database backup started for all components",
            "SAP_BASIS"
        ])

        component_ts = bk_start + timedelta(minutes=2)
        for comp_code, comp_name, min_gb, max_gb in COMPONENTS:
            size_gb = round(random.uniform(min_gb, max_gb), 1)
            # バックアップ用時間 (分): データ量に比例
            dur_min = int(size_gb / 20) + random.randint(3, 8)

            # START phase
            rows.append([
                component_ts.strftime("%Y-%m-%d %H:%M:%S"),
                backup_id, "FULL", comp_code, "SAPPROD",
                comp_name, "RUNNING", 0, 0,
                f"{comp_name} backup started", "SAP_BASIS"
            ])

            # COMPLETED phase
            end_ts = component_ts + timedelta(minutes=dur_min)
            rows.append([
                end_ts.strftime("%Y-%m-%d %H:%M:%S"),
                backup_id, "FULL", comp_code, "SAPPROD",
                comp_name, "COMPLETED", size_gb, dur_min * 60,
                f"{comp_name} backup completed successfully",
                "SAP_BASIS"
            ])
            component_ts = end_ts + timedelta(minutes=1)

        # 全体完了
        rows.append([
            component_ts.strftime("%Y-%m-%d %H:%M:%S"),
            backup_id, "FULL", "SYSTEM", "SAPPROD",
            "ALL", "COMPLETED", 0, 0,
            "Database backup completed successfully",
            "SAP_BASIS"
        ])

        backup_id_counter += 1

    write_raw_csv(
        path,
        ["# SAP S/4HANA - Transaction DB13 (Database Backup)",
         "# Report:   SAP HANA Backup Detail Log (per component per phase)",
         "# Source:   HANA backint + DBACOCKPIT backup catalog",
         "# Period:   FY2025 sampled 25 days",
         "# Export:   2026-02-19 08:00:00 JST"],
        "タイムスタンプ,バックアップID,バックアップ種別,コンポーネント,システム,コンポーネント名,フェーズ,サイズGB,所要秒数,ログメッセージ,実行ユーザ",
        rows,
        footer_lines=[f"# Records: {len(rows)} (25 days x 8 events: START + 3 components x 2 phases + END)"]
    )
    print(f"Created: {path.name} ({len(rows)} rows)")


# ============================================================
# 5. ITGC-OM-002 Zabbix インシデントタイムライン
# ============================================================
def gen_itgc_om_002_detailed():
    path = BASE / "ITGC" / "Zabbix_IncidentDetection_Log_FY2025.csv"
    random.seed(505)

    INCIDENTS = [
        ("2025-04-12", "03:15", "MEDIUM", "SAP_PROD_HOST", "CPU_UTILIZATION_HIGH",
         "夜間バッチ遅延（標準原価計算）", "db-lock"),
        ("2025-05-03", "10:22", "LOW", "WEB_FRONT_01", "HTTP_RESPONSE_SLOW",
         "ポータル画面の応答遅延", "network"),
        ("2025-06-08", "14:22", "LOW", "DESKTOP_VDI", "LOGIN_ERROR",
         "特定ユーザでSAP GUI接続エラー", "client-cache"),
        ("2025-06-18", "08:03", "LOW", "MAIL_SERVER_01", "QUEUE_BACKLOG",
         "メール配信遅延", "queue-bloat"),
        ("2025-07-11", "19:45", "MEDIUM", "SAP_APP_02", "MEMORY_THRESHOLD",
         "アプリサーバーメモリ閾値超過", "memory-leak"),
        ("2025-07-24", "11:30", "LOW", "PRINT_SERVER", "PRINT_QUEUE_STUCK",
         "印刷キュー詰まり", "driver-issue"),
        ("2025-08-15", "10:05", "MEDIUM", "NETWORK_CORE", "LINK_FLAP",
         "WMS-SAP間連携停止", "network-hw"),
        ("2025-08-28", "16:30", "LOW", "SAP_APP_01", "LOG_THRESHOLD",
         "ログ出力閾値超過", "logging-config"),
        ("2025-09-04", "09:12", "LOW", "DESKTOP_VDI", "LOGIN_ERROR",
         "VDIログイン遅延", "ad-sync"),
        ("2025-09-29", "13:45", "LOW", "STORAGE_ARRAY", "DISK_READ_SLOW",
         "ストレージ応答遅延", "disk-controller"),
        ("2025-10-08", "07:15", "LOW", "MAIL_SERVER_02", "QUEUE_BACKLOG",
         "メール配信遅延（朝方）", "external-relay"),
        ("2025-10-22", "16:45", "LOW", "PRINT_SERVER", "PRINT_QUEUE_STUCK",
         "印刷キュー詰まり（営業部）", "driver-issue"),
        ("2025-11-07", "22:30", "HIGH", "SAP_HANA_DB", "MEMORY_ALERT",
         "HANA DBメモリ不足警告", "query-overload"),
        ("2025-11-18", "14:20", "LOW", "BACKUP_SERVER", "BACKUP_DELAY",
         "バックアップ実行遅延", "tape-drive"),
        ("2025-12-03", "11:05", "LOW", "DESKTOP_VDI", "LOGIN_ERROR",
         "VDI認証遅延", "ldap-latency"),
        ("2025-12-20", "02:30", "LOW", "SAP_APP_02", "DISK_THRESHOLD",
         "アプリサーバーディスク閾値", "log-retention"),
        ("2026-01-14", "15:50", "LOW", "NETWORK_EDGE", "LATENCY_HIGH",
         "WAN遅延", "external-isp"),
        ("2026-02-08", "09:40", "LOW", "PRINT_SERVER", "PRINT_QUEUE_STUCK",
         "印刷キュー詰まり（購買部）", "driver-issue"),
    ]

    rows = []
    for idx, (dt_str, time_str, sev, host, trigger, msg, root_cause) in enumerate(INCIDENTS, 1):
        base_ts = datetime.strptime(f"{dt_str} {time_str}", "%Y-%m-%d %H:%M")
        incident_id = f"INC-{dt_str.replace('-', '')}-{idx:03d}"

        # Timeline events
        dur_min = {"HIGH": 180, "MEDIUM": 60, "LOW": 20}[sev]

        # 1. DETECTED
        rows.append([
            base_ts.strftime("%Y-%m-%d %H:%M:%S"),
            incident_id, "DETECTED", sev, host, trigger,
            f"[TRIGGER] {msg}", "ZABBIX_AGENT"
        ])

        # 2. NOTIFIED (1-3 min later)
        ts_notify = base_ts + timedelta(minutes=random.randint(1, 3))
        rows.append([
            ts_notify.strftime("%Y-%m-%d %H:%M:%S"),
            incident_id, "NOTIFIED", sev, host, trigger,
            f"Alert sent to on-call team via email/SMS",
            "ZABBIX_SERVER"
        ])

        # 3. ACKNOWLEDGED (3-15 min)
        ts_ack = ts_notify + timedelta(minutes=random.randint(3, 15))
        rows.append([
            ts_ack.strftime("%Y-%m-%d %H:%M:%S"),
            incident_id, "ACKNOWLEDGED", sev, host, trigger,
            f"Incident acknowledged by on-call engineer",
            random.choice(["IT002 吉田 雅彦", "IT003 加藤 洋子", "IT001 岡田 宏"])
        ])

        # 4. INVESTIGATING
        ts_inv = ts_ack + timedelta(minutes=random.randint(5, 30))
        rows.append([
            ts_inv.strftime("%Y-%m-%d %H:%M:%S"),
            incident_id, "INVESTIGATING", sev, host, trigger,
            f"Root cause candidate: {root_cause}",
            "IT002 吉田 雅彦"
        ])

        # 5. RESOLVED
        ts_res = base_ts + timedelta(minutes=dur_min)
        rows.append([
            ts_res.strftime("%Y-%m-%d %H:%M:%S"),
            incident_id, "RESOLVED", sev, host, trigger,
            f"Resolution applied: {root_cause} mitigation",
            "IT002 吉田 雅彦"
        ])

        # 6. CLOSED (review completed)
        ts_close = ts_res + timedelta(hours=random.randint(1, 24))
        rows.append([
            ts_close.strftime("%Y-%m-%d %H:%M:%S"),
            incident_id, "CLOSED", sev, host, trigger,
            f"Incident closed after post-mortem review",
            "IT001 岡田 宏"
        ])

    write_raw_csv(
        path,
        ["# Zabbix Monitoring Platform - Incident Timeline Log",
         "# Source:   zabbix_server.history_log + event + action",
         "# Period:   FY2025 (exhaustive 18 incidents)",
         "# Export:   2026-02-19 09:00:00 JST",
         "# Severity: HIGH / MEDIUM / LOW"],
        "タイムスタンプ,インシデントID,イベント種別,重大度,ホスト,トリガー,メッセージ,アクター",
        rows,
        footer_lines=[f"# Records: {len(rows)} (18 incidents x 6 timeline events)"]
    )
    print(f"Created: {path.name} ({len(rows)} rows)")


# ============================================================
# 6. ITGC-AC-002 SAP SUIM 4四半期分
# ============================================================
def gen_itgc_ac_002_quarterly():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HF = PatternFill("solid", fgColor="1F4E78")
    HFONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
    BFONT = Font(name="Yu Gothic", size=10)
    C_ = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L_ = Alignment(horizontal="left", vertical="center", wrap_text=True)
    T_ = Side("thin", color="888888")
    BRD = Border(left=T_, right=T_, top=T_, bottom=T_)
    FILL_NG = PatternFill("solid", fgColor="FCE4D6")

    # 既存の Q3 ファイルを削除
    old = BASE / "ITGC" / "SAP_SUIM_ActiveUserList_2025Q3.xlsx"
    if old.exists():
        old.unlink()

    USERS_BASE = [
        ("CEO001", "山本 健一", "代表取締役", "ALL_READ", ""),
        ("CFO001", "渡辺 正博", "管理本部", "FI_MGR, CO_MGR, ALL_READ", ""),
        ("ACC001", "佐藤 一郎", "経理部", "FI_MGR, CO_MGR", ""),
        ("ACC002", "高橋 美咲", "経理部", "FI_SUP, GL_POST", ""),
        ("ACC003", "伊藤 健太", "経理部", "CO_SUP", ""),
        ("ACC004", "中村 真理", "経理部", "FI_USER, GL_POST", ""),
        ("ACC005", "小川 由紀", "経理部", "FI_USER", ""),
        ("ACC006", "石井 健", "経理部", "FI_USER, AP_USER", ""),
        ("SLS001", "田中 太郎", "営業本部", "SD_MGR", ""),
        ("SLS002", "斎藤 次郎", "営業本部", "SD_SUP", ""),
        ("SLS003", "藤田 修", "営業本部", "SD_SUP", ""),
        ("SLS004", "松本 香織", "営業本部", "SD_USER", ""),
        ("SLS005", "井上 大輔", "営業本部", "SD_USER", ""),
        ("PUR001", "木村 浩二", "購買部", "MM_MGR, PO_APPROVE", ""),
        ("PUR002", "林 真由美", "購買部", "MM_SUP, PO_APPROVE", ""),
        ("PUR003", "清水 智明", "購買部", "MM_USER, PO_CREATE", ""),
        ("PUR004", "山田 純一", "購買部", "MM_USER, PO_CREATE, PO_APPROVE",
         "※ SoD違反: PO_CREATE+PO_APPROVE (要改善)"),
        ("MFG001", "森 和雄", "製造本部", "PP_MGR", ""),
        ("MFG002", "池田 昌夫", "製造本部", "PP_SUP", ""),
        ("WHS001", "橋本 明", "製造本部", "WM_SUP", ""),
        ("IT001", "岡田 宏", "情シス部", "BASIS, ALL_READ", "特権ID"),
        ("IT002", "吉田 雅彦", "情シス部", "BASIS", "特権ID"),
        ("IT003", "加藤 洋子", "情シス部", "DEVELOPER", ""),
        ("IT004", "西田 徹", "情シス部", "HELPDESK", ""),
        ("HR001", "近藤 文子", "人事部", "HR_MGR", ""),
        ("HR002", "野村 淳", "人事部", "HR_USER", ""),
        ("GA001", "前田 美香", "総務部", "GA_MGR", ""),
        ("IA001", "長谷川 剛", "内部監査室", "ALL_READ", ""),
        ("IA002", "大塚 美穂", "内部監査室", "ALL_READ", ""),
    ]

    # 四半期別の拠出タイミング
    QUARTERS = [
        ("2025Q1", date(2025, 6, 30), "2025-06-30 14:15:00"),
        ("2025Q2", date(2025, 9, 30), "2025-09-30 14:22:12"),
        ("2025Q3", date(2025, 12, 31), "2025-12-31 14:30:08"),
        ("2025Q4", date(2026, 3, 31), "2026-03-31 14:45:45"),
    ]

    for q_label, base_date, output_ts in QUARTERS:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SUIM_ユーザ一覧"

        ws.cell(row=1, column=1, value=f"SAP SUIM / Active User List Export - {q_label}")
        ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

        meta = [
            ("出力日時", output_ts),
            ("出力者", "IT003 加藤 洋子"),
            ("Transaction", "SUIM → User → Users by Logon Date"),
            ("抽出条件", "Status = Active / Client = 100 (Production)"),
            ("Validity", f"Valid on {base_date.strftime('%Y-%m-%d')}"),
        ]
        for i, (k, v) in enumerate(meta):
            r = 2 + i
            ws.cell(row=r, column=1, value=k).font = BFONT
            ws.cell(row=r, column=1).border = BRD
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
            ws.cell(row=r, column=3, value=v).font = BFONT
            ws.cell(row=r, column=3).border = BRD
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)

        headers = ["ユーザID", "氏名", "所属部門", "付与ロール", "有効期限",
                   "最終ログイン日", "作成日", "備考"]
        hr = 2 + len(meta) + 1
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=hr, column=i, value=h)
            c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

        # ユーザ行（各四半期の時点で有効なユーザ）
        random.seed(hash(q_label))
        users_this_q = list(USERS_BASE)  # Base 29

        # Q2以降は退職者除外
        retirees = {"2025Q1": [], "2025Q2": ["E0095-田中一郎-製造部", "E0094-山口次郎-経理部"],
                    "2025Q3": ["E0099-退職者A-営業本部", "E0095", "E0094"],
                    "2025Q4": ["E0099", "E0095", "E0094", "E0098-退職者B-購買部", "E0093"]}

        for r_idx, (uid, name, dept, roles, note) in enumerate(users_this_q):
            r = hr + 1 + r_idx
            last_login = base_date - timedelta(days=random.randint(0, 5))
            create_date = date(random.randint(2010, 2023), random.randint(1, 12),
                               random.randint(1, 28))
            row_data = [uid, name, dept, roles, "無期限",
                        last_login.strftime("%Y-%m-%d"),
                        create_date.strftime("%Y-%m-%d"), note]
            for c_i, v in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c_i, value=v)
                cell.font = BFONT; cell.border = BRD
                if c_i in (1, 3, 5, 6, 7):
                    cell.alignment = C_
                else:
                    cell.alignment = L_
            if note and ("違反" in note or "特権ID" in note):
                for c_i in range(1, 9):
                    ws.cell(row=r, column=c_i).fill = FILL_NG

        # 退職者を Q3/Q4 に追加（停止遅延を示すため）
        if q_label == "2025Q3":
            # SLS099 退職者A (退職 9/30、停止 10/11 → Q3末時点でまだ残存可能性あり)
            # 実際はQ3末(12/31)時点で既に停止済みだが、Q2末(9/30)時点では存在していた
            pass
        if q_label == "2025Q2":
            # Q2末時点で SLS099 はまだアクティブ（退職前）
            # Q1/Q2時点では SLS099 は有効だったがQ2末以降退職
            r = hr + 1 + len(users_this_q)
            row_data = ["SLS099", "退職者A", "営業本部(予定退職)", "SD_USER", "無期限",
                        (base_date - timedelta(days=2)).strftime("%Y-%m-%d"),
                        "2018-04-01", "2025/9/30 退職予定"]
            for c_i, v in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c_i, value=v)
                cell.font = BFONT; cell.border = BRD
                if c_i in (1, 3, 5, 6, 7):
                    cell.alignment = C_
                else:
                    cell.alignment = L_

        widths = [10, 14, 16, 30, 12, 14, 14, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = f"A{hr + 1}"

        out = BASE / "ITGC" / f"SAP_SUIM_ActiveUserList_{q_label}.xlsx"
        wb.save(out)
        print(f"Created: {out.name}")


if __name__ == "__main__":
    gen_fcrp_001_detailed()
    gen_plc_i_007_detailed()
    gen_plc_p_006_detailed()
    gen_itgc_om_001_detailed()
    gen_itgc_om_002_detailed()
    gen_itgc_ac_002_quarterly()
    print("\nAll files enhanced to proper granularity.")
