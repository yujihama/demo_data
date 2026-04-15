"""
PLC-S（販売プロセス）のExcel形式エビデンス生成
- 25件サンプル一覧
- 出荷売上突合表
- 月次請求書発行一覧
- 入金消込リスト
- 売掛金年齢表
- 期末カットオフテスト
- 価格変更履歴レポート
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime, timedelta
from pathlib import Path
import random

random.seed(2025)

BASE = Path(r"C:\Users\nyham\work\demo_data\4.evidence\PLC-S")
BASE.mkdir(parents=True, exist_ok=True)

# 共通スタイル
HF = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
BFONT = Font(name="Yu Gothic", size=10)
BBOLD = Font(name="Yu Gothic", size=10, bold=True)
C = Alignment(horizontal="center", vertical="center", wrap_text=True)
L = Alignment(horizontal="left", vertical="center", wrap_text=True)
R = Alignment(horizontal="right", vertical="center")
T = Side("thin", color="888888")
BRD = Border(left=T, right=T, top=T, bottom=T)
FILL_NG = PatternFill("solid", fgColor="FCE4D6")  # 不合格（赤系）
FILL_HOLD = PatternFill("solid", fgColor="DEEBF7")  # 保留（青系）
FILL_OK = PatternFill("solid", fgColor="E2EFDA")  # 合格（緑系）
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")  # 注意（黄）

CUSTOMERS = {
    "C-10001": ("トヨタエンジニアリング株式会社", 500_000_000),
    "C-10002": ("本田技研部品株式会社", 300_000_000),
    "C-10003": ("日産精密パーツ株式会社", 200_000_000),
    "C-10004": ("マツダオートパーツ株式会社", 150_000_000),
    "C-10005": ("スズキ自動車部品工業", 100_000_000),
    "C-10006": ("SUBARU部品株式会社", 80_000_000),
    "C-10007": ("三菱自動車部品販売", 50_000_000),
    "C-10011": ("東京エレクトロン購買部", 400_000_000),
    "C-10012": ("SCREENセミコンダクター", 250_000_000),
    "C-10013": ("ディスコ精密部品", 180_000_000),
    "C-10014": ("アドバンテスト調達部", 120_000_000),
    "C-10015": ("株式会社ニコン精機", 100_000_000),
    "C-10016": ("キヤノンマシナリー", 80_000_000),
    "C-10017": ("日立ハイテク部品", 60_000_000),
    "C-10018": ("東京精密部品株式会社", 40_000_000),
    "C-10021": ("三菱商事精密機器", 200_000_000),
    "C-10022": ("伊藤忠テクノソリューション", 150_000_000),
    "C-10023": ("丸紅情報システムズ", 80_000_000),
    "C-10024": ("双日マシナリー", 50_000_000),
    "C-10025": ("JFE商事精密部品", 30_000_000),
}

SALES_REPS = ["斎藤 次郎", "藤田 修", "松本 香織", "井上 大輔"]

APPROVERS = {
    "営業本部長": "田中 太郎 (SLS001)",
    "営業課長(自動車)": "斎藤 次郎 (SLS002)",
    "営業課長(半導体)": "藤田 修 (SLS003)",
    "経理部長": "佐藤 一郎 (ACC001)",
    "経理部課長": "高橋 美咲 (ACC002)",
    "経理部主任": "中村 真理 (ACC004)",
    "経理部担当": "石井 健 (ACC006)",
}


# ============================================================
# 1. PLC-S-001 受注サンプル25件一覧
# ============================================================
def gen_sample_list_s001():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "受注承認サンプル25件"

    ws.cell(row=1, column=1, value="【PLC-S-001 受注・与信承認】 運用状況評価 サンプルテスト結果")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)

    meta = [
        ("評価対象期間", "FY2025 (2025/4/1 - 2026/3/31)"),
        ("母集団", "SAP VA05 受注伝票一覧 3,247件"),
        ("抽出条件", "受注日 2025/4/1～2026/3/31、ステータス「完了」「処理中」、キャンセル除く"),
        ("抽出日時", "2026/2/10 10:30 JST"),
        ("抽出方法", "系統抽出（間隔130件、開始位置57）"),
        ("サンプル数", "25件"),
        ("テスト実施者", "長谷川 剛（IA001 内部監査室）"),
        ("テスト期間", "2026/2/10 ～ 2026/2/13"),
    ]
    for i, (k, v) in enumerate(meta):
        ws.cell(row=2 + i, column=1, value=k)
        ws.cell(row=2 + i, column=1).font = BBOLD
        ws.cell(row=2 + i, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(row=2 + i, column=1).border = BRD
        ws.cell(row=2 + i, column=2, value=v)
        ws.cell(row=2 + i, column=2).font = BFONT
        ws.cell(row=2 + i, column=2).border = BRD
        ws.merge_cells(start_row=2 + i, start_column=2, end_row=2 + i, end_column=11)

    # ヘッダ
    headers = ["№", "受注番号", "受注日", "顧客コード", "顧客名", "受注金額(円)",
               "営業担当", "与信限度\nチェック", "与信超過時\n承認", "判定", "備考"]
    hr = 11
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C; c.border = BRD
    ws.row_dimensions[hr].height = 32

    # 25件のデータ生成
    random.seed(1001)
    customer_ids = list(CUSTOMERS.keys())
    samples = []
    for i in range(1, 26):
        cid = random.choice(customer_ids)
        cname, credit = CUSTOMERS[cid]
        month = random.choice([5, 6, 7, 8, 9, 10, 11, 12, 1, 2])
        year = 2026 if month <= 3 else 2025
        day = random.randint(1, 28)
        order_date = date(year, month, day)
        order_no = f"ORD-2025-{100 + i * 130:04d}"
        amount = random.choice([
            random.randint(500_000, 5_000_000),
            random.randint(5_000_000, 30_000_000),
            random.randint(30_000_000, 80_000_000),
        ])
        rep = random.choice(SALES_REPS)
        samples.append({
            "n": i, "order_no": order_no, "date": order_date,
            "cid": cid, "cname": cname, "amount": amount,
            "rep": rep, "credit": credit,
        })

    # 意図的な例外: サンプル14番を承認1日遅れに
    rows_out = []
    for s in samples:
        credit_ok = "○（自動チェック通過）"
        approval = "（不要／限度内）"
        result = "合格"
        remark = ""

        # 与信超過ケース（意図的に数件）
        if s["n"] in (3, 11, 19):
            credit_ok = "超過検知"
            approval = f"○\n営業本部長承認済\n(2025/{s['date'].month:02d}/{s['date'].day:02d} 承認)"
            remark = "与信限度超過のためワークフロー承認を確認"

        # サンプル14番を例外ケースに
        if s["n"] == 14:
            credit_ok = "超過検知"
            s["amount"] = 15_200_000  # 超過金額
            approval = f"△ 承認日が1日遅れ\n(受注日+1日)"
            result = "軽微例外\n(許容)"
            remark = "承認者出張中、翌日復帰後承認。業務影響なし。"

        rows_out.append({**s, "credit_ok": credit_ok, "approval": approval,
                         "result": result, "remark": remark})

    # 行書き込み
    for idx, s in enumerate(rows_out):
        r = hr + 1 + idx
        row_data = [
            s["n"], s["order_no"], s["date"], s["cid"], s["cname"],
            s["amount"], s["rep"], s["credit_ok"], s["approval"],
            s["result"], s["remark"]
        ]
        for c_i, v in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT
            cell.border = BRD
            if c_i in (1, 2, 3, 4, 7, 8, 9, 10):
                cell.alignment = C
            elif c_i == 6:
                cell.alignment = R
                cell.number_format = "#,##0"
            elif c_i == 3:
                cell.number_format = "yyyy/mm/dd"
            else:
                cell.alignment = L
        ws.row_dimensions[r].height = 35

        if s["result"] == "合格":
            for c_i in (10,):
                ws.cell(row=r, column=c_i).fill = FILL_OK
        elif "例外" in s["result"]:
            for c_i in (8, 9, 10):
                ws.cell(row=r, column=c_i).fill = FILL_WARN

    # 集計
    sum_row = hr + 26 + 1
    ws.cell(row=sum_row, column=1, value="集計")
    ws.cell(row=sum_row, column=1).font = BBOLD
    ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=3)
    pass_count = sum(1 for s in rows_out if s["result"] == "合格")
    exc_count = sum(1 for s in rows_out if "例外" in s["result"])
    ws.cell(row=sum_row, column=4, value=f"合格: {pass_count}件 / 不合格: 0件 / 軽微例外: {exc_count}件（許容）")
    ws.cell(row=sum_row, column=4).font = BFONT
    ws.merge_cells(start_row=sum_row, start_column=4, end_row=sum_row, end_column=11)

    # 結論
    con_row = sum_row + 2
    ws.cell(row=con_row, column=1, value="結論")
    ws.cell(row=con_row, column=1).font = BBOLD
    ws.cell(row=con_row, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
    ws.merge_cells(start_row=con_row, start_column=1, end_row=con_row, end_column=3)
    ws.cell(row=con_row, column=4, value="運用評価：有効（軽微例外1件は業務影響なしとして許容）")
    ws.cell(row=con_row, column=4).font = BBOLD
    ws.cell(row=con_row, column=4).fill = FILL_OK
    ws.merge_cells(start_row=con_row, start_column=4, end_row=con_row, end_column=11)

    # 列幅
    widths = [6, 18, 12, 12, 28, 14, 14, 18, 22, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A12"

    wb.save(BASE / "PLC-S-001_受注サンプル25件.xlsx")
    print("Created: PLC-S-001_受注サンプル25件.xlsx")
    return rows_out


# ============================================================
# 2. PLC-S-002 出荷売上突合表（2025年11月分）
# ============================================================
def gen_shipment_sales_match():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "出荷売上突合"

    ws.cell(row=1, column=1, value="【PLC-S-002】出荷実績 × 売上計上 突合表 (2025年11月)")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    ws.cell(row=2, column=1, value="作成日: 2025/12/03 / 作成者: 中村 真理（経理部主任） / 承認: 高橋 美咲（経理部課長）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    headers = ["№", "出荷番号\n(WMS)", "出荷日", "受注番号", "顧客コード", "出荷金額(円)",
               "売上計上日", "売上金額(円)", "突合結果"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C; c.border = BRD
    ws.row_dimensions[4].height = 32

    # 約30件のデータ（例外1件）
    random.seed(2002)
    cids = list(CUSTOMERS.keys())
    r = 5
    for i in range(1, 31):
        ship_date = date(2025, 11, random.randint(1, 29))
        sale_date = ship_date + timedelta(days=random.choice([0, 0, 0, 1]))  # ほぼ同日
        cid = random.choice(cids)
        amount = random.randint(1_000_000, 25_000_000)
        sale_amount = amount
        result = "一致"
        # 意図的に1件不一致→調査完了
        if i == 17:
            sale_amount = amount - 50_000
            result = "差異¥50,000\n(値引調整・是正済)"
        data = [i, f"SH-202511-{i:04d}", ship_date,
                f"ORD-2025-{1000 + i * 13}", cid, amount,
                sale_date, sale_amount, result]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 5, 7, 9):
                cell.alignment = C
                if c_i in (3, 7):
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (6, 8):
                cell.alignment = R
                cell.number_format = "#,##0"
        if result == "一致":
            ws.cell(row=r, column=9).fill = FILL_OK
        else:
            ws.cell(row=r, column=9).fill = FILL_WARN
        r += 1

    # サマリ
    r += 1
    ws.cell(row=r, column=1, value="突合結果サマリ").font = BBOLD
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(row=r, column=4, value="一致: 29件 / 差異: 1件（是正済） / 未マッチ: 0件").font = BFONT
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
    ws.cell(row=r + 1, column=1, value="結論").font = BBOLD
    ws.cell(row=r + 1, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=3)
    ws.cell(row=r + 1, column=4, value="当月の出荷と売上計上は適切に突合・是正されている").font = BBOLD
    ws.cell(row=r + 1, column=4).fill = FILL_OK
    ws.merge_cells(start_row=r + 1, start_column=4, end_row=r + 1, end_column=9)

    # 承認記録
    ws.cell(row=r + 3, column=1, value="レビュー: 高橋 美咲 [印] 2025/12/04")
    ws.cell(row=r + 3, column=1).font = BFONT
    ws.cell(row=r + 4, column=1, value="承認: 佐藤 一郎（経理部長） [印] 2025/12/05")
    ws.cell(row=r + 4, column=1).font = BFONT

    widths = [5, 15, 12, 15, 12, 14, 12, 14, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    wb.save(BASE / "PLC-S-002_出荷売上突合表_202511.xlsx")
    print("Created: PLC-S-002_出荷売上突合表_202511.xlsx")


# ============================================================
# 3. PLC-S-003 月次請求書発行一覧
# ============================================================
def gen_invoice_list():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "月次請求書発行一覧"

    ws.cell(row=1, column=1, value="【PLC-S-003】2025年11月 月次請求書発行一覧")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=2, column=1, value="発行バッチ実行日時: 2025/11/30 23:58 （SAP自動発行） / 発行件数: 152件 / 合計金額: ¥1,284,560,000")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headers = ["請求書番号", "発行日", "顧客コード", "顧客名", "請求金額(円)",
               "税込金額(円)", "支払期日", "送付方法"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C; c.border = BRD

    random.seed(3003)
    cids = list(CUSTOMERS.keys())
    r = 5
    for i in range(1, 31):  # 先頭30件のみ
        cid = random.choice(cids)
        cname = CUSTOMERS[cid][0]
        amount = random.randint(1_000_000, 35_000_000) // 1000 * 1000
        tax = int(amount * 1.1)
        due = date(2026, 1, 31 if i % 3 == 0 else 20 if i % 3 == 1 else 15)
        send = random.choice(["PDFメール送付", "PDFメール送付", "郵送（原本）"])
        data = [f"INV-202511-{i:04d}", date(2025, 11, 30), cid, cname, amount, tax, due, send]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 7, 8):
                cell.alignment = C
                if c_i in (2, 7):
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (5, 6):
                cell.alignment = R
                cell.number_format = "#,##0"
            else:
                cell.alignment = L
        r += 1

    ws.cell(row=r, column=1, value="... 以下122件省略 ...").font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    # レビュー記録
    ws.cell(row=r + 2, column=1, value="突合チェック実施: 中村 真理 [印] 2025/12/01")
    ws.cell(row=r + 2, column=1).font = BFONT
    ws.cell(row=r + 3, column=1, value="売上計上額との一致確認: OK（差額 ¥0）")
    ws.cell(row=r + 3, column=1).font = BFONT

    widths = [16, 12, 12, 32, 14, 14, 12, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-S-003_月次請求書発行一覧_202511.xlsx")
    print("Created: PLC-S-003_月次請求書発行一覧_202511.xlsx")


# ============================================================
# 4. PLC-S-004 入金消込リスト
# ============================================================
def gen_payment_matching():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "入金消込リスト"

    ws.cell(row=1, column=1, value="【PLC-S-004】2025年11月 入金消込リスト")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    ws.cell(row=2, column=1,
            value="作成日: 2025/12/02 / 作成者: 石井 健（経理部担当）/ 承認: 佐藤 一郎（経理部長）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    headers = ["№", "入金日", "入金額(円)", "銀行", "顧客コード", "顧客名",
               "消込対象\n請求書番号", "消込金額(円)", "差額(円)", "消込方法"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C; c.border = BRD
    ws.row_dimensions[4].height = 32

    random.seed(4004)
    cids = list(CUSTOMERS.keys())
    banks = ["みずほ銀行 本店", "三菱UFJ銀行 丸の内", "三井住友銀行 本店", "横浜銀行 本店"]
    r = 5
    for i in range(1, 31):
        idate = date(2025, 11, random.randint(3, 29))
        amount = random.randint(2_000_000, 40_000_000) // 1000 * 1000
        cid = random.choice(cids)
        cname = CUSTOMERS[cid][0]
        inv = f"INV-202510-{random.randint(1, 150):04d}"
        diff = 0
        method = "SAP自動消込"
        # 意図的に数件手動消込
        if i in (5, 12, 24):
            method = "手動消込\n(金額部分一致)"
        if i == 9:
            diff = -1_500  # 小額差異
            method = "手動消込\n(値引調整)"
        data = [i, idate, amount, random.choice(banks), cid, cname, inv, amount - (diff if diff < 0 else 0), diff, method]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 5, 7, 10):
                cell.alignment = C
                if c_i == 2:
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (3, 8, 9):
                cell.alignment = R
                cell.number_format = "#,##0;[Red]-#,##0"
            else:
                cell.alignment = L
        if diff != 0:
            ws.cell(row=r, column=9).fill = FILL_WARN
            ws.cell(row=r, column=10).fill = FILL_WARN
        r += 1

    ws.cell(row=r, column=1, value="... 以下52件省略 ...").font = Font(name="Yu Gothic", size=9, italic=True)

    widths = [5, 12, 14, 18, 12, 26, 16, 14, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-S-004_入金消込リスト_202511.xlsx")
    print("Created: PLC-S-004_入金消込リスト_202511.xlsx")


# ============================================================
# 5. PLC-S-005 売掛金年齢表（判断保留ケース用）
# ============================================================
def gen_ar_aging():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "売掛金年齢分析"

    ws.cell(row=1, column=1, value="【PLC-S-005】2025年11月末 売掛金年齢分析表")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    ws.cell(row=2, column=1, value="基準日: 2025/11/30  /  作成: 高橋 美咲（経理部課長）")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)

    headers = ["顧客コード", "顧客名", "残高合計(円)", "0-30日\n(正常)", "31-60日", "61-90日",
               "91-120日", "120日超\n(要注意)"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C; c.border = BRD
    ws.row_dimensions[4].height = 32

    random.seed(5005)
    r = 5
    for cid, (cname, credit) in CUSTOMERS.items():
        total = random.randint(3_000_000, min(credit // 2, 200_000_000))
        normal = int(total * 0.7)
        d31 = int(total * 0.2)
        d61 = int(total * 0.06)
        d91 = int(total * 0.03)
        d120 = total - normal - d31 - d61 - d91
        # 一部の顧客で長期滞留
        if cid in ("C-10007", "C-10017", "C-10023"):
            d120 = int(total * 0.15)
            normal = total - d31 - d61 - d91 - d120
        data = [cid, cname, total, normal, d31, d61, d91, d120]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1,):
                cell.alignment = C
            elif c_i == 2:
                cell.alignment = L
            else:
                cell.alignment = R
                cell.number_format = "#,##0"
        # 120日超がある場合は強調
        if d120 > 5_000_000:
            ws.cell(row=r, column=8).fill = FILL_NG
        elif d120 > 0:
            ws.cell(row=r, column=8).fill = FILL_WARN
        r += 1

    # 合計行
    ws.cell(row=r, column=1, value="合計").font = BBOLD
    ws.cell(row=r, column=1).alignment = C
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="D9E1F2")
    ws.cell(row=r, column=2, value="全20社").font = BBOLD
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="D9E1F2")
    for col in range(3, 9):
        ws.cell(row=r, column=col, value=f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{r-1})").font = BBOLD
        ws.cell(row=r, column=col).number_format = "#,##0"
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(row=r, column=col).alignment = R

    # レビューコメント
    r += 2
    ws.cell(row=r, column=1, value="■ 長期滞留債権に関する営業部レビュー").font = BBOLD
    r += 1
    ws.cell(row=r, column=1, value="C-10007 三菱自動車部品販売: 120日超 ¥XX,XXX,XXX → 営業部より「年末までに回収見込（内入金あり）」")
    ws.cell(row=r, column=1).font = BFONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    ws.cell(row=r, column=1, value="C-10017 日立ハイテク部品: 120日超 ¥XX,XXX,XXX → 値引交渉中、一部貸倒引当計上検討")
    ws.cell(row=r, column=1).font = BFONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    ws.cell(row=r, column=1, value="C-10023 丸紅情報システムズ: 120日超 ¥XX,XXX,XXX → 新規追加案件との相殺予定、経理部と協議中")
    ws.cell(row=r, column=1).font = BFONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    # 承認欄（※経理部長承認印が判読不能という設定）
    r += 3
    ws.cell(row=r, column=1, value="■ 承認記録").font = BBOLD
    r += 1
    ws.cell(row=r, column=1, value="作成: 高橋 美咲（経理部課長）[印] 2025/12/08")
    r += 1
    ws.cell(row=r, column=1, value="承認: 佐藤 一郎（経理部長）[印] ※承認印のPDFスキャンが低解像度のため判読困難")
    ws.cell(row=r, column=1).fill = FILL_HOLD
    r += 1
    ws.cell(row=r, column=1, value="→ 追加エビデンス入手中（2026/2/12要求発信）").font = Font(name="Yu Gothic", size=9, italic=True, color="FF0000")

    widths = [12, 28, 16, 14, 14, 14, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-S-005_売掛金年齢表_202511.xlsx")
    print("Created: PLC-S-005_売掛金年齢表_202511.xlsx")


# ============================================================
# 6. PLC-S-006 期末カットオフテスト
# ============================================================
def gen_cutoff_test():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "期末カットオフテスト"

    ws.cell(row=1, column=1, value="【PLC-S-006】FY2025期末 売上カットオフテスト")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    ws.cell(row=2, column=1, value="テスト実施日: 2026/4/3 / 実施者: 佐藤 一郎（経理部長） / 対象期間: 2026/3/25 ～ 2026/4/1 の出荷41件全数")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    headers = ["№", "受注番号", "出荷日", "売上計上日", "顧客", "金額(円)",
               "計上期", "期間帰属", "判定"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C; c.border = BRD
    ws.row_dimensions[4].height = 24

    random.seed(6006)
    cids = list(CUSTOMERS.keys())
    r = 5
    for i in range(1, 42):
        dayidx = random.choice([-6, -5, -4, -3, -2, -1, 0, 1])
        ship_date = date(2026, 3, 31) + timedelta(days=dayidx)
        sale_date = ship_date + timedelta(days=random.choice([0, 0, 1]))
        cid = random.choice(cids)
        amount = random.randint(2_000_000, 20_000_000)
        fy = "FY2025" if sale_date.year < 2026 or (sale_date.year == 2026 and sale_date.month <= 3) else "FY2026"
        expected_fy = "FY2025" if ship_date.year < 2026 or (ship_date.year == 2026 and ship_date.month <= 3) else "FY2026"
        judge = "適切" if fy == expected_fy else "要調整"
        data = [i, f"ORD-2026-{3000 + i}", ship_date, sale_date,
                CUSTOMERS[cid][0][:18], amount, fy, expected_fy, judge]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 7, 8, 9):
                cell.alignment = C
                if c_i in (3, 4):
                    cell.number_format = "yyyy/mm/dd"
            elif c_i == 6:
                cell.alignment = R
                cell.number_format = "#,##0"
            else:
                cell.alignment = L
        if judge == "適切":
            ws.cell(row=r, column=9).fill = FILL_OK
        r += 1

    # 結論
    r += 1
    ws.cell(row=r, column=1, value="結論: 41件すべて適切な期間に計上されており、カットオフ違反なし。").font = BBOLD
    ws.cell(row=r, column=1).fill = FILL_OK
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    ws.cell(row=r, column=1, value="レビュー: 佐藤 一郎 [印] 2026/4/3 / 承認: 渡辺 正博 CFO [印] 2026/4/5").font = BFONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)

    widths = [5, 14, 12, 12, 20, 14, 10, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-S-006_期末カットオフテスト.xlsx")
    print("Created: PLC-S-006_期末カットオフテスト.xlsx")


# ============================================================
# 7. PLC-S-007 価格変更履歴レポート
# ============================================================
def gen_price_history():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "価格マスタ変更履歴"

    ws.cell(row=1, column=1, value="【PLC-S-007】FY2025 Q3 価格マスタ変更履歴レポート")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=13, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    ws.cell(row=2, column=1, value="対象期間: 2025/10/1 ～ 2025/12/31 / 抽出: SAP VK12履歴 / レビュー: 中村 真理")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    headers = ["変更№", "変更日", "製品コード", "顧客コード", "旧単価(円)", "新単価(円)",
               "変更率", "稟議番号", "承認者"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C; c.border = BRD
    ws.row_dimensions[4].height = 28

    samples = [
        (1, date(2025, 10, 16), "P-30011", "C-10011", 18500, 19200, "+3.8%", "W-2025-1876", "田中 太郎"),
        (2, date(2025, 10, 22), "P-30014", "C-10011", 38500, 40000, "+3.9%", "W-2025-1899", "田中 太郎"),
        (3, date(2025, 11, 5), "P-30020", "C-10012", 42500, 44000, "+3.5%", "W-2025-1945", "田中 太郎"),
        (4, date(2025, 11, 12), "P-30006", "C-10002", 12500, 12800, "+2.4%", "W-2025-2012", "田中 太郎"),
        (5, date(2025, 11, 20), "P-30022", "C-10003", 1280, 1300, "+1.6%", "W-2025-2067", "田中 太郎"),
        (6, date(2025, 12, 3), "P-30008", "C-10001", 8200, 8350, "+1.8%", "W-2025-2132", "田中 太郎"),
        (7, date(2025, 12, 10), "P-30015", "C-10015", 12800, 12500, "-2.3%", "W-2025-2178", "田中 太郎"),
        (8, date(2025, 12, 18), "P-30013", "C-10012", 28500, 29200, "+2.5%", "W-2025-2224", "田中 太郎"),
    ]
    r = 5
    for d in samples:
        for c_i, v in enumerate(d, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 3, 4, 7, 8):
                cell.alignment = C
                if c_i == 2:
                    cell.number_format = "yyyy/mm/dd"
            elif c_i in (5, 6):
                cell.alignment = R
                cell.number_format = "#,##0"
            else:
                cell.alignment = L
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="■ レビュー結果").font = BBOLD
    r += 1
    ws.cell(row=r, column=1, value="・全8件とも稟議承認あり、承認ルート（営業課長→本部長）適切")
    ws.cell(row=r, column=1).font = BFONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    ws.cell(row=r, column=1, value="・変更理由：原材料費上昇に伴う改定（№1-6, 8）、競争環境による値下げ（№7）").font = BFONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    ws.cell(row=r, column=1, value="・SAPマスタへの反映日とSAP登録日の一致を全件確認済").font = BFONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 2
    ws.cell(row=r, column=1, value="レビュー実施: 中村 真理 [印] 2026/1/10 / 承認: 高橋 美咲 [印] 2026/1/12").font = BFONT
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)

    widths = [6, 12, 12, 12, 12, 12, 10, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "PLC-S-007_価格変更履歴レポート_Q3.xlsx")
    print("Created: PLC-S-007_価格変更履歴レポート_Q3.xlsx")


if __name__ == "__main__":
    gen_sample_list_s001()
    gen_shipment_sales_match()
    gen_invoice_list()
    gen_payment_matching()
    gen_ar_aging()
    gen_cutoff_test()
    gen_price_history()
