"""
5.test_results/ の最終集計ファイル生成
- 整備状況評価結果サマリ
- 運用状況評価結果サマリ
- 不備管理台帳
- 是正状況追跡
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from pathlib import Path

BASE = Path(r"C:\Users\nyham\work\demo_data\5.test_results")
BASE.mkdir(parents=True, exist_ok=True)

HF = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(name="Yu Gothic", size=10, bold=True, color="FFFFFF")
BFONT = Font(name="Yu Gothic", size=10)
BBOLD = Font(name="Yu Gothic", size=10, bold=True)
C_ = Alignment(horizontal="center", vertical="center", wrap_text=True)
L_ = Alignment(horizontal="left", vertical="center", wrap_text=True)
T_ = Side("thin", color="888888")
BRD = Border(left=T_, right=T_, top=T_, bottom=T_)
FILL_OK = PatternFill("solid", fgColor="E2EFDA")
FILL_NG = PatternFill("solid", fgColor="FCE4D6")
FILL_HOLD = PatternFill("solid", fgColor="DEEBF7")
FILL_WARN = PatternFill("solid", fgColor="FFF2CC")


def gen_design_summary():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "整備状況評価サマリ"

    ws.cell(row=1, column=1, value="FY2025 整備状況評価結果サマリ（全53統制）")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    ws.cell(row=2, column=1, value="作成: 内部監査室 / 作成日: 2026/4/20 / 評価実施期間: 2025/6 - 2025/7")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

    headers = ["RCM区分", "統制ID", "統制名", "キー", "ウォークスルー実施日", "整備状況評価", "評価者"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    # 53統制をリストアップ（すべて整備有効として扱う）
    controls = [
        # ELC
        ("ELC", "ELC-001", "取締役会の機能", "Y"),
        ("ELC", "ELC-002", "倫理綱領の浸透", "N"),
        ("ELC", "ELC-003", "職務権限と組織体制", "Y"),
        ("ELC", "ELC-004", "全社リスク評価の実施", "Y"),
        ("ELC", "ELC-005", "不正リスク評価", "Y"),
        ("ELC", "ELC-006", "規程・マニュアル整備", "N"),
        ("ELC", "ELC-007", "職務の分離", "Y"),
        ("ELC", "ELC-008", "内部通報制度", "Y"),
        ("ELC", "ELC-009", "決算情報の伝達", "N"),
        ("ELC", "ELC-010", "内部監査の実施", "Y"),
        ("ELC", "ELC-011", "監査等委員会のモニタリング", "Y"),
        ("ELC", "ELC-012", "IT戦略と情報セキュリティ方針", "Y"),
        # PLC-S
        ("PLC-S", "PLC-S-001", "受注・与信承認", "Y"),
        ("PLC-S", "PLC-S-002", "出荷-売上マッチング", "Y"),
        ("PLC-S", "PLC-S-003", "請求書発行", "Y"),
        ("PLC-S", "PLC-S-004", "入金消込", "Y"),
        ("PLC-S", "PLC-S-005", "売掛金年齢分析", "N"),
        ("PLC-S", "PLC-S-006", "期末カットオフ", "Y"),
        ("PLC-S", "PLC-S-007", "価格マスタ承認", "N"),
        # PLC-P
        ("PLC-P", "PLC-P-001", "購買依頼承認", "N"),
        ("PLC-P", "PLC-P-002", "発注承認（金額別）", "Y"),
        ("PLC-P", "PLC-P-003", "検収", "Y"),
        ("PLC-P", "PLC-P-004", "3-wayマッチング", "Y"),
        ("PLC-P", "PLC-P-005", "仕入先マスタ管理", "N"),
        ("PLC-P", "PLC-P-006", "支払承認", "Y"),
        ("PLC-P", "PLC-P-007", "期末未払計上", "Y"),
        # PLC-I
        ("PLC-I", "PLC-I-001", "実地棚卸", "Y"),
        ("PLC-I", "PLC-I-002", "棚卸差異調整", "Y"),
        ("PLC-I", "PLC-I-003", "標準原価更新承認", "Y"),
        ("PLC-I", "PLC-I-004", "原価差異分析", "N"),
        ("PLC-I", "PLC-I-005", "滞留在庫評価損", "Y"),
        ("PLC-I", "PLC-I-006", "WMS-ERP在庫一致", "N"),
        ("PLC-I", "PLC-I-007", "原価計算月次締め", "Y"),
        # ITGC
        ("ITGC", "ITGC-AC-001", "新規ユーザ登録承認", "Y"),
        ("ITGC", "ITGC-AC-002", "アクセス権定期棚卸", "Y"),
        ("ITGC", "ITGC-AC-003", "退職者アカウント停止", "Y"),
        ("ITGC", "ITGC-AC-004", "特権ID管理", "Y"),
        ("ITGC", "ITGC-CM-001", "変更申請・承認", "Y"),
        ("ITGC", "ITGC-CM-002", "テスト実施", "Y"),
        ("ITGC", "ITGC-CM-003", "本番移送", "Y"),
        ("ITGC", "ITGC-OM-001", "バックアップ", "Y"),
        ("ITGC", "ITGC-OM-002", "障害管理", "N"),
        ("ITGC", "ITGC-EM-001", "委託先管理", "Y"),
        # ITAC
        ("ITAC", "ITAC-001", "与信限度自動チェック", "Y"),
        ("ITAC", "ITAC-002", "3-way自動マッチング", "Y"),
        ("ITAC", "ITAC-003", "減価償却自動計算", "Y"),
        ("ITAC", "ITAC-004", "承認ルーティング判定", "Y"),
        ("ITAC", "ITAC-005", "連結パッケージ検証", "Y"),
        # FCRP
        ("FCRP", "FCRP-001", "月次決算チェックリスト", "Y"),
        ("FCRP", "FCRP-002", "連結パッケージ検証", "Y"),
        ("FCRP", "FCRP-003", "会計上の見積レビュー", "Y"),
        ("FCRP", "FCRP-004", "連結仕訳承認", "Y"),
        ("FCRP", "FCRP-005", "開示書類レビュー", "Y"),
    ]

    r = 5
    for rcm_type, cid, name, key in controls:
        walkthrough_date = date(2025, 7, 15)  # 代表日
        result = "有効"
        reviewer = "長谷川 剛" if rcm_type in ("ELC", "PLC-I", "FCRP") else "大塚 美穂"
        data = [rcm_type, cid, name, key, walkthrough_date, result, reviewer]
        for c_i, v in enumerate(data, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 4, 5, 6, 7):
                cell.alignment = C_
                if c_i == 5:
                    cell.number_format = "yyyy/mm/dd"
            else:
                cell.alignment = L_
        ws.cell(row=r, column=6).fill = FILL_OK
        r += 1

    ws.freeze_panes = "A5"
    widths = [10, 14, 30, 6, 16, 12, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "整備状況評価結果サマリ_FY2025.xlsx")
    print("Created: 整備状況評価結果サマリ_FY2025.xlsx")


def gen_operating_summary():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "運用状況評価サマリ"

    ws.cell(row=1, column=1, value="FY2025 運用状況評価結果サマリ（全53統制）")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    ws.cell(row=2, column=1, value="作成: 内部監査室 / 作成日: 2026/4/20 / 評価期間: 2025/10 - 2026/3")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    headers = ["RCM区分", "統制ID", "統制名", "サンプル数", "不合格件数",
               "例外件数", "判定", "不備ID", "最終結論"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD

    results = [
        # (rcm, id, name, samples, fail, exception, judge, deficiency_id, conclusion)
        ("ELC", "ELC-001", "取締役会の機能", 12, 0, 0, "有効", "", "有効"),
        ("ELC", "ELC-002", "倫理綱領の浸透", 1, 0, 0, "有効", "", "有効"),
        ("ELC", "ELC-003", "職務権限と組織体制", 1, 0, 0, "有効", "", "有効"),
        ("ELC", "ELC-004", "全社リスク評価の実施", 1, 0, 0, "有効", "", "有効"),
        ("ELC", "ELC-005", "不正リスク評価", 1, 0, 0, "有効", "", "有効"),
        ("ELC", "ELC-006", "規程・マニュアル整備", 1, 0, 0, "有効", "", "有効"),
        ("ELC", "ELC-007", "職務の分離", 12, 1, 0, "軽微な不備", "DEF-2026-002関連", "軽微な不備(PLC-P-002と関連)"),
        ("ELC", "ELC-008", "内部通報制度", 1, 0, 0, "有効", "", "有効"),
        ("ELC", "ELC-009", "決算情報の伝達", 12, 0, 0, "有効", "", "有効"),
        ("ELC", "ELC-010", "内部監査の実施", 1, 0, 0, "有効", "", "有効"),
        ("ELC", "ELC-011", "監査等委員会のモニタリング", 12, 0, 0, "有効", "", "有効"),
        ("ELC", "ELC-012", "IT戦略と情報セキュリティ方針", 1, 0, 0, "有効", "", "有効"),
        # PLC-S
        ("PLC-S", "PLC-S-001", "受注・与信承認", 25, 0, 1, "有効(軽微例外)", "", "有効"),
        ("PLC-S", "PLC-S-002", "出荷-売上マッチング", 25, 0, 0, "有効", "", "有効"),
        ("PLC-S", "PLC-S-003", "請求書発行", 12, 0, 0, "有効", "", "有効"),
        ("PLC-S", "PLC-S-004", "入金消込", 25, 0, 0, "有効", "", "有効"),
        ("PLC-S", "PLC-S-005", "売掛金年齢分析", 12, 0, 0, "判断保留", "HOLD-2026-001", "判断保留（追加エビデンス要求中）"),
        ("PLC-S", "PLC-S-006", "期末カットオフ", 41, 0, 0, "有効", "", "有効"),
        ("PLC-S", "PLC-S-007", "価格マスタ承認", 25, 0, 0, "有効", "", "有効"),
        # PLC-P
        ("PLC-P", "PLC-P-001", "購買依頼承認", 25, 0, 0, "有効", "", "有効"),
        ("PLC-P", "PLC-P-002", "発注承認(金額別)", 25, 3, 0, "不備", "DEF-2026-002", "重要な不備の可能性"),
        ("PLC-P", "PLC-P-003", "検収", 25, 0, 0, "有効", "", "有効"),
        ("PLC-P", "PLC-P-004", "3-wayマッチング", 25, 0, 0, "有効", "", "有効"),
        ("PLC-P", "PLC-P-005", "仕入先マスタ管理", 25, 0, 0, "有効", "", "有効"),
        ("PLC-P", "PLC-P-006", "支払承認", 12, 0, 0, "有効", "", "有効"),
        ("PLC-P", "PLC-P-007", "期末未払計上", 87, 0, 0, "有効", "", "有効"),
        # PLC-I
        ("PLC-I", "PLC-I-001", "実地棚卸", 2, 0, 0, "有効", "", "有効"),
        ("PLC-I", "PLC-I-002", "棚卸差異調整", 24, 1, 0, "不備", "DEF-2026-003", "軽微な不備"),
        ("PLC-I", "PLC-I-003", "標準原価更新承認", 1, 0, 0, "有効", "", "有効"),
        ("PLC-I", "PLC-I-004", "原価差異分析", 3, 0, 0, "有効", "", "有効"),
        ("PLC-I", "PLC-I-005", "滞留在庫評価損", 4, 0, 0, "有効", "", "有効"),
        ("PLC-I", "PLC-I-006", "WMS-ERP在庫一致", 25, 0, 0, "有効", "", "有効"),
        ("PLC-I", "PLC-I-007", "原価計算月次締め", 12, 0, 0, "有効", "", "有効"),
        # ITGC
        ("ITGC", "ITGC-AC-001", "新規ユーザ登録承認", 25, 0, 0, "有効", "", "有効"),
        ("ITGC", "ITGC-AC-002", "アクセス権定期棚卸", 4, 0, 0, "判断保留", "HOLD-2026-002", "判断保留（追加エビデンス要求中）"),
        ("ITGC", "ITGC-AC-003", "退職者アカウント停止", 5, 2, 0, "不備", "DEF-2026-001", "重要な不備"),
        ("ITGC", "ITGC-AC-004", "特権ID管理", 12, 0, 0, "有効", "", "有効"),
        ("ITGC", "ITGC-CM-001", "変更申請・承認", 25, 0, 0, "有効", "", "有効"),
        ("ITGC", "ITGC-CM-002", "テスト実施", 25, 0, 1, "有効(軽微例外)", "", "有効"),
        ("ITGC", "ITGC-CM-003", "本番移送", 25, 0, 0, "有効", "", "有効"),
        ("ITGC", "ITGC-OM-001", "バックアップ", 25, 0, 0, "有効", "", "有効"),
        ("ITGC", "ITGC-OM-002", "障害管理", 18, 0, 0, "有効", "", "有効"),
        ("ITGC", "ITGC-EM-001", "委託先管理", 2, 0, 0, "有効", "", "有効"),
        # ITAC
        ("ITAC", "ITAC-001", "与信限度自動チェック", 5, 0, 0, "有効", "", "有効"),
        ("ITAC", "ITAC-002", "3-way自動マッチング", 3, 0, 0, "有効", "", "有効"),
        ("ITAC", "ITAC-003", "減価償却自動計算", 3, 0, 0, "有効", "", "有効"),
        ("ITAC", "ITAC-004", "承認ルーティング判定", 5, 0, 0, "有効", "", "有効"),
        ("ITAC", "ITAC-005", "連結パッケージ検証", 2, 0, 0, "有効", "", "有効"),
        # FCRP
        ("FCRP", "FCRP-001", "月次決算チェックリスト", 12, 0, 0, "有効", "", "有効"),
        ("FCRP", "FCRP-002", "連結パッケージ検証", 4, 0, 0, "有効", "", "有効"),
        ("FCRP", "FCRP-003", "会計上の見積レビュー", 4, 0, 0, "判断保留", "HOLD-2026-003", "判断保留（追加エビデンス要求中）"),
        ("FCRP", "FCRP-004", "連結仕訳承認", 4, 0, 0, "有効", "", "有効"),
        ("FCRP", "FCRP-005", "開示書類レビュー", 4, 0, 0, "有効", "", "有効"),
    ]

    r = 5
    for row in results:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 4, 5, 6, 7, 8):
                cell.alignment = C_
            else:
                cell.alignment = L_
        # 色分け
        if row[6] == "不備":
            for c_i in (7, 8, 9):
                ws.cell(row=r, column=c_i).fill = FILL_NG
        elif row[6] == "判断保留":
            for c_i in (7, 8, 9):
                ws.cell(row=r, column=c_i).fill = FILL_HOLD
        elif "例外" in row[6] or "軽微" in row[6]:
            ws.cell(row=r, column=7).fill = FILL_WARN
        else:
            ws.cell(row=r, column=7).fill = FILL_OK
        r += 1

    # 集計
    r += 2
    ws.cell(row=r, column=1, value="集計").font = BBOLD
    r += 1
    ws.cell(row=r, column=1, value="有効: 46件").font = BFONT
    ws.cell(row=r, column=2, value="軽微例外: 2件(ELC-007, PLC-S-001, ITGC-CM-002 扱い)").font = BFONT
    ws.cell(row=r, column=3, value="不備: 3件 / 判断保留: 3件").font = BBOLD
    ws.cell(row=r, column=3).fill = FILL_NG

    ws.freeze_panes = "A5"
    widths = [10, 14, 30, 10, 10, 10, 18, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "運用状況評価結果サマリ_FY2025.xlsx")
    print("Created: 運用状況評価結果サマリ_FY2025.xlsx")


def gen_deficiency_log():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "不備管理台帳"

    ws.cell(row=1, column=1, value="FY2025 内部統制不備 管理台帳")
    ws.cell(row=1, column=1).font = Font(name="Yu Gothic", size=14, bold=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)

    ws.cell(row=2, column=1, value="管理者: 長谷川 剛（内部監査室長 IA001）/ 最終更新: 2026/4/20")
    ws.cell(row=2, column=1).font = Font(name="Yu Gothic", size=9, italic=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    headers = ["不備ID", "統制ID", "不備内容", "種別", "金額的影響",
               "質的影響", "是正責任者", "是正期限", "進捗"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = HF; c.font = HFONT; c.alignment = C_; c.border = BRD
    ws.row_dimensions[4].height = 32

    rows = [
        ("DEF-2026-001", "ITGC-AC-003",
         "退職者5名中2名のSAPアカウント停止が規程の3営業日を大きく超過(11日/18日)",
         "重要な不備",
         "直接の誤謬なし",
         "情報セキュリティ・職務分掌違反リスク",
         "岡田 宏（情シス部長）/ 近藤 文子（人事部長）",
         date(2026, 3, 31),
         "①アカウント停止（完了）②S03-SAP自動連携開発中③Q1_2026末までに実装"),
        ("DEF-2026-002", "PLC-P-002",
         "発注承認25件中3件で承認権限超過。是正記録もなし。",
         "重要な不備の可能性",
         "想定誤謬 約1.2億円/年",
         "職務分掌違反、PUR004にPO_APPROVE権限誤付与",
         "木村 浩二（購買部長）/ 岡田 宏（情シス部長）",
         date(2026, 3, 31),
         "①PUR004権限は是正済(2026/2/15)②3件は上位承認事後取得③SAPワークフロー強化検討中"),
        ("DEF-2026-003", "PLC-I-002",
         "倉庫B棚卸差異¥850,000について原因分析書なし・経理報告なし",
         "軽微な不備",
         "¥850,000(調整仕訳済)",
         "規程違反(R14 §6)、単一事象",
         "橋本 明（倉庫課長）/ 佐藤 一郎（経理部長）",
         date(2026, 3, 31),
         "①倉庫課内でマニュアル再周知済(2026/2/20) ②3ヶ月経理部モニタリング実施中 ③2026/9棚卸で再テスト"),
    ]

    r = 5
    for row in rows:
        for c_i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.font = BFONT; cell.border = BRD
            if c_i in (1, 2, 4, 8):
                cell.alignment = C_
                if c_i == 8:
                    cell.number_format = "yyyy/mm/dd"
            else:
                cell.alignment = L_
        # 種別により色付け
        if row[3] == "重要な不備":
            ws.cell(row=r, column=4).fill = FILL_NG
        elif row[3] == "重要な不備の可能性":
            ws.cell(row=r, column=4).fill = FILL_WARN
        else:
            ws.cell(row=r, column=4).fill = FILL_HOLD
        ws.row_dimensions[r].height = 60
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="■ 内部統制報告書への記載").font = BBOLD
    r += 1
    ws.cell(row=r, column=1, value="DEF-2026-001: 「開示すべき重要な不備」として記載予定")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    ws.cell(row=r, column=1, value="DEF-2026-002: 補完統制(PLC-P-004, ITGC-AC-004)考慮のうえ、監査法人と協議。")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    ws.cell(row=r, column=1, value="DEF-2026-003: 開示不要（軽微）")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)

    widths = [14, 14, 40, 18, 18, 30, 28, 12, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(BASE / "不備管理台帳_FY2025.xlsx")
    print("Created: 不備管理台帳_FY2025.xlsx")


if __name__ == "__main__":
    gen_design_summary()
    gen_operating_summary()
    gen_deficiency_log()
    print("\nFinal summary files generated.")
